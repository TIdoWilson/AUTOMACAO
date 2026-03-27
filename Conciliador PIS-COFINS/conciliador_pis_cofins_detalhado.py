from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from unicodedata import normalize

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


PASTA_BASE = Path(__file__).resolve().parent
PASTA_ARQUIVOS = PASTA_BASE / "Arquivos"
TOLERANCIA = Decimal("0.10")
XLSX_FORMATO_MOEDA_BR = "[$R$-416] #,##0.00"
MIN_NOTA_SAIDA_RAZAO = 10000
MAX_DIGITOS_NOTA_ENTRADA_RAZAO = 7
MODO_PROCESSAMENTO = "AUTO"  # AUTO | RECUPERAR | RECOLHER

RE_DATA_VALOR_SALDO = re.compile(
    r"^(\d{1,2}/\d{2}/\d{4})\s+(-?\d{1,3}(?:\.\d{3})*,\d{2})\s+(-?\d{1,3}(?:\.\d{3})*,\d{2})\s+-\s+(.+)$",
    re.IGNORECASE,
)
RE_NF = re.compile(r"\b(?:NF|NFS-E)\s*:?\s*(\d{3,})\b", re.IGNORECASE)
RE_LINHA_RELATORIO = re.compile(r"^\s*(\d{3,})\s+\S+\s+\d{2}/\d{2}/\d{4}\S*\s+")
RE_MONEY_ANY = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2,4}")


@dataclass(frozen=True)
class RegistroRazao:
    nota: str
    data_lcto: str
    valor: Decimal
    historico: str
    arquivo: str


@dataclass(frozen=True)
class RegistroRelatorio:
    nota: str
    data_lcto: str
    valor_pis: Decimal
    valor_cofins: Decimal
    arquivo: str


def normalizar_ascii(s: str) -> str:
    return normalize("NFKD", s).encode("ASCII", "ignore").decode().upper()


def valor_br_para_decimal(s: str) -> Decimal:
    return Decimal(s.replace(".", "").replace(",", "."))


def normalizar_nota(s: str) -> str:
    return re.sub(r"\D", "", s).lstrip("0") or "0"


def ler_linhas_pdf(pdf_path: Path) -> list[str]:
    txt = "\n".join((pg.extract_text() or "") for pg in PdfReader(str(pdf_path)).pages)
    return [ln.strip() for ln in txt.splitlines() if ln.strip()]


def detectar_tipo_arquivo_por_conteudo(linhas: list[str]) -> tuple[str | None, str | None]:
    cab = normalizar_ascii("\n".join(linhas[:220]))
    cab_compacto = re.sub(r"\s+", " ", cab)

    if "RELATORIO DE APURACAO DE PIS E COFINS" in cab_compacto:
        if " - ENTRADAS - " in cab_compacto:
            return "entrada_relatorio", "AMBOS"
        if " - SAIDA - " in cab_compacto or " - SAIDAS - " in cab_compacto:
            return "saida_relatorio", "AMBOS"
        return None, None

    if "RELATORIO: RAZAO CONTABIL" not in cab_compacto and "RELATORIO:RAZAO CONTABIL" not in cab_compacto:
        return None, None

    if "PIS A RECUPERAR" in cab_compacto:
        return "entrada_razao", "PIS"
    if "PIS A RECOLHER" in cab_compacto:
        return "saida_razao", "PIS"
    if "COFINS A RECUPERAR" in cab_compacto:
        return "entrada_razao", "COFINS"
    if "COFINS A RECOLHER" in cab_compacto:
        return "saida_razao", "COFINS"

    return None, None


def parse_razao(pdf_path: Path, linhas: list[str]) -> list[RegistroRazao]:
    out: list[RegistroRazao] = []
    for ln in linhas:
        m = RE_DATA_VALOR_SALDO.match(ln)
        if not m:
            continue
        data_lcto, valor_txt, _, historico = m.groups()
        nf = RE_NF.search(historico)
        if not nf:
            continue
        nota = normalizar_nota(nf.group(1))
        valor = valor_br_para_decimal(valor_txt)
        if valor == 0:
            continue
        out.append(
            RegistroRazao(
                nota=nota,
                data_lcto=data_lcto,
                valor=valor,
                historico=historico[:220],
                arquivo=pdf_path.name,
            )
        )
    return out


def filtrar_razao_por_layout(registros: list[RegistroRazao], movimento: str) -> tuple[list[RegistroRazao], int]:
    filtrados: list[RegistroRazao] = []
    descartados = 0
    for r in registros:
        if not r.nota.isdigit():
            descartados += 1
            continue
        if movimento == "saida" and int(r.nota) < MIN_NOTA_SAIDA_RAZAO:
            descartados += 1
            continue
        if movimento == "entrada" and len(r.nota) > MAX_DIGITOS_NOTA_ENTRADA_RAZAO:
            descartados += 1
            continue
        filtrados.append(r)
    return filtrados, descartados


def parse_relatorio(pdf_path: Path, linhas: list[str]) -> list[RegistroRelatorio]:
    out: list[RegistroRelatorio] = []
    for ln in linhas:
        if not RE_LINHA_RELATORIO.match(ln):
            continue
        partes = ln.split()
        if len(partes) < 4:
            continue
        nota = normalizar_nota(partes[0])
        m_data = re.search(r"\d{2}/\d{2}/\d{4}", ln)
        data_lcto = m_data.group(0) if m_data else ""
        valores = RE_MONEY_ANY.findall(ln)
        if len(valores) < 9:
            continue
        # No layout dos exemplos, os campos de tributo seguem posicao estavel.
        # indice 5 -> valor PIS; indice 8 -> valor COFINS.
        try:
            valor_pis = valor_br_para_decimal(valores[5])
            valor_cofins = valor_br_para_decimal(valores[8])
        except Exception:
            continue
        out.append(
            RegistroRelatorio(
                nota=nota,
                data_lcto=data_lcto,
                valor_pis=valor_pis,
                valor_cofins=valor_cofins,
                arquivo=pdf_path.name,
            )
        )
    return out


def somar_por_nota_razao(registros: list[RegistroRazao]) -> dict[str, Decimal]:
    out: dict[str, Decimal] = defaultdict(Decimal)
    for r in registros:
        out[r.nota] += r.valor
    return dict(out)


def somar_por_nota_relatorio(registros: list[RegistroRelatorio], tributo: str) -> dict[str, Decimal]:
    out: dict[str, Decimal] = defaultdict(Decimal)
    for r in registros:
        if tributo == "PIS":
            out[r.nota] += r.valor_pis
        else:
            out[r.nota] += r.valor_cofins
    return dict(out)


def autoajustar_colunas(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            txt = "" if val is None else str(val)
            if len(txt) > max_len:
                max_len = len(txt)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 70))


def filtrar_exemplo(lista: list[object], limite: int = 3) -> str:
    if not lista:
        return ""
    return " | ".join(str(x) for x in lista[:limite])


def escrever_aba_inconsistencias(ws, linhas: list[list[object]]) -> None:
    ws.append(
        [
            "TIPO",
            "NOTA",
            "VALOR_RAZAO",
            "VALOR_RELATORIO",
            "DIFERENCA",
            "EXEMPLOS_RAZAO",
            "EXEMPLOS_RELATORIO",
        ]
    )
    for row in linhas:
        ws.append(row)
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=5):
        for c in row:
            if isinstance(c.value, (int, float)):
                c.number_format = XLSX_FORMATO_MOEDA_BR
    autoajustar_colunas(ws)


def escrever_xlsx(caminho: Path, nome_aba_pis: str, linhas_pis: list[list[object]], nome_aba_cofins: str, linhas_cofins: list[list[object]]) -> Path:
    wb = Workbook()
    ws_pis = wb.active
    ws_pis.title = nome_aba_pis
    escrever_aba_inconsistencias(ws_pis, linhas_pis)

    ws_cofins = wb.create_sheet(nome_aba_cofins)
    escrever_aba_inconsistencias(ws_cofins, linhas_cofins)

    wb.save(caminho)
    return caminho


def descobrir_pasta_pdfs() -> Path:
    candidatas = [
        PASTA_BASE / "Entrada",
        PASTA_BASE / "Arquivo de exemplo",
        PASTA_BASE,
    ]
    for pasta in candidatas:
        if pasta.exists() and any(pasta.glob("*.pdf")):
            return pasta
    raise RuntimeError(
        "Nao foi encontrada pasta com PDFs. Coloque os arquivos em "
        f"'{PASTA_BASE}\\Entrada' ou '{PASTA_BASE}\\Arquivo de exemplo'."
    )


def obter_arquivo_unico(lista: list[Path], descricao: str) -> Path:
    if not lista:
        raise RuntimeError(f"Arquivo obrigatorio nao encontrado para: {descricao}")
    if len(lista) > 1:
        nomes = ", ".join(p.name for p in lista)
        raise RuntimeError(f"Encontrado mais de um arquivo para {descricao}: {nomes}")
    return lista[0]


def montar_linhas_inconsistencia(
    mapa_razao: dict[str, Decimal],
    mapa_relatorio: dict[str, Decimal],
    exemplos_razao: dict[str, list[str]],
    exemplos_relatorio: dict[str, list[str]],
) -> list[list[object]]:
    out: list[list[object]] = []
    notas = sorted(set(mapa_razao) | set(mapa_relatorio))
    for nota in notas:
        vr = mapa_razao.get(nota, Decimal("0"))
        vl = mapa_relatorio.get(nota, Decimal("0"))
        dif = vr - vl

        if nota not in mapa_relatorio:
            tipo = "SO_NO_RAZAO"
        elif nota not in mapa_razao:
            tipo = "SO_NO_RELATORIO"
        elif abs(dif) > TOLERANCIA:
            tipo = "VALOR_DIVERGENTE"
        else:
            continue

        out.append(
            [
                tipo,
                nota,
                float(vr),
                float(vl),
                float(dif),
                filtrar_exemplo(exemplos_razao.get(nota, [])),
                filtrar_exemplo(exemplos_relatorio.get(nota, [])),
            ]
        )
    return out


def main() -> None:
    pasta = descobrir_pasta_pdfs()
    if not pasta.exists():
        raise FileNotFoundError(f"Pasta nao encontrada: {pasta}")

    pdfs = sorted(pasta.glob("*.pdf"))
    if not pdfs:
        raise RuntimeError("Nenhum PDF encontrado na pasta selecionada.")

    arquivos: dict[str, dict[str, list[Path]] | list[Path]] = {
        "entrada_razao": {"PIS": [], "COFINS": []},
        "saida_razao": {"PIS": [], "COFINS": []},
        "entrada_relatorio": [],
        "saida_relatorio": [],
    }
    cache_linhas: dict[Path, list[str]] = {}

    for pdf in pdfs:
        linhas = ler_linhas_pdf(pdf)
        cache_linhas[pdf] = linhas
        tipo, tributo = detectar_tipo_arquivo_por_conteudo(linhas)
        if tipo is None:
            continue

        if tipo in {"entrada_razao", "saida_razao"} and tributo in {"PIS", "COFINS"}:
            casted = arquivos[tipo]
            casted[tributo].append(pdf)  # type: ignore[index]
        elif tipo in {"entrada_relatorio", "saida_relatorio"}:
            casted_rel = arquivos[tipo]
            casted_rel.append(pdf)  # type: ignore[union-attr]

    movimentos: list[tuple[str, str]] = []
    if MODO_PROCESSAMENTO in {"AUTO", "RECUPERAR"}:
        movimentos.append(("entrada", "RECUPERAR"))
    if MODO_PROCESSAMENTO in {"AUTO", "RECOLHER"}:
        movimentos.append(("saida", "RECOLHER"))

    gerados: list[Path] = []

    for mov, rotulo in movimentos:
        chave_razao = "entrada_razao" if mov == "entrada" else "saida_razao"
        chave_rel = "entrada_relatorio" if mov == "entrada" else "saida_relatorio"

        razao_pis_pdf = obter_arquivo_unico(arquivos[chave_razao]["PIS"], f"{rotulo} Razao PIS")  # type: ignore[index]
        razao_cof_pdf = obter_arquivo_unico(arquivos[chave_razao]["COFINS"], f"{rotulo} Razao COFINS")  # type: ignore[index]
        rel_pdf = obter_arquivo_unico(arquivos[chave_rel], f"{rotulo} Relatorio")  # type: ignore[arg-type]

        regs_razao_pis = parse_razao(razao_pis_pdf, cache_linhas[razao_pis_pdf])
        regs_razao_cof = parse_razao(razao_cof_pdf, cache_linhas[razao_cof_pdf])
        regs_relatorio = parse_relatorio(rel_pdf, cache_linhas[rel_pdf])

        regs_razao_pis, desc_pis = filtrar_razao_por_layout(regs_razao_pis, mov)
        regs_razao_cof, desc_cof = filtrar_razao_por_layout(regs_razao_cof, mov)

        if not regs_razao_pis or not regs_razao_cof:
            raise RuntimeError(f"Sem registros validos de Razao para {rotulo}.")
        if not regs_relatorio:
            raise RuntimeError(f"Sem registros validos no Relatorio para {rotulo}.")

        linhas_saida: dict[str, list[list[object]]] = {}
        for tributo, regs_razao in (("PIS", regs_razao_pis), ("COFINS", regs_razao_cof)):
            mapa_razao = somar_por_nota_razao(regs_razao)
            mapa_rel = somar_por_nota_relatorio(regs_relatorio, tributo)

            ex_razao: dict[str, list[str]] = defaultdict(list)
            ex_rel: dict[str, list[str]] = defaultdict(list)
            for r in regs_razao:
                ex_razao[r.nota].append(f"{r.data_lcto} {r.valor:.2f}")
            for r in regs_relatorio:
                v = r.valor_pis if tributo == "PIS" else r.valor_cofins
                ex_rel[r.nota].append(f"{r.data_lcto} {v:.2f}")

            linhas = montar_linhas_inconsistencia(mapa_razao, mapa_rel, dict(ex_razao), dict(ex_rel))
            linhas_saida[tributo] = linhas
            desc = desc_pis if tributo == "PIS" else desc_cof
            print(
                f"{rotulo} {tributo} -> Razao: {len(regs_razao)} (desc. {desc}) | "
                f"Relatorio: {len(regs_relatorio)} | Inconsistencias: {len(linhas)}"
            )

        PASTA_ARQUIVOS.mkdir(parents=True, exist_ok=True)
        out = PASTA_ARQUIVOS / f"Conciliacao_PIS_COFINS_{rotulo}_{datetime.now().strftime('%H-%M-%S_%d-%m-%Y')}.xlsx"
        escrever_xlsx(out, f"{rotulo}_PIS", linhas_saida["PIS"], f"{rotulo}_COFINS", linhas_saida["COFINS"])
        gerados.append(out)
        print(f"Arquivo gerado ({rotulo}): {out}")

    if not gerados:
        raise RuntimeError("Nenhum trio completo encontrado (PIS Razao + COFINS Razao + Relatorio).")


if __name__ == "__main__":
    main()
