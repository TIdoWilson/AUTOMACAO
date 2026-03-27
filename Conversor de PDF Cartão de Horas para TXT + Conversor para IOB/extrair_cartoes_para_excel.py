from __future__ import annotations

import argparse
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from pypdf import PdfReader


PASTA_BASE = Path(__file__).resolve().parent
PASTA_EXEMPLO = PASTA_BASE / "Arquivos de exemplo"
PASTA_ARQUIVOS_GERADOS = PASTA_BASE / "Arquivos gerados"
PASTA_ARQUIVOS_IOB = PASTA_ARQUIVOS_GERADOS / "IOB"

PADRAO_DATA = re.compile(r"^(\d{2}/\d{2}/\d{4}) - ([A-Z]{3})$")
PADRAO_HORA = re.compile(r"^\d{2,3}:\d{2}$")
PADRAO_BATIDA = re.compile(r"^\d{2}:\d{2} \([A-Z]\)$")
PADRAO_DATA_ZOOPET = re.compile(r"^([A-Z]?)\s*(\d{2}/\d{2})\s+([A-Za-zÀ-ÿ]{3}\.)\s+(.*)$")
PADRAO_DATA_CESUL = re.compile(r"^([A-Za-zÀ-ÿ]{3}), (\d{2}/\d{2}/\d{4})$")

LIGATURAS = {
    "\ufb00": "ff",
    "\ufb01": "fi",
    "\ufb02": "fl",
    "\ufb03": "ffi",
    "\ufb04": "ffl",
}

MAPA_POSICOES_BATIDA = {
    1: "Entrada manha",
    2: "Saida almoco",
    3: "Volta almoco",
    4: "Saida final",
    5: "Entrada extra",
    6: "Saida extra",
}

EVENTOS_IOB = {
    "geral_horas_normais": 0,
    "geral_extra_50": 0,
    "geral_extra_100": 0,
    "geral_extra_diurna": 0,
    "geral_extra_noturna": 0,
    "geral_adicional_noturno": 0,
    "geral_hora_noturna_reduzida": 0,
    "geral_falta_atraso": 0,
}

TIPOS_EVENTO_IOB = {
    "geral_horas_normais": "Horas normais do periodo",
    "geral_extra_50": "Hora extra 50%",
    "geral_extra_100": "Hora extra 100%",
    "geral_extra_diurna": "Hora extra diurna",
    "geral_extra_noturna": "Hora extra noturna",
    "geral_adicional_noturno": "Adicional noturno",
    "geral_hora_noturna_reduzida": "Hora noturna reduzida",
    "geral_falta_atraso": "Falta e atraso/ausencia",
}

ORDEM_TIPOS_EVENTO_IOB = [
    "geral_horas_normais",
    "geral_extra_50",
    "geral_extra_100",
    "geral_extra_diurna",
    "geral_extra_noturna",
    "geral_adicional_noturno",
    "geral_hora_noturna_reduzida",
    "geral_falta_atraso",
]


@dataclass
class PaginaProcessada:
    resumo: dict
    lancamentos: list[dict]


def normalizar_texto(texto: str) -> str:
    for origem, destino in LIGATURAS.items():
        texto = texto.replace(origem, destino)
    return texto.replace("\xa0", " ").replace("\r", "")


def extrair_campo(texto: str, rotulo: str) -> str:
    padrao = rf"{re.escape(rotulo)}:\s*\n?\s*(.+)"
    encontrado = re.search(padrao, texto)
    return encontrado.group(1).strip() if encontrado else ""


def extrair_periodo(texto: str) -> tuple[str, str]:
    encontrado = re.search(r"DE (\d{2}/\d{2}/\d{4}) ATÉ (\d{2}/\d{2}/\d{4})", texto)
    if not encontrado:
        return "", ""
    return encontrado.group(1), encontrado.group(2)


def extrair_periodo_generico(texto: str) -> tuple[str, str]:
    encontrado = re.search(r"Período:\s*(\d{2}/\d{2}/\d{4})\s*a\s*(\d{2}/\d{2}/\d{4})", texto)
    if encontrado:
        return encontrado.group(1), encontrado.group(2)
    encontrado = re.search(r"De (\d{2}/\d{2}/\d{4}) até (\d{2}/\d{2}/\d{4})", texto)
    if encontrado:
        return encontrado.group(1), encontrado.group(2)
    return extrair_periodo(texto)


def converter_data(valor: str) -> datetime | None:
    if not valor:
        return None
    return datetime.strptime(valor, "%d/%m/%Y")


def eh_hora_texto(valor: str) -> bool:
    return bool(re.match(r"^-?\d{2,3}:\d{2}$", valor or ""))


def eh_token_horario(valor: str) -> bool:
    return bool(re.match(r"^(--:--|-?\d{2,3}:\d{2})$", valor or ""))


def hora_positiva(valor: str) -> bool:
    if not eh_hora_texto(valor):
        return False
    if valor.startswith("-"):
        return False
    return valor not in {"00:00", "--:--"}


def listar_linhas(texto: str) -> list[str]:
    linhas = []
    for linha in texto.splitlines():
        valor = linha.strip()
        if valor == "":
            continue
        linhas.append(valor)
    return linhas


def separar_bloco_diario(linhas: list[str]) -> tuple[list[dict], list[str]]:
    lancamentos = []
    totais = []

    try:
        inicio = next(i for i, linha in enumerate(linhas) if PADRAO_DATA.match(linha))
        fim = linhas.index("TOTAIS")
    except StopIteration:
        return lancamentos, totais
    except ValueError:
        return lancamentos, totais

    i = inicio
    while i < fim:
        cabecalho = linhas[i]
        correspondencia = PADRAO_DATA.match(cabecalho)
        if not correspondencia:
            i += 1
            continue

        data_referencia = correspondencia.group(1)
        dia_semana = correspondencia.group(2)
        i += 1

        bloco = []
        while i < fim and not PADRAO_DATA.match(linhas[i]):
            bloco.append(linhas[i])
            i += 1

        previsto = bloco[0] if bloco else ""
        registros = []
        metricas = []
        observacoes = []

        for item in bloco[1:]:
            if PADRAO_BATIDA.match(item) or item == "Falta":
                registros.append(item)
            elif PADRAO_HORA.match(item):
                metricas.append(item)
            else:
                observacoes.append(item)

        if previsto in {"Folga", "-"}:
            status = previsto
        elif "Falta" in registros:
            status = "Falta parcial"
        else:
            status = "Trabalhado"

        linha = {
            "data": data_referencia,
            "dia_semana": dia_semana,
            "previsto": previsto,
            "status": status,
            "registro_1": registros[0] if len(registros) > 0 else "",
            "registro_2": registros[1] if len(registros) > 1 else "",
            "registro_3": registros[2] if len(registros) > 2 else "",
            "registro_4": registros[3] if len(registros) > 3 else "",
            "registro_5": registros[4] if len(registros) > 4 else "",
            "registro_6": registros[5] if len(registros) > 5 else "",
            "horas_total_dia": metricas[0] if metricas else "",
            "adicional_1": metricas[1] if len(metricas) > 1 else "",
            "adicional_2": metricas[2] if len(metricas) > 2 else "",
            "adicional_3": metricas[3] if len(metricas) > 3 else "",
            "qtd_batidas_validas": len([registro for registro in registros if registro != "Falta"]),
            "tem_inclusao_manual": "Sim" if any("(I)" in registro for registro in registros) else "Nao",
            "tem_pre_assinalado": "Sim" if any("(P)" in registro for registro in registros) else "Nao",
            "observacoes": " | ".join(observacoes),
            "iob_extra_50": "",
            "iob_extra_diurna": "",
            "iob_extra_noturna": "",
            "iob_falta_atraso": "",
            "iob_zoopet_extra_diurna": "",
            "iob_zoopet_extra_noturna": "",
            "iob_zoopet_ausencia_diurna": "",
            "iob_zoopet_ausencia_noturna": "",
            "iob_cesul_he50": "",
            "iob_cesul_he100": "",
            "iob_cesul_adic_noturno": "",
            "iob_cesul_hn_reduzida": "",
        }
        if status == "Falta parcial":
            linha["iob_falta_atraso"] = linha["adicional_1"]
        else:
            linha["iob_extra_50"] = linha["adicional_1"]
            linha["iob_extra_diurna"] = linha["adicional_2"]
            linha["iob_extra_noturna"] = linha["adicional_3"]
        lancamentos.append(linha)

    i_totais = fim + 1
    while i_totais < len(linhas):
        item = linhas[i_totais]
        if item.startswith("(I)=Incluído") or item.startswith("(I)=Incluido"):
            break
        if PADRAO_HORA.match(item):
            totais.append(item)
        i_totais += 1

    return lancamentos, totais


def montar_resumo_base(arquivo_pdf: Path, pagina: int, resumo: dict, lancamentos: list[dict]) -> PaginaProcessada:
    for lancamento in lancamentos:
        lancamento["arquivo_pdf"] = arquivo_pdf.name
        lancamento["pagina_pdf"] = pagina
        lancamento["empresa"] = resumo["empresa"]
        lancamento["cnpj"] = resumo["cnpj"]
        lancamento["funcionario"] = resumo["funcionario"]
        lancamento["matricula"] = resumo["matricula"]
        lancamento["cargo"] = resumo["cargo"]
        lancamento["departamento"] = resumo["departamento"]
    return PaginaProcessada(resumo=resumo, lancamentos=lancamentos)


def processar_pagina_faenello(texto: str, arquivo_pdf: Path, pagina: int) -> PaginaProcessada:
    linhas = listar_linhas(texto)
    periodo_inicial, periodo_final = extrair_periodo_generico(texto)
    lancamentos, totais = separar_bloco_diario(linhas)

    resumo = {
        "arquivo_pdf": arquivo_pdf.name,
        "pagina_pdf": pagina,
        "layout_origem": "faenello",
        "empresa": extrair_campo(texto, "NOME DA EMPRESA"),
        "cnpj": extrair_campo(texto, "CNPJ DA EMPRESA"),
        "inscricao_estadual": extrair_campo(texto, "INSCRIÇÃO ESTADUAL DA EMPRESA"),
        "funcionario": extrair_campo(texto, "NOME DO FUNCIONÁRIO"),
        "cpf": extrair_campo(texto, "CPF DO FUNCIONÁRIO"),
        "pis": extrair_campo(texto, "PIS DO FUNCIONÁRIO"),
        "ctps": extrair_campo(texto, "CTPS DO FUNCIONÁRIO"),
        "data_admissao": extrair_campo(texto, "DATA DE ADMISSÃO DO FUNCIONÁRIO"),
        "cargo": extrair_campo(texto, "NOME DO CARGO"),
        "matricula": extrair_campo(texto, "NÚMERO DE MATRÍCULA"),
        "departamento": extrair_campo(texto, "NOME DO DEPARTAMENTO"),
        "periodo_inicial": periodo_inicial,
        "periodo_final": periodo_final,
        "horas_normais_mes": totais[0] if totais else "",
        "total_adicional_1": totais[1] if len(totais) > 1 else "",
        "total_adicional_2": totais[2] if len(totais) > 2 else "",
        "total_adicional_3": totais[3] if len(totais) > 3 else "",
        "totais_brutos": " | ".join(totais),
    }

    return montar_resumo_base(arquivo_pdf, pagina, resumo, lancamentos)


def processar_pagina_zoopet(texto: str, arquivo_pdf: Path, pagina: int) -> PaginaProcessada:
    linhas = listar_linhas(texto)
    periodo_inicial, periodo_final = extrair_periodo_generico(texto)

    empresa = linhas[0] if linhas else ""
    cnpj = ""
    for linha in linhas[:5]:
        encontrado = re.search(r"(\d{14})CNPJ", linha)
        if encontrado:
            cnpj = encontrado.group(1)
            break

    funcionario = ""
    departamento = ""
    data_admissao = ""
    cargo = ""
    pis = ""

    try:
        indice_cargo = linhas.index("Cargo: CTPS:")
        funcionario = linhas[indice_cargo + 1] if len(linhas) > indice_cargo + 1 else ""
        departamento = linhas[indice_cargo + 2] if len(linhas) > indice_cargo + 2 else ""
        data_admissao = linhas[indice_cargo + 3] if len(linhas) > indice_cargo + 3 else ""
        cargo = linhas[indice_cargo + 4] if len(linhas) > indice_cargo + 4 else ""
    except ValueError:
        pass

    for linha in linhas:
        if linha.startswith("PIS:"):
            pis = linha.replace("PIS:", "").strip()
            break

    ano_inicial = periodo_inicial[-4:] if periodo_inicial else ""
    ano_final = periodo_final[-4:] if periodo_final else ""
    mes_inicial = periodo_inicial[3:5] if periodo_inicial else ""

    lancamentos = []
    totais = []
    for linha in linhas:
        if linha.startswith("Totais:"):
            totais = re.findall(r"\d{3}:\d{2}|\d{2}:\d{2}", linha)
            continue

        correspondencia = PADRAO_DATA_ZOOPET.match(linha)
        if not correspondencia:
            continue

        marcador = correspondencia.group(1)
        data_curta = correspondencia.group(2)
        dia_semana = correspondencia.group(3).replace(".", "").upper()[:3]
        resto = correspondencia.group(4).split()

        if len(resto) < 10:
            continue

        ano = ano_inicial if data_curta[3:5] == mes_inicial else ano_final
        data_completa = f"{data_curta}/{ano}" if ano else data_curta

        previsto_tokens = resto[0:4]
        qtd_registros_esperados = len([token for token in previsto_tokens if token != "--:--"])
        registro_tokens = resto[4 : 4 + qtd_registros_esperados]
        metrica_tokens = []
        observacao_tokens = []
        encontrou_observacao = False
        for token in resto[4 + qtd_registros_esperados :]:
            if not encontrou_observacao and eh_token_horario(token):
                metrica_tokens.append(token)
            else:
                encontrou_observacao = True
                observacao_tokens.append(token)
        observacoes = " ".join(observacao_tokens)

        previsto = " ".join([token for token in previsto_tokens if token != "--:--"]).strip() or "-"
        registros_validos = [token for token in registro_tokens if token != "--:--"]

        if marcador == "F":
            status = "Feriado"
        elif marcador == "A":
            status = "Afastamento"
        elif all(token == "--:--" for token in registro_tokens):
            status = "Sem registro"
        else:
            status = "Trabalhado"

        lancamentos.append(
            {
                "data": data_completa,
                "dia_semana": dia_semana,
                "previsto": previsto,
                "status": status,
                "registro_1": registros_validos[0] if len(registros_validos) > 0 else "",
                "registro_2": registros_validos[1] if len(registros_validos) > 1 else "",
                "registro_3": registros_validos[2] if len(registros_validos) > 2 else "",
                "registro_4": registros_validos[3] if len(registros_validos) > 3 else "",
                "registro_5": "",
                "registro_6": "",
                "horas_total_dia": metrica_tokens[2] if len(metrica_tokens) > 2 and hora_positiva(metrica_tokens[2]) else "",
                "adicional_1": metrica_tokens[4] if len(metrica_tokens) > 4 and hora_positiva(metrica_tokens[4]) else "",
                "adicional_2": "",
                "adicional_3": "",
                "qtd_batidas_validas": len(registros_validos),
                "tem_inclusao_manual": "Nao",
                "tem_pre_assinalado": "Nao",
                "observacoes": observacoes,
                "iob_extra_50": "",
                "iob_extra_diurna": "",
                "iob_extra_noturna": "",
                "iob_falta_atraso": "",
                "iob_zoopet_extra_diurna": metrica_tokens[4] if len(metrica_tokens) > 4 and hora_positiva(metrica_tokens[4]) else "",
                "iob_zoopet_extra_noturna": metrica_tokens[5] if len(metrica_tokens) > 5 and hora_positiva(metrica_tokens[5]) else "",
                "iob_zoopet_ausencia_diurna": metrica_tokens[6] if len(metrica_tokens) > 6 and hora_positiva(metrica_tokens[6]) else "",
                "iob_zoopet_ausencia_noturna": metrica_tokens[7] if len(metrica_tokens) > 7 and hora_positiva(metrica_tokens[7]) else "",
                "iob_cesul_he50": "",
                "iob_cesul_he100": "",
                "iob_cesul_adic_noturno": "",
                "iob_cesul_hn_reduzida": "",
            }
        )

    resumo = {
        "arquivo_pdf": arquivo_pdf.name,
        "pagina_pdf": pagina,
        "layout_origem": "zoopet",
        "empresa": empresa,
        "cnpj": cnpj,
        "inscricao_estadual": "",
        "funcionario": funcionario,
        "cpf": "",
        "pis": pis,
        "ctps": "",
        "data_admissao": data_admissao,
        "cargo": cargo,
        "matricula": "",
        "departamento": departamento,
        "periodo_inicial": periodo_inicial,
        "periodo_final": periodo_final,
        "horas_normais_mes": totais[0] if totais else "",
        "total_adicional_1": totais[2] if len(totais) > 2 else "",
        "total_adicional_2": totais[5] if len(totais) > 5 else "",
        "total_adicional_3": "",
        "iob_resumo_zoopet_extra_diurna": totais[2] if len(totais) > 2 and hora_positiva(totais[2]) else "",
        "iob_resumo_zoopet_extra_noturna": totais[3] if len(totais) > 3 and hora_positiva(totais[3]) else "",
        "iob_resumo_zoopet_ausencia_diurna": totais[4] if len(totais) > 4 and hora_positiva(totais[4]) else "",
        "iob_resumo_zoopet_ausencia_noturna": totais[5] if len(totais) > 5 and hora_positiva(totais[5]) else "",
        "totais_brutos": " | ".join(totais),
    }

    return montar_resumo_base(arquivo_pdf, pagina, resumo, lancamentos)


def processar_pagina_cesul(texto: str, arquivo_pdf: Path, pagina: int) -> PaginaProcessada:
    linhas = listar_linhas(texto)
    periodo_inicial, periodo_final = extrair_periodo_generico(texto)
    rotulos_cesul = {
        "Unidade de negócio",
        "CNPJ",
        "Inscrição estadual",
        "Nome",
        "PIS",
        "CPF",
        "Cargo",
        "Matrícula",
        "Data de admissão",
        "Equipe",
    }

    def valor_apos(rotulo: str) -> str:
        try:
            indice = linhas.index(rotulo)
            if len(linhas) <= indice + 1:
                return ""
            valor = linhas[indice + 1]
            return "" if valor in rotulos_cesul else valor
        except ValueError:
            return ""

    empresa = valor_apos("Unidade de negócio")
    cnpj = valor_apos("CNPJ")
    funcionario = valor_apos("Nome")
    pis = valor_apos("PIS")
    cpf = valor_apos("CPF")
    cargo = valor_apos("Cargo")
    matricula = valor_apos("Matrícula")
    data_admissao = valor_apos("Data de admissão")
    departamento = valor_apos("Equipe")

    lancamentos = []
    totais = []
    try:
        inicio = next(i for i, linha in enumerate(linhas) if PADRAO_DATA_CESUL.match(linha))
        fim = linhas.index("TOTAIS")
    except StopIteration:
        inicio = -1
        fim = -1
    except ValueError:
        inicio = -1
        fim = -1

    if inicio >= 0 and fim > inicio:
        i = inicio
        while i < fim:
            correspondencia = PADRAO_DATA_CESUL.match(linhas[i])
            if not correspondencia:
                i += 1
                continue

            dia_semana = correspondencia.group(1).upper()[:3]
            data_completa = correspondencia.group(2)
            i += 1

            bloco = []
            while i < fim and not PADRAO_DATA_CESUL.match(linhas[i]):
                bloco.append(linhas[i])
                i += 1

            tempos = [item for item in bloco if eh_hora_texto(item)]
            textos = [item for item in bloco if not eh_hora_texto(item)]

            registros = tempos[0:4]
            status = "Trabalhado"
            if textos:
                status = textos[0]
            elif all(item == "00:00" for item in registros):
                status = "Sem registro"

            observacoes = []
            if len(tempos) >= 14:
                observacoes.append(f"Intervalo={tempos[6]}")
                observacoes.append(f"Normais={tempos[7]}")
                observacoes.append(f"HE50={tempos[8]}")
                observacoes.append(f"HE100={tempos[9]}")
                observacoes.append(f"AdicNot={tempos[10]}")
                observacoes.append(f"HNRed={tempos[11]}")
                observacoes.append(f"Saldo={tempos[13]}")
            if len(textos) > 1:
                observacoes.extend(textos[1:])

            lancamentos.append(
                {
                    "data": data_completa,
                    "dia_semana": dia_semana,
                    "previsto": "",
                    "status": status,
                    "registro_1": registros[0] if len(registros) > 0 and registros[0] != "00:00" else "",
                    "registro_2": registros[1] if len(registros) > 1 and registros[1] != "00:00" else "",
                    "registro_3": registros[2] if len(registros) > 2 and registros[2] != "00:00" else "",
                    "registro_4": registros[3] if len(registros) > 3 and registros[3] != "00:00" else "",
                    "registro_5": "",
                    "registro_6": "",
                    "horas_total_dia": tempos[12] if len(tempos) > 12 and hora_positiva(tempos[12]) else "",
                    "adicional_1": tempos[8] if len(tempos) > 8 and hora_positiva(tempos[8]) else "",
                    "adicional_2": tempos[9] if len(tempos) > 9 and hora_positiva(tempos[9]) else "",
                    "adicional_3": "",
                    "qtd_batidas_validas": len([registro for registro in registros if registro != "00:00"]),
                    "tem_inclusao_manual": "Nao",
                    "tem_pre_assinalado": "Nao",
                    "observacoes": " | ".join(observacoes),
                    "iob_extra_50": "",
                    "iob_extra_diurna": "",
                    "iob_extra_noturna": "",
                    "iob_falta_atraso": "",
                    "iob_zoopet_extra_diurna": "",
                    "iob_zoopet_extra_noturna": "",
                    "iob_zoopet_ausencia_diurna": "",
                    "iob_zoopet_ausencia_noturna": "",
                    "iob_cesul_he50": tempos[8] if len(tempos) > 8 and hora_positiva(tempos[8]) else "",
                    "iob_cesul_he100": tempos[9] if len(tempos) > 9 and hora_positiva(tempos[9]) else "",
                    "iob_cesul_adic_noturno": tempos[10] if len(tempos) > 10 and hora_positiva(tempos[10]) else "",
                    "iob_cesul_hn_reduzida": tempos[11] if len(tempos) > 11 and hora_positiva(tempos[11]) else "",
                }
            )

        totais = [item for item in linhas[fim + 1 :] if PADRAO_HORA.match(item)]

    resumo = {
        "arquivo_pdf": arquivo_pdf.name,
        "pagina_pdf": pagina,
        "layout_origem": "cesul",
        "empresa": empresa,
        "cnpj": cnpj,
        "inscricao_estadual": "",
        "funcionario": funcionario,
        "cpf": cpf,
        "pis": pis,
        "ctps": "",
        "data_admissao": data_admissao,
        "cargo": cargo,
        "matricula": matricula,
        "departamento": departamento,
        "periodo_inicial": periodo_inicial,
        "periodo_final": periodo_final,
        "horas_normais_mes": totais[3] if len(totais) > 3 else (totais[0] if totais else ""),
        "total_adicional_1": totais[4] if len(totais) > 4 else "",
        "total_adicional_2": totais[5] if len(totais) > 5 else "",
        "total_adicional_3": totais[8] if len(totais) > 8 else "",
        "iob_resumo_cesul_he50": totais[4] if len(totais) > 4 and hora_positiva(totais[4]) else "",
        "iob_resumo_cesul_he100": totais[5] if len(totais) > 5 and hora_positiva(totais[5]) else "",
        "iob_resumo_cesul_adic_noturno": totais[6] if len(totais) > 6 and hora_positiva(totais[6]) else "",
        "iob_resumo_cesul_hn_reduzida": totais[7] if len(totais) > 7 and hora_positiva(totais[7]) else "",
        "totais_brutos": " | ".join(totais),
    }

    return montar_resumo_base(arquivo_pdf, pagina, resumo, lancamentos)


def processar_pagina(texto: str, arquivo_pdf: Path, pagina: int) -> PaginaProcessada:
    texto = normalizar_texto(texto)
    if "Relatório de freqüência individual" in texto or "Relatório de frequência individual" in texto:
        return processar_pagina_zoopet(texto, arquivo_pdf, pagina)
    if "Espelho de Ponto Eletrônico" in texto:
        return processar_pagina_cesul(texto, arquivo_pdf, pagina)
    return processar_pagina_faenello(texto, arquivo_pdf, pagina)


def processar_pdf(caminho_pdf: Path) -> tuple[list[dict], list[dict]]:
    reader = PdfReader(str(caminho_pdf))
    resumos = []
    lancamentos = []

    for indice, pagina in enumerate(reader.pages, start=1):
        texto = pagina.extract_text() or ""
        if not texto.strip():
            continue
        pagina_processada = processar_pagina(texto, caminho_pdf, indice)
        resumos.append(pagina_processada.resumo)
        lancamentos.extend(pagina_processada.lancamentos)

    return resumos, lancamentos


def montar_resumo_final(resumos: list[dict], lancamentos: list[dict]) -> list[dict]:
    por_funcionario = defaultdict(list)
    for linha in lancamentos:
        chave = (linha["arquivo_pdf"], linha["pagina_pdf"], linha["matricula"])
        por_funcionario[chave].append(linha)

    saida = []
    for resumo in resumos:
        chave = (resumo["arquivo_pdf"], resumo["pagina_pdf"], resumo["matricula"])
        dias = por_funcionario.get(chave, [])
        registro = dict(resumo)
        registro["dias_trabalhados"] = sum(1 for dia in dias if dia["status"] == "Trabalhado")
        registro["dias_folga"] = sum(1 for dia in dias if dia["status"] == "Folga")
        registro["dias_sem_escala"] = sum(1 for dia in dias if dia["status"] == "-")
        registro["dias_com_falta"] = sum(1 for dia in dias if dia["status"] == "Falta parcial")
        registro["dias_com_inclusao_manual"] = sum(1 for dia in dias if dia["tem_inclusao_manual"] == "Sim")
        registro["dias_com_pre_assinalado"] = sum(1 for dia in dias if dia["tem_pre_assinalado"] == "Sim")
        registro["dias_com_adicionais"] = sum(
            1 for dia in dias if any(dia[chave_adicional] for chave_adicional in ("adicional_1", "adicional_2", "adicional_3"))
        )
        saida.append(registro)

    saida.sort(key=lambda item: (item["empresa"], item["funcionario"]))
    return saida


def montar_ocorrencias(lancamentos: list[dict]) -> list[dict]:
    ocorrencias = []
    for linha in lancamentos:
        motivos = []
        if linha["status"] in {"Folga", "-", "Falta parcial"}:
            motivos.append(linha["status"])
        if linha["tem_inclusao_manual"] == "Sim":
            motivos.append("Inclusao manual")
        if linha["tem_pre_assinalado"] == "Sim":
            motivos.append("Pre-assinalado")
        if any(linha[chave] for chave in ("adicional_1", "adicional_2", "adicional_3")):
            motivos.append("Horas adicionais")
        if linha["observacoes"]:
            motivos.append("Observacao no espelho")

        if not motivos:
            continue

        ocorrencia = dict(linha)
        ocorrencia["motivos"] = " | ".join(motivos)
        ocorrencias.append(ocorrencia)

    ocorrencias.sort(key=lambda item: (item["funcionario"], converter_data(item["data"]) or datetime.min))
    return ocorrencias


def lancamento_esta_vazio(linha: dict) -> bool:
    if linha.get("status") not in {"", "-", "Sem registro"}:
        return False
    if linha.get("previsto") not in {"", "-"}:
        return False
    if any(linha.get(chave) for chave in ("registro_1", "registro_2", "registro_3", "registro_4", "registro_5", "registro_6")):
        return False
    if any(linha.get(chave) for chave in ("horas_total_dia", "adicional_1", "adicional_2", "adicional_3")):
        return False
    if linha.get("observacoes"):
        return False
    return True


def montar_batidas(linha: dict) -> str:
    registros = [linha.get(chave, "") for chave in ("registro_1", "registro_2", "registro_3", "registro_4", "registro_5", "registro_6")]
    return " | ".join([registro for registro in registros if registro])


def montar_extras(linha: dict) -> str:
    extras = []
    if linha.get("adicional_1"):
        extras.append(f"Extra 1: {linha['adicional_1']}")
    if linha.get("adicional_2"):
        extras.append(f"Extra 2: {linha['adicional_2']}")
    if linha.get("adicional_3"):
        extras.append(f"Extra 3: {linha['adicional_3']}")
    return " | ".join(extras)


def montar_sinalizadores(linha: dict) -> str:
    flags = []
    if linha.get("tem_inclusao_manual") == "Sim":
        flags.append("Inclusao manual")
    if linha.get("tem_pre_assinalado") == "Sim":
        flags.append("Pre-assinalado")
    return " | ".join(flags)


def contar_horarios(texto: str) -> int:
    if not texto:
        return 0
    return len(re.findall(r"\d{2}:\d{2}", texto))


def listar_batidas_registradas(linha: dict) -> list[str]:
    registros = [linha.get(chave, "") for chave in ("registro_1", "registro_2", "registro_3", "registro_4", "registro_5", "registro_6")]
    return [registro for registro in registros if registro and registro != "Falta"]


def montar_faltas_batida(linha: dict) -> str:
    faltantes = []
    esperadas = min(contar_horarios(linha.get("previsto", "")), 6)
    for indice in range(1, esperadas + 1):
        valor = linha.get(f"registro_{indice}", "")
        if not valor or valor == "Falta":
            faltantes.append(MAPA_POSICOES_BATIDA[indice])
    return " | ".join(faltantes)


def linha_tem_problema(linha: dict) -> bool:
    if linha.get("status") in {"Falta parcial", "Afastamento", "Sem registro"}:
        return True
    if montar_faltas_batida(linha):
        return True
    if montar_extras(linha):
        return True
    if linha.get("observacoes"):
        return True
    if linha.get("tem_inclusao_manual") == "Sim" or linha.get("tem_pre_assinalado") == "Sim":
        return True
    return False


def obter_gravidade_painel(linha: dict) -> tuple[int, str, str]:
    faltas = montar_faltas_batida(linha)
    extras = montar_extras(linha)
    observacoes = linha.get("observacoes", "")
    status = linha.get("status", "")

    if faltas:
        return 1, "Alta", "Vermelho"
    if status in {"Falta parcial", "Afastamento", "Sem registro"}:
        return 2, "Media", "Azul"
    if observacoes:
        return 3, "Media", "Amarelo"
    if extras:
        return 4, "Baixa", "Verde"
    return 5, "Baixa", ""


def preparar_painel_visual(linha: dict) -> dict:
    batidas = montar_batidas(linha)
    faltas = montar_faltas_batida(linha)
    extras = linha.get("extras", "")
    observacoes = linha["observacoes"]
    ordem_gravidade, gravidade, cor_problema = obter_gravidade_painel(linha)
    tipo_problema = []
    if faltas:
        tipo_problema.append("Falta")
    if extras:
        tipo_problema.append("Extra")
    if observacoes:
        tipo_problema.append("Observacao")
    if linha["status"] in {"Falta parcial", "Afastamento", "Sem registro"}:
        tipo_problema.append("Status")
    return {
        "funcionario": linha["funcionario"],
        "matricula": linha["matricula"],
        "data": linha["data"],
        "dia_semana": linha["dia_semana"],
        "status": linha["status"],
        "ordem_gravidade": ordem_gravidade,
        "gravidade": gravidade,
        "cor_problema": cor_problema,
        "tipo_problema": " | ".join(tipo_problema),
        "bateu": batidas,
        "faltas_batida": faltas,
        "horas_total": linha["horas_total_dia"],
        "extras": extras,
        "observacoes": observacoes,
    }


def preparar_resumo_visual(linha: dict) -> dict:
    return {
        "empresa": linha["empresa"],
        "funcionario": linha["funcionario"],
        "matricula": linha["matricula"],
        "cargo": linha["cargo"],
        "departamento": linha["departamento"],
        "periodo": f"{linha['periodo_inicial']} a {linha['periodo_final']}".strip(" a "),
        "admissao": linha["data_admissao"],
        "horas_normais": linha["horas_normais_mes"],
        "extras": " | ".join([valor for valor in (linha["total_adicional_1"], linha["total_adicional_2"], linha["total_adicional_3"]) if valor]),
        "dias_trabalhados": linha["dias_trabalhados"],
        "dias_folga": linha["dias_folga"],
        "dias_falta": linha["dias_com_falta"],
        "dias_adicionais": linha["dias_com_adicionais"],
        "cpf": linha["cpf"],
        "pis": linha["pis"],
    }


def preparar_lancamento_visual(linha: dict) -> dict:
    return {
        "funcionario": linha["funcionario"],
        "matricula": linha["matricula"],
        "data": linha["data"],
        "dia_semana": linha["dia_semana"],
        "status": linha["status"],
        "previsto": linha["previsto"],
        "batidas": montar_batidas(linha),
        "horas_total": linha["horas_total_dia"],
        "extras": montar_extras(linha),
        "sinalizadores": montar_sinalizadores(linha),
        "observacoes": linha["observacoes"],
    }


def preparar_ocorrencia_visual(linha: dict) -> dict:
    return {
        "funcionario": linha["funcionario"],
        "matricula": linha["matricula"],
        "data": linha["data"],
        "dia_semana": linha["dia_semana"],
        "motivos": linha["motivos"],
        "status": linha["status"],
        "batidas": montar_batidas(linha),
        "horas_total": linha["horas_total_dia"],
        "extras": montar_extras(linha),
        "observacoes": linha["observacoes"],
    }


def escrever_tabela(aba, nome_tabela: str, colunas: list[tuple[str, str]], linhas: list[dict], nome_aba: str = "") -> None:
    colunas_visiveis = []
    for chave, titulo in colunas:
        if chave != "matricula":
            colunas_visiveis.append((chave, titulo))
            continue
        if any((linha.get(chave, "") not in ("", None)) for linha in linhas):
            colunas_visiveis.append((chave, titulo))

    aba.append([titulo for _, titulo in colunas_visiveis])
    for linha in linhas:
        valores = []
        for chave, _ in colunas_visiveis:
            valor = linha.get(chave, "")
            if chave in {"data", "data_admissao", "data_admissao", "admissao"} and valor:
                valor = converter_data(valor)
            valores.append(valor)
        aba.append(valores)

    ultima_linha = aba.max_row
    ultima_coluna = aba.max_column
    referencia = f"A1:{aba.cell(row=ultima_linha, column=ultima_coluna).coordinate}"
    if ultima_linha >= 2:
        tabela = Table(displayName=nome_tabela, ref=referencia)
        estilo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tabela.tableStyleInfo = estilo
        aba.add_table(tabela)
    else:
        aba.auto_filter.ref = referencia

    cabecalho_fill = PatternFill("solid", fgColor="1F4E78")
    cabecalho_fonte = Font(color="FFFFFF", bold=True)

    for celula in aba[1]:
        celula.fill = cabecalho_fill
        celula.font = cabecalho_fonte
        celula.alignment = Alignment(horizontal="center", vertical="center")

    aba.freeze_panes = "A2"
    aba.sheet_view.showGridLines = False

    for coluna in aba.columns:
        largura = max(len(str(celula.value or "")) for celula in coluna) + 2
        aba.column_dimensions[coluna[0].column_letter].width = min(max(largura, 10), 28)

    for linha in aba.iter_rows(min_row=2):
        for celula in linha:
            if isinstance(celula.value, datetime):
                celula.number_format = "dd/mm/yyyy"
            else:
                celula.alignment = Alignment(vertical="top")

    if nome_aba == "Painel":
        aplicar_cores_painel(aba)


def aplicar_cores_painel(aba) -> None:
    colunas = {aba.cell(row=1, column=col).value: col for col in range(1, aba.max_column + 1)}
    fill_falta = PatternFill("solid", fgColor="FDE9D9")
    fill_extra = PatternFill("solid", fgColor="E2F0D9")
    fill_obs = PatternFill("solid", fgColor="FFF2CC")
    fill_status = PatternFill("solid", fgColor="D9EAF7")

    def valor_coluna(row: int, titulo: str) -> str:
        indice = colunas.get(titulo)
        if not indice:
            return ""
        return aba.cell(row=row, column=indice).value or ""

    for row in range(2, aba.max_row + 1):
        status = valor_coluna(row, "Status")
        faltas_batida = valor_coluna(row, "Faltas de batida")
        extras = valor_coluna(row, "Extras")
        observacoes = valor_coluna(row, "Observacoes")
        fill_linha = None

        if faltas_batida:
            fill_linha = fill_falta
        elif extras:
            fill_linha = fill_extra
        elif observacoes:
            fill_linha = fill_obs
        elif status and status != "Trabalhado":
            fill_linha = fill_status

        if fill_linha:
            for col in range(1, aba.max_column + 1):
                aba.cell(row=row, column=col).fill = fill_linha


def preparar_faltas_batida_visual(linha: dict) -> dict:
    return {
        "funcionario": linha["funcionario"],
        "matricula": linha["matricula"],
        "data": linha["data"],
        "dia_semana": linha["dia_semana"],
        "status": linha["status"],
        "faltas_batida": montar_faltas_batida(linha),
        "horarios_batidos": montar_batidas(linha),
        "horas_total": linha["horas_total_dia"],
        "observacoes": linha["observacoes"],
    }


def somar_horas(lista_horas: list[str]) -> str:
    total_minutos = 0
    for valor in lista_horas:
        if not hora_positiva(valor):
            continue
        horas, minutos = valor.split(":")
        total_minutos += int(horas) * 60 + int(minutos)
    horas_total = total_minutos // 60
    minutos_total = total_minutos % 60
    return f"{horas_total:03d}{minutos_total:02d}"


def formatar_codigo_iob(valor: str) -> str:
    digitos = re.sub(r"\D", "", valor or "")
    return digitos.zfill(5)[-5:] if digitos else "00000"


def formatar_vinculo_iob(resumo: dict) -> str:
    if formatar_codigo_iob(resumo.get("matricula", "")) != "00000":
        return " " * 20
    base = (resumo.get("pis") or resumo.get("cpf") or resumo.get("funcionario") or "").strip()
    return base[:20].ljust(20)


def montar_linha_iob(codigo: str, evento: int, horas_hhhmm: str, vinculo: str) -> str:
    evento3 = "000"
    evento5 = "00000"
    if evento <= 999:
        evento3 = f"{evento:03d}"
    else:
        evento5 = f"{evento:05d}"
    return f"{codigo}|{evento3}|{horas_hhhmm}|00000000000|{vinculo}|{evento5}|"


def somar_minutos_horas(lista_horas: list[str]) -> int:
    total_minutos = 0
    for valor in lista_horas:
        if not hora_positiva(valor):
            continue
        horas, minutos = valor.split(":")
        total_minutos += int(horas) * 60 + int(minutos)
    return total_minutos


def coletar_totais_genericos_iob_por_funcionario(resumo: dict, lancamentos: list[dict]) -> dict[str, int]:
    mapeamentos = [
        ("geral_horas_normais", [resumo.get("horas_normais_mes", "")]),
        ("geral_extra_50", [resumo.get("iob_resumo_cesul_he50", "")] + [linha.get("iob_extra_50", "") for linha in lancamentos]),
        ("geral_extra_100", [resumo.get("iob_resumo_cesul_he100", "")]),
        ("geral_extra_diurna", [resumo.get("iob_resumo_zoopet_extra_diurna", "")] + [linha.get("iob_extra_diurna", "") for linha in lancamentos]),
        ("geral_extra_noturna", [resumo.get("iob_resumo_zoopet_extra_noturna", "")] + [linha.get("iob_extra_noturna", "") for linha in lancamentos]),
        ("geral_adicional_noturno", [resumo.get("iob_resumo_cesul_adic_noturno", "")]),
        ("geral_hora_noturna_reduzida", [resumo.get("iob_resumo_cesul_hn_reduzida", "")]),
        (
            "geral_falta_atraso",
            [resumo.get("iob_resumo_zoopet_ausencia_diurna", ""), resumo.get("iob_resumo_zoopet_ausencia_noturna", "")]
            + [linha.get("iob_falta_atraso", "") for linha in lancamentos],
        ),
    ]
    totais = {}
    for chave_evento, lista_horas in mapeamentos:
        totais[chave_evento] = somar_minutos_horas(lista_horas)
    return totais


def formatar_minutos_iob(total_minutos: int) -> str:
    if total_minutos <= 0:
        return "00000"
    horas_total = total_minutos // 60
    minutos_total = total_minutos % 60
    return f"{horas_total:03d}{minutos_total:02d}"


def agregar_eventos_iob(resumo: dict, lancamentos: list[dict], eventos_iob: dict[str, int]) -> list[tuple[int, str]]:
    minutos_por_evento = defaultdict(int)
    totais_genericos = coletar_totais_genericos_iob_por_funcionario(resumo, lancamentos)

    for chave_evento, total_minutos in totais_genericos.items():
        evento = eventos_iob.get(chave_evento, 0)
        if evento <= 0:
            continue
        if total_minutos <= 0:
            continue
        minutos_por_evento[evento] += total_minutos

    eventos = []
    for evento, total_minutos in minutos_por_evento.items():
        eventos.append((evento, formatar_minutos_iob(total_minutos)))
    return sorted(eventos, key=lambda item: item[0])


def resumir_totais_eventos_iob(resumos: list[dict], lancamentos: list[dict]) -> list[dict]:
    lancamentos_por_pagina = defaultdict(list)
    for linha in lancamentos:
        chave = (linha["arquivo_pdf"], linha["pagina_pdf"])
        lancamentos_por_pagina[chave].append(linha)

    totais_gerais = defaultdict(int)
    for resumo in resumos:
        chave = (resumo["arquivo_pdf"], resumo["pagina_pdf"])
        totais_funcionario = coletar_totais_genericos_iob_por_funcionario(resumo, lancamentos_por_pagina.get(chave, []))
        for chave_evento, total_minutos in totais_funcionario.items():
            totais_gerais[chave_evento] += total_minutos

    saida = []
    for chave_evento in ORDEM_TIPOS_EVENTO_IOB:
        total_minutos = totais_gerais.get(chave_evento, 0)
        saida.append(
            {
                "chave": chave_evento,
                "descricao": TIPOS_EVENTO_IOB.get(chave_evento, chave_evento),
                "minutos": total_minutos,
                "hhhmm": formatar_minutos_iob(total_minutos),
                "possui_valor": total_minutos > 0,
            }
        )
    return saida


def normalizar_eventos_iob(eventos_iob: dict[str, int] | None = None) -> dict[str, int]:
    base = dict(EVENTOS_IOB)
    if not eventos_iob:
        return base
    for chave, valor in eventos_iob.items():
        if chave not in base:
            continue
        try:
            base[chave] = int(valor)
        except (TypeError, ValueError):
            continue
    return base


def gerar_linhas_iob(resumos: list[dict], lancamentos: list[dict], eventos_iob: dict[str, int] | None = None) -> list[str]:
    eventos_iob_normalizados = normalizar_eventos_iob(eventos_iob)
    lancamentos_por_pagina = defaultdict(list)
    for linha in lancamentos:
        chave = (linha["arquivo_pdf"], linha["pagina_pdf"])
        lancamentos_por_pagina[chave].append(linha)

    registros = []
    for resumo in resumos:
        chave = (resumo["arquivo_pdf"], resumo["pagina_pdf"])
        eventos = agregar_eventos_iob(resumo, lancamentos_por_pagina.get(chave, []), eventos_iob_normalizados)
        codigo = formatar_codigo_iob(resumo.get("matricula", ""))
        vinculo = formatar_vinculo_iob(resumo)
        for evento, horas in eventos:
            registros.append({"codigo": codigo, "vinculo": vinculo, "linha": montar_linha_iob(codigo, evento, horas, vinculo)})

    registros.sort(key=lambda item: (item["codigo"], item["vinculo"]))
    return [item["linha"] for item in registros]


def gerar_txt_iob(
    resumos: list[dict],
    lancamentos: list[dict],
    saida: Path,
    eventos_iob: dict[str, int] | None = None,
) -> int:
    linhas = gerar_linhas_iob(resumos, lancamentos, eventos_iob)
    saida.write_text("\n".join(linhas), encoding="utf-8")
    return len(linhas)


def gerar_excel(resumos: list[dict], lancamentos: list[dict], saida: Path) -> None:
    lancamentos_filtrados = [linha for linha in lancamentos if not lancamento_esta_vazio(linha)]
    resumo_final = montar_resumo_final(resumos, lancamentos_filtrados)
    ocorrencias = montar_ocorrencias(lancamentos_filtrados)
    resumo_visual = [preparar_resumo_visual(linha) for linha in resumo_final]

    workbook = Workbook()
    aba_padrao = workbook.active
    workbook.remove(aba_padrao)

    aba_painel = workbook.create_sheet("Painel", 0)
    colunas_painel = [
        ("funcionario", "Funcionario"),
        ("matricula", "Matricula"),
        ("data", "Data"),
        ("dia_semana", "Dia"),
        ("status", "Status"),
        ("gravidade", "Gravidade"),
        ("cor_problema", "Cor do problema"),
        ("tipo_problema", "Tipo do problema"),
        ("bateu", "Horarios batidos"),
        ("faltas_batida", "Faltas de batida"),
        ("horas_total", "Horas do dia"),
        ("extras", "Extras"),
        ("observacoes", "Observacoes"),
    ]
    painel_base = [linha for linha in lancamentos_filtrados if linha_tem_problema(linha)]
    painel_visual = [preparar_painel_visual({**linha, "extras": montar_extras(linha)}) for linha in painel_base]
    painel_visual = sorted(
        painel_visual,
        key=lambda item: (
            item["funcionario"],
            converter_data(item["data"]) or datetime.min,
            item["ordem_gravidade"],
        ),
    )
    escrever_tabela(aba_painel, "TabelaPainel", colunas_painel, painel_visual, nome_aba="Painel")

    aba_resumo = workbook.create_sheet("Resumo")
    colunas_resumo = [
        ("empresa", "Empresa"),
        ("funcionario", "Funcionario"),
        ("matricula", "Matricula"),
        ("cargo", "Cargo"),
        ("departamento", "Departamento"),
        ("periodo", "Periodo"),
        ("admissao", "Admissao"),
        ("horas_normais", "Horas normais"),
        ("extras", "Totais extras"),
        ("dias_trabalhados", "Dias trabalhados"),
        ("dias_folga", "Folgas"),
        ("dias_falta", "Faltas"),
        ("dias_adicionais", "Dias com extras"),
        ("cpf", "CPF"),
        ("pis", "PIS"),
    ]
    escrever_tabela(aba_resumo, "TabelaResumo", colunas_resumo, resumo_visual, nome_aba="Resumo")

    aba_lancamentos = workbook.create_sheet("Lancamentos")
    colunas_lancamentos = [
        ("funcionario", "Funcionario"),
        ("matricula", "Matricula"),
        ("data", "Data"),
        ("dia_semana", "Dia"),
        ("status", "Status"),
        ("previsto", "Jornada prevista"),
        ("batidas", "Batidas"),
        ("horas_total", "Horas do dia"),
        ("extras", "Extras"),
        ("sinalizadores", "Sinalizadores"),
        ("observacoes", "Observacoes"),
    ]
    lancamentos_ordenados = sorted(
        lancamentos_filtrados,
        key=lambda item: (item["funcionario"], converter_data(item["data"]) or datetime.min),
    )
    lancamentos_visuais = [preparar_lancamento_visual(linha) for linha in lancamentos_ordenados]
    escrever_tabela(aba_lancamentos, "TabelaLancamentos", colunas_lancamentos, lancamentos_visuais, nome_aba="Lancamentos")

    aba_ocorrencias = workbook.create_sheet("Ocorrencias")
    colunas_ocorrencias = [
        ("funcionario", "Funcionario"),
        ("matricula", "Matricula"),
        ("data", "Data"),
        ("dia_semana", "Dia"),
        ("motivos", "Motivos"),
        ("status", "Status"),
        ("batidas", "Batidas"),
        ("horas_total", "Horas do dia"),
        ("extras", "Extras"),
        ("observacoes", "Observacoes"),
    ]
    ocorrencias_visuais = [preparar_ocorrencia_visual(linha) for linha in ocorrencias]
    escrever_tabela(aba_ocorrencias, "TabelaOcorrencias", colunas_ocorrencias, ocorrencias_visuais, nome_aba="Ocorrencias")

    aba_faltas = workbook.create_sheet("Faltas de batida")
    colunas_faltas = [
        ("funcionario", "Funcionario"),
        ("matricula", "Matricula"),
        ("data", "Data"),
        ("dia_semana", "Dia"),
        ("status", "Status"),
        ("faltas_batida", "Faltas de batida"),
        ("horarios_batidos", "Horarios batidos"),
        ("horas_total", "Horas do dia"),
        ("observacoes", "Observacoes"),
    ]
    faltas_base = [linha for linha in lancamentos_filtrados if montar_faltas_batida(linha)]
    faltas_base = sorted(
        faltas_base,
        key=lambda item: (item["funcionario"], converter_data(item["data"]) or datetime.min),
    )
    faltas_visuais = [preparar_faltas_batida_visual(linha) for linha in faltas_base]
    escrever_tabela(aba_faltas, "TabelaFaltasBatida", colunas_faltas, faltas_visuais, nome_aba="Faltas de batida")

    workbook.save(saida)


def normalizar_nome_arquivo(nome: str) -> str:
    nome_limpo = Path(nome).stem
    nome_limpo = re.sub(r'[<>:"/\\\\|?*]+', "_", nome_limpo)
    nome_limpo = re.sub(r"\s+", " ", nome_limpo).strip()
    return nome_limpo


def montar_nome_saida(resumos: list[dict], arquivo_pdf: Path) -> str:
    base_pdf = normalizar_nome_arquivo(arquivo_pdf.name)
    if not resumos:
        return base_pdf

    resumo = resumos[0]
    empresa = normalizar_nome_arquivo(resumo.get("empresa", "") or base_pdf)
    periodo = resumo.get("periodo_final") or resumo.get("periodo_inicial") or ""

    if re.match(r"^\d{2}/\d{2}/\d{4}$", periodo):
        dt = datetime.strptime(periodo, "%d/%m/%Y")
        return f"{empresa} - {dt.month:02d}-{dt.year}"

    return empresa


def ler_argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", action="store_true", help="Gera tambem o Excel de conferencia.")
    return parser.parse_args()


def main() -> None:
    args = ler_argumentos()
    arquivos_pdf = sorted(
        {arquivo for arquivo in PASTA_EXEMPLO.iterdir() if arquivo.is_file() and arquivo.suffix.lower() == ".pdf"},
        key=lambda item: item.name.lower(),
    )
    if not arquivos_pdf:
        raise FileNotFoundError(f"Nenhum PDF encontrado em: {PASTA_EXEMPLO}")

    PASTA_ARQUIVOS_GERADOS.mkdir(exist_ok=True)
    PASTA_ARQUIVOS_IOB.mkdir(exist_ok=True)
    total_funcionarios = 0
    total_lancamentos = 0
    total_registros_iob = 0

    for arquivo_pdf in arquivos_pdf:
        resumo_pdf, lancamentos_pdf = processar_pdf(arquivo_pdf)
        nome_base = montar_nome_saida(resumo_pdf, arquivo_pdf)
        caminho_iob = PASTA_ARQUIVOS_IOB / f"{nome_base}.txt"
        qtd_registros_iob = gerar_txt_iob(resumo_pdf, lancamentos_pdf, caminho_iob)
        print(f"TXT IOB gerado com sucesso em: {caminho_iob}")
        print(f"Registros IOB gerados neste arquivo: {qtd_registros_iob}")
        total_registros_iob += qtd_registros_iob

        if args.excel:
            caminho_saida = PASTA_ARQUIVOS_GERADOS / f"{nome_base}.xlsx"
            gerar_excel(resumo_pdf, lancamentos_pdf, caminho_saida)
            print(f"Excel gerado com sucesso em: {caminho_saida}")

        total_funcionarios += len(resumo_pdf)
        total_lancamentos += len(lancamentos_pdf)
        print(f"Funcionarios processados neste arquivo: {len(resumo_pdf)}")
        print(f"Lancamentos diarios processados neste arquivo: {len(lancamentos_pdf)}")

    print(f"Total de arquivos TXT IOB gerados: {len(arquivos_pdf)}")
    print(f"Total de registros IOB gerados: {total_registros_iob}")
    print(f"Total de funcionarios processados: {total_funcionarios}")
    print(f"Total de lancamentos diarios processados: {total_lancamentos}")


if __name__ == "__main__":
    main()
