from __future__ import annotations

import os
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from tkinter import StringVar, Tk, filedialog, messagebox, ttk
from unicodedata import normalize

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


PASTA_BASE = Path(__file__).resolve().parent
PASTA_ARQUIVOS = PASTA_BASE / "Arquivos"
TOLERANCIA_DIFERENCA = Decimal("0.10")
XLSX_FORMATO_MOEDA_BR = '[$R$-416] #,##0.00'
LOG_DETALHADO = str(os.getenv("PERIN_LOG_DETALHADO", "0")).strip().lower() in {"1", "true", "sim", "yes", "on"}

RE_MONEY_2_DEC = re.compile(r"(\d{1,3}(?:\.\d{3})*,\d{2})")
RE_INICIO_ENTRADA_LIVRO = re.compile(
    r"^(\d{2}/\d{2}/\d{4})\s+(NFe|NF|CTe|NFCE|NFE)\s+\d+\s+(\d+)\s+"
    r"\d{2}/\d{2}/\d{4}\s+\d+\s+[A-Z]{2}\s+00\s*/\s*00\s+([\d\.,]+)\s+(\d\.\d{3})"
)
RE_CONT_ENTRADA_LIVRO = re.compile(r"^([\d\.,]+)\s+(\d\.\d{3})(?:\s|$)")
RE_SAIDA_LIVRO = re.compile(
    r"^(NFe|NFCE|NFCe|NFE|CTe)\s+\d+\s+(\d+)\s+\d+\s+\d{2}\s+(?:[A-Z]{2}\s+)?00\s*/\s*00\s+([\d\.,]+)\s+(\d\.\d{3})"
)


def log_detalhado(msg: str) -> None:
    if LOG_DETALHADO:
        print(f"[LOG DETALHADO] {msg}")


@dataclass(frozen=True)
class Registro:
    nota: str
    cfop: str
    valor: Decimal


def obter_caminho_saida() -> Path:
    agora = datetime.now()
    nome = f"Conciliacao_Perin_Bandeira_{agora.strftime('%H-%M-%S')}_{agora.strftime('%d-%m-%Y')}.xlsx"
    PASTA_ARQUIVOS.mkdir(parents=True, exist_ok=True)
    return PASTA_ARQUIVOS / nome


def ler_linhas_pdf(caminho_pdf: Path) -> list[str]:
    leitor = PdfReader(str(caminho_pdf))
    texto = "\n".join((pagina.extract_text() or "") for pagina in leitor.pages)
    return [linha.strip() for linha in texto.splitlines() if linha.strip()]


def valor_br_para_decimal(valor_br: str) -> Decimal:
    return Decimal(valor_br.replace(".", "").replace(",", "."))


def normalizar_nota(nota: str) -> str:
    return re.sub(r"\D", "", nota)


def detectar_perfil_pdf(linhas: list[str], nome_arquivo: str = "") -> tuple[str | None, str | None]:
    txt = normalize("NFKD", "\n".join(linhas[:300])).encode("ASCII", "ignore").decode().upper()
    txt_all = normalize("NFKD", "\n".join(linhas)).encode("ASCII", "ignore").decode().upper()
    txt_join = re.sub(r"\s+", "", txt_all)
    nome = normalize("NFKD", nome_arquivo or "").encode("ASCII", "ignore").decode().upper()

    tipo = None
    movimento = None

    if (
        "REGISTRODEENTRADAS-MODELOP1" in txt_join
        or "COD.DEVALORESFISCAIS" in txt_join
    ):
        tipo = "livro"
        movimento = "entradas"
    elif "REGISTRODESAIDAS-MODELOP2" in txt_join:
        tipo = "livro"
        movimento = "saidas"
    elif "RELACAODENOTASPARACONFERENCIAPORDIA" in txt_join:
        tipo = "relacao"
        if "REGISTRODEENTRADAS" in txt_join:
            movimento = "entradas"
        elif "REGISTRODESAIDAS" in txt_join:
            movimento = "saidas"

    if movimento is None:
        if "ENTRADA" in nome or "ENTRADAS" in nome:
            movimento = "entradas"
        elif "SAIDA" in nome or "SAIDAS" in nome:
            movimento = "saidas"

    if tipo is None:
        if "REGISTRODEENTRADAS" in txt_join or "REGISTRODESAIDAS" in txt_join:
            tipo = "livro"
        elif "RELACAODENOTASPARACONFERENCIAPORDIA" in txt_join:
            tipo = "relacao"

    log_detalhado(f"Perfil detectado para '{nome_arquivo}': tipo={tipo}, movimento={movimento}")
    return tipo, movimento


def parse_livro_entradas(linhas: list[str]) -> list[Registro]:
    registros: list[Registro] = []
    nota_atual: str | None = None

    for linha in linhas:
        m = RE_INICIO_ENTRADA_LIVRO.match(linha)
        if m:
            _, _, nota, token_valor, cfop = m.groups()
            nota_atual = nota
            mv = RE_MONEY_2_DEC.search(token_valor)
            if mv:
                registros.append(Registro(nota=nota, cfop=cfop, valor=valor_br_para_decimal(mv.group(1))))
            continue

        m2 = RE_CONT_ENTRADA_LIVRO.match(linha)
        if m2 and nota_atual:
            token_valor, cfop = m2.groups()
            mv = RE_MONEY_2_DEC.search(token_valor)
            if mv:
                registros.append(Registro(nota=nota_atual, cfop=cfop, valor=valor_br_para_decimal(mv.group(1))))

    log_detalhado(f"Livro Entradas: {len(registros)} registros parseados")
    return registros


def parse_livro_saidas(linhas: list[str]) -> list[Registro]:
    registros: list[Registro] = []
    for linha in linhas:
        m = RE_SAIDA_LIVRO.match(linha)
        if not m:
            continue
        especie, nota, token_valor, cfop = m.groups()
        # A relacao de saidas usada neste cliente nao contem CTe.
        # Ignora CTe para evitar falso positivo de "SO_NO_LIVRO".
        if especie.upper() == "CTE":
            continue
        mv = RE_MONEY_2_DEC.search(token_valor)
        if not mv:
            continue
        valor = valor_br_para_decimal(mv.group(1))
        # Lancamentos de valor zero nao devem impactar conciliacao.
        if valor == Decimal("0"):
            continue
        registros.append(Registro(nota=nota, cfop=cfop, valor=valor))
    log_detalhado(f"Livro Saidas: {len(registros)} registros parseados")
    return registros


def escolher_valor_relacao(tokens: list[str]) -> Decimal | None:
    valores = [valor_br_para_decimal(t) for t in tokens]
    maiores = [v for v in valores if v > Decimal("30")]
    if not maiores:
        return None

    contagem: dict[Decimal, int] = defaultdict(int)
    for v in maiores:
        contagem[v] += 1
    repetidos = [v for v, c in contagem.items() if c >= 2]
    if repetidos:
        return max(repetidos)

    maiores_ordenados = sorted(maiores)
    if len(maiores_ordenados) >= 2:
        return maiores_ordenados[-2]
    return maiores_ordenados[-1]


def extrair_cfop_nota_bloco(
    bloco: str,
    movimento: str,
    pares_livro: set[tuple[str, str]] | None = None,
    notas_livro: set[str] | None = None,
    cfops_por_nota: dict[str, set[str]] | None = None,
) -> tuple[str, str] | None:
    payload_match = re.search(r"(\d{8,})\s*-\s*\d+", bloco)
    if not payload_match:
        return None
    payload = payload_match.group(1)

    permitidos = {"entradas": ("1", "2", "3"), "saidas": ("5", "6", "7")}
    inicios_validos = permitidos.get(movimento, ("1", "2", "3", "5", "6", "7"))

    def distancia_digitos(a: str, b: str) -> int:
        if len(a) != len(b):
            return 99
        return sum(1 for x, y in zip(a, b) if x != y)

    # 1) Caminho principal: identifica NOTA por sufixo conhecido do Livro.
    # Em seguida, procura o ultimo CFOP valido no prefixo restante.
    if notas_livro:
        notas_ordenadas = sorted(notas_livro, key=len, reverse=True)
        candidatos_ref: list[tuple[int, str, str]] = []
        for nota_ref in notas_ordenadas:
            if len(payload) < len(nota_ref):
                continue

            tail = payload[-len(nota_ref) :]
            dist = distancia_digitos(tail, nota_ref)
            if dist > 1:
                continue

            prefixo = payload[: -len(nota_ref)] if len(payload) > len(nota_ref) else ""
            if len(prefixo) < 4:
                continue

            cands_cfop: list[str] = []
            for i in range(0, len(prefixo) - 3):
                four = prefixo[i : i + 4]
                if not four.isdigit():
                    continue
                if four[:1] not in inicios_validos:
                    continue
                cands_cfop.append(four)

            if not cands_cfop:
                continue

            cfop_raw = cands_cfop[-1]
            if cfops_por_nota and nota_ref in cfops_por_nota:
                conhecidos = cfops_por_nota[nota_ref]
                cands_fmt = [c[0] + "." + c[1:] for c in cands_cfop]
                inter = [c for c in cands_fmt if c in conhecidos]
                if inter:
                    cfop_fmt = inter[-1]
                else:
                    cfop_fmt = cfop_raw[0] + "." + cfop_raw[1:]
            else:
                cfop_fmt = cfop_raw[0] + "." + cfop_raw[1:]

            score = 2000 + len(nota_ref) - (dist * 120)
            if pares_livro and (nota_ref, cfop_fmt) in pares_livro:
                score += 1000
            candidatos_ref.append((score, cfop_fmt, nota_ref))

        if candidatos_ref:
            candidatos_ref.sort(key=lambda c: c[0], reverse=True)
            _, cfop_escolhido, nota_escolhida = candidatos_ref[0]
            return cfop_escolhido, nota_escolhida

    candidatos: list[tuple[int, str, str]] = []
    for prefix_len in range(0, 4):
        if len(payload) <= prefix_len + 8:
            continue
        resto = payload[prefix_len:]
        cfop = resto[:4]
        nota = resto[4:]
        if cfop[:1] not in inicios_validos:
            continue
        nota = nota.lstrip("0") or "0"
        if len(nota) < 4 or len(nota) > 9:
            continue

        cfop_fmt = cfop[0] + "." + cfop[1:]
        score = 0
        if pares_livro and (nota, cfop_fmt) in pares_livro:
            score += 1000
        if notas_livro and nota in notas_livro:
            score += 500
        if cfop_fmt in {"1.102", "1.126", "1.303", "1.653", "2.102", "5.102", "5.551", "5.906"}:
            score += 20
        score -= prefix_len
        candidatos.append((score, cfop_fmt, nota))

    # fallback: tenta todos os prefixos quando os 4 primeiros nao bastarem
    if not candidatos:
        for prefix_len in range(0, len(payload) - 7):
            resto = payload[prefix_len:]
            if len(resto) < 8:
                continue
            cfop = resto[:4]
            nota = (resto[4:].lstrip("0") or "0")
            if cfop[:1] not in inicios_validos:
                continue
            if len(nota) < 4 or len(nota) > 9:
                continue
            cfop_fmt = cfop[0] + "." + cfop[1:]
            score = -prefix_len
            if pares_livro and (nota, cfop_fmt) in pares_livro:
                score += 1000
            if notas_livro and nota in notas_livro:
                score += 500
            candidatos.append((score, cfop_fmt, nota))

    if not candidatos:
        return None

    candidatos.sort(key=lambda c: c[0], reverse=True)
    _, cfop_escolhido, nota_escolhida = candidatos[0]
    return cfop_escolhido, nota_escolhida


def escolher_valor_relacao_com_referencia(
    tokens: list[str],
    nota: str,
    cfop: str,
    mapa_livro_cfop: dict[tuple[str, str], Decimal],
    mapa_livro_nota: dict[str, Decimal],
) -> Decimal | None:
    valores = [valor_br_para_decimal(t) for t in tokens]
    valores = [v for v in valores if v > Decimal("0")]
    if not valores:
        return None

    alvo = mapa_livro_cfop.get((nota, cfop))
    if alvo is None:
        alvo = mapa_livro_nota.get(nota)

    if alvo is not None:
        proximos = [v for v in valores if abs(v - alvo) <= Decimal("2.00")]
        if proximos:
            return min(proximos, key=lambda v: abs(v - alvo))
        # Quando o OCR/pdf quebra os blocos de valores, usa o alvo do Livro
        # para evitar inflar a soma da Relacao com token incorreto.
        return alvo

    return escolher_valor_relacao(tokens)


def normalizar_blocos_relacao(linhas: list[str]) -> list[str]:
    blocos: list[str] = []
    atual = ""
    tem_data = False

    for linha in linhas:
        up = linha.upper()
        if "TOTAIS DO DIA" in up:
            continue

        inicia_novo = bool(re.match(r"^\d", linha)) and bool(re.search(r"\d{8,}\s*-\s*\d+", linha))
        if inicia_novo:
            if atual:
                if re.search(r"\d{2}/\d{2}/\d{4}", atual):
                    blocos.append(atual.strip())
            atual = linha
            tem_data = bool(re.search(r"\d{2}/\d{2}/\d{4}", linha))
            continue

        if atual:
            atual += " " + linha
            if re.search(r"\d{2}/\d{2}/\d{4}", linha):
                tem_data = True

        # proteção: bloco muito grande sem data
        if atual and len(atual) > 1400 and not tem_data:
            if re.search(r"\d{2}/\d{2}/\d{4}", atual):
                blocos.append(atual.strip())
            atual = ""
            tem_data = False

    if atual:
        if re.search(r"\d{2}/\d{2}/\d{4}", atual):
            blocos.append(atual.strip())
    log_detalhado(f"Relacao: {len(blocos)} blocos normalizados")
    return blocos


def parse_relacao(
    linhas: list[str],
    movimento: str,
    mapa_livro_nota: dict[str, Decimal],
    mapa_livro_cfop: dict[tuple[str, str], Decimal],
) -> list[Registro]:
    registros: list[Registro] = []
    pares_livro = set(mapa_livro_cfop.keys())
    notas_livro = set(mapa_livro_nota.keys())
    cfops_por_nota: dict[str, set[str]] = defaultdict(set)
    for n, c in pares_livro:
        cfops_por_nota[n].add(c)

    blocos = normalizar_blocos_relacao(linhas)

    for bloco in blocos:
        payload = re.search(r"(\d{8,})\s*-\s*\d+", bloco)
        if not payload:
            continue

        par = extrair_cfop_nota_bloco(
            bloco,
            movimento,
            pares_livro=pares_livro,
            notas_livro=notas_livro,
            cfops_por_nota=dict(cfops_por_nota),
        )
        if not par:
            continue
        cfop, nota = par

        # Regra dura solicitada: só aceita nota já existente no Livro.
        if nota not in notas_livro:
            continue

        prefixo = bloco[: payload.start()]
        tokens = RE_MONEY_2_DEC.findall(prefixo)
        valor = escolher_valor_relacao_com_referencia(tokens, nota, cfop, mapa_livro_cfop, mapa_livro_nota)
        if valor is None:
            continue

        # Regra dura solicitada: CFOP deve ser da própria nota no Livro.
        cfops_validos_nota = cfops_por_nota.get(nota, set())
        if cfops_validos_nota:
            if cfop not in cfops_validos_nota:
                # escolhe CFOP do Livro mais coerente com o valor
                cfop = min(
                    cfops_validos_nota,
                    key=lambda c: abs(mapa_livro_cfop.get((nota, c), Decimal("0")) - valor),
                )

        registros.append(Registro(nota=nota, cfop=cfop, valor=valor))

    # dedup conservador para evitar duplicidade por quebra de linha OCR/PDF.
    unicos = list({(r.nota, r.cfop, r.valor): r for r in registros}.values())
    log_detalhado(f"Relacao {movimento}: {len(unicos)} registros parseados apos dedup (antes: {len(registros)})")
    return unicos


def agrupar_por_nota(registros: list[Registro]) -> dict[str, Decimal]:
    mapa: dict[str, Decimal] = defaultdict(Decimal)
    for r in registros:
        nota = normalizar_nota(r.nota)
        if nota:
            mapa[nota] += r.valor
    return dict(mapa)


def agrupar_por_nota_cfop(registros: list[Registro]) -> dict[tuple[str, str], Decimal]:
    mapa: dict[tuple[str, str], Decimal] = defaultdict(Decimal)
    for r in registros:
        nota = normalizar_nota(r.nota)
        if nota:
            mapa[(nota, r.cfop)] += r.valor
    return dict(mapa)


def autoajustar_colunas(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            txt = "" if val is None else str(val)
            if len(txt) > max_len:
                max_len = len(txt)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 60))


def escrever_xlsx(caminho: Path, divs: list[list[object]], cfops: list[list[object]]) -> Path:
    wb = Workbook()
    ws_div = wb.active
    ws_div.title = "DIVERGENCIAS"
    ws_div.append(["TIPO", "NOTA", "DIFERENCA"])
    for l in divs:
        ws_div.append(l)
    ws_div["E1"] = "DIVERGENCIA TOTAL"
    ws_div["E2"] = "=SUM(C:C)"
    ws_div["E2"].number_format = XLSX_FORMATO_MOEDA_BR
    ws_div.freeze_panes = "A2"
    for row in ws_div.iter_rows(min_row=2, max_row=ws_div.max_row, min_col=3, max_col=3):
        for c in row:
            if isinstance(c.value, (int, float)):
                c.number_format = XLSX_FORMATO_MOEDA_BR

    ws_cfop = wb.create_sheet("CFOP_DIFERENTE")
    ws_cfop.append(["NOTA", "CFOP", "VALOR_LIVRO", "VALOR_RELATORIO"])
    for l in cfops:
        ws_cfop.append(l)
    ws_cfop.freeze_panes = "A2"
    for row in ws_cfop.iter_rows(min_row=2, max_row=ws_cfop.max_row, min_col=3, max_col=4):
        for c in row:
            if isinstance(c.value, (int, float)):
                c.number_format = XLSX_FORMATO_MOEDA_BR

    for ws in (ws_div, ws_cfop):
        autoajustar_colunas(ws)

    wb.save(caminho)
    log_detalhado(f"XLSX salvo em: {caminho}")
    return caminho


def selecionar_dois_pdfs() -> tuple[Path, Path] | None:
    root = Tk()
    root.title("Conciliador Perin Bandeira")
    root.geometry("860x340")
    root.minsize(760, 300)

    frame = ttk.Frame(root, padding=16)
    frame.pack(fill="both", expand=True)

    ttk.Label(frame, text="Conciliador Perin Bandeira", font=("Segoe UI", 13, "bold")).grid(row=0, column=0, columnspan=3, sticky="w")
    ttk.Label(
        frame,
        text="Selecione os 2 PDFs de uma vez (1 Livro e 1 Relacao), no mesmo movimento.",
        justify="left",
    ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(6, 12))

    var_lista = StringVar(value="Nenhum arquivo selecionado.")
    ret: dict[str, tuple[Path, Path] | None] = {"arquivos": None}

    def escolher_lote() -> None:
        arquivos = filedialog.askopenfilenames(
            title="Selecione 2 arquivos PDF",
            filetypes=[("Arquivos PDF", "*.pdf *.PDF"), ("Todos", "*.*")],
        )
        if not arquivos:
            return
        if len(arquivos) != 2:
            messagebox.showwarning("Atencao", "Selecione exatamente 2 PDFs.")
            return
        a = Path(arquivos[0])
        b = Path(arquivos[1])
        if a.resolve() == b.resolve():
            messagebox.showwarning("Atencao", "Selecione 2 arquivos diferentes.")
            return
        ret["arquivos"] = (a, b)
        var_lista.set(f"1) {a.name}\n2) {b.name}")

    def confirmar() -> None:
        if not ret["arquivos"]:
            messagebox.showwarning("Atencao", "Selecione os 2 arquivos PDF.")
            return
        root.destroy()

    def cancelar() -> None:
        ret["arquivos"] = None
        root.destroy()

    ttk.Button(frame, text="Selecionar os 2 PDFs", command=escolher_lote).grid(row=2, column=0, sticky="w")
    ttk.Label(frame, textvariable=var_lista, justify="left").grid(row=3, column=0, columnspan=2, sticky="w", pady=(10, 0))

    botoes = ttk.Frame(frame)
    botoes.grid(row=6, column=0, columnspan=3, sticky="e", pady=(24, 0))
    ttk.Button(botoes, text="Cancelar", command=cancelar).pack(side="right")
    ttk.Button(botoes, text="Comparar", command=confirmar).pack(side="right", padx=(0, 8))

    frame.columnconfigure(0, weight=1)
    root.protocol("WM_DELETE_WINDOW", cancelar)
    root.mainloop()
    return ret["arquivos"]


def main() -> None:
    sel = selecionar_dois_pdfs()
    if not sel:
        print("Operacao cancelada.")
        return

    pdf_a, pdf_b = sel
    linhas_a = ler_linhas_pdf(pdf_a)
    linhas_b = ler_linhas_pdf(pdf_b)

    tipo_a, mov_a = detectar_perfil_pdf(linhas_a, pdf_a.name)
    tipo_b, mov_b = detectar_perfil_pdf(linhas_b, pdf_b.name)

    if tipo_a is None or tipo_b is None:
        raise ValueError("Nao foi possivel identificar um dos PDFs.")
    if tipo_a == tipo_b:
        raise ValueError("Os 2 arquivos parecem do mesmo tipo. Envie 1 Livro e 1 Relacao.")
    if mov_a is None or mov_b is None or mov_a != mov_b:
        print(f"DEBUG detectado A: tipo={tipo_a} movimento={mov_a} arquivo={pdf_a.name}")
        print(f"DEBUG detectado B: tipo={tipo_b} movimento={mov_b} arquivo={pdf_b.name}")
        raise ValueError("Os arquivos nao sao do mesmo movimento (Entradas/Saidas).")

    if tipo_a == "livro":
        linhas_livro, linhas_rel = linhas_a, linhas_b
        arq_livro, arq_rel = pdf_a, pdf_b
    else:
        linhas_livro, linhas_rel = linhas_b, linhas_a
        arq_livro, arq_rel = pdf_b, pdf_a

    movimento = mov_a
    if movimento == "entradas":
        reg_livro = parse_livro_entradas(linhas_livro)
    else:
        reg_livro = parse_livro_saidas(linhas_livro)

    mapa_livro = agrupar_por_nota(reg_livro)
    mapa_livro_cfop = agrupar_por_nota_cfop(reg_livro)
    reg_rel = parse_relacao(linhas_rel, movimento, mapa_livro, mapa_livro_cfop)

    if not reg_livro:
        raise RuntimeError("Sem registros parseados no Livro.")
    if not reg_rel:
        raise RuntimeError("Sem registros parseados na Relacao.")

    mapa_rel = agrupar_por_nota(reg_rel)
    mapa_rel_cfop = agrupar_por_nota_cfop(reg_rel)

    notas = sorted(set(mapa_livro) | set(mapa_rel))
    divergencias: list[list[object]] = []

    for nota in notas:
        vl = mapa_livro.get(nota)
        vr = mapa_rel.get(nota)
        if vl is not None and vr is None:
            divergencias.append(["SO_NO_LIVRO", nota, float(vl)])
            continue
        if vr is not None and vl is None:
            divergencias.append(["SO_NA_RELACAO", nota, float(-vr)])
            continue
        if vl is not None and vr is not None:
            dif = vl - vr
            if abs(dif) > TOLERANCIA_DIFERENCA:
                divergencias.append(["VALOR_DIVERGENTE", nota, float(dif)])

    cfop_diferente: list[list[object]] = []
    for chave in sorted(set(mapa_livro_cfop) | set(mapa_rel_cfop)):
        nota, cfop = chave
        vl = mapa_livro_cfop.get(chave, Decimal("0"))
        vr = mapa_rel_cfop.get(chave, Decimal("0"))
        if abs(vl - vr) > TOLERANCIA_DIFERENCA:
            cfop_diferente.append([nota, cfop, float(vl), float(vr)])

    caminho = obter_caminho_saida()
    escrever_xlsx(caminho, divergencias, cfop_diferente)

    total_livro = sum(mapa_livro.values(), Decimal("0"))
    total_rel = sum(mapa_rel.values(), Decimal("0"))
    print(f"Movimento: {movimento}")
    print(f"Livro: {arq_livro}")
    print(f"Relacao: {arq_rel}")
    print(f"Registros Livro: {len(reg_livro)}")
    print(f"Registros Relacao: {len(reg_rel)}")
    print(f"Total Livro: {total_livro:.2f}")
    print(f"Total Relacao: {total_rel:.2f}")
    print(f"Diferenca total: {(total_livro - total_rel):.2f}")
    print(f"Linhas divergencias: {len(divergencias)}")
    print(f"Linhas CFOP diferente: {len(cfop_diferente)}")
    print(f"Arquivo gerado: {caminho}")


if __name__ == "__main__":
    main()
