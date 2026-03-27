import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

import pandas as pd
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

MATCH_TEXT = (
    "Contratacao de Emprestimo Consignado no ambito do Programa Credito do Trabalhador."
)
MATCH_TEXT_NORM = ""

BASE_ROOT = Path("W:/DOCUMENTOS ESCRITORIO/RH/DET")
EXTRAS_ROOT = Path("W:/DOCUMENTOS ESCRITORIO/RH/AUTOMATIZADO/EMPR\u00c9STIMO")
OUTPUT_ROOT = Path(
    "W:/DOCUMENTOS ESCRITORIO/INSTALACAO SISTEMA/python/PROJETO RH LEONARDO/3 - DET/Comparador Parcelas Excels/Comparados"
)
LISTA_ROOT = Path("W:/DOCUMENTOS ESCRITORIO/RH/AUTOMATIZADO/EMPR\u00c9STIMO")
SCRIPT_DIR = Path(__file__).resolve().parent


def checked_log_path_for_ref(year: int, month: int) -> Path:
    return SCRIPT_DIR.parent / f"Empresas checadas - {year}-{month:02d}.xlsx"


def normalize_doc(value: str) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    return re.sub(r"\D", "", text)


def normalize_match_text(value: str) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Za-z0-9 ]", " ", text)
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def read_base_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() in {".xlsx", ".xls"}:
        return pd.read_excel(path, header=None, dtype=str)
    for enc in ("utf-8-sig", "utf-8"):
        try:
            return pd.read_csv(
                path,
                header=None,
                dtype=str,
                sep=None,
                engine="python",
                encoding=enc,
            )
        except UnicodeDecodeError:
            continue
    return pd.read_csv(
        path,
        header=None,
        dtype=str,
        sep=None,
        engine="python",
        encoding="latin1",
    )


def read_csv_structured(path: Path) -> pd.DataFrame:
    for enc in ("utf-8-sig", "utf-8"):
        try:
            return pd.read_csv(
                path,
                sep=";",
                dtype=str,
                encoding=enc,
            )
        except UnicodeDecodeError:
            continue
    return pd.read_csv(
        path,
        sep=";",
        dtype=str,
        encoding="latin1",
        engine="python",
    )


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def _find_header_indices(values: list[str]) -> dict[str, int | None]:
    normalized = [normalize_match_text(v) for v in values]

    def find_index(keywords: list[str]) -> int | None:
        for i, val in enumerate(normalized):
            for kw in keywords:
                if kw and kw in val:
                    return i
        return None

    return {
        "cnpj": find_index(["cnpj", "cpf", "documento"]),
        "nome": find_index(["empresa", "razao", "nome"]),
        "desc": find_index(["titulo", "assunto", "descricao", "descr"]),
        "data": find_index(["data", "dt"]),
    }


def collect_base_rows(path: Path) -> list[tuple[str, str, str, str]]:
    df = read_base_table(path)
    if df.empty:
        return []

    header_values = [str(value) for value in df.iloc[0].fillna("").tolist()]
    indices = _find_header_indices(header_values)
    has_header = indices["cnpj"] is not None and (
        indices["desc"] is not None or indices["data"] is not None or indices["nome"] is not None
    )

    cnpj_idx = indices["cnpj"] if indices["cnpj"] is not None else 0
    nome_idx = indices["nome"] if indices["nome"] is not None else 1
    desc_idx = indices["desc"] if indices["desc"] is not None else 3
    data_idx = indices["data"] if indices["data"] is not None else 2

    rows = []
    seen = set()
    start_row = 1 if has_header else 0
    for row in df.iloc[start_row:].itertuples(index=False):
        cnpj_raw = _clean_cell(row[cnpj_idx]) if len(row) > cnpj_idx else ""
        nome = _clean_cell(row[nome_idx]) if len(row) > nome_idx else ""
        data_contrato = _clean_cell(row[data_idx]) if len(row) > data_idx else ""
        desc = _clean_cell(row[desc_idx]) if len(row) > desc_idx else ""

        if not cnpj_raw and not nome:
            continue
        if normalize_match_text(desc) != MATCH_TEXT_NORM:
            continue

        cnpj_norm = normalize_doc(cnpj_raw)
        if not cnpj_norm or cnpj_norm in seen:
            continue

        seen.add(cnpj_norm)
        rows.append((cnpj_raw, nome, data_contrato, cnpj_norm))
    return rows


def find_column(columns: list[str], candidates: list[str]) -> str | None:
    normalized = {col.lower(): col for col in columns}
    for candidate in candidates:
        if candidate.lower() in normalized:
            return normalized[candidate.lower()]
    return None


def collect_extra_cnpjs(paths: list[Path]) -> set[str]:
    cnpjs = set()
    for path in paths:
        df = read_csv_structured(path)
        if df.empty:
            continue
        columns = [str(col).strip() for col in df.columns]
        cnpj_cols = []
        for candidate in [
            "numeroInscricaoEstabelecimento",
            "numeroInscricaoEmpregador",
            "cnpj",
            "cnpjEmpregador",
        ]:
            col = find_column(columns, [candidate])
            if col:
                cnpj_cols.append(col)
        if not cnpj_cols:
            continue
        for col in cnpj_cols:
            for value in df[col].fillna(""):
                norm = normalize_doc(value)
                if norm:
                    cnpjs.add(norm)
    return cnpjs


def load_checked_statuses(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    wb = load_workbook(path)
    ws = wb.active
    statuses = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cnpj = normalize_doc(row[0]) if row and len(row) > 0 else ""
        status = str(row[2]).strip().lower() if row and len(row) > 2 and row[2] else ""
        if cnpj:
            statuses[cnpj] = status
    return statuses


def load_checked_rows(path: Path) -> list[tuple[str, str, str, str]]:
    if not path.exists():
        return []
    wb = load_workbook(path)
    ws = wb.active
    rows = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        cnpj_raw = str(row[0]).strip() if row and len(row) > 0 and row[0] else ""
        nome = str(row[1]).strip() if row and len(row) > 1 and row[1] else ""
        cnpj_norm = normalize_doc(cnpj_raw)
        if not cnpj_norm or cnpj_norm in seen:
            continue
        seen.add(cnpj_norm)
        rows.append((cnpj_raw, nome, "", cnpj_norm))
    return rows


def write_output(
    rows: list[tuple[str, str, str, str]],
    extras: set[str] | None,
    checked_statuses: dict[str, str],
    output_path: Path,
    year: int,
    month: int,
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparacao"
    ws.append([
        "CNPJ/CPF BASE",
        "NOME BASE",
        "DATA CONTRATA\u00c7\u00c3O EMPR\u00c9STIMO",
        "BAIXOU PARCELA",
    ])

    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for cnpj_raw, nome, data_contrato, cnpj_norm in rows:
        checked_status = checked_statuses.get(cnpj_norm, "")
        if extras is None:
            if checked_status == "checado_ok":
                status = "OK"
            elif checked_status == "checado_sem_csv":
                status = "SEM CSV"
            elif checked_status == "checado_erro":
                status = "ERRO NO SITE DA PREFEITURA"
            else:
                status = "SEM PROCURA\u00c7\u00c3O"
        else:
            if cnpj_norm in extras:
                status = "OK"
            else:
                if checked_status == "checado_erro":
                    status = "ERRO NO SITE DA PREFEITURA"
                elif checked_status:
                    status = "SEM CSV"
                else:
                    status = "SEM PROCURA\u00c7\u00c3O"

        ws.append([cnpj_raw, nome, data_contrato, status])
        status_cell = ws.cell(row=ws.max_row, column=4)
        if status in ("OK", "SEM CSV"):
            status_cell.fill = ok_fill
        elif status in ("ERRO NO SITE DA PREFEITURA", "SEM PROCURA\u00c7\u00c3O"):
            status_cell.fill = error_fill

    wb.save(output_path)
    return save_copy_to_lista(output_path, year, month)


def save_copy_to_lista(output_path: Path, year: int, month: int) -> Path:
    dest_dir = LISTA_ROOT / str(year) / f"{month:02d}" / "Lista de Empresas"
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest_path = dest_dir / output_path.name
    shutil.copy2(output_path, dest_path)
    return dest_path


def month_dirs_for_ref(year: int, month: int) -> tuple[Path, Path]:
    base_dir = BASE_ROOT / f"{month:02d}-{year}"
    extras_dir = EXTRAS_ROOT / str(year) / f"{month:02d}"
    if not base_dir.exists():
        base_dir = BASE_ROOT
    if not extras_dir.exists():
        extras_dir = EXTRAS_ROOT
    return base_dir, extras_dir


def output_dir_for_ref(year: int, month: int) -> Path:
    dest_dir = OUTPUT_ROOT / str(year) / f"{month:02d}"
    dest_dir.mkdir(parents=True, exist_ok=True)
    return dest_dir


def default_output_name(year: int, month: int) -> str:
    return f"Comparacao Parcelas {month:02d}-{year}.xlsx"


def _parse_cli_args(argv: list[str]) -> tuple[int, list[int], bool]:
    now = datetime.now()
    year = now.year
    months = [now.month]
    use_checked_only = False

    for arg in argv[1:]:
        low = arg.strip().lower()
        if low.startswith("--ano="):
            year = int(arg.split("=", 1)[1].strip())
        elif low.startswith("--mes="):
            months = [int(arg.split("=", 1)[1].strip())]
        elif low.startswith("--meses="):
            raw = arg.split("=", 1)[1].strip()
            months = [int(x.strip()) for x in raw.split(",") if x.strip()]
        elif low == "--usar-checadas-only":
            use_checked_only = True

    months = sorted(set(months))
    if not months:
        raise SystemExit("Mes invalido.")
    for month in months:
        if month < 1 or month > 12:
            raise SystemExit("Mes invalido. Use 1 a 12.")
    if year < 2000 or year > 2100:
        raise SystemExit("Ano invalido. Use --ano=AAAA")

    return year, months, use_checked_only


def pick_default_base_file(base_dir: Path) -> str:
    candidates = []
    for pattern in ("*.xlsx", "*.xls", "*.csv"):
        candidates.extend(base_dir.glob(pattern))
    if len(candidates) == 1:
        return str(candidates[0])
    return ""


def _pick_base_file_for_cli(base_dir: Path) -> Path:
    candidates = []
    for pattern in ("*.xlsx", "*.xls", "*.csv"):
        candidates.extend(base_dir.glob(pattern))
    if not candidates and base_dir != BASE_ROOT:
        for pattern in ("*.xlsx", "*.xls", "*.csv"):
            candidates.extend(BASE_ROOT.glob(pattern))
    if not candidates:
        raise RuntimeError(f"Nenhum arquivo base encontrado em: {base_dir}")
    if len(candidates) == 1:
        return candidates[0]
    candidates = sorted(candidates, key=lambda p: p.name.lower())
    return candidates[0]


def run_comparison_for_ref(year: int, month: int, use_checked_only: bool = False) -> tuple[Path, Path]:
    base_dir, extras_dir = month_dirs_for_ref(year, month)
    base_path = _pick_base_file_for_cli(base_dir)
    checked_log_path = checked_log_path_for_ref(year, month)

    extras = None
    if not use_checked_only:
        extras = list(extras_dir.glob("*.csv")) + list(extras_dir.glob("*.CSV"))
        if not extras:
            raise RuntimeError(f"Nenhum CSV encontrado na pasta: {extras_dir}")

    output_dir = output_dir_for_ref(year, month)
    output_path = output_dir / default_output_name(year, month)
    if output_path.exists():
        output_path.unlink()

    base_rows = collect_base_rows(base_path)
    if not base_rows:
        base_rows = load_checked_rows(checked_log_path)
    if not base_rows:
        raise RuntimeError("Nenhuma empresa no arquivo base e nem no log mensal para comparacao.")

    checked_statuses = load_checked_statuses(checked_log_path)
    extra_cnpjs = None
    if extras is not None:
        extra_cnpjs = collect_extra_cnpjs(extras)
        if not extra_cnpjs:
            raise RuntimeError("Nenhum CNPJ encontrado nos CSVs para comparar.")

    lista_path = write_output(base_rows, extra_cnpjs, checked_statuses, output_path, year, month)
    return output_path, lista_path


def main() -> None:
    global MATCH_TEXT_NORM
    MATCH_TEXT_NORM = normalize_match_text(MATCH_TEXT)

    if len(sys.argv) > 1:
        year, months, use_checked_only = _parse_cli_args(sys.argv)
        for month in months:
            output_path, lista_path = run_comparison_for_ref(year, month, use_checked_only)
            print(f"[ok] comparacao gerada: {output_path}")
            print(f"[ok] copia lista: {lista_path}")
        return

    now = datetime.now()
    year_gui = now.year
    month_gui = now.month

    root = tk.Tk()
    root.title("Comparador de Parcelas")
    root.resizable(False, False)

    base_dir_default, extras_dir_default = month_dirs_for_ref(year_gui, month_gui)

    base_var = tk.StringVar(value=pick_default_base_file(base_dir_default))
    extras_var = tk.StringVar(value=str(extras_dir_default))
    output_name = default_output_name(year_gui, month_gui)
    use_checked_only_var = tk.BooleanVar(value=False)

    def choose_base_file() -> None:
        path = filedialog.askopenfilename(
            title="Selecione o arquivo base",
            initialdir=str(base_dir_default),
            filetypes=[
                ("Planilhas", "*.xlsx;*.xls;*.csv"),
                ("Todos os arquivos", "*.*"),
            ],
        )
        if path:
            base_var.set(path)

    def choose_extras_folder() -> None:
        path = filedialog.askdirectory(
            title="Selecione a pasta com os CSVs",
            initialdir=str(extras_dir_default),
        )
        if path:
            extras_var.set(path)

    def run_comparison() -> None:
        base_path_text = base_var.get().strip()
        extras_dir_text = extras_var.get().strip()
        use_checked_only = use_checked_only_var.get()

        if not base_path_text:
            messagebox.showwarning("Comparador", "Selecione o arquivo base.")
            return
        base_path = Path(base_path_text)
        if not base_path.exists() or not base_path.is_file():
            messagebox.showwarning("Comparador", "Arquivo base invalido.")
            return

        extras = None
        if not use_checked_only:
            if not extras_dir_text:
                messagebox.showwarning("Comparador", "Selecione a pasta dos CSVs.")
                return
            extras_dir = Path(extras_dir_text)
            if not extras_dir.exists() or not extras_dir.is_dir():
                messagebox.showwarning("Comparador", "Pasta de CSVs invalida.")
                return

            extras = list(extras_dir.glob("*.csv")) + list(extras_dir.glob("*.CSV"))
            if not extras:
                messagebox.showwarning("Comparador", "Nenhum CSV encontrado na pasta.")
                return

        output_dir = output_dir_for_ref(year_gui, month_gui)
        output_path = output_dir / output_name
        if output_path.exists():
            output_path.unlink()

        try:
            base_rows = collect_base_rows(base_path)
            if not base_rows:
                messagebox.showwarning(
                    "Comparador",
                    "Nenhuma empresa no arquivo base com o texto da contratacao.",
                )
                return

            checked_statuses = load_checked_statuses(checked_log_path_for_ref(year_gui, month_gui))
            extra_cnpjs = None
            if extras is not None:
                extra_cnpjs = collect_extra_cnpjs(extras)
                if not extra_cnpjs:
                    messagebox.showwarning(
                        "Comparador",
                        "Nenhum CNPJ encontrado nos CSVs para comparar.",
                    )
                    return

            lista_path = write_output(base_rows, extra_cnpjs, checked_statuses, output_path, year_gui, month_gui)
        except Exception as exc:
            messagebox.showerror("Comparador", f"Erro ao gerar comparacao: {exc}")
            return

        messagebox.showinfo(
            "Comparador",
            f"Comparacao gerada com sucesso.\nArquivo: {output_path}\nCopia: {lista_path}",
        )

    padding = {"padx": 8, "pady": 6}

    tk.Label(root, text="Arquivo base (DET):").grid(row=0, column=0, sticky="w", **padding)
    tk.Entry(root, textvariable=base_var, width=60).grid(row=0, column=1, **padding)
    tk.Button(root, text="Procurar", command=choose_base_file).grid(
        row=0, column=2, **padding
    )

    tk.Label(root, text="Pasta dos CSVs (EMPR\u00c9STIMO):").grid(
        row=1, column=0, sticky="w", **padding
    )
    extras_entry = tk.Entry(root, textvariable=extras_var, width=60)
    extras_entry.grid(row=1, column=1, **padding)
    extras_button = tk.Button(root, text="Procurar", command=choose_extras_folder)
    extras_button.grid(row=1, column=2, **padding)

    def toggle_csv_state() -> None:
        state = "disabled" if use_checked_only_var.get() else "normal"
        extras_entry.configure(state=state)
        extras_button.configure(state=state)

    tk.Checkbutton(
        root,
        text="Usar somente Empresas checadas (sem CSVs)",
        variable=use_checked_only_var,
        command=toggle_csv_state,
    ).grid(row=2, column=0, columnspan=3, sticky="w", **padding)
    toggle_csv_state()

    tk.Label(root, text=f"Sa\u00edda fixa: {output_name}").grid(
        row=3, column=0, columnspan=3, sticky="w", **padding
    )

    tk.Button(root, text="Gerar compara\u00e7\u00e3o", command=run_comparison, width=20).grid(
        row=4, column=1, pady=10
    )

    root.mainloop()


if __name__ == "__main__":
    main()
