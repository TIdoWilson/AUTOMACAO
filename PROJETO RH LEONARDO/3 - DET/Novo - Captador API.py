import argparse
import csv
import json
import ssl
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import quote_plus
from urllib.request import Request, urlopen

from playwright.sync_api import sync_playwright

from chrome_9222 import PORT, chrome_9222

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
except Exception:
    Workbook = None
    load_workbook = None

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_OUT_DIR = BASE_DIR / "Resultados" / "Captador API"
DEFAULT_EXCEL_FILE = BASE_DIR / "Empresas" / "6-13-14.xlsx"
DEFAULT_URL_TEMPLATE = (
    "https://papiecotrab.dataprev.gov.br/credito-trabalhador-empregador/v1.0.0/"
    "dados-consignacoes-empregador?codigoInscricao={codigo_inscricao}&numeroInscricao={numero_inscricao}&competencia={competencia}"
)
DEFAULT_INIT_URL = "https://servicos.mte.gov.br/empregador/#/credito-do-trabalhador/arquivo-emprestimo"

SSL_CONTEXT = ssl.create_default_context()
try:
    SSL_CONTEXT.minimum_version = ssl.TLSVersion.TLSv1_2
except AttributeError:
    SSL_CONTEXT.options |= ssl.OP_NO_SSLv2 | ssl.OP_NO_SSLv3
    SSL_CONTEXT.options |= ssl.OP_NO_TLSv1 | ssl.OP_NO_TLSv1_1


def _now_tag() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _pick(record: dict, *keys: str) -> str:
    for key in keys:
        if key in record and record[key] is not None:
            return _clean_text(record[key])
    return ""


def _records_from_payload(payload: Any) -> list[dict]:
    if isinstance(payload, list):
        return [x for x in payload if isinstance(x, dict)]
    if isinstance(payload, dict):
        if isinstance(payload.get("erros"), list):
            return []
        for key in ("data", "items", "result", "results"):
            value = payload.get(key)
            if isinstance(value, list):
                return [x for x in value if isinstance(x, dict)]
        return [payload] if any(payload.values()) else []
    return []


def _is_sem_registros_payload(payload: Any) -> bool:
    if not isinstance(payload, dict):
        return False
    erros = payload.get("erros")
    if not isinstance(erros, list):
        return False
    for item in erros:
        if not isinstance(item, dict):
            continue
        if _clean_text(item.get("codigo")).upper() == "LF":
            return True
    return False


def _extract_api_error(payload: Any) -> tuple[str, str]:
    if not isinstance(payload, dict):
        return "", ""
    erros = payload.get("erros")
    if not isinstance(erros, list) or not erros:
        return "", ""
    first = erros[0]
    if not isinstance(first, dict):
        return "", ""
    codigo = _clean_text(first.get("codigo")).upper()
    mensagem = _clean_text(first.get("mensagem"))
    return codigo, mensagem


def _parse_header_values(header_values: list[str]) -> dict[str, str]:
    headers: dict[str, str] = {}
    for raw in header_values:
        if ":" not in raw:
            continue
        k, v = raw.split(":", 1)
        k = k.strip()
        v = v.strip()
        if k:
            headers[k] = v
    return headers


def _read_values_file(path: Path) -> list[str]:
    if not path.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {path}")
    values: list[str] = []
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            value = line.strip()
            if value:
                values.append(value)
    return values


def _read_excel_values(path: Path, sheet_name: str | None, col_ref: str) -> list[str]:
    if load_workbook is None:
        raise SystemExit("openpyxl nao instalado. Instale: pip install openpyxl")
    if not path.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {path}")

    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise SystemExit(f"Aba nao encontrada no Excel: {sheet_name}")
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        values: list[str] = []
        col = col_ref.strip().upper()
        for row_idx in range(1, ws.max_row + 1):
            cell = ws[f"{col}{row_idx}"]
            raw = _clean_text(cell.value)
            if not raw:
                continue
            if raw.lower() in {"cnpj", "cpf", "inscricao", "numeroinscricao", "numero inscricao"}:
                continue
            values.append("".join(ch for ch in raw if ch.isdigit()) or raw)
        return values
    finally:
        wb.close()


def _build_urls(template: str, values: list[str], competencia: str, codigo_inscricao: str) -> list[str]:
    urls = []
    for value in values:
        numero = _clean_text(value).replace('"', "")
        built = template
        built = built.replace('"{valor}"', "{valor}")
        built = built.replace('"{numero_inscricao}"', "{numero_inscricao}")
        built = built.replace('"{competencia}"', "{competencia}")
        built = built.replace('"XXX"', "XXX")
        built = built.replace('"DATA"', "DATA")
        built = built.replace("{valor}", quote_plus(numero))
        built = built.replace("{numero_inscricao}", quote_plus(numero))
        built = built.replace("{competencia}", quote_plus(competencia))
        built = built.replace("{codigo_inscricao}", quote_plus(codigo_inscricao))
        built = built.replace("XXX", quote_plus(numero))
        built = built.replace("DATA", quote_plus(competencia))
        built = built.replace('"', "")
        urls.append(built)
    return urls


def _fetch_json(url: str, timeout_s: float, headers: dict[str, str]) -> Any:
    req_headers = {"Accept": "application/json"}
    req_headers.update(headers)
    req = Request(url=url, headers=req_headers, method="GET")
    with urlopen(req, timeout=timeout_s, context=SSL_CONTEXT) as resp:
        body = resp.read()
    text = body.decode("utf-8", errors="replace")
    return json.loads(text)


def _fetch_json_chrome(page, url: str, timeout_s: float) -> Any:
    timeout_ms = int(timeout_s * 1000)
    page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)

    text = ""
    for js_expr in (
        "() => document.body ? (document.body.innerText || '').trim() : ''",
        "() => document.documentElement ? (document.documentElement.innerText || '').trim() : ''",
    ):
        try:
            text = page.evaluate(js_expr) or ""
        except Exception:
            text = ""
        if text:
            break

    if not text:
        raise ValueError("Resposta vazia no navegador.")
    return json.loads(text)


def _parse_json_from_open_page(page) -> Any:
    text = ""
    for js_expr in (
        "() => document.body ? (document.body.innerText || '').trim() : ''",
        "() => document.documentElement ? (document.documentElement.innerText || '').trim() : ''",
    ):
        try:
            text = page.evaluate(js_expr) or ""
        except Exception:
            text = ""
        if text:
            break
    if not text:
        raise ValueError("Resposta vazia na aba.")
    return json.loads(text)


def _map_row(record: dict, url: str) -> dict[str, str]:
    return {
        "url_origem": url,
        "banco_codigo": _pick(record, "ifConcessora.codigo", "ifConcessoraCodigo"),
        "banco_nome": _pick(record, "ifConcessora.descricao", "ifConcessoraDescricao"),
        "contrato": _pick(record, "contrato"),
        "cpf": _pick(record, "cpf"),
        "matricula": _pick(record, "matricula"),
        "nome_trabalhador": _pick(record, "nomeTrabalhador"),
        "nome_empregador": _pick(record, "nomeEmpregador"),
        "inscricao_empregador_tipo": _pick(record, "inscricaoEmpregador.descricao", "inscricaoEmpregadorDescricao"),
        "inscricao_empregador_numero": _pick(record, "numeroInscricaoEmpregador", "inscricaoEmpregadorNumero"),
        "inscricao_estab_tipo": _pick(record, "inscricaoEstabelecimento.descricao", "inscricaoEstabelecimentoDescricao"),
        "inscricao_estab_numero": _pick(record, "numeroInscricaoEstabelecimento", "inscricaoEstabelecimentoNumero"),
        "competencia": _pick(record, "competencia"),
        "competencia_inicio_desconto": _pick(record, "competenciaInicioDesconto"),
        "competencia_fim_desconto": _pick(record, "competenciaFimDesconto"),
        "data_inicio_contrato": _pick(record, "dataInicioContrato"),
        "data_fim_contrato": _pick(record, "dataFimContrato"),
        "total_parcelas": _pick(record, "totalParcelas"),
        "valor_parcela": _pick(record, "valorParcela"),
        "valor_emprestimo": _pick(record, "valorEmprestimo"),
        "valor_liberado": _pick(record, "valorLiberado"),
        "qtd_pagamentos": _pick(record, "qtdPagamentos"),
        "qtd_escrituracoes": _pick(record, "qtdEscrituracoes"),
        "categoria_codigo": _pick(record, "categoriaTrabalhador.codigo", "categoriaTrabalhadorCodigo"),
        "categoria_descricao": _pick(record, "categoriaTrabalhador.descricao", "categoriaTrabalhadorDescricao"),
    }


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        with path.open("w", newline="", encoding="utf-8") as f:
            f.write("")
        return

    headers = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _collect_urls(args: argparse.Namespace) -> list[str]:
    urls = list(args.urls or [])
    competencia = args.competencia or datetime.now().strftime("%Y%m")
    codigo_inscricao = str(args.codigo_inscricao or "1").strip()

    if args.urls_file:
        urls.extend(_read_values_file(Path(args.urls_file)))
    if args.url_template and args.values_file:
        values = _read_values_file(Path(args.values_file))
        urls.extend(_build_urls(args.url_template, values, competencia, codigo_inscricao))
    if args.url_template and args.excel_file:
        values = _read_excel_values(Path(args.excel_file), args.excel_sheet, args.excel_col)
        urls.extend(_build_urls(args.url_template, values, competencia, codigo_inscricao))
    if not urls:
        raise SystemExit("Informe URLs via --urls, --urls-file ou --url-template + --values-file.")

    seen = set()
    out = []
    for url in urls:
        u = url.strip()
        if not u or u in seen:
            continue
        seen.add(u)
        out.append(u)
    if args.limit and args.limit > 0:
        out = out[: args.limit]
    return out


def _process_urls_via_tabs(
    args: argparse.Namespace,
    headers: dict[str, str],
    urls: list[str],
    raw_dir: Path,
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, Any]], list[dict[str, str]]]:
    all_rows: list[dict[str, str]] = []
    errors: list[dict[str, str]] = []
    summary: list[dict[str, Any]] = []
    pr_sem_procuracao: list[dict[str, str]] = []
    if not urls:
        return all_rows, errors, summary, pr_sem_procuracao

    target = args.init_url or "about:blank"
    with sync_playwright() as p:
        browser = chrome_9222(p, PORT)
        context = browser.contexts[0] if browser.contexts else browser.new_context()
        if headers:
            context.set_extra_http_headers(headers)
        init_page = context.new_page()
        print(f"[info] abrindo pagina inicial para bootstrap: {target}")
        try:
            init_page.goto(target, wait_until="domcontentloaded", timeout=int(args.timeout * 1000))
        except Exception as exc:
            print(f"[erro] falha ao abrir {target} durante bootstrap: {exc}")
            raise

        if args.esperar_login:
            input("[acao] navegador aberto. Pressione ENTER apos confirmar o certificado...")

        max_tabs = len(urls) if args.bootstrap_max_tabs <= 0 else min(len(urls), args.bootstrap_max_tabs)
        opened: list[tuple[int, str, Any]] = []
        print(f"[info] abrindo {max_tabs} aba(s) com os dados...")
        for idx, url in enumerate(urls[:max_tabs], start=1):
            tab = context.new_page()
            try:
                tab.goto(url, wait_until="domcontentloaded", timeout=int(args.timeout * 1000))
                opened.append((idx, url, tab))
            except Exception as exc:
                errors.append({"url": url, "erro": repr(exc)})
                print(f"[warn] falha ao abrir {url} -> {exc}")
                try:
                    tab.close()
                except Exception:
                    pass
        if max_tabs > 0 and opened:
            input("[acao] abas abertas. Pressione ENTER para extrair os JSONs...")
        total = len(opened)
        for pos, url, tab in opened:
            print(f"[{pos}/{total}] extraindo: {url}")
            try:
                payload = _parse_json_from_open_page(tab)
                err_code, err_msg = _extract_api_error(payload)
                if _is_sem_registros_payload(payload):
                    print("[ok] sem registros (LF). ignorado no resultado final.")
                    continue
                if err_code == "PR":
                    cnpj = _extract_cnpj_from_url(url)
                    pr_sem_procuracao.append({"cnpj": cnpj, "url": url, "mensagem": err_msg})
                    print(f"[warn] sem procuracao para CNPJ {cnpj}.")
                    continue
                raw_path = raw_dir / f"resposta_{pos:03d}.json"
                _write_json(raw_path, payload)
                records = _records_from_payload(payload)
                rows = [_map_row(rec, url) for rec in records]
                all_rows.extend(rows)
                summary.append({"url": url, "qtd_registros": len(rows), "raw_file": str(raw_path)})
                print(f"[ok] registros: {len(rows)}")
            except (json.JSONDecodeError, ValueError) as exc:
                errors.append({"url": url, "erro": str(exc)})
                print(f"[erro] {url} -> {exc}")
            except Exception as exc:
                errors.append({"url": url, "erro": repr(exc)})
                print(f"[erro] {url} -> {exc}")
            finally:
                try:
                    tab.close()
                except Exception:
                    pass
        try:
            init_page.close()
        except Exception:
            pass
        browser.close()
    return all_rows, errors, summary, pr_sem_procuracao


def _extract_cnpj_from_url(url: str) -> str:
    marker = "numeroInscricao="
    if marker not in url:
        return ""
    frag = url.split(marker, 1)[1]
    cnpj = frag.split("&", 1)[0]
    return cnpj.replace("%22", "").replace('"', "").strip()


def _write_excel_compacto(
    path: Path,
    rows: list[dict[str, str]],
    pr_sem_procuracao: list[dict[str, str]] | None = None,
) -> None:
    if Workbook is None:
        raise SystemExit("openpyxl nao instalado. Instale: pip install openpyxl")
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    def _auto_fit(ws, min_w: int = 10, max_w: int = 55) -> None:
        for col_cells in ws.columns:
            col_letter = None
            for probe in col_cells:
                if hasattr(probe, "column_letter"):
                    col_letter = probe.column_letter
                    break
            if not col_letter:
                continue
            max_len = 0
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 2))

    headers = [
        "CNPJ",
        "Empresa",
        "Codigo Banco",
        "Banco",
        "Contrato",
        "CPF",
        "Matricula",
        "Trabalhador",
        "Inicio Contrato",
        "Fim Contrato",
        "Inicio Desconto",
        "Fim Desconto",
        "Total Parcelas",
        "Valor Parcela",
        "Valor Emprestimo",
        "Valor Liberado",
        "Competencia",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    simple_rows: list[list[str]] = []
    if rows:
        for r in rows:
            cnpj = (
                r.get("inscricao_estab_numero")
                or r.get("inscricao_empregador_numero")
                or _extract_cnpj_from_url(r.get("url_origem", ""))
            )
            cnpj = "".join(ch for ch in str(cnpj or "") if ch.isdigit())
            simple_rows.append(
                [
                    cnpj,
                    r.get("nome_empregador", ""),
                    r.get("banco_codigo", ""),
                    r.get("banco_nome", ""),
                    r.get("contrato", ""),
                    r.get("cpf", ""),
                    r.get("matricula", ""),
                    r.get("nome_trabalhador", ""),
                    r.get("data_inicio_contrato", ""),
                    r.get("data_fim_contrato", ""),
                    r.get("competencia_inicio_desconto", ""),
                    r.get("competencia_fim_desconto", ""),
                    r.get("total_parcelas", ""),
                    r.get("valor_parcela", ""),
                    r.get("valor_emprestimo", ""),
                    r.get("valor_liberado", ""),
                    r.get("competencia", ""),
                ]
            )

        simple_rows.sort(key=lambda x: (x[1], x[0], x[7]))
        for line in simple_rows:
            ws.append(line)
    else:
        ws.append(["sem_dados", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _auto_fit(ws)

    # Aba adicional em formato de pequenas tabelas por Empresa/CNPJ.
    ws_tab = wb.create_sheet("Tabelado")
    table_headers = [
        "Trabalhador",
        "CPF",
        "Banco",
        "Contrato",
        "Valor Parcela",
        "Total Parcelas",
        "Inicio Desconto",
        "Fim Desconto",
        "Competencia",
    ]

    if simple_rows:
        groups: dict[tuple[str, str], list[list[str]]] = {}
        for row in simple_rows:
            # row indices:
            # 0 CNPJ, 1 Empresa, 2 CodBanco, 3 Banco, 4 Contrato, 5 CPF, 6 Matricula, 7 Trabalhador,
            # 8 InicioContrato, 9 FimContrato, 10 InicioDesconto, 11 FimDesconto, 12 TotalParcelas,
            # 13 ValorParcela, 14 ValorEmprestimo, 15 ValorLiberado, 16 Competencia
            key = (row[1], row[0])  # (Empresa, CNPJ)
            groups.setdefault(key, []).append(row)

        r = 1
        for (empresa, cnpj) in sorted(groups.keys(), key=lambda k: (k[0], k[1])):
            bloco = groups[(empresa, cnpj)]
            titulo = f"{empresa} | CNPJ: {cnpj}"
            ws_tab.cell(row=r, column=1, value=titulo)
            ws_tab.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(table_headers))
            title_cell = ws_tab.cell(row=r, column=1)
            title_cell.font = Font(bold=True)
            title_cell.fill = PatternFill("solid", fgColor="FCE4D6")

            r += 1
            for c, h in enumerate(table_headers, start=1):
                hc = ws_tab.cell(row=r, column=c, value=h)
                hc.font = Font(bold=True)
                hc.fill = PatternFill("solid", fgColor="D9E1F2")

            r += 1
            for row in bloco:
                ws_tab.append([
                    row[7],   # Trabalhador
                    row[5],   # CPF
                    row[3],   # Banco
                    row[4],   # Contrato
                    row[13],  # Valor Parcela
                    row[12],  # Total Parcelas
                    row[10],  # Inicio Desconto
                    row[11],  # Fim Desconto
                    row[16],  # Competencia
                ])
                r += 1

            r += 1
    else:
        ws_tab.append(["sem_dados"])

    _auto_fit(ws_tab)

    # Aba de CNPJs sem procuracao (erro PR)
    ws_pr = wb.create_sheet("Sem Procuracao")
    ws_pr.append(["CNPJ", "URL", "Mensagem"])
    for c in range(1, 4):
        hc = ws_pr.cell(row=1, column=c)
        hc.font = Font(bold=True)
        hc.fill = PatternFill("solid", fgColor="F8CBAD")

    if pr_sem_procuracao:
        for item in pr_sem_procuracao:
            ws_pr.append(
                [
                    item.get("cnpj", ""),
                    item.get("url", ""),
                    item.get("mensagem", ""),
                ]
            )
    else:
        ws_pr.append(["sem_dados", "", ""])
    ws_pr.freeze_panes = "A2"
    ws_pr.auto_filter.ref = ws_pr.dimensions
    _auto_fit(ws_pr)
    wb.save(path)


def main() -> int:
    ap = argparse.ArgumentParser(description="Captador de JSON por URL para consignacoes (API).")
    ap.add_argument("--urls", nargs="*", help="Lista de URLs.")
    ap.add_argument("--urls-file", help="TXT com 1 URL por linha.")
    ap.add_argument(
        "--url-template",
        default=DEFAULT_URL_TEMPLATE,
        help="Template com {numero_inscricao}/{competencia} (ou XXX/DATA).",
    )
    ap.add_argument("--values-file", help="TXT com valores para substituir no template.")
    ap.add_argument("--excel-file", default=str(DEFAULT_EXCEL_FILE), help="Excel com lista de inscricoes.")
    ap.add_argument("--excel-sheet", default=None, help="Nome da aba do Excel (padrao: primeira).")
    ap.add_argument("--excel-col", default="A", help="Coluna da inscricao no Excel. Ex.: A")
    ap.add_argument("--codigo-inscricao", default="1", help="codigoInscricao da URL (padrao: 1).")
    ap.add_argument(
        "--competencia",
        default=datetime.now().strftime("%Y%m"),
        help="Competencia no formato YYYYMM (padrao: mes atual).",
    )
    ap.add_argument("--header", action="append", default=[], help="Header extra. Ex.: Authorization: Bearer TOKEN")
    ap.add_argument("--timeout", type=float, default=30.0, help="Timeout em segundos por requisicao.")
    ap.add_argument("--limit", type=int, default=0, help="Limita quantidade de URLs/CNPJs processados.")
    ap.add_argument("--via-chrome", action="store_true", help="Faz as consultas via Chrome (CDP).")
    ap.add_argument("--init-url", default=DEFAULT_INIT_URL, help="Pagina para abrir antes da captura.")
    ap.add_argument("--esperar-login", action="store_true", help="Abre o navegador e aguarda ENTER para iniciar.")
    ap.add_argument(
        "--abrir-abas-primeiro",
        action="store_true",
        help="Abre todas as URLs em abas e extrai depois, mantendo abas abertas.",
    )
    ap.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR), help="Pasta de saida.")
    ap.add_argument(
        "--no-bootstrap-chrome",
        dest="bootstrap_chrome",
        action="store_false",
        default=True,
        help="Nao abre o Chrome antes do loop principal (o padrao abre para confirmar o certificado).",
    )
    ap.add_argument(
        "--bootstrap-max-tabs",
        type=int,
        default=0,
        help="Numero maximo de abas abertas no bootstrap (0 = abre todas).",
    )
    args = ap.parse_args()

    headers = _parse_header_values(args.header)
    urls = _collect_urls(args)

    run_dir = Path(args.out_dir) / _now_tag()
    raw_dir = run_dir / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)

    all_rows: list[dict[str, str]] = []
    errors: list[dict[str, str]] = []
    summary: list[dict[str, Any]] = []
    pr_sem_procuracao: list[dict[str, str]] = []

    if args.bootstrap_chrome:
        (
            all_rows,
            errors,
            summary,
            pr_sem_procuracao,
        ) = _process_urls_via_tabs(args, headers, urls, raw_dir)
    elif args.via_chrome:
        with sync_playwright() as p:
            browser = chrome_9222(p, PORT)
            context = browser.contexts[0] if browser.contexts else browser.new_context()
            if headers:
                context.set_extra_http_headers(headers)
            page = context.new_page()

            target = args.init_url or "about:blank"
            print(f"[info] abrindo pagina inicial: {target}")
            page.goto(target, wait_until="domcontentloaded", timeout=int(args.timeout * 1000))

            if args.esperar_login:
                input("[acao] navegador aberto. Pressione ENTER para iniciar a captura... ")

            if args.abrir_abas_primeiro:
                opened: list[tuple[int, str, Any]] = []
                print(f"[info] abrindo {len(urls)} URL(s) em abas...")
                for idx, url in enumerate(urls, start=1):
                    tab = page if idx == 1 else context.new_page()
                    print(f"[{idx}/{len(urls)}] abrindo aba: {url}")
                    try:
                        tab.goto(url, wait_until="domcontentloaded", timeout=int(args.timeout * 1000))
                        opened.append((idx, url, tab))
                    except Exception as exc:
                        errors.append({"url": url, "erro": repr(exc)})
                        print(f"[erro] abrir aba {url} -> {exc}")

                print("[info] extraindo JSON das abas abertas...")
                for idx, url, tab in opened:
                    print(f"[{idx}/{len(urls)}] extraindo: {url}")
                    try:
                        payload = _parse_json_from_open_page(tab)
                        err_code, err_msg = _extract_api_error(payload)
                        if _is_sem_registros_payload(payload):
                            print("[ok] sem registros (LF). ignorado no resultado final.")
                            continue
                        if err_code == "PR":
                            cnpj = _extract_cnpj_from_url(url)
                            pr_sem_procuracao.append(
                                {"cnpj": cnpj, "url": url, "mensagem": err_msg}
                            )
                            print(f"[warn] sem procuracao para CNPJ {cnpj}.")
                            continue
                        raw_path = raw_dir / f"resposta_{idx:03d}.json"
                        _write_json(raw_path, payload)
                        records = _records_from_payload(payload)
                        rows = [_map_row(rec, url) for rec in records]
                        all_rows.extend(rows)
                        summary.append({"url": url, "qtd_registros": len(rows), "raw_file": str(raw_path)})
                        print(f"[ok] registros: {len(rows)}")
                    except (json.JSONDecodeError, ValueError) as exc:
                        errors.append({"url": url, "erro": str(exc)})
                        print(f"[erro] {url} -> {exc}")
                    except Exception as exc:
                        errors.append({"url": url, "erro": repr(exc)})
                        print(f"[erro] {url} -> {exc}")
            else:
                for idx, url in enumerate(urls, start=1):
                    print(f"[{idx}/{len(urls)}] consultando: {url}")
                    try:
                        payload = _fetch_json_chrome(page, url=url, timeout_s=args.timeout)
                        err_code, err_msg = _extract_api_error(payload)
                        if _is_sem_registros_payload(payload):
                            print("[ok] sem registros (LF). ignorado no resultado final.")
                            continue
                        if err_code == "PR":
                            cnpj = _extract_cnpj_from_url(url)
                            pr_sem_procuracao.append(
                                {"cnpj": cnpj, "url": url, "mensagem": err_msg}
                            )
                            print(f"[warn] sem procuracao para CNPJ {cnpj}.")
                            continue
                        raw_path = raw_dir / f"resposta_{idx:03d}.json"
                        _write_json(raw_path, payload)
                        records = _records_from_payload(payload)
                        rows = [_map_row(rec, url) for rec in records]
                        all_rows.extend(rows)
                        summary.append({"url": url, "qtd_registros": len(rows), "raw_file": str(raw_path)})
                        print(f"[ok] registros: {len(rows)}")
                    except (json.JSONDecodeError, ValueError) as exc:
                        errors.append({"url": url, "erro": str(exc)})
                        print(f"[erro] {url} -> {exc}")
                    except Exception as exc:
                        errors.append({"url": url, "erro": repr(exc)})
                        print(f"[erro] {url} -> {exc}")
    else:
        for idx, url in enumerate(urls, start=1):
            print(f"[{idx}/{len(urls)}] consultando: {url}")
            try:
                payload = _fetch_json(url, timeout_s=args.timeout, headers=headers)
                err_code, err_msg = _extract_api_error(payload)
                if _is_sem_registros_payload(payload):
                    print("[ok] sem registros (LF). ignorado no resultado final.")
                    continue
                if err_code == "PR":
                    cnpj = _extract_cnpj_from_url(url)
                    pr_sem_procuracao.append(
                        {"cnpj": cnpj, "url": url, "mensagem": err_msg}
                    )
                    print(f"[warn] sem procuracao para CNPJ {cnpj}.")
                    continue
                raw_path = raw_dir / f"resposta_{idx:03d}.json"
                _write_json(raw_path, payload)
                records = _records_from_payload(payload)
                rows = [_map_row(rec, url) for rec in records]
                all_rows.extend(rows)
                summary.append({"url": url, "qtd_registros": len(rows), "raw_file": str(raw_path)})
                print(f"[ok] registros: {len(rows)}")
            except (HTTPError, URLError, TimeoutError, json.JSONDecodeError, ValueError) as exc:
                errors.append({"url": url, "erro": str(exc)})
                print(f"[erro] {url} -> {exc}")
            except Exception as exc:
                errors.append({"url": url, "erro": repr(exc)})
                print(f"[erro] {url} -> {exc}")

    _write_csv(run_dir / "consolidado.csv", all_rows)
    _write_json(run_dir / "consolidado.json", all_rows)
    _write_json(run_dir / "resumo.json", summary)
    _write_csv(run_dir / "erros.csv", errors)
    _write_csv(run_dir / "sem_procuracao_pr.csv", pr_sem_procuracao)
    _write_excel_compacto(run_dir / "consolidado.xlsx", all_rows, pr_sem_procuracao)

    print(f"[fim] pasta de saida: {run_dir}")
    print(f"[fim] total urls: {len(urls)} | sucesso: {len(summary)} | erros: {len(errors)}")
    print(f"[fim] total registros consolidados: {len(all_rows)}")
    print(f"[fim] total sem procuracao (PR): {len(pr_sem_procuracao)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
