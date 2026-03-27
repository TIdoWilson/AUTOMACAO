
import re
import sys
import time
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

from playwright.sync_api import TimeoutError as PWTimeout
from playwright.sync_api import sync_playwright

from chrome_9222 import PORT, chrome_9222

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None
if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

SELECAO_URL = "https://servicos.mte.gov.br/empregador/#/selecao-empresa-vinculada"
CREDITO_URL = "https://servicos.mte.gov.br/empregador/#/credito-do-trabalhador/arquivo-emprestimo"
LOGIN_URL = "https://servicos.mte.gov.br/empregador/#/login"

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_ROOT = Path(r"W:\DOCUMENTOS ESCRITORIO\RH\AUTOMATIZADO\EMPRÉSTIMO")
RESULTS_DIR = BASE_DIR / "Resultados"

WAIT_SHORT = 0.4
WAIT_MED = 1.0
WAIT_LONG = 15.0
MAX_RETRIES = 3


def _sanitize_filename(name: str) -> str:
    invalid = ["<", ">", ":", "\"", "/", "\\", "|", "?", "*"]
    for ch in invalid:
        name = name.replace(ch, "-")
    return " ".join(name.split()).strip()


def _normalize_cnpj(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def _parse_exec_args(argv: list[str]) -> tuple[int, int, bool, bool, bool]:
    now = datetime.now()
    ano_exec = now.year
    mes_exec = now.month
    forcar_todas = False
    sobrescrever = False
    resetar_log = False

    for arg in argv[1:]:
        low = arg.strip().lower()
        if low.startswith("--ano="):
            ano_exec = int(arg.split("=", 1)[1].strip())
        elif low.startswith("--mes="):
            mes_exec = int(arg.split("=", 1)[1].strip())
        elif low == "--forcar-todas":
            forcar_todas = True
        elif low == "--sobrescrever":
            sobrescrever = True
        elif low == "--resetar-log":
            resetar_log = True

    if mes_exec < 1 or mes_exec > 12:
        raise SystemExit("Mes invalido. Use --mes=1 ate --mes=12")
    if ano_exec < 2000 or ano_exec > 2100:
        raise SystemExit("Ano invalido. Use --ano=AAAA")

    return ano_exec, mes_exec, forcar_todas, sobrescrever, resetar_log


def _cnpj_regex_from_digits(digits: str) -> re.Pattern:
    if len(digits) != 14:
        return re.compile(re.escape(digits))
    pattern = (
        rf"{digits[0:2]}\D*{digits[2:5]}\D*{digits[5:8]}"
        rf"\D*{digits[8:12]}\D*{digits[12:14]}"
    )
    return re.compile(pattern)


def _cnpj_regex_any() -> re.Pattern:
    return re.compile(r"\b\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}\b")


def read_sylk_companies(path: Path, col: int = 1) -> list[dict]:
    if not path.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {path}")
    text = None
    for enc in ("latin-1", "utf-8"):
        try:
            text = path.read_text(encoding=enc, errors="strict")
            break
        except Exception:
            continue
    if text is None:
        text = path.read_text(encoding="latin-1", errors="ignore")

    cells = {}
    for line in text.splitlines():
        if not line.startswith("C;"):
            continue
        x_m = re.search(r";X(\d+)", line)
        y_m = re.search(r";Y(\d+)", line)
        if not x_m or not y_m:
            continue
        col_idx = int(x_m.group(1))
        row_idx = int(y_m.group(1))
        k_m = re.search(r';K"([^"]*)"', line)
        if not k_m:
            k_m = re.search(r";K([^;]+)", line)
        if not k_m:
            continue
        value = k_m.group(1).strip()
        if value:
            cells[(row_idx, col_idx)] = value

    rows = sorted({r for (r, _) in cells.keys()})
    companies = []
    seen = set()
    for r in rows:
        value = cells.get((r, col))
        if not value:
            continue
        lowered = value.strip().lower()
        if lowered in ("empresa", "razao social", "razao"):
            continue
        if value not in seen:
            companies.append({"row": r, "cnpj": value, "name": ""})
            seen.add(value)
    return companies


def _xlsx_read_shared_strings(z: zipfile.ZipFile, ns: dict) -> list[str]:
    shared = []
    if "xl/sharedStrings.xml" in z.namelist():
        ss = ET.fromstring(z.read("xl/sharedStrings.xml"))
        for si in ss.findall("main:si", ns):
            t = si.find("main:t", ns)
            if t is None:
                parts = [x.text or "" for x in si.findall(".//main:t", ns)]
                shared.append("".join(parts))
            else:
                shared.append(t.text or "")
    return shared


def _xlsx_find_sheet_path(z: zipfile.ZipFile) -> str:
    sheet_path = "xl/worksheets/sheet1.xml"
    if sheet_path in z.namelist():
        return sheet_path
    for name in z.namelist():
        if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
            return name
    return sheet_path


def _xlsx_cell_value(cell: ET.Element, shared: list[str], ns: dict) -> str:
    if cell is None:
        return ""
    if cell.attrib.get("t") == "inlineStr":
        t = cell.find(".//main:t", ns)
        return (t.text or "").strip() if t is not None else ""
    v = cell.find("main:v", ns)
    if v is None or (v.text or "") == "":
        return ""
    value = v.text or ""
    if cell.attrib.get("t") == "s":
        try:
            return (shared[int(value)] or "").strip()
        except Exception:
            return value.strip()
    return value.strip()


def _xlsx_parse_rows(path: Path) -> tuple[list[dict], list[str], str]:
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(path) as z:
        shared = _xlsx_read_shared_strings(z, ns)
        sheet_path = _xlsx_find_sheet_path(z)
        sheet = ET.fromstring(z.read(sheet_path))

    rows = []
    for row in sheet.findall("main:sheetData/main:row", ns):
        row_idx = int(row.attrib.get("r", "0") or "0")
        cells = {}
        for c in row.findall("main:c", ns):
            ref = c.attrib.get("r", "")
            col = re.sub(r"\d+", "", ref)
            if not col:
                continue
            cells[col] = _xlsx_cell_value(c, shared, ns)
        rows.append({"row": row_idx, "cells": cells})
    return rows, shared, sheet_path


def read_xlsx_companies(path: Path, col: int = 1) -> list[dict]:
    if not path.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {path}")
    rows, _, _ = _xlsx_parse_rows(path)
    companies = []
    seen = set()
    col_letter = chr(ord("A") + col - 1)
    for row in rows:
        cnpj = (row["cells"].get(col_letter) or "").strip()
        if not cnpj:
            continue
        lowered = cnpj.lower()
        if lowered in ("empresa", "empresas", "razao social", "razao", "cnpj"):
            continue
        status = (row["cells"].get("C") or "").strip().lower()
        if status == "fora da lista":
            continue
        name = (row["cells"].get("B") or "").strip()
        if cnpj not in seen:
            companies.append({"row": row["row"], "cnpj": cnpj, "name": name})
            seen.add(cnpj)
    return companies


def read_company_list(path: Path, col: int = 1) -> list[dict]:
    if path.suffix.lower() == ".xlsx":
        return read_xlsx_companies(path, col=col)
    return read_sylk_companies(path, col=col)


def _col_letter(col: int) -> str:
    return chr(ord("A") + col - 1)


def checked_log_path_for_ref(year: int, month: int) -> Path:
    return BASE_DIR / f"Empresas checadas - {year}-{month:02d}.xlsx"


def mark_company_status_xlsx(path: Path, row_idx: int, status: str, col: int = 2) -> None:
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    status = status.strip()
    col_letter = _col_letter(col)

    with zipfile.ZipFile(path) as z:
        shared = _xlsx_read_shared_strings(z, ns)
        sheet_path = _xlsx_find_sheet_path(z)
        sheet = ET.fromstring(z.read(sheet_path))
        shared_xml = z.read("xl/sharedStrings.xml") if "xl/sharedStrings.xml" in z.namelist() else None
        other_files = {
            name: z.read(name)
            for name in z.namelist()
            if name not in {sheet_path, "xl/sharedStrings.xml"}
        }

    if status in shared:
        status_idx = shared.index(status)
    else:
        shared.append(status)
        status_idx = len(shared) - 1

    row_el = None
    for row in sheet.findall("main:sheetData/main:row", ns):
        if int(row.attrib.get("r", "0") or "0") == row_idx:
            row_el = row
            break
    if row_el is None:
        row_el = ET.SubElement(sheet.find("main:sheetData", ns), f"{{{ns['main']}}}row", {"r": str(row_idx)})

    cell_ref = f"{col_letter}{row_idx}"
    cell_el = None
    for c in row_el.findall("main:c", ns):
        if c.attrib.get("r") == cell_ref:
            cell_el = c
            break
    if cell_el is None:
        cell_el = ET.SubElement(row_el, f"{{{ns['main']}}}c", {"r": cell_ref, "t": "s"})
    else:
        cell_el.attrib["t"] = "s"

    v = cell_el.find("main:v", ns)
    if v is None:
        v = ET.SubElement(cell_el, f"{{{ns['main']}}}v")
    v.text = str(status_idx)

    if shared_xml is None:
        sst = ET.Element(
            f"{{{ns['main']}}}sst",
            {
                "count": str(len(shared)),
                "uniqueCount": str(len(shared)),
            },
        )
    else:
        sst = ET.fromstring(shared_xml)
        sst.attrib["count"] = str(len(shared))
        sst.attrib["uniqueCount"] = str(len(shared))
        for si in sst.findall("main:si", ns):
            sst.remove(si)
    for text in shared:
        si = ET.SubElement(sst, f"{{{ns['main']}}}si")
        t = ET.SubElement(si, f"{{{ns['main']}}}t")
        t.text = text

    tmp_path = path.with_suffix(".tmp")
    with zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
        z.writestr(sheet_path, ET.tostring(sheet, encoding="utf-8", xml_declaration=True))
        z.writestr("xl/sharedStrings.xml", ET.tostring(sst, encoding="utf-8", xml_declaration=True))
        for name, data in other_files.items():
            z.writestr(name, data)
    tmp_path.replace(path)


def wait_for_login(page):
    print("[info] assumindo sessao ja logada no navegador (porta 9222).")
    page.goto(SELECAO_URL, wait_until="domcontentloaded")
    close_warning_if_present(page)
    if "/login" in page.url:
        raise RuntimeError("sessao_nao_logada")


def click_by_text(page, texts: list[str], timeout_ms: int = 8000) -> bool:
    for txt in texts:
        expr = (
            "//*[self::button or self::a or self::span or self::div]"
            f"[contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
            f" 'abcdefghijklmnopqrstuvwxyz'), '{txt.lower()}')]"
        )
        try:
            el = page.locator(expr).first
            el.wait_for(state="visible", timeout=timeout_ms)
            el.click(timeout=timeout_ms)
            return True
        except Exception:
            continue
    return False


def close_warning_if_present(page) -> bool:
    try:
        btn = page.locator("button.br-button._footerAction_slfpf_24.secondary:has-text('Fechar')").first
        if btn.is_visible():
            btn.click(timeout=2000)
            time.sleep(WAIT_SHORT)
            return True
    except Exception:
        return False
    return False


def is_unexpected_error_page(page) -> bool:
    try:
        if "erro-inesperado" in (page.url or ""):
            return True
    except Exception:
        pass
    try:
        if page.locator("text=Falha inesperada").first.is_visible():
            return True
    except Exception:
        pass
    try:
        if page.locator("text=Ocorreu um erro durante a operação").first.is_visible():
            return True
    except Exception:
        return False
    return False


def select_company(page, cnpj: str):
    page.wait_for_load_state("domcontentloaded")
    close_warning_if_present(page)
    digits = _normalize_cnpj(cnpj)
    if not digits:
        raise RuntimeError(f"CNPJ vazio: {cnpj}")

    try:
        locator = page.get_by_text(_cnpj_regex_from_digits(digits)).first
        locator.wait_for(state="visible", timeout=10000)
        locator.scroll_into_view_if_needed()
        locator.click(timeout=3000)
        try:
            page.wait_for_load_state("domcontentloaded", timeout=15000)
        except Exception:
            pass
        return
    except Exception:
        pass

    raise RuntimeError(f"Empresa nao encontrada na lista: {cnpj}")


def _pick_br_select_option(page, label_text: str, option_text: str) -> bool:
    try:
        label = page.locator(f"label:has-text('{label_text}')").first
        label.wait_for(state="visible", timeout=5000)
        br_select = label.locator("xpath=ancestor::div[contains(@class,'br-select')][1]")
        if not br_select.count():
            return False
        br_select.locator("button[aria-label='Exibir lista']").first.click(timeout=3000)
        opt = br_select.locator(f"label:has-text('{option_text}')").first
        opt.wait_for(state="visible", timeout=5000)
        opt.click(timeout=3000)
        return True
    except Exception:
        return False


def set_ano_mes(page, year: int, month: int):
    year_str = str(year)
    month_names = [
        "Janeiro",
        "Fevereiro",
        "Março",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro",
    ]
    month_name = month_names[month - 1]
    year_labels = ["Ano", "Ano-base", "Ano base"]
    month_labels = [
        "Mês",
        "Mes",
        "Competência",
        "Competencia",
        "Referência",
        "Referencia",
    ]
    if not any(_pick_br_select_option(page, lbl, year_str) for lbl in year_labels):
        raise RuntimeError("Nao consegui selecionar o ano pelo menu.")
    if not any(_pick_br_select_option(page, lbl, month_name) for lbl in month_labels):
        try:
            br_selects = page.locator("div.br-select")
            for i in range(br_selects.count()):
                bs = br_selects.nth(i)
                try:
                    bs.locator("button[aria-label='Exibir lista']").first.click(timeout=2000)
                    opt = bs.locator(f"label:has-text('{month_name}')").first
                    if opt.is_visible():
                        opt.click(timeout=2000)
                        return
                except Exception:
                    continue
        except Exception:
            pass
        raise RuntimeError("Nao consegui selecionar o mes pelo menu.")

def get_unique_path(dest: Path) -> Path:
    if not dest.exists():
        return dest
    i = 2
    while True:
        alt = dest.with_name(f"{dest.stem} ({i}){dest.suffix}")
        if not alt.exists():
            return alt
        i += 1


def click_consultar(page) -> None:
    try:
        page.locator("button.br-button.mt-md-4.primary:has-text('Consultar')").first.click(timeout=8000)
        return
    except Exception:
        pass
    click_by_text(page, ["Consultar"])


def _extract_companies_from_text(text: str) -> list[dict]:
    cnpj_re = _cnpj_regex_any()
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    ignore = {
        "cnpj",
        "empresa",
        "razao social",
        "razão social",
        "razao",
        "razão",
    }
    entries = []
    for i, line in enumerate(lines):
        for m in cnpj_re.finditer(line):
            cnpj = m.group(0)
            name = ""
            leftover = line.replace(cnpj, "").strip(" -:|")
            if leftover and not cnpj_re.search(leftover):
                if leftover.lower() not in ignore:
                    name = leftover
            if not name:
                j = i - 1
                while j >= 0:
                    prev = lines[j]
                    if not prev:
                        j -= 1
                        continue
                    if cnpj_re.search(prev):
                        j -= 1
                        continue
                    if prev.lower() in ignore:
                        j -= 1
                        continue
                    name = prev
                    break
            entries.append({"cnpj": cnpj, "name": name})
    return entries


def collect_companies_from_site(page, max_scrolls: int = 25) -> list[dict]:
    seen = {}
    no_change = 0
    last_count = 0
    for _ in range(max_scrolls):
        try:
            text = page.inner_text("body")
        except Exception:
            text = page.content() or ""
        for item in _extract_companies_from_text(text):
            norm = _normalize_cnpj(item["cnpj"])
            if not norm:
                continue
            existing = seen.get(norm)
            if existing is None:
                seen[norm] = {"cnpj": item["cnpj"], "name": item.get("name") or ""}
            elif not existing.get("name") and item.get("name"):
                existing["name"] = item["name"]
        if len(seen) == last_count:
            no_change += 1
            if no_change >= 3:
                break
        else:
            no_change = 0
        last_count = len(seen)
        try:
            page.mouse.wheel(0, 2000)
        except Exception:
            pass
        try:
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        except Exception:
            pass
        time.sleep(WAIT_SHORT)
    return list(seen.values())


def _ensure_openpyxl():
    if Workbook is None or load_workbook is None:
        raise SystemExit("openpyxl nao encontrado. Instale com: pip install openpyxl")


def _mes_ref_data(valor) -> str:
    if not valor:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m")
    texto = str(valor).strip()
    if not texto:
        return ""
    try:
        dt = datetime.strptime(texto, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%Y-%m")
    except Exception:
        pass
    try:
        dt = datetime.fromisoformat(texto)
        return dt.strftime("%Y-%m")
    except Exception:
        return ""


def load_checked_log(path: Path) -> tuple["Workbook", "Worksheet", dict, dict, dict]:
    _ensure_openpyxl()
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "checados"
        ws.append(["CNPJ", "Empresa", "Status", "Data/Hora"])
    row_map = {}
    status_map = {}
    mes_ref_map = {}
    for row in ws.iter_rows(min_row=2):
        cnpj_val = row[0].value
        if not cnpj_val:
            continue
        norm = _normalize_cnpj(str(cnpj_val))
        if not norm:
            continue
        row_map[norm] = row[0].row
        status = row[2].value if len(row) >= 3 else ""
        status_map[norm] = (str(status).strip().lower() if status else "")
        data_hora = row[3].value if len(row) >= 4 else ""
        mes_ref_map[norm] = _mes_ref_data(data_hora)
    return wb, ws, row_map, status_map, mes_ref_map


def update_checked_log(
    ws,
    row_map: dict,
    status_map: dict,
    mes_ref_map: dict,
    cnpj: str,
    name: str,
    status: str,
) -> None:
    norm = _normalize_cnpj(cnpj)
    if not norm:
        return
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row_idx = row_map.get(norm)
    if row_idx is None:
        ws.append([cnpj, name, status, now])
        row_map[norm] = ws.max_row
        status_map[norm] = status.strip().lower()
        mes_ref_map[norm] = _mes_ref_data(now)
        return

    current_status = status_map.get(norm, "")
    if status.strip().lower() == "encontrado" and current_status:
        if name:
            ws.cell(row=row_idx, column=2, value=name)
        return

    ws.cell(row=row_idx, column=1, value=cnpj)
    if name:
        ws.cell(row=row_idx, column=2, value=name)
    ws.cell(row=row_idx, column=3, value=status)
    ws.cell(row=row_idx, column=4, value=now)
    status_map[norm] = status.strip().lower()
    mes_ref_map[norm] = _mes_ref_data(now)


def results_path_for_ref(year: int, month: int) -> Path:
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    return RESULTS_DIR / f"Resultados - {year}-{month:02d}.xlsx"


def load_results_log(path: Path) -> tuple["Workbook", "Worksheet", dict]:
    _ensure_openpyxl()
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "resultados"
        ws.append(["CNPJ", "Status", "Data/Hora"])
    row_map = {}
    for row in ws.iter_rows(min_row=2):
        cnpj_val = row[0].value
        if not cnpj_val:
            continue
        norm = _normalize_cnpj(str(cnpj_val))
        if norm:
            row_map[norm] = row[0].row
    return wb, ws, row_map


def update_results_log(ws, row_map: dict, cnpj: str, status: str) -> None:
    norm = _normalize_cnpj(cnpj)
    if not norm:
        return
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row_idx = row_map.get(norm)
    if row_idx is None:
        ws.append([cnpj, status, now])
        row_map[norm] = ws.max_row
        return
    ws.cell(row=row_idx, column=1, value=cnpj)
    ws.cell(row=row_idx, column=2, value=status)
    ws.cell(row=row_idx, column=3, value=now)


def _output_dir_for_ref(year: int, month: int, output_root: Path) -> Path:
    dest_dir = output_root / str(year) / f"{month:02d}"
    dest_dir.mkdir(parents=True, exist_ok=True)
    return dest_dir


def download_csv_v14(
    page,
    company: str,
    output_root: Path,
    year_ref: int,
    month_ref: int,
    sobrescrever: bool = False,
) -> bool:
    dest_dir = _output_dir_for_ref(year_ref, month_ref, output_root)
    safe_company = _sanitize_filename(company)
    base_name = f"{safe_company} - {year_ref}-{month_ref:02d}.csv"
    dest_base = dest_dir / base_name
    if sobrescrever and dest_base.exists():
        dest_base.unlink()
    dest = dest_base if sobrescrever else get_unique_path(dest_base)

    end_time = time.time() + 5.0
    while time.time() < end_time:
        if is_unexpected_error_page(page):
            raise RuntimeError("erro_inesperado")
        try:
            if page.locator("button.br-button:has-text('CSV v1.4')").first.is_visible():
                break
        except Exception:
            pass
        time.sleep(0.4)

    if is_unexpected_error_page(page):
        raise RuntimeError("erro_inesperado")

    try:
        with page.expect_download(timeout=60000) as dl_info:
            try:
                page.locator("button.br-button:has-text('CSV v1.4')").first.click(timeout=10000)
            except Exception:
                if not click_by_text(page, ["CSV v1.4", "CSV 1.4", "CSV"]):
                    if is_unexpected_error_page(page):
                        raise RuntimeError("erro_inesperado")
                    return False
        download = dl_info.value
        download.save_as(str(dest))
        print(f"[ok] csv salvo: {dest}")
        return True
    except Exception:
        return False


def main():
    ano_exec, mes_exec, forcar_todas, sobrescrever, resetar_log = _parse_exec_args(sys.argv)
    if forcar_todas and not sobrescrever:
        sobrescrever = True
    checked_log_path = checked_log_path_for_ref(ano_exec, mes_exec)
    results_path = results_path_for_ref(ano_exec, mes_exec)
    if resetar_log:
        forcar_todas = True
        sobrescrever = True
        if checked_log_path.exists():
            checked_log_path.unlink()
        if results_path.exists():
            results_path.unlink()

    with sync_playwright() as p:
        browser = chrome_9222(p, PORT)
        context = browser.contexts[0] if browser.contexts else browser.new_context()
        page = context.new_page()
        wait_for_login(page)

        page.goto(SELECAO_URL)
        close_warning_if_present(page)
        companies = collect_companies_from_site(page)
        if not companies:
            print("[info] nenhuma empresa encontrada na lista do site.")
            return

        wb_checked, ws_checked, row_map, status_map, mes_ref_map = load_checked_log(checked_log_path)
        for item in companies:
            update_checked_log(
                ws_checked,
                row_map,
                status_map,
                mes_ref_map,
                item.get("cnpj", ""),
                item.get("name", ""),
                "encontrado",
            )
        wb_checked.save(checked_log_path)

        mes_atual = f"{ano_exec:04d}-{mes_exec:02d}"
        if forcar_todas:
            pending = [item for item in companies if _normalize_cnpj(item.get("cnpj", ""))]
        else:
            pending = []
            for item in companies:
                norm = _normalize_cnpj(item.get("cnpj", ""))
                if not norm:
                    continue
                status = status_map.get(norm, "")
                mes_ref = mes_ref_map.get(norm, "")
                if status in ("", "encontrado"):
                    pending.append(item)
                    continue
                if mes_ref != mes_atual:
                    pending.append(item)
                    continue

        if not pending:
            print("[info] nenhuma empresa pendente para checagem.")
            return

        wb_results, ws_results, results_row_map = load_results_log(results_path)

        for idx, item in enumerate(pending, start=1):
            company = item.get("name") or item.get("cnpj") or ""
            cnpj = item.get("cnpj") or ""
            print(f"[info] empresa {idx}/{len(pending)}: {company} ({cnpj})")
            attempts = 0
            while True:
                attempts += 1
                if attempts > 1:
                    print(f"[info] tentando novamente {attempts}/{MAX_RETRIES} para {company} ({cnpj})")

                page.goto(SELECAO_URL)
                close_warning_if_present(page)
                try:
                    select_company(page, cnpj)
                except RuntimeError as exc:
                    print(f"[warn] falha ao selecionar: {company} ({cnpj}) -> {exc}")
                    update_checked_log(ws_checked, row_map, status_map, mes_ref_map, cnpj, company, "checado_erro")
                    update_results_log(ws_results, results_row_map, cnpj, "Nao possui arquivos")
                    wb_checked.save(checked_log_path)
                    wb_results.save(results_path)
                    break

                try:
                    page.wait_for_url("**/home", timeout=15000)
                except Exception:
                    pass
                page.goto(CREDITO_URL)
                close_warning_if_present(page)
                if is_unexpected_error_page(page):
                    print("[warn] erro inesperado no site. Repetindo tentativa...")
                    if attempts < MAX_RETRIES:
                        time.sleep(WAIT_MED)
                        continue
                    print("[erro] erro no site.")
                    update_checked_log(ws_checked, row_map, status_map, mes_ref_map, cnpj, company, "checado_erro")
                    update_results_log(ws_results, results_row_map, cnpj, "Erro inesperado no site")
                    wb_checked.save(checked_log_path)
                    wb_results.save(results_path)
                    break

                set_ano_mes(page, ano_exec, mes_exec)
                click_consultar(page)
                time.sleep(5.0)
                if is_unexpected_error_page(page):
                    print("[warn] erro inesperado no site apos consultar. Repetindo tentativa...")
                    if attempts < MAX_RETRIES:
                        time.sleep(WAIT_MED)
                        continue
                    print("[erro] erro no site.")
                    update_checked_log(ws_checked, row_map, status_map, mes_ref_map, cnpj, company, "checado_erro")
                    update_results_log(ws_results, results_row_map, cnpj, "Erro inesperado no site")
                    wb_checked.save(checked_log_path)
                    wb_results.save(results_path)
                    break

                try:
                    ok = download_csv_v14(page, company, OUTPUT_ROOT, ano_exec, mes_exec, sobrescrever=sobrescrever)
                except RuntimeError as exc:
                    if str(exc) == "erro_inesperado":
                        print("[warn] erro inesperado no site ao aguardar botao. Repetindo tentativa...")
                        if attempts < MAX_RETRIES:
                            time.sleep(WAIT_MED)
                            continue
                        print("[erro] erro no site.")
                        update_checked_log(ws_checked, row_map, status_map, mes_ref_map, cnpj, company, "checado_erro")
                        update_results_log(ws_results, results_row_map, cnpj, "Erro inesperado no site")
                        wb_checked.save(checked_log_path)
                        wb_results.save(results_path)
                        break
                    raise
                if not ok:
                    print("[info] nao tem parcela.")
                    update_checked_log(ws_checked, row_map, status_map, mes_ref_map, cnpj, company, "checado_sem_csv")
                    update_results_log(ws_results, results_row_map, cnpj, "Nao possui arquivos")
                else:
                    print("[ok] baixou.")
                    update_checked_log(ws_checked, row_map, status_map, mes_ref_map, cnpj, company, "checado_ok")
                    update_results_log(ws_results, results_row_map, cnpj, "Baixou")
                wb_checked.save(checked_log_path)
                wb_results.save(results_path)
                time.sleep(WAIT_MED)
                page.goto(SELECAO_URL)
                close_warning_if_present(page)
                break


if __name__ == "__main__":
    main()
