from __future__ import annotations

from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import tkinter as tk
from tkinter import filedialog
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent

SAIDA_HEADERS = [
    "numero_nfse",
    "data_emissao",
    "situacao",
    "razao_social_tomador",
    "CNPJ/CPF",
    "valor_total_servicos",
    "irrf_retido",
    "pis_retido",
    "cofins_retido",
    "csll_retido",
    "inss_retido",
    "issqn_retido",
    "valor_liquido_servico",
    "arquivo",
]

ENTRADA_HEADERS = [
    "numero_nfse",
    "data_emissao",
    "situacao",
    "razao_social_emitente",
    "cnpj_emitente",
    "valor_total_servicos",
    "irrf_retido",
    "pis_retido",
    "cofins_retido",
    "csll_retido",
    "inss_retido",
    "issqn_retido",
    "valor_liquido_servico",
    "arquivo",
]

CURRENCY_COLUMNS = {
    "valor_total_servicos",
    "irrf_retido",
    "pis_retido",
    "cofins_retido",
    "csll_retido",
    "inss_retido",
    "issqn_retido",
    "valor_liquido_servico",
}


def local_name(tag: str) -> str:
    return tag.split("}", 1)[-1]


def direct_child(element: ET.Element | None, tag_name: str) -> ET.Element | None:
    if element is None:
        return None
    for child in list(element):
        if local_name(child.tag) == tag_name:
            return child
    return None


def first_descendant(element: ET.Element | None, tag_name: str) -> ET.Element | None:
    if element is None:
        return None
    for child in element.iter():
        if local_name(child.tag) == tag_name:
            return child
    return None


def text_of(element: ET.Element | None) -> str:
    if element is None or element.text is None:
        return ""
    return element.text.strip()


def descendant_text(element: ET.Element | None, tag_name: str) -> str:
    return text_of(first_descendant(element, tag_name))


def decimal_or_zero(value: str) -> float:
    if not value:
        return 0.0
    normalized = value.replace(",", ".")
    try:
        return float(Decimal(normalized))
    except (InvalidOperation, ValueError):
        return 0.0


def parse_date(value: str):
    if not value:
        return ""
    raw = value.strip()
    try:
        if "T" in raw:
            return datetime.fromisoformat(raw).date()
        return datetime.fromisoformat(raw[:10]).date()
    except ValueError:
        return raw


def get_tax_amounts(inf_nfse: ET.Element | None, inf_dps: ET.Element | None) -> dict[str, float]:
    nf_values = direct_child(inf_nfse, "valores")
    dps_values = direct_child(inf_dps, "valores")
    trib = direct_child(dps_values, "trib")
    trib_mun = direct_child(trib, "tribMun")
    trib_fed = direct_child(trib, "tribFed")
    pis_cofins = direct_child(trib_fed, "piscofins")

    tp_ret_issqn = descendant_text(trib_mun, "tpRetISSQN")
    issqn_value = decimal_or_zero(descendant_text(nf_values, "vISSQN"))
    issqn_retido = issqn_value if tp_ret_issqn == "2" else 0.0

    return {
        "valor_total_servicos": decimal_or_zero(descendant_text(dps_values, "vServ")),
        "irrf_retido": decimal_or_zero(descendant_text(trib_fed, "vRetIRRF")),
        "pis_retido": decimal_or_zero(descendant_text(pis_cofins, "vPis")),
        "cofins_retido": decimal_or_zero(descendant_text(pis_cofins, "vCofins")),
        "csll_retido": decimal_or_zero(descendant_text(trib_fed, "vRetCSLL")),
        "inss_retido": decimal_or_zero(descendant_text(trib_fed, "vRetCP")),
        "issqn_retido": issqn_retido,
        "valor_liquido_servico": decimal_or_zero(descendant_text(nf_values, "vLiq")),
    }


def get_situacao(path: Path, inf_dps: ET.Element | None, inf_nfse: ET.Element | None) -> str:
    name = path.stem.lower()
    if "cancelada" in name:
        return "cancelada"
    if "substituida" in name:
        return "substituida"

    cstat = descendant_text(inf_nfse, "cStat")
    if first_descendant(inf_dps, "subst") is not None or cstat == "101":
        return "substituicao"
    return "normal"


def get_party_document(element: ET.Element | None) -> str:
    cnpj = descendant_text(element, "CNPJ")
    if cnpj:
        return cnpj
    return descendant_text(element, "CPF")


def parse_xml(path: Path) -> tuple[str, dict[str, object]]:
    root = ET.parse(path).getroot()
    inf_nfse = first_descendant(root, "infNFSe")
    inf_dps = first_descendant(root, "infDPS")
    emit = direct_child(inf_nfse, "emit")
    toma = direct_child(inf_dps, "toma")
    taxes = get_tax_amounts(inf_nfse, inf_dps)

    common = {
        "numero_nfse": descendant_text(inf_nfse, "nNFSe"),
        "data_emissao": parse_date(descendant_text(inf_dps, "dhEmi") or descendant_text(inf_nfse, "dhProc")),
        "situacao": get_situacao(path, inf_dps, inf_nfse),
        "arquivo": path.name,
        **taxes,
    }

    file_type = "saida" if path.name.upper().startswith("SAIDA") else "entrada"
    if file_type == "saida":
        common["razao_social_tomador"] = descendant_text(toma, "xNome")
        common["CNPJ/CPF"] = get_party_document(toma)
    else:
        common["razao_social_emitente"] = descendant_text(emit, "xNome")
        common["cnpj_emitente"] = get_party_document(emit)
    return file_type, common


def adjust_columns(worksheet, headers: list[str]) -> None:
    for index, header in enumerate(headers, start=1):
        max_length = len(header)
        for cell in worksheet[get_column_letter(index)]:
            value = cell.value
            if value is None:
                continue
            length = len(str(value))
            if length > max_length:
                max_length = length
        worksheet.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 40)


def populate_sheet(worksheet, headers: list[str], rows: list[dict[str, object]]) -> None:
    worksheet.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    date_column = headers.index("data_emissao") + 1
    for row_cells in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        row_cells[date_column - 1].number_format = "dd/mm/yyyy"
        for index, header in enumerate(headers, start=1):
            if header in CURRENCY_COLUMNS:
                row_cells[index - 1].number_format = "#,##0.00"

    adjust_columns(worksheet, headers)


def ask_xml_folder() -> Path | None:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    selected = filedialog.askdirectory(
        title="Selecione a pasta com os XML",
        initialdir=str(BASE_DIR),
    )
    root.destroy()
    if not selected:
        return None
    return Path(selected)


def main() -> None:
    xml_folder = ask_xml_folder()
    if xml_folder is None:
        print("Selecao cancelada pelo usuario.")
        return

    xml_files = sorted(xml_folder.glob("*.xml"))
    output_file = xml_folder / f"relatorio_nfse_{xml_folder.name}.xlsx"

    if not xml_files:
        print("Nenhum XML encontrado na pasta selecionada.")
        return

    saida_rows: list[dict[str, object]] = []
    entrada_rows: list[dict[str, object]] = []

    for path in xml_files:
        file_type, row = parse_xml(path)
        if file_type == "saida":
            saida_rows.append(row)
        else:
            entrada_rows.append(row)

    workbook = Workbook()
    ws_saida = workbook.active
    ws_saida.title = "Notas Saida"
    populate_sheet(ws_saida, SAIDA_HEADERS, saida_rows)

    ws_entrada = workbook.create_sheet("Notas Entrada")
    populate_sheet(ws_entrada, ENTRADA_HEADERS, entrada_rows)

    workbook.save(output_file)
    print(f"Arquivo gerado: {output_file}")
    print(f"Notas de saida: {len(saida_rows)}")
    print(f"Notas de entrada: {len(entrada_rows)}")


if __name__ == "__main__":
    main()
