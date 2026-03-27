from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path
from tkinter import Tk, filedialog
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ===============================
# CONFIG
# ===============================
CONFIG: Dict[str, Any] = {
    "NOME_ABA": "Planilha1",
    "ARQUIVO_ENTRADA": "",
    "USAR_SELETOR_ARQUIVO": True,
    "PASTA_INICIAL_SELETOR": r"W:/PASTA CLIENTES/STURMER & WULFF ADVOCACIA TRIBUTARIA/CONCILIACAO",
    "ARQUIVO_TXT_SAIDA": r"W:/PASTA CLIENTES/STURMER & WULFF ADVOCACIA TRIBUTARIA/CONCILIACAO/2026/01/DIARIO_SW_01_2026_ERP.txt",
    "ARQUIVO_PENDENCIAS": r"W:/PASTA CLIENTES/STURMER & WULFF ADVOCACIA TRIBUTARIA/CONCILIACAO/2026/01/DIARIO_SW_01_2026_pendencias.csv",
    "ARQUIVO_EXCLUSOES": r"W:/PASTA CLIENTES/STURMER & WULFF ADVOCACIA TRIBUTARIA/CONCILIACAO/2026/01/DIARIO_SW_01_2026_exclusoes.csv",
    "CABECALHOS_ESPERADOS": [
        "Nº seq.",
        "Nº transação",
        "Data de lançamento",
        "Série",
        "Nº doc.",
        "Cta.contáb./cód.PN",
        "Cta.cont./Nome PN",
        "Débito/crédito (MC)",
        "Observações",
        "Número de sequência salvo",
        "Nome da filial",
        "Marca",
        "Centro de Custos",
    ],
    "PALAVRAS_EXCLUSAO_DIRETA": [
        "notas fiscais de saida",
        "nota fiscal de saida",
        "recebimento",
        "a/r invoice",
    ],
    "PALAVRAS_CANCELAMENTO": [
        "cancelamento",
        "cancelado",
        "cancellation",
        "cancelled",
        "anular",
        "anulado",
        "anulada",
        "estorno",
        "estornar",
        "estornado",
        "reversal",
        "reversed",
    ],
    "PARTNER_PREFIX_TO_ACCOUNT": {
        "TG": "1.1.2.02.04.01",
        "CL": "1.1.2.01.01.01",
        "F": "2.1.1.01.01.01",
        "BF": "1.1.2.01.01.01",
    },
    "COMPORTAMENTO_SEM_CLASSIFICACAO": "pendencia",
    "TOLERANCIA_PAREAMENTO_VALOR": Decimal("0.01"),
    "TOLERANCIA_PAREAMENTO_DIAS": 10,
    "PONTUACAO_MINIMA_PAREAMENTO": 5,
    "PONTUACAO_MINIMA_POSSIVEL_PAR": 4,
    "TXT_LAYOUT": {
        "delimitador": ";",
        "campos": [
            "seq",
            "numero_transacao",
            "data",
            "serie",
            "numero_doc",
            "historico",
            "conta_final",
            "valor",
            "descricao_item",
            "nome_filial",
            "marca",
            "centro_custos",
        ],
        "separador_decimal": ".",
    },
}


@dataclass
class MatchResult:
    candidato: Optional[Dict[str, Any]]
    pontuacao: int
    motivos: List[str]
    confirmado: bool


def normalizar_texto(texto: Any) -> str:
    if texto is None:
        return ""
    texto_str = str(texto).strip().lower()
    texto_str = unicodedata.normalize("NFKD", texto_str)
    texto_str = "".join(ch for ch in texto_str if not unicodedata.combining(ch))
    texto_str = re.sub(r"\s+", " ", texto_str)
    return texto_str


def limpar_historico(texto: str) -> str:
    if not texto:
        return ""
    limpo = str(texto)
    limpo = re.sub(r"\b\d+(?:\.\d+){2,}\b", " ", limpo)
    limpo = re.sub(r"\b(?:CL|F|TG|BF)\s*\d+\b", " ", limpo, flags=re.IGNORECASE)
    limpo = re.sub(r"\s*-\s*", " - ", limpo)
    limpo = re.sub(r"(?:\s*-\s*){2,}", " - ", limpo)
    limpo = re.sub(r"\s+", " ", limpo)
    limpo = re.sub(r"^[\s\-]+|[\s\-]+$", "", limpo)
    return limpo


def to_str_limpo(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    if isinstance(valor, Decimal) and valor == valor.to_integral_value():
        return str(int(valor))
    return str(valor).strip()


def parse_decimal(valor: Any) -> Optional[Decimal]:
    if valor is None:
        return None
    if isinstance(valor, Decimal):
        return valor
    if isinstance(valor, int):
        return Decimal(valor)
    if isinstance(valor, float):
        return Decimal(str(valor))

    texto = str(valor).strip()
    if not texto:
        return None
    texto = re.sub(r"[^\d,.\-]", "", texto)
    if not texto or texto in {"-", ",", "."}:
        return None
    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return Decimal(texto)
    except InvalidOperation:
        return None


def parse_data(valor: Any) -> Optional[date]:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    if not texto:
        return None
    formatos = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y")
    for fmt in formatos:
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def data_para_str(valor: Optional[date]) -> str:
    return valor.strftime("%Y-%m-%d") if valor else ""


def localizar_linha_cabecalho(
    ws: Worksheet, cabecalhos_esperados: List[str]
) -> Tuple[int, Dict[str, int]]:
    cabecalhos_norm = {normalizar_texto(c): c for c in cabecalhos_esperados}
    ultimo_mapa: Dict[str, int] = {}
    for idx_linha in range(1, min(ws.max_row, 30) + 1):
        valores = [ws.cell(row=idx_linha, column=c).value for c in range(1, ws.max_column + 1)]
        mapa: Dict[str, int] = {}
        for col_idx, valor in enumerate(valores, start=1):
            nome_norm = normalizar_texto(valor)
            if nome_norm in cabecalhos_norm:
                mapa[cabecalhos_norm[nome_norm]] = col_idx
        ultimo_mapa = mapa
        if len(mapa) == len(cabecalhos_esperados):
            return idx_linha, mapa
    faltantes = [c for c in cabecalhos_esperados if c not in ultimo_mapa]
    raise ValueError(f"Cabeçalhos esperados não encontrados. Faltantes: {faltantes}")


def agrupar_lancamentos(
    ws: Worksheet, linha_cabecalho: int, colunas: Dict[str, int]
) -> List[Dict[str, Any]]:
    lancamentos: List[Dict[str, Any]] = []
    atual: Optional[Dict[str, Any]] = None
    contador_id = 0

    def valor_coluna(nome_coluna: str, idx_linha: int) -> Any:
        return ws.cell(row=idx_linha, column=colunas[nome_coluna]).value

    for idx_linha in range(linha_cabecalho + 1, ws.max_row + 1):
        seq_raw = valor_coluna("Nº seq.", idx_linha)
        nova_operacao = to_str_limpo(seq_raw) != ""

        if nova_operacao:
            if atual:
                lancamentos.append(atual)
            contador_id += 1
            atual = {
                "_id": contador_id,
                "seq": to_str_limpo(seq_raw),
                "numero_transacao": to_str_limpo(valor_coluna("Nº transação", idx_linha)),
                "data_lancamento": parse_data(valor_coluna("Data de lançamento", idx_linha)),
                "serie": to_str_limpo(valor_coluna("Série", idx_linha)),
                "numero_doc": to_str_limpo(valor_coluna("Nº doc.", idx_linha)),
                "observacao_principal": to_str_limpo(valor_coluna("Observações", idx_linha)),
                "nome_filial": to_str_limpo(valor_coluna("Nome da filial", idx_linha)),
                "marca": to_str_limpo(valor_coluna("Marca", idx_linha)),
                "centro_custos": to_str_limpo(valor_coluna("Centro de Custos", idx_linha)),
                "itens_contabeis": [],
                "_linha_inicio": idx_linha,
            }

        if atual is None:
            continue

        codigo = to_str_limpo(valor_coluna("Cta.contáb./cód.PN", idx_linha))
        descricao = to_str_limpo(valor_coluna("Cta.cont./Nome PN", idx_linha))
        valor = parse_decimal(valor_coluna("Débito/crédito (MC)", idx_linha))
        observacao_linha = to_str_limpo(valor_coluna("Observações", idx_linha))

        if not codigo and not descricao and valor is None:
            continue

        item = {
            "codigo_original": codigo,
            "descricao_original": descricao,
            "valor": valor,
            "observacao": observacao_linha,
            "marca": to_str_limpo(valor_coluna("Marca", idx_linha)) or atual["marca"],
            "centro_custos": to_str_limpo(valor_coluna("Centro de Custos", idx_linha))
            or atual["centro_custos"],
            "nome_filial": to_str_limpo(valor_coluna("Nome da filial", idx_linha))
            or atual["nome_filial"],
            "_linha_origem": idx_linha,
        }
        atual["itens_contabeis"].append(item)

        if not atual["observacao_principal"] and observacao_linha:
            atual["observacao_principal"] = observacao_linha

    if atual:
        lancamentos.append(atual)
    return lancamentos


def texto_observacoes_lancamento(lancamento: Dict[str, Any]) -> str:
    partes = [lancamento.get("observacao_principal", "")]
    partes.extend(item.get("observacao", "") for item in lancamento.get("itens_contabeis", []))
    return " | ".join(p for p in partes if p)


def contem_termo(texto: str, termos: Iterable[str]) -> bool:
    texto_norm = normalizar_texto(texto)
    return any(normalizar_texto(t) in texto_norm for t in termos)


def eh_lancamento_de_exclusao_direta(
    lancamento: Dict[str, Any], config: Dict[str, Any]
) -> bool:
    texto = texto_observacoes_lancamento(lancamento)
    return contem_termo(texto, config["PALAVRAS_EXCLUSAO_DIRETA"])


def identificar_lancamentos_cancelatorios(
    lancamentos: List[Dict[str, Any]], config: Dict[str, Any]
) -> List[Dict[str, Any]]:
    resultado = []
    for lanc in lancamentos:
        texto = texto_observacoes_lancamento(lanc)
        if contem_termo(texto, config["PALAVRAS_CANCELAMENTO"]):
            resultado.append(lanc)
    return resultado


def total_lancamento(lancamento: Dict[str, Any]) -> Decimal:
    total = Decimal("0")
    for item in lancamento.get("itens_contabeis", []):
        valor = item.get("valor")
        if isinstance(valor, Decimal):
            total += valor
    return total


def assinatura_itens_absoluta(lancamento: Dict[str, Any]) -> Tuple[Tuple[str, str], ...]:
    soma_por_conta: Dict[str, Decimal] = {}
    for item in lancamento.get("itens_contabeis", []):
        conta = normalizar_texto(item.get("codigo_original", ""))
        valor = item.get("valor")
        if not conta or not isinstance(valor, Decimal):
            continue
        soma_por_conta[conta] = soma_por_conta.get(conta, Decimal("0")) + abs(valor)
    assinatura = tuple(
        sorted((conta, str(valor.quantize(Decimal("0.01")))) for conta, valor in soma_por_conta.items())
    )
    return assinatura


def extrair_tokens_referencia(texto: str) -> set:
    if not texto:
        return set()
    tokens = set(re.findall(r"\b[A-Z]{1,4}\d{2,}\b", texto.upper()))
    tokens.update(re.findall(r"\b\d{3,}\b", texto))
    return tokens


def distancia_dias(data_a: Optional[date], data_b: Optional[date]) -> Optional[int]:
    if not data_a or not data_b:
        return None
    return abs((data_a - data_b).days)


def texto_sem_termos_cancelamento(texto: str, config: Dict[str, Any]) -> str:
    resultado = normalizar_texto(texto)
    for termo in config["PALAVRAS_CANCELAMENTO"]:
        termo_norm = normalizar_texto(termo)
        resultado = resultado.replace(termo_norm, " ")
    resultado = re.sub(r"\s+", " ", resultado).strip()
    return resultado


def pontuar_candidato_par(
    estorno: Dict[str, Any], candidato: Dict[str, Any], config: Dict[str, Any]
) -> Tuple[int, List[str]]:
    pontos = 0
    motivos: List[str] = []
    tol_valor: Decimal = config["TOLERANCIA_PAREAMENTO_VALOR"]
    tol_dias: int = config["TOLERANCIA_PAREAMENTO_DIAS"]

    if estorno.get("numero_doc") and estorno.get("numero_doc") == candidato.get("numero_doc"):
        pontos += 3
        motivos.append("numero_doc")
    if estorno.get("serie") and estorno.get("serie") == candidato.get("serie"):
        pontos += 1
        motivos.append("serie")

    dias = distancia_dias(estorno.get("data_lancamento"), candidato.get("data_lancamento"))
    if dias is not None and dias <= tol_dias:
        pontos += 1
        motivos.append("data_proxima")

    total_estorno = total_lancamento(estorno)
    total_candidato = total_lancamento(candidato)
    if abs(abs(total_estorno) - abs(total_candidato)) <= tol_valor:
        pontos += 2
        motivos.append("valor_modulo")
    if abs(total_estorno + total_candidato) <= tol_valor:
        pontos += 2
        motivos.append("valor_espelhado")

    ass_estorno = assinatura_itens_absoluta(estorno)
    ass_candidato = assinatura_itens_absoluta(candidato)
    if ass_estorno and ass_estorno == ass_candidato:
        pontos += 2
        motivos.append("composicao_itens")

    obs_estorno = texto_observacoes_lancamento(estorno)
    obs_candidato = texto_observacoes_lancamento(candidato)
    tokens_estorno = extrair_tokens_referencia(obs_estorno)
    tokens_candidato = extrair_tokens_referencia(obs_candidato)
    if tokens_estorno.intersection(tokens_candidato):
        pontos += 1
        motivos.append("token_referencia")

    base_estorno = texto_sem_termos_cancelamento(obs_estorno, config)
    base_candidato = texto_sem_termos_cancelamento(obs_candidato, config)
    if base_estorno and base_candidato:
        similaridade = SequenceMatcher(a=base_estorno, b=base_candidato).ratio()
        if similaridade >= 0.65:
            pontos += 1
            motivos.append("similaridade_observacao")
    return pontos, motivos


def localizar_par_do_estorno(
    lancamento_estorno: Dict[str, Any],
    lancamentos: List[Dict[str, Any]],
    config: Dict[str, Any],
    excluidos_ids: set,
) -> MatchResult:
    melhor: Optional[Dict[str, Any]] = None
    melhor_pontuacao = -1
    melhor_motivos: List[str] = []

    for candidato in lancamentos:
        if candidato["_id"] == lancamento_estorno["_id"]:
            continue
        if candidato["_id"] in excluidos_ids:
            continue

        pontuacao, motivos = pontuar_candidato_par(lancamento_estorno, candidato, config)

        # Penaliza quando ambos parecem estorno/cancelamento explicito.
        if contem_termo(texto_observacoes_lancamento(candidato), config["PALAVRAS_CANCELAMENTO"]):
            pontuacao -= 1

        if pontuacao > melhor_pontuacao:
            melhor = candidato
            melhor_pontuacao = pontuacao
            melhor_motivos = motivos

    confirmado = melhor is not None and melhor_pontuacao >= config["PONTUACAO_MINIMA_PAREAMENTO"]
    return MatchResult(melhor, melhor_pontuacao, melhor_motivos, confirmado)


def construir_registro_exclusao(
    lancamento: Dict[str, Any], motivo: str, transacao_par: str = ""
) -> Dict[str, Any]:
    return {
        "seq": lancamento.get("seq", ""),
        "numero_transacao": lancamento.get("numero_transacao", ""),
        "numero_doc": lancamento.get("numero_doc", ""),
        "data": data_para_str(lancamento.get("data_lancamento")),
        "observacao": lancamento.get("observacao_principal", ""),
        "motivo_exclusao": motivo,
        "transacao_par_vinculada": transacao_par,
    }


def marcar_pares_para_exclusao(
    lancamentos: List[Dict[str, Any]], config: Dict[str, Any]
) -> Tuple[set, List[Dict[str, Any]]]:
    excluidos_ids: set = set()
    exclusoes: List[Dict[str, Any]] = []
    exclusoes_por_id: Dict[int, Dict[str, Any]] = {}

    for lanc in lancamentos:
        if eh_lancamento_de_exclusao_direta(lanc, config):
            excluidos_ids.add(lanc["_id"])
            exclusoes_por_id[lanc["_id"]] = construir_registro_exclusao(
                lanc, "exclusao_direta_tipo_operacao"
            )

    cancelatorios = identificar_lancamentos_cancelatorios(lancamentos, config)
    for estorno in cancelatorios:
        if estorno["_id"] in excluidos_ids:
            continue

        match = localizar_par_do_estorno(estorno, lancamentos, config, excluidos_ids)
        if match.confirmado and match.candidato:
            par = match.candidato
            excluidos_ids.add(estorno["_id"])
            excluidos_ids.add(par["_id"])

            exclusoes_por_id[estorno["_id"]] = construir_registro_exclusao(
                estorno,
                f"cancelamento_estorno_anulacao (score={match.pontuacao}; {'/'.join(match.motivos)})",
                transacao_par=par.get("numero_transacao", ""),
            )
            exclusoes_por_id[par["_id"]] = construir_registro_exclusao(
                par,
                f"par_do_cancelamento_estorno_anulacao (score={match.pontuacao}; {'/'.join(match.motivos)})",
                transacao_par=estorno.get("numero_transacao", ""),
            )
        elif match.candidato and match.pontuacao >= config["PONTUACAO_MINIMA_POSSIVEL_PAR"]:
            par_possivel = match.candidato
            excluidos_ids.add(estorno["_id"])
            excluidos_ids.add(par_possivel["_id"])
            exclusoes_por_id[estorno["_id"]] = construir_registro_exclusao(
                estorno,
                f"possivel_par_estorno_sem_validacao (score={match.pontuacao}; {'/'.join(match.motivos)})",
                transacao_par=par_possivel.get("numero_transacao", ""),
            )
            exclusoes_por_id[par_possivel["_id"]] = construir_registro_exclusao(
                par_possivel,
                f"possivel_original_de_estorno_sem_validacao (score={match.pontuacao}; {'/'.join(match.motivos)})",
                transacao_par=estorno.get("numero_transacao", ""),
            )
        else:
            excluidos_ids.add(estorno["_id"])
            exclusoes_por_id[estorno["_id"]] = construir_registro_exclusao(
                estorno,
                f"cancelamento_estorno_sem_par_confirmado (score={match.pontuacao})",
            )

    exclusoes.extend(exclusoes_por_id.values())
    return excluidos_ids, exclusoes


def eh_conta_contabil(codigo: str) -> bool:
    return bool(re.fullmatch(r"\d+(?:\.\d+)+", codigo.strip()))


def eh_codigo_parceiro_generico(codigo: str) -> bool:
    codigo_limpo = re.sub(r"\s+", "", codigo).upper()
    return bool(re.fullmatch(r"[A-Z]{1,4}\d{2,}", codigo_limpo))


def substituir_codigo_parceiro_por_prefixo(
    codigo: str, config: Dict[str, Any]
) -> Tuple[Optional[str], Optional[str]]:
    codigo_limpo = re.sub(r"\s+", "", to_str_limpo(codigo)).upper()
    if not codigo_limpo:
        return None, "codigo_vazio"

    mapa = config["PARTNER_PREFIX_TO_ACCOUNT"]
    for prefixo in sorted(mapa.keys(), key=len, reverse=True):
        if codigo_limpo.startswith(prefixo.upper()):
            return mapa[prefixo], None

    if eh_conta_contabil(codigo_limpo):
        return codigo_limpo, None

    if eh_codigo_parceiro_generico(codigo_limpo):
        return None, f"prefixo_nao_mapeado:{codigo_limpo}"

    return None, f"codigo_sem_classificacao:{codigo_limpo}"


def resolver_conta_final_item(
    item: Dict[str, Any], config: Dict[str, Any]
) -> Tuple[Optional[Dict[str, Any]], Optional[Dict[str, Any]]]:
    codigo_original = item.get("codigo_original", "")
    conta_final, erro = substituir_codigo_parceiro_por_prefixo(codigo_original, config)

    if erro and config.get("COMPORTAMENTO_SEM_CLASSIFICACAO") == "usar_codigo_original":
        codigo_limpo = re.sub(r"\s+", "", to_str_limpo(codigo_original)).upper()
        if codigo_limpo and not eh_codigo_parceiro_generico(codigo_limpo):
            conta_final, erro = codigo_limpo, None

    if erro:
        pendencia = {
            "codigo_parceiro": item.get("codigo_original", ""),
            "nome_parceiro": item.get("descricao_original", ""),
            "valor": item.get("valor", ""),
            "motivo_da_pendencia": erro,
        }
        return None, pendencia

    if re.match(r"^(TG|CL|F|BF)", conta_final or "", re.IGNORECASE):
        pendencia = {
            "codigo_parceiro": item.get("codigo_original", ""),
            "nome_parceiro": item.get("descricao_original", ""),
            "valor": item.get("valor", ""),
            "motivo_da_pendencia": "codigo_de_parceiro_nao_substituido",
        }
        return None, pendencia

    if item.get("valor") is None:
        pendencia = {
            "codigo_parceiro": item.get("codigo_original", ""),
            "nome_parceiro": item.get("descricao_original", ""),
            "valor": "",
            "motivo_da_pendencia": "item_sem_valor",
        }
        return None, pendencia

    resolvido = {
        "conta_original": item.get("codigo_original", ""),
        "conta_final": conta_final,
        "descricao": item.get("descricao_original", ""),
        "valor": item.get("valor"),
        "observacao": item.get("observacao", ""),
        "marca": item.get("marca", ""),
        "centro_custos": item.get("centro_custos", ""),
        "nome_filial": item.get("nome_filial", ""),
    }
    return resolvido, None


def normalizar_lancamento(
    bruto: Dict[str, Any], config: Dict[str, Any], pendencias: List[Dict[str, Any]]
) -> Optional[Dict[str, Any]]:
    historico = limpar_historico(bruto.get("observacao_principal", ""))
    lancamento_normalizado = {
        "seq": bruto.get("seq", ""),
        "numero_transacao": bruto.get("numero_transacao", ""),
        "data": data_para_str(bruto.get("data_lancamento")),
        "serie": bruto.get("serie", ""),
        "numero_doc": bruto.get("numero_doc", ""),
        "historico": historico,
        "nome_filial": bruto.get("nome_filial", ""),
        "marca": bruto.get("marca", ""),
        "centro_custos": bruto.get("centro_custos", ""),
        "itens": [],
    }

    for item in bruto.get("itens_contabeis", []):
        item_resolvido, pendencia = resolver_conta_final_item(item, config)
        if pendencia:
            pendencias.append(
                {
                    "seq": bruto.get("seq", ""),
                    "numero_transacao": bruto.get("numero_transacao", ""),
                    "numero_doc": bruto.get("numero_doc", ""),
                    "data": data_para_str(bruto.get("data_lancamento")),
                    "observacao": bruto.get("observacao_principal", ""),
                    **pendencia,
                }
            )
            continue
        lancamento_normalizado["itens"].append(item_resolvido)

    if not lancamento_normalizado["itens"]:
        pendencias.append(
            {
                "seq": bruto.get("seq", ""),
                "numero_transacao": bruto.get("numero_transacao", ""),
                "numero_doc": bruto.get("numero_doc", ""),
                "data": data_para_str(bruto.get("data_lancamento")),
                "observacao": bruto.get("observacao_principal", ""),
                "codigo_parceiro": "",
                "nome_parceiro": "",
                "valor": "",
                "motivo_da_pendencia": "lancamento_sem_itens_validos",
            }
        )
        return None

    return lancamento_normalizado


def formatar_decimal_txt(valor: Decimal, separador_decimal: str) -> str:
    texto = f"{valor:.2f}"
    if separador_decimal != ".":
        texto = texto.replace(".", separador_decimal)
    return texto


def formatar_linha_erp(item: Dict[str, Any], lancamento: Dict[str, Any], config: Dict[str, Any]) -> str:
    layout = config["TXT_LAYOUT"]
    delimitador = layout["delimitador"]
    separador_decimal = layout.get("separador_decimal", ".")

    contexto = {
        "seq": lancamento.get("seq", ""),
        "numero_transacao": lancamento.get("numero_transacao", ""),
        "data": lancamento.get("data", ""),
        "serie": lancamento.get("serie", ""),
        "numero_doc": lancamento.get("numero_doc", ""),
        "historico": lancamento.get("historico", ""),
        "conta_original": item.get("conta_original", ""),
        "conta_final": item.get("conta_final", ""),
        "descricao_item": item.get("descricao", ""),
        "valor": formatar_decimal_txt(item.get("valor", Decimal("0")), separador_decimal),
        "observacao_item": item.get("observacao", ""),
        "nome_filial": item.get("nome_filial", "") or lancamento.get("nome_filial", ""),
        "marca": item.get("marca", "") or lancamento.get("marca", ""),
        "centro_custos": item.get("centro_custos", "") or lancamento.get("centro_custos", ""),
    }

    campos = layout["campos"]
    valores = []
    for campo in campos:
        valor = to_str_limpo(contexto.get(campo, ""))
        valor = valor.replace("\n", " ").replace("\r", " ").strip()
        valores.append(valor)
    return delimitador.join(valores)


def gerar_txt_erp(lancamentos: List[Dict[str, Any]], caminho_saida: Path, config: Dict[str, Any]) -> None:
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    with caminho_saida.open("w", encoding="utf-8", newline="\n") as arquivo:
        for lanc in lancamentos:
            for item in lanc.get("itens", []):
                linha = formatar_linha_erp(item, lanc, config)
                arquivo.write(linha + "\n")


def gerar_relatorio_pendencias(pendencias: List[Dict[str, Any]], caminho_saida: Path) -> None:
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    campos = [
        "seq",
        "numero_transacao",
        "numero_doc",
        "data",
        "observacao",
        "codigo_parceiro",
        "nome_parceiro",
        "valor",
        "motivo_da_pendencia",
    ]
    with caminho_saida.open("w", encoding="utf-8", newline="") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=campos, delimiter=";")
        writer.writeheader()
        for pend in pendencias:
            registro = dict(pend)
            valor = registro.get("valor")
            if isinstance(valor, Decimal):
                registro["valor"] = f"{valor:.2f}"
            writer.writerow(registro)


def gerar_relatorio_exclusoes(exclusoes: List[Dict[str, Any]], caminho_saida: Path) -> None:
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    campos = [
        "seq",
        "numero_transacao",
        "numero_doc",
        "data",
        "observacao",
        "motivo_exclusao",
        "transacao_par_vinculada",
    ]
    with caminho_saida.open("w", encoding="utf-8", newline="") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=campos, delimiter=";")
        writer.writeheader()
        for exc in exclusoes:
            writer.writerow(exc)


def selecionar_planilha_usuario(config: Dict[str, Any]) -> Path:
    pasta_inicial = config.get("PASTA_INICIAL_SELETOR") or str(Path.cwd())
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    caminho = filedialog.askopenfilename(
        title="Selecione a planilha contábil para conversão",
        initialdir=pasta_inicial,
        filetypes=[
            ("Planilhas Excel", "*.xlsx *.xlsm *.xltx *.xltm"),
            ("Todos os arquivos", "*.*"),
        ],
    )
    root.destroy()
    if not caminho:
        raise FileNotFoundError("Nenhuma planilha foi selecionada na janela de arquivo.")
    return Path(caminho)


def carregar_planilha(config: Dict[str, Any]) -> Tuple[Worksheet, Path]:
    if config.get("USAR_SELETOR_ARQUIVO", True):
        caminho_entrada = selecionar_planilha_usuario(config)
    else:
        caminho_entrada = Path(config["ARQUIVO_ENTRADA"])
    if not caminho_entrada.exists():
        raise FileNotFoundError(f"Arquivo de entrada não encontrado: {caminho_entrada}")

    wb = load_workbook(caminho_entrada, data_only=True, read_only=False)
    nome_aba = config["NOME_ABA"]
    if nome_aba not in wb.sheetnames:
        raise ValueError(f"Aba '{nome_aba}' não encontrada. Abas disponíveis: {wb.sheetnames}")
    return wb[nome_aba], caminho_entrada


def executar(config: Dict[str, Any]) -> Dict[str, Any]:
    pendencias: List[Dict[str, Any]] = []

    ws, caminho_entrada = carregar_planilha(config)
    linha_cabecalho, colunas = localizar_linha_cabecalho(ws, config["CABECALHOS_ESPERADOS"])
    lancamentos_brutos = agrupar_lancamentos(ws, linha_cabecalho, colunas)

    excluidos_ids, exclusoes = marcar_pares_para_exclusao(lancamentos_brutos, config)

    lancamentos_validos_brutos = [l for l in lancamentos_brutos if l["_id"] not in excluidos_ids]
    lancamentos_normalizados: List[Dict[str, Any]] = []

    for bruto in lancamentos_validos_brutos:
        try:
            normalizado = normalizar_lancamento(bruto, config, pendencias)
            if normalizado:
                lancamentos_normalizados.append(normalizado)
        except Exception as exc:  # noqa: BLE001
            pendencias.append(
                {
                    "seq": bruto.get("seq", ""),
                    "numero_transacao": bruto.get("numero_transacao", ""),
                    "numero_doc": bruto.get("numero_doc", ""),
                    "data": data_para_str(bruto.get("data_lancamento")),
                    "observacao": bruto.get("observacao_principal", ""),
                    "codigo_parceiro": "",
                    "nome_parceiro": "",
                    "valor": "",
                    "motivo_da_pendencia": f"erro_normalizacao:{exc}",
                }
            )

    gerar_txt_erp(lancamentos_normalizados, Path(config["ARQUIVO_TXT_SAIDA"]), config)
    gerar_relatorio_pendencias(pendencias, Path(config["ARQUIVO_PENDENCIAS"]))
    gerar_relatorio_exclusoes(exclusoes, Path(config["ARQUIVO_EXCLUSOES"]))

    return {
        "lancamentos_lidos": len(lancamentos_brutos),
        "lancamentos_excluidos": len(excluidos_ids),
        "lancamentos_exportados": len(lancamentos_normalizados),
        "pendencias": len(pendencias),
        "arquivo_entrada": str(caminho_entrada),
        "arquivo_txt": config["ARQUIVO_TXT_SAIDA"],
        "arquivo_pendencias": config["ARQUIVO_PENDENCIAS"],
        "arquivo_exclusoes": config["ARQUIVO_EXCLUSOES"],
    }


if __name__ == "__main__":
    resumo = executar(CONFIG)
    print("Processamento concluido:")
    for chave, valor in resumo.items():
        print(f"- {chave}: {valor}")

    # COMO USAR
    # 1) Ajuste o bloco CONFIG (saida, layout TXT e pasta inicial do seletor).
    # 2) Instale dependencias: pip install openpyxl
    # 3) Execute: python converter_diario_sw_para_erp.py
    # 4) Selecione a planilha .xlsx na janela que sera aberta.
    # 5) Revise os arquivos de pendencias/exclusoes para auditoria.
