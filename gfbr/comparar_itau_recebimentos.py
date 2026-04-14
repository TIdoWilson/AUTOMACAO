from __future__ import annotations

import argparse
from datetime import date
from difflib import SequenceMatcher
from itertools import combinations
import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SHEET_BANCOS = "somente bancos"
SHEET_RECEBIMENTOS = "somente recebimentos"

OUTPUT_COLUMNS_BANCOS = [
    "data referencial",
    "valor referencial",
    "bancos origem",
    "bancos conta contrapartida",
    "bancos detalhes",
]

OUTPUT_COLUMNS_RECEB = [
    "data referencial",
    "valor referencial",
    "receb banco",
    "receb numero nf",
    "receb clientes",
    "receb valor liquido",
]

DATE_NUMBER_FORMAT = "dd/mm/yyyy"
VALUE_NUMBER_FORMAT = "#,##0.00"
VALUE_TOLERANCE = 0.05
MULTI_PERNA_MAX_SIZE = 10
MULTI_PERNA_MAX_CANDIDATES = 20
SIMILARITY_THRESHOLD = 0.50

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
GRAY_FILL = PatternFill("solid", fgColor="E7E6E6")


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().strip()
    text = " ".join(text.split())
    return "" if text in {"nan", "none", "nat", "null"} else text


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    return df.rename(columns={c: normalize_text(c) for c in df.columns})


def series_or_empty(df: pd.DataFrame, column: str) -> pd.Series:
    if column in df.columns:
        return df[column]
    return pd.Series("", index=df.index)


def normalize_match_text(value: object) -> str:
    text = normalize_text(value)
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def join_text_parts(*parts: object) -> str:
    normalized_parts = [normalize_text(part) for part in parts if normalize_text(part)]
    return " ".join(normalized_parts).strip()


def join_display_parts(*parts: object) -> str:
    display_parts = []
    for part in parts:
        text = "" if part is None else str(part).strip()
        if normalize_text(text):
            display_parts.append(text)
    return " ".join(display_parts).strip()


def fallback_text(primary: object, *parts: object) -> str:
    primary_text = normalize_text(primary)
    if primary_text:
        return str(primary).strip()
    fallback = join_display_parts(*parts)
    return fallback


def text_similarity(left: object, right: object) -> float:
    left_text = normalize_match_text(left)
    right_text = normalize_match_text(right)
    if not left_text or not right_text:
        return 0.0

    ratio = SequenceMatcher(None, left_text, right_text).ratio()

    if left_text in right_text or right_text in left_text:
        ratio = max(ratio, 0.90)

    left_tokens = {token for token in left_text.split() if len(token) >= 3}
    right_tokens = {token for token in right_text.split() if len(token) >= 3}
    if left_tokens and right_tokens:
        token_score = len(left_tokens & right_tokens) / len(left_tokens | right_tokens)
        ratio = max(ratio, token_score)

    return ratio


def text_matches_receita_reembolsavel(text: object) -> bool:
    normalized = normalize_match_text(text)
    if not normalized:
        return False

    if "receit" in normalized and "reembols" in normalized:
        return True

    return SequenceMatcher(None, normalized, "receitas reembolsaveis").ratio() >= 0.72


def looks_like_account_only(text: object) -> bool:
    normalized = normalize_text(text)
    if not normalized:
        return False

    compact = re.sub(r"[\s\-]+", "", normalized)
    return bool(re.fullmatch(r"\d+(?:\.\d+){3,}", compact))


def select_input_file(title: str, initial_dir: Path, filetypes: list[tuple[str, str]]) -> Path:
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    selected = filedialog.askopenfilename(
        title=title,
        initialdir=str(initial_dir),
        filetypes=filetypes,
    )
    root.destroy()

    if not selected:
        raise SystemExit("Selecao de arquivo cancelada pelo usuario.")

    return Path(selected)


def select_processing_period() -> tuple[int, int]:
    import tkinter as tk
    from tkinter import simpledialog

    today = date.today()

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    try:
        month = simpledialog.askinteger(
            title="Periodo de processamento",
            prompt="Informe o mes de processamento (1-12):",
            parent=root,
            minvalue=1,
            maxvalue=12,
            initialvalue=today.month,
        )
        if month is None:
            raise SystemExit("Selecao do mes de processamento cancelada pelo usuario.")

        year = simpledialog.askinteger(
            title="Periodo de processamento",
            prompt="Informe o ano de processamento (AAAA):",
            parent=root,
            minvalue=2000,
            maxvalue=2100,
            initialvalue=today.year,
        )
        if year is None:
            raise SystemExit("Selecao do ano de processamento cancelada pelo usuario.")
    finally:
        root.destroy()

    return year, month


def select_output_file(title: str, initial_dir: Path, default_filename: str) -> Path:
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    selected = filedialog.asksaveasfilename(
        title=title,
        initialdir=str(initial_dir),
        initialfile=default_filename,
        defaultextension=".xlsx",
        filetypes=[("Planilha Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
    )
    root.destroy()

    if not selected:
        raise SystemExit("Selecao de arquivo de saida cancelada pelo usuario.")

    return Path(selected)


def find_sheet_name(excel_path: Path, target_name: str, engine: str | None = None) -> str:
    workbook = pd.ExcelFile(excel_path, engine=engine)
    normalized_target = normalize_text(target_name)

    for sheet in workbook.sheet_names:
        if normalize_text(sheet) == normalized_target:
            return sheet

    raise ValueError(
        f"Aba '{target_name}' nao encontrada em {excel_path.name}. "
        f"Abas disponiveis: {workbook.sheet_names}"
    )


def parse_excel_date_column(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    numeric_ratio = float(numeric.notna().mean())
    serial_ratio = float(numeric.between(20000, 80000).mean())

    if numeric_ratio >= 0.70 and serial_ratio >= 0.70:
        return pd.to_datetime(numeric, unit="D", origin="1899-12-30", errors="coerce")

    return pd.to_datetime(series, errors="coerce")


def prepare_bancos_itau(path: Path, year: int, month: int) -> pd.DataFrame:
    sheet_name = find_sheet_name(path, "itau")
    df = pd.read_excel(path, sheet_name=sheet_name)
    df = normalize_columns(df)

    required_columns = [
        "data de lancamento",
        "c/d (ml)",
    ]
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        raise ValueError(f"Colunas obrigatorias ausentes em {path.name}: {missing}")

    df["data_lancamento"] = pd.to_datetime(df["data de lancamento"], errors="coerce")
    df["valor"] = pd.to_numeric(df["c/d (ml)"], errors="coerce").round(2)
    df["detalhes"] = series_or_empty(df, "detalhes")
    df["origem"] = series_or_empty(df, "origem")
    df["no_transacao"] = series_or_empty(df, "no transacao")
    df["no_origem"] = series_or_empty(df, "no origem")
    df["conta_contrapartida"] = series_or_empty(df, "conta de contrapartida")

    filtered = df[
        (df["data_lancamento"].dt.year == year)
        & (df["data_lancamento"].dt.month == month)
        & (df["valor"] > 0)
    ].copy()

    filtered["data referencial"] = filtered["data_lancamento"].dt.normalize()
    filtered["valor referencial"] = filtered["valor"].round(2)
    filtered["bancos origem"] = filtered["origem"].fillna("").astype(str)
    filtered["bancos conta contrapartida"] = filtered["conta_contrapartida"].fillna("").astype(str)
    filtered["bancos detalhes"] = filtered["detalhes"].fillna("").astype(str)
    filtered["bancos no transacao"] = filtered["no_transacao"].fillna("").astype(str)
    filtered["bancos no origem"] = filtered["no_origem"].fillna("").astype(str)
    filtered["bancos detalhes"] = filtered.apply(
        lambda row: fallback_text(
            row["bancos detalhes"],
            row["bancos origem"],
            f"Nº transacao {row['bancos no transacao']}" if normalize_text(row["bancos no transacao"]) else "",
            f"Nº origem {row['bancos no origem']}" if normalize_text(row["bancos no origem"]) else "",
            row["bancos conta contrapartida"],
        ),
        axis=1,
    )
    filtered["bancos historico"] = (
        filtered["bancos origem"]
        + " "
        + filtered["bancos conta contrapartida"]
        + " "
        + filtered["bancos detalhes"]
        + " "
        + filtered["bancos no transacao"]
        + " "
        + filtered["bancos no origem"]
    ).map(normalize_match_text)
    filtered["bancos match_text"] = (
        filtered["bancos origem"] + " " + filtered["bancos detalhes"] + " " + filtered["bancos no origem"]
    ).map(normalize_match_text)
    filtered["bancos reembolsavel"] = filtered["bancos detalhes"].map(text_matches_receita_reembolsavel)
    filtered["bancos conta numerica"] = filtered["bancos conta contrapartida"].map(looks_like_account_only)

    return filtered[
        [
            "data referencial",
            "valor referencial",
            "bancos origem",
            "bancos conta contrapartida",
            "bancos detalhes",
            "bancos no transacao",
            "bancos no origem",
            "bancos historico",
            "bancos match_text",
            "bancos reembolsavel",
            "bancos conta numerica",
        ]
    ]


def prepare_recebimentos(path: Path, year: int, month: int) -> pd.DataFrame:
    sheet_name = find_sheet_name(path, "recebimentos", engine="pyxlsb")
    df = pd.read_excel(path, sheet_name=sheet_name, engine="pyxlsb")
    df = normalize_columns(df)

    required_columns = [
        "data recebimento",
        "valor recebido (r$)",
        "banco",
    ]
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        raise ValueError(f"Colunas obrigatorias ausentes em {path.name}: {missing}")

    df["data referencial"] = parse_excel_date_column(df["data recebimento"]).dt.normalize()
    df["valor referencial"] = pd.to_numeric(df["valor recebido (r$)"], errors="coerce").round(2)
    df["banco"] = df["banco"].astype(str)

    filtered = df[
        (df["data referencial"].dt.year == year)
        & (df["data referencial"].dt.month == month)
        & (df["valor referencial"].notna())
        & (df["banco"].str.contains("itau", case=False, na=False))
    ].copy()

    filtered["receb banco"] = filtered["banco"].fillna("").astype(str)
    filtered["receb numero nf"] = series_or_empty(filtered, "numero nf").fillna("").astype(str)
    filtered["receb clientes"] = series_or_empty(filtered, "clientes").fillna("").astype(str)
    filtered["receb clientes"] = filtered.apply(
        lambda row: fallback_text(
            row["receb clientes"],
            row["receb numero nf"],
            row["receb banco"],
        ),
        axis=1,
    )
    filtered["receb valor liquido"] = pd.to_numeric(series_or_empty(filtered, "valor liquido (r$)"), errors="coerce")
    filtered["receb historico"] = (
        filtered["receb banco"] + " " + filtered["receb numero nf"] + " " + filtered["receb clientes"]
    ).map(normalize_match_text)
    filtered["receb match_text"] = (
        filtered["receb clientes"] + " " + filtered["receb numero nf"] + " " + filtered["receb banco"]
    ).map(normalize_match_text)

    return filtered[
        [
            "data referencial",
            "valor referencial",
            "receb banco",
            "receb numero nf",
            "receb clientes",
            "receb valor liquido",
            "receb historico",
            "receb match_text",
        ]
    ]


def mark_exact_matches(bancos: pd.DataFrame, receb: pd.DataFrame) -> None:
    bancos_sorted = bancos.sort_values(["data referencial", "valor referencial"], na_position="last").copy()
    receb_sorted = receb.sort_values(["data referencial", "valor referencial"], na_position="last").copy()

    bancos_sorted["_key"] = bancos_sorted["data referencial"].dt.strftime("%Y-%m-%d") + "|" + bancos_sorted[
        "valor referencial"
    ].map(
        lambda value: f"{value:.2f}"
    )
    receb_sorted["_key"] = receb_sorted["data referencial"].dt.strftime("%Y-%m-%d") + "|" + receb_sorted[
        "valor referencial"
    ].map(
        lambda value: f"{value:.2f}"
    )

    bancos_sorted["_ord"] = bancos_sorted.groupby("_key").cumcount()
    receb_sorted["_ord"] = receb_sorted.groupby("_key").cumcount()

    matches = bancos_sorted.reset_index().merge(
        receb_sorted.reset_index(),
        on=["_key", "_ord"],
        suffixes=("_b", "_r"),
        how="inner",
    )

    bancos.loc[matches["index_b"], "_matched"] = True
    receb.loc[matches["index_r"], "_matched"] = True


def find_best_multi_perna_match(
    bank_row: pd.Series, candidate_receb: pd.DataFrame
) -> tuple[list[int], float] | None:
    if candidate_receb.shape[0] < 2:
        return None

    bank_value = float(bank_row["valor referencial"])
    bank_history = bank_row["bancos match_text"]

    candidate_receb = candidate_receb.copy()
    candidate_receb["_similaridade"] = candidate_receb["receb match_text"].map(lambda text: text_similarity(bank_history, text))
    candidate_receb = candidate_receb.sort_values(
        ["_similaridade", "valor referencial"], ascending=[False, False], na_position="last"
    ).head(MULTI_PERNA_MAX_CANDIDATES)

    best_indices: list[int] | None = None
    best_score = 0.0
    best_diff = None

    max_size = min(MULTI_PERNA_MAX_SIZE, candidate_receb.shape[0])
    for size in range(2, max_size + 1):
        for combo in combinations(candidate_receb.index.tolist(), size):
            combo_total = float(candidate_receb.loc[list(combo), "valor referencial"].sum())
            diff = abs(combo_total - bank_value)
            if diff > VALUE_TOLERANCE:
                continue

            combo_text = " ".join(candidate_receb.loc[list(combo), "receb match_text"].astype(str).tolist())
            score = max(text_similarity(bank_history, combo_text), float(candidate_receb.loc[list(combo), "_similaridade"].mean()))
            if score < SIMILARITY_THRESHOLD:
                continue

            if best_indices is None:
                best_indices = list(combo)
                best_score = score
                best_diff = diff
                continue

            assert best_diff is not None
            if score > best_score or (score == best_score and (diff < best_diff or (diff == best_diff and len(combo) < len(best_indices)))):
                best_indices = list(combo)
                best_score = score
                best_diff = diff

    if best_indices is None:
        return None

    return best_indices, best_score


def conciliate(bancos_df: pd.DataFrame, receb_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    bancos = bancos_df.copy().reset_index(drop=True)
    receb = receb_df.copy().reset_index(drop=True)

    bancos["_matched"] = False
    receb["_matched"] = False

    mark_exact_matches(bancos, receb)

    changed = True
    while changed:
        changed = False
        bancos_unmatched = bancos[~bancos["_matched"]].sort_values(
            ["data referencial", "valor referencial"], ascending=[True, False], na_position="last"
        )

        for bank_idx, bank_row in bancos_unmatched.iterrows():
            if bancos.at[bank_idx, "_matched"]:
                continue

            same_day_receb = receb[
                (~receb["_matched"])
                & (receb["data referencial"] == bank_row["data referencial"])
                & (receb["valor referencial"] <= float(bank_row["valor referencial"]) + VALUE_TOLERANCE)
            ]

            best = find_best_multi_perna_match(bank_row, same_day_receb)
            if best is None:
                continue

            matched_receb_indices, _ = best
            bancos.at[bank_idx, "_matched"] = True
            receb.loc[matched_receb_indices, "_matched"] = True
            changed = True

    bancos_unmatched = bancos[~bancos["_matched"]].copy()
    receb_unmatched = receb[~receb["_matched"]].copy()

    bancos_output = bancos_unmatched[
        [
            "data referencial",
            "valor referencial",
            "bancos origem",
            "bancos conta contrapartida",
            "bancos detalhes",
            "bancos reembolsavel",
            "bancos conta numerica",
        ]
    ].sort_values(["data referencial", "valor referencial"], na_position="last")

    receb_output = receb_unmatched[
        [
            "data referencial",
            "valor referencial",
            "receb banco",
            "receb numero nf",
            "receb clientes",
            "receb valor liquido",
        ]
    ].sort_values(["data referencial", "valor referencial"], na_position="last")

    return bancos_output, receb_output


def write_sheet(
    workbook: Workbook,
    sheet_name: str,
    headers: list[str],
    rows: pd.DataFrame,
    highlight_mode: str | None = None,
) -> None:
    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows.to_dict(orient="records"):
        worksheet.append([row.get(header, "") for header in headers])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for row_cells in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for index, header in enumerate(headers, start=1):
            cell = row_cells[index - 1]
            if "data referencial" == header:
                cell.number_format = DATE_NUMBER_FORMAT
            elif "valor referencial" == header or header == "receb valor liquido":
                cell.number_format = VALUE_NUMBER_FORMAT
            cell.alignment = Alignment(vertical="top")

        if highlight_mode == "bancos":
            row_index = row_cells[0].row - 2
            source_row = rows.iloc[row_index]
            fill = None
            if bool(source_row.get("bancos conta numerica", False)):
                fill = GRAY_FILL
            elif bool(source_row.get("bancos reembolsavel", False)):
                fill = YELLOW_FILL

            if fill is not None:
                for cell in row_cells:
                    cell.fill = fill

    widths = {
        "data referencial": 15,
        "valor referencial": 16,
        "bancos origem": 28,
        "bancos conta contrapartida": 24,
        "bancos detalhes": 40,
        "receb banco": 24,
        "receb numero nf": 18,
        "receb clientes": 32,
        "receb valor liquido": 18,
    }
    for index, header in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = widths.get(header, 18)


def save_output(output_path: Path, bancos_output: pd.DataFrame, receb_output: pd.DataFrame) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    write_sheet(workbook, SHEET_BANCOS, OUTPUT_COLUMNS_BANCOS, bancos_output, highlight_mode="bancos")
    write_sheet(workbook, SHEET_RECEBIMENTOS, OUTPUT_COLUMNS_RECEB, receb_output)

    workbook.save(output_path)


def run(bancos_path: Path, recebimentos_path: Path, output_path: Path | None, year: int, month: int) -> Path:
    bancos_df = prepare_bancos_itau(bancos_path, year, month)
    receb_df = prepare_recebimentos(recebimentos_path, year, month)

    bancos_output, receb_output = conciliate(bancos_df, receb_df)

    if output_path is None:
        output_path = bancos_path.parent / f"COMPARATIVO_ITAU_RECEBIMENTOS_{year:04d}-{month:02d}.xlsx"

    save_output(output_path, bancos_output, receb_output)
    return output_path


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Concilia entradas do ITAU (BANCOS SW) com a aba RECEBIMENTOS (STURMER)."
    )
    parser.add_argument(
        "--bancos",
        type=Path,
        default=None,
        help="Caminho da planilha BANCOS SW (.xlsx). Se nao informar, abre janela para selecionar.",
    )
    parser.add_argument(
        "--recebimentos",
        type=Path,
        default=None,
        help="Caminho da planilha STURMER RECEBIMENTOS (.xlsb). Se nao informar, abre janela para selecionar.",
    )
    parser.add_argument(
        "--saida",
        type=Path,
        default=None,
        help="Caminho do arquivo de saida (.xlsx). Se nao informar, abre janela para selecionar.",
    )
    return parser


def main() -> None:
    parser = build_argument_parser()
    args = parser.parse_args()

    default_dir = Path(__file__).resolve().parent

    bancos_path = args.bancos
    if bancos_path is None:
        bancos_path = select_input_file(
            title="Selecione a planilha BANCOS SW",
            initial_dir=default_dir,
            filetypes=[("Planilhas Excel", "*.xlsx *.xlsm *.xls"), ("Todos os arquivos", "*.*")],
        )

    recebimentos_path = args.recebimentos
    if recebimentos_path is None:
        recebimentos_path = select_input_file(
            title="Selecione a planilha de recebimentos",
            initial_dir=bancos_path.parent,
            filetypes=[("Planilhas Excel Binarias", "*.xlsb"), ("Todos os arquivos", "*.*")],
        )

    year, month = select_processing_period()

    output_path = args.saida
    if output_path is None:
        default_name = f"COMPARATIVO_ITAU_RECEBIMENTOS_{year:04d}-{month:02d}.xlsx"

        output_path = select_output_file(
            title="Salvar comparativo como",
            initial_dir=bancos_path.parent,
            default_filename=default_name,
        )

    output_path = run(bancos_path, recebimentos_path, output_path, year, month)
    print(f"Comparativo gerado em: {output_path}")


if __name__ == "__main__":
    main()
