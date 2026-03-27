import argparse
import datetime as dt
import html
import json
import logging
import os
import re
import shutil
import sys
import time
import traceback
import uuid
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import load_workbook, Workbook

# ----------------------------
# Utilidades gerais
# python .\receitanetbx_automation_runtime_fixed4.py --config .\config_runtime_doc.json
# ----------------------------

def now_ts() -> str:
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def resolve_config_path(explicit_path: str | None, script_dir: str) -> str:
    """Resolve o caminho do config.json.

    Ordem:
      1) --config (se informado)
      2) variável de ambiente RECEITANET_CONFIG (se existir)
      3) config_runtime_doc_v5.json (na pasta do script)
      4) config.json (na pasta do script)
      5) primeiro arquivo *.json (na pasta do script)
    """
    if explicit_path:
        return explicit_path

    env_path = os.environ.get("RECEITANET_CONFIG")
    if env_path:
        return env_path

    candidates = [
        os.path.join(script_dir, "config_runtime_doc_v5.json"),
        os.path.join(script_dir, "config.json"),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p

    try:
        for name in sorted(os.listdir(script_dir)):
            if name.lower().endswith(".json"):
                return os.path.join(script_dir, name)
    except Exception:
        pass

    raise FileNotFoundError(
        "Não encontrei config JSON. Coloque um 'config_runtime_doc_v5.json' ou 'config.json' na mesma pasta do script, "
        "ou informe --config, ou defina RECEITANET_CONFIG."
    )


def prompt_if_missing(args):
    """Quando rodar sem argumentos (F5 no VS Code), pergunta sistema e datas."""
    if args.sistema is None:
        args.sistema = int(input("Sistema (ex.: 1, 2, 7, 20): ").strip())
    if args.data_inicio is None:
        args.data_inicio = input("Data início (YYYY-MM-DD): ").strip()
    if args.data_fim is None:
        args.data_fim = input("Data fim (YYYY-MM-DD): ").strip()
    return args

def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)

def sanitize_cnpj(value: str) -> str:
    if value is None:
        return ""
    digits = re.sub(r"\D", "", str(value))
    return digits.zfill(14)

def is_matriz(cnpj14: str) -> bool:
    # CNPJ: 8 (raiz) + 4 (filial) + 2 (dv). Matriz costuma ser filial 0001.
    if len(cnpj14) != 14:
        return False
    return cnpj14[8:12] == "0001"

def safe_filename(name: str) -> str:
    name = re.sub(r"[<>:\"/\\|?*\x00-\x1F]", "_", name)
    name = name.strip().strip(".")
    return name or "arquivo"

def indent_xml(elem: ET.Element, level: int = 0) -> None:
    # Só para debug/dump legível
    i = "\n" + level * "  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "  "
        for child in elem:
            indent_xml(child, level + 1)
        if not child.tail or not child.tail.strip():
            child.tail = i
    if level and (not elem.tail or not elem.tail.strip()):
        elem.tail = i

# ----------------------------
# Config
# ----------------------------

@dataclass
class SoapConfig:
    namespace: str
    prefix: str = "ws"
    entrada_param: str = "entrada"
    cdata_entrada: bool = True
    use_soap_action: bool = True
    soap_action_prefix: str = ""  # ex: "urn:..." se necessário

@dataclass
class WsConfig:
    endpoint_url: str
    timeout_sec: int = 30
    verify_tls: bool = False
    soap: SoapConfig = None

@dataclass
class EfdIcmsIpiConfig:
    enabled: bool = False
    campo_todos_estabelecimentos: str = "Buscar Arquivos de Todos os Estabelecimentos"
    campo_ultimo_transmitido: str = "Último arquivo transmitido"
    campo_cnpj_estabelecimento: str = "CNPJ do Estabelecimento"
    campo_inscricao_estadual: str = "Inscrição Estadual"
    marcar_todos_estabelecimentos_quando_matriz: bool = True
    ultimo_transmitido_valor: str = "F"  # pegar todos do período

@dataclass
class JobConfig:
    xml_mode: str = ""
    cliente: Optional[Dict[str, Any]] = None
    perfil: str = "Procurador"
    tiponirepresentado: str = "cnpj"
    sistema: str = ""
    tipoarquivo: str = ""
    tipopesquisa: str = ""
    campos: List[Dict[str, str]] = None
    efd_icms_ipi: EfdIcmsIpiConfig = None

@dataclass
class PollingConfig:
    interval_sec: int = 60
    timeout_sec: int = 1200

@dataclass
class IoConfig:
    input_xlsx: str = "cnpjs.xlsx"
    output_dir: str = "saida"
    copy_downloaded_files: bool = True

@dataclass
class AppConfig:
    ws: WsConfig
    job: JobConfig
    polling: PollingConfig
    io: IoConfig

def load_config(path: str) -> AppConfig:
    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    soap = raw["ws"].get("soap", {})
    soap_cfg = SoapConfig(
        namespace=soap.get("namespace", "urn:ReceitanetBX_WSCliente"),
        prefix=soap.get("prefix", "ws"),
        entrada_param=soap.get("entrada_param", "entrada"),
        cdata_entrada=bool(soap.get("cdata_entrada", True)),
        use_soap_action=bool(soap.get("use_soap_action", True)),
        soap_action_prefix=soap.get("soap_action_prefix", ""),
    )

    ws_cfg = WsConfig(
        endpoint_url=raw["ws"]["endpoint_url"],
        timeout_sec=int(raw["ws"].get("timeout_sec", 30)),
        verify_tls=bool(raw["ws"].get("verify_tls", False)),
        soap=soap_cfg
    )

    efd_raw = (raw.get("job", {}) or {}).get("efd_icms_ipi", {}) or {}
    efd_cfg = EfdIcmsIpiConfig(
        enabled=bool(efd_raw.get("enabled", False)),
        campo_todos_estabelecimentos=efd_raw.get("campo_todos_estabelecimentos", "Buscar Arquivos de Todos os Estabelecimentos"),
        campo_ultimo_transmitido=efd_raw.get("campo_ultimo_transmitido", "Último arquivo transmitido"),
        campo_cnpj_estabelecimento=efd_raw.get("campo_cnpj_estabelecimento", "CNPJ do Estabelecimento"),
        campo_inscricao_estadual=efd_raw.get("campo_inscricao_estadual", "Inscrição Estadual"),
        marcar_todos_estabelecimentos_quando_matriz=bool(efd_raw.get("marcar_todos_estabelecimentos_quando_matriz", True)),
        ultimo_transmitido_valor=efd_raw.get("ultimo_transmitido_valor", "F"),
    )

    job_raw = raw.get("job", {}) or {}
    job_cfg = JobConfig(
        xml_mode=job_raw.get("xml_mode", ""),
        cliente=job_raw.get("cliente"),
        perfil=job_raw.get("perfil", "Procurador"),
        tiponirepresentado=job_raw.get("tiponirepresentado", "cnpj"),
        sistema=job_raw.get("sistema", ""),
        tipoarquivo=job_raw.get("tipoarquivo", ""),
        tipopesquisa=job_raw.get("tipopesquisa", ""),
        campos=job_raw.get("campos", []) or [],
        efd_icms_ipi=efd_cfg,
    )

    poll_raw = raw.get("polling", {}) or {}
    polling_cfg = PollingConfig(
        interval_sec=int(poll_raw.get("interval_sec", 60)),
        timeout_sec=int(poll_raw.get("timeout_sec", 1200)),
    )

    io_raw = raw.get("io", {}) or {}
    io_cfg = IoConfig(
        input_xlsx=io_raw.get("input_xlsx", "cnpjs.xlsx"),
        output_dir=io_raw.get("output_dir", "saida"),
        copy_downloaded_files=bool(io_raw.get("copy_downloaded_files", True)),
    )

    cfg = AppConfig(ws=ws_cfg, job=job_cfg, polling=polling_cfg, io=io_cfg)
    # guarda raw para facilitar sobrescritas em runtime
    setattr(cfg, 'raw', raw)
    return cfg

# ----------------------------
# Logging + Debug dumps
# ----------------------------

class DebugDumper:
    def __init__(self, base_dir: str):
        self.base_dir = base_dir
        ensure_dir(self.base_dir)

    def dump(self, kind: str, content: str, meta: Dict[str, Any]) -> str:
        ts = now_ts()
        suffix = safe_filename(kind)
        uid = uuid.uuid4().hex[:10]
        filename = f"{ts}_{suffix}_{uid}.txt"
        path = os.path.join(self.base_dir, filename)
        header = json.dumps(meta, ensure_ascii=False, indent=2)
        with open(path, "w", encoding="utf-8") as f:
            f.write(header)
            f.write("\n\n")
            f.write(content)
        return path

def setup_logger(out_dir: str) -> logging.Logger:
    ensure_dir(out_dir)
    logger = logging.getLogger("receitanetbx")
    logger.setLevel(logging.DEBUG)

    fmt = logging.Formatter(
        fmt="%(asctime)s.%(msecs)03d %(levelname)s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )

    # Console (INFO)
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    # File (DEBUG)
    fh = logging.FileHandler(os.path.join(out_dir, f"run_{now_ts()}.log"), encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    return logger


def resolve_placeholders(template: str, *, cnpj: str = "", row: Optional[Dict[str, Any]] = None,
                         runtime_vars: Optional[Dict[str, str]] = None) -> str:
    """Substitui placeholders simples no formato {{chave}}.
    - Sempre suporta {{cnpj}}
    - Suporta chaves vindas do XLSX (row) e do runtime (runtime_vars)
    """
    out = str(template)
    out = out.replace("{{cnpj}}", cnpj or "")
    if runtime_vars:
        for k, v in runtime_vars.items():
            out = out.replace(f"{{{{{k}}}}}", str(v))
    if row:
        for k, v in row.items():
            if v is None:
                continue
            out = out.replace(f"{{{{{k}}}}}", str(v))
    return out


def build_pesquisa_xml_cliente(cliente_cfg: Dict[str, Any], cnpj: str, runtime_vars: Optional[Dict[str, str]] = None, row: Optional[Dict[str, Any]] = None) -> str:
    root = ET.Element("pesquisa", attrib={
        "id": cliente_cfg["idpesquisa"],
        "idtipo": str(cliente_cfg["idtipo"])
    })

    ucfg = dict(cliente_cfg.get("usuario", {}))
    for k, v in list(ucfg.items()):
        if isinstance(v, str):
            ucfg[k] = v.replace("{{cnpj}}", cnpj)
    ET.SubElement(root, "usuario", attrib=ucfg)

    for c in cliente_cfg.get("campos", []):
        ET.SubElement(root, "campo", attrib={"id": c["id"], "valor": resolve_placeholders(c["valor"], cnpj=cnpj, row=row, runtime_vars=runtime_vars)})

    indent_xml(root)
    return ET.tostring(root, encoding="utf-8").decode("utf-8")

def build_pedido_xml_cliente(cliente_cfg: Dict[str, Any], cnpj: str, arquivo_ids: List[str], runtime_vars: Optional[Dict[str, str]] = None, row: Optional[Dict[str, Any]] = None) -> str:
    root = ET.Element("pedido", attrib={
        "id": "0",
        "idtipo": str(cliente_cfg["idtipo"]),
        "idpesquisa": cliente_cfg["idpesquisa"]
    })

    ucfg = dict(cliente_cfg.get("usuario", {}))
    for k, v in list(ucfg.items()):
        if isinstance(v, str):
            ucfg[k] = v.replace("{{cnpj}}", cnpj)
    ET.SubElement(root, "usuario", attrib=ucfg)

    arqs = ET.SubElement(root, "arquivos")
    for aid in arquivo_ids:
        ET.SubElement(arqs, "arquivo", attrib={"id": str(aid)})

    indent_xml(root)
    return ET.tostring(root, encoding="utf-8").decode("utf-8")

# ----------------------------
# XML builders / parsers conforme doc
# ----------------------------

def build_identificacao(job: JobConfig, nirepresentado: str) -> Dict[str, str]:
    return {
        "perfil": job.perfil,
        "sistema": job.sistema,
        "tipoarquivo": job.tipoarquivo,
        "tipopesquisa": job.tipopesquisa,
        "nirepresentado": nirepresentado,
        "tiponirepresentado": job.tiponirepresentado
    }

def build_pesquisa_xml(ident: Dict[str, str], campos: List[Dict[str, str]]) -> str:
    root = ET.Element("pesquisa")
    ET.SubElement(root, "identificacao", attrib=ident)
    for c in campos:
        ET.SubElement(root, "campo", attrib={"nome": c["nome"], "valor": c["valor"]})
    indent_xml(root)
    return ET.tostring(root, encoding="utf-8", xml_declaration=False).decode("utf-8")

def build_pedido_xml(ident: Dict[str, str], arquivo_ids: List[str],
                     termo_campos: Optional[List[Dict[str, str]]] = None,
                     pesquisa_campos: Optional[List[Dict[str, str]]] = None) -> str:
    """
    Conforme doc: <pedido> pode ter <termo> e/ou <pesquisa> e/ou <arquivos>,
    mas <pesquisa> não pode coexistir com <arquivos> (e vice-versa). :contentReference[oaicite:4]{index=4}
    """
    if pesquisa_campos and arquivo_ids:
        raise ValueError("Config inválida: não use <pesquisa> junto com <arquivos> no pedido.")

    root = ET.Element("pedido")
    ET.SubElement(root, "identificacao", attrib=ident)

    if termo_campos:
        termo = ET.SubElement(root, "termo")
        for c in termo_campos:
            ET.SubElement(termo, "campo", attrib={"nome": c["nome"], "valor": c["valor"]})

    if pesquisa_campos:
        pesq = ET.SubElement(root, "pesquisa")
        for c in pesquisa_campos:
            ET.SubElement(pesq, "campo", attrib={"nome": c["nome"], "valor": c["valor"]})

    if arquivo_ids:
        arquivos = ET.SubElement(root, "arquivos")
        for aid in arquivo_ids:
            ET.SubElement(arquivos, "arquivo", attrib={"id": str(aid)})

    indent_xml(root)
    return ET.tostring(root, encoding="utf-8").decode("utf-8")

def build_verificar_pedidos_xml(pedido_ids: List[int], atributos: bool = True) -> str:
    root = ET.Element("pedidos", attrib={"atributos": "true" if atributos else "false"})
    for pid in pedido_ids:
        ET.SubElement(root, "pedido", attrib={"id": str(pid)})
    indent_xml(root)
    return ET.tostring(root, encoding="utf-8").decode("utf-8")

def parse_msg(root: ET.Element) -> str:
    m = root.find(".//mensagem")
    if m is not None and m.text:
        return m.text.strip()
    mu = root.find(".//mensagemusuario")
    if mu is not None and (mu.text or "").strip():
        return (mu.text or "").strip()
    return ""

def parse_retorno_pesquisa(saida_xml: str) -> Tuple[List[str], str]:
    """
    Esperado:
    <retornopesquisa>
      <arquivos><arquivo id="..."/></arquivos>
      <mensagem>...</mensagem>
    </retornopesquisa>
    :contentReference[oaicite:5]{index=5}
    """
    ids: List[str] = []
    msg = ""
    root = ET.fromstring(saida_xml)
    for a in root.findall(".//arquivo"):
        aid = a.attrib.get("id")
        if aid:
            ids.append(aid)
    m = root.find(".//mensagem")
    if m is not None and m.text:
        msg = m.text.strip()
    return ids, msg

def parse_retorno_pedido(saida_xml: str) -> Tuple[int, str]:
    """
    Esperado:
    <retornopedido id="..."><mensagem>...</mensagem></retornopedido>
    id=0 em erro. :contentReference[oaicite:6]{index=6}
    """
    root = ET.fromstring(saida_xml)
    pid = int(root.attrib.get("id", "0"))
    msg = ""
    m = root.find(".//mensagem")
    if m is not None and m.text:
        msg = m.text.strip()
    return pid, msg

def parse_retorno_pedidos(saida_xml: str) -> Dict[int, Dict[str, Any]]:
    """
    Estrutura documentada em VerificarSituacaoPedidos. :contentReference[oaicite:7]{index=7}
    Retorna dict pedido_id -> dict(status, mensagem, arquivos[])
    """
    root = ET.fromstring(saida_xml)
    result: Dict[int, Dict[str, Any]] = {}
    for p in root.findall(".//pedido"):
        pid = int(p.attrib.get("id", "0"))
        if pid <= 0:
            continue
        pdata: Dict[str, Any] = {
            "situacao": p.attrib.get("situacao", ""),
            "mensagem": p.attrib.get("mensagem", ""),
            "sistema": p.attrib.get("sistema", ""),
            "tipoarquivo": p.attrib.get("tipoarquivo", ""),
            "datasolicitacao": p.attrib.get("datasolicitacao", ""),
            "dataprevista": p.attrib.get("dataprevista", ""),
            "arquivos": []
        }
        for a in p.findall(".//arquivo"):
            ad = dict(a.attrib)
            # atributos nome/valor/tipo
            attrs = []
            for at in a.findall(".//atributo"):
                attrs.append(dict(at.attrib))
            ad["atributos"] = attrs
            pdata["arquivos"].append(ad)
        result[pid] = pdata
    return result

# ----------------------------
# SOAP transport (genérico, ajustável)
# ----------------------------

class WsTransportError(RuntimeError):
    pass

class SoapTransport:
    """
    Implementa um cliente SOAP 1.1 genérico, mas com ponto de ajuste
    (namespace, nome do parâmetro de entrada, SOAPAction, etc).

    Como ainda não temos WSDL/endpoints, isso é o “esqueleto executável”.
    Quando acharmos o endpoint correto, você só ajusta config.
    """
    SOAPENV = "http://schemas.xmlsoap.org/soap/envelope/"

    def __init__(self, ws: WsConfig, logger: logging.Logger, dumper: DebugDumper):
        self.ws = ws
        self.log = logger
        self.dump = dumper
        self.session = requests.Session()

    def _build_envelope(self, method: str, entrada_xml: str) -> str:
        ns = self.ws.soap.namespace
        prefix = self.ws.soap.prefix
        entrada_param = self.ws.soap.entrada_param

        env = ET.Element(ET.QName(self.SOAPENV, "Envelope"), attrib={
            f"xmlns:{prefix}": ns
        })
        ET.SubElement(env, ET.QName(self.SOAPENV, "Header"))
        body = ET.SubElement(env, ET.QName(self.SOAPENV, "Body"))

        op = ET.SubElement(body, ET.QName(ns, method))  # <ws:PesquisarArquivos>...
        arg = ET.SubElement(op, ET.QName(ns, entrada_param))

        if self.ws.soap.cdata_entrada:
            # ElementTree não suporta CDATA direto; fazemos “manual” depois.
            arg.text = f"__CDATA_PLACEHOLDER__{uuid.uuid4().hex}__"
            placeholder = arg.text
            indent_xml(env)
            raw = ET.tostring(env, encoding="utf-8", xml_declaration=True).decode("utf-8")
            cdata = f"<![CDATA[{entrada_xml}]]>"
            return raw.replace(placeholder, cdata)
        else:
            # envia como texto (escapado)
            arg.text = entrada_xml
            indent_xml(env)
            return ET.tostring(env, encoding="utf-8", xml_declaration=True).decode("utf-8")

    def _parse_soap_response(self, soap_xml: str) -> Tuple[int, str]:
        """
        Heurística:
        - procura um elemento que pareça 'return' com dígitos
        - procura um elemento 'saida' (ou termina com 'saida') com XML dentro (texto escapado ou CDATA)
        """
        try:
            root = ET.fromstring(soap_xml)
        except Exception as e:
            raise WsTransportError(f"Resposta SOAP inválida (não é XML): {e}") from e

        # Busca por 'return' em qualquer namespace
        ret_val = None
        for el in root.iter():
            tag = el.tag
            local = tag.split("}")[-1] if "}" in tag else tag
            if local.lower() in ("return", "resultado", "result", "retorno"):
                if el.text and el.text.strip().isdigit():
                    ret_val = int(el.text.strip())
                    break

        # Busca por 'saida'
        saida_text = None
        for el in root.iter():
            tag = el.tag
            local = tag.split("}")[-1] if "}" in tag else tag
            if local.lower() == "saida" or local.lower().endswith("saida"):
                if el.text:
                    saida_text = el.text
                    break

        if ret_val is None:
            # fallback: tenta achar qualquer texto dígito único
            for el in root.iter():
                if el.text and el.text.strip().isdigit():
                    ret_val = int(el.text.strip())
                    break

        if ret_val is None:
            raise WsTransportError("Não consegui extrair o retorno int (0/1) do SOAP. Ajuste parsing/namespace/endpoint.")

        if saida_text is None:
            # Pode ocorrer se o WSDL expõe o retorno de 'saida' como outro nome.
            # Nesse caso, devolvemos string vazia e você vê o dump.
            saida_text = ""

        # Se veio XML escapado (&lt;retornopesquisa&gt;...), desfaz
        saida_text = html.unescape(saida_text).strip()

        # Se por acaso vier com BOM/ruídos, limpamos o começo
        saida_text = saida_text.lstrip("\ufeff")

        return ret_val, saida_text

    def call(self, method: str, entrada_xml: str) -> Tuple[int, str]:
        soap_env = self._build_envelope(method, entrada_xml)

        headers = {
            "Content-Type": "text/xml; charset=utf-8",
        }
        if self.ws.soap.use_soap_action:
            # muitos serviços SOAP 1.1 exigem SOAPAction; prefixo configurável
            headers["SOAPAction"] = f"{self.ws.soap.soap_action_prefix}{method}"

        meta = {
            "endpoint_url": self.ws.endpoint_url,
            "method": method,
            "timeout_sec": self.ws.timeout_sec
        }
        req_dump_path = self.dump.dump(f"request_{method}", soap_env, meta)
        self.log.debug(f"[DUMP] request salvo em: {req_dump_path}")

        try:
            r = self.session.post(
                self.ws.endpoint_url,
                data=soap_env.encode("utf-8"),
                headers=headers,
                timeout=self.ws.timeout_sec,
                verify=self.ws.verify_tls,
            )
        except Exception as e:
            raise WsTransportError(f"Falha HTTP ao chamar {method}: {e}") from e

        resp_text = r.text or ""
        resp_meta = {
            **meta,
            "http_status": r.status_code,
            "content_type": r.headers.get("Content-Type", "")
        }
        resp_dump_path = self.dump.dump(f"response_{method}", resp_text, resp_meta)
        self.log.debug(f"[DUMP] response salvo em: {resp_dump_path}")

        return self._parse_soap_response(resp_text)

# ----------------------------
# Camada de negócio (Pesquisar -> Solicitar -> Verificar)
# ----------------------------

class ReceitanetBXClient:
    def __init__(self, cfg: AppConfig, transport: SoapTransport, logger: logging.Logger):
        self.cfg = cfg
        self.t = transport
        self.log = logger

    def pesquisar_arquivos(self, cnpj: str, campos: List[Dict[str, str]]) -> Tuple[List[str], str]:
        # se xml_mode=cliente, ignora build_identificacao/build_pesquisa_xml
        if getattr(self.cfg.job, "xml_mode", "") == "cliente":
            cliente_cfg = getattr(self.cfg.job, "cliente", None)
            if cliente_cfg is None:
                raise WsTransportError("Config job.cliente ausente no config.json")
            entrada = build_pesquisa_xml_cliente(cliente_cfg, cnpj, runtime_vars=getattr(self.cfg, 'runtime_vars', None), row=getattr(self.cfg, 'current_row', None))
        else:
            ident = build_identificacao(self.cfg.job, cnpj)
            entrada = build_pesquisa_xml(ident, campos)

        ret, saida = self.t.call("PesquisarArquivos", entrada)
        if ret == 0:
            raise WsTransportError(f"PesquisarArquivos retornou 0. Saída: {saida}")
        ids, msg = parse_retorno_pesquisa(saida)
        return ids, msg

    def solicitar_arquivos_por_ids(self, cnpj: str, arquivo_ids: List[str]) -> Tuple[int, str]:
        if getattr(self.cfg.job, "xml_mode", "") == "cliente":
            cliente_cfg = getattr(self.cfg.job, "cliente", None)
            if cliente_cfg is None:
                raise WsTransportError("Config job.cliente ausente no config.json")
            entrada = build_pedido_xml_cliente(cliente_cfg, cnpj, arquivo_ids, runtime_vars=getattr(self.cfg, 'runtime_vars', None), row=getattr(self.cfg, 'current_row', None))
        else:
            ident = build_identificacao(self.cfg.job, cnpj)
            entrada = build_pedido_xml(ident, arquivo_ids=arquivo_ids)

        ret, saida = self.t.call("SolicitarArquivos", entrada)
        if ret == 0:
            raise WsTransportError(f"SolicitarArquivos retornou 0. Saída: {saida}")
        pid, msg = parse_retorno_pedido(saida)
        return pid, msg

    def verificar_pedidos(self, pedido_ids: List[int], atributos: bool = True) -> Dict[int, Dict[str, Any]]:
        entrada = build_verificar_pedidos_xml(pedido_ids, atributos=atributos)
        ret, saida = self.t.call("VerificarSituacaoPedidos", entrada)
        if ret == 0:
            raise WsTransportError(f"VerificarSituacaoPedidos retornou 0. Saída: {saida}")
        return parse_retorno_pedidos(saida)

# ----------------------------
# IO: XLSX
# ----------------------------

def read_cnpjs_xlsx(path: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path)
    ws = wb.active
    headers = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            headers = [str(x).strip().lower() if x else "" for x in row]
            continue
        if not any(row):
            continue
        item = {}
        for h, v in zip(headers, row):
            if not h:
                continue
            item[h] = v
        rows.append(item)
    return rows

def write_result_xlsx(path: str, resumo_rows: List[Dict[str, Any]], arquivo_rows: List[Dict[str, Any]]) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumo"
    ws2 = wb.create_sheet("Arquivos")

    def write_sheet(ws, rows: List[Dict[str, Any]]):
        if not rows:
            ws.append(["(sem dados)"])
            return
        cols = list(rows[0].keys())
        ws.append(cols)
        for r in rows:
            ws.append([r.get(c, "") for c in cols])

    write_sheet(ws1, resumo_rows)
    write_sheet(ws2, arquivo_rows)

    ensure_dir(os.path.dirname(path) or ".")
    wb.save(path)

# ----------------------------
# Montagem de campos por CNPJ (inclui regra matriz EFD ICMS/IPI)
# ----------------------------

def build_campos_for_cnpj(cfg: AppConfig, row: Dict[str, Any], cnpj: str, runtime_vars: Optional[Dict[str, str]] = None) -> List[Dict[str, str]]:
    campos = []
    # Copia do config base
    for c in cfg.job.campos:
        nome = c.get("nome", "")
        valor_tmpl = str(c.get("valor", ""))
        valor = resolve_placeholders(valor_tmpl, cnpj=cnpj, row=row, runtime_vars=runtime_vars)
        campos.append({"nome": nome, "valor": valor})

    # Regras EFD ICMS/IPI (opcional)
    efd = cfg.job.efd_icms_ipi
    if efd and efd.enabled:
        matriz = is_matriz(cnpj)
        if efd.marcar_todos_estabelecimentos_quando_matriz:
            campos.append({
                "nome": efd.campo_todos_estabelecimentos,
                "valor": "V" if matriz else "F"
            })
        # Sempre pegar todos do período
        campos.append({
            "nome": efd.campo_ultimo_transmitido,
            "valor": efd.ultimo_transmitido_valor  # "F"
        })

        # Se não for matriz, normalmente precisa indicar estabelecimento/IE (dependendo do sistema fim).
        if not matriz:
            campos.append({"nome": efd.campo_cnpj_estabelecimento, "valor": cnpj})
            ie = row.get("ie")
            if ie:
                campos.append({"nome": efd.campo_inscricao_estadual, "valor": str(ie)})

    return campos

# ----------------------------
# Execução principal
# ----------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=False, help="Caminho para config.json (opcional; autodetecta na pasta do script)")
    ap.add_argument("--dry-run", action="store_true", help="Não chama o WS; só valida leitura e monta XMLs")
    ap.add_argument("--debug", action="store_true", help="Gera dumps XML (request/response) em ./saida (sem subpastas por padrão)")

    # Runtime params (podem sobrescrever o config.json)
    ap.add_argument("--sistema", help="Código do sistema (conforme opção usada no Cliente ReceitanetBX / manual)")
    ap.add_argument("--data-inicio", dest="data_inicio", help="Data inicial (YYYY-MM-DD)")
    ap.add_argument("--data-fim", dest="data_fim", help="Data final (YYYY-MM-DD)")
    ap.add_argument("--extra", action="append", default=[],
                    help="Campos extras no formato chave=valor (repita para vários). Ex.: --extra cnpjEstabelecimento=61042627000108")
    args = ap.parse_args()
    # Se rodar no VS Code (sem args), faz prompt no console.
    if len(sys.argv) == 1:
        args = prompt_if_missing(args)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = resolve_config_path(args.config, script_dir)
    cfg = load_config(config_path)


    # ----------------------------
    # Resolve parâmetros em tempo de execução (sistema + período + extras)
    # ----------------------------

    def _prompt(label: str, default: str = "") -> str:
        suf = f" [{default}]" if default else ""
        v = input(f"{label}{suf}: ").strip()
        return v or default

    # Seleção do sistema: usa args, senão config.job.sistema_default, senão pergunta
    sistema_default = ""
    try:
        sistema_default = str(getattr(cfg.job, "sistema_default", "") or "")
    except Exception:
        sistema_default = ""
    sistema = (str(args.sistema).strip() if args.sistema is not None else "") or _prompt("Sistema (ex.: 1, 2, 7, 20)", sistema_default)
    if not sistema:
        raise SystemExit("Sistema não informado.")

    # Período: sempre solicitado, pois muda a cada execução
    data_inicio = (str(args.data_inicio).strip() if args.data_inicio is not None else "") or _prompt("Data início (YYYY-MM-DD)")
    data_fim = (str(args.data_fim).strip() if args.data_fim is not None else "") or _prompt("Data fim (YYYY-MM-DD)")

    # Extras (chave=valor)
    extras: Dict[str, str] = {}
    for kv in args.extra or []:
        if "=" not in kv:
            raise SystemExit(f"--extra inválido (use chave=valor): {kv}")
        k, v = kv.split("=", 1)
        extras[k.strip()] = v.strip()

    # Config: job.systems define idtipo/idpesquisa/campos por sistema (cliente mode).
    # Se não existir job.systems, usa o job.cliente atual (compatibilidade).
    systems_map = {}
    try:
        systems_map = cfg.raw.get("job", {}).get("systems", {})  # type: ignore[attr-defined]
    except Exception:
        systems_map = {}

    if systems_map:
        sys_cfg = systems_map.get(str(sistema))
        if not sys_cfg:
            raise SystemExit(f"Sistema '{sistema}' não encontrado em job.systems do config.json.")

        # Decide o modo de XML pelo conteúdo do sistema selecionado:
        # - "doc": usa <identificacao ...> e <campo nome="...">
        # - "cliente": replica o XML observado no ReceitanetBX Cliente (<pesquisa ...><campo id="...">)
        if isinstance(sys_cfg, dict) and all(k in sys_cfg for k in ("perfil", "sistema", "tipoarquivo", "tipopesquisa")):
            cfg.job.xml_mode = "doc"
            cfg.job.perfil = sys_cfg["perfil"]
            cfg.job.sistema = sys_cfg["sistema"]
            cfg.job.tipoarquivo = sys_cfg["tipoarquivo"]
            cfg.job.tipopesquisa = sys_cfg["tipopesquisa"]

            # Campos podem vir como {"nome":..} ou {"id":..}; normaliza para {"nome":..}
            norm_campos = []
            for c in (sys_cfg.get("campos") or []):
                if "nome" in c:
                    norm_campos.append({"nome": c["nome"], "valor": c.get("valor", "")})
                elif "id" in c:
                    norm_campos.append({"nome": c["id"], "valor": c.get("valor", "")})
            cfg.job.campos = norm_campos


            # Normalização: alguns ambientes usam tipopesquisa="exercicio" e esperam 1 campo (ano),
            # e não período. Se vierem campos de período, converte para campo "Exercício".
            try:
                if str(cfg.job.tipopesquisa).strip().lower() == "exercicio":
                    nomes = [c.get("nome","").strip().lower() for c in (cfg.job.campos or [])]
                    if any(n in ("data início","data inicio","data inicial","data fim","data final") for n in nomes):
                        ano = runtime_vars.get("exercicio","")
                        cfg.job.campos = [{"nome": "Exercício", "valor": f"{{{{exercicio}}}}"}]
            except Exception:
                pass
            # Cliente mode não é usado aqui
            cfg.job.cliente = None  # type: ignore[assignment]
        else:
            cfg.job.xml_mode = "cliente"
            cfg.job.cliente = sys_cfg  # type: ignore[assignment]
    else:
        # Modo compatível: cliente já fixo no config.json
        sys_cfg = getattr(cfg.job, "cliente", None)
        if sys_cfg is None:
            raise SystemExit("Config inválida: informe job.systems ou job.cliente.")

    # Variáveis runtime para substituir placeholders em campos ({{dataInicio}}, {{dataFim}}, etc.)
    runtime_vars: Dict[str, str] = {"dataInicio": data_inicio, "dataFim": data_fim, "exercicio": (data_inicio[:4] if data_inicio else ""), "sistema": str(sistema), **extras}
    setattr(cfg, "runtime_vars", runtime_vars)

    # Se o sistema selecionado exigir campos extras (placeholders), pede ao usuário agora (exceto se vierem do XLSX)
    try:
        _campos_cfg = (sys_cfg or {}).get("campos", []) if isinstance(sys_cfg, dict) else []
        needed_keys = set()
        for c in _campos_cfg:
            val = str(c.get("valor", ""))
            for k in re.findall(r"\{\{([A-Za-z0-9_]+)\}\}", val):
                needed_keys.add(k)
        # chaves que podem vir do XLSX: se existir coluna, será resolvido por linha
        xlsx_keys = set()
        # (ainda não lemos o XLSX aqui, então só ignora chaves padrão)
        ignore = {"cnpj", "dataInicio", "dataFim", "sistema"}
        for k in sorted(needed_keys - ignore):
            if k not in runtime_vars:
                runtime_vars[k] = _prompt(f"Valor para campo '{k}'")
        setattr(cfg, "runtime_vars", runtime_vars)
    except Exception:
        pass

    setattr(cfg, "runtime_vars", runtime_vars)


    out_dir = cfg.io.output_dir
    ensure_dir(out_dir)
    logger = setup_logger(out_dir)
    dumper = DebugDumper(os.path.join(out_dir, "debug"))

    logger.info("Iniciando automação ReceitanetBX")
    logger.info(f"Input XLSX: {cfg.io.input_xlsx}")
    logger.info(f"Output dir: {cfg.io.output_dir}")
    logger.info(f"WS endpoint_url: {cfg.ws.endpoint_url}")

    rows = read_cnpjs_xlsx(cfg.io.input_xlsx)
    if not rows:
        logger.error("Nenhum CNPJ encontrado no XLSX.")
        sys.exit(2)

    transport = SoapTransport(cfg.ws, logger, dumper)
    client = ReceitanetBXClient(cfg, transport, logger)

    resumo: List[Dict[str, Any]] = []
    arquivos_out: List[Dict[str, Any]] = []

    result_xlsx = os.path.join(out_dir, f"resultado_{now_ts()}.xlsx")

    for idx, row in enumerate(rows, start=1):
        setattr(cfg, 'current_row', row)
        raw_cnpj = row.get("cnpj", "")
        cnpj = sanitize_cnpj(str(raw_cnpj))
        if not cnpj or len(cnpj) != 14:
            logger.warning(f"[{idx}] CNPJ inválido: {raw_cnpj}")
            resumo.append({
                "idx": idx,
                "cnpj": str(raw_cnpj),
                "pesquisa_ok": 0,
                "download_ok": 0,
                "erro": "CNPJ inválido"
            })
            continue

        logger.info(f"[{idx}] CNPJ {cnpj} - pesquisando disponibilidade")
        campos = build_campos_for_cnpj(cfg, row, cnpj, runtime_vars=runtime_vars)

        # Monta XMLs para debug mesmo em dry-run
        try:
            ident = build_identificacao(cfg.job, cnpj)
            pesquisa_xml = build_pesquisa_xml(ident, campos)
            if dumper:
                dumper.dump("entrada_pesquisa_xml", pesquisa_xml, {"cnpj": cnpj})
        except Exception as e:
            logger.error(f"[{idx}] Falha ao montar XML de pesquisa: {e}")
            resumo.append({
                "idx": idx, "cnpj": cnpj,
                "pesquisa_ok": 0, "download_ok": 0,
                "erro": f"XML pesquisa: {e}"
            })
            continue

        if args.dry_run:
            resumo.append({
                "idx": idx, "cnpj": cnpj,
                "pesquisa_ok": 0, "download_ok": 0,
                "erro": "dry-run (não executado)"
            })
            continue

        pesquisa_ok = 0
        download_ok = 0
        pedido_id = 0
        pesquisa_msg = ""
        erro = ""

        try:
            arquivo_ids, pesquisa_msg = client.pesquisar_arquivos(cnpj, campos)
            pesquisa_ok = 1
            logger.info(f"[{idx}] Pesquisa OK. Arquivos encontrados: {len(arquivo_ids)}")

            if not arquivo_ids:
                # doc: pesquisa pode retornar sucesso com 0 arquivos. :contentReference[oaicite:8]{index=8}
                resumo.append({
                    "idx": idx,
                    "cnpj": cnpj,
                    "pesquisa_ok": 1,
                    "qtd_arquivos": 0,
                    "pedido_id": "",
                    "download_ok": 1,
                    "situacao_final": "sem_arquivos",
                    "mensagem": pesquisa_msg
                })
                continue

            # Solicita 1 pedido por CNPJ (lista de IDs)
            logger.info(f"[{idx}] Solicitando pedido (1 por CNPJ) com {len(arquivo_ids)} ids")
            pedido_id, ped_msg = client.solicitar_arquivos_por_ids(cnpj, arquivo_ids)

            if pedido_id <= 0:
                raise WsTransportError(f"Pedido retornou id inválido: {pedido_id}. msg={ped_msg}")

            logger.info(f"[{idx}] Pedido criado: {pedido_id}")
            # polling
            start = time.time()
            # Polling com intervalos diferentes por situação:
            # - processando: usa interval_processando_sec (ou interval_sec)
            # - disponivel/concluido: usa interval_disponivel_sec
            # OBS: se interval_disponivel_sec = 0, NÃO fazemos loop infinito.
            # Fazemos poucas checagens rápidas e, se ainda não baixou, seguimos para o próximo CNPJ.
            interval_base = max(1, int(getattr(cfg.polling, "interval_sec", 60)))
            interval_proc = max(1, int(getattr(cfg.polling, "interval_processando_sec", interval_base)))
            interval_disp = max(0, int(getattr(cfg.polling, "interval_disponivel_sec", 0)))  # permite 0 (sem espera)
            timeout = max(interval_base, int(getattr(cfg.polling, "timeout_sec", 1200)))
            max_attempts = max(1, int(getattr(cfg.polling, "max_attempts", 120)))
            # quando interval_disp=0, NÃO vamos ficar poluindo o log nem rodando infinitamente.
            # Por padrão faz apenas 1 checagem adicional em "disponivel/concluido" e segue.
            max_disp_attempts = max(1, int(getattr(cfg.polling, "max_attempts_disponivel", 1)))

            # Modo "sem espera": não faz sentido tentar centenas/milhares de vezes.
            # Forçamos limites duros para evitar loop apertado e log gigantesco.
            if interval_disp == 0:
                max_disp_attempts = 1
                max_attempts = min(max_attempts, 5)
                timeout = min(timeout, 5)

            final_situacao = ""
            arquivos_status: List[Dict[str, Any]] = []
            attempts = 0
            disp_attempts = 0
            last_situacao = None

            while True:
                attempts += 1
                if attempts > max_attempts:
                    final_situacao = "max_attempts"
                    break

                if time.time() - start > timeout:
                    final_situacao = "timeout"
                    break

                status_map = client.verificar_pedidos([pedido_id], atributos=True)
                pdata = status_map.get(pedido_id)
                if not pdata:
                    final_situacao = "inexistente"
                    break

                situacao = pdata.get("situacao", "")
                final_situacao = situacao or final_situacao
                arquivos_status = pdata.get("arquivos", []) or []

                # Critério de conclusão:
                # - sucesso total se TODOS arquivos estiverem 'baixado'
                # - finaliza com falha se pedido estiver erro/inativo/inexistente
                if situacao in ("erro", "inativo", "inexistente"):
                    break

                if arquivos_status:
                    situacoes_arquivos = [a.get("situacao", "") for a in arquivos_status]
                    if all(s == "baixado" for s in situacoes_arquivos):
                        break
                    if any(s in ("erro", "cancelado") for s in situacoes_arquivos):
                        # encerra mais cedo (parcial/erro)
                        break

                # Decide o próximo intervalo conforme situação/arquivos
                if situacao == "processando" or not arquivos_status:
                    next_sleep = interval_proc
                elif situacao in ("disponivel", "concluido"):
                    next_sleep = interval_disp
                    if interval_disp == 0:
                        disp_attempts += 1
                        if disp_attempts >= max_disp_attempts:
                            # não ficar em loop; registra a situação e segue
                            break
                else:
                    next_sleep = interval_base

                # Não dormir além do timeout
                elapsed = time.time() - start
                remaining = max(0, timeout - elapsed)
                if remaining <= 0:
                    break
                if next_sleep > remaining:
                    next_sleep = remaining

                # Log controlado (evita encher o terminal)
                if next_sleep and next_sleep > 0:
                    should_log = (attempts == 1) or (situacao != last_situacao) or (attempts % 10 == 0)
                else:
                    # sem espera: log apenas no início e na saída
                    should_log = (attempts == 1) or (attempts >= max_attempts) or (disp_attempts >= max_disp_attempts)
                last_situacao = situacao

                if next_sleep and next_sleep > 0:
                    if should_log:
                        logger.info(f"[{idx}] Pedido {pedido_id} situacao={situacao}. aguardando {int(next_sleep)}s...")
                    time.sleep(next_sleep)
                else:
                    # sem espera: não spamma; log só quando mudar/primeira vez
                    if should_log:
                        logger.info(
                            f"[{idx}] Pedido {pedido_id} situacao={situacao}. sem espera (intervalo=0, tentativa {disp_attempts}/{max_disp_attempts})."
                        )

            # Consolida arquivos
            ok_total = True
            baixados = 0
            for a in arquivos_status:
                a_id = a.get("id", "")
                a_sit = a.get("situacao", "")
                a_local = a.get("local", "")
                a_hash = a.get("hash", "")
                a_tipohash = a.get("tipohash", "")
                a_tamanho = a.get("tamanho", "")
                arquivos_out.append({
                    "idx": idx,
                    "cnpj": cnpj,
                    "pedido_id": pedido_id,
                    "arquivo_id": a_id,
                    "situacao": a_sit,
                    "local": a_local,
                    "hash": a_hash,
                    "tipohash": a_tipohash,
                    "tamanho": a_tamanho
                })

                if a_sit != "baixado":
                    ok_total = False
                else:
                    baixados += 1
                    # local só aparece se baixado (doc). :contentReference[oaicite:9]{index=9}
                    if cfg.io.copy_downloaded_files and a_local and os.path.exists(a_local):
                        # Copia diretamente para out_dir (sem subpastas)
                        ensure_dir(out_dir)
                        dest = os.path.join(out_dir, safe_filename(f"{cnpj}_{pedido_id}_" + os.path.basename(a_local)))
                        try:
                            shutil.copy2(a_local, dest)
                        except Exception as e:
                            logger.warning(f"[{idx}] Falha ao copiar {a_local} -> {dest}: {e}")

            download_ok = 1 if ok_total else 0

            resumo.append({
                "idx": idx,
                "cnpj": cnpj,
                "pesquisa_ok": pesquisa_ok,
                "qtd_arquivos": len(arquivo_ids),
                "pedido_id": pedido_id,
                "download_ok": download_ok,
                "arquivos_baixados": baixados,
                "situacao_final": final_situacao,
                "mensagem": pesquisa_msg
            })

        except Exception as e:
            erro = str(e)
            logger.error(f"[{idx}] ERRO: {erro}")
            logger.debug(traceback.format_exc())
            resumo.append({
                "idx": idx,
                "cnpj": cnpj,
                "pesquisa_ok": pesquisa_ok,
                "qtd_arquivos": "",
                "pedido_id": pedido_id or "",
                "download_ok": download_ok,
                "situacao_final": "erro",
                "mensagem": pesquisa_msg,
                "erro": erro
            })
            continue

    # grava XLSX final
    write_result_xlsx(result_xlsx, resumo, arquivos_out)
    logger.info(f"Relatório gerado: {result_xlsx}")
    logger.info(f"Dumps em: {os.path.join(out_dir, 'debug')}")

if __name__ == "__main__":
    main()
