from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from tkinter import StringVar, Tk, filedialog, messagebox, ttk

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


# ==============================
# Configuracoes (ajustar para o novo cliente/layout)
# ==============================
PASTA_BASE = Path(__file__).resolve().parent
PASTA_ARQUIVOS = PASTA_BASE / "Arquivos"
TOLERANCIA_DIFERENCA = Decimal("0.10")
XLSX_FORMATO_MOEDA_BR = "[$R$-416] #,##0.00"


# ==============================
# Regex base (substituir conforme o template novo)
# ==============================
RE_MONEY = re.compile(r"(\d{1,3}(?:\.\d{3})*,\d{2})")


@dataclass(frozen=True)
class Registro:
    nota: str
    cfop: str
    valor: Decimal


def valor_br_para_decimal(txt: str) -> Decimal:
    return Decimal(txt.replace(".", "").replace(",", "."))


def normalizar_nota(txt: str) -> str:
    return re.sub(r"\D", "", txt).lstrip("0") or "0"


def ler_linhas_pdf(pdf_path: Path) -> list[str]:
    texto = "\n".join((pg.extract_text() or "") for pg in PdfReader(str(pdf_path)).pages)
    return [ln.strip() for ln in texto.splitlines() if ln.strip()]


def identificar_tipo_pdf(linhas: list[str]) -> str | None:
    cabecalho = "\n".join(linhas[:80]).upper()
    # TODO: ajustar a deteccao do layout no novo template.
    if "REGISTRO DE ENTRADAS" in cabecalho:
        return "livro"
    if "RELATORIO" in cabecalho:
        return "relatorio"
    return None


def parse_livro(linhas: list[str]) -> list[Registro]:
    # TODO: implementar parse do Livro no novo layout.
    return []


def parse_relatorio(linhas: list[str]) -> list[Registro]:
    # TODO: implementar parse do Relatorio no novo layout.
    return []


def agrupar_por_nota(registros: list[Registro]) -> dict[str, Decimal]:
    out: dict[str, Decimal] = defaultdict(Decimal)
    for r in registros:
        out[normalizar_nota(r.nota)] += r.valor
    return dict(out)


def agrupar_por_nota_cfop(registros: list[Registro]) -> dict[tuple[str, str], Decimal]:
    out: dict[tuple[str, str], Decimal] = defaultdict(Decimal)
    for r in registros:
        out[(normalizar_nota(r.nota), r.cfop)] += r.valor
    return dict(out)


def autoajustar_colunas(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            valor = ws.cell(row=row_idx, column=col_idx).value
            txt = "" if valor is None else str(valor)
            if len(txt) > max_len:
                max_len = len(txt)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 60))


def escrever_xlsx(caminho: Path, divergencias: list[list[object]], cfops: list[list[object]]) -> Path:
    wb = Workbook()
    ws_div = wb.active
    ws_div.title = "DIVERGENCIAS"
    ws_div.append(["TIPO", "NOTA", "DIFERENCA"])
    for row in divergencias:
        ws_div.append(row)
    ws_div["E1"] = "DIVERGENCIA TOTAL"
    ws_div["E2"] = "=SUM(C:C)"
    ws_div["E2"].number_format = XLSX_FORMATO_MOEDA_BR
    ws_div.freeze_panes = "A2"

    ws_cfop = wb.create_sheet("CFOP_DIFERENTE")
    ws_cfop.append(["NOTA", "CFOP", "VALOR_LIVRO", "VALOR_RELATORIO"])
    for row in cfops:
        ws_cfop.append(row)
    ws_cfop.freeze_panes = "A2"

    for ws in (ws_div, ws_cfop):
        autoajustar_colunas(ws)
    wb.save(caminho)
    return caminho


def selecionar_dois_pdfs() -> tuple[Path, Path] | None:
    root = Tk()
    root.title("Template Conciliador")
    root.geometry("860x320")
    frame = ttk.Frame(root, padding=16)
    frame.pack(fill="both", expand=True)

    ttk.Label(frame, text="Template Conciliador", font=("Segoe UI", 13, "bold")).grid(row=0, column=0, sticky="w")
    ttk.Label(frame, text="Selecione 2 PDFs (1 Livro e 1 Relatorio).").grid(row=1, column=0, sticky="w", pady=(6, 10))

    var = StringVar(value="Nenhum arquivo selecionado.")
    result: dict[str, tuple[Path, Path] | None] = {"files": None}

    def escolher() -> None:
        arquivos = filedialog.askopenfilenames(
            title="Selecione 2 arquivos PDF",
            filetypes=[("Arquivos PDF", "*.pdf *.PDF"), ("Todos", "*.*")],
        )
        if not arquivos:
            return
        if len(arquivos) != 2:
            messagebox.showwarning("Atencao", "Selecione exatamente 2 PDFs.")
            return
        a, b = Path(arquivos[0]), Path(arquivos[1])
        if a.resolve() == b.resolve():
            messagebox.showwarning("Atencao", "Selecione 2 arquivos diferentes.")
            return
        result["files"] = (a, b)
        var.set(f"1) {a.name}\n2) {b.name}")

    def confirmar() -> None:
        if not result["files"]:
            messagebox.showwarning("Atencao", "Selecione os 2 PDFs.")
            return
        root.destroy()

    def cancelar() -> None:
        result["files"] = None
        root.destroy()

    ttk.Button(frame, text="Selecionar os 2 PDFs", command=escolher).grid(row=2, column=0, sticky="w")
    ttk.Label(frame, textvariable=var, justify="left").grid(row=3, column=0, sticky="w", pady=(10, 0))

    botoes = ttk.Frame(frame)
    botoes.grid(row=4, column=0, sticky="e", pady=(16, 0))
    ttk.Button(botoes, text="Cancelar", command=cancelar).pack(side="right")
    ttk.Button(botoes, text="Comparar", command=confirmar).pack(side="right", padx=(0, 8))

    root.protocol("WM_DELETE_WINDOW", cancelar)
    root.mainloop()
    return result["files"]


def main() -> None:
    selecao = selecionar_dois_pdfs()
    if not selecao:
        print("Operacao cancelada.")
        return

    pdf_a, pdf_b = selecao
    linhas_a = ler_linhas_pdf(pdf_a)
    linhas_b = ler_linhas_pdf(pdf_b)
    tipo_a = identificar_tipo_pdf(linhas_a)
    tipo_b = identificar_tipo_pdf(linhas_b)

    if tipo_a is None or tipo_b is None:
        raise ValueError("Nao foi possivel identificar os tipos dos PDFs.")
    if tipo_a == tipo_b:
        raise ValueError("Selecione 1 Livro e 1 Relatorio.")

    if tipo_a == "livro":
        linhas_livro, linhas_relatorio = linhas_a, linhas_b
    else:
        linhas_livro, linhas_relatorio = linhas_b, linhas_a

    reg_livro = parse_livro(linhas_livro)
    reg_relatorio = parse_relatorio(linhas_relatorio)
    if not reg_livro:
        raise RuntimeError("Livro sem registros parseados.")
    if not reg_relatorio:
        raise RuntimeError("Relatorio sem registros parseados.")

    mapa_livro = agrupar_por_nota(reg_livro)
    mapa_relatorio = agrupar_por_nota(reg_relatorio)
    mapa_livro_cfop = agrupar_por_nota_cfop(reg_livro)
    mapa_relatorio_cfop = agrupar_por_nota_cfop(reg_relatorio)

    divergencias: list[list[object]] = []
    for nota in sorted(set(mapa_livro) | set(mapa_relatorio)):
        vl = mapa_livro.get(nota)
        vr = mapa_relatorio.get(nota)
        if vl is not None and vr is None:
            divergencias.append(["SO_NO_LIVRO", nota, float(vl)])
            continue
        if vr is not None and vl is None:
            divergencias.append(["SO_NO_RELATORIO", nota, float(-vr)])
            continue
        if vl is not None and vr is not None:
            dif = vl - vr
            if abs(dif) > TOLERANCIA_DIFERENCA:
                divergencias.append(["VALOR_DIVERGENTE", nota, float(dif)])

    cfop_div: list[list[object]] = []
    for chave in sorted(set(mapa_livro_cfop) | set(mapa_relatorio_cfop)):
        nota, cfop = chave
        vl = mapa_livro_cfop.get(chave, Decimal("0"))
        vr = mapa_relatorio_cfop.get(chave, Decimal("0"))
        if abs(vl - vr) > TOLERANCIA_DIFERENCA:
            cfop_div.append([nota, cfop, float(vl), float(vr)])

    PASTA_ARQUIVOS.mkdir(parents=True, exist_ok=True)
    out = PASTA_ARQUIVOS / f"Conciliacao_TEMPLATE_{datetime.now().strftime('%H-%M-%S_%d-%m-%Y')}.xlsx"
    escrever_xlsx(out, divergencias, cfop_div)
    print(f"Arquivo gerado: {out}")


if __name__ == "__main__":
    main()
