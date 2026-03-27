from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from tkinter import StringVar, Tk, filedialog, messagebox, ttk
from unicodedata import normalize

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


# ==============================
# Configuracoes
# ==============================
PASTA_BASE = Path(__file__).resolve().parent
PASTA_ARQUIVOS = PASTA_BASE / "Arquivos"
TOLERANCIA_DIFERENCA = Decimal("0.10")
XLSX_FORMATO_MOEDA_BR = '[$R$-416] #,##0.00'


# ==============================
# Utilitarios
# ==============================
RE_MONEY_2_DEC = re.compile(r"^(\d{1,3}(?:\.\d{3})*,\d{2})")
RE_MONEY_ANY = re.compile(r"(\d{1,3}(?:\.\d{3})*,\d{2})")
RE_INICIO_LIVRO = re.compile(
    r"^(\d{2}/\d{2}/\d{4})\s+(NFe|NF|CTe)\s+\d+\s+(\d+)\s+"
    r"\d{2}/\d{2}/\d{4}\s+\d+\s+[A-Z]{2}\s+00\s*/\s*00\s+([\d\.,]+)\s+(\d\.\d{3})"
)
RE_CONTINUACAO_LIVRO = re.compile(r"^([\d\.,]+)\s+(\d\.\d{3})(?:\s|$)")
RE_INICIO_RELATORIO = re.compile(r"^(\d{2}/\d{2}/\d{4})(.*)$")
RE_CHAVE_RELATORIO = re.compile(r"^\s*(\d+)\s+(\d\.\d{3})\s+(.+)$")


@dataclass(frozen=True)
class Registro:
    nota: str
    cfop: str
    valor: Decimal


def obter_caminhos_saida(pasta_saida: Path) -> dict[str, Path]:
    agora = datetime.now()
    nome_xlsx = f"resultado_divergencias {agora.strftime('%H-%M-%S')} {agora.strftime('%d-%m-%Y')}.xlsx"
    return {
        "xlsx_divergencias": pasta_saida / nome_xlsx,
    }


def ler_linhas_pdf(caminho_pdf: Path) -> list[str]:
    leitor = PdfReader(str(caminho_pdf))
    texto = "\n".join((pagina.extract_text() or "") for pagina in leitor.pages)
    return [linha.strip() for linha in texto.splitlines() if linha.strip()]


def valor_br_para_decimal(valor_br: str) -> Decimal:
    return Decimal(valor_br.replace(".", "").replace(",", "."))


def normalizar_nota(nota: str) -> str:
    return re.sub(r"\D", "", nota)


def extrair_valor_duas_casas(token: str) -> Decimal | None:
    match = RE_MONEY_2_DEC.match(token)
    if not match:
        return None
    return valor_br_para_decimal(match.group(1))


def extrair_total_rodape_livro(linhas: list[str]) -> Decimal | None:
    for linha in linhas:
        if linha.startswith("Total Geral"):
            match = RE_MONEY_ANY.search(linha)
            if match:
                return valor_br_para_decimal(match.group(1))
    return None


def extrair_total_rodape_relatorio(linhas: list[str]) -> Decimal | None:
    for i, linha in enumerate(linhas):
        if linha.startswith("Registros.:"):
            for prox in linhas[i + 1 : i + 20]:
                match = RE_MONEY_ANY.search(prox)
                if match:
                    return valor_br_para_decimal(match.group(1))
    return None


def identificar_tipo_pdf(linhas: list[str]) -> str | None:
    cabecalho = normalize("NFKD", "\n".join(linhas[:60])).encode("ASCII", "ignore").decode().upper()
    texto_completo = normalize("NFKD", "\n".join(linhas)).encode("ASCII", "ignore").decode().upper()

    if "REGISTRO DE ENTRADAS - MODELO P1" in cabecalho or "TOTAL GERAL" in texto_completo:
        return "livro"

    if "RELATORIO DO ARQUIVO MAGN" in cabecalho and "REGISTROS.:" in texto_completo:
        return "tipo50"

    return None


def parse_livro_registro(linhas: list[str]) -> list[Registro]:
    registros: list[Registro] = []
    nota_atual: str | None = None

    for linha in linhas:
        inicio = RE_INICIO_LIVRO.match(linha)
        if inicio:
            _, _, nota, token_valor, cfop = inicio.groups()
            nota_atual = nota
            valor = extrair_valor_duas_casas(token_valor)
            if valor is not None:
                registros.append(Registro(nota=nota, cfop=cfop, valor=valor))
            continue

        continuacao = RE_CONTINUACAO_LIVRO.match(linha)
        if continuacao and nota_atual:
            token_valor, cfop = continuacao.groups()
            valor = extrair_valor_duas_casas(token_valor)
            if valor is not None:
                registros.append(Registro(nota=nota_atual, cfop=cfop, valor=valor))

    return registros


def parse_relatorio_tipo50(linhas: list[str]) -> list[Registro]:
    registros: list[Registro] = []

    for linha in linhas:
        inicio = RE_INICIO_RELATORIO.match(linha)
        if not inicio:
            continue

        _, resto = inicio.groups()
        chave = RE_CHAVE_RELATORIO.match(resto)
        if not chave:
            continue

        nota, cfop, sufixo = chave.groups()
        primeiro_token = sufixo.split()[0] if sufixo.split() else ""
        valor = extrair_valor_duas_casas(primeiro_token)
        if valor is None:
            continue
        registros.append(Registro(nota=nota, cfop=cfop, valor=valor))

    return registros


def agrupar_por_nota(registros: list[Registro]) -> dict[str, Decimal]:
    acumulado: dict[str, Decimal] = defaultdict(Decimal)
    for registro in registros:
        nota_normalizada = normalizar_nota(registro.nota)
        if nota_normalizada:
            acumulado[nota_normalizada] += registro.valor
    return dict(acumulado)


def agrupar_por_nota_cfop(registros: list[Registro]) -> dict[tuple[str, str], Decimal]:
    acumulado: dict[tuple[str, str], Decimal] = defaultdict(Decimal)
    for registro in registros:
        nota_normalizada = normalizar_nota(registro.nota)
        if nota_normalizada:
            acumulado[(nota_normalizada, registro.cfop)] += registro.valor
    return dict(acumulado)


def anular_pares_opostos_linhas(linhas: list[list[object]], idx_diferenca: int) -> list[list[object]]:
    # Remove pares +X/-X com mesmo valor absoluto (em centavos).
    saldos: dict[int, list[list[object]]] = {}
    resultado: list[list[object]] = []

    for linha in linhas:
        valor = linha[idx_diferenca]
        if not isinstance(valor, (int, float)):
            resultado.append(linha)
            continue
        cents = int(round(valor * 100))
        oposto = -cents
        if oposto in saldos and saldos[oposto]:
            saldos[oposto].pop()
            continue
        saldos.setdefault(cents, []).append(linha)

    for bucket in saldos.values():
        resultado.extend(bucket)
    return resultado


def anular_pares_opostos_cfop(linhas: list[list[object]]) -> list[list[object]]:
    # linhas: [nota, cfop, valor_livro, valor_relatorio]
    # Anula apenas dentro da mesma NOTA+CFOP para nao remover diferencas reais
    # entre CFOPs distintos da mesma nota (ex.: 106868).
    saldos: dict[tuple[str, str, int], list[list[object]]] = {}
    resultado: list[list[object]] = []

    for linha in linhas:
        nota = str(linha[0])
        cfop = str(linha[1])
        vl = linha[2]
        vr = linha[3]
        if not isinstance(vl, (int, float)) or not isinstance(vr, (int, float)):
            resultado.append(linha)
            continue

        diferenca = float(vl) - float(vr)
        cents = int(round(diferenca * 100))
        oposto = -cents
        chave_oposta = (nota, cfop, oposto)
        if chave_oposta in saldos and saldos[chave_oposta]:
            saldos[chave_oposta].pop()
            continue
        chave = (nota, cfop, cents)
        saldos.setdefault(chave, []).append(linha)

    for bucket in saldos.values():
        resultado.extend(bucket)
    return resultado


def dec_para_str(valor: Decimal) -> str:
    return f"{valor:.2f}"


def autoajustar_colunas(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            texto = "" if val is None else str(val)
            if len(texto) > max_len:
                max_len = len(texto)
        # margem extra para leitura, semelhante ao autofit do Excel
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 60))


def escrever_xlsx_divergencias(
    caminho: Path,
    linhas: list[list[str]],
    linhas_cfop: list[list[str]],
) -> Path:
    wb = Workbook()
    ws_div = wb.active
    ws_div.title = "DIVERGENCIAS"
    ws_div.append(["TIPO", "NOTA", "DIFERENCA"])
    for linha in linhas:
        ws_div.append(linha)
    ws_div["E1"] = "DIVERGENCIA TOTAL"
    ws_div["E2"] = "=SUM(C:C)"
    ws_div["E2"].number_format = XLSX_FORMATO_MOEDA_BR
    ws_div.freeze_panes = "A2"
    for row in ws_div.iter_rows(min_row=2, max_row=ws_div.max_row, min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = XLSX_FORMATO_MOEDA_BR

    ws_cfop = wb.create_sheet("CFOP_DIFERENTE")
    ws_cfop.append(["NOTA", "CFOP", "VALOR_LIVRO", "VALOR_RELATORIO"])
    for linha in linhas_cfop:
        ws_cfop.append(linha)
    ws_cfop.freeze_panes = "A2"
    for row in ws_cfop.iter_rows(min_row=2, max_row=ws_cfop.max_row, min_col=3, max_col=4):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = XLSX_FORMATO_MOEDA_BR

    # Forca todos os textos em maiusculo.
    for ws in (ws_div, ws_cfop):
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = cell.value.upper()
        autoajustar_colunas(ws)

    try:
        wb.save(caminho)
        return caminho
    except PermissionError:
        caminho_alternativo = caminho.with_name(f"{caminho.stem}_novo{caminho.suffix}")
        wb.save(caminho_alternativo)
        return caminho_alternativo


def selecionar_dois_pdfs() -> tuple[Path, Path] | None:
    root = Tk()
    root.title("Conciliador - Livro IOB x Balancete Tipo 50")
    root.geometry("860x360")
    root.minsize(780, 320)

    estilo = ttk.Style(root)
    try:
        estilo.theme_use("vista")
    except Exception:
        pass

    frame = ttk.Frame(root, padding=18)
    frame.pack(fill="both", expand=True)

    ttk.Label(
        frame,
        text="Comparador de PDF fiscal",
        font=("Segoe UI", 13, "bold"),
    ).grid(row=0, column=0, columnspan=3, sticky="w")
    ttk.Label(
        frame,
        text=(
            "Selecione 2 PDFs para comparar um Livro de Registros IOB com um Balancete Tipo 50.\n"
            "A identificacao e automatica por conteudo, sem depender do nome do arquivo."
        ),
        justify="left",
    ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(6, 14))

    caminho_a = StringVar()
    caminho_b = StringVar()
    resultado: dict[str, tuple[Path, Path] | None] = {"arquivos": None}

    def escolher(var: StringVar) -> None:
        arquivo = filedialog.askopenfilename(
            title="Selecione um arquivo PDF",
            filetypes=[("Arquivos PDF", "*.pdf *.PDF"), ("Todos os arquivos", "*.*")],
        )
        if arquivo:
            var.set(arquivo)

    def confirmar() -> None:
        if not caminho_a.get() or not caminho_b.get():
            messagebox.showwarning("Atencao", "Selecione os 2 arquivos PDF.")
            return

        arq_a = Path(caminho_a.get())
        arq_b = Path(caminho_b.get())

        if arq_a.resolve() == arq_b.resolve():
            messagebox.showwarning("Atencao", "Selecione 2 arquivos diferentes.")
            return

        resultado["arquivos"] = (arq_a, arq_b)
        root.destroy()

    def cancelar() -> None:
        resultado["arquivos"] = None
        root.destroy()

    ttk.Label(frame, text="Arquivo PDF 1").grid(row=2, column=0, sticky="w", pady=(0, 4))
    ttk.Entry(frame, textvariable=caminho_a, width=88).grid(row=3, column=0, sticky="ew")
    ttk.Button(frame, text="Selecionar", command=lambda: escolher(caminho_a)).grid(row=3, column=1, padx=(10, 0))

    ttk.Label(frame, text="Arquivo PDF 2").grid(row=4, column=0, sticky="w", pady=(16, 4))
    ttk.Entry(frame, textvariable=caminho_b, width=88).grid(row=5, column=0, sticky="ew")
    ttk.Button(frame, text="Selecionar", command=lambda: escolher(caminho_b)).grid(row=5, column=1, padx=(10, 0))

    botoes = ttk.Frame(frame)
    botoes.grid(row=6, column=0, columnspan=3, sticky="e", pady=(24, 0))
    ttk.Button(botoes, text="Cancelar", command=cancelar).pack(side="right")
    ttk.Button(botoes, text="Comparar", command=confirmar).pack(side="right", padx=(0, 10))

    frame.columnconfigure(0, weight=1)
    root.protocol("WM_DELETE_WINDOW", cancelar)
    root.mainloop()
    return resultado["arquivos"]


# ==============================
# Main
# ==============================
def main() -> None:
    selecao = selecionar_dois_pdfs()
    if not selecao:
        print("Operacao cancelada.")
        return

    pdf_1, pdf_2 = selecao
    if not pdf_1.exists() or not pdf_2.exists():
        raise FileNotFoundError("Um ou mais arquivos selecionados nao existem.")

    linhas_pdf_1 = ler_linhas_pdf(pdf_1)
    linhas_pdf_2 = ler_linhas_pdf(pdf_2)

    tipo_1 = identificar_tipo_pdf(linhas_pdf_1)
    tipo_2 = identificar_tipo_pdf(linhas_pdf_2)

    if tipo_1 is None or tipo_2 is None:
        raise ValueError(
            "Nao foi possivel identificar o tipo de um dos PDFs. "
            "Use um Livro de Registros IOB e um Balancete Tipo 50."
        )
    if tipo_1 == tipo_2:
        raise ValueError("Os dois arquivos parecem ser do mesmo tipo. Selecione 1 Livro e 1 Tipo 50.")

    if tipo_1 == "livro":
        pdf_livro = pdf_1
        pdf_relatorio = pdf_2
        linhas_livro = linhas_pdf_1
        linhas_relatorio = linhas_pdf_2
    else:
        pdf_livro = pdf_2
        pdf_relatorio = pdf_1
        linhas_livro = linhas_pdf_2
        linhas_relatorio = linhas_pdf_1

    PASTA_ARQUIVOS.mkdir(parents=True, exist_ok=True)
    caminhos_saida = obter_caminhos_saida(PASTA_ARQUIVOS)

    registros_livro = parse_livro_registro(linhas_livro)
    registros_relatorio = parse_relatorio_tipo50(linhas_relatorio)
    total_rodape_livro = extrair_total_rodape_livro(linhas_livro)
    total_rodape_relatorio = extrair_total_rodape_relatorio(linhas_relatorio)

    mapa_livro = agrupar_por_nota(registros_livro)
    mapa_relatorio = agrupar_por_nota(registros_relatorio)
    mapa_livro_cfop = agrupar_por_nota_cfop(registros_livro)
    mapa_relatorio_cfop = agrupar_por_nota_cfop(registros_relatorio)

    notas = set(mapa_livro) | set(mapa_relatorio)
    divergencias_xlsx: list[list[str]] = []
    divergencias_cfop_xlsx: list[list[str]] = []
    soma_impacto_divergencias = Decimal("0")

    for nota in sorted(notas):
        valor_livro = mapa_livro.get(nota)
        valor_relatorio = mapa_relatorio.get(nota)

        if valor_livro is not None and valor_relatorio is None:
            divergencias_xlsx.append(
                ["SO_NO_LIVRO", nota, float(valor_livro)]
            )
            soma_impacto_divergencias += valor_livro
            continue

        if valor_relatorio is not None and valor_livro is None:
            divergencias_xlsx.append(
                ["SO_NO_RELATORIO", nota, float(-valor_relatorio)]
            )
            soma_impacto_divergencias -= valor_relatorio
            continue

        if valor_livro is not None and valor_relatorio is not None:
            diferenca = valor_livro - valor_relatorio
            if abs(diferenca) <= TOLERANCIA_DIFERENCA:
                continue
            soma_impacto_divergencias += diferenca
            divergencias_xlsx.append(
                [
                    "VALOR_DIVERGENTE",
                    nota,
                    float(diferenca),
                ]
            )

    chaves_cfop = set(mapa_livro_cfop) | set(mapa_relatorio_cfop)
    for nota, cfop in sorted(chaves_cfop):
        valor_livro = mapa_livro_cfop.get((nota, cfop), Decimal("0"))
        valor_relatorio = mapa_relatorio_cfop.get((nota, cfop), Decimal("0"))
        # Remove linhas onde VALOR_RELATORIO (coluna D) arredondado em 2 casas e 0,00.
        if round(float(valor_relatorio), 2) == 0.0:
            continue
        diferenca = valor_livro - valor_relatorio
        if abs(diferenca) <= TOLERANCIA_DIFERENCA:
            continue
        divergencias_cfop_xlsx.append(
            [
                nota,
                cfop,
                float(valor_livro),
                float(valor_relatorio),
            ]
        )

    divergencias_xlsx = anular_pares_opostos_linhas(divergencias_xlsx, 2)
    divergencias_cfop_xlsx = anular_pares_opostos_cfop(divergencias_cfop_xlsx)

    total_parse_livro = sum(mapa_livro.values(), Decimal("0"))
    total_parse_relatorio = sum(mapa_relatorio.values(), Decimal("0"))
    diferenca_parse = total_parse_livro - total_parse_relatorio
    diferenca_rodape = None
    if total_rodape_livro is not None and total_rodape_relatorio is not None:
        diferenca_rodape = total_rodape_livro - total_rodape_relatorio

    xlsx_salvo = escrever_xlsx_divergencias(
        caminhos_saida["xlsx_divergencias"],
        divergencias_xlsx,
        divergencias_cfop_xlsx,
    )

    print(f"Livro identificado: {pdf_livro}")
    print(f"Tipo 50 identificado: {pdf_relatorio}")
    print(f"Registros parseados (Livro): {len(registros_livro)}")
    print(f"Registros parseados (Relatorio): {len(registros_relatorio)}")
    print(f"Total parseado Livro: {dec_para_str(total_parse_livro)}")
    print(f"Total parseado Relatorio: {dec_para_str(total_parse_relatorio)}")
    print(f"Diferenca parseada (Livro-Relatorio): {dec_para_str(diferenca_parse)}")
    if total_rodape_livro is not None:
        print(f"Total rodape Livro: {dec_para_str(total_rodape_livro)}")
    if total_rodape_relatorio is not None:
        print(f"Total rodape Relatorio: {dec_para_str(total_rodape_relatorio)}")
    if diferenca_rodape is not None:
        print(f"Diferenca rodape (Livro-Relatorio): {dec_para_str(diferenca_rodape)}")
    print(f"Tolerancia diferenca por nota: {dec_para_str(TOLERANCIA_DIFERENCA)}")
    print(f"Soma impacto divergencias: {dec_para_str(soma_impacto_divergencias)}")
    print(f"Total de linhas de discrepancia: {len(divergencias_xlsx)}")
    print(f"Total de linhas de CFOP divergente: {len(divergencias_cfop_xlsx)}")
    print(f"XLSX gerado: {xlsx_salvo}")


if __name__ == "__main__":
    main()
