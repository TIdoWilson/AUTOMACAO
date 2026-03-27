# -*- coding: utf-8 -*-
import os
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except ImportError:
    tk = None
    filedialog = None
    messagebox = None

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("Erro: modulo 'openpyxl' nao encontrado. Instale com: pip install openpyxl") from exc

try:
    import rarfile
except ImportError:
    rarfile = None


# ===================== CONFIG =====================
NOME_ARQUIVO_SAIDA = "relatorio_venda_presencial.xlsx"
COLUNAS_PLANILHA = ["nr da nota", "CFOPs", "presencial", "chave de acesso"]
IND_PRES_PRESENCIAL = {"1", "5"}
MAPA_IND_PRES = {
    "0": "NAO (0 - nao se aplica)",
    "1": "SIM (1 - operacao presencial)",
    "2": "NAO (2 - internet)",
    "3": "NAO (3 - teleatendimento)",
    "4": "NAO (4 - entrega em domicilio)",
    "5": "SIM (5 - presencial fora do estabelecimento)",
    "9": "NAO (9 - outros)",
}


# ===================== UTIL =====================
def nome_local_tag(tag):
    return str(tag).rsplit("}", 1)[-1]


def encontrar_primeiro(elemento, nome_tag):
    for item in elemento.iter():
        if nome_local_tag(item.tag) == nome_tag:
            return item
    return None


def encontrar_filhos(elemento, nome_tag):
    return [item for item in list(elemento) if nome_local_tag(item.tag) == nome_tag]


def texto_filho(elemento, nome_tag):
    if elemento is None:
        return ""

    for filho in list(elemento):
        if nome_local_tag(filho.tag) == nome_tag:
            return (filho.text or "").strip()
    return ""


def adicionar_se_nao_existir(lista, valor):
    if valor and valor not in lista:
        lista.append(valor)


def formatar_ind_pres(ind_pres):
    if not ind_pres:
        return "NAO ENCONTRADO"
    return MAPA_IND_PRES.get(ind_pres, f"DESCONHECIDO ({ind_pres})")


def definir_pasta_saida(arquivos_selecionados):
    pastas = [str(Path(caminho).resolve().parent) for caminho in arquivos_selecionados]
    pasta_comum = Path(os.path.commonpath(pastas))
    return pasta_comum


def definir_caminho_saida(pasta_saida):
    caminho_saida = pasta_saida / NOME_ARQUIVO_SAIDA
    if not caminho_saida.exists():
        return caminho_saida

    indice = 2
    while True:
        nome_tentativa = f"{caminho_saida.stem}_{indice}{caminho_saida.suffix}"
        tentativa = pasta_saida / nome_tentativa
        if not tentativa.exists():
            return tentativa
        indice += 1


def selecionar_arquivos():
    if tk is None or filedialog is None:
        entrada = input("Informe os caminhos dos arquivos .xml/.zip/.rar separados por ;\n").strip()
        if not entrada:
            return []
        return [item.strip('" ').strip() for item in entrada.split(";") if item.strip()]

    raiz = tk.Tk()
    raiz.withdraw()
    raiz.attributes("-topmost", True)

    arquivos = filedialog.askopenfilenames(
        title="Selecione os XMLs ou arquivos ZIP/RAR",
        filetypes=[
            ("Arquivos suportados", "*.xml *.zip *.rar"),
            ("XML", "*.xml"),
            ("ZIP", "*.zip"),
            ("RAR", "*.rar"),
            ("Todos os arquivos", "*.*"),
        ],
    )

    raiz.destroy()
    return list(arquivos)


def abrir_xmls_do_zip(caminho_zip):
    with zipfile.ZipFile(caminho_zip, "r") as arquivo_zip:
        for info in arquivo_zip.infolist():
            if info.is_dir() or not info.filename.lower().endswith(".xml"):
                continue
            yield f"{Path(caminho_zip).name}::{info.filename}", arquivo_zip.read(info)


def abrir_xmls_do_rar(caminho_rar):
    if rarfile is None:
        raise RuntimeError(
            "Arquivo .rar selecionado, mas o modulo 'rarfile' nao esta instalado. "
            "Instale com: pip install rarfile. Se necessario, instale tambem o utilitario 'unrar' ou 'bsdtar'."
        )

    try:
        with rarfile.RarFile(caminho_rar, "r") as arquivo_rar:
            for info in arquivo_rar.infolist():
                if info.isdir() or not info.filename.lower().endswith(".xml"):
                    continue
                with arquivo_rar.open(info) as xml_rar:
                    yield f"{Path(caminho_rar).name}::{info.filename}", xml_rar.read()
    except rarfile.Error as exc:
        raise RuntimeError(f"Falha ao ler o RAR '{caminho_rar}': {exc}") from exc


# ===================== XML =====================
def analisar_xml(xml_bytes):
    raiz = ET.fromstring(xml_bytes)
    inf_nfe = encontrar_primeiro(raiz, "infNFe")

    if inf_nfe is None:
        raise ValueError("Tag infNFe nao encontrada no XML.")

    ide = encontrar_primeiro(inf_nfe, "ide")
    numero_nota = texto_filho(ide, "nNF")
    ind_pres = texto_filho(ide, "indPres")

    cfops = []
    for det in encontrar_filhos(inf_nfe, "det"):
        prod = encontrar_primeiro(det, "prod")
        adicionar_se_nao_existir(cfops, texto_filho(prod, "CFOP"))

    chave_acesso = ""
    tag_chave = encontrar_primeiro(raiz, "chNFe")
    if tag_chave is not None:
        chave_acesso = (tag_chave.text or "").strip()

    if not chave_acesso:
        chave_acesso = inf_nfe.attrib.get("Id", "").replace("NFe", "").strip()

    return {
        "nr da nota": numero_nota,
        "CFOPs": ", ".join(cfops),
        "presencial": formatar_ind_pres(ind_pres),
        "chave de acesso": chave_acesso,
        "_eh_presencial": ind_pres in IND_PRES_PRESENCIAL,
    }


def carregar_xmls(caminhos_selecionados):
    registros = []
    erros = []

    for caminho in caminhos_selecionados:
        path = Path(caminho)
        sufixo = path.suffix.lower()

        try:
            if sufixo == ".xml":
                registros.append((path.name, path.read_bytes()))
            elif sufixo == ".zip":
                registros.extend(abrir_xmls_do_zip(path))
            elif sufixo == ".rar":
                registros.extend(abrir_xmls_do_rar(path))
            else:
                erros.append(f"Arquivo ignorado por extensao nao suportada: {path}")
        except Exception as exc:
            erros.append(f"Falha ao abrir '{path}': {exc}")

    return registros, erros


# ===================== PLANILHA =====================
def exportar_planilha(linhas, caminho_saida):
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas Presenciais"
    ws.append(COLUNAS_PLANILHA)

    for linha in linhas:
        ws.append([linha[coluna] for coluna in COLUNAS_PLANILHA])

    for indice_coluna, coluna in enumerate(COLUNAS_PLANILHA, start=1):
        maior = len(coluna)
        for celula in ws.iter_cols(min_col=indice_coluna, max_col=indice_coluna, min_row=2):
            for item in celula:
                maior = max(maior, len(str(item.value or "")))
        ws.column_dimensions[get_column_letter(indice_coluna)].width = min(maior + 2, 80)

    ws.freeze_panes = "A2"
    wb.save(caminho_saida)


# ===================== MAIN =====================
def main():
    arquivos_selecionados = selecionar_arquivos()
    if not arquivos_selecionados:
        print("Nenhum arquivo foi selecionado.")
        return

    registros_xml, erros = carregar_xmls(arquivos_selecionados)
    if not registros_xml:
        print("Nenhum XML foi encontrado nos arquivos selecionados.")
        if erros:
            print("\nErros encontrados:")
            for erro in erros:
                print(f"- {erro}")
        return

    linhas = []
    for origem_xml, xml_bytes in registros_xml:
        try:
            linhas.append(analisar_xml(xml_bytes))
        except Exception as exc:
            erros.append(f"Falha ao processar '{origem_xml}': {exc}")

    if not linhas:
        print("Nenhum XML valido foi processado.")
        if erros:
            print("\nErros encontrados:")
            for erro in erros:
                print(f"- {erro}")
        return

    linhas.sort(key=lambda item: (not item["_eh_presencial"], item["nr da nota"], item["chave de acesso"]))
    for linha in linhas:
        linha.pop("_eh_presencial", None)

    pasta_saida = definir_pasta_saida(arquivos_selecionados)
    caminho_saida = definir_caminho_saida(pasta_saida)
    exportar_planilha(linhas, caminho_saida)

    quantidade_presenciais = sum(1 for linha in linhas if str(linha["presencial"]).startswith("SIM"))

    print(f"Planilha gerada em: {caminho_saida}")
    print(f"XMLs processados com sucesso: {len(linhas)}")
    print(f"Notas com indicativo presencial: {quantidade_presenciais}")

    if erros:
        print("\nErros encontrados:")
        for erro in erros:
            print(f"- {erro}")

    if messagebox is not None:
        resumo = [
            f"Planilha gerada em:\n{caminho_saida}",
            f"\nXMLs processados: {len(linhas)}",
            f"\nVendas presenciais: {quantidade_presenciais}",
        ]
        if erros:
            resumo.append(f"\n\nErros: {len(erros)} (veja o console)")
        messagebox.showinfo("Consulta de venda presencial", "".join(resumo))


if __name__ == "__main__":
    main()
