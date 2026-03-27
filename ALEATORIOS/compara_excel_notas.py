import sys
import re
import csv
from decimal import Decimal
from pathlib import Path
import openpyxl
import pandas as pd

def to_decimal(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return Decimal(str(v)).quantize(Decimal("0.01"))
    s = str(v).strip().replace("\u00a0", " ").replace(" ", "")
    if not re.search(r"\d", s):
        return None
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return Decimal(s).quantize(Decimal("0.01"))
    except:
        return None

def norm_doc_from_cell(v):
    s = str(v).strip().upper().replace("-", "")
    if s.startswith("NFC"):
        return "NFCe"
    if s.startswith("NFE"):
        return "NFe"
    if s.startswith("NFCE"):
        return "NFCe"
    return s

def parse_livro_registro(path_xlsx: Path):
    wb = openpyxl.load_workbook(path_xlsx, data_only=True)

    ws = None
    for name in wb.sheetnames:
        tmp = wb[name]
        for r in range(1, min(tmp.max_row, 200) + 1):
            v = tmp.cell(r, 1).value
            if isinstance(v, str) and v.strip().upper().startswith("NF"):
                ws = tmp
                break
        if ws:
            break
    if ws is None:
        raise RuntimeError("Não encontrei dados (NF*) no arquivo do Livro.")

    # achar início dos dados
    start = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().upper().startswith("NF"):
            start = r
            break
    if start is None:
        raise RuntimeError("Não consegui identificar a primeira linha de NF.")

    current = None  # (doc, serie, numero)
    totals = {}     # total por nota (col 7/8)
    per_cfop = {}   # soma por nota+cfop

    def add_cfop(key, cfop, amt):
        kcf = (key[0], key[1], key[2], cfop)
        per_cfop[kcf] = per_cfop.get(kcf, Decimal("0.00")) + amt

    for r in range(start, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        if isinstance(c1, str) and ("TOTAL GERAL" in c1.upper() or "RESUMO" in c1.upper()):
            break

        # linha de cabeçalho da nota
        if isinstance(c1, str) and c1.strip().upper().startswith("NF"):
            doc = norm_doc_from_cell(c1)
            serie = ws.cell(r, 2).value
            num_fim = ws.cell(r, 4).value
            if isinstance(serie, (int, float)) and isinstance(num_fim, (int, float)):
                current = (doc, int(serie), int(num_fim))
                total = to_decimal(ws.cell(r, 7).value) or to_decimal(ws.cell(r, 8).value)
                if total is not None and total != Decimal("0.00"):
                    totals[current] = total

        # linha CFOP (pode ser continuação)
        cfop = ws.cell(r, 9).value
        if current is None or cfop is None:
            continue

        cfop_int = str(int(cfop)) if isinstance(cfop, (int, float)) else str(cfop).replace(".", "")

        # Valor por CFOP no Livro normalmente aparece como:
        # base (col 11) + isentas (col 14) + outras (col 15)
        base = to_decimal(ws.cell(r, 11).value) or Decimal("0.00")
        isentas = to_decimal(ws.cell(r, 14).value) or Decimal("0.00")
        outras = to_decimal(ws.cell(r, 15).value) or Decimal("0.00")

        amt = base + isentas + outras
        if amt != Decimal("0.00"):
            add_cfop(current, cfop_int, amt)

    return totals, per_cfop

def parse_modelo_p2a(path_xlsx: Path):
    wb = openpyxl.load_workbook(path_xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # achar início dos dados
    start = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().upper().startswith("NF") and isinstance(ws.cell(r, 3).value, (int, float)):
            start = r
            break
    if start is None:
        raise RuntimeError("Não consegui identificar a primeira linha de NF no Modelo P2/A.")

    current = None
    totals = {}
    per_cfop = {}

    def add_cfop(key, cfop, amt):
        kcf = (key[0], key[1], key[2], cfop)
        per_cfop[kcf] = per_cfop.get(kcf, Decimal("0.00")) + amt
        totals[key] = totals.get(key, Decimal("0.00")) + amt

    for r in range(start, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        if isinstance(c1, str) and ("TOTAL GERAL" in c1.upper() or "RESUMO" in c1.upper()):
            break

        if isinstance(c1, str) and c1.strip().upper().startswith("NF"):
            doc = norm_doc_from_cell(c1)
            serie = ws.cell(r, 2).value
            num_fim = ws.cell(r, 4).value
            if isinstance(serie, (int, float)) and isinstance(num_fim, (int, float)):
                current = (doc, int(serie), int(num_fim))

        cfop = ws.cell(r, 9).value
        val = to_decimal(ws.cell(r, 7).value)

        if current is None or cfop is None or val is None or val == Decimal("0.00"):
            continue

        cfop_int = str(int(cfop)) if isinstance(cfop, (int, float)) else str(cfop).replace(".", "")
        add_cfop(current, cfop_int, val)

    return totals, per_cfop

def compare_maps(map_a, map_b, key_cols):
    keys = sorted(set(map_a.keys()) | set(map_b.keys()))
    rows = []
    for k in keys:
        a = map_a.get(k)
        b = map_b.get(k)
        if a is None:
            rows.append((*k, None, b, None, "FALTA_NO_RELATORIO_1"))
        elif b is None:
            rows.append((*k, a, None, None, "FALTA_NO_RELATORIO_2"))
        else:
            diff = a - b
            if diff != Decimal("0.00"):
                rows.append((*k, a, b, diff, "DIVERGENTE"))

    cols = key_cols + ["valor_relatorio_1", "valor_relatorio_2", "diferenca", "status"]
    return pd.DataFrame(rows, columns=cols)

def main(xlsx1, xlsx2):
    xlsx1 = Path(xlsx1)
    xlsx2 = Path(xlsx2)

    tot1, cfop1 = parse_livro_registro(xlsx1)
    tot2, cfop2 = parse_modelo_p2a(xlsx2)

    df_nota = compare_maps(tot1, tot2, ["doc", "serie", "numero"])
    df_cfop = compare_maps(cfop1, cfop2, ["doc", "serie", "numero", "cfop"])

    df_nota.to_csv("dif_por_nota.csv", index=False, encoding="utf-8-sig")
    df_cfop.to_csv("dif_por_cfop.csv", index=False, encoding="utf-8-sig")

    with pd.ExcelWriter("resultado_comparacao.xlsx", engine="openpyxl") as writer:
        df_nota.to_excel(writer, sheet_name="Dif_por_nota", index=False)
        df_cfop.to_excel(writer, sheet_name="Dif_por_cfop", index=False)

    print(f"Notas comparadas: {len(tot1)} x {len(tot2)}")
    print(f"Diferenças por NOTA: {len(df_nota)}  -> dif_por_nota.csv")
    print(f"Diferenças por NOTA+CFOP: {len(df_cfop)} -> dif_por_cfop.csv")
    print("Arquivo Excel: resultado_comparacao.xlsx")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python compara_excel_notas.py <livro.xlsx> <p2a.xlsx>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
