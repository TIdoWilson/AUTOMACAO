import os
import re
import sys
import asyncio
from datetime import datetime, date, timedelta
from typing import Optional, Tuple
from pathlib import Path
from urllib.parse import quote

from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError

# MantÃ©m as dependÃªncias do relatÃ³rio original
from decimal import Decimal, InvalidOperation
from xml.etree import ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ===================== CONFIGURAÃ‡ÃƒO =====================

# Pasta base para salvar: W:\XML PREFEITURA\<EMPRESA>\<PASTA_PERIODO>\
BASE_DOWNLOAD_DIR = r"W:\XML PREFEITURA"

# PerÃ­odo (DD/MM/AAAA) â€” sempre do 1Âº ao Ãºltimo dia do mÃªs anterior ao corrente
def periodo_mes_anterior(hoje: Optional[date] = None) -> Tuple[str, str]:
    hoje = hoje or date.today()
    primeiro_dia_mes_atual = hoje.replace(day=1)
    ultimo_dia_mes_anterior = primeiro_dia_mes_atual - timedelta(days=1)
    primeiro_dia_mes_anterior = ultimo_dia_mes_anterior.replace(day=1)
    return (
        primeiro_dia_mes_anterior.strftime("%d/%m/%Y"),
        ultimo_dia_mes_anterior.strftime("%d/%m/%Y"),
    )

DATA_INICIO: Optional[str] = None
DATA_FIM: Optional[str] = None

# ConcorrÃªncia de downloads (XML/PDF) por pÃ¡gina
MAX_CONCORRENCIA_DOWNLOADS = 15

# Se True, rebaixa arquivo mesmo se jÃ¡ existir
SOBRESCREVER_ARQUIVOS = False

# Playwright
HEADLESS = False
TIMEOUT_MS = 60_000
TIMEOUT_LOGIN_MS = 10 * 60_000  # 10 min


# ===================== UTILITÃRIOS =====================

def sanitize_folder(name: str) -> str:
    """Remove caracteres invÃ¡lidos em nomes de pasta no Windows e normaliza espaÃ§os."""
    name = re.sub(r'[<>:"/\\|?*]+', ' ', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name if name else "Empresa_Desconhecida"


def _try_parse_date(s: str) -> datetime:
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    raise ValueError(f"Data invÃ¡lida: {s!r} (use DD/MM/AAAA)")


def pasta_periodo(data_inicio: str, data_fim: str) -> str:
    """Nomeia a pasta do perÃ­odo. Se for o mesmo mÃªs/ano -> 'MM-AAAA', senÃ£o -> 'PERIODO_YYYYMMDD_YYYYMMDD'."""
    di = _try_parse_date(data_inicio)
    df = _try_parse_date(data_fim)

    if di.year == df.year and di.month == df.month:
        return f"{di.month:02d}-{di.year}"
    return f"PERIODO_{di.strftime('%Y%m%d')}_{df.strftime('%Y%m%d')}"


def obter_periodo_usuario() -> Tuple[str, str]:
    """Solicita data inicial e final ao usuario e valida o periodo informado."""
    while True:
        data_inicio = input("Informe a data inicial (DD/MM/AAAA): ").strip()
        data_fim = input("Informe a data final (DD/MM/AAAA): ").strip()

        try:
            di = _try_parse_date(data_inicio)
            df = _try_parse_date(data_fim)
        except ValueError as e:
            print(f"[ERRO] {e}")
            continue

        if di > df:
            print("[ERRO] A data inicial nao pode ser maior que a data final.")
            continue

        return data_inicio, data_fim


def build_list_url(base: str, pg: int, executar, data_inicio: str, data_fim: str, busca: str = "") -> str:
    """
    Monta a URL do portal com:
      pg, executar (pode repetir), busca, datainicio, datafim

    - 'executar' pode ser str ou lista[str]. Se lista, repete o parÃ¢metro:
        executar=["1","1"]  -> ...&executar=1&executar=1
    - MantÃ©m vÃ­rgula em valores (ex.: '1,1')
    - Escapa '/' nas datas (vira %2F)
    """
    if isinstance(executar, (list, tuple)):
        executar_list = list(executar)
    else:
        executar_list = [str(executar)]

    busca_enc = quote(busca or "", safe="")
    di_enc = quote(data_inicio, safe="")
    df_enc = quote(data_fim, safe="")

    executar_part = "".join([f"&executar={quote(v, safe=',')}" for v in executar_list])

    return (
        f"{base}?"
        f"pg={pg}"
        f"{executar_part}"
        f"&busca={busca_enc}"
        f"&datainicio={di_enc}"
        f"&datafim={df_enc}"
    )

    return (
        f"{base}?"
        f"pg={pg}"
        f"&executar={executar_enc}"
        f"&busca={busca_enc}"
        f"&datainicio={di_enc}"
        f"&datafim={df_enc}"
    )


async def wait_table_or_empty(page):
    """Espera a tabela carregar (linhas ou estado vazio)."""
    await page.wait_for_selector("table.table", timeout=TIMEOUT_MS)
    await page.wait_for_selector("table.table tbody", timeout=TIMEOUT_MS)

    # Aguarda o corpo popular (linhas) ou algum sinal de conteÃºdo.
    # NÃ£o usamos `networkidle` porque o site pode manter requisiÃ§Ãµes em background.
    try:
        await page.wait_for_function(
            """() => {
                const tbody = document.querySelector('table.table tbody');
                if (!tbody) return false;
                const hasRow = tbody.querySelectorAll('tr').length > 0;
                const hasDownload = tbody.querySelector("a[href*='/Notas/Download/']") !== null;
                return hasRow || hasDownload;
            }""",
            timeout=3_000,
        )
    except PlaywrightTimeoutError:
        # Se nÃ£o detectar, segue com o que jÃ¡ carregou
        pass


async def get_last_page_number(page) -> int:
    """
    LÃª o mÃ¡ximo nÃºmero de pÃ¡gina no componente de paginaÃ§Ã£o.
    Se nÃ£o encontrar, assume 1.
    """
    try:
        texts = await page.eval_on_selector_all(
            "ul.pagination a",
            "els => els.map(e => (e.textContent || '').trim())"
        )
        nums = []
        for t in texts:
            if t.isdigit():
                nums.append(int(t))
        return max(nums) if nums else 1
    except Exception:
        return 1


async def obter_empresa_no_dashboard(page) -> str:
    """
    LÃª o parÃ¡grafo que contÃ©m 'Nome:' no bloco 'Meus dados' do Dashboard
    e retorna o texto apÃ³s 'Nome:' (o nome da empresa).
    """
    candidatos = [
        "section:has-text('Meus dados') p:has-text('Nome')",
        "div:has-text('Meus dados') p:has-text('Nome')",
        "p:has-text('Nome')",
        "xpath=(//p[contains(normalize-space(.),'Nome')])[1]",
    ]

    for sel in candidatos:
        try:
            p = page.locator(sel).first
            if await p.count() > 0:
                raw = (await p.inner_text()).strip()
                m = re.search(r"Nome\s*:\s*(.+)$", raw, flags=re.IGNORECASE)
                if m:
                    return sanitize_folder(m.group(1).strip())
        except Exception:
            pass

    # Fallbacks
    for label in ["RazÃ£o Social", "Razao Social", "Empresa"]:
        try:
            p = page.locator(f"p:has-text('{label}')").first
            if await p.count() > 0:
                raw = (await p.inner_text()).strip()
                parts = raw.split(":", 1)
                if len(parts) == 2:
                    return sanitize_folder(parts[1].strip())
        except Exception:
            pass

    return "Empresa_Desconhecida"


# ===================== DOWNLOAD CONCORRENTE =====================

async def baixar_binario(request, url: str, out_path: str, sem: asyncio.Semaphore, max_retries: int = 3) -> bool:
    """Baixa um arquivo via request.get (mesma sessÃ£o/cookies do navegador) com retries e limite de concorrÃªncia."""
    if (not SOBRESCREVER_ARQUIVOS) and os.path.exists(out_path):
        return True

    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    async with sem:
        for tent in range(1, max_retries + 1):
            try:
                resp = await request.get(url, timeout=TIMEOUT_MS)
                if not resp.ok:
                    raise RuntimeError(f"HTTP {resp.status}")
                data = await resp.body()
                with open(out_path, "wb") as f:
                    f.write(data)
                return True
            except Exception as e:
                if tent == max_retries:
                    print(f"âŒ Falhou ({tent}/{max_retries}): {url} -> {e}")
                    return False
                await asyncio.sleep(0.6 * tent)
    return False


async def processa_pagina_download(page, request, pasta_destino: str, prefixo: str, sem: asyncio.Semaphore) -> int:
    """
    Baixa todos os XMLs/PDFs da pÃ¡gina atual (sem filtrar por data â€” o filtro jÃ¡ vem na URL).
    Retorna o nÃºmero de downloads disparados.
    """
    await wait_table_or_empty(page)

    rows = page.locator("table.table tbody tr")
    n = await rows.count()
    if n == 0:
        return 0

    tarefas = []
    vistos = set()  # evita duplicar por algum bug de DOM

    for i in range(n):
        row = rows.nth(i)

        # SituaÃ§Ã£o (cancelada / substituida) pelo tooltip do Ã­cone na coluna de situaÃ§Ã£o
        status_suf = ""
        try:
            img = row.locator("td.td-situacao img").first
            tooltip = None
            if await img.count() > 0:
                tooltip = (await img.get_attribute("data-original-title")) or (await img.get_attribute("title"))
            if tooltip:
                t = tooltip.lower()
                if "cancelada" in t:
                    status_suf = " - cancelada"
                elif "substitui" in t:
                    status_suf = " - substituida"
        except Exception:
            pass

        # XML
        link_xml = row.locator("a[href*='/Notas/Download/NFSe/']").first
        if await link_xml.count() > 0:
            href = await link_xml.get_attribute("href")
            if href:
                url = href if href.startswith("http") else ("https://www.nfse.gov.br" + href)
                if url not in vistos:
                    vistos.add(url)
                    nome_padrao = href.split("/")[-1] + ".xml"
                    base, ext = os.path.splitext(nome_padrao)
                    novo_nome = f"{prefixo} {base}{status_suf}{ext}"
                    out_path = os.path.join(pasta_destino, novo_nome)
                    tarefas.append(baixar_binario(request, url, out_path, sem))

        # PDF
        link_pdf = row.locator("a[href*='/Notas/Download/DANFSe/']").first
        if await link_pdf.count() > 0:
            href = await link_pdf.get_attribute("href")
            if href:
                url = href if href.startswith("http") else ("https://www.nfse.gov.br" + href)
                if url not in vistos:
                    vistos.add(url)
                    nome_padrao = href.split("/")[-1] + ".pdf"
                    base, ext = os.path.splitext(nome_padrao)
                    novo_nome = f"{prefixo} {base}{status_suf}{ext}"
                    out_path = os.path.join(pasta_destino, novo_nome)
                    tarefas.append(baixar_binario(request, url, out_path, sem))

    if tarefas:
        await asyncio.gather(*tarefas)

    return len(tarefas)


async def baixar_todas_paginas(page, request, base_url: str, executar: str, pasta_destino: str, prefixo: str):
    """
    Percorre pg=1..N alterando a URL (sem clicar em 'fa-angle-right'),
    baixa tudo de cada pÃ¡gina e para quando a pÃ¡gina nÃ£o tiver nenhuma nota (sem links de download).
    """
    sem = asyncio.Semaphore(MAX_CONCORRENCIA_DOWNLOADS)
    pg = 1
    while True:
        url = build_list_url(base_url, pg, executar, DATA_INICIO, DATA_FIM, busca="")
        print(f"\n--- {prefixo} | PÃGINA {pg} ---")
        print(url)

        await page.goto(url, wait_until="domcontentloaded", timeout=TIMEOUT_MS)

        try:
            await wait_table_or_empty(page)
        except PlaywrightTimeoutError:
            print("âŒ Timeout ao carregar a tabela.")
            break

        rows = page.locator("table.table tbody tr")
        if await rows.count() == 0:
            print("âœ” Lista vazia. Encerrando paginaÃ§Ã£o.")
            break

        qtd = await processa_pagina_download(page, request, pasta_destino, prefixo=prefixo, sem=sem)
        print(f"âœ” Downloads disparados nesta pÃ¡gina: {qtd}")

        if qtd == 0:
            print("âœ” Nenhuma nota detectada nesta pÃ¡gina (sem links de download). Encerrando paginaÃ§Ã£o.")
            break

        pg += 1


# ========== FUNCOES PARA GERAR RELATORIO (IMPORTADAS DO gerar_relatorio_nfse.py) ==========

SAIDA_HEADERS = [
    "numero_nfse",
    "data_emissao",
    "situacao",
    "razao_social_tomador",
    "CNPJ/CPF",
    "valor_total_servicos",
    "irrf_retido",
    "pis_retido",
    "cofins_retido",
    "csll_retido",
    "inss_retido",
    "issqn_retido",
    "valor_liquido_servico",
    "arquivo",
]

ENTRADA_HEADERS = [
    "numero_nfse",
    "data_emissao",
    "situacao",
    "razao_social_emitente",
    "cnpj_emitente",
    "valor_total_servicos",
    "irrf_retido",
    "pis_retido",
    "cofins_retido",
    "csll_retido",
    "inss_retido",
    "issqn_retido",
    "valor_liquido_servico",
    "arquivo",
]

CURRENCY_COLUMNS = {
    "valor_total_servicos",
    "irrf_retido",
    "pis_retido",
    "cofins_retido",
    "csll_retido",
    "inss_retido",
    "issqn_retido",
    "valor_liquido_servico",
}


def local_name(tag: str) -> str:
    return tag.split("}", 1)[-1]


def direct_child(element: ET.Element | None, tag_name: str) -> ET.Element | None:
    if element is None:
        return None
    for child in list(element):
        if local_name(child.tag) == tag_name:
            return child
    return None


def first_descendant(element: ET.Element | None, tag_name: str) -> ET.Element | None:
    if element is None:
        return None
    for child in element.iter():
        if local_name(child.tag) == tag_name:
            return child
    return None


def text_of(element: ET.Element | None) -> str:
    if element is None or element.text is None:
        return ""
    return element.text.strip()


def descendant_text(element: ET.Element | None, tag_name: str) -> str:
    return text_of(first_descendant(element, tag_name))


def decimal_or_zero(value: str) -> float:
    if not value:
        return 0.0
    normalized = value.replace(",", ".")
    try:
        return float(Decimal(normalized))
    except (InvalidOperation, ValueError):
        return 0.0


def parse_date(value: str):
    if not value:
        return ""
    raw = value.strip()
    try:
        if "T" in raw:
            return datetime.fromisoformat(raw).date()
        return datetime.fromisoformat(raw[:10]).date()
    except ValueError:
        return raw


def get_tax_amounts(inf_nfse: ET.Element | None, inf_dps: ET.Element | None) -> dict[str, float]:
    nf_values = direct_child(inf_nfse, "valores")
    dps_values = direct_child(inf_dps, "valores")
    trib = direct_child(dps_values, "trib")
    trib_mun = direct_child(trib, "tribMun")
    trib_fed = direct_child(trib, "tribFed")
    pis_cofins = direct_child(trib_fed, "piscofins")

    tp_ret_issqn = descendant_text(trib_mun, "tpRetISSQN")
    issqn_value = decimal_or_zero(descendant_text(nf_values, "vISSQN"))
    issqn_retido = issqn_value if tp_ret_issqn == "2" else 0.0

    return {
        "valor_total_servicos": decimal_or_zero(descendant_text(dps_values, "vServ")),
        "irrf_retido": decimal_or_zero(descendant_text(trib_fed, "vRetIRRF")),
        "pis_retido": decimal_or_zero(descendant_text(pis_cofins, "vPis")),
        "cofins_retido": decimal_or_zero(descendant_text(pis_cofins, "vCofins")),
        "csll_retido": decimal_or_zero(descendant_text(trib_fed, "vRetCSLL")),
        "inss_retido": decimal_or_zero(descendant_text(trib_fed, "vRetCP")),
        "issqn_retido": issqn_retido,
        "valor_liquido_servico": decimal_or_zero(descendant_text(nf_values, "vLiq")),
    }


def get_situacao(path: Path, inf_dps: ET.Element | None, inf_nfse: ET.Element | None) -> str:
    name = path.stem.lower()
    if "cancelada" in name:
        return "cancelada"
    if "substituida" in name:
        return "substituida"

    cstat = descendant_text(inf_nfse, "cStat")
    if first_descendant(inf_dps, "subst") is not None or cstat == "101":
        return "substituicao"
    return "normal"


def get_party_document(element: ET.Element | None) -> str:
    cnpj = descendant_text(element, "CNPJ")
    if cnpj:
        return cnpj
    return descendant_text(element, "CPF")


def parse_xml(path: Path) -> tuple[str, dict[str, object]]:
    root = ET.parse(path).getroot()
    inf_nfse = first_descendant(root, "infNFSe")
    inf_dps = first_descendant(root, "infDPS")
    emit = direct_child(inf_nfse, "emit")
    toma = direct_child(inf_dps, "toma")
    taxes = get_tax_amounts(inf_nfse, inf_dps)

    common = {
        "numero_nfse": descendant_text(inf_nfse, "nNFSe"),
        "data_emissao": parse_date(descendant_text(inf_dps, "dhEmi") or descendant_text(inf_nfse, "dhProc")),
        "situacao": get_situacao(path, inf_dps, inf_nfse),
        "arquivo": path.name,
        **taxes,
    }

    file_type = "saida" if path.name.upper().startswith("SAIDA") else "entrada"
    if file_type == "saida":
        common["razao_social_tomador"] = descendant_text(toma, "xNome")
        common["CNPJ/CPF"] = get_party_document(toma)
    else:
        common["razao_social_emitente"] = descendant_text(emit, "xNome")
        common["cnpj_emitente"] = get_party_document(emit)
    return file_type, common


def adjust_columns(worksheet, headers: list[str]) -> None:
    for index, header in enumerate(headers, start=1):
        max_length = len(header)
        for cell in worksheet[get_column_letter(index)]:
            value = cell.value
            if value is None:
                continue
            length = len(str(value))
            if length > max_length:
                max_length = length
        worksheet.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 40)


def populate_sheet(worksheet, headers: list[str], rows: list[dict[str, object]]) -> None:
    worksheet.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    date_column = headers.index("data_emissao") + 1
    for row_cells in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        row_cells[date_column - 1].number_format = "dd/mm/yyyy"
        for index, header in enumerate(headers, start=1):
            if header in CURRENCY_COLUMNS:
                row_cells[index - 1].number_format = "#,##0.00"

    adjust_columns(worksheet, headers)


def gerar_relatorio(target_dir: Path):
    if not target_dir.exists() or not target_dir.is_dir():
        print(f"[ERRO] Pasta nao encontrada para gerar relatorio: {target_dir}")
        return

    xml_files = sorted(target_dir.glob("*.xml"))
    output_file = target_dir / f"relatorio_nfse_{target_dir.name}.xlsx"

    if not xml_files:
        print("[AVISO] Nenhum XML encontrado na pasta para gerar relatorio.")
        return

    saida_rows: list[dict[str, object]] = []
    entrada_rows: list[dict[str, object]] = []

    for path in xml_files:
        try:
            file_type, row = parse_xml(path)
        except Exception as e:
            print(f"[AVISO] Falha ao processar XML '{path.name}': {e}")
            continue

        if file_type == "saida":
            saida_rows.append(row)
        else:
            entrada_rows.append(row)

    workbook = Workbook()
    ws_saida = workbook.active
    ws_saida.title = "Notas Saida"
    populate_sheet(ws_saida, SAIDA_HEADERS, saida_rows)

    ws_entrada = workbook.create_sheet("Notas Entrada")
    populate_sheet(ws_entrada, ENTRADA_HEADERS, entrada_rows)

    workbook.save(output_file)
    print(f"[OK] Relatorio gerado: {output_file}")
    print(f"Notas de saida: {len(saida_rows)}")
    print(f"Notas de entrada: {len(entrada_rows)}")


# ===================== MAIN =====================

async def main():
    global DATA_INICIO, DATA_FIM

    DATA_INICIO, DATA_FIM = obter_periodo_usuario()
    print(f"Periodo selecionado: {DATA_INICIO} a {DATA_FIM}")

    periodo_dirname = pasta_periodo(DATA_INICIO, DATA_FIM)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS)
        context = await browser.new_context()
        page = await context.new_page()

        # Login (manual)
        await page.goto("https://www.nfse.gov.br/EmissorNacional/Login", wait_until="domcontentloaded", timeout=TIMEOUT_MS)
        print("Aguardando login...")

        # Aguarda o Dashboard
        await page.wait_for_url("**/Dashboard", timeout=TIMEOUT_LOGIN_MS)

        # Identifica a empresa
        empresa = await obter_empresa_no_dashboard(page)
        print(f"Empresa identificada: {empresa}")

        # Cria pasta base W:\XML PREFEITURA\<EMPRESA>\<PERIODO>
        pasta_empresa = os.path.join(BASE_DOWNLOAD_DIR, empresa)
        pasta_destino = os.path.join(pasta_empresa, periodo_dirname)
        os.makedirs(pasta_destino, exist_ok=True)

        # ===================== EMITIDAS =====================
        await baixar_todas_paginas(
            page=page,
            request=context.request,
            base_url="https://www.nfse.gov.br/EmissorNacional/Notas/Emitidas",
            executar=["1","1"],
            pasta_destino=pasta_destino,
            prefixo="SAIDA",
        )

        # ===================== RECEBIDAS =====================
        await baixar_todas_paginas(
            page=page,
            request=context.request,
            base_url="https://www.nfse.gov.br/EmissorNacional/Notas/Recebidas",
            executar="1,1",  # conforme exemplo informado
            pasta_destino=pasta_destino,
            prefixo="ENTRADA",
        )

        await browser.close()

    # RelatÃ³rio (usa os XMLs do perÃ­odo recÃ©m baixado)
    print(f"\nIniciando processamento do relatÃ³rio na pasta: {pasta_destino}")
    gerar_relatorio(Path(pasta_destino))


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        sys.exit(0)

