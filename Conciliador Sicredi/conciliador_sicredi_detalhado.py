from __future__ import annotations

import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from itertools import combinations
from pathlib import Path
from unicodedata import normalize

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


# ==============================
# Configuracoes
# ==============================
PASTA_BASE = Path(__file__).resolve().parent
PASTA_ARQUIVOS = PASTA_BASE / "Arquivos"
TOLERANCIA = Decimal("0.10")
XLSX_FORMATO_MOEDA_BR = "[$R$-416] #,##0.00"
MODO_EXPERIMENTAL_COMPENSAR_RAZAO = True


# ==============================
# Regex
# ==============================
RE_RAZAO_LANC = re.compile(
    r"^(\d{1,2}/\d{2}/\d{4})\s+(-?\d{1,3}(?:\.\d{3})*,\d{2})\s+(-?\d{1,3}(?:\.\d{3})*,\d{2})\s+-\s+(.+)$",
    re.IGNORECASE,
)
RE_EXT_HEAD = re.compile(r"^(\d{1,2}/\d{2}/\d{4})\s+(\d{1,3})\s+-\s+(.+)$", re.IGNORECASE)
RE_EXT_STAR = re.compile(r"^\*\s*(\d+)\s+(-?\d{1,3}(?:\.\d{3})*,\d{2})\s+([+-])\s+(-?\d{1,3}(?:\.\d{3})*,\d{2})$")
RE_EXT_STAR_INLINE = re.compile(r"\*\s*(\d+)\s+(-?\d{1,3}(?:\.\d{3})*,\d{2})\s+([+-])\s+(-?\d{1,3}(?:\.\d{3})*,\d{2})")
RE_EXT_DETALHE = re.compile(r"^(.+?)\s+(\d{1,10})\s+([0-9A-Z]{1,4})\s*-\s*(-?\d{1,3}(?:\.\d{3})*,\d{2})$")
RE_DOC_PARC = re.compile(r"\b(\d{1,10})[-/](\d{1,4}|ND|PR)\b", re.IGNORECASE)
RE_DOC_TIT = re.compile(r"\bTITULO\s+(\d{1,10})\b", re.IGNORECASE)
RE_DOC_CHEQUE = re.compile(r"\bCHEQUE\s+(\d{1,10})\b", re.IGNORECASE)
RE_DOC_ADIANT = re.compile(r"\bADIANTAMENTO:\s*(\d{1,10})\b", re.IGNORECASE)
RE_DOC_CAIXA = re.compile(r"\bLANCAMENTO\s+CAIXA:\s*(\d{1,10})\b", re.IGNORECASE)
RE_DOC_LANC_BANC = re.compile(r"\bLANCAMENTO\s+BANCARIO:\s*(\d{1,10})\b", re.IGNORECASE)
RE_DOC_HEAD_NUM = re.compile(r"^\s*(\d{6,10})\b")


@dataclass(frozen=True)
class RegistroRazao:
    data_lcto: str
    tipo: str
    valor: Decimal
    saldo: Decimal
    doc: str
    parc: str
    historico: str


@dataclass(frozen=True)
class RegistroExtrato:
    data_lcto: str
    tipo: str
    valor: Decimal
    doc: str
    parc: str
    origem: str  # principal | detalhe9803
    linha: str


def classificar_tipo_razao(historico: str, saldo: Decimal, saldo_anterior: Decimal | None) -> str:
    h = normalizar_ascii(historico)

    # 1) Palavras-chave mais confiaveis
    if "PAGAMENTO TITULO" in h or "BAIXA ADIANTAMENTO" in h or "REF SALARIOS" in h:
        return "SAIDA"
    if "RECEBIMENTO TITULO" in h or "DEPOSITO CHEQUE" in h:
        return "ENTRADA"

    # 2) Marcadores contabeis no final da linha (ex.: D D, D C, C D)
    m = re.search(r"\b([DC])\s*([DC])\s*$", h)
    if m:
        par = m.group(1) + m.group(2)
        if par in {"DD", "CD"}:
            return "ENTRADA"
        if par in {"DC", "CC"}:
            return "SAIDA"

    # 3) Fallback por variacao de saldo
    if saldo_anterior is None:
        return "ENTRADA"
    if saldo > saldo_anterior:
        return "ENTRADA"
    if saldo < saldo_anterior:
        return "SAIDA"
    return "ENTRADA"


def normalizar_ascii(s: str) -> str:
    return normalize("NFKD", s).encode("ASCII", "ignore").decode().upper()


def valor_br_para_decimal(s: str) -> Decimal:
    return Decimal(s.replace(".", "").replace(",", "."))


def dec_para_str(v: Decimal) -> str:
    return f"{v:.2f}"


def ler_linhas_pdf(caminho_pdf: Path) -> list[str]:
    txt = "\n".join((pg.extract_text() or "") for pg in PdfReader(str(caminho_pdf)).pages)
    return [ln.strip() for ln in txt.splitlines() if ln.strip()]


def descobrir_pasta_pdfs() -> Path:
    candidatas = [
        PASTA_BASE / "Entrada",
        PASTA_BASE / "Arquivos de Exemplo",
        PASTA_BASE,
    ]
    for pasta in candidatas:
        if pasta.exists() and any(pasta.glob("*.pdf")):
            return pasta
    raise RuntimeError(
        "Nao foi encontrada pasta com PDFs. Coloque os arquivos em "
        f"'{PASTA_BASE}\\Entrada' ou '{PASTA_BASE}\\Arquivos de Exemplo'."
    )


def detectar_tipo_pdf(linhas: list[str]) -> str | None:
    cab = normalizar_ascii("\n".join(linhas[:160]))
    if "RELATORIO: RAZAO CONTABIL" in cab and "BCO SICREDI" in cab:
        return "razao"
    if "EXTRATO BANCARIO" in cab and "SICREDI" in cab:
        return "extrato"
    return None


def extrair_doc_parc(texto: str) -> tuple[str, str]:
    m = RE_DOC_PARC.search(texto)
    if m:
        return m.group(1), m.group(2).upper()
    m = RE_DOC_TIT.search(texto)
    if m:
        return m.group(1), ""
    m = RE_DOC_CHEQUE.search(texto)
    if m:
        return m.group(1), ""
    m = RE_DOC_ADIANT.search(texto)
    if m:
        return m.group(1), ""
    m = RE_DOC_CAIXA.search(texto)
    if m:
        return m.group(1), ""
    m = RE_DOC_LANC_BANC.search(texto)
    if m:
        return m.group(1), ""
    m = RE_DOC_HEAD_NUM.search(texto)
    if m:
        return m.group(1), ""
    return "", ""


def parse_razao(linhas: list[str]) -> list[RegistroRazao]:
    out: list[RegistroRazao] = []
    saldo_anterior: Decimal | None = None
    for ln in linhas:
        m = RE_RAZAO_LANC.match(ln)
        if not m:
            continue
        data_lcto, valor_txt, saldo_txt, historico = m.groups()
        valor = valor_br_para_decimal(valor_txt)
        saldo = valor_br_para_decimal(saldo_txt)
        if valor == 0:
            continue

        tipo = classificar_tipo_razao(historico, saldo, saldo_anterior)

        doc, parc = extrair_doc_parc(historico)
        out.append(
            RegistroRazao(
                data_lcto=data_lcto,
                tipo=tipo,
                valor=valor,
                saldo=saldo,
                doc=doc,
                parc=parc,
                historico=historico[:220],
            )
        )
        saldo_anterior = saldo
    return out


def parse_extrato(linhas: list[str]) -> list[RegistroExtrato]:
    out: list[RegistroExtrato] = []
    vistos: set[tuple[str, str, Decimal, str, str, str]] = set()
    data_atual = ""
    tipo_atual = ""
    linha_head = ""
    doc_head = ""
    parc_head = ""
    bloco_9803 = False

    for ln in linhas:
        m_head = RE_EXT_HEAD.match(ln)
        if m_head:
            data_atual, _, pos = m_head.groups()
            pos_up = normalizar_ascii(pos)

            # Resolve variacoes de cabecalho do extrato (incluindo estornos)
            if pos_up.startswith("ESTORNO DE ENTRADA "):
                tipo_atual = "SAIDA"
                resto = pos[len("ESTORNO DE ENTRADA ") :]
            elif pos_up.startswith("ESTORNO DE SAIDA "):
                tipo_atual = "ENTRADA"
                resto = pos[len("ESTORNO DE SAIDA ") :]
            elif pos_up.startswith("ENTRADA "):
                tipo_atual = "ENTRADA"
                resto = pos[len("ENTRADA ") :]
            elif pos_up.startswith("SAIDA "):
                tipo_atual = "SAIDA"
                resto = pos[len("SAIDA ") :]
            else:
                # Cabecalho nao reconhecido como movimento util
                tipo_atual = ""
                continue

            linha_head = resto
            doc_head, parc_head = extrair_doc_parc(resto)
            resto_up = normalizar_ascii(resto)
            bloco_9803 = "9803" in resto_up or "REFERENTE TITULOS" in resto_up

            # Alguns extratos trazem '* seq valor +/- saldo' na mesma linha do cabecalho.
            ms_in = RE_EXT_STAR_INLINE.search(ln)
            if ms_in and not bloco_9803:
                _, valor_txt, _, _ = ms_in.groups()
                valor = valor_br_para_decimal(valor_txt)
                chave = (data_atual, tipo_atual.upper(), valor, doc_head, parc_head, "principal")
                if chave not in vistos:
                    out.append(
                        RegistroExtrato(
                            data_lcto=data_atual,
                            tipo=tipo_atual.upper(),
                            valor=valor,
                            doc=doc_head,
                            parc=parc_head,
                            origem="principal",
                            linha=linha_head[:220],
                        )
                    )
                    vistos.add(chave)
            continue

        m_star = RE_EXT_STAR.match(ln)
        if m_star and data_atual and tipo_atual:
            if bloco_9803:
                continue
            _, valor_txt, _, _ = m_star.groups()
            valor = valor_br_para_decimal(valor_txt)
            chave = (data_atual, tipo_atual.upper(), valor, doc_head, parc_head, "principal")
            if chave not in vistos:
                out.append(
                    RegistroExtrato(
                        data_lcto=data_atual,
                        tipo=tipo_atual.upper(),
                        valor=valor,
                        doc=doc_head,
                        parc=parc_head,
                        origem="principal",
                        linha=linha_head[:220],
                    )
                )
                vistos.add(chave)
            continue

        if data_atual and tipo_atual:
            m_det = RE_EXT_DETALHE.match(ln)
            if not m_det:
                if bloco_9803 and (RE_EXT_HEAD.match(ln) or ln.startswith("*")):
                    bloco_9803 = False
                continue
            _, doc, parc, valor_txt = m_det.groups()
            valor = valor_br_para_decimal(valor_txt)
            origem = "detalhe9803" if bloco_9803 else "detalhe_livre"
            parc_fmt = parc.upper()
            chave = (data_atual, tipo_atual.upper(), valor, doc, parc_fmt, origem)
            if chave not in vistos:
                out.append(
                    RegistroExtrato(
                        data_lcto=data_atual,
                        tipo=tipo_atual.upper(),
                        valor=valor,
                        doc=doc,
                        parc=parc_fmt,
                        origem=origem,
                        linha=ln[:220],
                    )
                )
                vistos.add(chave)

    return out


def chave_doc_parc(r_doc: str, r_parc: str) -> tuple[str, str]:
    return r_doc.strip(), r_parc.strip().upper()


def conciliar(reg_razao: list[RegistroRazao], reg_extrato: list[RegistroExtrato]) -> tuple[list[tuple[RegistroRazao, RegistroExtrato]], list[RegistroRazao], list[RegistroExtrato]]:
    matches: list[tuple[RegistroRazao, RegistroExtrato]] = []
    usados_extrato: set[int] = set()

    # indice forte: doc+parc+valor (ignora tipo para evitar falso negativo contabil)
    idx_forte: dict[tuple[str, str, Decimal], list[int]] = defaultdict(list)
    for i, e in enumerate(reg_extrato):
        idx_forte[(e.doc, e.parc, e.valor)].append(i)

    # 1) casamento por documento/parcela exatos
    usados_razao: set[int] = set()
    for i, r in enumerate(reg_razao):
        if not r.doc:
            continue
        key = (r.doc, r.parc, r.valor)
        cand = idx_forte.get(key, [])
        j = next((x for x in cand if x not in usados_extrato), None)
        if j is None:
            continue
        usados_razao.add(i)
        usados_extrato.add(j)
        matches.append((r, reg_extrato[j]))

    # 2) fallback por data+valor (ignora tipo)
    idx_fallback: dict[tuple[str, Decimal], list[int]] = defaultdict(list)
    for j, e in enumerate(reg_extrato):
        if j in usados_extrato:
            continue
        idx_fallback[(e.data_lcto, e.valor)].append(j)

    for i, r in enumerate(reg_razao):
        if i in usados_razao:
            continue
        key = (r.data_lcto if len(r.data_lcto) == 10 else f"0{r.data_lcto}", r.valor)
        cand = idx_fallback.get(key, [])
        livres = [x for x in cand if x not in usados_extrato]
        j = livres[0] if len(livres) == 1 else None
        if j is None:
            # tenta data sem zero inicial de dia
            d = r.data_lcto
            if d.startswith("0"):
                d = d[1:]
            key2 = (d, r.valor)
            cand = idx_fallback.get(key2, [])
            livres = [x for x in cand if x not in usados_extrato]
            j = livres[0] if len(livres) == 1 else None
        if j is None:
            continue
        usados_razao.add(i)
        usados_extrato.add(j)
        matches.append((r, reg_extrato[j]))

    # 2.5) Fallback por identificador de "lancamento bancario" no historico do Razao
    idx_lanc_banc: dict[tuple[str, Decimal], list[int]] = defaultdict(list)
    for j, e in enumerate(reg_extrato):
        if j in usados_extrato:
            continue
        # doc do extrato pode vir com 1 digito sufixo (ex.: 1192621)
        if e.doc:
            idx_lanc_banc[(e.doc, e.valor)].append(j)
            if len(e.doc) >= 6:
                idx_lanc_banc[(e.doc[:-1], e.valor)].append(j)

    for i, r in enumerate(reg_razao):
        if i in usados_razao:
            continue
        m = RE_DOC_LANC_BANC.search(r.historico)
        if not m:
            continue
        doc_lanc = m.group(1)
        cand = idx_lanc_banc.get((doc_lanc, r.valor), [])
        j = next((x for x in cand if x not in usados_extrato), None)
        if j is None:
            continue
        usados_razao.add(i)
        usados_extrato.add(j)
        matches.append((r, reg_extrato[j]))

    # 3) Match por soma de parcelas no Razao para um unico lancamento no Extrato
    nao_r_idx = [i for i in range(len(reg_razao)) if i not in usados_razao]
    nao_e_idx = [j for j in range(len(reg_extrato)) if j not in usados_extrato]

    grupo_razao_doc: dict[str, list[int]] = defaultdict(list)
    for i in nao_r_idx:
        r = reg_razao[i]
        if r.doc:
            grupo_razao_doc[r.doc].append(i)

    for j in nao_e_idx:
        if j in usados_extrato:
            continue
        e = reg_extrato[j]
        if not e.doc:
            continue
        cand_idx = [i for i in grupo_razao_doc.get(e.doc, []) if i not in usados_razao]
        if len(cand_idx) < 2:
            continue

        best_combo: tuple[int, ...] | None = None
        for tam in range(2, min(5, len(cand_idx)) + 1):
            for combo in combinations(cand_idx, tam):
                soma = sum((reg_razao[i].valor for i in combo), Decimal("0"))
                if abs(soma - e.valor) <= TOLERANCIA:
                    best_combo = combo
                    break
            if best_combo:
                break

        if best_combo:
            usados_extrato.add(j)
            for i in best_combo:
                usados_razao.add(i)
                # registra o pareamento de cada parcela com o mesmo lancamento do extrato
                matches.append((reg_razao[i], e))

    nao_conciliado_razao = [r for i, r in enumerate(reg_razao) if i not in usados_razao]
    nao_conciliado_extrato = [e for j, e in enumerate(reg_extrato) if j not in usados_extrato]
    return matches, nao_conciliado_razao, nao_conciliado_extrato


def autoajustar_colunas(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            txt = "" if val is None else str(val)
            if len(txt) > max_len:
                max_len = len(txt)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 70))


def sugerir_motivo_inconsistencia_razao(r: RegistroRazao) -> str:
    h = normalizar_ascii(r.historico)
    if "ESTORNO" in h:
        return "ESTORNO_SEM_CORRESPONDENTE_NO_EXTRATO"
    if "CANCELAMENTO DE ADIANTAMENTO" in h:
        return "CANCELAMENTO_ADIANTAMENTO_SEM_LANCAMENTO_EQUIVALENTE"
    if "CONFORME LANCAMENTO BANCARIO" in h:
        return "LANCAMENTO_CONTABIL_REFERENCIAL_NAO_ENCONTRADO"
    if "LANCAMENTO CAIXA" in h:
        return "MOVIMENTO_CAIXA_SEM_LANCAMENTO_BANCARIO_EQUIVALENTE"
    if "ADIANTAMENTO CFE TITULO" in h:
        return "ADIANTAMENTO_SEM_CORRESPONDENTE_NO_EXTRATO"
    if "RECEBIMENTO TITULO" in h or "PAGAMENTO TITULO" in h:
        return "TITULO_NAO_LOCALIZADO_NO_EXTRATO"
    return "VERIFICAR_HISTORICO_MANUALMENTE"


def categoria_historico_razao(h: str) -> str:
    t = normalizar_ascii(h)
    if "ESTORNO" in t:
        return "ESTORNO"
    if "CANCELAMENTO DE ADIANTAMENTO" in t:
        return "CANCELAMENTO_ADIANTAMENTO"
    if "ADIANTAMENTO CFE TITULO" in t:
        return "ADIANTAMENTO"
    if "RECEBIMENTO TITULO" in t:
        return "RECEBIMENTO_TITULO"
    if "PAGAMENTO TITULO" in t:
        return "PAGAMENTO_TITULO"
    if "LANCAMENTO CAIXA" in t:
        return "LANCAMENTO_CAIXA"
    if "CONFORME LANCAMENTO BANCARIO" in t:
        return "LANCAMENTO_BANCARIO"
    return "OUTROS"


def aplicar_compensacao_interna_razao(nao_razao: list[RegistroRazao]) -> tuple[list[RegistroRazao], int]:
    """
    Remove pares internos ENTRADA/SAIDA que se anulam no Razao
    (mesmo valor e mesma referencia/categoria), para reduzir ruído
    de estornos/cancelamentos internos sem reflexo no extrato.
    """
    usados: set[int] = set()
    idx: dict[tuple[str, str, Decimal], list[int]] = defaultdict(list)
    idx_sem_ref: dict[tuple[str, str, str, Decimal], list[int]] = defaultdict(list)
    categorias_permitidas_sem_ref = {
        "ESTORNO",
        "CANCELAMENTO_ADIANTAMENTO",
    }

    for i, r in enumerate(nao_razao):
        ref = f"{r.doc}/{r.parc}".strip("/") if r.doc else ""
        if ref:
            idx[(ref, r.tipo, r.valor)].append(i)
        else:
            cat = categoria_historico_razao(r.historico)
            if cat in categorias_permitidas_sem_ref:
                idx_sem_ref[(cat, r.data_lcto, r.tipo, r.valor)].append(i)

    # 1) por referencia exata
    refs = {k[0] for k in idx.keys()}
    for ref in refs:
        vals = {k[2] for k in idx.keys() if k[0] == ref}
        for v in vals:
            ent = idx.get((ref, "ENTRADA", v), [])
            sai = idx.get((ref, "SAIDA", v), [])
            qtd = min(len(ent), len(sai))
            for a, b in zip(ent[:qtd], sai[:qtd]):
                usados.add(a)
                usados.add(b)

    # 2) sem referencia: por categoria+valor
    cat_datas = {(k[0], k[1]) for k in idx_sem_ref.keys()}
    for cat, data_lcto in cat_datas:
        vals = {k[3] for k in idx_sem_ref.keys() if k[0] == cat and k[1] == data_lcto}
        for v in vals:
            ent = idx_sem_ref.get((cat, data_lcto, "ENTRADA", v), [])
            sai = idx_sem_ref.get((cat, data_lcto, "SAIDA", v), [])
            qtd = min(len(ent), len(sai))
            for a, b in zip(ent[:qtd], sai[:qtd]):
                usados.add(a)
                usados.add(b)

    filtrado = [r for i, r in enumerate(nao_razao) if i not in usados]
    return filtrado, len(usados)


def referencia_registro(doc: str, parc: str, data_lcto: str, tipo: str) -> str:
    ref = f"{doc}/{parc}".strip("/") if doc else ""
    if ref:
        return ref
    # fallback para lancamentos sem doc/parcela
    return f"SEM_REF_{data_lcto}_{tipo}"


def reduzir_linhas_conectadas(linhas: list[list[object]]) -> list[list[object]]:
    """
    Remove duplicidades conectadas no relatorio final.
    Regra: se houver duas linhas com mesma data/tipo/diferenca (somente no Razao),
    sendo uma SEM_REF e outra com referencia real, mantem apenas a referencia real.
    """
    idx_por_chave: dict[tuple[str, str, Decimal], list[int]] = defaultdict(list)
    for i, row in enumerate(linhas):
        ref = str(row[0] or "")
        data_base = str(row[1] or "")
        tipo = str(row[2] or "")
        razao_total = Decimal(str(row[3]))
        extrato_total = Decimal(str(row[4]))
        dif = Decimal(str(row[5]))
        # foco apenas em diferencas "so no razao"
        if extrato_total != 0 or razao_total == 0:
            continue
        chave = (data_base, tipo, dif.copy_abs())
        if ref.startswith("SEM_REF_") or ref:
            idx_por_chave[chave].append(i)

    remover: set[int] = set()
    for _, idxs in idx_por_chave.items():
        if len(idxs) < 2:
            continue
        sem_ref = [i for i in idxs if str(linhas[i][0]).startswith("SEM_REF_")]
        com_ref = [i for i in idxs if not str(linhas[i][0]).startswith("SEM_REF_")]
        if sem_ref and com_ref:
            # mantem as linhas com referencia real e remove duplicatas SEM_REF
            remover.update(sem_ref)

    return [row for i, row in enumerate(linhas) if i not in remover]


def extrair_base_referencia(ref: str) -> str:
    if not ref or ref.startswith("SEM_REF_"):
        return ""
    if "/" in ref:
        return ref.split("/", 1)[0].strip()
    return ref.strip()


def agrupar_por_documento_base(linhas: list[list[object]]) -> list[list[object]]:
    """
    Agrupa diferencas conectadas por documento-base quando houver parcelas.
    Ex.: 12345/01 + 12345/02 -> 12345/*
    """
    grupos: dict[tuple[str, str, str], list[int]] = defaultdict(list)
    for i, row in enumerate(linhas):
        ref = str(row[0] or "")
        base = extrair_base_referencia(ref)
        if not base:
            continue
        # agrupa apenas referencias de parcela para evitar colapsar lancamentos nao relacionados
        if "/" not in ref:
            continue
        tipo = str(row[2] or "")
        motivo = str(row[8] or "")
        grupos[(base, tipo, motivo)].append(i)

    saida: list[list[object]] = []
    consumidos: set[int] = set()

    for (_, _, _), idxs in grupos.items():
        if len(idxs) < 2:
            continue
        idxs_ord = sorted(idxs, key=lambda k: str(linhas[k][1]))
        base = extrair_base_referencia(str(linhas[idxs_ord[0]][0] or ""))
        data_base = str(linhas[idxs_ord[0]][1] or "")
        tipo = str(linhas[idxs_ord[0]][2] or "")
        motivo = str(linhas[idxs_ord[0]][8] or "")
        razao_total = Decimal("0")
        extrato_total = Decimal("0")
        qtd_razao = 0
        qtd_extrato = 0

        for i in idxs_ord:
            consumidos.add(i)
            row = linhas[i]
            razao_total += Decimal(str(row[3]))
            extrato_total += Decimal(str(row[4]))
            qtd_razao += int(row[6])
            qtd_extrato += int(row[7])

        dif = razao_total - extrato_total
        saida.append(
            [
                f"{base}/*",
                data_base,
                tipo,
                float(razao_total),
                float(extrato_total),
                float(dif),
                qtd_razao,
                qtd_extrato,
                motivo,
            ]
        )

    for i, row in enumerate(linhas):
        if i not in consumidos:
            saida.append(row)

    return saida


def normalizar_data_chave(data_lcto: str) -> str:
    p = data_lcto.split("/")
    if len(p) != 3:
        return data_lcto
    d, m, y = p
    return f"{int(d):02d}/{m}/{y}"


def buscar_possiveis_valores_extrato(
    reg_extrato: list[RegistroExtrato],
    data_base: str,
    tipo: str,
    valor_alvo: Decimal,
    limite_resultados: int = 20,
) -> list[tuple[str, Decimal]]:
    """
    Lista valores candidatos no extrato (mesma data/tipo) para analise manual
    quando o valor de referencia pode estar dividido.
    """
    if valor_alvo <= 0:
        return []

    data_ref = normalizar_data_chave(str(data_base))
    candidatos = [
        e
        for e in reg_extrato
        if normalizar_data_chave(e.data_lcto) == data_ref and e.tipo == tipo and e.valor > 0 and e.valor <= valor_alvo + TOLERANCIA
    ]
    if not candidatos:
        return []

    unicos: dict[tuple[str, str], Decimal] = {}
    for c in candidatos:
        ref = c.doc or "SEM_DOC"
        parc = c.parc or ""
        chave = (ref, parc)
        if chave not in unicos:
            unicos[chave] = c.valor

    saida = [((f"{d}/{p}".strip("/")), v) for (d, p), v in unicos.items()]
    saida.sort(key=lambda t: (t[1], t[0]), reverse=True)
    return saida[:limite_resultados]


def buscar_candidatos_exatos_extrato(
    reg_extrato: list[RegistroExtrato],
    data_base: str,
    tipo: str,
    valor_alvo: Decimal,
    limite_resultados: int = 40,
) -> list[tuple[str, Decimal]]:
    """
    Retorna todos os candidatos exatos (mesma data, tipo e valor)
    para demarcar casos ambíguos sem escolher correspondência automática.
    """
    data_ref = normalizar_data_chave(str(data_base))
    unicos: dict[tuple[str, str], Decimal] = {}
    for e in reg_extrato:
        if normalizar_data_chave(e.data_lcto) != data_ref:
            continue
        if e.tipo != tipo:
            continue
        if abs(e.valor - valor_alvo) > TOLERANCIA:
            continue
        ref = e.doc or "SEM_DOC"
        parc = e.parc or ""
        chave = (ref, parc)
        if chave not in unicos:
            unicos[chave] = e.valor
    saida = [((f"{d}/{p}".strip("/")), v) for (d, p), v in unicos.items()]
    saida.sort(key=lambda t: (t[1], t[0]), reverse=True)
    return saida[:limite_resultados]


def escrever_xlsx(
    caminho: Path,
    reg_razao: list[RegistroRazao],
    reg_extrato: list[RegistroExtrato],
    matches: list[tuple[RegistroRazao, RegistroExtrato]],
    nao_razao: list[RegistroRazao],
    nao_extrato: list[RegistroExtrato],
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "DIFERENCAS"
    ws.append(["REFERENCIA", "DATA_BASE", "TIPO", "RAZAO_TOTAL", "EXTRATO_TOTAL", "DIFERENCA", "QTD_RAZAO", "QTD_EXTRATO", "MOTIVO"])

    acumulado: dict[tuple[str, str], dict[str, object]] = {}
    for r in nao_razao:
        ref = referencia_registro(r.doc, r.parc, r.data_lcto, r.tipo)
        key = (ref, r.tipo)
        if key not in acumulado:
            acumulado[key] = {
                "data_base": r.data_lcto,
                "razao_total": Decimal("0"),
                "extrato_total": Decimal("0"),
                "qtd_razao": 0,
                "qtd_extrato": 0,
                "motivos": Counter(),
            }
        item = acumulado[key]
        item["razao_total"] = item["razao_total"] + r.valor  # type: ignore[operator]
        item["qtd_razao"] = item["qtd_razao"] + 1  # type: ignore[operator]
        item["motivos"][sugerir_motivo_inconsistencia_razao(r)] += 1  # type: ignore[index]

    for e in nao_extrato:
        ref = referencia_registro(e.doc, e.parc, e.data_lcto, e.tipo)
        key = (ref, e.tipo)
        if key not in acumulado:
            acumulado[key] = {
                "data_base": e.data_lcto,
                "razao_total": Decimal("0"),
                "extrato_total": Decimal("0"),
                "qtd_razao": 0,
                "qtd_extrato": 0,
                "motivos": Counter(),
            }
        item = acumulado[key]
        item["extrato_total"] = item["extrato_total"] + e.valor  # type: ignore[operator]
        item["qtd_extrato"] = item["qtd_extrato"] + 1  # type: ignore[operator]
        item["motivos"]["LANCAMENTO_SEM_CORRESPONDENTE_NO_RAZAO"] += 1  # type: ignore[index]

    linhas: list[list[object]] = []
    for (ref, tipo), item in acumulado.items():
        razao_total: Decimal = item["razao_total"]  # type: ignore[assignment]
        extrato_total: Decimal = item["extrato_total"]  # type: ignore[assignment]
        dif = razao_total - extrato_total
        if abs(dif) <= TOLERANCIA:
            continue
        motivos: Counter = item["motivos"]  # type: ignore[assignment]
        if razao_total != 0 and extrato_total == 0:
            motivo_principal = motivos.most_common(1)[0][0] if motivos else "SO_NO_RAZAO"
        elif extrato_total != 0 and razao_total == 0:
            motivo_principal = "SO_NO_EXTRATO"
        else:
            motivo_principal = "VALOR_TOTAL_DIVERGENTE"
        linhas.append(
            [
                ref,
                item["data_base"],
                tipo,
                float(razao_total),
                float(extrato_total),
                float(dif),
                item["qtd_razao"],
                item["qtd_extrato"],
                motivo_principal,
            ]
        )

    if not linhas:
        ws.append(["SEM_DIFERENCAS", "", "", 0.0, 0.0, 0.0, 0, 0, ""])
    else:
        linhas = reduzir_linhas_conectadas(linhas)
        linhas = agrupar_por_documento_base(linhas)
        linhas.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[2]), float(r[5])))
        for row in linhas:
            ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=6):
        for c in row:
            if isinstance(c.value, (int, float)):
                c.number_format = XLSX_FORMATO_MOEDA_BR
    ws.freeze_panes = "A2"
    autoajustar_colunas(ws)

    # Aba simplificada: apenas referencias no formato doc/parcela (NFS-e erradas)
    ws_nf = wb.create_sheet("NFSE_ERRADAS")
    ws_nf.append(["REFERENCIA", "DATA_BASE", "TIPO", "VALOR_ERRADO", "MOTIVO"])

    ws_poss = wb.create_sheet("POSSIVEIS_VALORES")
    ws_poss.append(["REFERENCIA_ERRADA", "DATA_BASE", "TIPO", "VALOR_ERRADO", "CRITERIO", "REF_EXTRATO", "VALOR_EXTRATO"])

    for row in linhas:
        ref = str(row[0] or "")
        if "/" not in ref or ref.startswith("SEM_REF_"):
            continue
        valor_errado = Decimal(str(row[5]))
        ws_nf.append([row[0], row[1], row[2], float(valor_errado), row[8]])

        candidatos_exatos = buscar_candidatos_exatos_extrato(
            reg_extrato=reg_extrato,
            data_base=str(row[1] or ""),
            tipo=str(row[2] or ""),
            valor_alvo=valor_errado.copy_abs(),
        )
        if len(candidatos_exatos) > 1:
            for ref_ext, val_ext in candidatos_exatos:
                ws_poss.append([row[0], row[1], row[2], float(valor_errado), "MULTIPLOS_EXATOS", ref_ext, float(val_ext)])

        possiveis = buscar_possiveis_valores_extrato(
            reg_extrato=reg_extrato,
            data_base=str(row[1] or ""),
            tipo=str(row[2] or ""),
            valor_alvo=valor_errado.copy_abs(),
        )
        for ref_ext, val_ext in possiveis:
            ws_poss.append([row[0], row[1], row[2], float(valor_errado), "POSSIVEL_DIVIDIDO", ref_ext, float(val_ext)])
    for c in ws_nf["D"][1:]:
        if isinstance(c.value, (int, float)):
            c.number_format = XLSX_FORMATO_MOEDA_BR
    ws_nf.freeze_panes = "A2"
    autoajustar_colunas(ws_nf)

    for c in ws_poss["D"][1:]:
        if isinstance(c.value, (int, float)):
            c.number_format = XLSX_FORMATO_MOEDA_BR
    for c in ws_poss["G"][1:]:
        if isinstance(c.value, (int, float)):
            c.number_format = XLSX_FORMATO_MOEDA_BR
    ws_poss.freeze_panes = "A2"
    autoajustar_colunas(ws_poss)

    wb.save(caminho)
    return caminho


def main() -> None:
    pasta = descobrir_pasta_pdfs()
    pdfs = sorted(pasta.glob("*.pdf"))
    if len(pdfs) < 2:
        raise RuntimeError("Sao necessarios pelo menos 2 PDFs (Razao e Extrato).")

    razao_pdfs: list[Path] = []
    extrato_pdfs: list[Path] = []
    cache: dict[Path, list[str]] = {}

    for p in pdfs:
        linhas = ler_linhas_pdf(p)
        cache[p] = linhas
        tipo = detectar_tipo_pdf(linhas)
        if tipo == "razao":
            razao_pdfs.append(p)
        elif tipo == "extrato":
            extrato_pdfs.append(p)

    if len(razao_pdfs) != 1 or len(extrato_pdfs) != 1:
        nomes_razao = ", ".join(p.name for p in razao_pdfs) or "nenhum"
        nomes_extrato = ", ".join(p.name for p in extrato_pdfs) or "nenhum"
        raise RuntimeError(
            "Identificacao de PDFs ambigua/incompleta. "
            f"Razao encontrados ({len(razao_pdfs)}): {nomes_razao}. "
            f"Extrato encontrados ({len(extrato_pdfs)}): {nomes_extrato}."
        )

    razao_pdf = razao_pdfs[0]
    extrato_pdf = extrato_pdfs[0]

    reg_razao = parse_razao(cache[razao_pdf])
    reg_extrato = parse_extrato(cache[extrato_pdf])
    matches, nao_razao, nao_extrato = conciliar(reg_razao, reg_extrato)

    compensados = 0
    if MODO_EXPERIMENTAL_COMPENSAR_RAZAO:
        nao_razao, compensados = aplicar_compensacao_interna_razao(nao_razao)

    PASTA_ARQUIVOS.mkdir(parents=True, exist_ok=True)
    out = PASTA_ARQUIVOS / f"Conciliacao_Sicredi_{datetime.now().strftime('%H-%M-%S_%d-%m-%Y')}.xlsx"
    escrever_xlsx(out, reg_razao, reg_extrato, matches, nao_razao, nao_extrato)

    print(f"Razao identificado: {razao_pdf}")
    print(f"Extrato identificado: {extrato_pdf}")
    print(f"Lancamentos Razao: {len(reg_razao)}")
    print(f"Lancamentos Extrato: {len(reg_extrato)}")
    print(f"Conciliados: {len(matches)}")
    if MODO_EXPERIMENTAL_COMPENSAR_RAZAO:
        print(f"Compensados internos Razao (experimental): {compensados}")
    print(f"Nao conciliado Razao: {len(nao_razao)}")
    print(f"Nao conciliado Extrato: {len(nao_extrato)}")
    print(f"Arquivo gerado: {out}")


if __name__ == "__main__":
    main()
