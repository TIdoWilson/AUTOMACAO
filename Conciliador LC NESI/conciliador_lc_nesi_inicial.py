from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from tkinter import StringVar, Tk, filedialog, messagebox, ttk
from unicodedata import normalize

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


PASTA_BASE = Path(__file__).resolve().parent
PASTA_ARQUIVOS = PASTA_BASE / "Arquivos"
TOLERANCIA_DIFERENCA = Decimal("0.10")
XLSX_FORMATO_MOEDA_BR = '[$R$-416] #,##0.00'

RE_MONEY = re.compile(r"(\d{1,3}(?:\.\d{3})*,\d{2})")
RE_CFOP_PONTUADO = re.compile(r"\b([1-7]\.\d{3})\b")
RE_CFOP_SECO = re.compile(r"\b([1-7]\d{3})\b")

RE_LIVRO_ENTRADA = re.compile(
    r"^(\d{2}/\d{2}/\d{4})\s+(NFE|NFCE|CTE|NS)\s+\d+\s+(\d+)\s+"
    r"\d{2}/\d{2}/\d{4}\s+\d+\s+[A-Z]{2}\s+\d{2}\s*/\s*00\s+([\d\.,]+)\s+([1-7]\.\d{3})"
)
RE_LIVRO_ENTRADA_CONT = re.compile(r"^([\d\.,]+)\s+([1-7]\.\d{3})(?:\s|$)")
RE_LIVRO_SAIDA = re.compile(
    r"^(NFE|NFCE|CTE|NS)\s+\d+\s+(\d+)\s+\d+\s+\d{2}\s+(?:[A-Z]{2}\s+)?\d{2}\s*/\s*00\s+([\d\.,]+)\s+([1-7]\.\d{3})"
)
RE_LIVRO_SAIDA_CONT = re.compile(r"^([\d\.,]+)\s+([1-7]\.\d{3})\s+ICMS\b")


@dataclass(frozen=True)
class Registro:
    nota: str
    cfop: str
    valor: Decimal


def normalizar_texto(s: str) -> str:
    return normalize("NFKD", s).encode("ASCII", "ignore").decode().upper()


def valor_br_para_decimal(s: str) -> Decimal:
    return Decimal(s.replace(".", "").replace(",", "."))


def normalizar_nota(s: str) -> str:
    return re.sub(r"\D", "", s).lstrip("0") or "0"


def ler_linhas_pdf(pdf_path: Path) -> list[str]:
    txt = "\n".join((pg.extract_text() or "") for pg in PdfReader(str(pdf_path)).pages)
    return [ln.strip() for ln in normalizar_texto(txt).splitlines() if ln.strip()]


def distancia_digitos(a: str, b: str) -> int:
    if len(a) != len(b):
        return 99
    return sum(1 for x, y in zip(a, b) if x != y)


def escolher_nota_por_sufixo(payload: str, notas_livro: set[str], max_dist: int = 1) -> str | None:
    melhor: tuple[int, str] | None = None
    for nota in notas_livro:
        if len(nota) < 3 or len(payload) < len(nota):
            continue
        tail = payload[-len(nota) :]
        dist = distancia_digitos(tail, nota)
        if dist > max_dist:
            continue
        score = dist * 100 - len(nota)
        cand = (score, nota)
        if melhor is None or cand < melhor:
            melhor = cand
    return melhor[1] if melhor else None


def detectar_tipo_movimento(linhas: list[str]) -> tuple[str | None, str | None]:
    txt = "".join(linhas[:250])
    tipo = None
    movimento = None

    if "REGISTRO DE ENTRADAS - MODELO P1" in txt:
        tipo = "livro"
        movimento = "entradas"
    elif "REGISTRO DE SAIDAS - MODELO P2" in txt:
        tipo = "livro"
        movimento = "saidas"
    elif "REGISTROS DE ENTRADAS FS-001" in txt:
        tipo = "relatorio"
        movimento = "entradas"
    elif "REGISTROS DE SAIDAS FS-002" in txt:
        tipo = "relatorio"
        movimento = "saidas"

    return tipo, movimento


def parse_livro_entradas(linhas: list[str]) -> list[Registro]:
    registros: list[Registro] = []
    nota_atual: str | None = None

    for ln in linhas:
        m = RE_LIVRO_ENTRADA.match(ln)
        if m:
            _, especie, nota, tok, cfop = m.groups()
            nota_atual = normalizar_nota(nota)
            valor = valor_br_para_decimal(RE_MONEY.search(tok).group(1)) if RE_MONEY.search(tok) else None
            if valor is None or valor == 0:
                continue
            if especie == "CTE":
                continue
            registros.append(Registro(nota=nota_atual, cfop=cfop, valor=valor))
            continue

        c = RE_LIVRO_ENTRADA_CONT.match(ln)
        if c and nota_atual:
            tok, cfop = c.groups()
            mv = RE_MONEY.search(tok)
            if not mv:
                continue
            valor = valor_br_para_decimal(mv.group(1))
            if valor == 0:
                continue
            registros.append(Registro(nota=nota_atual, cfop=cfop, valor=valor))

    return registros


def parse_livro_saidas(linhas: list[str]) -> list[Registro]:
    registros: list[Registro] = []
    nota_atual: str | None = None
    for ln in linhas:
        if "CANCELADA" in ln:
            nota_atual = None
            continue
        m = RE_LIVRO_SAIDA.match(ln)
        if not m:
            c = RE_LIVRO_SAIDA_CONT.match(ln)
            if c and nota_atual:
                tok, cfop = c.groups()
                mv = RE_MONEY.search(tok)
                if mv:
                    valor = valor_br_para_decimal(mv.group(1))
                    if valor > 0:
                        registros.append(Registro(nota=nota_atual, cfop=cfop, valor=valor))
            continue
        especie, nota, tok, cfop = m.groups()
        if especie == "CTE":
            nota_atual = None
            continue
        mv = RE_MONEY.search(tok)
        if not mv:
            nota_atual = None
            continue
        valor = valor_br_para_decimal(mv.group(1))
        if valor == 0:
            nota_atual = None
            continue
        nota_atual = normalizar_nota(nota)
        registros.append(Registro(nota=nota_atual, cfop=cfop, valor=valor))
    return registros


def parse_relatorio_entradas(
    linhas: list[str],
    notas_livro: set[str],
    cfops_livro_por_nota: dict[str, set[str]],
) -> list[Registro]:
    registros: list[Registro] = []

    for ln in linhas:
        if "REGISTROS DE ENTRADAS FS-001" in ln or "TOTAL" in ln or "TRANSPORTAR" in ln:
            continue
        if "NOTA FISCAL CANCELADA" in ln:
            continue

        m = re.match(
            r"^\d{2}/\d{2}/\d{4}(NFE|NFCE|CTE|NS)\s+\d+\s+(\d{6,})\s+\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\d{2}/\d{2}/\d{4}\s+[A-Z]{2}\s+(\d{1,3}(?:\.\d{3})*,\d{2})\s+([1-3]\.\d{3})$",
            ln,
        )
        if not m:
            continue
        especie, payload, valor_txt, cfop = m.groups()
        if especie == "CTE":
            continue

        nota = escolher_nota_por_sufixo(payload, notas_livro, max_dist=0)
        if not nota:
            continue
        # Mantem CFOP consistente com o Livro quando houver mapeamento da nota.
        cfops_validos = cfops_livro_por_nota.get(nota)
        if cfops_validos and cfop not in cfops_validos:
            continue
        valor = valor_br_para_decimal(valor_txt)
        if valor > 0:
            registros.append(Registro(nota=nota, cfop=cfop, valor=valor))

    return registros


def parse_relatorio_saidas(linhas: list[str], notas_livro: set[str]) -> list[Registro]:
    registros: list[Registro] = []

    for ln in linhas:
        if "REGISTROS DE SAIDAS FS-002" in ln or "TOTAL" in ln or "TRANSPORTAR" in ln:
            continue
        if "NOTA FISCAL CANCELADA" in ln:
            continue

        # Linha principal de nota no FS-002 (NFE/NFCE).
        linha_principal = re.search(r"\d{2}/\d{2}/\d{4}(NFE|NFCE)", ln)
        if linha_principal:
            end_digits = re.search(r"(\d{4,})\s*$", ln)
            candidate_note = normalizar_nota(end_digits.group(1)) if end_digits else None
            if not candidate_note or candidate_note not in notas_livro:
                if end_digits:
                    candidate_note = escolher_nota_por_sufixo(end_digits.group(1), notas_livro, max_dist=0)
            if not candidate_note:
                continue

            cfops = [c[0] + "." + c[1:] for c in RE_CFOP_SECO.findall(ln) if c.startswith(("5", "6", "7"))]
            money = RE_MONEY.findall(ln)
            if not money or not cfops:
                continue
            cfop = cfops[0]
            # No FS-002, o primeiro valor da linha principal corresponde ao valor contabil da nota.
            valor = valor_br_para_decimal(money[0])
            if valor > 0:
                registros.append(Registro(nota=candidate_note, cfop=cfop, valor=valor))

    return registros


def agrupar_por_nota(registros: list[Registro]) -> dict[str, Decimal]:
    out: dict[str, Decimal] = defaultdict(Decimal)
    for r in registros:
        out[r.nota] += r.valor
    return dict(out)


def agrupar_por_nota_cfop(registros: list[Registro]) -> dict[tuple[str, str], Decimal]:
    out: dict[tuple[str, str], Decimal] = defaultdict(Decimal)
    for r in registros:
        out[(r.nota, r.cfop)] += r.valor
    return dict(out)


def autoajustar_colunas(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            txt = "" if val is None else str(val)
            if len(txt) > max_len:
                max_len = len(txt)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 60))


def escrever_xlsx(caminho: Path, divs: list[list[object]], cfops: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "DIVERGENCIAS"
    ws.append(["TIPO", "NOTA", "DIFERENCA"])
    for r in divs:
        ws.append(r)
    ws["E1"] = "DIVERGENCIA TOTAL"
    ws["E2"] = "=SUM(C:C)"
    ws["E2"].number_format = XLSX_FORMATO_MOEDA_BR
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for c in row:
            if isinstance(c.value, (int, float)):
                c.number_format = XLSX_FORMATO_MOEDA_BR
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("CFOP_DIFERENTE")
    ws2.append(["NOTA", "CFOP", "VALOR_LIVRO", "VALOR_RELATORIO"])
    for r in cfops:
        ws2.append(r)
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=3, max_col=4):
        for c in row:
            if isinstance(c.value, (int, float)):
                c.number_format = XLSX_FORMATO_MOEDA_BR
    ws2.freeze_panes = "A2"

    autoajustar_colunas(ws)
    autoajustar_colunas(ws2)
    wb.save(caminho)


def selecionar_dois_pdfs() -> tuple[Path, Path] | None:
    root = Tk()
    root.title("Conciliador LC NESI")
    root.geometry("860x320")
    frame = ttk.Frame(root, padding=16)
    frame.pack(fill="both", expand=True)

    ttk.Label(frame, text="Conciliador LC NESI", font=("Segoe UI", 13, "bold")).grid(row=0, column=0, sticky="w")
    ttk.Label(frame, text="Selecione 2 PDFs de uma vez (1 Livro e 1 Relatorio FS).").grid(row=1, column=0, sticky="w", pady=(6, 10))

    var = StringVar(value="Nenhum arquivo selecionado.")
    result: dict[str, tuple[Path, Path] | None] = {"files": None}

    def escolher() -> None:
        arquivos = filedialog.askopenfilenames(
            title="Selecione 2 arquivos PDF",
            filetypes=[("Arquivos PDF", "*.pdf *.PDF"), ("Todos", "*.*")],
        )
        if not arquivos:
            return
        if len(arquivos) != 2:
            messagebox.showwarning("Atencao", "Selecione exatamente 2 PDFs.")
            return
        a, b = Path(arquivos[0]), Path(arquivos[1])
        if a.resolve() == b.resolve():
            messagebox.showwarning("Atencao", "Selecione 2 arquivos diferentes.")
            return
        result["files"] = (a, b)
        var.set(f"1) {a.name}\n2) {b.name}")

    def confirmar() -> None:
        if not result["files"]:
            messagebox.showwarning("Atencao", "Selecione os 2 PDFs.")
            return
        root.destroy()

    def cancelar() -> None:
        result["files"] = None
        root.destroy()

    ttk.Button(frame, text="Selecionar os 2 PDFs", command=escolher).grid(row=2, column=0, sticky="w")
    ttk.Label(frame, textvariable=var, justify="left").grid(row=3, column=0, sticky="w", pady=(10, 0))

    b = ttk.Frame(frame)
    b.grid(row=4, column=0, sticky="e", pady=(16, 0))
    ttk.Button(b, text="Cancelar", command=cancelar).pack(side="right")
    ttk.Button(b, text="Comparar", command=confirmar).pack(side="right", padx=(0, 8))

    root.protocol("WM_DELETE_WINDOW", cancelar)
    root.mainloop()
    return result["files"]


def processar_pair(pdf_a: Path, pdf_b: Path) -> Path:
    linhas_a = ler_linhas_pdf(pdf_a)
    linhas_b = ler_linhas_pdf(pdf_b)
    tipo_a, mov_a = detectar_tipo_movimento(linhas_a)
    tipo_b, mov_b = detectar_tipo_movimento(linhas_b)

    if tipo_a is None or tipo_b is None:
        raise ValueError("Nao foi possivel identificar os modelos dos PDFs.")
    if tipo_a == tipo_b:
        raise ValueError("Selecione 1 Livro e 1 Relatorio FS.")
    if mov_a != mov_b:
        raise ValueError("Os arquivos devem ser do mesmo movimento (Entradas ou Saidas).")
    if mov_a is None:
        raise ValueError("Nao foi possivel identificar o movimento.")

    if tipo_a == "livro":
        linhas_livro, linhas_rel = linhas_a, linhas_b
    else:
        linhas_livro, linhas_rel = linhas_b, linhas_a

    if mov_a == "entradas":
        reg_livro = parse_livro_entradas(linhas_livro)
    else:
        reg_livro = parse_livro_saidas(linhas_livro)
    if not reg_livro:
        raise RuntimeError("Livro sem registros parseados.")

    mapa_livro = agrupar_por_nota(reg_livro)
    mapa_livro_cfop = agrupar_por_nota_cfop(reg_livro)
    notas_livro = set(mapa_livro.keys())
    cfops_livro_por_nota: dict[str, set[str]] = defaultdict(set)
    for n, c in mapa_livro_cfop:
        cfops_livro_por_nota[n].add(c)

    if mov_a == "entradas":
        reg_rel = parse_relatorio_entradas(linhas_rel, notas_livro, dict(cfops_livro_por_nota))
    else:
        reg_rel = parse_relatorio_saidas(linhas_rel, notas_livro)
    if not reg_rel:
        raise RuntimeError("Relatorio FS sem registros parseados.")

    mapa_rel = agrupar_por_nota(reg_rel)
    mapa_rel_cfop = agrupar_por_nota_cfop(reg_rel)

    divergencias: list[list[object]] = []
    for nota in sorted(set(mapa_livro) | set(mapa_rel)):
        vl = mapa_livro.get(nota)
        vr = mapa_rel.get(nota)
        if vl is not None and vr is None:
            divergencias.append(["SO_NO_LIVRO", nota, float(vl)])
            continue
        if vr is not None and vl is None:
            divergencias.append(["SO_NO_RELATORIO", nota, float(-vr)])
            continue
        if vl is not None and vr is not None:
            dif = vl - vr
            if abs(dif) > TOLERANCIA_DIFERENCA:
                divergencias.append(["VALOR_DIVERGENTE", nota, float(dif)])

    # Comparacao de CFOP mais conservadora para modelo FS:
    # somente notas com CFOP unico no Livro e no Relatorio.
    cfops_livro_por_nota: dict[str, set[str]] = defaultdict(set)
    for n, c in mapa_livro_cfop:
        cfops_livro_por_nota[n].add(c)
    cfops_rel_por_nota: dict[str, set[str]] = defaultdict(set)
    for n, c in mapa_rel_cfop:
        cfops_rel_por_nota[n].add(c)

    cfop_div: list[list[object]] = []
    notas_cfop_validas = {
        n
        for n in set(cfops_livro_por_nota) & set(cfops_rel_por_nota)
        if len(cfops_livro_por_nota[n]) == 1 and len(cfops_rel_por_nota[n]) == 1
    }
    for nota in sorted(notas_cfop_validas):
        cf_liv = next(iter(cfops_livro_por_nota[nota]))
        cf_rel = next(iter(cfops_rel_por_nota[nota]))
        vl = mapa_livro_cfop.get((nota, cf_liv), Decimal("0"))
        vr = mapa_rel_cfop.get((nota, cf_rel), Decimal("0"))
        if cf_liv != cf_rel or abs(vl - vr) > TOLERANCIA_DIFERENCA:
            cfop_div.append([nota, f"{cf_liv}|{cf_rel}", float(vl), float(vr)])

    PASTA_ARQUIVOS.mkdir(parents=True, exist_ok=True)
    out = PASTA_ARQUIVOS / f"Conciliacao_LC_NESI_{datetime.now().strftime('%H-%M-%S_%d-%m-%Y')}.xlsx"
    escrever_xlsx(out, divergencias, cfop_div)

    print(f"Movimento: {mov_a}")
    print(f"Registros Livro: {len(reg_livro)}")
    print(f"Registros Relatorio: {len(reg_rel)}")
    print(f"Total Livro: {sum(mapa_livro.values(), Decimal('0')):.2f}")
    print(f"Total Relatorio: {sum(mapa_rel.values(), Decimal('0')):.2f}")
    print(f"Diferenca: {(sum(mapa_livro.values(), Decimal('0')) - sum(mapa_rel.values(), Decimal('0'))):.2f}")
    print(f"Linhas divergencias: {len(divergencias)}")
    print(f"Linhas CFOP: {len(cfop_div)}")
    print(f"Arquivo gerado: {out}")
    return out


def main() -> None:
    sel = selecionar_dois_pdfs()
    if not sel:
        print("Operacao cancelada.")
        return
    processar_pair(sel[0], sel[1])


if __name__ == "__main__":
    main()
