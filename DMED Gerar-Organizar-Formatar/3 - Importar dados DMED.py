# -*- coding: utf-8 -*-
"""
DMED - Importacao automatizada em lote
"""

from __future__ import annotations

import sys
import time
import unicodedata
import re
from datetime import datetime
from pathlib import Path

import pyautogui

try:
    from pywinauto import Desktop
except Exception:
    Desktop = None

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill
except Exception:
    Workbook = None
    load_workbook = None
    PatternFill = None

try:
    import pytesseract
except Exception:
    pytesseract = None

# ===================== CONFIGURACAO =====================
TITULO_JANELA_EXATO = "Dmed 2026 - Versão 1.0"
TITULO_JANELA_REGEX = r".*Dmed 2026.*"
TITULO_SUBJANELA_IMPORTACAO_REGEX = r".*Dmed 2026 - Declara..o de Servi.os M.dicos e de Sa.de.*"
TITULO_DIALOGO_ABRIR_REGEX = r".*Abrir declara..o para importa..o.*"
TITULO_JANELA_POS_CTRL_G_REGEX = r".*Dmed 2026 - Declara..o de Servi.os M.dicos e de Sa.de.*"
TITULO_RELATORIO_IMPORTACAO_REGEX = r".*Relat.rio de importa..o.*"

PASTA_ARQUIVOS_BASE = (
    r"W:\DOCUMENTOS ESCRITORIO\INSTALACAO SISTEMA\python\DMED Gerar-Organizar-Formatar\arquivos base"
)
NOME_PLANILHA_REGISTRO = "TXT_Importados_DMED.xlsx"
NOME_PRINT_ERRO_PREFIXO = "print_erro_gravacao_dmed"
TESSERACT_EXE = ""  # ex.: r"C:\Program Files\Tesseract-OCR\tesseract.exe"
STATUS_OK = "OK"
STATUS_ERRO = "ERRO"
STATUS_FALHA = "FALHA"
SUFIXO_TXT_CORRIGIDO = "_corrigido"
ARQUIVOS_TXT_IGNORADOS = {
    "dmed movimentos.txt",
    "dmed_corretor_erros_detectados.txt",
    "dmed_corretor_alterados.txt",
}

COORD_IMPORTAR_DADOS = (736, 606)
COORD_BOTAO_PASTA = (1279, 425)
COORD_NOME_ARQUIVO = (987, 630)
COORD_FOCO_FINAL = (565, 24)
COORD_DIGITAR_PR = (759, 426)

DELAY_ANTES_DO_CLIQUE = 0.4
TIMEOUT_SUBJANELA = 10
TIMEOUT_DIALOGO_ABRIR = 10
TIMEOUT_JANELA_POS_IMPORTACAO = 10
DELAY_ANTES_ALT_O = 1.5
TIMEOUT_RELATORIO_TENTATIVA_1 = 6
TIMEOUT_RELATORIO_TENTATIVA_2 = 5
TIMEOUT_JANELA_POS_CTRL_G = 10
TIMEOUT_RESET_TELA_INICIAL = 10
INTERVALO_OCR = 1.0
ENTERS_APOS_SEGUNDO_ALT_A = 2
TIMEOUT_FECHAR_SUBJANELA_CONCLUSAO = 30
INTERVALO_ENTER_CONCLUSAO = 3.0
ROI_ERRO_RELATIVA = (8, 105, 150, 235)  # area da coluna dos icones/labels de status
LIMIAR_PIXELS_VERMELHOS = 45
# ========================================================


def validar_dependencias() -> bool:
    if Desktop is None:
        print("Dependencia ausente: pywinauto. Instale: pip install pywinauto")
        return False
    if Workbook is None or load_workbook is None:
        print("Dependencia ausente: openpyxl. Instale: pip install openpyxl")
        return False
    if PatternFill is None:
        print("Dependencia ausente: openpyxl.styles. Reinstale: pip install openpyxl")
        return False
    if pytesseract is None:
        print("Dependencia ausente: pytesseract. Instale: pip install pytesseract")
        return False
    if TESSERACT_EXE.strip():
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_EXE.strip()
    return True


def _normalizar_texto(texto: str) -> str:
    sem_acentos = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return sem_acentos.upper()


def _extrair_texto_ocr(imagem) -> str:
    try:
        return pytesseract.image_to_string(imagem, lang="por")
    except Exception:
        return pytesseract.image_to_string(imagem)


def _preprocessar_imagem_ocr(imagem):
    img = imagem.convert("L")
    w, h = img.size
    img = img.resize((w * 2, h * 2))
    img = img.point(lambda p: 255 if p > 180 else 0)
    return img


def _contar_pixels_vermelhos(img, limiar: int = LIMIAR_PIXELS_VERMELHOS) -> int:
    rgb = img.convert("RGB")
    count = 0
    px = rgb.load()
    w, h = rgb.size
    for y in range(h):
        for x in range(w):
            r, g, b = px[x, y]
            if r >= 165 and g <= 125 and b <= 125 and (r - max(g, b)) >= 35:
                count += 1
                if count >= limiar:
                    return count
    return count


def detectar_icone_erro_por_cor(dlg_janela) -> bool:
    try:
        rect = dlg_janela.rectangle()
        x1, y1, x2, y2 = ROI_ERRO_RELATIVA
        rx = rect.left + x1
        ry = rect.top + y1
        rw = max(1, x2 - x1)
        rh = max(1, y2 - y1)
        img = pyautogui.screenshot(region=(rx, ry, rw, rh))
        vermelhos = _contar_pixels_vermelhos(img)
        return vermelhos >= LIMIAR_PIXELS_VERMELHOS
    except Exception:
        return False


def caminho_planilha_registro() -> Path:
    return Path(__file__).resolve().parent / NOME_PLANILHA_REGISTRO


def _garantir_cabecalho_planilha(ws) -> None:
    if ws.max_row == 0:
        ws.append(["DataHora", "ArquivoTXT", "CaminhoCompleto", "Status"])
        return
    cab = [str(c.value or "").strip() for c in ws[1]]
    if not cab:
        ws.append(["DataHora", "ArquivoTXT", "CaminhoCompleto", "Status"])
        return
    if len(cab) < 4:
        ws.cell(row=1, column=4, value="Status")


def carregar_txts_registrados() -> set[str]:
    arquivo = caminho_planilha_registro()
    if not arquivo.exists():
        return set()
    wb = load_workbook(arquivo)
    ws = wb.active
    _garantir_cabecalho_planilha(ws)
    registrados: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        caminho = str(row[2] or "").strip()
        if caminho:
            registrados.add(str(Path(caminho)))
    return registrados


def registrar_txt_importado(caminho_txt: str, status: str) -> None:
    arquivo = caminho_planilha_registro()
    if arquivo.exists():
        wb = load_workbook(arquivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Importados"
    _garantir_cabecalho_planilha(ws)

    txt = Path(caminho_txt)
    datahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([datahora, txt.name, str(txt), status])

    linha = ws.max_row
    if status == STATUS_ERRO:
        cor = PatternFill(fill_type="solid", fgColor="FFF4CCCC")
    else:
        cor = PatternFill(fill_type="solid", fgColor="FFE2F0D9")
    for col in range(1, 5):
        ws.cell(row=linha, column=col).fill = cor

    wb.save(arquivo)
    print(f"Registro adicionado em: {arquivo}")


def proximo_txt_pendente() -> Path | None:
    pasta = Path(PASTA_ARQUIVOS_BASE)
    if not pasta.exists():
        print(f"Pasta nao encontrada: {PASTA_ARQUIVOS_BASE}")
        return None
    candidatos = []
    for p in sorted([x for x in pasta.glob("*.txt") if x.is_file()]):
        if p.name.lower() in ARQUIVOS_TXT_IGNORADOS:
            continue
        candidatos.append(p)

    if not candidatos:
        print("Nenhum .txt encontrado para importar.")
        return None

    # Para cada base logica, prioriza o arquivo _corrigido.txt quando existir.
    escolhidos_por_base: dict[str, Path] = {}
    for txt in candidatos:
        stem = txt.stem
        if stem.endswith(SUFIXO_TXT_CORRIGIDO):
            base_logica = stem[: -len(SUFIXO_TXT_CORRIGIDO)]
            atual = escolhidos_por_base.get(base_logica)
            if atual is None or not atual.stem.endswith(SUFIXO_TXT_CORRIGIDO):
                escolhidos_por_base[base_logica] = txt
        else:
            base_logica = stem
            escolhidos_por_base.setdefault(base_logica, txt)

    registrados = carregar_txts_registrados()
    for base_logica in sorted(escolhidos_por_base.keys()):
        txt = escolhidos_por_base[base_logica]
        caminho_txt = str(txt)
        caminho_original = str(txt.with_name(f"{base_logica}.txt"))
        caminho_corrigido = str(txt.with_name(f"{base_logica}{SUFIXO_TXT_CORRIGIDO}.txt"))

        # Evita reimportar a mesma base logica se original/corrigido ja estiver registrado.
        if caminho_txt in registrados:
            continue
        if caminho_original in registrados or caminho_corrigido in registrados:
            continue

        return txt
    return None


def focar_janela_dmed(timeout: int = 5):
    fim = time.time() + timeout
    while time.time() < fim:
        try:
            dlg_exata = Desktop(backend="uia").window(title=TITULO_JANELA_EXATO)
            if dlg_exata.exists(timeout=0.5):
                dlg_exata.set_focus()
                return dlg_exata
        except Exception:
            pass

        try:
            dlg_regex = Desktop(backend="uia").window(title_re=TITULO_JANELA_REGEX)
            if dlg_regex.exists(timeout=0.5):
                dlg_regex.set_focus()
                return dlg_regex
        except Exception:
            pass

        try:
            for w in Desktop(backend="uia").windows():
                titulo = (w.window_text() or "").upper()
                if "DMED 2026" in titulo and w.is_visible():
                    w.set_focus()
                    return w
        except Exception:
            pass
        time.sleep(0.2)
    return None


def esperar_subjanela_importacao():
    try:
        dlg = Desktop(backend="uia").window(title_re=TITULO_SUBJANELA_IMPORTACAO_REGEX)
        if not dlg.exists(timeout=TIMEOUT_SUBJANELA):
            return None
        dlg.set_focus()
        return dlg
    except Exception:
        return None


def subjanela_importacao_esta_aberta() -> bool:
    try:
        dlg = Desktop(backend="uia").window(title_re=TITULO_SUBJANELA_IMPORTACAO_REGEX)
        return dlg.exists(timeout=0.3)
    except Exception:
        return False


def esperar_dialogo_abrir():
    try:
        dlg = Desktop(backend="uia").window(title_re=TITULO_DIALOGO_ABRIR_REGEX)
        if not dlg.exists(timeout=TIMEOUT_DIALOGO_ABRIR):
            return None
        dlg.set_focus()
        return dlg
    except Exception:
        return None


def esperar_janela_relatorio(timeout_segundos: int):
    try:
        dlg = Desktop(backend="uia").window(title_re=TITULO_RELATORIO_IMPORTACAO_REGEX)
        if not dlg.exists(timeout=timeout_segundos):
            return None
        dlg.set_focus()
        return dlg
    except Exception:
        return None


def esperar_janela_pos_ctrl_g():
    try:
        dlg = Desktop(backend="uia").window(title_re=TITULO_JANELA_POS_CTRL_G_REGEX)
        if not dlg.exists(timeout=TIMEOUT_JANELA_POS_CTRL_G):
            return None
        dlg.set_focus()
        return dlg
    except Exception:
        return None


def detectar_erro_e_printar_janela(dlg_janela) -> bool:
    fim = time.time() + TIMEOUT_JANELA_POS_CTRL_G
    while time.time() < fim:
        rect = dlg_janela.rectangle()
        img = pyautogui.screenshot(region=(rect.left, rect.top, rect.width(), rect.height()))

        # Regra unica: detectar pela "fotinho" vermelha de ERRO na area configurada.
        if detectar_icone_erro_por_cor(dlg_janela):
            pasta_script = Path(__file__).resolve().parent
            carimbo = datetime.now().strftime("%Y%m%d_%H%M%S")
            destino = pasta_script / f"{NOME_PRINT_ERRO_PREFIXO}_{carimbo}.png"
            img.save(destino)
            print(f"ERRO detectado por icone/cor. Print salvo em: {destino}")
            return True
        time.sleep(INTERVALO_OCR)
    return False


def concluir_sem_erro() -> None:
    pyautogui.hotkey("alt", "a")
    time.sleep(0.2)
    pyautogui.click(*COORD_DIGITAR_PR)
    time.sleep(0.15)
    pyautogui.hotkey("ctrl", "a")
    pyautogui.write("PR", interval=0.01)
    pyautogui.press("enter")
    time.sleep(0.2)
    pyautogui.hotkey("alt", "a")
    time.sleep(0.2)

    # Sequencia reforcada: 2 enters imediatos + 1 apos 3s.
    pyautogui.press("enter")
    time.sleep(0.2)
    pyautogui.press("enter")
    time.sleep(INTERVALO_ENTER_CONCLUSAO)
    pyautogui.press("enter")

    # Continua tentando ate a subjanela sumir (com limite de seguranca).
    fim = time.time() + TIMEOUT_FECHAR_SUBJANELA_CONCLUSAO
    while subjanela_importacao_esta_aberta() and time.time() < fim:
        time.sleep(INTERVALO_ENTER_CONCLUSAO)
        pyautogui.press("enter")

    if subjanela_importacao_esta_aberta():
        print("Aviso: subjanela de importacao ainda aberta apos tentativas de Enter.")
    else:
        print("Fluxo de conclusao executado e subjanela fechada.")


def fechar_verificacao_com_alt_c() -> None:
    pyautogui.click(*COORD_FOCO_FINAL)
    time.sleep(0.1)
    pyautogui.hotkey("alt", "c")
    print("Janela de verificacao fechada com Alt+C.")


def validar_tela_inicial_por_ocr(dlg_principal) -> bool:
    fim = time.time() + TIMEOUT_RESET_TELA_INICIAL
    while time.time() < fim:
        rect = dlg_principal.rectangle()
        img = pyautogui.screenshot(region=(rect.left, rect.top, rect.width(), rect.height()))
        texto = _normalizar_texto(_extrair_texto_ocr(img))
        if "IMPORTAR DADOS" in texto or "NOVA DECLARACAO" in texto:
            return True
        time.sleep(INTERVALO_OCR)
    return False


def resetar_para_tela_inicial_e_validar() -> bool:
    pyautogui.click(*COORD_FOCO_FINAL)
    time.sleep(0.1)
    pyautogui.hotkey("alt", "d")
    time.sleep(0.1)
    pyautogui.hotkey("ctrl", "f")
    print("Reset para tela inicial acionado com Alt+D e Ctrl+F.")

    time.sleep(0.8)
    dlg = focar_janela_dmed(timeout=2)
    if dlg is None:
        print("Falha: janela principal do DMED nao encontrada apos reset.")
        return False

    ok_inicio = validar_tela_inicial_por_ocr(dlg)
    if ok_inicio:
        print("Tela inicial do DMED confirmada por OCR.")
        return True

    print("Falha: apos reset, tela inicial nao foi confirmada por OCR.")
    return False


def processar_txt(txt_importacao: Path) -> str:
    # Garante foco no DMED no inicio de cada rodada (inclusive a primeira).
    dlg = focar_janela_dmed()
    if dlg is None:
        print("Janela do DMED nao encontrada.")
        return STATUS_FALHA

    time.sleep(DELAY_ANTES_DO_CLIQUE)
    pyautogui.click(*COORD_IMPORTAR_DADOS)
    print(f"Clique em Importar dados: {COORD_IMPORTAR_DADOS}")

    dlg_sub = esperar_subjanela_importacao()
    if dlg_sub is None:
        print("Subjanela de importacao nao apareceu.")
        return STATUS_FALHA

    time.sleep(0.25)
    pyautogui.click(*COORD_BOTAO_PASTA)
    print(f"Clique no botao pasta: {COORD_BOTAO_PASTA}")

    dlg_abrir = esperar_dialogo_abrir()
    if dlg_abrir is None:
        print("Dialogo para abrir arquivo nao apareceu.")
        return STATUS_FALHA

    time.sleep(0.2)
    pyautogui.click(*COORD_NOME_ARQUIVO)
    time.sleep(0.15)
    pyautogui.hotkey("ctrl", "a")
    pyautogui.write(str(txt_importacao), interval=0.01)
    pyautogui.press("enter")
    print(f"Arquivo enviado para importacao: {txt_importacao}")

    time.sleep(0.2)
    pyautogui.hotkey("alt", "a")
    if esperar_subjanela_importacao() is None:
        print("Janela de importacao nao apareceu apos Alt+A.")
        return STATUS_FALHA

    time.sleep(DELAY_ANTES_ALT_O)
    pyautogui.hotkey("alt", "o")
    dlg_rel = esperar_janela_relatorio(TIMEOUT_RELATORIO_TENTATIVA_1)
    dlg_ctrl_g = None

    if dlg_rel is None:
        pyautogui.hotkey("alt", "o")
        dlg_rel = esperar_janela_relatorio(TIMEOUT_RELATORIO_TENTATIVA_2)
        if dlg_rel is None:
            print("Relatorio nao apareceu. Tentando seguir com Ctrl+G.")
            pyautogui.click(*COORD_FOCO_FINAL)
            time.sleep(0.2)
            pyautogui.hotkey("ctrl", "g")
            dlg_ctrl_g = esperar_janela_pos_ctrl_g()
            if dlg_ctrl_g is None:
                print("Falha: nem relatorio nem janela apos Ctrl+G apareceram.")
                return STATUS_FALHA
    else:
        time.sleep(0.2)
        pyautogui.hotkey("alt", "f4")
        print("Relatorio fechado com Alt+F4.")

    if dlg_ctrl_g is None:
        time.sleep(0.2)
        pyautogui.click(*COORD_FOCO_FINAL)
        time.sleep(0.2)
        pyautogui.hotkey("ctrl", "g")
        dlg_ctrl_g = esperar_janela_pos_ctrl_g()

    if dlg_ctrl_g is None:
        print("Janela apos Ctrl+G nao apareceu.")
        return STATUS_FALHA

    time.sleep(0.2)
    tem_erro = detectar_erro_e_printar_janela(dlg_ctrl_g)
    if tem_erro:
        fechar_verificacao_com_alt_c()
        print("Importacao com ERRO detectado. Marcando no registro e seguindo para o proximo.")
        return STATUS_ERRO

    concluir_sem_erro()
    return STATUS_OK


def main() -> int:
    if not validar_dependencias():
        return 1

    pyautogui.FAILSAFE = True
    pyautogui.PAUSE = 0.08

    while True:
        txt_pendente = proximo_txt_pendente()
        if txt_pendente is None:
            print("Nao ha mais .txt pendente para importar.")
            return 0

        print(f"Iniciando importacao do TXT: {txt_pendente}")
        status = processar_txt(txt_pendente)
        if status == STATUS_FALHA:
            return 1

        registrar_txt_importado(str(txt_pendente), status)
        if not resetar_para_tela_inicial_e_validar():
            return 1


if __name__ == "__main__":
    sys.exit(main())

