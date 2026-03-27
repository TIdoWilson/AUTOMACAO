# -*- coding: utf-8 -*-
"""
DMED - Parte 1 (resumo extremo)
"""

from __future__ import annotations

import os
import time
from datetime import datetime
import ctypes
import re
import unicodedata

import pyautogui

try:
    from pywinauto import Desktop
except Exception as exc:
    Desktop = None
    print(f"Falha ao importar pywinauto: {exc}")

# ===================== CONFIGURACAO =====================
TITULO_JANELA = "Fiscal"
COORD_EMPRESA_SISTEMA = (1637, 58)
COORD_FOCO = (1600, 200)
COORD_CHECKBOX_FALLBACK = (815, 622)
COORD_SALVAR_FLUXO_FOCO = (671, 326)
COORD_SALVAR_FLUXO_BOTAO_SALVAR = (402, 83)
COORD_SALVAR_FLUXO_VOLTAR = (467, 85)
COORD_SALVAR_FLUXO_CONFIRMAR = (1745, 1007)
BACKSPACES_EMPRESA = 5
DELAY_POS_ALT_MGM = 0.2
TIMEOUT_AVISO_1 = 10
TIMEOUT_AVISO_2 = 5
DELAY_ENTRE_TEXTO_E_ENTER = 0.1
DELAY_APOS_ALT_G = 0.2
DELAY_APOS_VOLTAR = 0.1

DIR_BASE_TEMPLATE = r"W:\DOCUMENTOS ESCRITORIO\INSTALACAO SISTEMA\python\DMED Gerar-Organizar-Formatar\arquivos base"
LISTA_TEMPLATE = r"W:\DECLARAÇÕES\DMED\DMED {ano}\LISTA.xlsx"
NOME_ARQUIVO_LISTA_NEGRA = "Lista Negra.xlsx"
# ========================================================


def focar_janela_fiscal():
    if Desktop is None:
        print("pywinauto nao disponivel. Foque manualmente a janela 'Fiscal'.")
        return None

    try:
        dlg = Desktop(backend="uia").window(title_re=f".*{TITULO_JANELA}.*")
        if not dlg.exists(timeout=3):
            print("Janela 'Fiscal' nao encontrada. Foque manualmente.")
            return None
        dlg.set_focus()
        try:
            dlg.maximize()
        except Exception:
            pass
        return dlg
    except Exception as exc:
        print(f"Falha ao focar a janela 'Fiscal': {exc}")
        return None


def clicar(x: int, y: int) -> None:
    pyautogui.click(x, y)


def escrever(texto: str, intervalo: float = 0.02) -> None:
    pyautogui.write(texto, interval=intervalo)


def escrever_unicode(texto: str, intervalo: float = 0.01) -> None:
    KEYEVENTF_UNICODE = 0x0004
    KEYEVENTF_KEYUP = 0x0002

    class KEYBDINPUT(ctypes.Structure):
        _fields_ = [
            ("wVk", ctypes.c_ushort),
            ("wScan", ctypes.c_ushort),
            ("dwFlags", ctypes.c_ulong),
            ("time", ctypes.c_ulong),
            ("dwExtraInfo", ctypes.c_void_p),
        ]

    class INPUT(ctypes.Structure):
        _fields_ = [("type", ctypes.c_ulong), ("ki", KEYBDINPUT)]

    def _send_char(ch: str) -> None:
        code = ord(ch)
        inp_down = INPUT(1, KEYBDINPUT(0, code, KEYEVENTF_UNICODE, 0, None))
        inp_up = INPUT(1, KEYBDINPUT(0, code, KEYEVENTF_UNICODE | KEYEVENTF_KEYUP, 0, None))
        ctypes.windll.user32.SendInput(1, ctypes.byref(inp_down), ctypes.sizeof(inp_down))
        ctypes.windll.user32.SendInput(1, ctypes.byref(inp_up), ctypes.sizeof(inp_up))

    for ch in texto:
        _send_char(ch)
        if intervalo:
            time.sleep(intervalo)


def _set_clipboard_text(texto: str) -> None:
    CF_UNICODETEXT = 13
    GMEM_MOVEABLE = 0x0002

    buf = ctypes.create_unicode_buffer(texto)
    tamanho = ctypes.sizeof(buf)

    if not ctypes.windll.user32.OpenClipboard(None):
        raise RuntimeError("Nao foi possivel abrir a area de transferencia.")
    try:
        ctypes.windll.user32.EmptyClipboard()
        hglobal = ctypes.windll.kernel32.GlobalAlloc(GMEM_MOVEABLE, tamanho)
        if not hglobal:
            raise RuntimeError("Falha ao alocar memoria para a area de transferencia.")
        lp = ctypes.windll.kernel32.GlobalLock(hglobal)
        if not lp:
            raise RuntimeError("Falha ao travar memoria da area de transferencia.")
        ctypes.memmove(lp, buf, tamanho)
        ctypes.windll.kernel32.GlobalUnlock(hglobal)
        ctypes.windll.user32.SetClipboardData(CF_UNICODETEXT, hglobal)
    finally:
        ctypes.windll.user32.CloseClipboard()


def colar_focado(texto: str) -> bool:
    try:
        _set_clipboard_text(texto)
        pyautogui.hotkey("ctrl", "a")
        pyautogui.hotkey("ctrl", "v")
        return True
    except Exception:
        return False


def mandar_backspaces(qtd: int) -> None:
    for _ in range(qtd):
        pyautogui.press("backspace")


def enter_com_delay(qtd: int = 1) -> None:
    for _ in range(qtd):
        pyautogui.press("enter")
        time.sleep(DELAY_ENTRE_TEXTO_E_ENTER)


def delay_pos_alt_g() -> None:
    time.sleep(DELAY_APOS_ALT_G)


def alt_mgm() -> None:
    pyautogui.keyDown("alt")
    pyautogui.press("m")
    pyautogui.press("g")
    pyautogui.press("m")
    pyautogui.keyUp("alt")


def garantir_checkbox(dlg) -> None:
    clicar(*COORD_CHECKBOX_FALLBACK)


def _achar_aviso_sistema(timeout: float):
    if Desktop is None:
        time.sleep(timeout)
        return None
    titulo_regex = r".*Aviso.*Sistema.*"
    end_time = time.time() + timeout
    while time.time() < end_time:
        dlg = _achar_aviso_sistema_uma_varredura(titulo_regex)
        if dlg is not None:
            return dlg

        time.sleep(0.1)
    return None


def _achar_aviso_sistema_uma_varredura(titulo_regex: str):
    # Tenta localizar qualquer variacao de "Aviso ... Sistema" em UIA e win32.
    for backend in ("uia", "win32"):
        try:
            dlg = Desktop(backend=backend).window(title_re=titulo_regex)
            if dlg.exists(timeout=0.1):
                return dlg
        except Exception:
            pass

        try:
            for w in Desktop(backend=backend).windows():
                try:
                    if not w.is_visible():
                        continue
                    titulo = (w.window_text() or "").upper()
                    if "AVISO" in titulo and "SISTEMA" in titulo:
                        return w
                except Exception:
                    continue
        except Exception:
            pass
    return None


def _normalizar_texto(texto: str) -> str:
    sem_acentos = unicodedata.normalize("NFKD", texto or "").encode("ascii", "ignore").decode("ascii")
    return sem_acentos.upper()


def _obter_texto_aviso(dlg) -> str:
    textos: list[str] = []
    try:
        titulo = dlg.window_text() or ""
        if titulo:
            textos.append(titulo)
    except Exception:
        pass

    try:
        for ctrl in dlg.descendants():
            try:
                t = (ctrl.window_text() or "").strip()
                if t:
                    textos.append(t)
            except Exception:
                continue
    except Exception:
        pass

    return " ".join(textos).strip()


def _confirmar_aviso(dlg) -> None:
    if dlg is None:
        return
    try:
        btn_ok = dlg.child_window(title="OK", control_type="Button")
        if btn_ok.exists(timeout=1):
            btn_ok.click_input()
            return
    except Exception:
        pass
    try:
        btn_ok = dlg.child_window(title="OK")
        if btn_ok.exists(timeout=0.5):
            btn_ok.click_input()
            return
    except Exception:
        pass
    pyautogui.press("enter")


def _esperar_e_confirmar_aviso(timeout: float) -> bool:
    dlg = _achar_aviso_sistema(timeout)
    if dlg is None:
        return False
    _confirmar_aviso(dlg)
    return True


def registrar_na_lista_negra(numero_empresa: str, justificativa: str) -> None:
    try:
        from openpyxl import load_workbook, Workbook
    except Exception:
        print("openpyxl nao encontrado para registrar lista negra.")
        return

    caminho_lista_negra = os.path.join(os.path.dirname(os.path.abspath(__file__)), NOME_ARQUIVO_LISTA_NEGRA)

    try:
        if os.path.exists(caminho_lista_negra):
            wb = load_workbook(caminho_lista_negra)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "ListaNegra"
            ws.append(["Numero", "Justificativa"])

        numero_txt = str(numero_empresa or "").strip()
        if not numero_txt:
            return

        for row in ws.iter_rows(min_row=2, values_only=True):
            n = str((row[0] if row else "") or "").strip()
            if n == numero_txt:
                return

        ws.append([numero_txt, justificativa])
        wb.save(caminho_lista_negra)
        print(f"Adicionado na lista negra: {numero_txt} - {justificativa}")
    except Exception as exc:
        print(f"Falha ao registrar lista negra: {exc}")


def _fluxo_salvar_por_coordenadas(caminho: str, nome_arquivo: str) -> None:
    clicar(*COORD_SALVAR_FLUXO_FOCO)
    clicar(*COORD_SALVAR_FLUXO_BOTAO_SALVAR)
    _esperar_janela_salvar()
    caminho_completo = os.path.join(caminho, nome_arquivo)
    escrever(caminho_completo)
    enter_com_delay()
    clicar(*COORD_SALVAR_FLUXO_CONFIRMAR)
    clicar(*COORD_SALVAR_FLUXO_VOLTAR)


def _esperar_janela_salvar() -> None:
    # Aguarda a janela/dialogo de salvar aparecer antes de digitar o caminho.
    # Sem timeout: fica aguardando o tempo necessario.
    if Desktop is None:
        # Fallback minimo sem deteccao de janela.
        time.sleep(3.0)
        return

    while True:
        # Se aparecer aviso antes da janela de salvar, confirma e continua aguardando.
        dlg_aviso = _achar_aviso_sistema_uma_varredura(r".*Aviso.*Sistema.*")
        if dlg_aviso is not None:
            _confirmar_aviso(dlg_aviso)
            time.sleep(0.2)

        try:
            for w in Desktop(backend="win32").windows():
                if not w.is_visible():
                    continue
                titulo = (w.window_text() or "").upper()
                if "SALVAR" in titulo or "SAVE" in titulo:
                    return
        except Exception:
            pass

        try:
            for w in Desktop(backend="uia").windows():
                if not w.is_visible():
                    continue
                titulo = (w.window_text() or "").upper()
                if "SALVAR" in titulo or "SAVE" in titulo:
                    return
        except Exception:
            pass

        time.sleep(0.2)


def registrar_movimento(dir_dest: str, empresa: str) -> None:
    pasta_pai = dir_dest.rstrip("\\/").rsplit("\\", 1)[0]
    caminho_txt = f"{pasta_pai}\\DMED Movimentos.txt"
    try:
        with open(caminho_txt, "a", encoding="utf-8") as f_out:
            f_out.write(f"{empresa}\n")
    except Exception as exc:
        print(f"Falha ao registrar movimento em {caminho_txt}: {exc}")


def _sanitizar_nome_arquivo(texto: str) -> str:
    # Remove caracteres invalidos de nome de arquivo no Windows, sem substituicao.
    texto = re.sub(r'[\\/:*?"<>|]+', "", str(texto or ""))
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto.rstrip(". ")


def montar_base_arquivo(numero_empresa: str, nome_empresa: str) -> str:
    num = _sanitizar_nome_arquivo(numero_empresa)
    nome = _sanitizar_nome_arquivo(nome_empresa)
    if nome and nome != num:
        return f"{nome} - {num}"
    return num


def carregar_lista_negra() -> set[str]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return set()

    caminho_lista_negra = os.path.join(os.path.dirname(os.path.abspath(__file__)), NOME_ARQUIVO_LISTA_NEGRA)
    if not os.path.exists(caminho_lista_negra):
        return set()

    bloqueados: set[str] = set()
    try:
        wb = load_workbook(caminho_lista_negra, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            numero = str(row[0] or "").strip()
            if numero:
                bloqueados.add(numero)
    except Exception as exc:
        print(f"Falha ao ler lista negra: {exc}")
        return set()
    return bloqueados


def carregar_empresas_lista(caminho_lista: str) -> list[tuple[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        print("openpyxl nao encontrado. Instale para ler a LISTA.xlsx.")
        return []

    if not os.path.exists(caminho_lista):
        print(f"Lista nao encontrada: {caminho_lista}")
        return []

    wb = load_workbook(caminho_lista, data_only=True)
    ws = wb.active
    empresas: list[tuple[str, str]] = []

    # Padrao novo: numero na coluna C e nome na coluna D (ao lado do numero).
    # Compatibilidade: se D estiver vazio, usa coluna A como nome.
    for row in ws.iter_rows(min_row=1, values_only=True):
        if len(row) < 3:
            continue
        numero = row[2]
        if numero is None:
            continue
        numero_txt = str(numero).strip()
        if not numero_txt:
            continue

        nome_lado = str(row[3]).strip() if len(row) >= 4 and row[3] is not None else ""
        nome_col_a = str(row[0]).strip() if len(row) >= 1 and row[0] is not None else ""
        nome = nome_lado or nome_col_a
        empresas.append((numero_txt, nome))

    return empresas


def processar_empresa(numero_empresa: str, nome_empresa: str, dir_dest: str, ano_atual: int, dlg) -> None:
    base_arquivo = montar_base_arquivo(numero_empresa, nome_empresa)
    nome_arquivo = f"{base_arquivo}.txt"
    nome_arquivo_slk = f"{base_arquivo}.slk"

    clicar(*COORD_EMPRESA_SISTEMA)
    mandar_backspaces(BACKSPACES_EMPRESA)
    escrever(numero_empresa)
    enter_com_delay(2)
    clicar(*COORD_FOCO)

    alt_mgm()

    time.sleep(DELAY_POS_ALT_MGM)

    escrever(str(ano_atual))
    enter_com_delay()

    escrever(str(ano_atual - 1))
    enter_com_delay()

    enter_com_delay(3)

    escrever(dir_dest)
    enter_com_delay()

    escrever(nome_arquivo)
    enter_com_delay()

    garantir_checkbox(dlg)

    pyautogui.hotkey("alt", "g")
    delay_pos_alt_g()

    _esperar_e_confirmar_aviso(TIMEOUT_AVISO_1)
    delay_pos_alt_g()
    dlg_aviso_2 = _achar_aviso_sistema(TIMEOUT_AVISO_2)
    if dlg_aviso_2 is not None:
        texto_aviso_2 = _normalizar_texto(_obter_texto_aviso(dlg_aviso_2))
        if "RELATORIO EM BRANCO" in texto_aviso_2:
            registrar_na_lista_negra(numero_empresa, "nao tem relatorio")
            _confirmar_aviso(dlg_aviso_2)
            pyautogui.press("enter")
            print(f"Relatorio em branco para empresa {numero_empresa}. Indo para a proxima.")
            return
        _confirmar_aviso(dlg_aviso_2)
    delay_pos_alt_g()
    _fluxo_salvar_por_coordenadas(dir_dest, nome_arquivo_slk)
    time.sleep(DELAY_APOS_VOLTAR)


def main() -> None:
    import argparse

    p = argparse.ArgumentParser()
    p.add_argument("--empresa", "-e", default="")
    args = p.parse_args()

    ano_atual = datetime.now().year

    dir_dest = DIR_BASE_TEMPLATE.format(ano=ano_atual)
    caminho_lista = LISTA_TEMPLATE.format(ano=ano_atual)

    dlg = focar_janela_fiscal()
    time.sleep(0.3)

    empresas: list[tuple[str, str]] = []
    if args.empresa:
        numero_manual = str(args.empresa).strip()
        empresas = [(numero_manual, "")]
    else:
        empresas = carregar_empresas_lista(caminho_lista)
    lista_negra = carregar_lista_negra()

    for numero_empresa, nome_empresa in empresas:
        if not numero_empresa:
            continue
        if numero_empresa in lista_negra:
            print(f"Empresa na lista negra. Pulando: {numero_empresa}")
            continue
        base_arquivo = montar_base_arquivo(numero_empresa, nome_empresa)
        caminho_txt = os.path.join(dir_dest, f"{base_arquivo}.txt")
        caminho_slk = os.path.join(dir_dest, f"{base_arquivo}.slk")
        if os.path.exists(caminho_slk):
            print(f"Ja existe SLK para {base_arquivo}. Pulando.")
            continue
        if os.path.exists(caminho_txt):
            print(f"TXT ja existe para {base_arquivo}. Gerando apenas o SLK.")
        processar_empresa(numero_empresa, nome_empresa, dir_dest, ano_atual, dlg)


if __name__ == "__main__":
    main()
