import os
import sys
import time
import base64
import gzip
import shutil
import sqlite3
from datetime import datetime

import urllib3
from requests_pkcs12 import post
from requests.exceptions import HTTPError, RequestException
from lxml import etree
from openpyxl import load_workbook

# ==========================
# Configurações gerais
# ==========================

# URL do serviço de Distribuição DF-e - Ambiente Nacional (produção)
WS_URL = "https://www1.nfe.fazenda.gov.br/NFeDistribuicaoDFe/NFeDistribuicaoDFe.asmx"

# Ambiente: 1 = Produção, 2 = Homologação
TP_AMB = "1"

# Código da UF autora -> PR = 41
CODIGO_UF = "41"

# Intervalo entre requisições por empresa (para evitar bloqueio)
INTERVALO_SEGUNDOS = 2

# Máximo de iterações por empresa (seguro para não cair em loop infinito)
MAX_ITERACOES_POR_EMPRESA = 200

# Desativar warning de certificado self-signed
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ==========================
# Utilitários gerais
# ==========================

def get_script_dir() -> str:
    """Retorna a pasta onde o script está sendo executado."""
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def normalizar_cnpj(cnpj: str) -> str:
    """Remove caracteres não numéricos do CNPJ."""
    return "".join(filter(str.isdigit, cnpj or ""))


def normalizar_nsu(nsu: str) -> str:
    """Formata NSU com 15 dígitos, preenchendo com zeros à esquerda."""
    nsu = "".join(filter(str.isdigit, str(nsu or "")))
    if not nsu:
        nsu = "0"
    nsu_int = int(nsu)
    return f"{nsu_int:015d}"


def slug_nome_empresa(nome: str) -> str:
    """
    Nome "seguro" para pasta da empresa.
    Mantém espaços, só tira caracteres estranhos.
    """
    if not nome:
        return "empresa sem nome"
    permitido = " -_"
    slug = "".join(c for c in nome if c.isalnum() or c in permitido).strip()
    return slug or "empresa sem nome"


def criar_pasta_empresa(base_dir: str, nome_empresa: str) -> str:
    """Cria a pasta NFE/<empresa> e retorna o caminho."""
    pasta_base_nfe = os.path.join(base_dir, "NFE")
    os.makedirs(pasta_base_nfe, exist_ok=True)

    pasta_empresa = os.path.join(pasta_base_nfe, slug_nome_empresa(nome_empresa))
    os.makedirs(pasta_empresa, exist_ok=True)

    return pasta_empresa


# ==========================
# Banco de dados (SQLite)
# ==========================

def init_db(conn: sqlite3.Connection):
    """Cria a tabela empresas se não existir."""
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS empresas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            cnpj TEXT NOT NULL UNIQUE,
            caminho_certificado TEXT NOT NULL,
            senha_certificado TEXT,
            nsu_atual INTEGER NOT NULL DEFAULT 0,
            ativo INTEGER NOT NULL DEFAULT 1,
            criado_em TEXT,
            atualizado_em TEXT
        );
        """
    )
    conn.commit()


def backup_db_if_exists(db_path: str):
    """Cria um backup do banco antes de alterar, se existir."""
    if not os.path.isfile(db_path):
        return

    base_dir = os.path.dirname(db_path)
    backup_dir = os.path.join(base_dir, "BACKUP_DB")
    os.makedirs(backup_dir, exist_ok=True)

    data_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"certificados_{data_str}.db"
    backup_path = os.path.join(backup_dir, backup_name)
    shutil.copy2(db_path, backup_path)
    print(f"Backup do banco criado em: {backup_path}")


def importar_xlsx_para_db(conn: sqlite3.Connection, base_dir: str, log_geral: list):
    """
    Importa dados de certificados.xlsx para o banco, somente se:
    - existir o XLSX
    - e a tabela empresas estiver vazia
    """
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM empresas;")
    (qtd,) = cur.fetchone()
    if qtd > 0:
        return  # já tem dados, não importa

    caminho_xlsx = os.path.join(base_dir, "certificados.xlsx")
    if not os.path.isfile(caminho_xlsx):
        return

    print("Importando dados de certificados.xlsx para certificados.db ...")
    log_geral.append("Importando dados de certificados.xlsx para o banco SQLite.")

    wb = load_workbook(caminho_xlsx)
    ws = wb.active
    header = [cell.value for cell in ws[1]]

    def get_idx(col_name):
        try:
            return header.index(col_name)
        except ValueError:
            return None

    idx_nome = get_idx("NOME")
    idx_cnpj = get_idx("CNPJ")
    idx_caminho = get_idx("CAMINHO CERTIFICADO")
    idx_senha = get_idx("SENHA CERTIFICADO")
    idx_nsu = get_idx("NSU")

    inseridos = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        nome = str(row[idx_nome] or "").strip() if idx_nome is not None else ""
        cnpj = str(row[idx_cnpj] or "").strip() if idx_cnpj is not None else ""
        caminho = str(row[idx_caminho] or "").strip() if idx_caminho is not None else ""
        senha = str(row[idx_senha] or "").strip() if idx_senha is not None else ""
        nsu = str(row[idx_nsu] or "").strip() if idx_nsu is not None else "0"

        if not cnpj:
            continue

        try:
            nsu_int = int("".join(filter(str.isdigit, nsu))) if nsu else 0
        except ValueError:
            nsu_int = 0

        agora = datetime.now().isoformat(timespec="seconds")
        cur.execute(
            """
            INSERT OR IGNORE INTO empresas (nome, cnpj, caminho_certificado, senha_certificado, nsu_atual, ativo, criado_em, atualizado_em)
            VALUES (?, ?, ?, ?, ?, 1, ?, ?);
            """,
            (nome, normalizar_cnpj(cnpj), caminho, senha, nsu_int, agora, agora),
        )
        inseridos += 1

    conn.commit()
    log_geral.append(f"{inseridos} empresa(s) importada(s) do XLSX para o banco.")


def carregar_empresas_db(conn: sqlite3.Connection) -> list:
    """Retorna lista de empresas ativas como dicionários."""
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, nome, cnpj, caminho_certificado, senha_certificado, nsu_atual, ativo
        FROM empresas
        WHERE ativo = 1
        ORDER BY id;
        """
    )
    rows = cur.fetchall()
    empresas = []
    for r in rows:
        empresas.append(
            {
                "id": r[0],
                "nome": r[1] or "",
                "cnpj": r[2] or "",
                "caminho_certificado": r[3] or "",
                "senha_certificado": r[4] or "",
                "nsu_atual": r[5] or 0,
                "ativo": r[6],
            }
        )
    return empresas


def atualizar_nsu_db(conn: sqlite3.Connection, empresa_id: int, nsu_str: str, log_geral: list):
    """
    Atualiza nsu_atual da empresa no banco, garantindo que nunca diminui.
    """
    cur = conn.cursor()
    cur.execute("SELECT nsu_atual FROM empresas WHERE id = ?;", (empresa_id,))
    row = cur.fetchone()
    if row is None:
        return

    nsu_atual = row[0] or 0

    try:
        nsu_novo = int("".join(filter(str.isdigit, str(nsu_str)))) if nsu_str else 0
    except ValueError:
        nsu_novo = nsu_atual  # ignorar valor estranho

    # Atualiza sempre que o NSU for diferente do que está gravado
    if nsu_novo == nsu_atual:
        # nada mudou, não precisa atualizar
        return

    agora = datetime.now().isoformat(timespec="seconds")
    cur.execute(
        "UPDATE empresas SET nsu_atual = ?, atualizado_em = ? WHERE id = ?;",
        (nsu_novo, agora, empresa_id),
    )
    conn.commit()
    log_geral.append(f"  NSU atualizado no banco: {nsu_atual} -> {nsu_novo}")


# ==========================
# WebService NFeDistribuicaoDFe
# ==========================

def montar_envelope_distnsu(cnpj: str, ult_nsu: str) -> str:
    """Monta o envelope SOAP distNSU."""
    cnpj = normalizar_cnpj(cnpj)
    ult_nsu = normalizar_nsu(ult_nsu)

    envelope = f"""<?xml version="1.0" encoding="utf-8"?>
<soap12:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
                 xmlns:xsd="http://www.w3.org/2001/XMLSchema"
                 xmlns:soap12="http://www.w3.org/2003/05/soap-envelope">
  <soap12:Body>
    <nfeDistDFeInteresse xmlns="http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe">
      <nfeDadosMsg>
        <distDFeInt xmlns="http://www.portalfiscal.inf.br/nfe" versao="1.01">
          <tpAmb>{TP_AMB}</tpAmb>
          <cUFAutor>{CODIGO_UF}</cUFAutor>
          <CNPJ>{cnpj}</CNPJ>
          <distNSU>
            <ultNSU>{ult_nsu}</ultNSU>
          </distNSU>
        </distDFeInt>
      </nfeDadosMsg>
    </nfeDistDFeInteresse>
  </soap12:Body>
</soap12:Envelope>"""
    return envelope


def enviar_requisicao(caminho_cert: str, senha_cert: str, envelope: str) -> str:
    """Chama o WebService com certificado PKCS12 e retorna resposta."""
    headers = {
        "Content-Type": "application/soap+xml; charset=utf-8",
        "SOAPAction": "http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe/nfeDistDFeInteresse",
    }

    try:
        resp = post(
            WS_URL,
            data=envelope.encode("utf-8"),
            headers=headers,
            pkcs12_filename=caminho_cert,
            pkcs12_password=senha_cert,
            timeout=60,
            verify=False,
        )
        resp.raise_for_status()
        return resp.text
    except HTTPError as e:
        raise RuntimeError(f"Erro HTTP ao chamar o serviço: {e}") from e
    except RequestException as e:
        raise RuntimeError(f"Erro de requisição ao chamar o serviço: {e}") from e


def decodificar_doczip(doczip_element) -> str:
    """Decodifica docZip (Base64 + GZip) e retorna XML interno."""
    conteudo_b64 = doczip_element.text
    if not conteudo_b64:
        raise ValueError("Elemento docZip sem conteúdo.")

    binario = base64.b64decode(conteudo_b64)
    xml_bytes = gzip.decompress(binario)
    return xml_bytes.decode("utf-8")


def extrair_chave_de_xml(xml_str: str) -> str:
    """Extrai chNFe de diferentes tipos de XML (resNFe, procNFe, eventos)."""
    try:
        root = etree.fromstring(xml_str.encode("utf-8"))
    except Exception:
        return ""

    ns = {"nfe": "http://www.portalfiscal.inf.br/nfe"}

    ch = root.find(".//nfe:chNFe", namespaces=ns)
    if ch is not None and ch.text:
        return ch.text.strip()

    ch = root.find(".//nfe:infProt/nfe:chNFe", namespaces=ns)
    if ch is not None and ch.text:
        return ch.text.strip()

    ch = root.find(".//nfe:procEventoNFe//nfe:chNFe", namespaces=ns)
    if ch is not None and ch.text:
        return ch.text.strip()

    ch = root.find(".//nfe:infEvento/nfe:chNFe", namespaces=ns)
    if ch is not None and ch.text:
        return ch.text.strip()

    return ""


def salvar_xml(pasta_empresa: str, xml_str: str, chave: str, indice: int) -> str:
    """
    Salva o XML em arquivo na pasta da empresa.
    Nome: <chave>.xml   (se não houver chave: SEMCHAVE_<indice>.xml)
    """
    if not chave:
        chave = f"SEMCHAVE_{indice:05d}"

    nome_arquivo = f"{chave}.xml"
    caminho_arquivo = os.path.join(pasta_empresa, nome_arquivo)

    with open(caminho_arquivo, "w", encoding="utf-8") as f:
        f.write(xml_str)

    return caminho_arquivo


def processar_resposta_distnsu(xml_resposta: str, pasta_empresa: str, log_msgs: list) -> dict:
    """
    Processa resposta distNSU.
    Salva APENAS XMLs procNFe.
    """
    resultado = {
        "cStat": "",
        "xMotivo": "",
        "ultNSU": None,
        "maxNSU": None,
        "qtd_documentos": 0,
    }

    ns = {
        "soap": "http://www.w3.org/2003/05/soap-envelope",
        "ws": "http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe",
        "nfe": "http://www.portalfiscal.inf.br/nfe",
    }

    try:
        root = etree.fromstring(xml_resposta.encode("utf-8"))
    except Exception as e:
        msg = f"Falha ao parsear XML de resposta SOAP: {e}"
        log_msgs.append(msg)
        raise RuntimeError(msg) from e

    ret = root.find(".//nfe:retDistDFeInt", namespaces=ns)
    if ret is None:
        msg = "Elemento retDistDFeInt não encontrado na resposta."
        log_msgs.append(msg)
        raise RuntimeError(msg)

    cStat = ret.findtext("nfe:cStat", namespaces=ns) or ""
    xMotivo = ret.findtext("nfe:xMotivo", namespaces=ns) or ""
    ultNSU = ret.findtext("nfe:ultNSU", namespaces=ns)
    maxNSU = ret.findtext("nfe:maxNSU", namespaces=ns)

    resultado["cStat"] = cStat
    resultado["xMotivo"] = xMotivo
    resultado["ultNSU"] = ultNSU
    resultado["maxNSU"] = maxNSU

    log_msgs.append(f"  cStat={cStat} | xMotivo={xMotivo} | ultNSU={ultNSU} | maxNSU={maxNSU}")

    if cStat == "137":
        return resultado

    if cStat != "138":
        log_msgs.append(f"  cStat diferente de 137/138: {cStat} - {xMotivo}")
        return resultado

    doczips = ret.findall(".//nfe:docZip", namespaces=ns)
    indice = 1
    for dz in doczips:
        schema = dz.get("schema", "desconhecido")

        # 🔥 somente procNFe
        if "procNFe" not in schema:
            continue

        try:
            xml_interno = decodificar_doczip(dz)
            chave = extrair_chave_de_xml(xml_interno)
            caminho_xml = salvar_xml(pasta_empresa, xml_interno, chave, indice)
            log_msgs.append(f"  XML salvo (procNFe): {caminho_xml} | chNFe={chave}")
            indice += 1
            resultado["qtd_documentos"] += 1
        except Exception as e:
            log_msgs.append(f"  Erro ao processar docZip procNFe (schema={schema}): {e}")

    return resultado


def escrever_log_geral(base_dir: str, log_msgs: list):
    """Escreve log geral."""
    if not log_msgs:
        return

    data_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_log = f"LOG_NFE_{data_str}.txt"
    caminho_log = os.path.join(base_dir, nome_log)

    with open(caminho_log, "w", encoding="utf-8") as f:
        f.write("Log geral de processamento NFeDistribuicaoDFe\n")
        f.write(f"Data/Hora: {datetime.now().isoformat()}\n")
        f.write("=" * 80 + "\n\n")
        for linha in log_msgs:
            f.write(linha + "\n")

    print(f"Log geral gravado em: {caminho_log}")


# ==========================
# Fluxo por empresa
# ==========================

def processar_empresa(base_dir: str, empresa: dict, log_geral: list, conn: sqlite3.Connection):
    """
    Processa uma empresa:
    - Cria pasta NFE/<empresa>
    - Faz consultas distNSU
    - Atualiza nsu_atual no banco
    """
    empresa_id = empresa["id"]
    nome_empresa = (empresa["nome"] or "").strip() or "(sem nome)"
    cnpj = (empresa["cnpj"] or "").strip()
    caminho_cert = (empresa["caminho_certificado"] or "").strip()
    senha_cert = (empresa["senha_certificado"] or "").strip()
    nsu_inicial = str(empresa.get("nsu_atual", 0))

    log_geral.append("=" * 80)
    log_geral.append(f"Empresa: {nome_empresa}")
    log_geral.append(f"CNPJ: {cnpj}")
    log_geral.append(f"Certificado: {caminho_cert}")
    log_geral.append(f"NSU inicial: {nsu_inicial}")

    if not cnpj:
        msg = "[IGNORANDO] CNPJ vazio."
        print(f"[{nome_empresa}] {msg}")
        log_geral.append("  " + msg)
        return

    pasta_empresa = criar_pasta_empresa(base_dir, nome_empresa)

    if not caminho_cert:
        msg = "Caminho do certificado não informado."
        print(f"[{nome_empresa}] {msg}")
        log_geral.append("  " + msg)
        return

    if not os.path.isfile(caminho_cert):
        msg = f"Certificado não encontrado: {caminho_cert}"
        print(f"[{nome_empresa}] {msg}")
        log_geral.append("  " + msg)
        return

    ult_nsu = normalizar_nsu(nsu_inicial)
    iteracoes = 0
    total_documentos = 0

    try:
        while iteracoes < MAX_ITERACOES_POR_EMPRESA:
            iteracoes += 1
            log_geral.append(f"  --- Iteração {iteracoes} | ultNSU atual: {ult_nsu} ---")
            print(f"[{nome_empresa}] Iteração {iteracoes} | ultNSU={ult_nsu}")

            envelope = montar_envelope_distnsu(cnpj, ult_nsu)
            try:
                resposta_soap = enviar_requisicao(caminho_cert, senha_cert, envelope)
            except RuntimeError as e:
                msg = f"Erro ao enviar requisição: {e}"
                print(f"[{nome_empresa}] {msg}")
                log_geral.append("  " + msg)
                break

            resultado = processar_resposta_distnsu(resposta_soap, pasta_empresa, log_geral)

            cStat = resultado["cStat"]
            xMotivo = resultado["xMotivo"]
            ultNSU_ret = resultado["ultNSU"]
            maxNSU_ret = resultado["maxNSU"]
            qtd_docs = resultado["qtd_documentos"]
            total_documentos += qtd_docs

            if cStat == "137":
                msg = "Nenhum documento localizado (cStat=137). Encerrando."
                print(f"[{nome_empresa}] {msg}")
                log_geral.append("  " + msg)
                break

            if cStat not in ("137", "138"):
                msg = f"cStat={cStat} - {xMotivo}. Encerrando empresa."
                print(f"[{nome_empresa}] {msg}")
                log_geral.append("  " + msg)
                break

            if ultNSU_ret:
                ult_nsu = normalizar_nsu(ultNSU_ret)

            if ultNSU_ret and maxNSU_ret and ultNSU_ret == maxNSU_ret:
                msg = f"ultNSU atingiu maxNSU ({ultNSU_ret}). Nada mais a buscar."
                print(f"[{nome_empresa}] {msg}")
                log_geral.append("  " + msg)
                break

            if qtd_docs == 0:
                msg = "Nenhum procNFe retornado nessa iteração. Encerrando."
                print(f"[{nome_empresa}] {msg}")
                log_geral.append("  " + msg)
                break

            time.sleep(INTERVALO_SEGUNDOS)

    except Exception as e:
        msg = f"Erro inesperado: {e}"
        print(f"[{nome_empresa}] {msg}")
        log_geral.append("  " + msg)

    log_geral.append(f"  Processamento concluído para '{nome_empresa}'.")
    log_geral.append(f"  Total de procNFe recebidos: {total_documentos}")
    log_geral.append(f"  NSU final utilizado (texto): {ult_nsu}")

    atualizar_nsu_db(conn, empresa_id, ult_nsu, log_geral)


# ==========================
# main
# ==========================

def main():
    base_dir = get_script_dir()
    print(f"Pasta do script: {base_dir}")

    db_path = os.path.join(base_dir, "certificados.db")

    # Backup do banco, se já existir
    backup_db_if_exists(db_path)

    conn = sqlite3.connect(db_path)
    init_db(conn)

    log_geral = []

    # Importa dados do XLSX se banco estiver vazio (primeira vez)
    importar_xlsx_para_db(conn, base_dir, log_geral)

    empresas = carregar_empresas_db(conn)
    if not empresas:
        print("Nenhuma empresa ativa encontrada no banco.")
        log_geral.append("Nenhuma empresa ativa encontrada no banco.")
        escrever_log_geral(base_dir, log_geral)
        conn.close()
        sys.exit(0)

    print(f"{len(empresas)} empresa(s) ativa(s) encontrada(s) no banco.")

    for empresa in empresas:
        processar_empresa(base_dir, empresa, log_geral, conn)
        print("-" * 80)

    conn.close()
    escrever_log_geral(base_dir, log_geral)


if __name__ == "__main__":
    main()
