import os
import sys
import time
import base64
import gzip
from datetime import datetime

import urllib3
from requests_pkcs12 import post
from requests.exceptions import HTTPError, RequestException
from lxml import etree
from openpyxl import load_workbook

# ==========================
# Configurações gerais
# ==========================

# URL do serviço de Distribuição DF-e - Ambiente Nacional (produção)
WS_URL = "https://www1.nfe.fazenda.gov.br/NFeDistribuicaoDFe/NFeDistribuicaoDFe.asmx"

# Ambiente: 1 = Produção, 2 = Homologação
TP_AMB = "1"

# Código da UF autora -> PR = 41
CODIGO_UF = "41"

# Intervalo entre requisições por empresa (para evitar bloqueio)
INTERVALO_SEGUNDOS = 2

# Máximo de iterações por empresa (seguro para não cair em loop infinito)
MAX_ITERACOES_POR_EMPRESA = 200

# Desativar warning de certificado self-signed
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ==========================
# Utilitários
# ==========================

def get_script_dir() -> str:
    """Retorna a pasta onde o script está sendo executado."""
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        # Executável (pyinstaller, etc.)
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def normalizar_cnpj(cnpj: str) -> str:
    """Remove caracteres não numéricos do CNPJ."""
    return "".join(filter(str.isdigit, cnpj or ""))


def normalizar_nsu(nsu: str) -> str:
    """
    Ajusta o NSU para o formato numérico de 15 dígitos, preenchendo com zeros à esquerda.
    Se vier vazio ou inválido, retorna '000000000000000'.
    """
    nsu = "".join(filter(str.isdigit, str(nsu or "")))
    if not nsu:
        nsu = "0"
    nsu_int = int(nsu)
    return f"{nsu_int:015d}"


def slug_nome_empresa(nome: str) -> str:
    """
    Nome "seguro" para pasta da empresa.
    Agora mantemos espaços (sem trocar por _), apenas removendo caracteres estranhos.
    """
    if not nome:
        return "empresa sem nome"
    permitido = " -_"
    slug = "".join(c for c in nome if c.isalnum() or c in permitido).strip()
    return slug or "empresa sem nome"


def criar_pasta_empresa(base_dir: str, nome_empresa: str) -> str:
    """Cria a pasta NFE/<empresa> e retorna o caminho."""
    pasta_base_nfe = os.path.join(base_dir, "NFE")
    os.makedirs(pasta_base_nfe, exist_ok=True)

    pasta_empresa = os.path.join(pasta_base_nfe, slug_nome_empresa(nome_empresa))
    os.makedirs(pasta_empresa, exist_ok=True)

    return pasta_empresa


def montar_envelope_distnsu(cnpj: str, ult_nsu: str) -> str:
    """
    Monta o envelope SOAP para o serviço NFeDistribuicaoDFe, modo distNSU,
    no mesmo padrão do serviço por chave (ambiente nacional).
    """
    cnpj = normalizar_cnpj(cnpj)
    ult_nsu = normalizar_nsu(ult_nsu)

    envelope = f"""<?xml version="1.0" encoding="utf-8"?>
<soap12:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
                 xmlns:xsd="http://www.w3.org/2001/XMLSchema"
                 xmlns:soap12="http://www.w3.org/2003/05/soap-envelope">
  <soap12:Body>
    <nfeDistDFeInteresse xmlns="http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe">
      <nfeDadosMsg>
        <distDFeInt xmlns="http://www.portalfiscal.inf.br/nfe" versao="1.01">
          <tpAmb>{TP_AMB}</tpAmb>
          <cUFAutor>{CODIGO_UF}</cUFAutor>
          <CNPJ>{cnpj}</CNPJ>
          <distNSU>
            <ultNSU>{ult_nsu}</ultNSU>
          </distNSU>
        </distDFeInt>
      </nfeDadosMsg>
    </nfeDistDFeInteresse>
  </soap12:Body>
</soap12:Envelope>"""
    return envelope


def enviar_requisicao(caminho_cert: str, senha_cert: str, envelope: str) -> str:
    """
    Envia a requisição SOAP ao serviço de distribuição DF-e usando certificado PKCS12.
    Retorna o XML da resposta como string.
    """
    headers = {
        "Content-Type": "application/soap+xml; charset=utf-8",
        "SOAPAction": "http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe/nfeDistDFeInteresse",
    }

    try:
        resp = post(
            WS_URL,
            data=envelope.encode("utf-8"),
            headers=headers,
            pkcs12_filename=caminho_cert,
            pkcs12_password=senha_cert,
            timeout=60,
            verify=False,  # dependendo do ambiente, pode ser True
        )
        resp.raise_for_status()
        return resp.text
    except HTTPError as e:
        raise RuntimeError(f"Erro HTTP ao chamar o serviço: {e}") from e
    except RequestException as e:
        raise RuntimeError(f"Erro de requisição ao chamar o serviço: {e}") from e


def decodificar_doczip(doczip_element) -> str:
    """
    Recebe um elemento <docZip>, decodifica Base64 + GZip e retorna o XML interno como string.
    """
    conteudo_b64 = doczip_element.text
    if not conteudo_b64:
        raise ValueError("Elemento docZip sem conteúdo.")

    binario = base64.b64decode(conteudo_b64)
    xml_bytes = gzip.decompress(binario)
    return xml_bytes.decode("utf-8")


def extrair_chave_de_xml(xml_str: str) -> str:
    """
    Tenta extrair a chave da NF-e (chNFe) de diferentes tipos de XML.
    Retorna "" se não conseguir.
    """
    try:
        root = etree.fromstring(xml_str.encode("utf-8"))
    except Exception:
        return ""

    ns = {"nfe": "http://www.portalfiscal.inf.br/nfe"}

    # 1) resNFe
    ch = root.find(".//nfe:chNFe", namespaces=ns)
    if ch is not None and ch.text:
        return ch.text.strip()

    # 2) nfeProc (procNFe) - pega da infProt
    ch = root.find(".//nfe:infProt/nfe:chNFe", namespaces=ns)
    if ch is not None and ch.text:
        return ch.text.strip()

    # 3) procEventoNFe
    ch = root.find(".//nfe:procEventoNFe//nfe:chNFe", namespaces=ns)
    if ch is not None and ch.text:
        return ch.text.strip()

    # 4) infEvento/chNFe
    ch = root.find(".//nfe:infEvento/nfe:chNFe", namespaces=ns)
    if ch is not None and ch.text:
        return ch.text.strip()

    return ""


def salvar_xml(pasta_empresa: str, xml_str: str, chave: str, indice: int) -> str:
    """
    Salva o XML em arquivo na pasta da empresa.
    AGORA: nome do arquivo SOMENTE com a chave (ex.: 3519...xml).
    Se não tiver chave, usa SEMCHAVE_<indice>.xml
    """
    if not chave:
        chave = f"SEMCHAVE_{indice:05d}"

    nome_arquivo = f"{chave}.xml"
    caminho_arquivo = os.path.join(pasta_empresa, nome_arquivo)

    with open(caminho_arquivo, "w", encoding="utf-8") as f:
        f.write(xml_str)

    return caminho_arquivo


def processar_resposta_distnsu(xml_resposta: str, pasta_empresa: str, log_msgs: list) -> dict:
    """
    Processa a resposta SOAP de uma chamada distNSU.
    Salva APENAS os XMLs procNFe na pasta da empresa.
    Retorna dict com:
        {
            "cStat": str,
            "xMotivo": str,
            "ultNSU": str ou None,
            "maxNSU": str ou None,
            "qtd_documentos": int   # apenas procNFe contados
        }
    """
    resultado = {
        "cStat": "",
        "xMotivo": "",
        "ultNSU": None,
        "maxNSU": None,
        "qtd_documentos": 0,
    }

    ns = {
        "soap": "http://www.w3.org/2003/05/soap-envelope",
        "ws": "http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe",
        "nfe": "http://www.portalfiscal.inf.br/nfe",
    }

    try:
        root = etree.fromstring(xml_resposta.encode("utf-8"))
    except Exception as e:
        msg = f"Falha ao parsear XML de resposta SOAP: {e}"
        log_msgs.append(msg)
        raise RuntimeError(msg) from e

    ret = root.find(".//nfe:retDistDFeInt", namespaces=ns)
    if ret is None:
        msg = "Elemento retDistDFeInt não encontrado na resposta."
        log_msgs.append(msg)
        raise RuntimeError(msg)

    cStat = ret.findtext("nfe:cStat", namespaces=ns) or ""
    xMotivo = ret.findtext("nfe:xMotivo", namespaces=ns) or ""
    ultNSU = ret.findtext("nfe:ultNSU", namespaces=ns)
    maxNSU = ret.findtext("nfe:maxNSU", namespaces=ns)

    resultado["cStat"] = cStat
    resultado["xMotivo"] = xMotivo
    resultado["ultNSU"] = ultNSU
    resultado["maxNSU"] = maxNSU

    log_msgs.append(f"  cStat={cStat} | xMotivo={xMotivo} | ultNSU={ultNSU} | maxNSU={maxNSU}")

    # 137 = nenhum documento localizado
    if cStat == "137":
        return resultado

    if cStat != "138":
        # Outros códigos
        msg = f"  cStat diferente de 137/138: {cStat} - {xMotivo}"
        log_msgs.append(msg)
        return resultado

    # Processar somente docZip do tipo procNFe
    doczips = ret.findall(".//nfe:docZip", namespaces=ns)
    indice = 1
    for dz in doczips:
        schema = dz.get("schema", "desconhecido")

        # 🔥 FILTRO: só baixa procNFe
        # Ex.: "procNFe_v4.00.xsd"
        if "procNFe" not in schema:
            continue

        try:
            xml_interno = decodificar_doczip(dz)
            chave = extrair_chave_de_xml(xml_interno)
            caminho_xml = salvar_xml(pasta_empresa, xml_interno, chave, indice)
            log_msgs.append(f"  XML salvo (procNFe): {caminho_xml} | chNFe={chave}")
            indice += 1
            resultado["qtd_documentos"] += 1
        except Exception as e:
            log_msgs.append(f"  Erro ao processar docZip procNFe (schema={schema}): {e}")

    return resultado


def escrever_log_geral(base_dir: str, log_msgs: list):
    """
    Escreve um único arquivo de log com tudo (todas as empresas).
    Arquivo: LOG_NFE_<data>.txt na pasta do script.
    """
    if not log_msgs:
        return

    data_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_log = f"LOG_NFE_{data_str}.txt"
    caminho_log = os.path.join(base_dir, nome_log)

    with open(caminho_log, "w", encoding="utf-8") as f:
        f.write("Log geral de processamento NFeDistribuicaoDFe\n")
        f.write(f"Data/Hora: {datetime.now().isoformat()}\n")
        f.write("=" * 80 + "\n\n")
        for linha in log_msgs:
            f.write(linha + "\n")

    print(f"Log geral gravado em: {caminho_log}")


# ==========================
# XLSX: leitura e escrita
# ==========================

def carregar_empresas_xlsx(caminho_xlsx: str) -> list:
    """
    Lê o arquivo certificados.xlsx e retorna uma lista de dicionários.
    Colunas usados:
    NOME | CNPJ | CAMINHO CERTIFICADO | SENHA CERTIFICADO | NSU
    """
    wb = load_workbook(caminho_xlsx)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    empresas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        empresa = dict(zip(header, row))
        empresas.append(empresa)

    return empresas


def salvar_empresas_xlsx(caminho_xlsx: str, empresas: list):
    """
    Atualiza o XLSX sobrescrevendo os valores existentes,
    mantendo a mesma estrutura e cabeçalho.
    """
    wb = load_workbook(caminho_xlsx)
    ws = wb.active

    header = [cell.value for cell in ws[1]]

    # Apaga todas as linhas exceto cabeçalho
    ws.delete_rows(2, ws.max_row)

    # Regrava as linhas
    for empresa in empresas:
        row = [empresa.get(col) for col in header]
        ws.append(row)

    wb.save(caminho_xlsx)
    print(f"Planilha atualizada: {caminho_xlsx}")


# ==========================
# Fluxo principal por empresa
# ==========================

def processar_empresa(base_dir: str, empresa_row: dict, log_geral: list):
    """
    Processa uma empresa:
    - Cria pasta NFE/<empresa>
    - Faz consultas distNSU sucessivas
    - Atualiza NSU da empresa_row
    - Escreve logs no log_geral
    """

    def get_val(row, *keys):
        """Tenta pegar o valor de uma das chaves informadas."""
        for k in keys:
            if k in row and row[k] is not None:
                return row[k]
        return ""

    # Aceita tanto os nomes originais quanto os da sua planilha
    nome_empresa = str(get_val(empresa_row, "Nome da empresa", "NOME", "Nome") or "").strip()
    cnpj = str(get_val(empresa_row, "CNPJ", "cnpj") or "").strip()
    caminho_cert = str(get_val(empresa_row, "Caminho para o certificado", "CAMINHO CERTIFICADO", "CAMINHO CERT") or "").strip()
    senha_cert = str(get_val(empresa_row, "senha certificado", "SENHA CERTIFICADO", "Senha Certificado") or "").strip()
    nsu_inicial = str(get_val(empresa_row, "NSU", "nsu") or "").strip()

    if not nome_empresa:
        nome_empresa = "(sem nome)"

    log_geral.append("=" * 80)
    log_geral.append(f"Empresa: {nome_empresa}")
    log_geral.append(f"CNPJ: {cnpj}")
    log_geral.append(f"Certificado: {caminho_cert}")
    log_geral.append(f"NSU inicial: {nsu_inicial or '0'}")

    # Ignorar linhas onde o CNPJ está vazio
    if not cnpj:
        msg = "[IGNORANDO] Linha ignorada porque o CNPJ está vazio."
        print(f"[{nome_empresa}] {msg}")
        log_geral.append("  " + msg)
        return

    pasta_empresa = criar_pasta_empresa(base_dir, nome_empresa)

    if not caminho_cert:
        msg = "Caminho do certificado não informado na planilha."
        print(f"[{nome_empresa}] {msg}")
        log_geral.append("  " + msg)
        return

    if not os.path.isfile(caminho_cert):
        msg = f"Certificado não encontrado: {caminho_cert}"
        print(f"[{nome_empresa}] {msg}")
        log_geral.append("  " + msg)
        return

    ult_nsu = normalizar_nsu(nsu_inicial)
    iteracoes = 0
    total_documentos = 0

    try:
        while iteracoes < MAX_ITERACOES_POR_EMPRESA:
            iteracoes += 1
            log_geral.append(f"  --- Iteração {iteracoes} | ultNSU atual: {ult_nsu} ---")
            print(f"[{nome_empresa}] Iteração {iteracoes} | ultNSU={ult_nsu}")

            envelope = montar_envelope_distnsu(cnpj, ult_nsu)
            try:
                resposta_soap = enviar_requisicao(caminho_cert, senha_cert, envelope)
            except RuntimeError as e:
                msg = f"Erro ao enviar requisição: {e}"
                print(f"[{nome_empresa}] {msg}")
                log_geral.append("  " + msg)
                break

            resultado = processar_resposta_distnsu(resposta_soap, pasta_empresa, log_geral)

            cStat = resultado["cStat"]
            xMotivo = resultado["xMotivo"]
            ultNSU_ret = resultado["ultNSU"]
            maxNSU_ret = resultado["maxNSU"]
            qtd_docs = resultado["qtd_documentos"]
            total_documentos += qtd_docs

            if cStat == "137":
                msg = "Nenhum documento localizado (cStat=137). Encerrando."
                print(f"[{nome_empresa}] {msg}")
                log_geral.append("  " + msg)
                break

            if cStat not in ("137", "138"):
                msg = f"cStat={cStat} - {xMotivo}. Encerrando empresa."
                print(f"[{nome_empresa}] {msg}")
                log_geral.append("  " + msg)
                break

            if ultNSU_ret:
                ult_nsu = normalizar_nsu(ultNSU_ret)

            if ultNSU_ret and maxNSU_ret and ultNSU_ret == maxNSU_ret:
                msg = f"ultNSU atingiu maxNSU ({ultNSU_ret}). Nada mais a buscar."
                print(f"[{nome_empresa}] {msg}")
                log_geral.append("  " + msg)
                break

            if qtd_docs == 0:
                msg = "Nenhum procNFe retornado nessa iteração. Encerrando."
                print(f"[{nome_empresa}] {msg}")
                log_geral.append("  " + msg)
                break

            time.sleep(INTERVALO_SEGUNDOS)

    except Exception as e:
        msg = f"Erro inesperado no processamento da empresa '{nome_empresa}': {e}"
        print(f"[{nome_empresa}] {msg}")
        log_geral.append("  " + msg)

    log_geral.append(f"  Processamento concluído para '{nome_empresa}'.")
    log_geral.append(f"  Total de procNFe recebidos: {total_documentos}")
    log_geral.append(f"  NSU final utilizado: {ult_nsu}")

    try:
        empresa_row["NSU"] = str(int(ult_nsu))
    except Exception:
        empresa_row["NSU"] = ult_nsu


# ==========================
# main
# ==========================

def main():
    base_dir = get_script_dir()
    print(f"Pasta do script: {base_dir}")

    caminho_xlsx = os.path.join(base_dir, "certificados.xlsx")
    if not os.path.isfile(caminho_xlsx):
        print(f"XLSX não encontrado na pasta do script: {caminho_xlsx}")
        print("Crie o arquivo com as colunas (linha 1 - cabeçalho):")
        print("NOME | CNPJ | CAMINHO CERTIFICADO | SENHA CERTIFICADO | NSU")
        sys.exit(1)

    empresas = carregar_empresas_xlsx(caminho_xlsx)

    if not empresas:
        print("Nenhuma empresa encontrada na planilha.")
        sys.exit(0)

    print(f"{len(empresas)} empresa(s) encontrada(s) no XLSX.")

    log_geral = []

    for empresa in empresas:
        processar_empresa(base_dir, empresa, log_geral)
        print("-" * 80)

    salvar_empresas_xlsx(caminho_xlsx, empresas)
    print("NSUs atualizados no XLSX.")

    escrever_log_geral(base_dir, log_geral)


if __name__ == "__main__":
    main()
