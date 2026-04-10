from __future__ import annotations

import argparse
import hashlib
import json
import re
import subprocess
import sys
import tempfile
import time
from collections import Counter
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

from chrome_9222 import PORT, chrome_9222

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except Exception:  # pragma: no cover
    tk = None
    filedialog = None
    messagebox = None

if sys.platform.startswith("win"):
    import ctypes


SITE_URL = "https://www.nfe.fazenda.gov.br/portal/consultaRecaptcha.aspx?tipoConsulta=resumo&tipoConteudo=7PhJ%20gAVw2g="
KEY_RE = re.compile(r"\D+")
DOWNLOAD_TIMEOUT_FIRST_MS = 10 * 60 * 1000
DOWNLOAD_TIMEOUT_NEXT_MS = 180 * 1000
CONSULT_SELECTOR = (
    "input[id*='btnConsultar'], input[name*='btnConsultar'], "
    "button[id*='btnConsultar'], button[name*='btnConsultar'], "
    "a[id*='btnConsultar'], a[name*='btnConsultar']"
)
DOWNLOAD_SELECTOR = (
    "input[id*='btnDownload'], input[name*='btnDownload'], "
    "button[id*='btnDownload'], button[name*='btnDownload'], "
    "a[id*='btnDownload'], a[name*='btnDownload']"
)


def log(message: str) -> None:
    now = time.strftime("%H:%M:%S")
    print(f"[{now}] {message}", flush=True)


def normalize_digits(value: object) -> str:
    return KEY_RE.sub("", "" if value is None else str(value).strip())


def normalize_header(value: object) -> str:
    text = ("" if value is None else str(value)).strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def pick_file(title: str) -> Optional[Path]:
    if tk is None or filedialog is None:
        return None
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    selected = filedialog.askopenfilename(
        title=title,
        filetypes=[("Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
    )
    root.destroy()
    return Path(selected) if selected else None


def pick_folder(title: str, initial_dir: Optional[Path] = None) -> Optional[Path]:
    if tk is None or filedialog is None:
        return None
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    selected = filedialog.askdirectory(
        title=title,
        initialdir=str(initial_dir) if initial_dir else None,
        mustexist=False,
    )
    root.destroy()
    return Path(selected) if selected else None


def copy_to_clipboard(text: str) -> None:
    if sys.platform.startswith("win"):
        escaped = text.replace("'", "''")
        subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-Command",
                f"Set-Clipboard -Value '{escaped}'",
            ],
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        return

    if tk is not None:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        root.clipboard_clear()
        root.clipboard_append(text)
        root.update()
        root.destroy()
        return

    subprocess.run(
        ["powershell", "-NoProfile", "-Command", f"Set-Clipboard -Value @'\n{text}\n'@"],
        check=False,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    counter = 2
    while True:
        candidate = parent / f"{stem}_{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def save_watchdog_state(state_path: Path, data: dict) -> None:
    state_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def load_watchdog_state(state_path: Path) -> dict:
    return json.loads(state_path.read_text(encoding="utf-8"))


def clear_watchdog_state(state_path: Path) -> None:
    try:
        state_path.unlink()
    except FileNotFoundError:
        pass
    except Exception:
        pass


def state_path_for_excel(excel_path: Path) -> Path:
    digest = hashlib.md5(str(excel_path.resolve()).encode("utf-8"), usedforsecurity=False).hexdigest()
    return Path(tempfile.gettempdir()) / f"baixar_xml_fsist_watchdog_{digest}.json"


def resolve_excel_path(arg_path: Optional[Path]) -> Path:
    if arg_path is not None:
        return arg_path
    if tk is None or filedialog is None:
        raise SystemExit("Tkinter indisponivel. Informe o arquivo com --excel.")
    selected = pick_file("Selecione o Excel com as chaves")
    if not selected:
        raise SystemExit("Operacao cancelada: arquivo Excel nao selecionado.")
    return selected


def resolve_output_dir(arg_path: Optional[Path], excel_path: Path) -> Path:
    if arg_path is not None:
        return arg_path
    if tk is not None and filedialog is not None:
        selected = pick_folder("Selecione a pasta para salvar os XMLs", excel_path.parent)
        if selected:
            return selected
    return excel_path.parent / f"{excel_path.stem}_XML"


def load_keys_from_workbook(workbook_path: Path) -> list[str]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        best_sheet = None
        best_col = None
        best_score = -1

        header_aliases = {
            "chave",
            "chave de acesso",
            "chaves",
            "chave nota",
            "chave da nota",
            "chave da nf",
            "chave nf",
            "chave nfe",
            "chave cte",
        }

        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=20):
                for cell in row:
                    header = normalize_header(cell.value)
                    if header in header_aliases:
                        best_sheet = ws
                        best_col = cell.column
                        best_score = 10_000
                        break
                if best_col is not None:
                    break
            if best_col is not None:
                break

            column_scores = Counter()
            for row in ws.iter_rows():
                for cell in row:
                    digits = normalize_digits(cell.value)
                    if len(digits) == 44:
                        column_scores[cell.column] += 1
            if column_scores:
                col, score = column_scores.most_common(1)[0]
                if score > best_score:
                    best_sheet = ws
                    best_col = col
                    best_score = score

        keys: list[str] = []
        seen = set()

        if best_sheet is not None and best_col is not None:
            for row in best_sheet.iter_rows(min_row=1):
                if best_col > len(row):
                    continue
                digits = normalize_digits(row[best_col - 1].value)
                if len(digits) != 44 or digits in seen:
                    continue
                seen.add(digits)
                keys.append(digits)

        if keys:
            return keys

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    digits = normalize_digits(cell.value)
                    if len(digits) != 44 or digits in seen:
                        continue
                    seen.add(digits)
                    keys.append(digits)

        return keys
    finally:
        wb.close()


def prepare_single_page(context):
    pages = [page for page in context.pages if not page.is_closed()]
    if not pages:
        return context.new_page()

    page = pages[0]
    for other in pages[1:]:
        try:
            other.close()
        except Exception:
            pass
    try:
        page.bring_to_front()
    except Exception:
        pass
    return page


def wait_for_ready(page, navigate: bool = True):
    last_error = None
    for attempt in range(2):
        try:
            if navigate or attempt == 1:
                page.goto(SITE_URL, wait_until="domcontentloaded", timeout=30_000)
            page.locator("#ctl00_ContentPlaceHolder1_txtChaveAcessoResumo").wait_for(
                state="visible",
                timeout=30_000,
            )
            return page
        except Exception as exc:
            last_error = exc
            time.sleep(1)
    raise RuntimeError("Nao foi possivel abrir a pagina de consulta da NF-e.") from last_error


def fill_key(page, key: str) -> None:
    field = page.locator("#ctl00_ContentPlaceHolder1_txtChaveAcessoResumo")
    field.wait_for(state="visible", timeout=30_000)
    field.click(timeout=10_000)
    try:
        field.press("Control+A")
    except Exception:
        pass
    copy_to_clipboard(key)
    page.wait_for_timeout(150)
    field.press("Control+V")
    page.wait_for_timeout(150)
    if normalize_digits(field.input_value()) != key:
        field.press_sequentially(key, delay=5)
    if normalize_digits(field.input_value()) != key:
        field.fill(key)
    field.press("Tab")
    page.wait_for_timeout(467)


def resolve_consulta_page(context, fallback_page):
    deadline = time.time() + 60
    best_page = fallback_page

    while time.time() < deadline:
        pages = [p for p in context.pages if not p.is_closed()]
        for candidate in reversed(pages):
            best_page = candidate
            try:
                current_url = (candidate.url or "").lower()
            except Exception:
                current_url = ""

            if "consultaresumo" in current_url:
                return candidate
        time.sleep(0.25)

    raise RuntimeError(
        "Nao foi possivel detectar mudanca de URL para consultaResumo apos clicar em Consultar."
    )


def resolve_consulta_page_after_manual(context, fallback_page):
    deadline = time.time() + 60
    best_page = fallback_page

    while time.time() < deadline:
        pages = [p for p in context.pages if not p.is_closed()]
        for candidate in reversed(pages):
            best_page = candidate
            try:
                current_url = (candidate.url or "").lower()
            except Exception:
                current_url = ""

            if "consultaresumo" in current_url:
                return candidate

            try:
                if candidate.locator(DOWNLOAD_SELECTOR).first.is_visible(timeout=250):
                    return candidate
            except Exception:
                pass
        time.sleep(0.25)

    raise RuntimeError(
        "Nao foi possivel detectar a tela de resumo apos a etapa manual da primeira nota."
    )


def wait_user_after_manual_first_download(index: int) -> None:
    message = (
        f"NOTA {index}: apos clicar em Consultar, faca o primeiro download manual agora "
        "(mesmo que salve na pasta errada), confirme o aviso e selecione o certificado digital. "
        "Depois confirme para o script detectar a tela de resumo e baixar novamente na pasta correta."
    )
    print(message, flush=True)
    input(f"{message}\nPressione Enter para continuar...")


def register_passive_dialog_logging(context, index: int):
    handlers = []

    def on_dialog(dialog):
        try:
            log(f"NOTA {index}: dialog detectado durante etapa manual -> {dialog.type}: {dialog.message}")
        except Exception:
            log(f"NOTA {index}: dialog detectado durante etapa manual")

    for candidate in [p for p in context.pages if not p.is_closed()]:
        try:
            candidate.on("dialog", on_dialog)
            handlers.append((candidate, on_dialog))
        except Exception:
            pass
    return handlers


def unregister_passive_dialog_logging(handlers) -> None:
    for candidate, handler in handlers:
        try:
            candidate.remove_listener("dialog", handler)
        except Exception:
            pass


def bring_consulta_page_to_front(context, fallback_page):
    pages = [p for p in context.pages if not p.is_closed()]
    for candidate in reversed(pages):
        try:
            current_url = (candidate.url or "").lower()
        except Exception:
            current_url = ""
        if "consultaresumo" in current_url:
            try:
                candidate.bring_to_front()
            except Exception:
                pass
            return candidate

    try:
        fallback_page.bring_to_front()
    except Exception:
        pass
    return fallback_page


def register_auto_accept_dialog(page, index: int):
    def on_dialog(dialog):
        try:
            log(f"NOTA {index}: dialog detectado -> {dialog.type}: {dialog.message}")
        except Exception:
            log(f"NOTA {index}: dialog detectado durante clique em Download")
        try:
            dialog.accept()
            log(f"NOTA {index}: dialog aceito automaticamente")
        except Exception as exc:
            log(f"NOTA {index}: falha ao aceitar dialog automaticamente: {exc}")

    page.on("dialog", on_dialog)
    return on_dialog


def unregister_auto_accept_dialog(page, handler) -> None:
    try:
        page.remove_listener("dialog", handler)
    except Exception:
        pass


def native_click_screen_point(x: float, y: float) -> None:
    if not sys.platform.startswith("win"):
        raise RuntimeError("Clique nativo de tela disponivel apenas no Windows.")
    ctypes.windll.user32.SetCursorPos(int(round(x)), int(round(y)))
    ctypes.windll.user32.mouse_event(0x0002, 0, 0, 0, 0)
    ctypes.windll.user32.mouse_event(0x0004, 0, 0, 0, 0)


def get_element_screen_point(page, button):
    handle = button.element_handle()
    if handle is None:
        return None
    return page.evaluate(
        """
        (el) => {
            const rect = el.getBoundingClientRect();
            const borderX = Math.max((window.outerWidth - window.innerWidth) / 2, 0);
            const borderY = Math.max(window.outerHeight - window.innerHeight - borderX, 0);
            return {
                x: window.screenX + borderX + rect.left + (rect.width / 2),
                y: window.screenY + borderY + rect.top + (rect.height / 2)
            };
        }
        """,
        handle,
    )


def click_download_button(page, button, index: int) -> None:
    page.bring_to_front()
    log(f"NOTA {index}: foco no Chrome antes do clique em Download")
    button.scroll_into_view_if_needed(timeout=5_000)
    log(f"NOTA {index}: botao Download rolado para a tela")

    screen_point = None
    try:
        screen_point = get_element_screen_point(page, button)
    except Exception:
        screen_point = None

    if screen_point is not None:
        x = screen_point["x"]
        y = screen_point["y"]
        log(f"NOTA {index}: clicando em Download por coordenada nativa ({x:.1f}, {y:.1f})")
        native_click_screen_point(x, y)
        return

    log(f"NOTA {index}: sem coordenada nativa, usando click no elemento")
    button.click(timeout=30_000)


def consult_and_download(context, page, output_dir: Path, key: str, index: int, first_note: bool) -> tuple[Path, object]:
    log(f"NOTA {index}: iniciando consulta para chave {key}")
    consult_button = page.locator(CONSULT_SELECTOR).first
    log(f"NOTA {index}: aguardando botao Consultar")
    consult_button.wait_for(state="visible", timeout=30_000)
    log(f"NOTA {index}: clicando em Consultar")
    consult_button.click(timeout=30_000)
    if first_note:
        log(f"NOTA {index}: pausa manual antes da deteccao de URL")
        dialog_handlers = register_passive_dialog_logging(context, index)
        try:
            wait_user_after_manual_first_download(index)
            log(f"NOTA {index}: input recebido, retomando fluxo automatico")
            page = bring_consulta_page_to_front(context, page)
            log(f"NOTA {index}: foco retornado ao Chrome")
            log(f"NOTA {index}: reconhecendo tela de resumo apos etapa manual")
            page = resolve_consulta_page_after_manual(context, page)
        finally:
            unregister_passive_dialog_logging(dialog_handlers)
    else:
        log(f"NOTA {index}: aguardando mudanca de URL para consultaResumo")
        page = resolve_consulta_page(context, page)
    log(f"NOTA {index}: URL confirmada -> {page.url}")
    page.wait_for_timeout(500)

    button = page.locator(DOWNLOAD_SELECTOR).first
    log(f"NOTA {index}: aguardando botao Download visivel")
    button.wait_for(state="visible", timeout=60_000)

    timeout_ms = DOWNLOAD_TIMEOUT_FIRST_MS if first_note else DOWNLOAD_TIMEOUT_NEXT_MS
    try:
        dialog_handler = register_auto_accept_dialog(page, index)
        try:
            with page.expect_download(timeout=timeout_ms) as download_info:
                click_download_button(page, button, index)
                log(f"NOTA {index}: clique automatico enviado, aguardando evento de download")
            download = download_info.value
        finally:
            unregister_auto_accept_dialog(page, dialog_handler)
        log(f"NOTA {index}: download detectado ({download.suggested_filename})")
        target = unique_path(output_dir / f"{index:04d}_{key}.xml")
        log(f"NOTA {index}: salvando XML em {target}")
        download.save_as(str(target))
        log(f"NOTA {index}: XML salvo com sucesso")
        return target, page
    except PlaywrightTimeoutError as exc:
        log(f"NOTA {index}: timeout aguardando download")
        raise RuntimeError(
            f"Timeout aguardando o download da nota {index} ({key}). "
            "Se o Chrome abriu a selecao de certificado, conclua a selecao e tente novamente."
        ) from exc


def reset_for_next(page):
    page.goto(SITE_URL, wait_until="domcontentloaded", timeout=30_000)
    page.locator("#ctl00_ContentPlaceHolder1_txtChaveAcessoResumo").wait_for(state="visible", timeout=30_000)
    return page


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Baixa XMLs do portal da NF-e a partir de chaves em Excel.")
    parser.add_argument("--excel", type=Path, help="Arquivo Excel com as chaves.")
    parser.add_argument("--saida", type=Path, help="Pasta para salvar os XMLs.")
    parser.add_argument("--resume-state", type=Path, help="Arquivo de estado usado pelo watchdog.")
    return parser.parse_args(argv)


def main() -> None:
    args = parse_args(sys.argv[1:])
    resume_state = args.resume_state

    if resume_state is not None:
        if not resume_state.exists():
            raise SystemExit(f"Arquivo de estado nao encontrado: {resume_state}")
        resume_data = load_watchdog_state(resume_state)
        excel_path = resolve_excel_path(args.excel) if args.excel is not None else Path(resume_data["excel_path"])
        output_dir = resolve_output_dir(args.saida, excel_path) if args.saida is not None else Path(
            resume_data["output_dir"]
        )
        start_index = int(resume_data.get("index", 0))
        start_stage = str(resume_data.get("stage", "consult"))
    else:
        excel_path = resolve_excel_path(args.excel)
        output_dir = resolve_output_dir(args.saida, excel_path)
        output_dir.mkdir(parents=True, exist_ok=True)
        start_index = 0
        start_stage = "fill"
        clear_watchdog_state(state_path_for_excel(excel_path))

    if not excel_path.exists():
        raise SystemExit(f"Arquivo Excel nao encontrado: {excel_path}")

    output_dir.mkdir(parents=True, exist_ok=True)
    state_path = state_path_for_excel(excel_path)

    keys = load_keys_from_workbook(excel_path)
    if not keys:
        raise SystemExit("Nenhuma chave valida de 44 digitos foi encontrada no Excel.")
    if start_index < 0 or start_index >= len(keys):
        raise SystemExit("Estado de retomada invalido.")

    with sync_playwright() as p:
        browser = chrome_9222(p, PORT)
        context = browser.contexts[0]
        page = prepare_single_page(context)
        page = wait_for_ready(page, navigate=(start_stage == "fill"))

        if start_stage == "fill":
            fill_key(page, keys[start_index])
            save_watchdog_state(
                state_path,
                {
                    "excel_path": str(excel_path),
                    "output_dir": str(output_dir),
                    "index": start_index,
                    "stage": "consult",
                },
            )
            return

        if start_stage != "consult":
            raise SystemExit("Estado de retomada invalido.")

        _, page = consult_and_download(
            context,
            page,
            output_dir,
            keys[start_index],
            start_index + 1,
            first_note=(start_index == 0),
        )
        next_index = start_index + 1
        if next_index >= len(keys):
            clear_watchdog_state(state_path)
            print(f"Arquivos salvos em: {output_dir}")
            print(f"Total processado: {start_index + 1}")
            return

        page = reset_for_next(page)
        fill_key(page, keys[next_index])
        save_watchdog_state(
            state_path,
            {
                "excel_path": str(excel_path),
                "output_dir": str(output_dir),
                "index": next_index,
                "stage": "consult",
            },
        )


if __name__ == "__main__":
    main()
