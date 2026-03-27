# -*- coding: utf-8 -*-
"""
DCTFWEB — versão Playwright (CDP attach) baseada no seu 1.2.3 (Selenium).
- Anexa no Chrome existente via CDP
- Procura elementos em iframes (recursivo)
- Mantém planilhamento incremental e lógica de paginação/recibos

Observação:
- Playwright lida com downloads via expect_download + save_as.
"""

import os
import re
import sys
import time
from datetime import date, datetime, timedelta
from calendar import monthrange
from pathlib import Path
from chrome_9222 import chrome_9222, PORT
import pandas as pd
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError

# ------------------------ CONFIG ------------------------
ALVO_TITULO_CONTEM = "eCAC - Centro Virtual de Atendimento"

DOWNLOAD_ROOT = r"T:\testes"

EXCEL_REFERENCIA_PATH = (
    r"W:\DOCUMENTOS ESCRITORIO\INSTALACAO SISTEMA\python\CONFERIDOR DCTFWEB"
    r"\Consulta DCTFWEB WILSON MARCOS LOPES.xlsx"
)

REF_KEY_COLS = [
    "Número de Identificação",
    "Período de Apuração",
    "Data Transmissão",
    "Situação",
]
REF_KEYS = set()

SEL_IMG_PROCESSANDO = "img.image-processamento, img[class*='image-processamento']"
SEL_TABELA_LISTA = "table#ctl00_cphConteudo_tabelaListagemDctf_GridViewDctfs"
SEL_DROPD_CATEG_BTN = "button.btn.dropdown-toggle.btn-default[data-id='ctl00_cphConteudo_ddlCategoriaDeclaracao']"
SEL_DROPD_OUTORGANTES_BTN = "button.btn.dropdown-toggle.btn-default[data-id='ctl00_cphConteudo_ddlOutorgantes']"

TIMEOUT_PADRAO_MS = 20_000

CHROME_DEBUG_ADDR = os.environ.get("CHROME_DEBUG_ADDR", "http://localhost:9222")
# -------------------------------------------------------

# ------------------------ MODOS ------------------------
DO_PLANILHA = False
DO_RECIBOS = False
DATE_OVERRIDE = None  # dict: {'ap_ini','ap_fim','tx_ini','tx_fim'} como datetime.date


class SessaoExpiradaException(Exception):
    pass


def log(msg: str):
    print(msg, flush=True)


def _parse_br_date(s):
    try:
        return datetime.strptime(s, "%d/%m/%Y").date()
    except Exception:
        return None


def _last_day_of_month(d: date) -> date:
    last = monthrange(d.year, d.month)[1]
    return d.replace(day=last)


def _last_day_of_next_month(d: date) -> date:
    if d.month == 12:
        first_next = date(d.year + 1, 1, 1)
    else:
        first_next = date(d.year, d.month + 1, 1)
    last = monthrange(first_next.year, first_next.month)[1]
    return first_next.replace(day=last)


def _ask_mode_and_dates():
    global DO_PLANILHA, DO_RECIBOS, DATE_OVERRIDE

    print("Selecione o modo de execução:")
    print("  1) Apenas planilhar")
    print("  2) Baixar recibos")
    print("  3) Ambos (planilhar + recibos)")
    escolha = input("Digite 1, 2 ou 3 e pressione Enter: ").strip()

    if escolha not in {"1", "2", "3"}:
        print("Opção inválida. Encerrando.")
        sys.exit(1)

    if escolha == "1":
        DO_PLANILHA = True
        DO_RECIBOS = False
        di_str = input("Informe a Data Inicial (dd/mm/aaaa): ").strip()
        ap_ini = _parse_br_date(di_str)
        if ap_ini is None:
            print("Data inválida. Use o formato dd/mm/aaaa. Encerrando.")
            sys.exit(1)
        ap_fim = _last_day_of_month(ap_ini)
        tx_ini = ap_ini
        tx_fim = _last_day_of_next_month(ap_ini)
        DATE_OVERRIDE = {"ap_ini": ap_ini, "ap_fim": ap_fim, "tx_ini": tx_ini, "tx_fim": tx_fim}
        print("Datas definidas:")
        print(f"  - Período Apuração: {ap_ini.strftime('%d/%m/%Y')} a {ap_fim.strftime('%d/%m/%Y')}")
        print(f"  - Data Transmissão: {tx_ini.strftime('%d/%m/%Y')} a {tx_fim.strftime('%d/%m/%Y')}")
    elif escolha == "2":
        DO_PLANILHA = False
        DO_RECIBOS = True
        DATE_OVERRIDE = None
    else:
        DO_PLANILHA = True
        DO_RECIBOS = True
        DATE_OVERRIDE = None


# ---------------- Playwright helpers (frames) ----------------

def _all_frames(page):
    # inclui main frame e iframes aninhados
    return page.frames


def locator_any_frame(page, selector: str):
    """
    Retorna (frame, locator) do primeiro frame onde o selector existir.
    """
    for fr in _all_frames(page):
        loc = fr.locator(selector)
        try:
            if loc.count() > 0:
                return fr, loc
        except Exception:
            continue
    raise PWTimeoutError(f"Não achei selector em nenhum frame: {selector}")


def wait_visible_any_frame(page, selector: str, timeout_ms=TIMEOUT_PADRAO_MS):
    end = time.time() + (timeout_ms / 1000)
    last_err = None
    while time.time() < end:
        try:
            fr, loc = locator_any_frame(page, selector)
            loc.first.wait_for(state="visible", timeout=500)
            return fr, loc.first
        except Exception as e:
            last_err = e
            time.sleep(0.1)
    raise PWTimeoutError(f"Timeout aguardando visível: {selector}. Último erro: {last_err}")


def click_any_frame(page, selector: str, timeout_ms=TIMEOUT_PADRAO_MS):
    fr, loc = wait_visible_any_frame(page, selector, timeout_ms=timeout_ms)
    loc.click()
    return fr


# def wait_idle_custom(page, quiet_ms=700, timeout_ms=30_000):
    """
    Similar ao seu wait_until_idle (spinner + XHR/fetch pendentes) :contentReference[oaicite:3]{index=3}
    """
    # instala hooks uma vez no main frame e também tenta nos iframes (melhor esforço)
    install_js = r"""
    (function(){
      if (window.__idleHooksInstalled) return;
      window.__idleHooksInstalled = true;
      var Open = XMLHttpRequest.prototype.open;
      var Send = XMLHttpRequest.prototype.send;
      window.__pendingXHR = 0;
      XMLHttpRequest.prototype.open = function(){ return Open.apply(this, arguments); };
      XMLHttpRequest.prototype.send = function(){
        window.__pendingXHR++;
        this.addEventListener('loadend', function(){ window.__pendingXHR--; }, {once:true});
        return Send.apply(this, arguments);
      };
      if (window.fetch){
        var _fetch = window.fetch;
        window.__pendingFetch = 0;
        window.fetch = function(){
          window.__pendingFetch++;
          try{
            return _fetch.apply(this, arguments).finally(function(){ window.__pendingFetch--; });
          }catch(e){
            window.__pendingFetch--;
            throw e;
          }
        };
      } else {
        window.__pendingFetch = 0;
      }
    })();
    """

    for fr in _all_frames(page):
        try:
            fr.evaluate(install_js)
        except Exception:
            pass

    check_js = r"""
      const spinner = document.querySelector("img.image-processamento, img[class*='image-processamento']");
      const ready = document.readyState === 'complete' || document.readyState === 'interactive';
      const jqIdle = (window.jQuery && typeof jQuery.active === 'number') ? (jQuery.active === 0) : true;
      const xhrIdle = (window.__pendingXHR||0) === 0;
      const fetchIdle = (window.__pendingFetch||0) === 0;
      return (!!ready) && (!spinner) && jqIdle && xhrIdle && fetchIdle;
    """

    end = time.time() + timeout_ms / 1000
    stable_since = None
    while time.time() < end:
        ok_any = False
        # se qualquer frame ainda estiver “ocupado”, considera não-idle
        all_ok = True
        for fr in _all_frames(page):
            try:
                ok = bool(fr.evaluate(check_js))
                ok_any = True
                if not ok:
                    all_ok = False
                    break
            except Exception:
                # frames cross-origin podem falhar; ignora
                continue

        if ok_any and all_ok:
            now = time.time()
            stable_since = stable_since or now
            if (now - stable_since) * 1000 >= quiet_ms:
                return True
        else:
            stable_since = None

        time.sleep(0.08)

    raise PWTimeoutError("Página não ficou idle a tempo.")


# ---------------- Regras de negócio (iguais) ----------------

def checar_sessao_expirada(page):
    html = ""
    try:
        html = page.content()
    except Exception:
        return
    if "Sessão Expirada" in html or "Sessao Expirada" in html or "sessao-expirada" in html:
        raise SessaoExpiradaException("Sessão Expirada no eCAC detectada.")

def datas_mes_anterior_e_hoje():
    hoje = date.today()
    primeiro_deste = hoje.replace(day=1)
    ultimo_mes = primeiro_deste - timedelta(days=1)
    return ultimo_mes.replace(day=1), ultimo_mes, hoje

def _set_input_value(frame, selector_or_locator, value: str):
    if isinstance(selector_or_locator, str):
        loc = frame.locator(selector_or_locator).first
    else:
        loc = selector_or_locator
    try:
        loc.fill(value)
    except Exception:
        # fallback via JS
        frame.evaluate(
            """(sel, v)=>{
                const el = document.querySelector(sel);
                if(!el) return false;
                el.value = v;
                el.dispatchEvent(new Event('input', {bubbles:true}));
                el.dispatchEvent(new Event('change', {bubbles:true}));
                return true;
            }""",
            loc._selector, value  # internals; se te incomodar, troque por selector string sempre
        )

def preencher_datas(page, override=None):
    if override:
        d_ap_ini = override["ap_ini"]
        d_ap_fim = override["ap_fim"]
        d_tx_ini = override["tx_ini"]
        d_tx_fim = override["tx_fim"]
    else:
        d_ap_ini, d_ap_fim, d_tx_fim = datas_mes_anterior_e_hoje()
        d_tx_ini = d_ap_ini

    dd = lambda d: d.strftime("%d/%m/%Y")
    log("Preenchendo datas…")

    # tenta por ids conhecidos (como no seu) :contentReference[oaicite:4]{index=4}
    campos = [
        (["#txtDataInicio", "input[id*='txtPeriodoApuracaoInicial']", "input[id*='txtPeriodoApuracaoInicio']"], dd(d_ap_ini), "Período Apuração Inicial"),
        (["#txtDataFinal", "input[id*='txtPeriodoApuracaoFinal']", "input[id*='txtPeriodoApuracaoFim']"], dd(d_ap_fim), "Período Apuração Final"),
        (["input[id*='txtDataTransmissaoInicial']", "input[id*='txtDtTransmissaoInicial']"], dd(d_tx_ini), "Data Transmissão Inicial"),
        (["input[id*='txtDataTransmissaoFinal']", "input[id*='txtDtTransmissaoFinal']"], dd(d_tx_fim), "Data Transmissão Final"),
    ]

    for selectors, val, label in campos:
        ok = False
        for sel in selectors:
            try:
                fr, loc = wait_visible_any_frame(page, sel, timeout_ms=60_000)
                loc.fill(val)
                log(f"  - {label} = {val}")
                ok = True
                break
            except Exception:
                continue
        if not ok:
            log(f"  ! {label}: não consegui preencher.")

def selecionar_categoria_indice5(page):
    log("Selecionando Categoria índice 5…")
    # Mantém sua estratégia “bootstrap-select” :contentReference[oaicite:5]{index=5}
    try:
        click_any_frame(page, SEL_DROPD_CATEG_BTN, timeout_ms=10_000)
        fr, _ = locator_any_frame(page, "div.open ul.dropdown-menu.inner")
        # limpar e selecionar índice 4 (equivalente ao seu clique)
        try:
            fr.locator("button.actions-btn.bs-deselect-all").first.click()
            time.sleep(0.1)
        except Exception:
            pass
        fr.locator("div.open ul.dropdown-menu.inner li[data-original-index='4'] a").first.click()
        # fecha dropdown
        click_any_frame(page, SEL_DROPD_CATEG_BTN, timeout_ms=5_000)
        log("  - Categoria marcada via bootstrap-select.")
        return
    except Exception:
        pass

    # fallback: select nativo
    try:
        fr, loc = wait_visible_any_frame(page, "select[id*='ddlCategoriaDeclaracao']", timeout_ms=8_000)
        # select_index(5) (Playwright não tem direto; usa evaluate)
        fr.evaluate(
            """(sel)=>{
                const el = document.querySelector(sel);
                if(!el) return false;
                el.selectedIndex = 5;
                el.dispatchEvent(new Event('change', {bubbles:true}));
                return true;
            }""",
            "select[id*='ddlCategoriaDeclaracao']"
        )
        log("  - Categoria marcada no <select> nativo (fallback).")
    except Exception as e:
        log(f"  ! Não consegui marcar categoria índice 5: {e}")

def marcar_sou_procurador(page):
    log("Marcando 'Sou Procurador'…")
    seletor_cb = (
        "input[type='checkbox'][id*='ListarOutorgantes'], "
        "input[type='checkbox'][id*='chkListarOutorgantes'], "
        "input[type='checkbox'][id='ctl00_cphConteudo_chkListarOutorgantes']"
    )
    fr, loc = wait_visible_any_frame(page, seletor_cb, timeout_ms=20_000)
    loc.click()
    # aguarda processamento (como seu fluxo) :contentReference[oaicite:6]{index=6}
    try:
        wait_visible_any_frame(page, SEL_IMG_PROCESSANDO, timeout_ms=2_000)
    except Exception:
        pass
    page.wait_for_load_state("domcontentloaded", timeout=60_000)
    wait_visible_any_frame(page, SEL_DROPD_OUTORGANTES_BTN, timeout_ms=20_000)

def selecionar_outorgantes_apenas_cnpj(page):
    log("Outorgantes: filtrar 'CNPJ' e selecionar 'Todas'…")
    page.wait_for_load_state("domcontentloaded", timeout=60_000)
    click_any_frame(page, SEL_DROPD_OUTORGANTES_BTN, timeout_ms=10_000)

    fr, _ = locator_any_frame(page, "div.open ul.dropdown-menu")
    # limpar
    try:
        fr.locator("button.actions-btn.bs-deselect-all").first.click()
        time.sleep(0.1)
    except Exception:
        pass

    # busca "CNPJ"
    try:
        s = fr.locator("div.dropdown-menu .bs-searchbox input").first
        s.fill("CNPJ")
        time.sleep(0.2)
        try:
            fr.locator("button.actions-btn.bs-select-all").first.click()
        except Exception:
            # clica manualmente nos itens visíveis
            items = fr.locator("div.open ul.dropdown-menu.inner li:not(.hidden):not(.disabled) a")
            n = items.count()
            for i in range(n):
                try:
                    items.nth(i).click()
                    time.sleep(0.02)
                except Exception:
                    pass
    except Exception:
        # fallback: tenta aba "CNPJ"
        try:
            fr.locator("//div[contains(@class,'open')]//a[normalize-space()='CNPJ' or contains(.,'CNPJ')]").first.click()
            time.sleep(0.1)
            fr.locator("button.actions-btn.bs-select-all").first.click()
        except Exception:
            pass

    # fecha
    click_any_frame(page, SEL_DROPD_OUTORGANTES_BTN, timeout_ms=5_000)
    page.wait_for_load_state("domcontentloaded", timeout=60_000)

def clicar_acesso_govbr(page, timeout_ms=20000):
    """
    Clica no botão/imagem 'Acesso Gov BR'.
    Usa múltiplos seletores para garantir robustez.
    """
    seletores = [
        "input[type='image'][alt='Acesso Gov BR']",
        "input[type='image'][onclick*='govBr']",
        "input[type='image'][src*='gov-br.png']",
        "//input[@type='image' and contains(@alt,'Gov')]",
    ]

    ultimo_erro = None

    for sel in seletores:
        try:
            # suporta CSS e XPath
            if sel.startswith("//"):
                loc = page.locator(f"xpath={sel}").first
            else:
                loc = page.locator(sel).first

            loc.wait_for(state="visible", timeout=timeout_ms)
            loc.click()
            return True
        except Exception as e:
            ultimo_erro = e
            continue

    raise PWTimeoutError(
        f"Não foi possível clicar no botão 'Acesso Gov BR'. Último erro: {ultimo_erro}"
    )

def clicar_checkbox_div_em_frames(page, timeout_ms=20000, marcar=True):
    """
    Marca checkbox ARIA (hCaptcha / Gov.br) em iframes.
    Clica UMA VEZ e aguarda confirmação real.
    """

    seletores = [
        "div[role='checkbox']#checkbox",
        "div[role='checkbox'][aria-labelledby='a11y-label']",
        "div[role='checkbox']",
    ]

    fim = time.time() + (timeout_ms / 1000)
    clicou = False
    ultimo_erro = None

    while time.time() < fim:
        for frame in page.frames:
            for sel in seletores:
                try:
                    loc = frame.locator(sel).first
                    if loc.count() == 0:
                        continue

                    loc.wait_for(state="visible", timeout=800)

                    estado = loc.get_attribute("aria-checked")
                    estado_bool = estado == "true"

                    # já marcado
                    if estado_bool and marcar:
                        return True

                    # clique UMA vez
                    if not clicou:
                        try:
                            loc.click()
                        except Exception:
                            loc.focus()
                            page.keyboard.press("Space")
                        clicou = True
                        # após clicar, sai do loop interno para aguardar estabilização
                        break

                except Exception as e:
                    ultimo_erro = e
                    continue

        # após clique, aguarda mudança estrutural (iframe costuma recarregar)
        if clicou:
            try:
                page.wait_for_timeout(1200)
                # se checkbox sumiu, é SUCESSO (captcha validou)
                if not any(
                    frame.locator("div[role='checkbox']").count() > 0
                    for frame in page.frames
                ):
                    return True
            except Exception:
                pass

        time.sleep(0.2)

    raise PWTimeoutError(
        f"Checkbox clicado mas não confirmou estado a tempo. Último erro: {ultimo_erro}"
    )

def wait_idle_app(page, timeout_ms=30000):
    page.wait_for_load_state("domcontentloaded", timeout=timeout_ms)
    page.wait_for_load_state("networkidle", timeout=timeout_ms)

def clicar_filtrar_e_aguardar(page):
    log("Filtrar…")
    # tenta __doPostBack direto (como você fazia) :contentReference[oaicite:7]{index=7}
    ok = False
    for fr in _all_frames(page):
        try:
            ok = bool(fr.evaluate(
                "(()=>{ if (typeof __doPostBack==='function'){ __doPostBack('ctl00$cphConteudo$imageFiltrar',''); return true;} return false; })()"
            ))
            if ok:
                break
        except Exception:
            continue

    if not ok:
        # clique no botão/âncora
        try:
            click_any_frame(page, "a#ctl00_cphConteudo_imageFiltrar, a[id*='imageFiltrar']", timeout_ms=10_000)
        except Exception:
            click_any_frame(page, "a.btn.btn-icon-sec.btn-exibe-processando.image-pesquisar", timeout_ms=10_000)

    try:
        wait_visible_any_frame(page, SEL_IMG_PROCESSANDO, timeout_ms=2_000)
    except Exception:
        pass

    wait_idle_app(page)
    wait_visible_any_frame(page, SEL_TABELA_LISTA, timeout_ms=30_000)

def _extrair_dados_de_tr_playwright(tr_loc):
    tds = tr_loc.locator("td")
    if tds.count() < 9:
        return {}

    def gt(i):
        try:
            return (tds.nth(i).inner_text() or "").strip()
        except Exception:
            return ""

    tipo_ni = gt(0)
    numero_id = gt(1)
    periodo_apur = gt(2)
    dt_transm = gt(3)
    categoria = gt(4)
    origem = gt(5)
    tipo = gt(6)

    try:
        situacao = (tr_loc.locator(".coluna-lblSituacao span").first.inner_text() or "").strip()
    except Exception:
        situacao = gt(7)

    valor_apurado = gt(8)
    try:
        saldo_pagar = (tr_loc.locator("span[id*='lblSaldoPagar']").first.inner_text() or "").strip()
    except Exception:
        saldo_pagar = gt(9) if tds.count() > 9 else ""

    return {
        "Tipo NI": tipo_ni,
        "Número de Identificação": numero_id,
        "Período de Apuração": periodo_apur,
        "Data Transmissão": dt_transm,
        "Categoria": categoria,
        "Origem": origem,
        "Tipo": tipo,
        "Situação": situacao,
        "Valor Apurado": valor_apurado,
        "Saldo a Pagar": saldo_pagar,
    }

def coletar_tabela_pagina(page):
    page.wait_for_load_state("domcontentloaded", timeout=60_000)
    log("Coletando linhas da página…")
    fr, tbl = wait_visible_any_frame(page, SEL_TABELA_LISTA, timeout_ms=60_000)

    rows = tbl.locator("tr")
    n = rows.count()
    dados = []
    # pula cabeçalho
    for i in range(1, n):
        tr = rows.nth(i)
        try:
            d = _extrair_dados_de_tr_playwright(tr)
            if d:
                dados.append(d)
        except Exception:
            continue

    log(f"Linhas coletadas: {len(dados)}")
    return dados

def carregar_excel_referencia():
    global REF_KEYS
    REF_KEYS = set()

    if not os.path.exists(EXCEL_REFERENCIA_PATH):
        log(f"Excel de referência NÃO encontrado: {EXCEL_REFERENCIA_PATH}")
        log("Nenhuma linha será pulada com base nesse arquivo.")
        return

    try:
        df_ref = pd.read_excel(EXCEL_REFERENCIA_PATH)
    except Exception as e:
        log(f"Erro ao ler Excel de referência: {e}")
        log("Nenhuma linha será pulada com base nesse arquivo.")
        return

    faltando = [c for c in REF_KEY_COLS if c not in df_ref.columns]
    if faltando:
        log(f"Excel de referência não contém as colunas esperadas: {faltando}")
        log("Nenhuma linha será pulada com base nesse arquivo.")
        return

    for _, row in df_ref.iterrows():
        chave = tuple(str(row[c]).strip() for c in REF_KEY_COLS)
        REF_KEYS.add(chave)

    log(f"Excel de referência carregado. Linhas cadastradas: {len(REF_KEYS)}")

def linha_ja_processada(dados_linha: dict) -> bool:
    if not REF_KEYS:
        return False
    chave = tuple(str(dados_linha.get(c, "")).strip() for c in REF_KEY_COLS)
    return chave in REF_KEYS

def salvar_em_excel_incremental(basepath, nome_titular, dados_pagina):
    if not dados_pagina:
        return None

    caminho = Path(EXCEL_REFERENCIA_PATH)
    caminho.parent.mkdir(parents=True, exist_ok=True)

    df = pd.DataFrame(dados_pagina)
    for col in ["Valor Apurado", "Saldo a Pagar"]:
        if col in df.columns:
            df[col] = (
                df[col].astype(str)
                .str.replace(".", "", regex=False)
                .str.replace(",", ".", regex=False)
                .apply(lambda x: pd.to_numeric(x, errors="coerce"))
            )

    sheet = "DCTFWeb"

    if caminho.exists():
        from openpyxl import load_workbook
        wb = load_workbook(caminho)
        startrow = wb[sheet].max_row if sheet in wb.sheetnames else 0

        with pd.ExcelWriter(caminho, engine="openpyxl", mode="a", if_sheet_exists="overlay") as w:
            df.to_excel(w, index=False, header=(startrow == 0), startrow=startrow, sheet_name=sheet)
    else:
        with pd.ExcelWriter(caminho, engine="openpyxl", mode="w") as w:
            df.to_excel(w, index=False, sheet_name=sheet)

    log(f"Excel (incremental) salvo/atualizado em: {caminho}")
    return str(caminho)

def _ensure_dir(p):
    Path(p).mkdir(parents=True, exist_ok=True)
    return str(Path(p))

def baixar_recibo_linha(page, link_locator, timeout_ms=30000):
    """
    Baixa recibo clicando no link REAL.
    Compatível com ASP.NET AJAX (DCTFWeb).
    """

    # garante que está visível e habilitado
    link_locator.wait_for(state="visible", timeout=timeout_ms)

    # remove disabled / bloqueios visuais
    page.evaluate(
        """(el)=>{
            el.removeAttribute('disabled');
            el.classList.remove('disabled');
            el.style.pointerEvents = 'auto';
        }""",
        link_locator,
    )

    # captura download real
    with page.expect_download(timeout=timeout_ms) as download_info:
        try:
            # 1️⃣ clique humano
            link_locator.click()
        except Exception:
            # 2️⃣ fallback: dispatch de evento real
            page.evaluate(
                """(el)=>{
                    el.dispatchEvent(new MouseEvent('click', {
                        bubbles: true,
                        cancelable: true,
                        view: window
                    }));
                }""",
                link_locator,
            )

    download = download_info.value
    return download

def ir_para_proxima_pagina_e_aguardar(page):
    """
    Similar ao seu ir_para_proxima_pagina_e_aguardar :contentReference[oaicite:8]{index=8}
    """
    try:
        fr, img = wait_visible_any_frame(page, "img[id*='imgNextPage']", timeout_ms=5_000)
    except Exception:
        return False

    # ancestor <a> e checa disabled
    a = img.locator("xpath=ancestor::a[1]").first
    cls = (a.get_attribute("class") or "").lower()
    if "disabled" in cls:
        return False

    a.click()
    # aguarda estabilizar
    try:
        page.wait_for_load_state("domcontentloaded", timeout=TIMEOUT_PADRAO_MS)
    except Exception:
        pass
    page.wait_for_load_state("domcontentloaded", timeout=60_000)
    wait_visible_any_frame(page, SEL_TABELA_LISTA, timeout_ms=30_000)
    return True

def selecionar_aba_por_titulo(browser, alvo_contem: str):
    """
    No Playwright você pega 'pages' do context.
    Aqui seleciona a primeira aba cujo título contenha alvo_contem.
    """
    alvo = alvo_contem.lower()
    for ctx in browser.contexts:
        for p in ctx.pages:
            try:
                title = (p.title() or "").strip()
            except Exception:
                continue
            log(f"- '{title}'")
            if alvo in title.lower():
                log(f"Aba selecionada: '{title}'")
                return p
    log("Aba alvo não encontrada; usando a primeira aba disponível.")
    # fallback
    for ctx in browser.contexts:
        if ctx.pages:
            return ctx.pages[0]
    raise RuntimeError("Nenhuma aba encontrada no browser CDP.")

def extrair_nome_titular(page):
    try:
        fr, el = wait_visible_any_frame(page, "#informacao-perfil", timeout_ms=5_000)
        txt = (el.inner_text() or "").strip()
        m = re.search(r"Titular\s*\(.*?\)\s*[\d\.\-\/]+\s*-\s*(.+)$", txt, flags=re.IGNORECASE | re.MULTILINE)
        if m:
            return m.group(1).strip()
        if " - " in txt:
            return txt.split(" - ", 1)[-1].strip()
    except Exception:
        pass
    return "TITULAR"

def acionar_botao_prosseguir(page, timeout_ms=30000):
    """
    Aciona o botão 'Prosseguir' (ASP.NET / __doPostBack),
    mesmo quando está marcado como disabled.
    Funciona dentro de iframes.
    """

    seletores = [
        "a#ctl00_cphConteudo_btnProsseguir",
        "a[id$='btnProsseguir']",
        "//a[contains(@id,'btnProsseguir')]",
        "//a[normalize-space()='Prosseguir']",
    ]

    ultimo_erro = None
    fim = page.context._impl_obj._loop.time() + (timeout_ms / 1000)

    while page.context._impl_obj._loop.time() < fim:
        for frame in page.frames:
            for sel in seletores:
                try:
                    loc = (
                        frame.locator(f"xpath={sel}").first
                        if sel.startswith("//")
                        else frame.locator(sel).first
                    )

                    if loc.count() == 0:
                        continue

                    loc.wait_for(state="attached", timeout=500)

                    # 1) remove disabled (se existir)
                    frame.evaluate(
                        """(el)=>{
                            if (!el) return;
                            el.removeAttribute('disabled');
                            el.classList.remove('disabled');
                        }""",
                        loc,
                    )

                    # 2) tenta clique direto
                    try:
                        loc.click()
                        return True
                    except Exception:
                        pass

                    # 3) dispara __doPostBack manualmente (mais confiável)
                    ok = frame.evaluate(
                        """(el)=>{
                            const href = el.getAttribute('href') || '';
                            const m = href.match(/__doPostBack\\('([^']+)'\\s*,\\s*'([^']*)'\\)/);
                            if (m && typeof __doPostBack === 'function') {
                                __doPostBack(m[1], m[2] || '');
                                return true;
                            }
                            return false;
                        }""",
                        loc,
                    )
                    if ok:
                        return True

                except Exception as e:
                    ultimo_erro = e
                    continue

        # espera curta antes de tentar novamente
        page.wait_for_timeout(300)

    raise PWTimeoutError(
        f"Não foi possível acionar o botão 'Prosseguir'. Último erro: {ultimo_erro}"
    )

def main():
    _ask_mode_and_dates()

    with sync_playwright() as p:
        browser = chrome_9222(p, PORT)   # conecta ou inicia Chrome
        context = browser.contexts[0]    # contexto persistente
        page = context.new_page()
        
        page.goto("https://det.sit.trabalho.gov.br/login?r=%2Fservicos")
        page.wait_for_load_state("domcontentloaded")
        page.click("text=Entrar com")
        page.wait_for_url("https://det.sit.trabalho.gov.br/servicos")
        page.goto('https://cav.receita.fazenda.gov.br/ecac/Aplicacao.aspx?id=10015&origem=maisacessados')
        clicar_acesso_govbr(page)
        page.wait_for_url("https://cav.receita.fazenda.gov.br/ecac/")
        page.goto('https://cav.receita.fazenda.gov.br/ecac/Aplicacao.aspx?id=10015&origem=maisacessados')


        clicar_checkbox_div_em_frames(page, marcar=True)
        
        acionar_botao_prosseguir(page)       
                
        # pega/seleciona a aba do eCAC (como seu selecionar_aba_por_titulo) :contentReference[oaicite:9]{index=9}
        log("Enumerando abas:")
        page = selecionar_aba_por_titulo(browser, ALVO_TITULO_CONTEM)

        preencher_datas(page, override=DATE_OVERRIDE)
        selecionar_categoria_indice5(page)
        marcar_sou_procurador(page)
        selecionar_outorgantes_apenas_cnpj(page)
        clicar_filtrar_e_aguardar(page)

        nome_titular = extrair_nome_titular(page)
        basepath = os.path.dirname(os.path.abspath(__file__))

        carregar_excel_referencia()

        pagina = 1
        while True:
            log(f"== Página {pagina} ==")
            try:
                checar_sessao_expirada(page)

                if DO_PLANILHA:
                    dados_pagina = coletar_tabela_pagina(page)
                    salvar_em_excel_incremental(basepath, nome_titular, dados_pagina)

                if DO_RECIBOS:
                    baixar_recibo_linha(page, num_pagina_atual=pagina)

                if not ir_para_proxima_pagina_e_aguardar(page):
                    break
                pagina += 1

            except SessaoExpiradaException:
                log("Sessão Expirada no ECAC detectada.")
                print("\n*** Sessão Expirada no ECAC ***")
                print(f"A conferência travou na PÁGINA {pagina}.")
                input(
                    "Refaça o login/navegação no eCAC até a mesma tela e página,\n"
                    "depois pressione ENTER aqui para continuar a conferência..."
                )
                try:
                    wait_idle_app(page)
                except Exception:
                    pass
                wait_idle_app(page)
                continue

        log("Processo concluído.")
        # não fecha o Chrome do usuário (CDP attach); só desconecta
        browser.close()


if __name__ == "__main__":
    main()