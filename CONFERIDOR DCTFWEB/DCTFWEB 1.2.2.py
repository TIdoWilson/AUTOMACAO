# -*- coding: utf-8 -*-
"""
Versão 1.2.2 — modos de execução + escrita incremental no Excel + sem prints de debug.
Ajuste: em "Apenas planilhar", o script percorre TODAS as páginas (abas/paginação) até o fim.
"""
# (conteúdo completo abaixo, sem omissões)

import os
import re
import sys
import time
from datetime import date, datetime, timedelta
from calendar import monthrange
from pathlib import Path

import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, StaleElementReferenceException
)

# ------------------------ CONFIG ------------------------
ALVO_TITULO_CONTEM = "eCAC - Centro Virtual de Atendimento"
WIN_CLASS = "Chrome_WidgetWin_1"
WIN_PID = 2416  # ajuste se necessário

DOWNLOAD_ROOT = r"T:\\testes"

# Caminho fixo do Excel de referência (já salvo anteriormente)
EXCEL_REFERENCIA_PATH = (
    r"W:\DOCUMENTOS ESCRITORIO\INSTALACAO SISTEMA\python\CONFERIDOR DCTFWEB"
    r"\Consulta DCTFWEB WILSON MARCOS LOPES.xlsx"
)

# Colunas que usamos para comparar se a linha é "igual"
REF_KEY_COLS = [
    "Número de Identificação",
    "Período de Apuração",
    "Data Transmissão",
    "Situação",
]
REF_KEYS = set()  # conjunto de chaves já existentes no Excel de referência

class SessaoExpiradaException(Exception):
    """Levanta quando a página do eCAC indica 'Sessão Expirada'."""
    pass


SEL_IMG_PROCESSANDO = "img.image-processamento, img[class*='image-processamento']"
SEL_TABELA_LISTA = "table#ctl00_cphConteudo_tabelaListagemDctf_GridViewDctfs"
SEL_DROPD_CATEG_BTN = "button.btn.dropdown-toggle.btn-default[data-id='ctl00_cphConteudo_ddlCategoriaDeclaracao']"
SEL_DROPD_OUTORGANTES_BTN = "button.btn.dropdown-toggle.btn-default[data-id='ctl00_cphConteudo_ddlOutorgantes']"

T_ESP_CURTO = 0.10
T_ESP_MEDIO = 0.20
T_ESP_LONGO = 0.40
TIMEOUT_PADRAO = 20
# -------------------------------------------------------

# ------------------------ MODOS ------------------------
DO_PLANILHA = False
DO_RECIBOS = False
DATE_OVERRIDE = None  # dict: {'ap_ini','ap_fim','tx_ini','tx_fim'} como datetime.date

def _ask_mode_and_dates():
    global DO_PLANILHA, DO_RECIBOS, DATE_OVERRIDE

    print("Selecione o modo de execução:")
    print("  1) Apenas planilhar")
    print("  2) Baixar recibos")
    print("  3) Ambos (planilhar + recibos)")
    escolha = input("Digite 1, 2 ou 3 e pressione Enter: ").strip()

    if escolha not in {"1", "2", "3"}:
        print("Opção inválida. Encerrando.")
        sys.exit(1)

    if escolha == "1":
        DO_PLANILHA = True
        DO_RECIBOS = False
        di_str = input("Informe a Data Inicial (dd/mm/aaaa): ").strip()
        ap_ini = _parse_br_date(di_str)
        if ap_ini is None:
            print("Data inválida. Use o formato dd/mm/aaaa. Encerrando.")
            sys.exit(1)
        ap_fim = _last_day_of_month(ap_ini)
        tx_ini = ap_ini
        tx_fim = _last_day_of_next_month(ap_ini)
        DATE_OVERRIDE = {"ap_ini": ap_ini, "ap_fim": ap_fim, "tx_ini": tx_ini, "tx_fim": tx_fim}
        print("Datas definidas:")
        print(f"  - Período Apuração: {ap_ini.strftime('%d/%m/%Y')} a {ap_fim.strftime('%d/%m/%Y')}")
        print(f"  - Data Transmissão: {tx_ini.strftime('%d/%m/%Y')} a {tx_fim.strftime('%d/%m/%Y')}")
    elif escolha == "2":
        DO_PLANILHA = False
        DO_RECIBOS = True
        DATE_OVERRIDE = None
    else:
        DO_PLANILHA = True
        DO_RECIBOS = True
        DATE_OVERRIDE = None

def _parse_br_date(s):
    try:
        return datetime.strptime(s, "%d/%m/%Y").date()
    except Exception:
        return None

def _last_day_of_month(d: date) -> date:
    last = monthrange(d.year, d.month)[1]
    return d.replace(day=last)

def _last_day_of_next_month(d: date) -> date:
    if d.month == 12:
        first_next = date(d.year + 1, 1, 1)
    else:
        first_next = date(d.year, d.month + 1, 1)
    last = monthrange(first_next.year, first_next.month)[1]
    return first_next.replace(day=last)

# ------------------------ Logging ----------------------
def log(msg):
    print(msg, flush=True)

# ---------- Windows focus (opcional) ----------
def focar_janela():
    try:
        from pywinauto import Desktop
        app = Desktop(backend="uia")
        wins = [w for w in app.windows(class_name=WIN_CLASS) if w.process_id() == WIN_PID]
        alvo = wins[0] if wins else None
        if alvo:
            alvo.set_focus()
            try: alvo.restore()
            except Exception: pass
            log("Janela focada via pywinauto.")
        else:
            log("Janela não encontrada via pywinauto (prosseguindo).")
    except Exception as e:
        log(f"pywinauto indisponível ({e}); prosseguindo.")

# ---------- Chrome attach + modo rápido ----------
DEBUGGER_ADDRESS = os.environ.get("CHROME_DEBUG_ADDR", "localhost:9222")

def anexar_chrome_existente():
    opts = webdriver.ChromeOptions()
    opts.add_experimental_option("debuggerAddress", DEBUGGER_ADDRESS)
    try:
        driver = webdriver.Chrome(options=opts)
        return driver
    except Exception as e:
        msg = (
            f"\\nNão consegui conectar ao Chrome em {DEBUGGER_ADDRESS}.\\n"
            f"Verifique:\\n"
            f"  1) Inicie o Chrome assim (feche antes todas as janelas):\\n"
            f'     \"C:\\\\Program Files\\\\Google\\\\Chrome\\\\Application\\\\chrome.exe\" '
            f'--remote-debugging-port=9222 --user-data-dir=\"C:\\\\ChromeProfileDCTF\"\\n'
            f"  2) Confirme a porta: Test-NetConnection localhost -Port 9222 (deve ser True)\\n"
            f"  3) Firewall/antivírus: libere localhost:{DEBUGGER_ADDRESS.split(':')[-1]}\\n"
            f"  4) Não use outra instância do Chrome sem a flag (ele pode anexar na errada)\\n"
            f"Se usar outra porta, defina CHROME_DEBUG_ADDR (ex.: 'localhost:9223')."
        )
        raise RuntimeError(msg) from e

def sanity_check_tabs(driver):
    try:
        print("Abas abertas:")
        for h in driver.window_handles:
            driver.switch_to.window(h)
            print(" -", driver.title)
        driver.switch_to.window(driver.window_handles[-1])
    except Exception:
        pass

def esperar_domcarregado(driver, timeout=TIMEOUT_PADRAO):
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("return document.readyState") in ("interactive", "complete")
    )
    time.sleep(T_ESP_MEDIO)

def configurar_modo_rapido(driver, bloquear_recursos=False):
    try:
        driver.implicitly_wait(0)
    except Exception:
        pass
    try:
        driver.set_script_timeout(10)
    except Exception:
        pass
    try:
        driver.execute_cdp_cmd("Network.enable", {})
    except Exception as e:
        log(f"CDP (Network) indisponível: {e}")

# ---------- IFRAME cache ----------
FRAME_CACHE = {}

def reset_frame_cache():
    FRAME_CACHE.clear()
    log("FRAME_CACHE limpo.")

def prime_frame_cache(driver):
    chaves = [SEL_TABELA_LISTA, SEL_DROPD_OUTORGANTES_BTN, SEL_DROPD_CATEG_BTN]
    for css in chaves:
        try:
            find_fast(driver, By.CSS_SELECTOR, css, timeout=3)
        except Exception:
            pass
    log("FRAME_CACHE reaquecido (prime).")

def _switch_to_frame_path(driver, path):
    driver.switch_to.default_content()
    for idx in path:
        frs = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")
        if idx >= len(frs): raise TimeoutException("Caminho de frame inválido.")
        driver.switch_to.frame(frs[idx])

def _find_in_path(driver, by, selector, path):
    _switch_to_frame_path(driver, path)
    return WebDriverWait(driver, 1).until(EC.presence_of_element_located((by, selector)))

def find_fast(driver, by, selector, timeout=TIMEOUT_PADRAO, visible=False):
    end = time.time() + timeout
    last_err = None

    if selector in FRAME_CACHE:
        try:
            el = _find_in_path(driver, by, selector, FRAME_CACHE[selector])
            if visible: WebDriverWait(driver, 0.5).until(EC.visibility_of(el))
            return el, FRAME_CACHE[selector]
        except Exception as e:
            last_err = e
            FRAME_CACHE.pop(selector, None)
            driver.switch_to.default_content()

    while time.time() < end:
        try:
            driver.switch_to.default_content()
            el = WebDriverWait(driver, 0.3).until(EC.presence_of_element_located((by, selector)))
            if visible: WebDriverWait(driver, 0.5).until(EC.visibility_of(el))
            return el, []
        except Exception as e:
            last_err = e

        try:
            driver.switch_to.default_content()
            frames_lvl0 = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")
            for i0, fr0 in enumerate(frames_lvl0):
                try:
                    driver.switch_to.frame(fr0)
                    try:
                        el = WebDriverWait(driver, 0.3).until(EC.presence_of_element_located((by, selector)))
                        if visible: WebDriverWait(driver, 0.5).until(EC.visibility_of(el))
                        FRAME_CACHE[selector] = [i0]; return el, FRAME_CACHE[selector]
                    except Exception:
                        pass
                    subframes = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")
                    for i1, fr1 in enumerate(subframes):
                        try:
                            driver.switch_to.frame(fr1)
                            el = WebDriverWait(driver, 0.3).until(EC.presence_of_element_located((by, selector)))
                            if visible: WebDriverWait(driver, 0.5).until(EC.visibility_of(el))
                            FRAME_CACHE[selector] = [i0,i1]; return el, FRAME_CACHE[selector]
                        except Exception:
                            pass
                        finally:
                            driver.switch_to.parent_frame()
                finally:
                    driver.switch_to.default_content()
        except Exception as e:
            last_err = e

    raise TimeoutException(f"Não achei '{selector}' em {timeout}s (último erro: {last_err})")

def click_fast(driver, by, selector, timeout=TIMEOUT_PADRAO):
    el, path = find_fast(driver, by, selector, timeout=timeout, visible=True)
    _switch_to_frame_path(driver, path)
    try: el.click()
    except Exception: driver.execute_script("arguments[0].click();", el)
    driver.switch_to.default_content()

def wait_appear(driver, css, timeout=TIMEOUT_PADRAO):
    end = time.time() + timeout
    while time.time() < end:
        try:
            _, path = find_fast(driver, By.CSS_SELECTOR, css, timeout=0.6)
            _switch_to_frame_path(driver, path)
            el = driver.find_element(By.CSS_SELECTOR, css)
            if el.is_displayed():
                driver.switch_to.default_content(); return True
        except Exception:
            pass
        time.sleep(0.1)
    raise TimeoutException(f"Elemento não apareceu: {css}")

def wait_disappear(driver, css, timeout=TIMEOUT_PADRAO):
    end = time.time() + timeout
    while time.time() < end:
        try:
            _, path = find_fast(driver, By.CSS_SELECTOR, css, timeout=0.3)
            _switch_to_frame_path(driver, path)
            el = driver.find_element(By.CSS_SELECTOR, css)
            vis = el.is_displayed()
            driver.switch_to.default_content()
            if not vis: return True
        except Exception:
            return True
        time.sleep(0.1)
    raise TimeoutException(f"Elemento não sumiu: {css}")

# ======== IDLE TRACKER / WAITS ========
def _install_activity_hooks(driver):
    js = r"""
    (function(){
      if (window.__idleHooksInstalled) return;
      window.__idleHooksInstalled = true;
      var Open = XMLHttpRequest.prototype.open;
      var Send = XMLHttpRequest.prototype.send;
      window.__pendingXHR = 0;
      XMLHttpRequest.prototype.open = function(){ return Open.apply(this, arguments); };
      XMLHttpRequest.prototype.send = function(){
        window.__pendingXHR++;
        this.addEventListener('loadend', function(){ window.__pendingXHR--; }, {once:true});
        return Send.apply(this, arguments);
      };
      if (window.fetch){
        var _fetch = window.fetch;
        window.__pendingFetch = 0;
        window.fetch = function(){
          window.__pendingFetch++;
          try{
            return _fetch.apply(this, arguments).finally(function(){ window.__pendingFetch--; });
          }catch(e){
            window.__pendingFetch--;
            throw e;
          }
        };
      } else {
        window.__pendingFetch = 0;
      }
    })();
    """
    try: driver.execute_script(js)
    except Exception: pass

def _is_page_idle(driver):
    js = r"""
      const spinner = document.querySelector("img.image-processamento, img[class*='image-processamento']");
      const ready = document.readyState === 'complete';
      const jqIdle = (window.jQuery && typeof jQuery.active === 'number') ? (jQuery.active === 0) : true;
      const xhrIdle = (window.__pendingXHR||0) === 0;
      const fetchIdle = (window.__pendingFetch||0) === 0;
      return (!!ready) && (!spinner) && jqIdle && xhrIdle && fetchIdle;
    """
    try: return bool(driver.execute_script(js))
    except Exception: return False

def wait_until_idle(driver, quiet_ms=700, timeout=30):
    _install_activity_hooks(driver)
    end = time.time() + timeout
    stable_since = None
    while time.time() < end:
        ok = _is_page_idle(driver)
        now = time.time()
        if ok:
            stable_since = stable_since or now
            if (now - stable_since)*1000 >= quiet_ms:
                return True
        else:
            stable_since = None
        time.sleep(0.08)
    raise TimeoutException("Página não ficou idle a tempo.")

# ---------- Auxiliares de grid/linha ----------
def _grid_row_count(driver):
    tbl, path = find_fast(driver, By.CSS_SELECTOR, SEL_TABELA_LISTA, timeout=TIMEOUT_PADRAO, visible=True)
    _switch_to_frame_path(driver, path)
    try:
        rows = tbl.find_elements(By.CSS_SELECTOR, "tr")
        return max(len(rows)-1, 0), path
    finally:
        driver.switch_to.default_content()

def wait_grid_ready(driver, timeout=30):
    end = time.time() + timeout
    last_count, stable_ticks = None, 0
    while time.time() < end:
        wait_until_idle(driver, quiet_ms=300, timeout=timeout)
        try:
            count, _ = _grid_row_count(driver)
        except Exception:
            time.sleep(0.1); continue
        if last_count is None or count != last_count:
            last_count, stable_ticks = count, 0
        else:
            stable_ticks += 1
            if stable_ticks >= 5:
                return True
        time.sleep(0.1)
    raise TimeoutException("Grid não estabilizou a tempo.")

def _data_rows_xpath():
    return ("//table[@id='ctl00_cphConteudo_tabelaListagemDctf_GridViewDctfs']"
            "//tbody/tr[not(.//img[contains(@id,'imgNextPage')]) "
            "and not(.//a[contains(@id,'lnkPagina')]) "
            "and not(contains(@class,'paginacao')) "
            "and count(td)>=8]")

def _data_row_count_and_path(driver):
    tbl, path = find_fast(driver, By.CSS_SELECTOR, SEL_TABELA_LISTA, timeout=TIMEOUT_PADRAO, visible=True)
    _switch_to_frame_path(driver, path)
    try:
        rows = driver.find_elements(By.XPATH, _data_rows_xpath())
        return len(rows), path
    finally:
        driver.switch_to.default_content()

def _get_data_row(driver, path_tbl, k):
    _switch_to_frame_path(driver, path_tbl)
    rows = driver.find_elements(By.XPATH, _data_rows_xpath())
    if k < 1 or k > len(rows):
        driver.switch_to.default_content()
        raise IndexError(f"linha {k} inexistente (tem {len(rows)})")
    tr = driver.find_element(By.XPATH, f"({_data_rows_xpath()})[{k}]")
    return tr

def wait_row_services_ready(driver, path_tbl, row_index, timeout=8):
    end = time.time() + timeout
    hasIcon = False
    while time.time() < end:
        try:
            tr = _get_data_row(driver, path_tbl, row_index)
            hasSpinner = bool(tr.find_elements(By.CSS_SELECTOR, "img.image-processamento, img[class*='image-processamento']"))
            aguardeTxt = ("aguarde" in (tr.text or "").lower())
            hasIcon = bool(tr.find_elements(By.CSS_SELECTOR,
                        "a[title='Visualizar Recibo'], a.image-tabela-visualizar-recibo, a[id*='lkbVisualizarRecibo']"))
            driver.switch_to.default_content()
            if not hasSpinner and not aguardeTxt:
                return hasIcon
        except IndexError:
            driver.switch_to.default_content()
            return False
        except Exception:
            driver.switch_to.default_content()
        time.sleep(0.12)
    return hasIcon

def _exec_postback_js(driver, target, argument):
    return driver.execute_script(
        "if (typeof __doPostBack==='function'){ __doPostBack(arguments[0], arguments[1]||''); return true; } return false;",
        target, argument or ""
    )

def _form_postback_fallback(driver, target, argument):
    js = """
    var f = document.forms && document.forms[0];
    if(!f) return false;
    var t = document.getElementById('__EVENTTARGET') || f.querySelector('input[name="__EVENTTARGET"]');
    var a = document.getElementById('__EVENTARGUMENT') || f.querySelector('input[name="__EVENTARGUMENT"]');
    if(!t || !a) return false;
    t.value = arguments[0];
    a.value = arguments[1] || '';
    if (typeof f.submit === 'function') { f.submit(); return true; }
    var ev = document.createEvent('Event'); ev.initEvent('submit', true, true); return f.dispatchEvent(ev);
    """
    return bool(driver.execute_script(js, target, argument or ""))

def executar_postback_scoped(driver, frame_path, target, argument=""):
    try:
        if _exec_postback_js(driver, target, argument):
            return True
    except Exception:
        pass
    try:
        _switch_to_frame_path(driver, frame_path or [])
        ok = _exec_postback_js(driver, target, argument)
        driver.switch_to.default_content()
        if ok:
            return True
    except Exception:
        driver.switch_to.default_content()
    try:
        _switch_to_frame_path(driver, frame_path or [])
        ok = _form_postback_fallback(driver, target, argument)
        driver.switch_to.default_content()
        if ok:
            return True
    except Exception:
        driver.switch_to.default_content()
    return False

# ---------- Ações ----------
def selecionar_aba_por_titulo(driver, alvo_contem):
    log("Enumerando abas:")
    escolhido = None
    for h in driver.window_handles:
        driver.switch_to.window(h)
        title = (driver.title or "").strip()
        log(f"- '{title}'")
        if alvo_contem.lower() in title.lower(): escolhido = h
    if escolhido:
        driver.switch_to.window(escolhido)
        log(f"Aba selecionada: '{driver.title}'")
    else:
        log("Aba alvo não encontrada; seguindo na atual.")

def datas_mes_anterior_e_hoje():
    hoje = date.today()
    primeiro_deste = hoje.replace(day=1)
    ultimo_mes = primeiro_deste - timedelta(days=1)
    return ultimo_mes.replace(day=1), ultimo_mes, hoje

def _ensure_dir(p):
    Path(p).mkdir(parents=True, exist_ok=True)
    return str(Path(p))

def _current_files(path_dir):
    try: return {f for f in os.listdir(path_dir) if not f.endswith(".crdownload")}
    except FileNotFoundError: return set()

def _set_download_dir(driver, path_dir):
    try:
        driver.execute_cdp_cmd("Page.setDownloadBehavior", {"behavior":"allow","downloadPath":path_dir})
    except Exception:
        try:
            driver.execute_cdp_cmd("Browser.setDownloadBehavior", {"behavior":"allow","downloadPath":path_dir})
        except Exception as e:
            log(f"  ! setDownloadBehavior falhou: {e}")

def _wait_new_file(path_dir, before_set, timeout=120):
    end = time.time() + timeout
    while time.time() < end:
        now = _current_files(path_dir)
        new = {f for f in (now - before_set) if not f.endswith(".tmp") and not f.endswith(".partial")}
        if new:
            if not any(n.endswith(".crdownload") for n in os.listdir(path_dir)):
                return list(new)[0]
        time.sleep(0.25)
    raise TimeoutException(f"Timeout esperando download em {path_dir}")

def executar_postback(driver, event_target, event_argument=""):
    driver.switch_to.default_content()
    js = """
      if (typeof __doPostBack === 'function') { __doPostBack(arguments[0], arguments[1]||''); return true; }
      return false;
    """
    ok = driver.execute_script(js, event_target, event_argument)
    if not ok: raise RuntimeError("__doPostBack não disponível.")

def _set_input_value(driver, el, value):
    try:
        el.clear()
    except Exception:
        pass
    try:
        el.send_keys(value)
    except Exception:
        driver.execute_script("arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('change'));", el, value)

def _find_input_by_label_text(driver, label_contains: str, timeout=10):
    txt = label_contains.lower()
    xpaths = [
        f"//label[contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚÂÊÔÃÕÇ','abcdefghijklmnopqrstuvwxyzáéíóúâêôãõç'), '{txt}')]/following::input[1]",
        f"//*[self::span or self::strong][contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚÂÊÔÃÕÇ','abcdefghijklmnopqrstuvwxyzáéíóúâêôãõç'), '{txt}')]/following::input[1]",
        f"//td[contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚÂÊÔÃÕÇ','abcdefghijklmnopqrstuvwxyzáéíóúâêôãõç'), '{txt}')]/following-sibling::td//input[1]",
    ]
    last = None
    for xp in xpaths:
        try:
            el, path = find_fast(driver, By.XPATH, xp, timeout=timeout, visible=True)
            return el, path
        except Exception as e:
            last = e
    raise TimeoutException(f"Campo com rótulo '{label_contains}' não encontrado (último erro: {last})")

def preencher_datas(driver, override=None):
    if override:
        d_ap_ini = override["ap_ini"]
        d_ap_fim = override["ap_fim"]
        d_tx_ini = override["tx_ini"]
        d_tx_fim = override["tx_fim"]
    else:
        d_ap_ini, d_ap_fim, d_tx_fim = datas_mes_anterior_e_hoje()
        d_tx_ini = d_ap_ini
    dd = lambda d: d.strftime("%d/%m/%Y")

    log("Preenchendo datas…")
    alvos = [
        ("Periodo Apuração Inicial",
         ["txtDataInicio", "txtPeriodoApuracaoInicial", "txtPeriodoApuracaoInicio"],
         "Período Apuração Inicial", dd(d_ap_ini)),
        ("Periodo Apuração Final",
         ["txtDataFinal", "txtPeriodoApuracaoFinal", "txtPeriodoApuracaoFim"],
         "Período Apuração Final", dd(d_ap_fim)),
        ("Data Transmissão Inicial",
         ["txtDataTransmissaoInicial", "txtDtTransmissaoInicial"],
         "Data Transmissão Inicial", dd(d_tx_ini)),
        ("Data Transmissão Final",
         ["txtDataTransmissaoFinal", "txtDtTransmissaoFinal"],
         "Data Transmissão Final", dd(d_tx_fim)),
    ]

    for nome, ids, rotulo, valor in alvos:
        ok = False
        for cid in ids:
            try:
                el, path = find_fast(driver, By.CSS_SELECTOR, f"input[id='{cid}'], input[id*='{cid}']", timeout=3)
                _switch_to_frame_path(driver, path)
                _set_input_value(driver, el, valor)
                driver.switch_to.default_content()
                log(f"  - {nome} = {valor}")
                ok = True
                break
            except Exception:
                driver.switch_to.default_content()
        if not ok:
            try:
                el, path = _find_input_by_label_text(driver, rotulo, timeout=6)
                _switch_to_frame_path(driver, path)
                _set_input_value(driver, el, valor)
                driver.switch_to.default_content()
                log(f"  - {nome} (label) = {valor}")
                ok = True
            except Exception as e:
                driver.switch_to.default_content()
                log(f"  ! {nome}: não consegui preencher ({e})")

def selecionar_categoria_indice5(driver):
    log("Selecionando Categoria índice 5…")
    try:
        click_fast(driver, By.CSS_SELECTOR, SEL_DROPD_CATEG_BTN)
        try:
            ul, path = find_fast(driver, By.CSS_SELECTOR, "div.open ul.dropdown-menu.inner", timeout=6)
            _switch_to_frame_path(driver, path)
            try:
                driver.find_element(By.CSS_SELECTOR, "button.actions-btn.bs-deselect-all").click()
                time.sleep(0.08)
            except Exception:
                pass
            try:
                driver.find_element(By.CSS_SELECTOR, "div.open ul.dropdown-menu.inner li[data-original-index='4'] a").click()
            except Exception:
                driver.find_element(By.XPATH, "//div_contains(@class,'open')]//li/a/span[contains(., 'Geral')]/..").click()
        finally:
            driver.switch_to.default_content()
            try:
                click_fast(driver, By.CSS_SELECTOR, SEL_DROPD_CATEG_BTN)
            except Exception:
                pass
        log("  - Categoria marcada via bootstrap-select.")
        return
    except Exception:
        driver.switch_to.default_content()
    try:
        sel, path = find_fast(driver, By.CSS_SELECTOR, "select[id*='ddlCategoriaDeclaracao']", timeout=8, visible=True)
        _switch_to_frame_path(driver, path)
        try:
            Select(sel).select_by_index(5)
        except Exception:
            driver.execute_script(
                "arguments[0].selectedIndex = 5; arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", sel
            )
        driver.switch_to.default_content()
        log("  - Categoria marcada no <select> nativo (fallback).")
    except Exception as e:
        driver.switch_to.default_content()
        log(f"  ! Não consegui marcar categoria índice 5: {e}")

def marcar_sou_procurador(driver):
    log("Marcando 'Sou Procurador'…")
    seletor_cb = ("input[type='checkbox'][id*='ListarOutorgantes'], "
                  "input[type='checkbox'][id*='chkListarOutorgantes'], "
                  "input[type='checkbox'][id='ctl00_cphConteudo_chkListarOutorgantes']")
    try:
        el, path = find_fast(driver, By.CSS_SELECTOR, seletor_cb, timeout=20)
    except TimeoutException:
        el, path = find_fast(
            driver, By.XPATH,
            "//label[contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚÂÊÔÃÕÇ','abcdefghijklmnopqrstuvwxyzáéíóúâêôãõç'), 'sou procurador')]",
            timeout=20
        )
    _switch_to_frame_path(driver, path)
    try:
        try:
            el.click()
        except Exception:
            driver.execute_script("arguments[0].click();", el)
    finally:
        driver.switch_to.default_content()

    try: wait_appear(driver, SEL_IMG_PROCESSANDO, timeout=2)
    except Exception: pass
    try: wait_disappear(driver, SEL_IMG_PROCESSANDO, timeout=TIMEOUT_PADRAO)
    except Exception: pass
    wait_until_idle(driver, quiet_ms=500, timeout=30)

    find_fast(driver, By.CSS_SELECTOR, SEL_DROPD_OUTORGANTES_BTN, timeout=TIMEOUT_PADRAO, visible=True)

    reset_frame_cache()
    prime_frame_cache(driver)

def selecionar_outorgantes_apenas_cnpj(driver):
    log("Outorgantes: filtrar 'CNPJ' e selecionar 'Todas'…")
    click_fast(driver, By.CSS_SELECTOR, SEL_DROPD_OUTORGANTES_BTN)
    try:
        ul, path = find_fast(driver, By.CSS_SELECTOR, "div.open ul.dropdown-menu.inner", timeout=10)
        _switch_to_frame_path(driver, path)
        try:
            btn_none = driver.find_element(By.CSS_SELECTOR, "button.actions-btn.bs-deselect-all")
            btn_none.click(); time.sleep(0.10)
        except Exception:
            pass
        usou_busca = False
        try:
            search_input = driver.find_element(By.CSS_SELECTOR, "div.dropdown-menu .bs-searchbox input")
            search_input.clear(); time.sleep(0.05)
            search_input.send_keys("CNPJ"); usou_busca = True
            end = time.time() + 5
            while time.time() < end:
                vis = driver.find_elements(By.CSS_SELECTOR,
                    "div.open ul.dropdown-menu.inner li:not(.hidden):not(.disabled):not(.divider):not(.dropdown-header)")
                if vis: break
                time.sleep(0.05)
            try:
                btn_all = driver.find_element(By.CSS_SELECTOR, "button.actions-btn.bs-select-all")
                btn_all.click(); time.sleep(0.10)
            except Exception:
                for a in driver.find_elements(By.CSS_SELECTOR,
                    "div.open ul.dropdown-menu.inner li:not(.hidden):not(.disabled):not(.divider):not(.dropdown-header) a"):
                    try: a.click(); time.sleep(0.02)
                    except Exception: pass
        except Exception:
            usou_busca = False
        if not usou_busca:
            try:
                aba_cnpj = driver.find_element(By.XPATH, "//div[contains(@class,'open')]//a[normalize-space()='CNPJ' or contains(.,'CNPJ')]")
                aba_cnpj.click(); time.sleep(0.10)
            except Exception:
                pass
            try:
                btn_all = driver.find_element(By.CSS_SELECTOR, "button.actions-btn.bs-select-all")
                btn_all.click(); time.sleep(0.10)
            except Exception:
                for a in driver.find_elements(By.CSS_SELECTOR,
                    "div.open ul.dropdown-menu.inner li:not(.disabled):not(.divider):not(.dropdown-header) a"):
                    lbl = ((a.text or "") + " " + (a.get_attribute("title") or a.get_attribute("data-subtext") or "")).lower()
                    if "cnpj" in lbl:
                        try: a.click(); time.sleep(0.02)
                        except Exception: pass
    finally:
        driver.switch_to.default_content()
        click_fast(driver, By.CSS_SELECTOR, SEL_DROPD_OUTORGANTES_BTN)

    wait_until_idle(driver, quiet_ms=300, timeout=30)
    reset_frame_cache()
    prime_frame_cache(driver)

def clicar_filtrar_e_aguardar(driver):
    log("Filtrar (postback direto)…")
    try:
        executar_postback(driver, "ctl00$cphConteudo$imageFiltrar", "")
    except Exception:
        try:
            click_fast(driver, By.CSS_SELECTOR, "a#ctl00_cphConteudo_imageFiltrar, a[id*='imageFiltrar']")
        except Exception:
            click_fast(driver, By.CSS_SELECTOR, "a.btn.btn-icon-sec.btn-exibe-processando.image-pesquisar")
    try: wait_appear(driver, SEL_IMG_PROCESSANDO, timeout=2)
    except Exception: pass
    try: wait_disappear(driver, SEL_IMG_PROCESSANDO, timeout=TIMEOUT_PADRAO)
    except Exception: pass
    wait_until_idle(driver, quiet_ms=500, timeout=30)
    wait_grid_ready(driver, timeout=30)
    reset_frame_cache()
    prime_frame_cache(driver)

def checar_sessao_expirada(driver):
    """
    Verifica se a página atual do eCAC mostra 'Sessão Expirada'.
    Se sim, lança SessaoExpiradaException.
    """
    try:
        html = driver.page_source or ""
    except Exception:
        return
    if "Sessão Expirada" in html or "Sessao Expirada" in html or "sessao-expirada" in html:
        raise SessaoExpiradaException("Sessão Expirada no eCAC detectada.")


def _extrair_dados_de_tr(tr):
    """
    Recebe um <tr> da tabela de listagem e devolve um dict com as colunas
    usadas tanto para Excel quanto para comparação.
    """
    tds = tr.find_elements(By.CSS_SELECTOR, "td")
    if len(tds) < 9:
        return {}

    gt = lambda el: (el.text or "").strip()

    tipo_ni = gt(tds[0])
    numero_id = gt(tds[1])
    periodo_apur = gt(tds[2])
    dt_transm = gt(tds[3])
    categoria = gt(tds[4])
    origem = gt(tds[5])
    tipo = gt(tds[6])

    try:
        situacao = tr.find_element(By.CSS_SELECTOR, ".coluna-lblSituacao span").text.strip()
    except Exception:
        situacao = gt(tds[7])

    valor_apurado = gt(tds[8])
    try:
        saldo_pagar = tr.find_element(By.CSS_SELECTOR, "span[id*='lblSaldoPagar']").text.strip()
    except Exception:
        saldo_pagar = gt(tds[9]) if len(tds) > 9 else ""

    return {
        "Tipo NI": tipo_ni,
        "Número de Identificação": numero_id,
        "Período de Apuração": periodo_apur,
        "Data Transmissão": dt_transm,
        "Categoria": categoria,
        "Origem": origem,
        "Tipo": tipo,
        "Situação": situacao,
        "Valor Apurado": valor_apurado,
        "Saldo a Pagar": saldo_pagar,
    }


def coletar_tabela_pagina(driver):
    wait_grid_ready(driver, timeout=30)
    log("Coletando linhas da página…")
    dados = []
    tbl, path = find_fast(driver, By.CSS_SELECTOR, SEL_TABELA_LISTA, timeout=30, visible=True)
    _switch_to_frame_path(driver, path)
    try:
        linhas = tbl.find_elements(By.CSS_SELECTOR, "tr")[1:]
        for tr in linhas:
            try:
                linha_dict = _extrair_dados_de_tr(tr)
                if linha_dict:
                    dados.append(linha_dict)
            except StaleElementReferenceException:
                continue
    finally:
        driver.switch_to.default_content()
    log(f"Linhas coletadas: {len(dados)}")
    return dados


def carregar_excel_referencia():
    """
    Lê o Excel de referência já salvo (WILSON MARCOS LOPES) e
    preenche REF_KEYS com as chaves (Número + Período + Data + Situação).
    """
    global REF_KEYS

    REF_KEYS = set()
    if not os.path.exists(EXCEL_REFERENCIA_PATH):
        log(f"Excel de referência NÃO encontrado: {EXCEL_REFERENCIA_PATH}")
        log("Nenhuma linha será pulada com base nesse arquivo.")
        return

    try:
        df_ref = pd.read_excel(EXCEL_REFERENCIA_PATH)
    except Exception as e:
        log(f"Erro ao ler Excel de referência: {e}")
        log("Nenhuma linha será pulada com base nesse arquivo.")
        return

    faltando = [c for c in REF_KEY_COLS if c not in df_ref.columns]
    if faltando:
        log(f"Excel de referência não contém as colunas esperadas: {faltando}")
        log("Nenhuma linha será pulada com base nesse arquivo.")
        return

    for _, row in df_ref.iterrows():
        chave = tuple(str(row[c]).strip() for c in REF_KEY_COLS)
        REF_KEYS.add(chave)

    log(f"Excel de referência carregado. Linhas cadastradas: {len(REF_KEYS)}")


def linha_ja_processada(dados_linha: dict) -> bool:
    """
    Verifica se a linha (dict) já está presente no Excel de referência,
    usando as colunas definidas em REF_KEY_COLS.
    """
    if not REF_KEYS:
        return False
    chave = tuple(str(dados_linha.get(c, "")).strip() for c in REF_KEY_COLS)
    return chave in REF_KEYS


# ============ Escrita incremental no Excel por página ============
def salvar_em_excel_incremental(basepath, nome_titular, dados_pagina):
    if not dados_pagina: 
        return None
    caminho = Path(basepath) / f"Consulta DCTFWEB {nome_titular or 'TITULAR'}.xlsx"
    df = pd.DataFrame(dados_pagina)
    for col in ["Valor Apurado", "Saldo a Pagar"]:
        if col in df.columns:
            df[col] = (df[col].astype(str).str.replace(".","",regex=False)
                                   .str.replace(",",".",regex=False)
                                   .apply(lambda x: pd.to_numeric(x, errors="coerce")))
    modo = "a" if caminho.exists() else "w"
    if caminho.exists():
        with pd.ExcelWriter(caminho, engine="openpyxl", mode="a", if_sheet_exists="overlay") as w:
            sheet = "DCTFWeb"
            from openpyxl import load_workbook
            wb = load_workbook(caminho)
            startrow = wb[sheet].max_row if sheet in wb.sheetnames else 0
            df.to_excel(w, index=False, header=(startrow == 0), startrow=startrow, sheet_name=sheet)
    else:
        with pd.ExcelWriter(caminho, engine="openpyxl", mode="w") as w:
            df.to_excel(w, index=False, sheet_name="DCTFWeb")

        sheet = "DCTFWeb"
        if modo == "w":
            df.to_excel(w, index=False, sheet_name=sheet)
        else:
            from openpyxl import load_workbook
            wb = load_workbook(caminho)
            startrow = wb[sheet].max_row if sheet in wb.sheetnames else 0
            df.to_excel(w, index=False, header=(startrow==0), startrow=startrow, sheet_name=sheet)
    log(f"Excel (incremental) salvo/atualizado em: {caminho}")
    return str(caminho)

def baixar_recibos_pagina(driver, num_pagina_atual=None):
    log("Baixando recibos da página (robusto por XPath)…")
    checar_sessao_expirada(driver)
    wait_grid_ready(driver, timeout=30)
    total_rows, path_tbl = _data_row_count_and_path(driver)
    baixados = 0
    for idx in range(1, total_rows + 1):
        try:
            hasIcon = wait_row_services_ready(driver, path_tbl, idx, timeout=5)

            # Pegar a linha completa para montar o dict de dados
            tr = _get_data_row(driver, path_tbl, idx)
            dados_linha = _extrair_dados_de_tr(tr)

            # Se a linha já existe no Excel de referência, pula tudo
            if linha_ja_processada(dados_linha):
                log(
                    f"  - Página {num_pagina_atual or '?'} Linha {idx}: "
                    "já está idêntica no Excel de referência; pulando download do recibo."
                )
                driver.switch_to.default_content()
                continue

            destino = _ensure_dir(DOWNLOAD_ROOT)
            _set_download_dir(driver, destino)
            before = _current_files(destino)

            if not hasIcon:
                log(f"  - Linha {idx}: sem ícone de recibo (pulando).")
                driver.switch_to.default_content()
                continue

            # A partir daqui, reaproveita o 'tr' para achar o link de recibo
            lnk = None
            for sel in (
                "a[title='Visualizar Recibo']",
                "a.image-tabela-visualizar-recibo",
                "a[id*='lkbVisualizarRecibo']",
            ):
                try:
                    lnk = tr.find_element(By.CSS_SELECTOR, sel)
                    break
                except NoSuchElementException:
                    lnk = None

            href = lnk.get_attribute("href") if lnk else ""
            driver.switch_to.default_content()

            m = re.search(r"__doPostBack\('([^']+)'\,'([^']*)'\)", href or "")
            disparou = False
            if m:
                disparou = executar_postback_scoped(
                    driver, path_tbl, m.group(1), (m.group(2) or "")
                )
            if not disparou:
                tr = _get_data_row(driver, path_tbl, idx)
                try:
                    lnk = None
                    for sel in (
                        "a[title='Visualizar Recibo']",
                        "a.image-tabela-visualizar-recibo",
                        "a[id*='lkbVisualizarRecibo']",
                    ):
                        try:
                            lnk = tr.find_element(By.CSS_SELECTOR, sel)
                            break
                        except NoSuchElementException:
                            lnk = None
                    if lnk is None:
                        raise NoSuchElementException(
                            "Link de recibo sumiu no fallback de clique."
                        )
                    try:
                        lnk.click()
                    except Exception:
                        driver.execute_script("arguments[0].click();", lnk)
                finally:
                    driver.switch_to.default_content()

            try:
                wait_appear(driver, SEL_IMG_PROCESSANDO, timeout=1.2)
            except Exception:
                pass
            try:
                wait_disappear(driver, SEL_IMG_PROCESSANDO, timeout=TIMEOUT_PADRAO)
            except Exception:
                pass
            wait_until_idle(driver, quiet_ms=400, timeout=3)

            novo = _wait_new_file(destino, before, timeout=20)
            baixados += 1
            log(
                f"  - (pág {num_pagina_atual or '?'} linha {idx}/{total_rows}) "
                f"recibo salvo: {os.path.join(destino, novo)}"
            )

            checar_sessao_expirada(driver)

        except IndexError:
            driver.switch_to.default_content()
            log(
                f"  - Linha {idx}: não existe (grid mudou). Recalculando contagem…"
            )
            total_rows, path_tbl = _data_row_count_and_path(driver)
            continue
        except SessaoExpiradaException:
            # Propaga para o main tratar
            driver.switch_to.default_content()
            raise
        except Exception as e:
            driver.switch_to.default_content()
            try:
                checar_sessao_expirada(driver)
            except SessaoExpiradaException:
                raise
            log(f"  ! Linha {idx}: {e}")
            continue

    log(f"Recibos baixados nesta página: {baixados}")

def ir_para_proxima_pagina_e_aguardar(driver):
    try:
        btn, path = find_fast(driver, By.CSS_SELECTOR, "img[id*='imgNextPage']", timeout=5)
        _switch_to_frame_path(driver, path)
        a = btn.find_element(By.XPATH, "./ancestor::a[1]")
        disabled = "disabled" in (a.get_attribute("class") or "").lower()
        if disabled:
            driver.switch_to.default_content()
            return False
        try:
            a.click()
        except Exception:
            driver.execute_script("arguments[0].click();", a)
        driver.switch_to.default_content()
    except TimeoutException:
        return False

    esperar_domcarregado(driver)
    wait_until_idle(driver, quiet_ms=500, timeout=3)
    wait_grid_ready(driver, timeout=30)
    reset_frame_cache()
    prime_frame_cache(driver)
    return True

def extrair_nome_titular(driver):
    try:
        el, path = find_fast(driver, By.CSS_SELECTOR, "#informacao-perfil", timeout=5)
        _switch_to_frame_path(driver, path)
        txt = (el.text or "").strip()
        driver.switch_to.default_content()
        m = re.search(r"Titular\s*\(.*?\)\s*[\d\.\-\/]+\s*-\s*(.+)$", txt, flags=re.IGNORECASE|re.MULTILINE)
        if m: return m.group(1).strip()
        if " - " in txt: return txt.split(" - ", 1)[-1].strip()
    except Exception:
        pass
    return "TITULAR"

# ------------------------ MAIN ------------------------
def main():
    _ask_mode_and_dates()

    focar_janela()
    driver = anexar_chrome_existente()
    sanity_check_tabs(driver)
    configurar_modo_rapido(driver)
    log("Anexado ao Chrome existente.")
    esperar_domcarregado(driver)
    wait_until_idle(driver, quiet_ms=400, timeout=30)
    reset_frame_cache()
    prime_frame_cache(driver)
    selecionar_aba_por_titulo(driver, ALVO_TITULO_CONTEM)

    preencher_datas(driver, override=DATE_OVERRIDE)
    selecionar_categoria_indice5(driver)
    marcar_sou_procurador(driver)
    selecionar_outorgantes_apenas_cnpj(driver)
    clicar_filtrar_e_aguardar(driver)

    nome_titular = extrair_nome_titular(driver)
    basepath = os.path.dirname(os.path.abspath(__file__))
    # Carrega o Excel de referência (para pular linhas já processadas)
    carregar_excel_referencia()


    # ---- Loop de paginação (sempre avança até o fim) ----
    pagina = 1
    while True:
        log(f"== Página {pagina} ==")
        try:
            checar_sessao_expirada(driver)

            if DO_PLANILHA:
                dados_pagina = coletar_tabela_pagina(driver)
                salvar_em_excel_incremental(basepath, nome_titular, dados_pagina)

            if DO_RECIBOS:
                baixar_recibos_pagina(driver, num_pagina_atual=pagina)

            # Tenta ir para a próxima página
            if not ir_para_proxima_pagina_e_aguardar(driver):
                break
            pagina += 1

        except SessaoExpiradaException:
            log("Sessão Expirada no ECAC detectada.")
            print("\n*** Sessão Expirada no ECAC ***")
            print(f"A conferência travou na PÁGINA {pagina}.")
            input(
                "Refaça o login/navegação no eCAC até a mesma tela e página,\n"
                "depois pressione ENTER aqui para continuar a conferência..."
            )

            # Depois que o usuário religar a sessão no navegador, normaliza de novo
            esperar_domcarregado(driver)
            wait_until_idle(driver, quiet_ms=400, timeout=60)
            reset_frame_cache()
            prime_frame_cache(driver)

            # NÃO altera 'pagina' e NÃO chama próxima página: volta para o topo do while
            continue

    log("Processo concluído.")


if __name__ == "__main__":
    main()
