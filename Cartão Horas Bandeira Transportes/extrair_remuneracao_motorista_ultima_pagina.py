from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

# =========================
# Configuracoes
# =========================
PASTA_BASE = Path(__file__).resolve().parent
PASTA_EXEMPLO = PASTA_BASE / "Arquivos de exemplo"
PASTA_GERADOS = PASTA_BASE / "Arquivos gerados"
ARQUIVO_LISTA_FUNCIONARIOS = PASTA_BASE / "Lista funcionários.xlsx"

CAMPOS_BLOCO_PRINCIPAL = [
    "valor_frete_sem_impostos",
    "combustivel",
    "faturamento_liquido",
    "comissao_motorista",
    "dsr",
    "diarias_refeicoes",
    "horas_extras",
    "total_remuneracao_mensal",
]

ROTULO_CAMPO = {
    "valor_frete_sem_impostos": "(+) Valor do Frete sem impostos",
    "combustivel": "(-) Combustivel",
    "faturamento_liquido": "(=) Faturamento LIQUIDO",
    "comissao_motorista": "Comissao do Motorista",
    "dsr": "Descanso Semanal Remunerado (DSR)",
    "diarias_refeicoes": "Diarias Refeicoes",
    "horas_extras": "Horas Extras",
    "total_remuneracao_mensal": "TOTAL DE REMUNERACAO MENSAL DO MOTORISTA",
    "salario": "Salario",
    "estadia_tempo_espera": "Estadia/Tempo de espera",
    "premio_disco_tacografo": "Premio disco tacografo",
}

EVENTO_POR_CAMPO = {
    "valor_frete_sem_impostos": 901,
    "combustivel": 902,
    "faturamento_liquido": 903,
    "comissao_motorista": 904,
    "dsr": 905,
    "diarias_refeicoes": 906,
    "horas_extras": 907,
    "total_remuneracao_mensal": 908,
    "salario": 909,
    "estadia_tempo_espera": 910,
    "premio_disco_tacografo": 911,
}

PADRAO_VALOR_BR = re.compile(r"^-?\d{1,3}(?:\.\d{3})*,\d{2}$|^-?\d+,\d{2}$")
PADRAO_PERIODO = re.compile(r"Per[ií]odo:\s*(\d{2}/\d{2}/\d{4})\s*a\s*(\d{2}/\d{2}/\d{4})", re.IGNORECASE)
PADRAO_CPF = re.compile(r"CPF:\s*(\d{11})", re.IGNORECASE)
PADRAO_MATRICULA = re.compile(r"Matr[ií]cula:\s*([0-9]+)", re.IGNORECASE)


@dataclass
class ResultadoExtracao:
    empresa: str
    periodo_inicial: str
    periodo_final: str
    empregado: str
    cpf: str
    matricula: str
    origem_matricula: str
    matricula_foi_heuristica: bool
    matricula_candidatas: list[str]
    valores: dict[str, str]


def normalizar_texto(texto: str) -> str:
    return texto.replace("\xa0", " ").replace("\r", "")


def listar_linhas(texto: str) -> list[str]:
    return [linha.strip() for linha in texto.splitlines() if linha.strip()]


def extrair_ultima_pagina(pdf_path: Path) -> str:
    reader = PdfReader(str(pdf_path))
    if not reader.pages:
        raise ValueError("PDF sem paginas.")
    return normalizar_texto(reader.pages[-1].extract_text() or "")


def eh_valor_br(valor: str) -> bool:
    return bool(PADRAO_VALOR_BR.match(valor or ""))


def decimal_para_br(valor: Decimal) -> str:
    texto = f"{valor:,.2f}"
    return texto.replace(",", "#").replace(".", ",").replace("#", ".")


def valor_br_para_decimal(valor: str) -> Decimal:
    bruto = (valor or "").strip().replace(".", "").replace(",", ".")
    try:
        return Decimal(bruto)
    except InvalidOperation as exc:
        raise ValueError(f"Valor monetario invalido: {valor}") from exc


def formatar_codigo_iob(valor: str) -> str:
    digitos = re.sub(r"\D", "", valor or "")
    return digitos.zfill(5)[-5:] if digitos else "00000"


def normalizar_nome(valor: str) -> str:
    base = (valor or "").upper().strip()
    substituicoes = str.maketrans(
        {
            "Á": "A",
            "À": "A",
            "Â": "A",
            "Ã": "A",
            "Ä": "A",
            "É": "E",
            "È": "E",
            "Ê": "E",
            "Ë": "E",
            "Í": "I",
            "Ì": "I",
            "Î": "I",
            "Ï": "I",
            "Ó": "O",
            "Ò": "O",
            "Ô": "O",
            "Õ": "O",
            "Ö": "O",
            "Ú": "U",
            "Ù": "U",
            "Û": "U",
            "Ü": "U",
            "Ç": "C",
        }
    )
    base = base.translate(substituicoes)
    base = re.sub(r"\s+", " ", base)
    return base


def normalizar_cpf(valor: str) -> str:
    return re.sub(r"\D", "", valor or "")


def carregar_lista_funcionarios(caminho: Path) -> list[dict]:
    if not caminho.exists():
        return []

    wb = load_workbook(caminho, data_only=True, read_only=True)
    registros: list[dict] = []

    for ws in wb.worksheets:
        if ws.max_row < 2 or ws.max_column < 2:
            continue

        coluna_codigo = 1
        coluna_nome = 2
        coluna_nome_completo = 3 if ws.max_column >= 3 else 2
        coluna_cpf = ws.max_column

        for r in range(2, ws.max_row + 1):
            codigo_raw = ws.cell(r, coluna_codigo).value
            nome_raw = ws.cell(r, coluna_nome).value
            nome_completo_raw = ws.cell(r, coluna_nome_completo).value
            cpf_raw = ws.cell(r, coluna_cpf).value

            codigo = formatar_codigo_iob(str(codigo_raw or ""))
            nome = str(nome_raw or "").strip()
            nome_completo = str(nome_completo_raw or "").strip()
            cpf = normalizar_cpf(str(cpf_raw or ""))

            if codigo == "00000":
                continue
            if not nome and not nome_completo and not cpf:
                continue

            registros.append(
                {
                    "codigo": codigo,
                    "nome": nome,
                    "nome_norm": normalizar_nome(nome),
                    "nome_completo": nome_completo,
                    "nome_completo_norm": normalizar_nome(nome_completo),
                    "cpf": cpf,
                    "aba": ws.title,
                    "linha": r,
                }
            )

    return registros


def buscar_matricula_na_lista(nome_pdf: str, cpf_pdf: str, caminho_lista: Path) -> tuple[str, str]:
    registros = carregar_lista_funcionarios(caminho_lista)
    if not registros:
        return "", "", False, []

    cpf_pdf_norm = normalizar_cpf(cpf_pdf)
    nome_pdf_norm = normalizar_nome(nome_pdf)

    def escolher_preferido(candidatos: list[dict]) -> dict:
        # Em duplicidade, prioriza a ultima linha da planilha (registro mais recente)
        return sorted(candidatos, key=lambda item: (item["linha"], item["codigo"]))[-1]

    if cpf_pdf_norm:
        matches_cpf = [item for item in registros if item["cpf"] == cpf_pdf_norm]
        if len(matches_cpf) == 1:
            item = matches_cpf[0]
            return item["codigo"], f"lista:{caminho_lista.name}:{item['aba']}!L{item['linha']} (CPF)", False, [item["codigo"]]
        if len(matches_cpf) > 1:
            match_nome = [
                item
                for item in matches_cpf
                if nome_pdf_norm
                and nome_pdf_norm in {item["nome_norm"], item["nome_completo_norm"]}
            ]
            if len(match_nome) == 1:
                item = match_nome[0]
                return item["codigo"], f"lista:{caminho_lista.name}:{item['aba']}!L{item['linha']} (CPF+NOME)", False, [item["codigo"]]
            candidatos = match_nome if match_nome else matches_cpf
            item = escolher_preferido(candidatos)
            codigos = sorted({x["codigo"] for x in candidatos})
            return (
                item["codigo"],
                f"lista:{caminho_lista.name}:{item['aba']}!L{item['linha']} (CPF-DUP:ULTIMA_LINHA)",
                True,
                codigos,
            )

    if nome_pdf_norm:
        matches_nome = [
            item
            for item in registros
            if nome_pdf_norm in {item["nome_norm"], item["nome_completo_norm"]}
        ]
        if len(matches_nome) == 1:
            item = matches_nome[0]
            return item["codigo"], f"lista:{caminho_lista.name}:{item['aba']}!L{item['linha']} (NOME)", False, [item["codigo"]]
        if len(matches_nome) > 1:
            item = escolher_preferido(matches_nome)
            codigos = sorted({x["codigo"] for x in matches_nome})
            return (
                item["codigo"],
                f"lista:{caminho_lista.name}:{item['aba']}!L{item['linha']} (NOME-DUP:ULTIMA_LINHA)",
                True,
                codigos,
            )

    return "", "", False, []


def formatar_vinculo_iob(resultado: ResultadoExtracao) -> str:
    base = (resultado.cpf or resultado.empregado or "").strip()
    return base[:20].ljust(20)


def decimal_para_centavos_iob(valor: Decimal) -> str:
    centavos = int((valor * 100).quantize(Decimal("1")))
    if centavos < 0:
        centavos = 0
    return str(centavos).zfill(11)[-11:]


def montar_linha_iob(codigo: str, evento: int, valor_centavos: str, vinculo: str) -> str:
    evento3 = "000"
    evento5 = "00000"
    if evento <= 999:
        evento3 = f"{evento:03d}"
    else:
        evento5 = f"{evento:05d}"
    return f"{codigo}|{evento3}|00000|{valor_centavos}|{vinculo}|{evento5}|"


def extrair_metadados(linhas: list[str]) -> tuple[str, str, str, str, str, str]:
    empresa = linhas[0] if linhas else ""

    periodo_inicial = ""
    periodo_final = ""
    texto_compacto = "\n".join(linhas)
    m_periodo = PADRAO_PERIODO.search(texto_compacto)
    if m_periodo:
        periodo_inicial, periodo_final = m_periodo.group(1), m_periodo.group(2)

    empregado = ""
    for i, linha in enumerate(linhas):
        if "EMPREGADO EMPREGADOR" in linha.upper() and i + 1 < len(linhas):
            empregado = linhas[i + 1].strip()
            break

    cpf = ""
    m_cpf = PADRAO_CPF.search(texto_compacto)
    if m_cpf:
        cpf = m_cpf.group(1)

    matricula = ""
    m_matricula = PADRAO_MATRICULA.search(texto_compacto)
    if m_matricula:
        matricula = m_matricula.group(1)

    return empresa, periodo_inicial, periodo_final, empregado, cpf, matricula


def extrair_bloco_principal(linhas: list[str]) -> dict[str, str]:
    indice_total = -1
    for i, linha in enumerate(linhas):
        if "TOTAL DE REMUNERACAO MENSAL DO MOTORISTA" in linha.upper() or "TOTAL DE REMUNERAÇÃO MENSAL DO MOTORISTA" in linha.upper():
            indice_total = i
            break

    if indice_total < 0:
        raise ValueError("Nao foi encontrado o bloco principal de remuneracao na ultima pagina.")

    valores = []
    for linha in linhas[indice_total + 1 :]:
        if linha.upper().startswith("VISTO:"):
            break
        if eh_valor_br(linha):
            valores.append(linha)

    if len(valores) < len(CAMPOS_BLOCO_PRINCIPAL):
        raise ValueError(
            f"Quantidade de valores insuficiente no bloco principal: esperado {len(CAMPOS_BLOCO_PRINCIPAL)}, encontrado {len(valores)}."
        )

    saida = {}
    for campo, valor in zip(CAMPOS_BLOCO_PRINCIPAL, valores):
        saida[campo] = valor
    return saida


def extrair_campos_linha_unica(linhas: list[str]) -> dict[str, str]:
    saida: dict[str, str] = {}
    for linha in linhas:
        m_salario = re.search(r"Sal[aá]rio\s+(-?\d{1,3}(?:\.\d{3})*,\d{2}|-?\d+,\d{2})", linha, re.IGNORECASE)
        if m_salario:
            saida["salario"] = m_salario.group(1)

        m_estadia = re.search(r"Estadia/Tempo de espera\s+(-?\d{1,3}(?:\.\d{3})*,\d{2}|-?\d+,\d{2})", linha, re.IGNORECASE)
        if m_estadia:
            saida["estadia_tempo_espera"] = m_estadia.group(1)

        m_premio = re.search(r"Premio disco tac[oó]grafo\s+(-?\d{1,3}(?:\.\d{3})*,\d{2}|-?\d+,\d{2})", linha, re.IGNORECASE)
        if m_premio:
            saida["premio_disco_tacografo"] = m_premio.group(1)

    for chave in ("salario", "estadia_tempo_espera", "premio_disco_tacografo"):
        saida.setdefault(chave, "0,00")

    return saida


def extrair_dados_ultima_pagina(pdf_path: Path, matricula_forcada: str = "") -> ResultadoExtracao:
    texto = extrair_ultima_pagina(pdf_path)
    linhas = listar_linhas(texto)

    empresa, periodo_inicial, periodo_final, empregado, cpf, matricula = extrair_metadados(linhas)
    valores = {}
    valores.update(extrair_bloco_principal(linhas))
    valores.update(extrair_campos_linha_unica(linhas))

    matricula_final = (matricula_forcada or matricula or "").strip()
    origem_matricula = ""
    matricula_foi_heuristica = False
    matricula_candidatas: list[str] = []
    if matricula_forcada:
        origem_matricula = "manual"
        matricula_candidatas = [formatar_codigo_iob(matricula_forcada)]
    elif matricula:
        origem_matricula = "pdf"
        matricula_candidatas = [formatar_codigo_iob(matricula)]
    else:
        matricula_lista, origem_lista, foi_heuristica, candidatas = buscar_matricula_na_lista(empregado, cpf, ARQUIVO_LISTA_FUNCIONARIOS)
        if matricula_lista:
            matricula_final = matricula_lista
            origem_matricula = origem_lista
            matricula_foi_heuristica = foi_heuristica
            matricula_candidatas = candidatas

    return ResultadoExtracao(
        empresa=empresa,
        periodo_inicial=periodo_inicial,
        periodo_final=periodo_final,
        empregado=empregado,
        cpf=cpf,
        matricula=matricula_final,
        origem_matricula=origem_matricula,
        matricula_foi_heuristica=matricula_foi_heuristica,
        matricula_candidatas=matricula_candidatas,
        valores=valores,
    )


def montar_linhas_txt(resultado: ResultadoExtracao) -> list[str]:
    # Mantem o padrao de linha do script original: codigo|evento|hhhmm|valor|vinculo|evento5|
    codigo = formatar_codigo_iob(resultado.matricula)
    vinculo = formatar_vinculo_iob(resultado)
    linhas_txt = []
    for campo in [
        *CAMPOS_BLOCO_PRINCIPAL,
        "salario",
        "estadia_tempo_espera",
        "premio_disco_tacografo",
    ]:
        evento = EVENTO_POR_CAMPO.get(campo, 0)
        if evento <= 0:
            continue
        valor = resultado.valores.get(campo, "0,00")
        valor_centavos = decimal_para_centavos_iob(valor_br_para_decimal(valor))
        linhas_txt.append(montar_linha_iob(codigo, evento, valor_centavos, vinculo))
    return linhas_txt


def montar_preview(resultado: ResultadoExtracao, linhas_txt: list[str]) -> dict:
    itens = []
    total = Decimal("0")
    for campo, valor in resultado.valores.items():
        evento = EVENTO_POR_CAMPO.get(campo, 0)
        dec = valor_br_para_decimal(valor)
        total += dec
        itens.append(
            {
                "campo": campo,
                "rotulo": ROTULO_CAMPO.get(campo, campo),
                "evento": evento,
                "valor": decimal_para_br(dec),
                "valor_decimal": float(dec),
                "valor_centavos_iob": decimal_para_centavos_iob(dec),
            }
        )

    return {
        "empresa": resultado.empresa,
        "periodo_inicial": resultado.periodo_inicial,
        "periodo_final": resultado.periodo_final,
        "empregado": resultado.empregado,
        "cpf": resultado.cpf,
        "matricula": resultado.matricula,
        "origem_matricula": resultado.origem_matricula,
        "matricula_foi_heuristica": resultado.matricula_foi_heuristica,
        "matricula_candidatas": resultado.matricula_candidatas,
        "permite_edicao_matricula": True,
        "matricula_encontrada": bool(resultado.matricula.strip()),
        "requer_matricula_usuario": not bool(resultado.matricula.strip()),
        "mensagem_matricula": (
            ""
            if resultado.matricula.strip() and not resultado.matricula_foi_heuristica
            else (
                "Matricula sugerida automaticamente com duplicidade na lista. Confira/edite antes de regenerar."
                if resultado.matricula.strip() and resultado.matricula_foi_heuristica
                else "Matricula nao encontrada no PDF/lista. Informe a matricula para gerar o TXT final com codigo correto."
            )
        ),
        "itens": itens,
        "total_itens": len(itens),
        "soma_itens": decimal_para_br(total),
        "linhas_txt": linhas_txt,
    }


def processar_pdf_para_txt(pdf_path: Path, pasta_saida: Path | None = None, matricula_forcada: str = "") -> tuple[Path, Path]:
    resultado = extrair_dados_ultima_pagina(pdf_path, matricula_forcada=matricula_forcada)
    linhas_txt = montar_linhas_txt(resultado)
    preview = montar_preview(resultado, linhas_txt)

    destino = pasta_saida or PASTA_GERADOS
    destino.mkdir(parents=True, exist_ok=True)

    nome_base = pdf_path.stem
    caminho_txt = destino / f"{nome_base}.txt"
    caminho_preview = destino / f"{nome_base}_preview.json"

    caminho_txt.write_text("\n".join(linhas_txt), encoding="utf-8")
    caminho_preview.write_text(json.dumps(preview, ensure_ascii=False, indent=2), encoding="utf-8")

    return caminho_txt, caminho_preview


def ler_argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extrai exclusivamente a ultima pagina do PDF e gera TXT + preview.")
    parser.add_argument("--pdf", type=Path, help="Caminho do PDF de entrada.")
    parser.add_argument("--saida", type=Path, help="Pasta de saida para TXT/preview.")
    parser.add_argument("--matricula", type=str, help="Matricula informada manualmente (opcional).")
    return parser.parse_args()


def main() -> None:
    args = ler_argumentos()

    if args.pdf:
        pdf = args.pdf
    else:
        pdfs = sorted(PASTA_EXEMPLO.glob("*.pdf"), key=lambda p: p.name.lower())
        if not pdfs:
            raise FileNotFoundError(f"Nenhum PDF encontrado em: {PASTA_EXEMPLO}")
        pdf = pdfs[0]

    if not pdf.exists():
        raise FileNotFoundError(f"PDF nao encontrado: {pdf}")

    caminho_txt, caminho_preview = processar_pdf_para_txt(pdf, args.saida, matricula_forcada=(args.matricula or ""))
    print(f"TXT gerado: {caminho_txt}")
    print(f"Preview JSON gerado: {caminho_preview}")


if __name__ == "__main__":
    main()
