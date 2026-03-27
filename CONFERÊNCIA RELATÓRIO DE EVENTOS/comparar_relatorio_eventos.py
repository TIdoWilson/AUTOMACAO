#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
Compara a planilha-base da folha com o relatorio oficial de eventos.

Uso:
    python "W:\DOCUMENTOS ESCRITORIO\INSTALACAO SISTEMA\python\CONFERÊNCIA RELATÓRIO DE EVENTOS\comparar_relatorio_eventos.py" `
        --arquivo-base "C:\Users\User\OneDrive\Área de Trabalho\Folhas para o robô novo\FOLHA - WILSON - jatonet.xlsm" `
        --arquivo-oficial "W:\DOCUMENTOS ESCRITORIO\INSTALACAO SISTEMA\python\CONFERÊNCIA RELATÓRIO DE EVENTOS\Oficiais\jatonet oficial.xlsx"
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from collections import Counter
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError as exc:
    print("Erro: modulo 'openpyxl' nao encontrado. Instale com: pip install openpyxl")
    raise SystemExit(1) from exc


# =========================
# Configuracao
# =========================
COLUNA_NOME_BASE = 2
LINHA_CABECALHO_1 = 3
LINHA_CABECALHO_2 = 4
LINHA_INICIAL_BASE = 5
LINHA_INICIAL_OFICIAL = 6
TOLERANCIA_NOME = 0.82
MARGEM_MINIMA_NOME = 0.05

MAPEAMENTO_EVENTOS = [
    {"coluna": 6, "evento": "SALARIO NORMAL"},
    {"coluna": 10, "evento": "PREMIO"},
    {"coluna": 11, "evento": "ADICIONAL DE PERIC"},
    {
        "coluna": 12,
        "evento": "DESCONTO / EMPRESTIMOS",
        "tipo": "presenca",
        "eventos_oficial": [
            "EMPRESTIMO CREDITO",
            "EMPRESTIMO CRED TR",
            "EMPRESTIMO CREDITO TRABALHADOR",
        ],
    },
    {"coluna": 18, "evento": "SOBRE AVISO", "coluna_referencia": 17},
    {"coluna": 19, "evento": "DSR SOBRE AVISO"},
    {"coluna": 22, "evento": "HORAS EXTRAS 050%", "coluna_referencia": 20},
    {"coluna": 23, "evento": "DSR SOBRE VARIAVEI"},
    {"coluna": 24, "evento": "HORAS EXTRAS 100%", "coluna_referencia": 21},
    {"coluna": 25, "evento": "DSR SOBRE VARIAVEI"},
    {"coluna": 31, "evento": "DESC DE FALTAS INJ", "coluna_referencia": 30},
    {"coluna": 35, "evento": "DESCONTO INSS"},
    {"coluna": 40, "evento": "IRPF"},
    {"coluna": 44, "evento": "VALE REFEIÇÃO"},
]

COLUNAS_AUXILIARES_BASE = {
    1,
    2,
    3,
    4,
    5,
    8,
    9,
    13,
    14,
    15,
    16,
    26,
    27,
    28,
    32,
    34,
    36,
    37,
    38,
    39,
    41,
    42,
    43,
    45,
    46,
    47,
}

COLUNAS_IGNORADAS_SEM_MAPEAMENTO = {
    7,
}

ALIASES_EVENTOS_OFICIAL = {
    "DESCONTO INSS": ["DESCONTO INSS"],
    "IRPF": ["IRPF", "DESCONTO I.R.R.F."],
}

EVENTOS_COMPARAR_POR_SOMA = {
    "IRPF",
}

EVENTOS_COMPARAR_POR_SOMA_BASE = {
    "DSR SOBRE VARIAVEI",
}

EVENTOS_AFASTAMENTO_ACIDENTE = {
    "AUX ACIDENTE PAGO",
    "AUX ACIDENTE PAGO INSS",
}


# =========================
# Modelos
# =========================
@dataclass(frozen=True)
class Evento:
    origem: str
    funcionario: str
    descricao: str
    valor: Decimal
    referencia: str = ""

    @property
    def chave(self) -> tuple[str, Decimal]:
        return normalizar_texto(self.descricao), self.valor


@dataclass
class FuncionarioBase:
    nome: str
    eventos: list[Evento]


@dataclass
class FuncionarioOficial:
    nome: str
    eventos: list[Evento]


@dataclass(frozen=True)
class PareamentoNome:
    nome_base: str
    nome_oficial: str
    score: float
    metodo: str


@dataclass(frozen=True)
class Divergencia:
    tipo: str
    funcionario_base: str
    funcionario_oficial: str
    evento_base: str
    evento_oficial: str
    valor_base: str
    valor_oficial: str
    referencia_base: str
    referencia_oficial: str


@dataclass(frozen=True)
class LinhaConferencia:
    funcionario_base: str
    funcionario_oficial: str
    status: str
    evento: str
    valor_base: str
    valor_oficial: str
    referencia_base: str
    referencia_oficial: str


# =========================
# Utilitarios
# =========================
def remover_acentos(texto: str) -> str:
    return unicodedata.normalize("NFKD", texto or "").encode("ascii", "ignore").decode("ascii")


def normalizar_texto(texto: str) -> str:
    texto = remover_acentos(texto).upper()
    texto = re.sub(r"[^A-Z0-9%/ ]+", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def tokenizar_nome(texto: str) -> list[str]:
    tokens = normalizar_texto(texto).split()
    return [token for token in tokens if len(token) > 2 and token not in {"FILHO", "NETO"}]


def nome_canonico(texto: str) -> str:
    return " ".join(tokenizar_nome(texto))


def parse_decimal(valor: object) -> Decimal | None:
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        numero = Decimal(str(valor)).quantize(Decimal("0.01"))
        if numero == 0:
            return None
        return numero

    texto = str(valor).strip()
    if not texto:
        return None
    texto = texto.replace("R$", "").replace(".", "").replace(" ", "").replace(",", ".")
    if texto in {"-", ""}:
        return None

    try:
        numero = Decimal(texto)
    except InvalidOperation:
        return None

    if numero == 0:
        return None
    return numero.quantize(Decimal("0.01"))


def parse_texto_celula(valor: object) -> str:
    if valor is None:
        return ""
    if isinstance(valor, str):
        return valor.strip()
    return str(valor).strip()


def formatar_decimal(valor: Decimal | None) -> str:
    if valor is None:
        return ""
    texto = f"{valor:.2f}"
    inteiro, decimal = texto.split(".")
    grupos: list[str] = []
    while inteiro:
        grupos.append(inteiro[-3:])
        inteiro = inteiro[:-3]
    return ".".join(reversed(grupos)) + "," + decimal


def cabecalho_coluna(worksheet, coluna: int) -> str:
    topo = parse_texto_celula(worksheet.cell(LINHA_CABECALHO_1, coluna).value)
    subtitulo = parse_texto_celula(worksheet.cell(LINHA_CABECALHO_2, coluna).value)
    return " / ".join(parte for parte in (topo, subtitulo) if parte)


def score_nomes(nome_base: str, nome_oficial: str) -> tuple[float, str]:
    base_norm = normalizar_texto(nome_base)
    oficial_norm = normalizar_texto(nome_oficial)
    if base_norm == oficial_norm:
        return 1.0, "exato"

    base_canonico = nome_canonico(nome_base)
    oficial_canonico = nome_canonico(nome_oficial)
    if base_canonico and base_canonico == oficial_canonico:
        return 1.0, "canonico"

    tokens_base = tokenizar_nome(nome_base)
    tokens_oficial = tokenizar_nome(nome_oficial)
    if not tokens_base or not tokens_oficial:
        return 0.0, "sem_tokens"

    conjunto_base = set(tokens_base)
    conjunto_oficial = set(tokens_oficial)
    comuns = conjunto_base & conjunto_oficial
    if not comuns:
        ratio = SequenceMatcher(None, base_canonico, oficial_canonico).ratio()
        if ratio >= 0.84 and tokens_base[0] == tokens_oficial[0]:
            return ratio, "sequencia"
        return 0.0, "sem_intersecao"

    cobertura = len(comuns) / min(len(conjunto_base), len(conjunto_oficial))
    jaccard = len(comuns) / len(conjunto_base | conjunto_oficial)
    primeiro = 0.1 if tokens_base[0] == tokens_oficial[0] else 0.0
    ultimo = 0.1 if tokens_base[-1].startswith(tokens_oficial[-1]) or tokens_oficial[-1].startswith(tokens_base[-1]) else 0.0
    ratio = SequenceMatcher(None, base_canonico, oficial_canonico).ratio()
    score = min(max((0.45 * cobertura) + (0.2 * jaccard) + (0.25 * ratio) + primeiro + ultimo, ratio), 1.0)

    if cobertura == 1 and (primeiro > 0 or ultimo > 0):
        score = max(score, 0.93)
        return score, "tokens"

    if ratio >= 0.9 and primeiro > 0:
        return score, "sequencia"

    return score, "similaridade"


def iterar_eventos_linha_oficial(valores: list[object], funcionario: str) -> Iterable[Evento]:
    blocos = ((2, 3, 4, 5), (7, 8, 9, 10))
    for idx_codigo, idx_descricao, idx_referencia, idx_valor in blocos:
        descricao = parse_texto_celula(valores[idx_descricao])
        valor = parse_decimal(valores[idx_valor])
        if not descricao or valor is None:
            continue
        yield Evento(
            origem="oficial",
            funcionario=funcionario,
            descricao=descricao,
            valor=valor,
            referencia=parse_texto_celula(valores[idx_referencia]),
        )


# =========================
# Leitura
# =========================
def ler_base(caminho_arquivo: Path) -> tuple[dict[str, FuncionarioBase], list[str]]:
    workbook = load_workbook(caminho_arquivo, data_only=True, keep_vba=True)
    try:
        if "Folha de pagamento" not in workbook.sheetnames:
            raise ValueError("Aba 'Folha de pagamento' nao encontrada no arquivo-base.")

        worksheet = workbook["Folha de pagamento"]
        funcionarios: dict[str, FuncionarioBase] = {}
        colunas_nao_mapeadas: list[str] = []

        for linha in range(LINHA_INICIAL_BASE, worksheet.max_row + 1):
            nome = parse_texto_celula(worksheet.cell(linha, COLUNA_NOME_BASE).value)
            if not nome:
                continue

            salario = parse_decimal(worksheet.cell(linha, 6).value)
            if salario is None:
                continue

            eventos: list[Evento] = []
            for item in MAPEAMENTO_EVENTOS:
                valor = parse_decimal(worksheet.cell(linha, item["coluna"]).value)
                if valor is None:
                    continue

                referencia = ""
                coluna_referencia = item.get("coluna_referencia")
                if coluna_referencia:
                    referencia = parse_texto_celula(worksheet.cell(linha, coluna_referencia).value)

                eventos.append(
                    Evento(
                        origem="base",
                        funcionario=nome,
                        descricao=item["evento"],
                        valor=valor,
                        referencia=referencia,
                    )
                )

            if eventos:
                funcionarios[nome] = FuncionarioBase(nome=nome, eventos=eventos)

        colunas_mapeadas = {item["coluna"] for item in MAPEAMENTO_EVENTOS}
        for coluna in range(1, worksheet.max_column + 1):
            cabecalho = cabecalho_coluna(worksheet, coluna)
            if (
                not cabecalho
                or coluna in colunas_mapeadas
                or coluna in COLUNAS_AUXILIARES_BASE
                or coluna in COLUNAS_IGNORADAS_SEM_MAPEAMENTO
            ):
                continue
            tem_numero = False
            for linha in range(LINHA_INICIAL_BASE, worksheet.max_row + 1):
                if parse_decimal(worksheet.cell(linha, coluna).value) is not None:
                    tem_numero = True
                    break
            if tem_numero:
                colunas_nao_mapeadas.append(f"Coluna {coluna}: {cabecalho}")

        return funcionarios, colunas_nao_mapeadas
    finally:
        workbook.close()


def ler_oficial(caminho_arquivo: Path) -> dict[str, FuncionarioOficial]:
    workbook = load_workbook(caminho_arquivo, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        funcionarios: dict[str, list[Evento]] = {}
        nome_atual = ""
        eventos_atuais: list[tuple[str, Decimal, str]] = []

        def finalizar_bloco() -> None:
            nonlocal nome_atual, eventos_atuais
            nome_final = nome_atual.strip()
            if not nome_final or not eventos_atuais:
                nome_atual = ""
                eventos_atuais = []
                return

            eventos_convertidos = [
                Evento(
                    origem="oficial",
                    funcionario=nome_final,
                    descricao=descricao,
                    valor=valor,
                    referencia=referencia,
                )
                for descricao, valor, referencia in eventos_atuais
            ]
            funcionarios.setdefault(nome_final, []).extend(eventos_convertidos)
            nome_atual = ""
            eventos_atuais = []

        for linha in range(LINHA_INICIAL_OFICIAL, worksheet.max_row + 1):
            valores = [worksheet.cell(linha, coluna).value for coluna in range(1, 13)]
            primeira = parse_texto_celula(valores[0])
            segunda = parse_texto_celula(valores[1])
            terceira = parse_texto_celula(valores[2])
            quarta = parse_texto_celula(valores[3])

            if quarta.upper() == "DESCRIÇÃO" or quarta.upper() == "DESCRICAO":
                break

            if primeira:
                finalizar_bloco()
                nome_atual = segunda
            elif segunda:
                nome_atual = (nome_atual + " " + segunda).strip()

            if not nome_atual:
                continue

            eventos_linha = list(iterar_eventos_linha_oficial(valores, nome_atual))
            if not eventos_linha and not terceira and not quarta:
                continue

            for evento in eventos_linha:
                eventos_atuais.append((evento.descricao, evento.valor, evento.referencia))

        finalizar_bloco()

        return {
            nome: FuncionarioOficial(nome=nome, eventos=eventos)
            for nome, eventos in funcionarios.items()
            if eventos
        }
    finally:
        workbook.close()


# =========================
# Comparacao
# =========================
def parear_funcionarios(
    funcionarios_base: dict[str, FuncionarioBase],
    funcionarios_oficial: dict[str, FuncionarioOficial],
) -> tuple[list[PareamentoNome], list[str], list[str]]:
    pareamentos: list[PareamentoNome] = []
    bases_livres = set(funcionarios_base)
    oficiais_livres = set(funcionarios_oficial)

    for nome_oficial in sorted(funcionarios_oficial):
        candidatos: list[tuple[float, str, str]] = []
        for nome_base in bases_livres:
            score, metodo = score_nomes(nome_base, nome_oficial)
            if score <= 0:
                continue
            candidatos.append((score, metodo, nome_base))

        if not candidatos:
            continue

        candidatos.sort(key=lambda item: item[0], reverse=True)
        melhor_score, metodo, melhor_nome = candidatos[0]
        segundo_score = candidatos[1][0] if len(candidatos) > 1 else 0.0

        if melhor_score >= TOLERANCIA_NOME and (melhor_score - segundo_score) >= MARGEM_MINIMA_NOME:
            pareamentos.append(
                PareamentoNome(
                    nome_base=melhor_nome,
                    nome_oficial=nome_oficial,
                    score=melhor_score,
                    metodo=metodo,
                )
            )
            bases_livres.remove(melhor_nome)
            oficiais_livres.remove(nome_oficial)

    for nome_oficial in sorted(list(oficiais_livres)):
        tokens_oficial = tokenizar_nome(nome_oficial)
        if not tokens_oficial:
            continue

        candidatos_mesmo_primeiro_nome = []
        for nome_base in bases_livres:
            tokens_base = tokenizar_nome(nome_base)
            if not tokens_base or tokens_base[0] != tokens_oficial[0]:
                continue
            score, metodo = score_nomes(nome_base, nome_oficial)
            candidatos_mesmo_primeiro_nome.append((score, metodo, nome_base))

        if len(candidatos_mesmo_primeiro_nome) != 1:
            continue

        score, metodo, nome_base = candidatos_mesmo_primeiro_nome[0]
        if score < 0.55:
            continue

        pareamentos.append(
            PareamentoNome(
                nome_base=nome_base,
                nome_oficial=nome_oficial,
                score=score,
                metodo=f"{metodo}_primeiro_nome",
            )
        )
        bases_livres.remove(nome_base)
        oficiais_livres.remove(nome_oficial)

    return pareamentos, sorted(bases_livres), sorted(oficiais_livres)


def comparar_eventos(
    funcionarios_base: dict[str, FuncionarioBase],
    funcionarios_oficial: dict[str, FuncionarioOficial],
    pareamentos: list[PareamentoNome],
) -> tuple[list[Divergencia], list[LinhaConferencia]]:
    divergencias: list[Divergencia] = []
    conferencia: list[LinhaConferencia] = []
    eventos_mapeados = {normalizar_texto(item["evento"]) for item in MAPEAMENTO_EVENTOS}
    regras_presenca = [
        item for item in MAPEAMENTO_EVENTOS if str(item.get("tipo") or "").strip().lower() == "presenca"
    ]
    tolerancia_agregacao = Decimal("0.02")

    def ordenar_eventos(eventos: list[Evento]) -> list[Evento]:
        return sorted(
            eventos,
            key=lambda evento: (
                normalizar_texto(evento.descricao),
                evento.valor,
                normalizar_texto(evento.referencia),
            ),
        )

    for pareamento in pareamentos:
        eventos_base = list(funcionarios_base[pareamento.nome_base].eventos)
        eventos_oficial = list(funcionarios_oficial[pareamento.nome_oficial].eventos)
        por_desc_base: dict[str, list[Evento]] = {}
        por_desc_oficial_bruto: dict[str, list[Evento]] = {}

        eventos_aux_acidente = [
            evento
            for evento in eventos_oficial
            if normalizar_texto(evento.descricao) in EVENTOS_AFASTAMENTO_ACIDENTE
        ]
        eventos_oficial_relevantes = [
            evento
            for evento in eventos_oficial
            if normalizar_texto(evento.descricao) in eventos_mapeados
            or normalizar_texto(evento.descricao) in EVENTOS_AFASTAMENTO_ACIDENTE
        ]
        if eventos_aux_acidente and len(eventos_oficial_relevantes) == len(eventos_aux_acidente):
            valor_aux = sum((evento.valor for evento in eventos_aux_acidente), Decimal("0.00"))
            conferencia.append(
                LinhaConferencia(
                    funcionario_base=pareamento.nome_base,
                    funcionario_oficial=pareamento.nome_oficial,
                    status="OK",
                    evento="AUX. ACIDENTE PAGO INSS",
                    valor_base="",
                    valor_oficial=formatar_decimal(valor_aux),
                    referencia_base="",
                    referencia_oficial=" + ".join(
                        referencia
                        for referencia in [evento.referencia.strip() for evento in eventos_aux_acidente]
                        if referencia
                    ),
                )
            )
            continue

        salario_base = next((evento for evento in eventos_base if normalizar_texto(evento.descricao) == "SALARIO NORMAL"), None)
        salario_oficial = [evento for evento in eventos_oficial if normalizar_texto(evento.descricao) == "SALARIO NORMAL"]
        ferias_oficial = [evento for evento in eventos_oficial if normalizar_texto(evento.descricao) == "FERIAS NORMAIS"]
        if salario_base and salario_oficial and ferias_oficial:
            soma_oficial = sum((evento.valor for evento in salario_oficial + ferias_oficial), Decimal("0.00"))
            if abs(salario_base.valor - soma_oficial) <= tolerancia_agregacao:
                eventos_base = [evento for evento in eventos_base if evento is not salario_base]
                eventos_oficial = [
                    evento
                    for evento in eventos_oficial
                    if evento not in salario_oficial and evento not in ferias_oficial
                ]
                conferencia.append(
                    LinhaConferencia(
                        funcionario_base=pareamento.nome_base,
                        funcionario_oficial=pareamento.nome_oficial,
                        status="OK",
                        evento="SALARIO NORMAL (COM FERIAS)",
                        valor_base=formatar_decimal(salario_base.valor),
                        valor_oficial=formatar_decimal(soma_oficial),
                        referencia_base=salario_base.referencia,
                        referencia_oficial=" + ".join(
                            filtro for filtro in [salario_oficial[0].referencia.strip(), "FERIAS NORMAIS"] if filtro
                        ),
                    )
                )

        for regra in regras_presenca:
            descricao_base = regra["evento"]
            descricao_base_norm = normalizar_texto(descricao_base)
            base_regra = [evento for evento in eventos_base if normalizar_texto(evento.descricao) == descricao_base_norm]
            if not base_regra:
                continue

            nomes_oficial = {normalizar_texto(nome) for nome in regra.get("eventos_oficial", [])}

            def evento_casa_regra_presenca(evento: Evento) -> bool:
                descricao_normalizada = normalizar_texto(evento.descricao)
                for nome_oficial in nomes_oficial:
                    if (
                        descricao_normalizada == nome_oficial
                        or descricao_normalizada.startswith(nome_oficial)
                        or nome_oficial.startswith(descricao_normalizada)
                        or nome_oficial in descricao_normalizada
                        or descricao_normalizada in nome_oficial
                    ):
                        return True
                return False

            oficiais_regra = [evento for evento in eventos_oficial if evento_casa_regra_presenca(evento)]

            eventos_base = [evento for evento in eventos_base if normalizar_texto(evento.descricao) != descricao_base_norm]
            eventos_oficial = [evento for evento in eventos_oficial if not evento_casa_regra_presenca(evento)]

            evento_base = base_regra[0]
            if oficiais_regra:
                valor_oficial_regra = sum((evento.valor for evento in oficiais_regra), Decimal("0.00"))
                conferencia.append(
                    LinhaConferencia(
                        funcionario_base=pareamento.nome_base,
                        funcionario_oficial=pareamento.nome_oficial,
                        status="OK",
                        evento=descricao_base,
                        valor_base=formatar_decimal(evento_base.valor),
                        valor_oficial=formatar_decimal(valor_oficial_regra),
                        referencia_base=evento_base.referencia,
                        referencia_oficial=" + ".join(
                            referencia
                            for referencia in [evento.referencia.strip() for evento in oficiais_regra]
                            if referencia
                        ),
                    )
                )
                continue

            divergencias.append(
                Divergencia(
                    tipo="faltando_no_oficial",
                    funcionario_base=pareamento.nome_base,
                    funcionario_oficial=pareamento.nome_oficial,
                    evento_base=evento_base.descricao,
                    evento_oficial="",
                    valor_base=formatar_decimal(evento_base.valor),
                    valor_oficial="",
                    referencia_base=evento_base.referencia,
                    referencia_oficial="",
                )
            )
            conferencia.append(
                LinhaConferencia(
                    funcionario_base=pareamento.nome_base,
                    funcionario_oficial=pareamento.nome_oficial,
                    status="FALTANDO NO OFICIAL",
                    evento=evento_base.descricao,
                    valor_base=formatar_decimal(evento_base.valor),
                    valor_oficial="",
                    referencia_base=evento_base.referencia,
                    referencia_oficial="",
                )
            )

        for evento in eventos_base:
            por_desc_base.setdefault(normalizar_texto(evento.descricao), []).append(evento)
        for evento in eventos_oficial:
            por_desc_oficial_bruto.setdefault(normalizar_texto(evento.descricao), []).append(evento)

        descricoes = sorted(set(por_desc_base))
        for descricao_norm in descricoes:
            lista_base = ordenar_eventos(por_desc_base.get(descricao_norm, []))
            aliases = ALIASES_EVENTOS_OFICIAL.get((lista_base[0].descricao if lista_base else ""), [])
            descricoes_oficial = {descricao_norm}
            descricoes_oficial.update(normalizar_texto(alias) for alias in aliases)

            lista_oficial = []
            for descricao_oficial in descricoes_oficial:
                lista_oficial.extend(por_desc_oficial_bruto.get(descricao_oficial, []))
            lista_oficial = ordenar_eventos(lista_oficial)
            evento_legivel = lista_base[0].descricao

            if (
                evento_legivel in EVENTOS_COMPARAR_POR_SOMA
                and len(lista_base) == 1
                and len(lista_oficial) > 1
            ):
                soma_oficial = sum((evento.valor for evento in lista_oficial), Decimal("0.00"))
                referencia_oficial = " + ".join(
                    referencia for referencia in [evento.referencia.strip() for evento in lista_oficial] if referencia
                )
                evento_base = lista_base[0]
                referencias_compativeis = (
                    not evento_base.referencia.strip()
                    or not referencia_oficial
                    or evento_base.referencia.strip() == referencia_oficial
                )
                if abs(evento_base.valor - soma_oficial) <= tolerancia_agregacao and referencias_compativeis:
                    conferencia.append(
                        LinhaConferencia(
                            funcionario_base=pareamento.nome_base,
                            funcionario_oficial=pareamento.nome_oficial,
                            status="OK",
                            evento=evento_legivel,
                            valor_base=formatar_decimal(evento_base.valor),
                            valor_oficial=formatar_decimal(soma_oficial),
                            referencia_base=evento_base.referencia.strip(),
                            referencia_oficial=referencia_oficial,
                        )
                    )
                    continue

            if (
                evento_legivel in EVENTOS_COMPARAR_POR_SOMA_BASE
                and len(lista_base) > 1
                and len(lista_oficial) == 1
            ):
                soma_base = sum((evento.valor for evento in lista_base), Decimal("0.00"))
                referencia_base = " + ".join(
                    referencia for referencia in [evento.referencia.strip() for evento in lista_base] if referencia
                )
                evento_oficial = lista_oficial[0]
                referencias_compativeis = (
                    not referencia_base
                    or not evento_oficial.referencia.strip()
                    or referencia_base == evento_oficial.referencia.strip()
                )
                if abs(soma_base - evento_oficial.valor) <= tolerancia_agregacao and referencias_compativeis:
                    conferencia.append(
                        LinhaConferencia(
                            funcionario_base=pareamento.nome_base,
                            funcionario_oficial=pareamento.nome_oficial,
                            status="OK",
                            evento=evento_legivel,
                            valor_base=formatar_decimal(soma_base),
                            valor_oficial=formatar_decimal(evento_oficial.valor),
                            referencia_base=referencia_base,
                            referencia_oficial=evento_oficial.referencia.strip(),
                        )
                    )
                else:
                    divergencias.append(
                        Divergencia(
                            tipo="valor_ou_referencia_divergente",
                            funcionario_base=pareamento.nome_base,
                            funcionario_oficial=pareamento.nome_oficial,
                            evento_base=evento_legivel,
                            evento_oficial=evento_oficial.descricao,
                            valor_base=formatar_decimal(soma_base),
                            valor_oficial=formatar_decimal(evento_oficial.valor),
                            referencia_base=referencia_base,
                            referencia_oficial=evento_oficial.referencia.strip(),
                        )
                    )
                    conferencia.append(
                        LinhaConferencia(
                            funcionario_base=pareamento.nome_base,
                            funcionario_oficial=pareamento.nome_oficial,
                            status="DIVERGENTE",
                            evento=evento_legivel,
                            valor_base=formatar_decimal(soma_base),
                            valor_oficial=formatar_decimal(evento_oficial.valor),
                            referencia_base=referencia_base,
                            referencia_oficial=evento_oficial.referencia.strip(),
                        )
                    )
                continue

            quantidade_pares = min(len(lista_base), len(lista_oficial))

            for indice in range(quantidade_pares):
                evento_base = lista_base[indice]
                evento_oficial = lista_oficial[indice]
                referencia_base = evento_base.referencia.strip()
                referencia_oficial = evento_oficial.referencia.strip()
                referencias_compativeis = (
                    referencia_base == referencia_oficial
                    or not referencia_base
                    or not referencia_oficial
                )
                if evento_base.valor == evento_oficial.valor and referencias_compativeis:
                    conferencia.append(
                        LinhaConferencia(
                            funcionario_base=pareamento.nome_base,
                            funcionario_oficial=pareamento.nome_oficial,
                            status="OK",
                            evento=evento_legivel,
                            valor_base=formatar_decimal(evento_base.valor),
                            valor_oficial=formatar_decimal(evento_oficial.valor),
                            referencia_base=referencia_base,
                            referencia_oficial=referencia_oficial,
                        )
                    )
                    continue

                divergencias.append(
                    Divergencia(
                        tipo="valor_ou_referencia_divergente",
                        funcionario_base=pareamento.nome_base,
                        funcionario_oficial=pareamento.nome_oficial,
                        evento_base=evento_base.descricao,
                        evento_oficial=evento_oficial.descricao,
                        valor_base=formatar_decimal(evento_base.valor),
                        valor_oficial=formatar_decimal(evento_oficial.valor),
                        referencia_base=referencia_base,
                        referencia_oficial=referencia_oficial,
                    )
                )
                conferencia.append(
                    LinhaConferencia(
                        funcionario_base=pareamento.nome_base,
                        funcionario_oficial=pareamento.nome_oficial,
                        status="DIVERGENTE",
                        evento=evento_legivel,
                        valor_base=formatar_decimal(evento_base.valor),
                        valor_oficial=formatar_decimal(evento_oficial.valor),
                        referencia_base=referencia_base,
                        referencia_oficial=referencia_oficial,
                    )
                )

            for evento in lista_base[quantidade_pares:]:
                divergencias.append(
                    Divergencia(
                        tipo="faltando_no_oficial",
                        funcionario_base=pareamento.nome_base,
                        funcionario_oficial=pareamento.nome_oficial,
                        evento_base=evento.descricao,
                        evento_oficial="",
                        valor_base=formatar_decimal(evento.valor),
                        valor_oficial="",
                        referencia_base=evento.referencia,
                        referencia_oficial="",
                    )
                )
                conferencia.append(
                    LinhaConferencia(
                        funcionario_base=pareamento.nome_base,
                        funcionario_oficial=pareamento.nome_oficial,
                        status="FALTANDO NO OFICIAL",
                        evento=evento.descricao,
                        valor_base=formatar_decimal(evento.valor),
                        valor_oficial="",
                        referencia_base=evento.referencia,
                        referencia_oficial="",
                    )
                )

            for evento in lista_oficial[quantidade_pares:]:
                tipo = "a_mais_no_oficial"
                status_linha = "A MAIS NO OFICIAL"
                if descricao_norm not in eventos_mapeados:
                    tipo = "sem_mapeamento_na_base"
                    status_linha = "SEM MAPEAMENTO NA BASE"
                divergencias.append(
                    Divergencia(
                        tipo=tipo,
                        funcionario_base=pareamento.nome_base,
                        funcionario_oficial=pareamento.nome_oficial,
                        evento_base="",
                        evento_oficial=evento.descricao,
                        valor_base="",
                        valor_oficial=formatar_decimal(evento.valor),
                        referencia_base="",
                        referencia_oficial=evento.referencia,
                    )
                )
                conferencia.append(
                    LinhaConferencia(
                        funcionario_base=pareamento.nome_base,
                        funcionario_oficial=pareamento.nome_oficial,
                        status=status_linha,
                        evento=evento.descricao,
                        valor_base="",
                        valor_oficial=formatar_decimal(evento.valor),
                        referencia_base="",
                        referencia_oficial=evento.referencia,
                    )
                )

    return divergencias, conferencia


# =========================
# Saida
# =========================
def ajustar_larguras(worksheet) -> None:
    limites: dict[int, int] = {}
    for linha in worksheet.iter_rows():
        for celula in linha:
            valor = "" if celula.value is None else str(celula.value)
            limites[celula.column] = min(max(limites.get(celula.column, 0), len(valor) + 2), 45)
            celula.alignment = Alignment(vertical="center", wrap_text=True)
    for coluna, largura in limites.items():
        worksheet.column_dimensions[get_column_letter(coluna)].width = largura


def aplicar_layout_grade(worksheet) -> None:
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions


def get_column_letter(indice: int) -> str:
    letras = ""
    while indice:
        indice, resto = divmod(indice - 1, 26)
        letras = chr(65 + resto) + letras
    return letras


def salvar_saida(
    caminho_saida: Path,
    arquivo_base: Path,
    arquivo_oficial: Path,
    pareamentos: list[PareamentoNome],
    divergencias: list[Divergencia],
    conferencia: list[LinhaConferencia],
    bases_sem_pareamento: list[str],
    oficiais_sem_pareamento: list[str],
    colunas_nao_mapeadas: list[str],
) -> None:
    workbook = Workbook()
    divergencias_reais = [
        item for item in divergencias
        if item.tipo not in {"sem_mapeamento_na_base", "a_mais_no_oficial"}
    ]
    borda = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    fill_titulo = PatternFill(fill_type="solid", fgColor="1F4E78")
    fill_cab = PatternFill(fill_type="solid", fgColor="D9EAF7")
    fonte_branca = Font(color="FFFFFF", bold=True)
    fonte_negrito = Font(bold=True)

    resumo = workbook.active
    resumo.title = "Resumo"
    linhas_resumo = [
        ("Arquivo base", str(arquivo_base)),
        ("Arquivo oficial", str(arquivo_oficial)),
        ("Funcionarios base", len(pareamentos) + len(bases_sem_pareamento)),
        ("Funcionarios oficial", len(pareamentos) + len(oficiais_sem_pareamento)),
        ("Pareamentos automaticos", len(pareamentos)),
        ("Divergencias reais", len(divergencias_reais)),
        ("Linhas OK na conferencia", sum(1 for item in conferencia if item.status == "OK")),
        ("Base sem pareamento", len(bases_sem_pareamento)),
        ("Oficial sem pareamento", len(oficiais_sem_pareamento)),
        ("Colunas da base nao mapeadas", len(colunas_nao_mapeadas)),
    ]
    resumo.append(["RESUMO DA COMPARACAO", ""])
    resumo.merge_cells("A1:B1")
    for celula in resumo[1]:
        celula.fill = fill_titulo
        celula.font = fonte_branca
        celula.alignment = Alignment(horizontal="center", vertical="center")
    for linha in linhas_resumo:
        resumo.append(list(linha))
    for row in resumo.iter_rows(min_row=2, max_row=resumo.max_row, min_col=1, max_col=2):
        row[0].font = fonte_negrito
        for celula in row:
            celula.border = borda

    aba_nomes = workbook.create_sheet("Pareamento Nomes")
    aba_nomes.append(["Nome base", "Nome oficial", "Score", "Metodo"])
    for linha in aba_nomes[1]:
        linha.fill = fill_cab
        linha.font = fonte_negrito
        linha.border = borda
    for item in pareamentos:
        aba_nomes.append([item.nome_base, item.nome_oficial, round(item.score, 4), item.metodo])
    for row in aba_nomes.iter_rows(min_row=2, max_row=aba_nomes.max_row):
        for celula in row:
            celula.border = borda

    aba_conf = workbook.create_sheet("Conferencia")
    ordem_status = {
        "DIVERGENTE": 0,
        "FALTANDO NO OFICIAL": 1,
        "OK": 2,
    }
    conferencia_ordenada = sorted(
        conferencia,
        key=lambda item: (
            item.funcionario_base,
            ordem_status.get(item.status, 99),
            normalizar_texto(item.evento),
            item.valor_base,
            item.valor_oficial,
        ),
    )
    conferencia_ordenada = [
        item for item in conferencia_ordenada
        if item.status != "SEM MAPEAMENTO NA BASE"
    ]
    linha_conf = 1
    funcionario_atual = None
    for item in conferencia_ordenada:
        if item.funcionario_base != funcionario_atual:
            if funcionario_atual is not None:
                linha_conf += 2
            funcionario_atual = item.funcionario_base
            aba_conf.cell(linha_conf, 1, funcionario_atual)
            aba_conf.cell(linha_conf, 1).font = fonte_branca
            aba_conf.cell(linha_conf, 1).fill = fill_titulo
            aba_conf.cell(linha_conf, 1).border = borda
            aba_conf.cell(linha_conf, 1).alignment = Alignment(horizontal="left", vertical="center")
            aba_conf.merge_cells(start_row=linha_conf, start_column=1, end_row=linha_conf, end_column=3)
            for coluna in range(1, 4):
                aba_conf.cell(linha_conf, coluna).border = borda
            linha_conf += 1

            cabecalhos = ["Evento", "Valor Folha", "Valor Relatório"]
            for coluna, cabecalho in enumerate(cabecalhos, start=1):
                celula = aba_conf.cell(linha_conf, coluna, cabecalho)
                celula.fill = fill_cab
                celula.font = fonte_negrito
                celula.border = borda
                celula.alignment = Alignment(horizontal="center", vertical="center")
            linha_conf += 1

        aba_conf.cell(linha_conf, 1, item.evento)
        aba_conf.cell(linha_conf, 2, item.valor_base)
        aba_conf.cell(linha_conf, 3, item.valor_oficial)
        for coluna in range(1, 4):
            aba_conf.cell(linha_conf, coluna).border = borda

        if item.status == "OK":
            fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
        else:
            fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
        for coluna in range(1, 4):
            aba_conf.cell(linha_conf, coluna).fill = fill
        linha_conf += 1

    aba_div = workbook.create_sheet("Divergencias")
    aba_div.append(
        [
            "Tipo",
            "Funcionario base",
            "Funcionario oficial",
            "Evento base",
            "Valor base",
            "Ref. base",
            "Evento oficial",
            "Valor oficial",
            "Ref. oficial",
        ]
    )
    for linha in aba_div[1]:
        linha.fill = fill_cab
        linha.font = fonte_negrito
        linha.border = borda
    for item in divergencias_reais:
        aba_div.append(
            [
                item.tipo,
                item.funcionario_base,
                item.funcionario_oficial,
                item.evento_base,
                item.valor_base,
                item.referencia_base,
                item.evento_oficial,
                item.valor_oficial,
                item.referencia_oficial,
            ]
        )
    for row in aba_div.iter_rows(min_row=2, max_row=aba_div.max_row):
        for celula in row:
            celula.border = borda

    aba_sem_nome = workbook.create_sheet("Sem Pareamento")
    aba_sem_nome.append(["Tipo", "Nome"])
    for linha in aba_sem_nome[1]:
        linha.fill = fill_cab
        linha.font = fonte_negrito
        linha.border = borda
    for nome in bases_sem_pareamento:
        aba_sem_nome.append(["base", nome])
    for nome in oficiais_sem_pareamento:
        aba_sem_nome.append(["oficial", nome])
    for row in aba_sem_nome.iter_rows(min_row=2, max_row=aba_sem_nome.max_row):
        for celula in row:
            celula.border = borda

    aba_mapa = workbook.create_sheet("Base Nao Mapeada")
    aba_mapa.append(["Colunas com numeros nao comparadas"])
    aba_mapa["A1"].fill = fill_cab
    aba_mapa["A1"].font = fonte_negrito
    aba_mapa["A1"].border = borda
    for item in colunas_nao_mapeadas:
        aba_mapa.append([item])
    for row in aba_mapa.iter_rows(min_row=2, max_row=aba_mapa.max_row):
        for celula in row:
            celula.border = borda

    for planilha in workbook.worksheets:
        ajustar_larguras(planilha)
        if planilha.title in {"Pareamento Nomes", "Conferencia", "Divergencias", "Sem Pareamento", "Base Nao Mapeada"}:
            aplicar_layout_grade(planilha)

    ordem_abas = [
        "Resumo",
        "Conferencia",
        "Pareamento Nomes",
        "Divergencias",
        "Sem Pareamento",
        "Base Nao Mapeada",
    ]
    for indice, nome_aba in enumerate(ordem_abas):
        if nome_aba in workbook.sheetnames:
            workbook.move_sheet(workbook[nome_aba], offset=-workbook.index(workbook[nome_aba]) + indice)

    workbook.save(caminho_saida)


# =========================
# Main
# =========================
def montar_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Compara planilha-base de folha com relatorio oficial de eventos.")
    parser.add_argument("--arquivo-base", required=True, help="Caminho do arquivo-base (.xlsx/.xlsm).")
    parser.add_argument("--arquivo-oficial", required=True, help="Caminho do relatorio oficial (.xlsx/.xlsm).")
    parser.add_argument("--saida", help="Caminho do Excel de saida. Se omitido, salva na pasta do script.")
    return parser


def main() -> int:
    parser = montar_parser()
    args = parser.parse_args()

    arquivo_base = Path(args.arquivo_base)
    arquivo_oficial = Path(args.arquivo_oficial)

    if not arquivo_base.exists():
        print(f"Erro: arquivo-base nao encontrado: {arquivo_base}")
        return 1
    if not arquivo_oficial.exists():
        print(f"Erro: arquivo oficial nao encontrado: {arquivo_oficial}")
        return 1

    funcionarios_base, colunas_nao_mapeadas = ler_base(arquivo_base)
    funcionarios_oficial = ler_oficial(arquivo_oficial)
    pareamentos, bases_sem_pareamento, oficiais_sem_pareamento = parear_funcionarios(funcionarios_base, funcionarios_oficial)
    divergencias, conferencia = comparar_eventos(funcionarios_base, funcionarios_oficial, pareamentos)
    conferencia = [item for item in conferencia if item.status != "A MAIS NO OFICIAL"]

    if args.saida:
        caminho_saida = Path(args.saida)
    else:
        pasta_script = Path(__file__).resolve().parent
        nome_saida = f"comparacao_{arquivo_base.stem}_vs_{arquivo_oficial.stem}.xlsx"
        caminho_saida = pasta_script / nome_saida

    salvar_saida(
        caminho_saida=caminho_saida,
        arquivo_base=arquivo_base,
        arquivo_oficial=arquivo_oficial,
        pareamentos=pareamentos,
        divergencias=divergencias,
        conferencia=conferencia,
        bases_sem_pareamento=bases_sem_pareamento,
        oficiais_sem_pareamento=oficiais_sem_pareamento,
        colunas_nao_mapeadas=colunas_nao_mapeadas,
    )

    divergencias_reais = sum(
        1 for item in divergencias
        if item.tipo not in {"sem_mapeamento_na_base", "a_mais_no_oficial"}
    )
    sem_mapeamento = sum(1 for item in divergencias if item.tipo == "sem_mapeamento_na_base")
    print(f"Comparacao concluida: {caminho_saida}")
    print(f"Pareamentos automaticos: {len(pareamentos)}")
    print(f"Divergencias reais: {divergencias_reais}")
    print(f"Base sem pareamento: {len(bases_sem_pareamento)}")
    print(f"Oficial sem pareamento: {len(oficiais_sem_pareamento)}")
    print(f"Colunas nao mapeadas da base: {len(colunas_nao_mapeadas)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
