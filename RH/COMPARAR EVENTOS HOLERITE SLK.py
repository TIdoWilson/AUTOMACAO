#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
Comparar eventos de holerite em arquivo .slk
--------------------------------------------------
O script:
- Lê um ou mais arquivos .slk de holerite;
- Identifica competência, funcionário e eventos de cada holerite;
- Compara os eventos de uma competência anterior com outra competência atual;
- Lista, por funcionário, os eventos que existiam no mês anterior e não existem no mês atual;
- Salva um .txt com o resultado ao lado do primeiro arquivo informado.

Exemplo:
    python "W:\DOCUMENTOS ESCRITORIO\INSTALACAO SISTEMA\python\RH\COMPARAR EVENTOS HOLERITE SLK.py" `
        --arquivo "C:\Users\User\OneDrive\Documentos\Teste Holerite.slk" `
        --competencia-anterior 01/26 `
        --competencia-atual 02/26
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError as exc:
    print("Erro: modulo 'openpyxl' nao encontrado. Instale com: pip install openpyxl")
    raise SystemExit(1) from exc


# =========================
# Configuracao
# =========================
MESES_PT_BR = {
    "JANEIRO": "01",
    "FEVEREIRO": "02",
    "MARCO": "03",
    "MARÇO": "03",
    "ABRIL": "04",
    "MAIO": "05",
    "JUNHO": "06",
    "JULHO": "07",
    "AGOSTO": "08",
    "SETEMBRO": "09",
    "OUTUBRO": "10",
    "NOVEMBRO": "11",
    "DEZEMBRO": "12",
}
MESES_NUMERO_PARA_NOME = {
    "01": "Janeiro",
    "02": "Fevereiro",
    "03": "Marco",
    "04": "Abril",
    "05": "Maio",
    "06": "Junho",
    "07": "Julho",
    "08": "Agosto",
    "09": "Setembro",
    "10": "Outubro",
    "11": "Novembro",
    "12": "Dezembro",
}
TEXTOS_IGNORADOS = {
    "",
    "NOME",
    "NOME DO FUNCIONARIO",
    "NOME DO FUNCIONÁRIO",
    "CODIGO",
    "CÓDIGO",
    "CBO",
    "EMP.",
    "LOCAL",
    "DEPTO.",
    "SETOR",
    "SECAO",
    "SEÇÃO",
    "FL.",
    "GERAL",
    "CONTINUA",
    "RECIBO DE PAGAMENTO DE SALARIO",
    "RECIBO DE PAGAMENTO DE SALÁRIO",
    "RECIBO DE RETIRADAS",
    "DESCRICAO",
    "DESCRIÇÃO",
    "REFERENCIA",
    "REFERÊNCIA",
    "VENCIMENTOS",
    "DESCONTOS",
    "VALOR LIQUIDO",
    "VALOR LÍQUIDO",
    "TOTAL DE VENCIMENTOS",
    "TOTAL DE DESCONTOS",
    "SALARIO BASE",
    "SALÁRIO BASE",
    "SAL. CONTR. INSS",
    "BASE CALC. FGTS",
    "BASE CÁLC. FGTS",
    "FGTS DO MES",
    "FGTS DO MÊS",
    "BASE CALC. IRRF",
    "BASE CÁLC. IRRF",
    "FAIXA IRRF",
}


@dataclass
class Celula:
    ordem: int
    linha: int
    coluna: int
    valor: str


@dataclass
class EventoParcial:
    codigo: str = ""
    descricao: str = ""
    referencia: str = ""
    valor: str = ""
    coluna_valor: int | None = None


@dataclass(frozen=True)
class Evento:
    codigo: str
    descricao: str
    referencia: str = ""
    valor: str = ""
    tipo: str = ""

    @property
    def chave(self) -> tuple[str, str]:
        return normalizar_codigo(self.codigo), normalizar_texto(self.descricao)


@dataclass
class BlocoHolerite:
    competencia: str
    funcionario: str = ""
    linhas: dict[int, EventoParcial] = field(default_factory=dict)


# =========================
# Utilitarios
# =========================
def remover_acentos(texto: str) -> str:
    return unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")


def normalizar_texto(texto: str) -> str:
    texto = remover_acentos(texto or "").upper()
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def normalizar_codigo(texto: str) -> str:
    return re.sub(r"\D", "", texto or "")


def formatar_competencia(valor: str) -> str:
    valor = valor.strip()
    if re.fullmatch(r"\d{2}/\d{2}", valor):
        return valor
    if re.fullmatch(r"\d{2}/\d{4}", valor):
        return f"{valor[:2]}/{valor[-2:]}"
    if re.fullmatch(r"\d{6}", valor):
        return f"{valor[:2]}/{valor[-2:]}"
    raise ValueError(f"Competência inválida: {valor}")


def competencia_por_extenso(valor: str) -> str:
    competencia = formatar_competencia(valor)
    mes, ano_curto = competencia.split("/")
    ano_completo = f"20{ano_curto}"
    return f"{MESES_NUMERO_PARA_NOME[mes]} / {ano_completo}"


def normalizar_competencia_texto(texto_mes: str) -> str:
    texto_limpo = texto_mes.replace("Mês:", "").replace("Mes:", "").strip()
    texto_limpo = re.sub(r"\s+", " ", texto_limpo)

    match_numerico = re.search(r"(\d{2})/(\d{4})", texto_limpo)
    if match_numerico:
        return f"{match_numerico.group(1)}/{match_numerico.group(2)[-2:]}"

    match_texto = re.search(r"([A-Za-zÇÁÉÍÓÚÃÕÂÊÔà-ÿ]+)\s*/\s*(\d{4})", texto_limpo, flags=re.IGNORECASE)
    if not match_texto:
        raise ValueError(f"Não foi possível interpretar a competência: {texto_mes}")

    mes_nome = normalizar_texto(match_texto.group(1))
    ano = match_texto.group(2)
    if mes_nome not in MESES_PT_BR:
        raise ValueError(f"Mês não reconhecido: {match_texto.group(1)}")
    return f"{MESES_PT_BR[mes_nome]}/{ano[-2:]}"


def ler_linhas_slk(caminho_arquivo: Path) -> list[str]:
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return caminho_arquivo.read_text(encoding=encoding).splitlines()
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("slk", b"", 0, 1, f"Falha ao ler o arquivo: {caminho_arquivo}")


def extrair_valor_k(valor_bruto: str) -> str:
    valor_bruto = valor_bruto.strip()
    if valor_bruto.startswith('"') and valor_bruto.endswith('"'):
        return valor_bruto[1:-1]
    return valor_bruto


def parsear_celulas_slk(caminho_arquivo: Path) -> list[Celula]:
    padrao_f = re.compile(r"^F;(?:[^;]+;)*Y(\d+);X(\d+)$")
    padrao_c = re.compile(r"^C;(?:(?:Y(\d+));)?(?:(?:X(\d+));)?K(.+)$")

    linhas = ler_linhas_slk(caminho_arquivo)
    celulas: list[Celula] = []
    linha_atual = 0
    coluna_atual = 0

    for ordem, conteudo in enumerate(linhas):
        match_f = padrao_f.match(conteudo)
        if match_f:
            linha_atual = int(match_f.group(1))
            coluna_atual = int(match_f.group(2))
            continue

        match_c = padrao_c.match(conteudo)
        if not match_c:
            continue

        if match_c.group(1):
            linha_atual = int(match_c.group(1))
        if match_c.group(2):
            coluna_atual = int(match_c.group(2))

        if linha_atual <= 0 or coluna_atual <= 0:
            continue

        celulas.append(
            Celula(
                ordem=ordem,
                linha=linha_atual,
                coluna=coluna_atual,
                valor=extrair_valor_k(match_c.group(3)),
            )
        )

    return celulas


def eh_nome_funcionario(texto: str) -> bool:
    texto_norm = normalizar_texto(texto)
    if texto_norm in TEXTOS_IGNORADOS:
        return False
    if texto_norm.startswith("ADMISSAO:") or texto_norm.startswith("ADMISSAO:"):
        return False
    if texto_norm.startswith("BANCO :") or texto_norm.startswith("AGENCIA :"):
        return False
    if texto_norm.startswith("DEPOSITO EFETUADO") or texto_norm.startswith("DEPÓSITO EFETUADO"):
        return False
    if len(texto_norm) < 5:
        return False
    return any(char.isalpha() for char in texto_norm)


def eh_codigo_evento(texto: str) -> bool:
    codigo = normalizar_codigo(texto)
    return len(codigo) >= 4


def eh_descricao_evento(texto: str) -> bool:
    texto_norm = normalizar_texto(texto)
    if texto_norm in TEXTOS_IGNORADOS:
        return False
    if texto_norm.startswith("ADMISSAO:") or texto_norm.startswith("ADMISSAO:"):
        return False
    if texto_norm.startswith("BANCO :") or texto_norm.startswith("AGENCIA :"):
        return False
    if texto_norm.startswith("DEPOSITO EFETUADO") or texto_norm.startswith("DEPÓSITO EFETUADO"):
        return False
    return any(char.isalpha() for char in texto_norm)


def eh_valor_numerico(texto: str) -> bool:
    return bool(re.fullmatch(r"-?\d+(?:\.\d+)?", texto.strip()))


def eh_referencia_evento(texto: str) -> bool:
    texto_limpo = texto.strip()
    if not texto_limpo:
        return False

    texto_norm = normalizar_texto(texto_limpo)
    if texto_norm in TEXTOS_IGNORADOS:
        return False

    padroes_validos = (
        r"\d{1,2}/\d{1,2}",
        r"\d{1,2}:\d{2}",
        r"\d{1,3},\d{1,2}%",
        r"\d{1,3},\d{1,2}",
        r"\d{1,3}\.\d{1,2}",
        r"\d{1,2}/\d{2}",
        r"\d{1,2}/\d{4}",
        r"\d+/\d+",
    )
    return any(re.fullmatch(padrao, texto_limpo) for padrao in padroes_validos)


def tipo_evento_por_coluna(coluna: int | None) -> str:
    if coluna == 4:
        return "Vencimento"
    if coluna == 5:
        return "Desconto"
    if coluna == 3:
        return "Valor"
    return ""


def adicionar_eventos_do_bloco(bloco: BlocoHolerite, destino: dict[str, dict[str, dict[tuple[str, str], Evento]]]) -> None:
    if not bloco.funcionario:
        return

    funcionario = bloco.funcionario.strip()
    eventos_funcionario = destino.setdefault(bloco.competencia, {}).setdefault(funcionario, {})

    for parcial in bloco.linhas.values():
        if not eh_codigo_evento(parcial.codigo):
            continue
        if not eh_descricao_evento(parcial.descricao):
            continue
        if normalizar_texto(parcial.descricao) == normalizar_texto(funcionario):
            continue

        evento = Evento(
            codigo=normalizar_codigo(parcial.codigo),
            descricao=parcial.descricao.strip(),
            referencia=parcial.referencia.strip(),
            valor=parcial.valor.strip(),
            tipo=tipo_evento_por_coluna(parcial.coluna_valor),
        )
        eventos_funcionario[evento.chave] = evento


def extrair_eventos_por_competencia(caminhos_arquivos: list[Path]) -> dict[str, dict[str, dict[tuple[str, str], Evento]]]:
    resultado: dict[str, dict[str, dict[tuple[str, str], Evento]]] = {}

    for caminho_arquivo in caminhos_arquivos:
        celulas = parsear_celulas_slk(caminho_arquivo)
        bloco_atual: BlocoHolerite | None = None
        aguardando_funcionario = False

        for celula in celulas:
            valor = celula.valor.strip()
            if not valor:
                continue

            if valor.startswith("Mês:") or valor.startswith("Mes:"):
                if bloco_atual is not None:
                    adicionar_eventos_do_bloco(bloco_atual, resultado)
                bloco_atual = BlocoHolerite(competencia=normalizar_competencia_texto(valor))
                aguardando_funcionario = True
                continue

            if bloco_atual is None:
                continue

            if aguardando_funcionario and celula.coluna == 2 and eh_nome_funcionario(valor):
                bloco_atual.funcionario = valor.strip()
                aguardando_funcionario = False
                continue

            if not bloco_atual.funcionario:
                continue

            parcial = bloco_atual.linhas.setdefault(celula.linha, EventoParcial())

            if celula.coluna == 1 and eh_codigo_evento(valor):
                parcial.codigo = valor
                continue

            if celula.coluna == 2 and eh_descricao_evento(valor):
                parcial.descricao = valor
                continue

            if celula.coluna == 3 and eh_referencia_evento(valor):
                parcial.referencia = valor
                continue

            if eh_valor_numerico(valor) and celula.coluna in (3, 4, 5):
                parcial.valor = valor
                parcial.coluna_valor = celula.coluna

        if bloco_atual is not None:
            adicionar_eventos_do_bloco(bloco_atual, resultado)

    return resultado


def comparar_competencias(
    eventos_por_competencia: dict[str, dict[str, dict[tuple[str, str], Evento]]],
    competencia_anterior: str,
    competencia_atual: str,
) -> list[tuple[str, list[Evento]]]:
    competencia_anterior = formatar_competencia(competencia_anterior)
    competencia_atual = formatar_competencia(competencia_atual)

    if competencia_anterior not in eventos_por_competencia:
        competencias = ", ".join(sorted(eventos_por_competencia))
        raise ValueError(f"Competência anterior não encontrada: {competencia_anterior}. Competências lidas: {competencias}")

    if competencia_atual not in eventos_por_competencia:
        competencias = ", ".join(sorted(eventos_por_competencia))
        raise ValueError(f"Competência atual não encontrada: {competencia_atual}. Competências lidas: {competencias}")

    anterior = eventos_por_competencia[competencia_anterior]
    atual = eventos_por_competencia[competencia_atual]
    faltantes_por_funcionario: list[tuple[str, list[Evento]]] = []

    for funcionario in sorted(anterior):
        eventos_anteriores = anterior[funcionario]
        eventos_atuais = atual.get(funcionario, {})
        chaves_faltantes = sorted(set(eventos_anteriores) - set(eventos_atuais), key=lambda item: (item[1], item[0]))
        if not chaves_faltantes:
            continue
        faltantes_por_funcionario.append(
            (
                funcionario,
                [eventos_anteriores[chave] for chave in chaves_faltantes],
            )
        )

    return faltantes_por_funcionario


def montar_texto_saida(
    caminhos_arquivos: list[Path],
    competencia_anterior: str,
    competencia_atual: str,
    faltantes_por_funcionario: list[tuple[str, list[Evento]]],
) -> str:
    linhas = [
        "RELATÓRIO DE EVENTOS AUSENTES",
        f"Competência anterior: {competencia_anterior}",
        f"Competência atual: {competencia_atual}",
        f"Arquivos analisados: {len(caminhos_arquivos)}",
        "",
    ]

    for caminho_arquivo in caminhos_arquivos:
        linhas.append(str(caminho_arquivo))
    linhas.append("")

    if not faltantes_por_funcionario:
        linhas.append("Nenhum evento do mes anterior ficou ausente no mes atual.")
        return "\n".join(linhas) + "\n"

    for funcionario, eventos in faltantes_por_funcionario:
        linhas.append(f"FUNCIONÁRIO: {funcionario}")
        for evento in eventos:
            linhas.append(f"- {evento.codigo} | {evento.descricao}")
        linhas.append("")

    return "\n".join(linhas).rstrip() + "\n"


def salvar_excel_saida(
    caminho_saida: Path,
    caminhos_arquivos: list[Path],
    competencia_anterior: str,
    competencia_atual: str,
    faltantes_por_funcionario: list[tuple[str, list[Evento]]],
) -> None:
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Eventos Ausentes"

    preenchimento_titulo = PatternFill(fill_type="solid", fgColor="1F4E78")
    preenchimento_cabecalho = PatternFill(fill_type="solid", fgColor="D9EAF7")
    preenchimento_meta = PatternFill(fill_type="solid", fgColor="EEF4F8")
    fonte_branca = Font(color="FFFFFF", bold=True)
    fonte_negrito = Font(bold=True)
    borda_fina = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    ultima_coluna = 3

    linha_atual = 1
    planilha.cell(linha_atual, 1, "RELATÓRIO DE EVENTOS AUSENTES")
    planilha.cell(linha_atual, 1).font = Font(bold=True, color="FFFFFF", size=12)
    planilha.cell(linha_atual, 1).fill = preenchimento_titulo
    planilha.cell(linha_atual, 1).alignment = Alignment(horizontal="center", vertical="center")
    planilha.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=ultima_coluna)
    linha_atual += 2

    planilha.cell(linha_atual, 1, "Competência")
    planilha.cell(linha_atual, 2, competencia_por_extenso(competencia_anterior))
    planilha.cell(linha_atual, 3, competencia_por_extenso(competencia_atual))
    for coluna in (1, 2, 3):
        planilha.cell(linha_atual, coluna).font = fonte_negrito
        planilha.cell(linha_atual, coluna).fill = preenchimento_meta
        planilha.cell(linha_atual, coluna).border = borda_fina
        planilha.cell(linha_atual, coluna).alignment = Alignment(horizontal="center", vertical="center")
    linha_atual += 2

    if not faltantes_por_funcionario:
        planilha.cell(linha_atual, 1, "Nenhum evento do mes anterior ficou ausente no mes atual.")
    else:
        maior_quantidade_eventos = max(len(eventos) for _, eventos in faltantes_por_funcionario)

        planilha.cell(linha_atual, 1, "Funcionário")
        for indice_evento in range(maior_quantidade_eventos):
            planilha.cell(linha_atual, 2 + indice_evento, f"Evento {indice_evento + 1}")

        for coluna in range(1, 2 + maior_quantidade_eventos):
            celula = planilha.cell(linha_atual, coluna)
            celula.font = fonte_branca
            celula.fill = preenchimento_titulo
            celula.border = borda_fina
            celula.alignment = Alignment(horizontal="center", vertical="center")
        linha_atual += 1

        for funcionario, eventos in faltantes_por_funcionario:
            planilha.cell(linha_atual, 1, funcionario)
            planilha.cell(linha_atual, 1).alignment = Alignment(horizontal="left", vertical="center")

            for coluna in range(1, 2 + maior_quantidade_eventos):
                planilha.cell(linha_atual, coluna).border = borda_fina

            for indice_evento, evento in enumerate(eventos, start=2):
                planilha.cell(linha_atual, indice_evento, f"{evento.codigo} - {evento.descricao}")
                planilha.cell(linha_atual, indice_evento).alignment = Alignment(horizontal="left", vertical="center")

            linha_atual += 1

        planilha.column_dimensions["A"].width = 36
        for indice_evento in range(maior_quantidade_eventos):
            letra_coluna = chr(66 + indice_evento)
            planilha.column_dimensions[letra_coluna].width = 28

    for linha in planilha.iter_rows():
        for celula in linha:
            if celula.alignment.horizontal:
                celula.alignment = Alignment(
                    horizontal=celula.alignment.horizontal,
                    vertical="center",
                    wrap_text=True,
                )
            else:
                celula.alignment = Alignment(vertical="center", wrap_text=True)

    workbook.save(caminho_saida)


def listar_resumo(eventos_por_competencia: dict[str, dict[str, dict[tuple[str, str], Evento]]]) -> str:
    linhas = ["Competencias encontradas:"]
    for competencia in sorted(eventos_por_competencia):
        total_funcionarios = len(eventos_por_competencia[competencia])
        total_eventos = sum(len(eventos) for eventos in eventos_por_competencia[competencia].values())
        linhas.append(f"- {competencia}: {total_funcionarios} funcionario(s), {total_eventos} evento(s) unicos")
    return "\n".join(linhas)


# =========================
# Main
# =========================
def montar_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Compara eventos de holerite entre competencias em arquivo .slk")
    parser.add_argument(
        "--arquivo",
        dest="arquivos",
        action="append",
        required=True,
        help="Caminho do arquivo .slk. Pode repetir o parametro para informar mais de um arquivo.",
    )
    parser.add_argument("--competencia-anterior", required=True, help="Competencia anterior. Ex.: 01/26")
    parser.add_argument("--competencia-atual", required=True, help="Competencia atual. Ex.: 02/26")
    parser.add_argument(
        "--saida",
        help="Caminho do .xlsx de saida. Se nao informar, salva ao lado do primeiro arquivo.",
    )
    parser.add_argument(
        "--listar-competencias",
        action="store_true",
        help="Mostra as competencias encontradas antes da comparacao.",
    )
    return parser


def main() -> None:
    parser = montar_parser()
    args = parser.parse_args()

    caminhos_arquivos = [Path(item).expanduser() for item in args.arquivos]
    for caminho_arquivo in caminhos_arquivos:
        if not caminho_arquivo.exists():
            print(f"Arquivo nao encontrado: {caminho_arquivo}")
            sys.exit(1)

    competencia_anterior = formatar_competencia(args.competencia_anterior)
    competencia_atual = formatar_competencia(args.competencia_atual)

    eventos_por_competencia = extrair_eventos_por_competencia(caminhos_arquivos)

    if args.listar_competencias:
        print(listar_resumo(eventos_por_competencia))
        print("")

    faltantes_por_funcionario = comparar_competencias(
        eventos_por_competencia=eventos_por_competencia,
        competencia_anterior=competencia_anterior,
        competencia_atual=competencia_atual,
    )

    if args.saida:
        caminho_saida = Path(args.saida).expanduser()
    else:
        nome_saida = (
            f"{caminhos_arquivos[0].stem}_eventos_ausentes_"
            f"{competencia_anterior.replace('/', '-')}_vs_{competencia_atual.replace('/', '-')}.xlsx"
        )
        caminho_saida = caminhos_arquivos[0].with_name(nome_saida)

    texto_saida = montar_texto_saida(
        caminhos_arquivos=caminhos_arquivos,
        competencia_anterior=competencia_anterior,
        competencia_atual=competencia_atual,
        faltantes_por_funcionario=faltantes_por_funcionario,
    )
    salvar_excel_saida(
        caminho_saida=caminho_saida,
        caminhos_arquivos=caminhos_arquivos,
        competencia_anterior=competencia_anterior,
        competencia_atual=competencia_atual,
        faltantes_por_funcionario=faltantes_por_funcionario,
    )
    print(texto_saida)
    print(f"Arquivo gerado: {caminho_saida}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nOperacao cancelada pelo usuario.")
        sys.exit(1)
    except Exception as exc:
        print(f"Erro: {exc}")
        sys.exit(1)
