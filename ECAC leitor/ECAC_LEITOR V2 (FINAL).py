import re
import time
from datetime import datetime
from pathlib import Path
from copy import copy
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright


# =========================
# CONFIG
# =========================
ECAC_URL = "https://cav.receita.fazenda.gov.br/"

# e-CAC (IDs/seletores)
SEL_BTN_PERFIL = "#btnPerfil"
SEL_INPUT_CNPJ = "#txtNIPapel2"
SEL_BTN_CAIXA = "#btnCaixaPostal"
SEL_DIALOG_PERFIL = "#perfilAcesso"
SEL_BTN_ALTERAR_PJ = "input.submit[type='button'][value='Alterar'][onclick*=\"formPJ\"]"

# ngx-datatable
SEL_ROW = "datatable-body-row"
SEL_CELL = "datatable-body-cell"

# arquivos na mesma pasta do script
INPUT_BASENAME = "Exportacao de Empresas Simples"  # sem extensão

# Caixa Postal: page size (20 -> 100)
CAIXA_PAGE_SIZE_TARGET = 100

# Debug
DEBUG_DUMPS = True
DEBUG_DIRNAME = "_debug_ecac"
DEBUG_LOGS = True

# Automação
MAX_AUTOMATION_RETRIES_PER_DOC = 3
AUTOMATION_DELAY_SECONDS = 33

# Lista de sem procuração
SEM_PROCURACAO_TXT = "sem procuração.txt"


# =========================
# PADRÕES / FAILSAFE
# =========================
DENY_PATTERNS = [
    r"procura[cç][aã]o.*expir",
    r"procura[cç][aã]o.*venc",
    r"certificado.*expir",
    r"detentor.*certificado.*expir",
    r"acesso.*expir",
    r"n[aã]o.*poss[ií]vel.*acessar",
    r"n[aã]o.*autorizad",
    r"n[aã]o.*habilitad",
]
DENY_REGEX = re.compile("|".join(f"(?:{p})" for p in DENY_PATTERNS), re.IGNORECASE | re.DOTALL)

SEM_PROC_PATTERNS = [
    r"n[aã]o\s+existe\s+procura[cç][aã]o\s+eletr[oô]nica",
    r"n[aã]o\s+existe\s+procura[cç][aã]o\s+eletr[oô]nica\s+cadastrad",
    r"sem\s+procura[cç][aã]o\s+eletr[oô]nica",
]
SEM_PROC_REGEX = re.compile("|".join(f"(?:{p})" for p in SEM_PROC_PATTERNS), re.IGNORECASE | re.DOTALL)

BOT_PATTERNS = [
    r"suspeita\s+de\s+automatiza",
    r"atividade\s+suspeita",
    r"comportamento\s+suspeito",
    r"acesso\s+automatiz",
    r"uso\s+automatiz",
    r"detectamos\s+um\s+acesso\s+at[ií]pico",
    r"prote[cç][aã]o\s+contra\s+rob[oô]",
    r"verifica[cç][aã]o\s+de\s+seguran",
    r"valida[cç][aã]o\s+de\s+seguran",
    r"captcha",
    r"recaptcha",
    r"n[aã]o\s+sou\s+um\s+rob[oô]",
]
BOT_REGEX = re.compile("|".join(f"(?:{p})" for p in BOT_PATTERNS), re.IGNORECASE | re.DOTALL)


# =========================
# EXCEÇÕES
# =========================
class SkipDoc(RuntimeError):
    pass


class BotDetected(RuntimeError):
    pass


# =========================
# UTIL
# =========================
def only_digits(s: str) -> str:
    return re.sub(r"\D+", "", str(s or ""))


def format_doc(digits: str) -> str:
    d = only_digits(digits)
    if len(d) == 14:
        return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
    if len(d) == 11:
        return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"
    return digits


def normalize_doc_from_cell(value) -> str | None:
    if value is None:
        return None

    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        if "cpf" in s.lower() and "cnpj" in s.lower():
            return None

        if "e" in s.lower():
            try:
                d = Decimal(s)
                s = format(d.quantize(Decimal(1)), "f")
            except InvalidOperation:
                pass

        digits = only_digits(s)
    else:
        try:
            d = Decimal(str(value))
            digits = only_digits(format(d.quantize(Decimal(1)), "f"))
        except Exception:
            digits = only_digits(str(value))

    if not digits or len(digits) < 11:
        return None

    if len(digits) <= 11:
        digits = digits.zfill(11)
    elif len(digits) <= 14:
        digits = digits.zfill(14)
    else:
        return None

    return digits if len(digits) in (11, 14) else None


def find_input_excel(base_dir: Path) -> Path:
    for ext in (".xlsx", ".xls", ".xlsm"):
        p = base_dir / f"{INPUT_BASENAME}{ext}"
        if p.exists():
            return p
    raise FileNotFoundError(f"Não encontrei '{INPUT_BASENAME}.xlsx' (nem .xls/.xlsm) na pasta: {base_dir}")


def find_template_excel(base_dir: Path, input_path: Path) -> Path:
    candidates = []
    for ext in ("*.xlsx", "*.xlsm"):
        for p in base_dir.glob(ext):
            if p.resolve() == input_path.resolve():
                continue
            if p.name.lower().startswith("caixa_postal_ecac_"):
                continue
            if p.name.lower().startswith("controle_verificacao"):
                continue
            candidates.append(p)

    if not candidates:
        raise FileNotFoundError(
            "Não achei nenhum arquivo .xlsx/.xlsm de modelo na pasta.\n"
            "Coloque sua planilha de exemplo (template) na mesma pasta do script."
        )

    preferred = [p for p in candidates if any(k in p.name.lower() for k in ("modelo", "exemplo", "template"))]
    pick_from = preferred if preferred else candidates
    pick_from.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return pick_from[0]


def read_docs_from_input_excel(path: Path) -> list[str]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    docs = []
    seen = set()

    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        doc = normalize_doc_from_cell(v)
        if doc and doc not in seen:
            seen.add(doc)
            docs.append(doc)

    return docs


def read_doc_name_map_from_input_excel(path: Path) -> dict[str, str]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    mp: dict[str, str] = {}

    for r in range(2, ws.max_row + 1):
        doc_raw = ws.cell(r, 1).value
        name_raw = ws.cell(r, 2).value

        digits = normalize_doc_from_cell(doc_raw)
        if not digits:
            continue

        name = str(name_raw).strip() if name_raw is not None else ""
        if digits not in mp:
            mp[digits] = name
        else:
            if (not mp[digits]) and name:
                mp[digits] = name

    return mp


# =========================
# DEBUG
# =========================
def _debug_dir(base_dir: Path) -> Path:
    d = base_dir / DEBUG_DIRNAME
    d.mkdir(parents=True, exist_ok=True)
    return d


def dump_debug(page, base_dir: Path, label: str):
    if not DEBUG_DUMPS:
        return
    try:
        d = _debug_dir(base_dir)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_label = re.sub(r"[^a-zA-Z0-9_\-]+", "_", label)[:90]
        png = d / f"{ts}_{safe_label}.png"
        html = d / f"{ts}_{safe_label}.html"
        try:
            page.screenshot(path=str(png), full_page=True)
        except Exception:
            pass
        try:
            html.write_text(page.content() or "", encoding="utf-8")
        except Exception:
            pass
    except Exception:
        pass


# =========================
# ANTI-BOT / TEXT
# =========================
def _is_bot_text(txt: str) -> bool:
    return bool(txt and BOT_REGEX.search(txt))


def _is_deny_text(txt: str) -> bool:
    return bool(txt and DENY_REGEX.search(txt))


def _is_sem_procuracao_text(txt: str) -> bool:
    return bool(txt and SEM_PROC_REGEX.search(txt))


def iter_scopes(page):
    yield page
    try:
        for fr in page.frames:
            yield fr
    except Exception:
        pass


def scope_text(scope) -> str:
    try:
        return (scope.evaluate("() => document.body ? document.body.innerText : ''") or "").strip()
    except Exception:
        return ""


def _bot_guard_check_page(page, base_dir: Path, label: str):
    try:
        txt = scope_text(page)
        if _is_bot_text(txt):
            dump_debug(page, base_dir, f"bot_{label}")
            raise BotDetected("Suspeita de automação detectada na página.")
    except BotDetected:
        raise
    except Exception:
        return


def _append_sem_procuracao_txt(base_dir: Path, digits: str):
    p = base_dir / SEM_PROCURACAO_TXT
    formatted = format_doc(digits)

    existing = set()
    if p.exists():
        try:
            existing = set(x.strip() for x in p.read_text(encoding="utf-8").splitlines() if x.strip())
        except Exception:
            existing = set()

    if formatted not in existing:
        with p.open("a", encoding="utf-8") as f:
            f.write(formatted + "\n")


# =========================
# PLAYWRIGHT HELPERS
# =========================
def safe_click(page, selector: str, timeout=15000, retries: int = 1):
    last_err = None
    for _ in range(max(1, retries)):
        try:
            loc = page.locator(selector).first
            loc.wait_for(state="visible", timeout=timeout)
            loc.click()
            return
        except Exception as e:
            last_err = e
            try:
                page.wait_for_timeout(250)
            except Exception:
                time.sleep(0.25)
    raise last_err


def wait_input_settled(page, locator, digits: str):
    try:
        page.wait_for_timeout(200)
    except Exception:
        time.sleep(0.2)

    for _ in range(20):
        try:
            v = locator.input_value(timeout=500)
        except Exception:
            v = ""
        if (v == digits) or (only_digits(v) == digits):
            return
        try:
            page.wait_for_timeout(100)
        except Exception:
            time.sleep(0.1)


def _close_perfil_dialog(page):
    """
    Fecha o modal 'Alterar perfil de acesso'.

    Importante: o botão X normalmente NÃO fica dentro de #perfilAcesso.
    Ele fica no wrapper do jQuery UI: div.ui-dialog:has(#perfilAcesso)
    """
    try:
        wrapper = page.locator(f"div.ui-dialog:has({SEL_DIALOG_PERFIL})").first
        if wrapper.count() <= 0:
            wrapper = page.locator(SEL_DIALOG_PERFIL).first

        close_selectors = [
            ".ui-dialog-titlebar-close",
            "a.ui-dialog-titlebar-close",
            "button.ui-dialog-titlebar-close",
            ".ui-dialog-titlebar .ui-dialog-titlebar-close",
            ".ui-dialog-titlebar-close .ui-icon",
            "button[title*='Fechar' i]",
            "a[title*='Fechar' i]",
        ]

        for sel in close_selectors:
            try:
                btn = wrapper.locator(sel).first
                if btn.count() > 0:
                    btn.click(force=True, timeout=1500)
                    try:
                        page.locator(SEL_DIALOG_PERFIL).first.wait_for(state="hidden", timeout=6000)
                        return
                    except Exception:
                        pass
                    try:
                        page.locator(f"div.ui-dialog:has({SEL_DIALOG_PERFIL})").first.wait_for(state="hidden", timeout=6000)
                        return
                    except Exception:
                        pass
                    return
            except Exception:
                pass

        # fallback: clicar no canto superior direito do wrapper (onde costuma ficar o X)
        try:
            box = wrapper.bounding_box()
            if box:
                x = box["x"] + box["width"] - 12
                y = box["y"] + 12
                page.mouse.click(x, y)
                try:
                    page.locator(SEL_DIALOG_PERFIL).first.wait_for(state="hidden", timeout=4000)
                    return
                except Exception:
                    pass
        except Exception:
            pass

        # último fallback: ESC
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass

    except Exception:
        pass


def switch_profile_to_doc(page, digits: str, base_dir: Path):
    """
    Troca o perfil no e-CAC para o CNPJ/CPF informado.

    FAILSAFE:
      - Se aparecer a janelinha de erro (div.erro / p.mensagemErro), fecha no X e pausa 33s
      - "Não existe procuração eletrônica..." -> grava em 'sem procuração.txt', pausa 33s e pula
      - "procuração ... expirou" / acesso negado -> pausa 33s e pula
      - suspeita de automação -> BotDetected
    """
    _bot_guard_check_page(page, base_dir, f"before_perfil_{digits}")

    safe_click(page, SEL_BTN_PERFIL, timeout=15000, retries=2)

    dialog = page.locator(SEL_DIALOG_PERFIL).first
    dialog.wait_for(state="visible", timeout=15000)

    # Se o e-CAC levantar uma caixa de erro explícita dentro do diálogo (ex.: procuração expirada),
    # fecha, pausa 33s e pula.
    try:
        err_loc = dialog.locator("div.erro:visible p.mensagemErro").first
        if err_loc.count() > 0:
            msg_err = (err_loc.inner_text() or "").strip()
            _close_perfil_dialog(page)
            _pause_on_dialog_error(page, base_dir, digits, reason="dialog_erro", msg=msg_err)
            raise SkipDoc(msg_err or "Erro no diálogo de perfil.")
    except SkipDoc:
        raise
    except Exception:
        pass

    # checagem imediata
    try:
        t0 = (dialog.inner_text() or "").strip()

        if _is_bot_text(t0):
            dump_debug(page, base_dir, f"bot_perfil_{digits}")
            _close_perfil_dialog(page)
            raise BotDetected("Suspeita de automação no diálogo de perfil.")

        if _is_sem_procuracao_text(t0):
            _append_sem_procuracao_txt(base_dir, digits)
            _close_perfil_dialog(page)
            _pause_on_dialog_error(page, base_dir, digits, reason="sem_procuracao", msg=t0)
            raise SkipDoc("Sem procuração eletrônica (diálogo de perfil).")

        if _is_deny_text(t0):
            _close_perfil_dialog(page)
            _pause_on_dialog_error(page, base_dir, digits, reason="deny", msg=t0)
            raise SkipDoc("Procuração expirada / acesso negado (diálogo de perfil).")

    except (BotDetected, SkipDoc):
        raise
    except Exception:
        pass

    inp = dialog.locator(SEL_INPUT_CNPJ).first
    inp.wait_for(state="visible", timeout=15000)
    inp.click()
    inp.fill(digits)
    wait_input_settled(page, inp, digits)

    btn_pj = dialog.locator(SEL_BTN_ALTERAR_PJ).first
    if btn_pj.count() > 0:
        btn_pj.click()
    else:
        dialog.locator("input.submit[type='button'][value='Alterar']").last.click()

    try:
        dialog.wait_for(state="hidden", timeout=12000)
        return
    except Exception:
        pass

    try:
        t1 = (dialog.inner_text() or "").strip()
    except Exception:
        t1 = ""

    # pode ter levantado erro após clicar em Alterar
    try:
        err_loc2 = dialog.locator("div.erro:visible p.mensagemErro").first
        if err_loc2.count() > 0:
            msg_err2 = (err_loc2.inner_text() or "").strip()
            _close_perfil_dialog(page)
            _pause_on_dialog_error(page, base_dir, digits, reason="dialog_erro_after", msg=msg_err2)
            raise SkipDoc(msg_err2 or "Erro no diálogo de perfil (após alterar).")
    except SkipDoc:
        raise
    except Exception:
        pass

    if _is_bot_text(t1):
        dump_debug(page, base_dir, f"bot_perfil_after_{digits}")
        _close_perfil_dialog(page)
        raise BotDetected("Suspeita de automação após tentar trocar perfil.")

    if _is_sem_procuracao_text(t1):
        _append_sem_procuracao_txt(base_dir, digits)
        _close_perfil_dialog(page)
        _pause_on_dialog_error(page, base_dir, digits, reason="sem_procuracao_after", msg=t1)
        raise SkipDoc("Sem procuração eletrônica (ao trocar perfil).")

    if _is_deny_text(t1):
        _close_perfil_dialog(page)
        _pause_on_dialog_error(page, base_dir, digits, reason="deny_after", msg=t1)
        raise SkipDoc("Procuração expirada / acesso negado (ao trocar perfil).")

    dump_debug(page, base_dir, f"perfil_nao_fechou_{digits}")
    _close_perfil_dialog(page)
    _pause_on_dialog_error(page, base_dir, digits, reason="dialog_nao_fechou", msg=t1)
    raise RuntimeError("Não consegui trocar o perfil (diálogo não fechou).")
def _scope_has_any_table_like(scope) -> bool:
    try:
        if scope.locator("ngx-datatable").count() > 0:
            return True
        if scope.locator("datatable-body").count() > 0:
            return True
        if scope.locator(SEL_ROW).count() > 0:
            return True
        if scope.locator("table").count() > 0:
            return True
        if scope.locator("[role=grid]").count() > 0:
            return True
    except Exception:
        pass
    return False


def wait_for_caixa_any(page, base_dir: Path, timeout_ms: int = 120000):
    """
    Retorna (scope, state):
      - ok: encontrou tabela/grade
      - empty: estado vazio detectado
      - deny: acesso negado/expirado/sem procuração detectado
    """
    t0 = time.time()
    empty_patterns = re.compile(r"(nenhuma|nenhum).*(mensagem|item)|0\s*itens|sem\s+mensagens", re.IGNORECASE)

    while (time.time() - t0) < (timeout_ms / 1000):
        _bot_guard_check_page(page, base_dir, "wait_caixa")

        for sc in iter_scopes(page):
            try:
                txt = scope_text(sc)

                if _is_bot_text(txt):
                    dump_debug(page, base_dir, "bot_caixa")
                    raise BotDetected("Suspeita de automação na Caixa Postal.")

                if _is_sem_procuracao_text(txt) or _is_deny_text(txt):
                    return sc, "deny"

                if empty_patterns.search(txt or ""):
                    if _scope_has_any_table_like(sc):
                        return sc, "ok"
                    return sc, "empty"

                if _scope_has_any_table_like(sc):
                    return sc, "ok"
            except BotDetected:
                raise
            except Exception:
                continue

        try:
            page.wait_for_timeout(350)
        except Exception:
            time.sleep(0.35)

    dump_debug(page, base_dir, "timeout_caixa")
    return page, "deny"


def goto_caixa_postal(page, base_dir: Path, timeout_ms=120000):
    """
    Entra na Caixa Postal e retorna: (page, scope, state)
    """
    ctx = page.context
    new_page = None

    try:
        with ctx.expect_page(timeout=5000) as pinfo:
            try:
                safe_click(page, SEL_BTN_CAIXA, timeout=15000, retries=2)
            except Exception:
                link = page.locator("a[href*='/caixapostal']").first
                link.wait_for(state="visible", timeout=15000)
                link.click()
        new_page = pinfo.value
    except Exception:
        try:
            safe_click(page, SEL_BTN_CAIXA, timeout=15000, retries=2)
        except Exception:
            link = page.locator("a[href*='/caixapostal']").first
            link.wait_for(state="visible", timeout=15000)
            link.click()

    if new_page:
        page = new_page
        try:
            page.bring_to_front()
        except Exception:
            pass

    _bot_guard_check_page(page, base_dir, "after_click_caixa")

    try:
        page.wait_for_load_state("domcontentloaded", timeout=25000)
    except Exception:
        pass
    try:
        page.wait_for_url("**/caixapostal**", timeout=25000)
    except Exception:
        pass

    _bot_guard_check_page(page, base_dir, "after_url_caixa")

    scope, state = wait_for_caixa_any(page, base_dir=base_dir, timeout_ms=timeout_ms)
    return page, scope, state


# =========================
# PAGE SIZE (20 -> 100)  [CORRIGIDO: NÃO CLICA NO FILTRO]
# =========================
def _get_page_from_scope(scope):
    try:
        if hasattr(scope, "page"):
            return scope.page
    except Exception:
        pass
    return None


def _find_pagesize_ngselect(scope):
    """
    Encontra o ng-select do 'Exibir:' (tamanho da página).
    Evita pegar ng-select de filtro (ex.: 'Todos').
    """
    page = _get_page_from_scope(scope)

    # 1) âncora no texto "Exibir"
    for container in (scope, page) if page else (scope,):
        try:
            ex = container.locator("text=/\\bExibir\\b/i").first
            if ex.count() > 0:
                ns = ex.locator("xpath=following::ng-select[1]").first
                if ns.count() > 0:
                    val = ns.locator(".ng-value-label").first
                    if val.count() > 0:
                        t = (val.inner_text() or "").strip()
                        if t.isdigit():
                            return ns
                    return ns
        except Exception:
            pass

    # 2) no footer do datatable
    candidates = [
        "datatable-footer ng-select.ng-select-single",
        ".datatable-footer ng-select.ng-select-single",
        "datatable-footer ng-select",
        ".datatable-footer ng-select",
    ]
    for sel in candidates:
        try:
            ns = scope.locator(sel).first
            if ns.count() > 0:
                val = ns.locator(".ng-value-label").first
                if val.count() > 0:
                    t = (val.inner_text() or "").strip()
                    if t.isdigit():
                        return ns
        except Exception:
            pass

    # 3) último recurso: ng-select com valor 10/20/50/100
    try:
        all_ns = scope.locator("ng-select.ng-select-single")
        for i in range(min(all_ns.count(), 20)):
            ns = all_ns.nth(i)
            val = ns.locator(".ng-value-label").first
            if val.count() > 0:
                t = (val.inner_text() or "").strip()
                if t in ("10", "20", "50", "100"):
                    return ns
    except Exception:
        pass

    return None


def _click_ngselect_open(scope, ngsel) -> bool:
    if ngsel is None:
        return False
    try:
        arrow = ngsel.locator(".ng-arrow-wrapper").first
        if arrow.count() > 0:
            arrow.click()
            return True
    except Exception:
        pass
    try:
        ngsel.click()
        return True
    except Exception:
        return False


def set_caixa_page_size(scope, target: int, base_dir: Path, label: str = "") -> bool:
    """
    Ajusta o "Exibir:" para 100.
    (Corrigido para não abrir o filtro errado.)
    """
    target_txt = str(target)
    page = _get_page_from_scope(scope)

    ngsel = _find_pagesize_ngselect(scope)
    if ngsel is None:
        if DEBUG_DUMPS and page:
            dump_debug(page, base_dir, f"pagesize_ngselect_not_found_{label}")
        return False

    # já está no alvo?
    try:
        cur = (ngsel.locator(".ng-value-label").first.inner_text() or "").strip()
        if cur == target_txt:
            return True
    except Exception:
        pass

    if not _click_ngselect_open(scope, ngsel):
        if DEBUG_DUMPS and page:
            dump_debug(page, base_dir, f"pagesize_open_fail_{label}")
        return False

    option_label_selectors = [
        ".ng-dropdown-panel .ng-option .ng-option-label",
        ".ng-dropdown-panel .ng-option",
        "div.ng-option .ng-option-label",
        "div.ng-option",
        "span.ng-option-label",
        "[role=option]",
    ]

    clicked = False
    for container in (scope, page) if page else (scope,):
        for osel in option_label_selectors:
            try:
                opts = container.locator(osel)
                cnt = opts.count()
                for i in range(min(cnt, 120)):
                    o = opts.nth(i)
                    t = (o.inner_text() or "").strip()
                    if t == target_txt:
                        o.click()
                        clicked = True
                        break
                if clicked:
                    break
            except Exception:
                pass
        if clicked:
            break

    if not clicked:
        try:
            if page:
                page.keyboard.press("Escape")
        except Exception:
            pass
        if DEBUG_DUMPS and page:
            dump_debug(page, base_dir, f"pagesize_option_not_found_{label}_{target}")
        return False

    # espera valor mudar
    t0 = time.time()
    while time.time() - t0 < 8:
        try:
            cur = (ngsel.locator(".ng-value-label").first.inner_text() or "").strip()
            if cur == target_txt:
                if DEBUG_LOGS:
                    print(f"[CAIXA] page-size ajustado para {target} ({label})")
                return True
        except Exception:
            pass
        try:
            scope.wait_for_timeout(200)
        except Exception:
            time.sleep(0.2)

    if DEBUG_DUMPS and page:
        dump_debug(page, base_dir, f"pagesize_nochange_{label}_{target}")
    return False


# =========================
# PAGINAÇÃO
# =========================
def _get_caixa_pager_text(scope) -> str:
    """
    Lê o texto de paginação/range no rodapé.
    Ex.: "1-100 de 151 itens"
    """
    # 1) pega direto o texto do range (mais confiável)
    try:
        loc = scope.locator(r"text=/\d+\s*[-–]\s*\d+\s+de\s+\d+\s+itens/i").first
        if loc.count() > 0:
            return (loc.inner_text() or "").strip()
    except Exception:
        pass

    # 2) fallback: tenta o footer inteiro
    try:
        t = (scope.locator("datatable-footer").inner_text() or "").strip()
        if t:
            return t
    except Exception:
        pass

    # 3) último fallback: qualquer texto do rodapé/pager
    try:
        t = (scope.locator("datatable-footer, datatable-pager, .datatable-footer, .datatable-pager").first.inner_text() or "").strip()
        return t
    except Exception:
        return ""
    
def _parse_caixa_range(pager_text: str):
    if not pager_text:
        return None
    m = re.search(r"(\d+)\s*-\s*(\d+)\s*de\s*(\d+)\s*iten", pager_text, flags=re.I)
    if not m:
        return None
    a, b, c = int(m.group(1)), int(m.group(2)), int(m.group(3))
    return a, b, c


def _get_current_page_number(scope) -> int | None:
    sels = [
        ".datatable-pager li.active",
        ".datatable-pager li.active a",
        ".datatable-pager a.active",
        ".datatable-pager .pages .active",
        ".datatable-pager .pages a.active",
    ]
    for sel in sels:
        try:
            loc = scope.locator(sel).first
            if loc.count() > 0:
                t = (loc.inner_text() or "").strip()
                if t.isdigit():
                    return int(t)
        except Exception:
            pass
    return None


def _click_next_page(scope, base_dir: Path, label: str = "", timeout_ms: int = 20000) -> bool:
    """
    Clica em "Próxima página" e espera mudar o range do pager (ex.: 1-100 -> 101-151).
    Retorna False se não houver próxima página (botão desabilitado ou inexistente).
    """
    page = _get_page_from_scope(scope)

    before_txt = _get_caixa_pager_text(scope)
    before = _parse_caixa_range(before_txt) or None

    # Seletores atualizados (e-CAC atual: botão com aria-label="Página seguinte")
    selectors = [
        "button[aria-label='Página seguinte']",
        "button[aria-label*='Página seguinte' i]",
        "button.br-button.circle:has(i.fas.fa-angle-right)",
        "button:has(i.fas.fa-angle-right)",
        # Fallbacks antigos (mantém compatibilidade)
        "datatable-pager .pager .pages a[aria-label*='Next' i]",
        "datatable-pager .pager .pages a:has-text('>')",
        "datatable-pager .pager .pages a:has-text('›')",
        "datatable-pager a[rel='next']",
        "a[aria-label*='próxima' i]",
        "a[aria-label*='proxima' i]",
    ]

    btn = None
    for sel in selectors:
        try:
            cand = scope.locator(sel).first
            if cand.count() > 0:
                btn = cand
                break
        except Exception:
            continue

    if btn is None:
        return False

    # Se estiver desabilitado, não tem próxima página
    try:
        if btn.is_disabled():
            return False
    except Exception:
        pass

    try:
        aria_dis = (btn.get_attribute("aria-disabled") or "").strip().lower()
        if aria_dis in ("true", "1"):
            return False
    except Exception:
        pass

    # Clica
    try:
        btn.click(timeout=1500)
    except Exception:
        try:
            btn.click(force=True, timeout=1500)
        except Exception:
            dump_debug(page, base_dir, f"next_click_fail_{label}")
            return True  # tentou clicar, mas não confirmou; deixa o loop decidir

    # Espera o range mudar
    start = time.time()
    while (time.time() - start) * 1000 < timeout_ms:
        now_txt = _get_caixa_pager_text(scope)
        now = _parse_caixa_range(now_txt) or None

        # Se conseguimos parsear antes e agora, espera diferença real
        if before and now and now != before:
            return True

        # Se não parseou, mas o texto mudou, já é sinal forte
        if before_txt and now_txt and now_txt.strip() != before_txt.strip():
            return True

        try:
            page.wait_for_timeout(300)
        except Exception:
            time.sleep(0.3)

    dump_debug(page, base_dir, f"pager_timeout_{label}")
    return True
# =========================
# SCRAPING
# =========================
def norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("ª", "a").replace("º", "o")
    s = s.replace("ç", "c")
    s = s.replace("á", "a").replace("à", "a").replace("ã", "a").replace("â", "a")
    s = s.replace("é", "e").replace("ê", "e")
    s = s.replace("í", "i")
    s = s.replace("ó", "o").replace("ô", "o").replace("õ", "o")
    s = s.replace("ú", "u")
    return s


def _find_datatable_scroller(scope):
    for sel in (
        "datatable-body",
        ".datatable-body",
        ".datatable-body-wrapper",
        ".datatable-scroll",
        "[role=grid]",
        "[role=table]",
    ):
        try:
            loc = scope.locator(sel).first
            if loc.count() > 0:
                return loc
        except Exception:
            pass
    return None


def scrape_ng_datatable(scope, max_rounds=220, stable_rounds=8):
    headers = scope.locator("datatable-header-cell")
    hn = headers.count()
    header_map = {}
    if hn > 0:
        for i in range(hn):
            t = (headers.nth(i).inner_text() or "").strip()
            k = norm_header(t)
            if k:
                header_map[k] = i

    def idx_of(*keys, default=None):
        for k in keys:
            for hk, hv in header_map.items():
                if k in hk:
                    return hv
        return default

    col_aco = idx_of("acao", "acoes", default=1)
    col_rem = idx_of("remetente", default=2)
    col_ass = idx_of("assunto", default=3)
    col_env = idx_of("enviada em", default=4)
    col_exi = idx_of("exibicao ate", "exibição ate", "exibição até", default=5)
    col_lei = idx_of("data de 1a leitura", "data de 1ª leitura", default=6)
    col_des = idx_of("destinatario", default=7)
    col_idm = idx_of("id mensagem", "mensagem", "numero", "nº", "no", default=8)

    scroller = _find_datatable_scroller(scope) or scope.locator("body").first

    def cell_text(cells, idx):
        try:
            if idx is None:
                return ""
            return (cells.nth(idx).inner_text() or "").strip()
        except Exception:
            return ""

    def cell_link_text(cells, idx):
        try:
            if idx is None:
                return ""
            a = cells.nth(idx).locator("a").first
            if a.count() > 0:
                return (a.inner_text() or "").strip()
            return (cells.nth(idx).inner_text() or "").strip()
        except Exception:
            return ""

    results = {}
    stable = 0

    for _ in range(max_rounds):
        txt = scope_text(scope)
        if _is_bot_text(txt):
            raise BotDetected("Sinal de automação durante scraping (datatable).")

        rows = scope.locator(SEL_ROW)
        rc = rows.count()

        before = len(results)

        for i in range(rc):
            r = rows.nth(i)
            cells = r.locator(SEL_CELL)
            n = cells.count()
            if n < 4:
                continue

            msg = {
                "Ações": cell_text(cells, col_aco),
                "Remetente": cell_link_text(cells, col_rem),
                "Assunto": cell_link_text(cells, col_ass),
                "Enviada em:": cell_text(cells, col_env),
                "Exibição até:": cell_text(cells, col_exi),
                "Data de 1ª leitura": cell_text(cells, col_lei),
                "Destinatário": cell_text(cells, col_des),
                "ID mensagem": cell_text(cells, col_idm),
            }

            key = msg.get("ID mensagem") or f"{msg.get('Remetente','')}|{msg.get('Assunto','')}|{msg.get('Enviada em:','')}|{i}"
            if key and key not in results:
                results[key] = msg

        try:
            scope.evaluate("(el) => { el.scrollTop = el.scrollTop + el.clientHeight * 0.9; }", scroller)
        except Exception:
            try:
                scope.evaluate("(el) => { el.scrollTop = el.scrollTop + 1200; }", scroller)
            except Exception:
                pass

        try:
            scope.wait_for_timeout(300)
        except Exception:
            time.sleep(0.3)

        after = len(results)
        stable = stable + 1 if after == before else 0
        if stable >= stable_rounds:
            break

    return list(results.values())


def scrape_html_table(scope):
    try:
        table = scope.locator("table").first
        if table.count() <= 0:
            return []
        rows = table.locator("tr")
        rc = rows.count()
        if rc <= 1:
            return []

        head = rows.nth(0).locator("th,td")
        hc = head.count()
        headers = [norm_header((head.nth(i).inner_text() or "").strip()) for i in range(hc)]

        def idx(*keys, default=None):
            for k in keys:
                for i, h in enumerate(headers):
                    if k in h:
                        return i
            return default

        col_rem = idx("remetente", default=0)
        col_ass = idx("assunto", default=1)
        col_env = idx("enviada em", default=2)
        col_exi = idx("exibicao ate", "exibição até", default=3)
        col_idm = idx("id mensagem", "mensagem", "numero", default=4)

        out = []
        for r in range(1, rc):
            cells = rows.nth(r).locator("td")
            cc = cells.count()
            if cc <= 1:
                continue

            def g(i):
                if i is None or i >= cc:
                    return ""
                return (cells.nth(i).inner_text() or "").strip()

            out.append(
                {
                    "Ações": "",
                    "Remetente": g(col_rem),
                    "Assunto": g(col_ass),
                    "Enviada em:": g(col_env),
                    "Exibição até:": g(col_exi),
                    "Data de 1ª leitura": "",
                    "Destinatário": "",
                    "ID mensagem": g(col_idm),
                }
            )
        return out
    except Exception:
        return []


def scrape_role_grid(scope, max_rounds=220, stable_rounds=8):
    headers = scope.locator("[role=columnheader]")
    hn = headers.count()
    header_map = {}
    if hn > 0:
        for i in range(hn):
            t = (headers.nth(i).inner_text() or "").strip()
            k = norm_header(t)
            if k:
                header_map[k] = i

    def idx_of(*keys, default=None):
        for k in keys:
            for hk, hv in header_map.items():
                if k in hk:
                    return hv
        return default

    col_rem = idx_of("remetente", default=0)
    col_ass = idx_of("assunto", default=1)
    col_env = idx_of("enviada em", default=2)
    col_exi = idx_of("exibicao ate", "exibição até", default=3)
    col_idm = idx_of("id mensagem", "mensagem", "numero", default=4)

    scroller = _find_datatable_scroller(scope) or scope.locator("body").first

    def read_row(row):
        cells = row.locator("[role=gridcell]")
        n = cells.count()

        def g(i):
            if i is None or i >= n:
                return ""
            a = cells.nth(i).locator("a").first
            if a.count() > 0:
                return (a.inner_text() or "").strip()
            return (cells.nth(i).inner_text() or "").strip()

        return {
            "Ações": "",
            "Remetente": g(col_rem),
            "Assunto": g(col_ass),
            "Enviada em:": g(col_env),
            "Exibição até:": g(col_exi),
            "Data de 1ª leitura": "",
            "Destinatário": "",
            "ID mensagem": g(col_idm),
        }

    results = {}
    stable = 0
    for _ in range(max_rounds):
        rows = scope.locator("[role=row]")
        rc = rows.count()
        before = len(results)

        for i in range(rc):
            msg = read_row(rows.nth(i))
            key = msg.get("ID mensagem") or f"{msg.get('Remetente','')}|{msg.get('Assunto','')}|{msg.get('Enviada em:','')}|{i}"
            if key and key not in results and (msg.get("Remetente") or msg.get("Assunto")):
                results[key] = msg

        try:
            scope.evaluate("(el) => { el.scrollTop = el.scrollTop + el.clientHeight * 0.9; }", scroller)
        except Exception:
            try:
                scope.evaluate("(el) => { el.scrollTop = el.scrollTop + 1200; }", scroller)
            except Exception:
                pass

        try:
            scope.wait_for_timeout(300)
        except Exception:
            time.sleep(0.3)

        after = len(results)
        stable = stable + 1 if after == before else 0
        if stable >= stable_rounds:
            break

    return list(results.values())


def _reset_caixa_scroll_top(scope):
    for sel in (
        "datatable-body",
        ".datatable-body",
        ".datatable-body-wrapper",
        ".datatable-scroll",
        "[role=grid]",
        "[role=table]",
        "body",
    ):
        try:
            loc = scope.locator(sel).first
            if loc.count() > 0:
                scope.evaluate("(el) => { try { el.scrollTop = 0; } catch(e){} }", loc)
                return True
        except Exception:
            pass
    return False


def scrape_messages(scope, base_dir: Path, label: str = "", max_pages: int = 50):
    """
    - coleta mensagens na página atual
    - usa pager "1-100 de 151 itens" e avança páginas
    """
    def _collect_once():
        txt = scope_text(scope)
        if _is_bot_text(txt):
            raise BotDetected("Sinal de automação antes do scraping.")

        msgs = []
        try:
            msgs = scrape_ng_datatable(scope)
        except Exception:
            msgs = []

        if not msgs:
            msgs = scrape_html_table(scope)

        if not msgs:
            msgs = scrape_role_grid(scope)

        return msgs or []

    all_by_id = {}
    pages_done = 0

    while True:
        msgs = _collect_once()

        for d in msgs:
            k = (d.get("ID mensagem") or "").strip()
            if not k:
                k = f"{d.get('Remetente','')}|{d.get('Assunto','')}|{d.get('Enviada em:','')}|{len(all_by_id)+1}"
            if k not in all_by_id:
                all_by_id[k] = d

        pages_done += 1
        if pages_done >= max_pages:
            if DEBUG_LOGS:
                print(f"[CAIXA] limite de páginas atingido ({max_pages}) ({label}).")
            break

        advanced = _click_next_page(scope, base_dir, label=label, timeout_ms=20000)
        if not advanced:
            break

        try:
            scope.wait_for_timeout(900)
        except Exception:
            time.sleep(0.9)

        _reset_caixa_scroll_top(scope)

    if DEBUG_LOGS:
        print(f"[CAIXA] total coletado={len(all_by_id)} ({label})")
    return list(all_by_id.values())


# =========================
# TOTAL DO DIA (E TOTAL GERAL)
# =========================
def compute_total_today(msgs: list[dict]) -> int:
    """
    TOTAL = total de mensagens extraídas (o que o script já coleta).
    Isto preenche {{TOTAL}} no modelo e a coluna TOTAL no controle.
    """
    try:
        return int(len(msgs or []))
    except Exception:
        return 0

# =========================
# EXCEL OUTPUT (TEMPLATE {{CNPJ}} + {{NOME}} + {{TOTAL}})
# =========================
def find_row_by_exact_value(ws, value: str):
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(r, c).value or "").strip() == value:
                return r, c
    return None, None


def find_header_row(ws):
    for r in range(1, ws.max_row + 1):
        vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        if "remetente" in vals and "assunto" in vals:
            return r
    return None


def detect_template_rows(ws):
    row_cnpj, col_cnpj = find_row_by_exact_value(ws, "{{CNPJ}}")
    row_nome, col_nome = find_row_by_exact_value(ws, "{{NOME}}")  # opcional
    row_data, _ = find_row_by_exact_value(ws, "XXXX")
    row_header = find_header_row(ws)

    if not row_cnpj:
        raise RuntimeError("Modelo: não encontrei o marcador {{CNPJ}}.")
    if not row_data:
        raise RuntimeError("Modelo: não encontrei o marcador XXXX (linha que deve repetir).")
    if not row_header:
        raise RuntimeError("Modelo: não encontrei a linha de cabeçalhos (precisa ter 'Remetente' e 'Assunto').")

    row_blank = row_header + 1
    is_blank = True
    if row_blank <= ws.max_row:
        for c in range(1, ws.max_column + 1):
            if ws.cell(row_blank, c).value not in (None, ""):
                is_blank = False
                break
    if not is_blank:
        row_blank = row_data

    return (row_cnpj, col_cnpj, row_nome, col_nome, row_header, row_blank, row_data)


def find_merge_span_for_cell(ws, row: int, col: int):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_col, rng.max_col
    return col, col


def copy_cell_style(dst_cell, src_cell):
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy(src_cell.protection)


def copy_row_style(dst_ws, src_ws, src_row: int, dst_row: int, max_col: int):
    if src_ws.row_dimensions[src_row].height is not None:
        dst_ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height
    for c in range(1, max_col + 1):
        copy_cell_style(dst_ws.cell(dst_row, c), src_ws.cell(src_row, c))


def apply_column_widths(dst_ws, src_ws, max_col: int):
    for c in range(1, max_col + 1):
        col = get_column_letter(c)
        dst_ws.column_dimensions[col].width = src_ws.column_dimensions[col].width


def build_header_map(tpl_ws, header_row: int, max_col: int):
    m = {}
    for c in range(1, max_col + 1):
        v = tpl_ws.cell(header_row, c).value
        key = norm_header(str(v) if v is not None else "")
        if key:
            m[key] = c
    return m


def write_message_row(out_ws, row_idx: int, header_map: dict, msg: dict, fallback_max_col: int):
    def set_if_exists(header_keys, value):
        for hk in header_keys:
            col = header_map.get(hk)
            if col:
                out_ws.cell(row_idx, col).value = value
                return True
        return False

    ok_any = False
    ok_any |= set_if_exists(["acoes"], msg.get("Ações", ""))
    ok_any |= set_if_exists(["remetente"], msg.get("Remetente", ""))
    ok_any |= set_if_exists(["assunto"], msg.get("Assunto", ""))
    ok_any |= set_if_exists(["enviada em:", "enviada em"], msg.get("Enviada em:", ""))
    ok_any |= set_if_exists(["exibicao ate:", "exibicao ate", "exibição até:", "exibição até"], msg.get("Exibição até:", ""))
    ok_any |= set_if_exists(["data de 1a leitura", "data de 1ª leitura"], msg.get("Data de 1ª leitura", ""))
    ok_any |= set_if_exists(["destinatario", "destinatário"], msg.get("Destinatário", ""))
    ok_any |= set_if_exists(["id mensagem", "mensagem", "numero", "nº", "no"], msg.get("ID mensagem", ""))

    if not ok_any:
        vals = [
            msg.get("Ações", ""),
            msg.get("Remetente", ""),
            msg.get("Assunto", ""),
            msg.get("Enviada em:", ""),
            msg.get("Exibição até:", ""),
            msg.get("Data de 1ª leitura", ""),
            msg.get("Destinatário", ""),
            msg.get("ID mensagem", ""),
        ]
        for c in range(1, min(fallback_max_col, len(vals)) + 1):
            out_ws.cell(row_idx, c).value = vals[c - 1]


def build_output_excel(template_path: Path, blocks: list[dict], out_path: Path):
    tpl_wb = load_workbook(template_path)
    tpl_ws = tpl_wb.active
    max_col = tpl_ws.max_column

    row_cnpj, col_cnpj, row_nome, col_nome, row_header, row_blank, row_data = detect_template_rows(tpl_ws)
    header_map = build_header_map(tpl_ws, row_header, max_col)

    cnpj_merge_min, cnpj_merge_max = find_merge_span_for_cell(tpl_ws, row_cnpj, col_cnpj)

    nome_merge_min = nome_merge_max = None
    if row_nome and col_nome:
        nome_merge_min, nome_merge_max = find_merge_span_for_cell(tpl_ws, row_nome, col_nome)

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = tpl_ws.title

    apply_column_widths(out_ws, tpl_ws, max_col)

    row_idx = 1
    for block in blocks:
        doc = block["doc"]
        nome = block.get("nome", "") or ""
        total = int(block.get("total", 0) or 0)
        msgs = block["msgs"]

        # linha do cabeçalho do bloco (onde ficam {{CNPJ}}/{{NOME}}/{{TOTAL}} no modelo)
        copy_row_style(out_ws, tpl_ws, row_cnpj, row_idx, max_col)

        if cnpj_merge_max > cnpj_merge_min:
            out_ws.merge_cells(
                start_row=row_idx, start_column=cnpj_merge_min,
                end_row=row_idx, end_column=cnpj_merge_max
            )

        if nome_merge_min and nome_merge_max and (nome_merge_max > nome_merge_min):
            out_ws.merge_cells(
                start_row=row_idx, start_column=nome_merge_min,
                end_row=row_idx, end_column=nome_merge_max
            )

        # preenche placeholders em TODAS as células dessa linha (permite "TOTAL DE... {{TOTAL}}")
        for c in range(1, max_col + 1):
            v = tpl_ws.cell(row_cnpj, c).value
            if isinstance(v, str) and ("{{" in v and "}}" in v):
                v2 = v.replace("{{CNPJ}}", format_doc(doc))
                v2 = v2.replace("{{NOME}}", nome)
                v2 = v2.replace("{{TOTAL}}", str(total))
                out_ws.cell(row_idx, c).value = v2

        # também garante que o topo esquerdo do merge do CNPJ fique correto
        out_ws.cell(row_idx, cnpj_merge_min).value = format_doc(doc)

        # e garante o NOME na célula do marcador (se existir marcador exato)
        if row_nome and col_nome:
            out_ws.cell(row_idx, col_nome).value = nome

        row_idx += 1

        copy_row_style(out_ws, tpl_ws, row_blank, row_idx, max_col)
        row_idx += 1

        copy_row_style(out_ws, tpl_ws, row_header, row_idx, max_col)
        for c in range(1, max_col + 1):
            out_ws.cell(row_idx, c).value = tpl_ws.cell(row_header, c).value
        row_idx += 1

        copy_row_style(out_ws, tpl_ws, row_blank, row_idx, max_col)
        row_idx += 1

        for m in msgs:
            copy_row_style(out_ws, tpl_ws, row_data, row_idx, max_col)
            write_message_row(out_ws, row_idx, header_map, m, max_col)
            row_idx += 1

            copy_row_style(out_ws, tpl_ws, row_blank, row_idx, max_col)
            row_idx += 1

        copy_row_style(out_ws, tpl_ws, row_blank, row_idx, max_col)
        row_idx += 1

    out_wb.save(out_path)


# =========================
# CONTROLE AO VIVO (Excel)
# =========================
def live_safe_save_workbook(wb, control_path: Path) -> Path:
    control_path = Path(control_path)
    try:
        wb.save(control_path)
        return control_path
    except PermissionError:
        alt = control_path.with_name(control_path.stem + "__DESBLOQUEAR.xlsx")
        wb.save(alt)
        return alt
    except Exception:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt2 = control_path.with_name(control_path.stem + f"__FALLBACK_{ts}.xlsx")
        wb.save(alt2)
        return alt2


def live_find_col(ws, header_name: str, create_if_missing: bool = False) -> int | None:
    want = (header_name or "").strip().lower()
    if not want:
        return None

    maxc = max(1, ws.max_column)
    for c in range(1, maxc + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        if str(v).strip().lower() == want:
            return c

    if not create_if_missing:
        return None

    new_c = maxc + 1
    ws.cell(1, new_c).value = header_name
    return new_c


def live_init_control_workbook(base_dir: Path):
    path = Path(base_dir) / "controle_verificacao.xlsx"

    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()

    if "Para checar" in wb.sheetnames:
        ws_check = wb["Para checar"]
    else:
        ws_check = wb.active
        ws_check.title = "Para checar"

    if "Resultados" in wb.sheetnames:
        ws_results = wb["Resultados"]
    else:
        ws_results = wb.create_sheet("Resultados")

    if ws_check.max_row < 1 or ws_check.cell(1, 1).value is None:
        ws_check.append(["CNPJ/CPF"])
    if ws_results.max_row < 1 or ws_results.cell(1, 1).value is None:
        ws_results.append(["Data/Hora"])

    # Para checar
    live_find_col(ws_check, "CNPJ/CPF", create_if_missing=True)
    live_find_col(ws_check, "Empresa", create_if_missing=True)
    live_find_col(ws_check, "TOTAL", create_if_missing=True)  # <--- novo
    live_find_col(ws_check, "Status", create_if_missing=True)
    live_find_col(ws_check, "Última atualização", create_if_missing=True)
    live_find_col(ws_check, "Tentativas", create_if_missing=True)
    live_find_col(ws_check, "Observação", create_if_missing=True)

    # Resultados
    live_find_col(ws_results, "Data/Hora", create_if_missing=True)
    live_find_col(ws_results, "CNPJ/CPF", create_if_missing=True)
    live_find_col(ws_results, "Empresa", create_if_missing=True)
    live_find_col(ws_results, "Resultado", create_if_missing=True)
    live_find_col(ws_results, "Qtde Mensagens", create_if_missing=True)
    live_find_col(ws_results, "TOTAL do dia", create_if_missing=True)  # <--- novo (histórico)
    live_find_col(ws_results, "Observação", create_if_missing=True)

    saved_to = live_safe_save_workbook(wb, path)
    if DEBUG_LOGS:
        print(f"[LIVE] Controle: {saved_to.resolve()}")
    return wb, ws_check, ws_results, saved_to


def live_load_queue_from_control(wb, ws_check, control_path: Path, docs: list[str], name_map: dict[str, str]) -> list[str]:
    col_doc = live_find_col(ws_check, "CNPJ/CPF", create_if_missing=True)
    col_emp = live_find_col(ws_check, "Empresa", create_if_missing=True)
    col_st = live_find_col(ws_check, "Status", create_if_missing=True)

    existing = {}
    for r in range(2, ws_check.max_row + 1):
        v = ws_check.cell(r, col_doc).value
        if v:
            d = normalize_doc_from_cell(v)
            if d:
                existing[d] = r

    added = 0
    touched = 0

    for d in docs:
        if d not in existing:
            r = ws_check.max_row + 1
            ws_check.cell(r, col_doc).value = format_doc(d)
            ws_check.cell(r, col_emp).value = (name_map.get(d, "") or "")
            existing[d] = r
            added += 1
        else:
            r = existing[d]
            cur_emp = ws_check.cell(r, col_emp).value
            if (cur_emp is None or str(cur_emp).strip() == "") and (name_map.get(d, "") or ""):
                ws_check.cell(r, col_emp).value = name_map.get(d, "")
                touched += 1

    if added or touched:
        live_safe_save_workbook(wb, control_path)
        if DEBUG_LOGS:
            print(f"[LIVE] 'Para checar' atualizado: +{added} novos, {touched} empresas preenchidas.")

    def needs_run(st: str) -> bool:
        s = (st or "").strip().lower()
        if s in ("preenchido", "sem procuração"):
            return False
        return True

    queue = []
    for r in range(2, ws_check.max_row + 1):
        raw = ws_check.cell(r, col_doc).value
        if not raw:
            continue
        digits = normalize_doc_from_cell(raw)
        if not digits:
            continue
        st = ws_check.cell(r, col_st).value
        if not needs_run(st):
            continue
        queue.append((r, digits))

    queue.sort(key=lambda x: x[0])
    return [d for _, d in queue]


def live_set_check_status(wb, ws_check, control_path: Path, digits: str, empresa: str, status: str, note: str = "", inc_attempt: bool = True):
    col_doc = live_find_col(ws_check, "CNPJ/CPF", create_if_missing=True)
    col_emp = live_find_col(ws_check, "Empresa", create_if_missing=True)
    col_st = live_find_col(ws_check, "Status", create_if_missing=True)
    col_ts = live_find_col(ws_check, "Última atualização", create_if_missing=True)
    col_try = live_find_col(ws_check, "Tentativas", create_if_missing=True)
    col_obs = live_find_col(ws_check, "Observação", create_if_missing=True)

    key = only_digits(digits)
    row = None
    for r in range(2, ws_check.max_row + 1):
        v = ws_check.cell(r, col_doc).value
        if v and only_digits(str(v)) == key:
            row = r
            break

    if row is None:
        row = ws_check.max_row + 1
        ws_check.cell(row, col_doc).value = format_doc(digits)

    ws_check.cell(row, col_doc).value = format_doc(digits)
    if empresa:
        ws_check.cell(row, col_emp).value = empresa
    ws_check.cell(row, col_st).value = status
    ws_check.cell(row, col_ts).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if inc_attempt:
        try:
            cur = ws_check.cell(row, col_try).value
            cur = int(cur) if cur is not None else 0
        except Exception:
            cur = 0
        ws_check.cell(row, col_try).value = cur + 1

    if note:
        ws_check.cell(row, col_obs).value = (note or "")[:500]

    saved_to = live_safe_save_workbook(wb, control_path)
    if DEBUG_LOGS:
        print(f"[LIVE] Para checar: {format_doc(digits)} ({empresa}) -> {status} | salvo: {saved_to.name}")


def live_set_check_total(wb, ws_check, control_path: Path, digits: str, total_today: int):
    """
    NOVA FUNÇÃO:
    grava o TOTAL (total de mensagens do dia) na aba "Para checar", coluna TOTAL.
    """
    col_doc = live_find_col(ws_check, "CNPJ/CPF", create_if_missing=True)
    col_tot = live_find_col(ws_check, "TOTAL", create_if_missing=True)

    key = only_digits(digits)
    row = None
    for r in range(2, ws_check.max_row + 1):
        v = ws_check.cell(r, col_doc).value
        if v and only_digits(str(v)) == key:
            row = r
            break

    if row is None:
        row = ws_check.max_row + 1
        ws_check.cell(row, col_doc).value = format_doc(digits)

    ws_check.cell(row, col_tot).value = int(total_today or 0)

    saved_to = live_safe_save_workbook(wb, control_path)
    if DEBUG_LOGS:
        print(f"[LIVE] TOTAL (dia) atualizado: {format_doc(digits)} -> {int(total_today or 0)} | salvo: {saved_to.name}")


def live_append_result(wb, ws_results, control_path: Path, digits: str, empresa: str, result: str, qty: int = 0, total_today: int = 0, note: str = ""):
    col_dt = live_find_col(ws_results, "Data/Hora", create_if_missing=True)
    col_doc = live_find_col(ws_results, "CNPJ/CPF", create_if_missing=True)
    col_emp = live_find_col(ws_results, "Empresa", create_if_missing=True)
    col_res = live_find_col(ws_results, "Resultado", create_if_missing=True)
    col_qty = live_find_col(ws_results, "Qtde Mensagens", create_if_missing=True)
    col_tdy = live_find_col(ws_results, "TOTAL do dia", create_if_missing=True)
    col_obs = live_find_col(ws_results, "Observação", create_if_missing=True)

    r = ws_results.max_row + 1
    ws_results.cell(r, col_dt).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_results.cell(r, col_doc).value = format_doc(digits)
    ws_results.cell(r, col_emp).value = empresa or ""
    ws_results.cell(r, col_res).value = result
    ws_results.cell(r, col_qty).value = int(qty or 0)
    ws_results.cell(r, col_tdy).value = int(total_today or 0)
    ws_results.cell(r, col_obs).value = (note or "")[:500]

    saved_to = live_safe_save_workbook(wb, control_path)
    if DEBUG_LOGS:
        print(f"[LIVE] Resultados: {format_doc(digits)} ({empresa}) -> {result} | salvo: {saved_to.name}")


# =========================
# AUTOMATION PAUSE
# =========================
def _pause_on_dialog_error(page, base_dir: Path, digits: str, reason: str = "dialog_error", msg: str = ""):
    """
    Pausa de segurança (AUTOMATION_DELAY_SECONDS) para qualquer erro que levante a janelinha
    do diálogo de perfil/acesso (ex.: procuração expirada). Também tenta estabilizar voltando
    para a home do e-CAC depois da pausa.
    """
    try:
        dump_debug(page, base_dir, f"dialog_error_{reason}_{digits}")
    except Exception:
        pass

    extra = f" | {msg.strip()[:160]}" if msg else ""
    print(f"  -> ALERTA: erro no diálogo ({reason}). Pausando {AUTOMATION_DELAY_SECONDS}s{extra}")
    time.sleep(AUTOMATION_DELAY_SECONDS)

    try:
        page.goto(ECAC_URL, wait_until="domcontentloaded", timeout=30000)
    except Exception:
        try:
            page.reload(wait_until="domcontentloaded", timeout=30000)
        except Exception:
            pass

def _pause_and_defer_doc(page, base_dir: Path, digits: str):
    try:
        dump_debug(page, base_dir, f"bot_defer_{digits}")
    except Exception:
        pass

    print(f"  -> ALERTA: suspeita de automatização detectada. Pausando {AUTOMATION_DELAY_SECONDS}s e adiando esse CNPJ...")
    time.sleep(AUTOMATION_DELAY_SECONDS)

    try:
        page.goto(ECAC_URL, wait_until="domcontentloaded", timeout=30000)
    except Exception:
        try:
            page.reload(wait_until="domcontentloaded", timeout=30000)
        except Exception:
            pass


# =========================
# CHROME CDP
# =========================
def pick_ecac_page_from_open_tabs(context):
    pages = context.pages
    if not pages:
        return context.new_page()
    for p in pages:
        u = (p.url or "").lower()
        if "cav.receita" in u or "caixapostal" in u:
            return p
    return pages[0]


# =========================
# MAIN
# =========================
def main():
    base_dir = Path(__file__).resolve().parent

    input_excel = find_input_excel(base_dir)
    template_excel = find_template_excel(base_dir, input_excel)

    docs = read_docs_from_input_excel(input_excel)
    if not docs:
        raise RuntimeError("Não encontrei CPFs/CNPJs válidos na coluna A (linha 2 em diante).")

    name_map = read_doc_name_map_from_input_excel(input_excel)

    ctrl_wb, ws_check, ws_results, control_path = live_init_control_workbook(base_dir)
    queue = live_load_queue_from_control(ctrl_wb, ws_check, control_path, docs, name_map)

    bot_retries = {d: 0 for d in queue}
    blocks = []

    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp("http://127.0.0.1:9222")
        context = browser.contexts[0] if browser.contexts else browser.new_context()
        page = pick_ecac_page_from_open_tabs(context)

        print("\nConectei no Chrome aberto.")
        print("Deixe a aba do e-CAC aberta e logada, e pressione ENTER para começar.\n")
        input()

        try:
            if "cav.receita" not in (page.url or ""):
                page.goto(ECAC_URL, wait_until="domcontentloaded")
        except Exception:
            pass

        i = 0
        while i < len(queue):
            digits = queue[i]
            empresa = name_map.get(digits, "") or ""

            live_set_check_status(ctrl_wb, ws_check, control_path, digits, empresa, "em andamento", note="", inc_attempt=True)

            try:
                print(f"Processando {digits}...")

                _bot_guard_check_page(page, base_dir, f"start_doc_{digits}")

                switch_profile_to_doc(page, digits, base_dir)

                page, scope, state = goto_caixa_postal(page, base_dir, timeout_ms=120000)

                if state == "deny":
                    note = "Acesso negado/expirado detectado na tela"
                    dump_debug(page, base_dir, f"deny_caixa_{digits}")

                    blocks.append({"doc": digits, "nome": empresa, "total": 0, "msgs": []})
                    live_set_check_total(ctrl_wb, ws_check, control_path, digits, 0)
                    live_set_check_status(ctrl_wb, ws_check, control_path, digits, empresa, "pulado", note=note, inc_attempt=False)
                    live_append_result(ctrl_wb, ws_results, control_path, digits, empresa, "pulado", qty=0, total_today=0, note=note)

                    i += 1
                    continue

                if state == "empty":
                    blocks.append({"doc": digits, "nome": empresa, "total": 0, "msgs": []})
                    live_set_check_total(ctrl_wb, ws_check, control_path, digits, 0)
                    live_set_check_status(ctrl_wb, ws_check, control_path, digits, empresa, "preenchido", note="Caixa postal vazia", inc_attempt=False)
                    live_append_result(ctrl_wb, ws_results, control_path, digits, empresa, "preenchido", qty=0, total_today=0, note="Caixa postal vazia")

                    i += 1
                    continue

                # Ajusta exibição para 100 (corrigido p/ não abrir filtro)
                try:
                    set_caixa_page_size(scope, CAIXA_PAGE_SIZE_TARGET, base_dir, label=digits)
                except Exception:
                    pass

                msgs = scrape_messages(scope, base_dir, label=digits, max_pages=50)

                # TOTAL DO DIA (hoje) + total geral já extraído
                total_today = compute_total_today(msgs)
                if DEBUG_LOGS:
                    print(f"  -> OK ({len(msgs)} mensagens). TOTAL DO DIA: {total_today}")

                blocks.append({"doc": digits, "nome": empresa, "total": total_today, "msgs": msgs})

                # grava total ao vivo
                live_set_check_total(ctrl_wb, ws_check, control_path, digits, total_today)

                live_set_check_status(ctrl_wb, ws_check, control_path, digits, empresa, "preenchido", note=f"{len(msgs)} mensagens", inc_attempt=False)
                live_append_result(ctrl_wb, ws_results, control_path, digits, empresa, "preenchido", qty=len(msgs), total_today=total_today, note="")

                i += 1

            except SkipDoc as e:
                msg = str(e)
                blocks.append({"doc": digits, "nome": empresa, "total": 0, "msgs": []})
                live_set_check_total(ctrl_wb, ws_check, control_path, digits, 0)

                low = msg.lower()
                if "sem procura" in low:
                    live_set_check_status(ctrl_wb, ws_check, control_path, digits, empresa, "sem procuração", note=msg, inc_attempt=False)
                    live_append_result(ctrl_wb, ws_results, control_path, digits, empresa, "sem procuração", qty=0, total_today=0, note=msg)
                else:
                    live_set_check_status(ctrl_wb, ws_check, control_path, digits, empresa, "pulado", note=msg, inc_attempt=False)
                    live_append_result(ctrl_wb, ws_results, control_path, digits, empresa, "pulado", qty=0, total_today=0, note=msg)

                i += 1
                continue

            except BotDetected as e:
                note = str(e)
                bot_retries[digits] = bot_retries.get(digits, 0) + 1

                blocks.append({"doc": digits, "nome": empresa, "total": 0, "msgs": []})
                live_set_check_total(ctrl_wb, ws_check, control_path, digits, 0)

                live_set_check_status(ctrl_wb, ws_check, control_path, digits, empresa, "detectada automatização", note=note, inc_attempt=False)
                live_append_result(ctrl_wb, ws_results, control_path, digits, empresa, "detectada automatização", qty=0, total_today=0, note=note)

                if bot_retries[digits] > MAX_AUTOMATION_RETRIES_PER_DOC:
                    print(f"  -> PULADO DEFINITIVO (muitas suspeitas): {digits} | {note}")
                    i += 1
                    continue

                _pause_and_defer_doc(page, base_dir, digits)
                queue.append(queue.pop(i))
                print(f"  -> ADIADO: {digits} foi movido para o final (tentativa {bot_retries[digits]}).")
                continue

            except Exception as e:
                msg = str(e)
                print(f"[ERRO] {digits}: {msg}")

                try:
                    dump_debug(page, base_dir, f"erro_{digits}")
                except Exception:
                    pass

                blocks.append({"doc": digits, "nome": empresa, "total": 0, "msgs": []})
                live_set_check_total(ctrl_wb, ws_check, control_path, digits, 0)

                live_set_check_status(ctrl_wb, ws_check, control_path, digits, empresa, "erro", note=msg, inc_attempt=False)
                live_append_result(ctrl_wb, ws_results, control_path, digits, empresa, "erro", qty=0, total_today=0, note=msg)

                i += 1
                continue

        out_file = base_dir / f"caixa_postal_ecac_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        build_output_excel(template_excel, blocks, out_file)

        print(f"\nOK: gerado {out_file.resolve()}")
        print(f"Input usado: {input_excel.name}")
        print(f"Modelo usado: {template_excel.name}")
        print(f"Controle ao vivo: {control_path.resolve()}")
        print(f"Sem procuração: {(base_dir / SEM_PROCURACAO_TXT).resolve()}")
        if DEBUG_DUMPS:
            print(f"Debug (se houver): {base_dir / DEBUG_DIRNAME}")


if __name__ == "__main__":
    main()
