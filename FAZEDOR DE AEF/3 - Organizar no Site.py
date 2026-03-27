# 3 - Organizar no Site

# =========================
# Configuracoes
# =========================

import os
import sys
import time
import argparse
import shutil
import subprocess
from datetime import datetime
from pathlib import Path
import re
import unicodedata


BASE_DIR = r"W:\DOCUMENTOS ESCRITORIO\INSTALACAO SISTEMA\python\FAZEDOR DE AEF"
PASTA_ARQUIVOS = os.path.join(BASE_DIR, "Arquivos")
CAMINHO_EMPRESAS = os.path.join(BASE_DIR, "empresas.txt")
SCRIPT_4_MOVER = os.path.join(BASE_DIR, "4 - Mover AEF para Pasta Cliente.py")
EMPRESAS_COM_ETAPA_4 = {"22", "3", "143-1", "143-2", "201"}

# Site
URL_LOGIN = "https://aefhondabr.nx-services.com/hondabr/index.html#/login"

# Arquivos aceitos para leitura (gerados pelo "2 - Fromatador XLSX.py" e variacoes legadas)
PADRAO_ARQUIVO_FINAL = "final_{empresa}.xlsx"
PADRAO_ARQUIVO_AEF = "BALANCETE AEF {empresa}.xlsx"
NOME_ARQUIVO_BALANCETE = "balancete.xlsx"

NOME_LOG = "Log - Organizar no Site.txt"
PASTA_LOGS = os.path.join(BASE_DIR, "logs")
PASTA_RUNS = os.path.join(PASTA_LOGS, "runs")
PASTA_LEGACY = os.path.join(PASTA_LOGS, "legacy")
MANTER_ULTIMOS_RUNS_PADRAO = 30

# Pastas efetivas desta execucao (podem ser sobrescritas via _init_debug_run()).
RUN_ID = ""
RUN_DIR = ""
PASTA_PRINTS = os.path.join(PASTA_LOGS, "prints")
PASTA_DUMPS = os.path.join(PASTA_LOGS, "dumps")
PASTA_HTML = os.path.join(PASTA_LOGS, "html")

DELAY_ENTRE_ACOES = 0.2
PAUSA_ENTRE_EMPRESAS = 0.5

# Perfis de login (credenciais em ENV / .env)
PERFIS_LOGIN = ["LOBO", "CASTRO", "TELEMACO", "MOTOACAO", "RIO BRANCO"]
PERFIL_PADRAO = "RIO BRANCO"
DOTENV_PADROES = [
    Path(BASE_DIR) / ".env",
    Path(__file__).resolve().parent / ".env",
]

# Playwright
PLAYWRIGHT_HEADLESS = False
PLAYWRIGHT_TIMEOUT_MS = 30_000
# Padrao: login deve ser detectado automaticamente. Pause apos login somente se solicitado via flag.
PAUSAR_APOS_LOGIN = False
GOTO_TIMEOUT_MS = 90_000
TENTATIVAS_GOTO = 3
ESPERA_ENTRE_GOTO_S = 3
# Esperas curtas para sincronizacao de SPA (evita travar em long-polling).
ESPERA_DOM_MS = 8_000
ESPERA_NETWORKIDLE_MS = 1_500

# Formato numerico do site (BR): milhar com "." e decimal com ",".
# O campo costuma estar vazio, entao a deteccao automatica pode falhar.
SITE_DECIMAL_VIRGULA_PADRAO = True

# Retries (pagina de submissions as vezes demora a "popular" a tabela)
TENTATIVAS_CLICAR_EDITAR = 3
ESPERA_ENTRE_TENTATIVAS_S = 5

# Edicao (apos clicar em Editar)
TEXTO_MENU_VEICULOS_NOVOS = "Veiculos Novos"
OPCAO_VEICULOS_NOVOS = "Ativo"
TENTATIVAS_CONFIRMAR = 3
ESPERA_ENTRE_TENTATIVAS_CONFIRMAR_S = 3

# XLSX (valores calculados)
ABA_ATIVO = "Ativo"
ABA_PASSIVO = "Passivo"
ABA_DRE = "DRE"
COL_NIVEL = 1
COL_DESCRICAO = 2
COL_VALOR = 3
MAX_LINHAS_POR_TESTE = 0  # 0 = todas

# Scroll (telas longas podem carregar elementos sob demanda)
SCROLL_STEP_PX = 900
SCROLL_MAX_PASSOS_POR_CODIGO = 40

# DRE
DRE_COLUNAS_ALVO = [
    "TOTAL",
    "VN",
    "VED",
    "SEM USO ATUAL",
    "VSN",
    "CNH",
    "ATV",
    "PECAS",
    "ACESS",
    "BOUTIQUE",
    "OFICINA",
    "ADMINISTRATIVO",
    "ADM",  # alias
]

DRE_ESPERAR_TIMEOUT_MS = 60_000
DRE_ZOOM_PERCENT = 67  # zoom-out para ajudar a renderizar/mostrar a grade inteira
FALHAR_DIVERGENCIA_ATIVO_PASSIVO = False  # nao interrompe fluxo antes de chegar no DRE

# Mapa fixo do XLSX (observado em final_22.xlsx). Se mudar, ajusta aqui.
DRE_XLSX_COLS = {
    "TOTAL": 3,
    "VN": 5,
    "VED": 6,
    "SEM USO ATUAL": 7,
    "VSN": 8,
    "CNH": 9,
    "ATV": 10,
    "PF": 11,
    "PECAS": 12,
    "ACESS": 13,
    "BOUTIQUE": 14,
    "OFICINA": 15,
    "ADMINISTRATIVO": 16,
}

# =========================
# Utilitarios
# =========================


def _normalizar_empresa(texto: str) -> str:
    return texto.strip().lstrip("\ufeff")


def carregar_empresas(caminho: str) -> list[str]:
    if not os.path.isfile(caminho):
        print(f"ERRO: arquivo de empresas nao encontrado: {caminho}")
        sys.exit(1)

    with open(caminho, "r", encoding="utf-8") as arquivo:
        empresas = [_normalizar_empresa(linha) for linha in arquivo.readlines() if linha.strip()]

    if not empresas:
        print("ERRO: lista de empresas vazia.")
        sys.exit(1)

    return empresas


def localizar_arquivo_final(empresa: str) -> str | None:
    empresa = _normalizar_empresa(empresa)
    pasta_empresa = os.path.join(PASTA_ARQUIVOS, empresa)

    candidatos_nome = [
        PADRAO_ARQUIVO_AEF.format(empresa=empresa),
        PADRAO_ARQUIVO_FINAL.format(empresa=empresa),
        NOME_ARQUIVO_BALANCETE,
    ]

    # 1) Busca direta dentro da pasta da empresa (mais confiavel).
    for nome in candidatos_nome:
        caminho = os.path.join(pasta_empresa, nome)
        if os.path.isfile(caminho):
            return caminho

    # 2) Busca por comparacao normalizada na pasta da empresa.
    if os.path.isdir(pasta_empresa):
        try:
            alvos_norm = {
                _norm_header(PADRAO_ARQUIVO_AEF.format(empresa=empresa)),
                _norm_header(PADRAO_ARQUIVO_FINAL.format(empresa=empresa)),
                _norm_header(NOME_ARQUIVO_BALANCETE),
            }
            xlsx_empresa = []
            for nome in os.listdir(pasta_empresa):
                if not nome.lower().endswith(".xlsx"):
                    continue
                caminho = os.path.join(pasta_empresa, nome)
                xlsx_empresa.append(caminho)
                if _norm_header(nome) in alvos_norm:
                    return caminho
            # Se houver apenas um .xlsx na pasta da empresa, usa ele.
            if len(xlsx_empresa) == 1:
                return xlsx_empresa[0]
        except Exception:
            pass

    # 3) Fallback global: busca por nome final_<empresa>.xlsx / BALANCETE AEF <empresa>.xlsx.
    alvos_global = {
        _norm_header(PADRAO_ARQUIVO_AEF.format(empresa=empresa)),
        _norm_header(PADRAO_ARQUIVO_FINAL.format(empresa=empresa)),
    }
    for raiz, _, arquivos in os.walk(PASTA_ARQUIVOS):
        for nome in arquivos:
            if not nome.lower().endswith(".xlsx"):
                continue
            if _norm_header(nome) in alvos_global:
                return os.path.join(raiz, nome)

    return None


def _validar_xlsx_existente(caminho: str, contexto: str) -> str:
    """
    Padrao desta automacao: os arquivos sempre sao .xlsx (o pipeline anterior cuida disso).
    """
    caminho = (caminho or "").strip()
    if not caminho:
        raise RuntimeError(f"{contexto}: caminho vazio.")
    if not caminho.lower().endswith(".xlsx"):
        raise RuntimeError(f"{contexto}: esperado .xlsx (recebido: {caminho}).")
    if not os.path.isfile(caminho):
        raise RuntimeError(f"{contexto}: arquivo nao encontrado: {caminho}")
    return caminho


def caminho_log() -> str:
    return os.path.join(BASE_DIR, NOME_LOG)


def caminho_log_run() -> str:
    if RUN_DIR:
        return os.path.join(RUN_DIR, "run.log")
    return ""


def log_linha(msg: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    linha = f"[{ts}] {msg}"
    print(linha)
    with open(caminho_log(), "a", encoding="utf-8") as arq:
        arq.write(linha + "\n")
    try:
        p = caminho_log_run()
        if p:
            with open(p, "a", encoding="utf-8") as arq2:
                arq2.write(linha + "\n")
    except RuntimeError:
        raise
    except Exception:
        pass


def _slug(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    # Normaliza para nome de pasta/arquivo:
    # - remove acentos
    # - mantem somente A-Z, 0-9 e "_"
    t = unicodedata.normalize("NFKD", s)
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = t.upper()
    t = re.sub(r"[^A-Z0-9]+", "_", t).strip("_")
    return t[:60].lower()


def _init_debug_run(args) -> None:
    """
    Prepara um diretório por execucao para centralizar tudo de debug:
    - logs\\runs\\<run_id>\\run.log
    - logs\\runs\\<run_id>\\prints\\*.png
    - logs\\runs\\<run_id>\\dumps\\*.txt / *.json
    - logs\\runs\\<run_id>\\html\\*.html
    """
    global RUN_ID, RUN_DIR, PASTA_PRINTS, PASTA_DUMPS, PASTA_HTML

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    emp = _slug(getattr(args, "empresa", ""))
    comp = _slug(getattr(args, "competencia", ""))
    perfil = _slug(getattr(args, "perfil", ""))

    partes = [ts]
    if perfil:
        partes.append(perfil)
    if emp:
        partes.append(f"emp{emp}")
    if comp:
        partes.append(comp)
    RUN_ID = "_".join(partes)
    RUN_DIR = os.path.join(PASTA_RUNS, RUN_ID)

    PASTA_PRINTS = os.path.join(RUN_DIR, "prints")
    PASTA_DUMPS = os.path.join(RUN_DIR, "dumps")
    PASTA_HTML = os.path.join(RUN_DIR, "html")

    os.makedirs(PASTA_PRINTS, exist_ok=True)
    os.makedirs(PASTA_DUMPS, exist_ok=True)
    os.makedirs(PASTA_HTML, exist_ok=True)

    # Migra artefatos antigos que ficaram no logs\\prints (versoes antigas) para evitar confusao.
    # As execucoes atuais ja usam logs\\runs\\<run_id>\\...
    try:
        _migrar_logs_raiz_para_legacy(run_id=RUN_ID)
    except RuntimeError:
        raise
    except Exception:
        pass

    # Metadados basicos da execucao
    try:
        meta = {
            "run_id": RUN_ID,
            "timestamp": ts,
            "cwd": os.getcwd(),
            "args": vars(args),
        }
        with open(os.path.join(PASTA_DUMPS, "meta.json"), "w", encoding="utf-8") as f:
            import json

            json.dump(meta, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _limpar_runs_antigos(maneter_ultimos: int) -> list[str]:
    """
    Remove pastas antigas em logs\\runs, mantendo apenas os N runs mais recentes.
    Retorna lista das pastas removidas.
    """
    removidos: list[str] = []
    try:
        n = int(maneter_ultimos)
    except Exception:
        n = MANTER_ULTIMOS_RUNS_PADRAO
    if n < 1:
        n = 1

    try:
        if not os.path.isdir(PASTA_RUNS):
            return []
        dirs = []
        for nome in os.listdir(PASTA_RUNS):
            p = os.path.join(PASTA_RUNS, nome)
            if os.path.isdir(p):
                dirs.append(p)
        # Mais recente primeiro (mtime da pasta).
        dirs.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        antigos = dirs[n:]
        for p in antigos:
            try:
                shutil.rmtree(p)
                removidos.append(p)
            except Exception:
                continue
    except Exception:
        return removidos

    return removidos


def _migrar_logs_raiz_para_legacy(run_id: str) -> None:
    """
    Versoes antigas deste script salvavam screenshots diretamente em logs\\prints.
    Para evitar bloat/confusao, movemos esses arquivos para logs\\legacy\\<run_id>\\...

    Nao remove pastas de runs (isso e feito por _limpar_runs_antigos).
    """
    run_id = (run_id or "").strip()
    if not run_id:
        return

    srcs = [
        os.path.join(PASTA_LOGS, "prints"),
        os.path.join(PASTA_LOGS, "dumps"),
        os.path.join(PASTA_LOGS, "html"),
    ]

    # Somente migra se existir arquivo (nao mexe em pastas vazias).
    algum = False
    for s in srcs:
        try:
            if os.path.isdir(s) and any(os.path.isfile(os.path.join(s, n)) for n in os.listdir(s)):
                algum = True
                break
        except Exception:
            continue
    if not algum:
        return

    base_dst = os.path.join(PASTA_LEGACY, run_id)
    os.makedirs(base_dst, exist_ok=True)

    for src in srcs:
        if not os.path.isdir(src):
            continue
        nome = os.path.basename(src.rstrip("\\/")) or "outros"
        dst = os.path.join(base_dst, nome)
        os.makedirs(dst, exist_ok=True)

        try:
            for arq in os.listdir(src):
                p = os.path.join(src, arq)
                if not os.path.isfile(p):
                    continue
                # Evita sobrescrever em caso de nomes repetidos.
                alvo = os.path.join(dst, arq)
                if os.path.exists(alvo):
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    base, ext = os.path.splitext(arq)
                    alvo = os.path.join(dst, f"{base}_dup_{ts}{ext}")
                try:
                    shutil.move(p, alvo)
                except Exception:
                    continue
        except Exception:
            continue

        # Se ficou vazio, remove para reduzir poluicao visual.
        try:
            if not os.listdir(src):
                os.rmdir(src)
        except Exception:
            pass


def montar_tarefas(empresas: list[str]) -> list[tuple[str, str]]:
    tarefas: list[tuple[str, str]] = []
    for emp in empresas:
        caminho = localizar_arquivo_final(emp)
        if not caminho:
            log_linha(f"AVISO: nao encontrei arquivo final para a empresa {emp}.")
            continue
        tarefas.append((emp, caminho))
    return tarefas


# =========================
# Automacao (site)
# =========================


def _normalizar_perfil(perfil: str) -> str:
    # ENV nao aceita espacos: "RIO BRANCO" -> "RIO_BRANCO"
    return perfil.strip().upper().replace("-", "_").replace(" ", "_")


def _carregar_dotenv_se_existir() -> None:
    """
    Carrega variaveis de um .env se existir (sem sobrescrever ENV do Windows).
    Se python-dotenv nao estiver instalado, apenas ignora.
    """
    try:
        from dotenv import load_dotenv  # type: ignore
    except Exception:
        load_dotenv = None  # type: ignore

    for p in DOTENV_PADROES:
        try:
            if p.is_file():
                if load_dotenv is not None:
                    load_dotenv(dotenv_path=p, override=False)
                else:
                    # Fallback sem dependencia externa (servidor sem python-dotenv).
                    with open(p, "r", encoding="utf-8") as f:
                        for linha in f:
                            s = (linha or "").strip()
                            if not s or s.startswith("#") or "=" not in s:
                                continue
                            k, v = s.split("=", 1)
                            k = k.strip()
                            v = v.strip().strip('"').strip("'")
                            if k and k not in os.environ:
                                os.environ[k] = v
        except Exception:
            continue


def obter_credenciais_por_perfil(perfil: str) -> tuple[str, str]:
    """
    Procura credenciais do perfil em ENV (e opcionalmente .env).

    Padroes aceitos (ex.: perfil "RIO BRANCO" => "RIO_BRANCO"):
    - RIO_BRANCO_USUARIO / RIO_BRANCO_SENHA
    - RIO_BRANCO_USER / RIO_BRANCO_PASS
    - AEF_RIO_BRANCO_USUARIO / AEF_RIO_BRANCO_SENHA
    - AEF_RIO_BRANCO_USER / AEF_RIO_BRANCO_PASS
    - AEF_SITE_RIO_BRANCO_USUARIO / AEF_SITE_RIO_BRANCO_SENHA
    - AEF_SITE_RIO_BRANCO_USER / AEF_SITE_RIO_BRANCO_PASS
    """
    _carregar_dotenv_se_existir()

    pfx = _normalizar_perfil(perfil)

    candidatos = [
        (f"{pfx}_USUARIO", f"{pfx}_SENHA"),
        (f"{pfx}_USER", f"{pfx}_PASS"),
        (f"AEF_{pfx}_USUARIO", f"AEF_{pfx}_SENHA"),
        (f"AEF_{pfx}_USER", f"AEF_{pfx}_PASS"),
        (f"AEF_SITE_{pfx}_USUARIO", f"AEF_SITE_{pfx}_SENHA"),
        (f"AEF_SITE_{pfx}_USER", f"AEF_SITE_{pfx}_PASS"),
    ]

    for k_user, k_pass in candidatos:
        user = os.getenv(k_user, "").strip()
        senha = os.getenv(k_pass, "").strip()
        if user and senha:
            return user, senha

    chaves = []
    for k_user, k_pass in candidatos:
        chaves.append(f"{k_user} / {k_pass}")
    raise RuntimeError(
        "Credenciais nao encontradas no ENV/.env para o perfil "
        f"'{perfil}'. Chaves aceitas: " + "; ".join(chaves)
    )


def obter_perfil_por_empresa(empresa: str) -> str | None:
    """
    Resolve perfil de login por codigo da empresa usando ENV/.env.
    Ex.: AEF_SITE_RIO_BRANCO_CODIGO=201
    """
    _carregar_dotenv_se_existir()
    emp = _normalizar_empresa(empresa)
    if not emp:
        return None

    for perfil in PERFIS_LOGIN:
        pfx = _normalizar_perfil(perfil)
        chaves = [
            f"AEF_SITE_{pfx}_CODIGO",
            f"AEF_{pfx}_CODIGO",
            f"{pfx}_CODIGO",
        ]
        for k in chaves:
            v = (os.getenv(k) or "").strip()
            if v and _normalizar_empresa(v) == emp:
                return perfil
    return None


def _importar_dependencias_ui():
    # Mantido por compatibilidade: este script esta migrando para Playwright.
    return


def _importar_dependencias_playwright():
    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError  # noqa: F401
    except Exception as exc:
        print("ERRO: playwright nao encontrado.")
        print("Instale: pip install playwright")
        print("E depois: playwright install chromium")
        print(f"Detalhe: {exc}")
        raise


def _importar_dependencias_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except Exception as exc:
        print("ERRO: openpyxl nao encontrado.")
        print("Instale: pip install openpyxl")
        print(f"Detalhe: {exc}")
        raise


def _salvar_screenshot(page, nome_base: str) -> str | None:
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome = f"{nome_base}_{ts}.png"
        try:
            os.makedirs(PASTA_PRINTS, exist_ok=True)
            caminho = os.path.join(PASTA_PRINTS, nome)
        except Exception:
            # Fallback: se nao conseguir criar pasta, salva no BASE_DIR como antes.
            caminho = os.path.join(BASE_DIR, nome)
        page.screenshot(path=caminho, full_page=True)
        return caminho
    except Exception:
        return None


def _salvar_dump_texto(nome_base: str, texto: str) -> str | None:
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome = f"{nome_base}_{ts}.txt"
        try:
            os.makedirs(PASTA_DUMPS, exist_ok=True)
            caminho = os.path.join(PASTA_DUMPS, nome)
        except Exception:
            caminho = os.path.join(BASE_DIR, nome)
        with open(caminho, "w", encoding="utf-8") as f:
            f.write(texto or "")
        return caminho
    except Exception:
        return None


def _salvar_dump_json(nome_base: str, payload: object) -> str | None:
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome = f"{nome_base}_{ts}.json"
        try:
            os.makedirs(PASTA_DUMPS, exist_ok=True)
            caminho = os.path.join(PASTA_DUMPS, nome)
        except Exception:
            caminho = os.path.join(BASE_DIR, nome)
        import json

        with open(caminho, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        return caminho
    except Exception:
        return None


def _salvar_dump_html(nome_base: str, html: str) -> str | None:
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome = f"{nome_base}_{ts}.html"
        try:
            os.makedirs(PASTA_HTML, exist_ok=True)
            caminho = os.path.join(PASTA_HTML, nome)
        except Exception:
            caminho = os.path.join(BASE_DIR, nome)
        with open(caminho, "w", encoding="utf-8") as f:
            f.write(html or "")
        return caminho
    except Exception:
        return None


def _dump_estado_pagina(page, nome_base: str) -> None:
    """
    Dump generico do estado da pagina para debug (texto + html + screenshot).
    """
    try:
        _aguardar_carregar_pagina(page)
    except Exception:
        pass

    try:
        caminho = _salvar_screenshot(page, nome_base)
        if caminho:
            log_linha(f"Screenshot: {caminho}")
    except Exception:
        pass

    def _dump_frame(fr, sufixo: str) -> None:
        try:
            txt = fr.evaluate("() => (document.body && document.body.innerText) ? document.body.innerText : ''")
        except Exception:
            txt = ""
        if isinstance(txt, str) and txt.strip():
            ptxt = _salvar_dump_texto(f"{nome_base}_{sufixo}", txt)
            if ptxt:
                log_linha(f"Dump (texto): {ptxt}")

        try:
            html = fr.evaluate(
                "() => (document.documentElement && document.documentElement.outerHTML) ? document.documentElement.outerHTML : ''"
            )
        except Exception:
            html = ""
        if isinstance(html, str) and html.strip():
            phtml = _salvar_dump_html(f"{nome_base}_{sufixo}", html)
            if phtml:
                log_linha(f"Dump (html): {phtml}")

    # Dump do documento principal
    try:
        _dump_frame(page.main_frame, "page")
    except Exception:
        pass

    # Dump do frame "principal" de edicao (se existir). Em algumas telas, o conteudo fica dentro de iframe.
    try:
        fr2 = _esperar_frame_edicao(page, timeout_ms=2000)
        if fr2 is not None and fr2 != page.main_frame:
            _dump_frame(fr2, "frame")
    except Exception:
        pass


def _norm_txt(s: str) -> str:
    t = (s or "").strip()
    if not t:
        return ""
    t = unicodedata.normalize("NFKD", t)
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = t.upper()
    t = t.replace("\u00a0", " ")
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _norm_header(s: str) -> str:
    t = _norm_txt(s)
    t = re.sub(r"[^A-Z0-9]+", "", t)
    return t


def _e_nivel_l(texto: str) -> bool:
    t = (texto or "").strip().upper()
    if not t.startswith("L"):
        return False
    return t[1:].isdigit()


def _numero_nivel(texto: str) -> int:
    t = (texto or "").strip().upper()
    if not t.startswith("L"):
        return 999999
    try:
        return int(t[1:])
    except Exception:
        return 999999


def _e_codigo_linha_planilha(texto: str) -> bool:
    # Ex.: L1, L99, M1, M30...
    t = (texto or "").strip().upper()
    return bool(re.match(r"^[A-Z]{1,3}\d{1,4}$", t))


def _codigo_sort_key(texto: str) -> tuple[str, int]:
    t = (texto or "").strip().upper()
    m = re.match(r"^([A-Z]{1,3})(\d{1,4})$", t)
    if not m:
        return ("ZZZ", 999999)
    return (m.group(1), int(m.group(2)))


def ler_linhas_planilha_xlsx(caminho_xlsx: str, aba: str) -> list[dict]:
    """
    Le uma aba do XLSX usando data_only=True para pegar o VALOR calculado.
    Retorna lista de itens: {codigo, descricao, valor}.
    """
    _importar_dependencias_openpyxl()
    import openpyxl

    if not os.path.isfile(caminho_xlsx):
        raise RuntimeError(f"Arquivo XLSX nao encontrado: {caminho_xlsx}")

    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    if aba in wb.sheetnames:
        ws = wb[aba]
    else:
        ws = wb[wb.sheetnames[0]]

    itens: list[dict] = []
    for r in range(1, ws.max_row + 1):
        codigo = ws.cell(r, COL_NIVEL).value
        if codigo is None:
            continue
        codigo_txt = str(codigo).strip().upper()
        if not _e_codigo_linha_planilha(codigo_txt):
            continue

        descricao = ws.cell(r, COL_DESCRICAO).value
        valor = ws.cell(r, COL_VALOR).value

        if valor is None:
            # Sem valor calculado (pode acontecer se o arquivo nao foi salvo com calculo no Excel).
            continue
        if not isinstance(valor, (int, float)):
            continue

        itens.append(
            {
                "codigo": codigo_txt,
                "descricao": (str(descricao).strip() if descricao is not None else ""),
                "valor": float(valor),
            }
        )

    itens.sort(key=lambda x: _codigo_sort_key(x["codigo"]))
    return itens


def ler_dre_xlsx(caminho_xlsx: str) -> dict[str, dict[str, float]]:
    """
    Le a aba DRE e extrai valores por codigo (ex.: N16) e coluna (VN/VSN/...).
    Usa headers "rolantes": a cada linha de header encontrada, atualiza o mapa.
    """
    _importar_dependencias_openpyxl()
    import openpyxl

    if not os.path.isfile(caminho_xlsx):
        raise RuntimeError(f"Arquivo XLSX nao encontrado: {caminho_xlsx}")

    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    if ABA_DRE not in wb.sheetnames:
        raise RuntimeError(f"Aba '{ABA_DRE}' nao encontrada no XLSX.")
    ws = wb[ABA_DRE]

    alvo_norm = {_norm_header(x): x for x in DRE_COLUNAS_ALVO}

    # Sinônimos comuns do arquivo
    sinonimos = {
        "PECAS": ["PECAS", "PEÇAS"],
        "ACESS": ["ACESS", "ACESSORIOS", "ACESSÓRIOS"],
    }

    header_por_col: dict[int, str] = {}
    out: dict[str, dict[str, float]] = {}

    def _eh_linha_header(valores: list[object]) -> bool:
        hits = 0
        for v in valores:
            if not isinstance(v, str):
                continue
            hn = _norm_header(v)
            if hn in alvo_norm:
                hits += 1
            elif hn in ["ACESSORIOS"]:
                hits += 1
            elif hn in ["PECAS"]:
                hits += 1
        return hits >= 4

    for r in range(1, ws.max_row + 1):
        # Pega uma janela razoável de colunas (o arquivo tem até 29)
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]

        # Atualiza header map se a linha for header
        if _eh_linha_header(vals):
            header_por_col.clear()
            for idx0, v in enumerate(vals, start=1):
                if not isinstance(v, str):
                    continue
                hn = _norm_header(v)
                if hn in alvo_norm:
                    header_por_col[idx0] = hn
                    continue
                if hn == "PEÇAS" or hn == "PECAS":
                    header_por_col[idx0] = "PECAS"
                    continue
                if hn in ["ACESSORIOS", "ACESSÓRIOS"]:
                    header_por_col[idx0] = "ACESS"
                    continue
                if hn == "ACESS":
                    header_por_col[idx0] = "ACESS"
                    continue
            continue

        a1 = ws.cell(r, 1).value
        if a1 is None:
            continue
        if not isinstance(a1, str):
            a1 = str(a1)
        codigo = _norm_txt(a1)
        if not re.match(r"^N\d{1,3}$", codigo):
            continue
        if not header_por_col:
            continue

        dados: dict[str, float] = {}
        for col_idx, hnorm in header_por_col.items():
            if hnorm not in alvo_norm and hnorm not in ["PECAS", "ACESS"]:
                continue
            v = ws.cell(r, col_idx).value
            if v is None:
                continue
            if isinstance(v, str):
                tv = v.strip()
                if tv in ["-", "–"]:
                    v = 0
                else:
                    continue
            if not isinstance(v, (int, float)):
                continue
            key = hnorm
            # Normaliza keys para bater com DRE_COLUNAS_ALVO
            if key == "PEÇAS":
                key = "PECAS"
            dados[key] = float(v)

        if dados:
            out[codigo] = dados

    return out


def ler_dre_xlsx_com_descricoes(caminho_xlsx: str) -> dict[str, dict]:
    """
    Retorna:
    {
      "N16": {
        "descricao": "SalárioDiretor",
        "valores": {"VN": 0.0, "VSN": ...}
      },
      ...
    }
    """
    _importar_dependencias_openpyxl()
    import openpyxl

    if not os.path.isfile(caminho_xlsx):
        raise RuntimeError(f"Arquivo XLSX nao encontrado: {caminho_xlsx}")

    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    if ABA_DRE not in wb.sheetnames:
        raise RuntimeError(f"Aba '{ABA_DRE}' nao encontrada no XLSX.")
    ws = wb[ABA_DRE]

    # Layout do final_XX muda com frequencia.
    # Prioriza leitura dinamica por header (TOTAL/VN/VED/...) e usa colunas fixas como fallback.
    out: dict[str, dict] = {}
    codigos_vistos: dict[str, dict] = {}
    header_por_col: dict[int, str] = {}

    alvo_norm = {
        _norm_header("TOTAL"): "TOTAL",
        _norm_header("VN"): "VN",
        _norm_header("VED"): "VED",
        _norm_header("SEM USO ATUAL"): "SEM USO ATUAL",
        _norm_header("VSN"): "VSN",
        _norm_header("CNH"): "CNH",
        _norm_header("ATV"): "ATV",
        _norm_header("PF"): "PF",
        _norm_header("PECAS"): "PECAS",
        _norm_header("PEÇAS"): "PECAS",
        _norm_header("ACESS"): "ACESS",
        _norm_header("ACESSORIOS"): "ACESS",
        _norm_header("ACESSÓRIOS"): "ACESS",
        _norm_header("BOUTIQUE"): "BOUTIQUE",
        _norm_header("OFICINA"): "OFICINA",
        _norm_header("ADMINISTRATIVO"): "ADMINISTRATIVO",
        _norm_header("ADM"): "ADMINISTRATIVO",
    }

    def _eh_linha_header(vals: list[object]) -> bool:
        hits = 0
        for v in vals:
            if not isinstance(v, str):
                continue
            if _norm_header(v) in alvo_norm:
                hits += 1
        return hits >= 4

    def _ler_num(v) -> tuple[float | None, int]:
        """
        Retorna (valor, hint_sinal):
        - hint_sinal = -1: detectado negativo explicito ("-123")
        - hint_sinal = +1: detectado marcador no inicio ("- 123")
        - hint_sinal = 0: sem indicio explicito
        """
        if v is None:
            return None, 0
        if isinstance(v, (int, float)):
            return float(v), 0
        if isinstance(v, str):
            tv = v.strip()
            if tv in ["", "-", "–"]:
                return 0.0, 0
            t0 = tv.replace("\u2212", "-").replace("\u2013", "-").replace("\u2014", "-")
            hint = 0
            if re.match(r"^-\s+\S+", t0):
                hint = 1   # marcador
            elif re.match(r"^-\S+", t0):
                hint = -1  # negativo colado
            pv = _parse_float_tolerante(tv)
            return pv, hint
        return None, 0

    def _score_valores(vals: dict[str, float]) -> tuple[int, float]:
        # Preferir linha com mais colunas preenchidas e maior soma absoluta
        qtd = len(vals)
        soma = sum(abs(float(x)) for x in vals.values())
        return qtd, soma

    # 1) Leitura dinamica
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]

        if _eh_linha_header(vals):
            header_por_col.clear()
            for cidx, vv in enumerate(vals, start=1):
                if not isinstance(vv, str):
                    continue
                hn = _norm_header(vv)
                if hn in alvo_norm:
                    header_por_col[cidx] = alvo_norm[hn]
            continue

        a1 = ws.cell(r, 1).value
        if a1 is None:
            continue
        codigo = _norm_txt(str(a1))
        if not re.match(r"^N\d{1,3}$", codigo):
            continue

        descricao = (str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value is not None else "")
        codigos_vistos[codigo] = {"row": r, "descricao": descricao}

        valores: dict[str, float] = {}
        hints: dict[str, int] = {}
        if header_por_col:
            for cidx, col_name in header_por_col.items():
                vv, hint = _ler_num(ws.cell(r, cidx).value)
                if vv is None:
                    continue
                valores[col_name] = float(vv)
                hints[col_name] = int(hint)

        if valores:
            novo_score = _score_valores(valores)
            antigo = out.get(codigo)
            if not antigo:
                out[codigo] = {"descricao": descricao, "valores": valores, "hints": hints}
            else:
                antigo_vals = antigo.get("valores") or {}
                antigo_score = _score_valores(antigo_vals)
                if novo_score >= antigo_score:
                    out[codigo] = {"descricao": descricao, "valores": valores, "hints": hints}

    # 2) Fallback fixo (se a leitura dinamica nao conseguiu dados suficientes)
    if len(out) < 20:
        out_fix: dict[str, dict] = {}
        for r in range(1, ws.max_row + 1):
            a1 = ws.cell(r, 1).value
            b1 = ws.cell(r, 2).value
            if a1 is None:
                continue
            codigo = _norm_txt(str(a1))
            if not re.match(r"^N\d{1,3}$", codigo):
                continue

            descricao = ""
            if b1 is not None:
                descricao = str(b1).strip()
            codigos_vistos[codigo] = {"row": r, "descricao": descricao}

            valores: dict[str, float] = {}
            hints: dict[str, int] = {}
            for nome, col_idx in DRE_XLSX_COLS.items():
                if col_idx > ws.max_column:
                    continue
                vv, hint = _ler_num(ws.cell(r, col_idx).value)
                if vv is None:
                    continue
                valores[nome] = float(vv)
                hints[nome] = int(hint)

            if valores:
                novo_score = _score_valores(valores)
                antigo = out_fix.get(codigo)
                if not antigo:
                    out_fix[codigo] = {"descricao": descricao, "valores": valores, "hints": hints}
                else:
                    antigo_vals = antigo.get("valores") or {}
                    antigo_score = _score_valores(antigo_vals)
                    if novo_score >= antigo_score:
                        out_fix[codigo] = {"descricao": descricao, "valores": valores, "hints": hints}

        if len(out_fix) > len(out):
            out = out_fix

    # 3) Linhas de TOTAL (N33+ etc.) podem vir vazias no XLSX (celula em branco, nao 0).
    # Para nao deixar a tela com lixo antigo, garante ao menos TOTAL=0.0 nesses codigos vistos.
    for codigo, info in codigos_vistos.items():
        if codigo in out:
            continue
        try:
            r = int(info.get("row") or 0)
        except Exception:
            r = 0
        if r <= 0:
            continue
        descricao = str(info.get("descricao") or "").strip()
        # TOTAL oficial da planilha (coluna C). Se vazio, forca 0.
        v_total = ws.cell(r, 3).value
        if isinstance(v_total, str):
            tv = v_total.strip()
            if tv in ["", "-", "–"]:
                total_num = 0.0
            else:
                pv = _parse_float_tolerante(tv)
                total_num = float(pv) if pv is not None else 0.0
        elif isinstance(v_total, (int, float)):
            total_num = float(v_total)
        else:
            total_num = 0.0
        out[codigo] = {"descricao": descricao, "valores": {"TOTAL": total_num}, "hints": {"TOTAL": 0}}

    return out


def _sincronizar_xlsx_dre_readonly(caminho_xlsx: str, divergencias: list[dict]) -> int:
    """
    Sincroniza no XLSX oficial os valores readonly calculados pelo site (DRE),
    para manter o arquivo final aderente ao que o sistema efetivamente publica.
    Retorna a quantidade de celulas atualizadas.
    """
    if not divergencias:
        return 0

    _importar_dependencias_openpyxl()
    import openpyxl

    if not os.path.isfile(caminho_xlsx):
        return 0

    wb = openpyxl.load_workbook(caminho_xlsx, data_only=False)
    if ABA_DRE not in wb.sheetnames:
        return 0
    ws = wb[ABA_DRE]

    # Mapa codigo -> linha
    row_by_code: dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        c = ws.cell(r, 1).value
        if c is None:
            continue
        cod = _norm_txt(str(c))
        if re.match(r"^N\d{1,3}$", cod):
            row_by_code[cod] = r

    atualizadas = 0
    for d in divergencias:
        try:
            if (d.get("status") or "") != "skip_readonly":
                continue
            cod = _norm_txt(str(d.get("codigo") or ""))
            col = _norm_txt(str(d.get("coluna") or ""))
            obs = d.get("observado", None)
            if not cod or obs is None:
                continue
            if cod not in row_by_code:
                continue

            col_key = col
            if col_key == "ADM":
                col_key = "ADMINISTRATIVO"
            if col_key not in DRE_XLSX_COLS:
                continue

            r = row_by_code[cod]
            cidx = DRE_XLSX_COLS[col_key]
            atual = ws.cell(r, cidx).value
            try:
                atual_f = float(atual) if atual is not None else None
            except Exception:
                atual_f = None

            if atual_f is not None and _quase_igual(atual_f, float(obs)):
                continue

            ws.cell(r, cidx).value = float(obs)
            atualizadas += 1
        except Exception:
            continue

    if atualizadas > 0:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        try:
            bak = caminho_xlsx[:-5] + f".bak_{ts}.xlsx" if caminho_xlsx.lower().endswith(".xlsx") else caminho_xlsx + f".bak_{ts}"
            shutil.copy2(caminho_xlsx, bak)
            log_linha(f"Backup XLSX criado: {bak}")
        except Exception as exc:
            log_linha(f"AVISO: nao consegui criar backup do XLSX antes de salvar sync readonly: {exc}")
        try:
            wb.save(caminho_xlsx)
            log_linha(f"XLSX sincronizado com readonly do DRE: {atualizadas} celulas atualizadas em {caminho_xlsx}")
        except Exception as exc:
            # Fallback quando o arquivo oficial estiver bloqueado (ex.: aberto no Excel).
            try:
                alt = caminho_xlsx[:-5] + f".sync_{ts}.xlsx" if caminho_xlsx.lower().endswith(".xlsx") else caminho_xlsx + f".sync_{ts}"
                wb.save(alt)
                log_linha(
                    "AVISO: nao consegui salvar no XLSX oficial (possivel arquivo aberto/bloqueado). "
                    + f"Salvei copia sincronizada em: {alt}. Erro original: {exc}"
                )
            except Exception as exc2:
                log_linha(
                    "AVISO: falha ao salvar sincronizacao readonly do DRE. "
                    + f"Erro oficial: {exc} | Erro fallback: {exc2}"
                )

    return atualizadas


def _formatar_valor_para_site(valor: float, usa_virgula: bool) -> str:
    # Mantem 2 casas e sem separador de milhar.
    # IMPORTANTE: no DRE, se enviar "2264.33" (ponto), o campo pode interpretar como "226433".
    if abs(valor) < 0.005:
        s = "0"
    else:
        s = f"{valor:.2f}"
    if usa_virgula:
        s = s.replace(".", ",")
    return s


def _ajustar_valor_por_descricao_para_site(
    codigo: str,
    descricao: str,
    valor: float,
    forcar_positivo: bool = False,
) -> float:
    """
    Regras:
    - Se forcar_positivo=True, sempre envia/valida valor absoluto.
    - Para codigos L*/M*/N* (Ativo/Passivo/DRE), se a descricao tiver "(-)",
      enviar/validar no site sem sinal negativo.
    """
    cod = _norm_txt(codigo)
    desc = _norm_txt(descricao)
    try:
        v = float(valor)
    except Exception:
        return valor

    if forcar_positivo:
        return abs(v)
    if cod.startswith(("L", "M", "N")) and "(-)" in desc:
        return abs(v)
    return v


def _ajustar_valor_dre_para_input(codigo: str, descricao: str, valor: float) -> float:
    """
    Ajuste base para input do DRE.
    Regra operacional atual: tudo que vai para o site no DRE deve ser positivo.
    """
    v = _ajustar_valor_por_descricao_para_site(
        codigo=codigo,
        descricao=descricao,
        valor=valor,
        forcar_positivo=False,
    )
    return 0.0 if abs(v) <= 0.005 else abs(float(v))


def _forcar_eventos_pos_preencher_input(inp) -> None:
    """
    Alguns campos do site so disparam calculo/validacao no blur/keydown.
    Entao, apos preencher, forca um blur simples via TAB (sem depender de JS interno).
    """
    try:
        inp.press("Tab")
        time.sleep(0.05)
    except Exception:
        try:
            inp.evaluate(
                "(el) => { try { el.dispatchEvent(new Event('change', {bubbles:true})); } catch(e){}; try { el.blur(); } catch(e){} }"
            )
            time.sleep(0.05)
        except Exception:
            pass


def _escrever_input_robusto(inp, texto: str, valor_esperado: float | None = None) -> tuple[bool, str, float | None]:
    """
    Escreve em campo de forma resiliente para mascaras JS do AEF.
    Ordem:
    1) fill
    2) click + Ctrl+A + type
    3) set value via evaluate + dispatch input/change/blur
    Retorna: (ok, valor_lido_str, valor_lido_float)
    """
    valor_lido_str = ""
    valor_lido = None

    def _le():
        nonlocal valor_lido_str, valor_lido
        try:
            valor_lido_str = inp.input_value(timeout=1200) or ""
            valor_lido = _parse_float_tolerante(valor_lido_str)
        except Exception:
            valor_lido_str = ""
            valor_lido = None

    def _bate() -> bool:
        if valor_esperado is None:
            return True if valor_lido_str != "" else False
        if valor_lido is None:
            return False
        return _quase_igual(float(valor_esperado), float(valor_lido), tol_abs=0.03)

    # Alguns campos aceitam apenas ponto como decimal.
    variantes = [str(texto)]
    try:
        t = str(texto)
        if "," in t:
            variantes.append(t.replace(".", "").replace(",", "."))
    except Exception:
        pass
    # remove duplicadas preservando ordem
    variantes = list(dict.fromkeys(variantes))

    for vtxt in variantes:
        # 1) fill
        try:
            inp.fill(vtxt)
            _forcar_eventos_pos_preencher_input(inp)
            _le()
            if _bate():
                return True, valor_lido_str, valor_lido
        except Exception:
            pass

        # 2) type
        try:
            inp.click()
            inp.press("Control+A")
            inp.type(vtxt, delay=18)
            _forcar_eventos_pos_preencher_input(inp)
            _le()
            if _bate():
                return True, valor_lido_str, valor_lido
        except Exception:
            pass

        # 3) JS set + eventos
        try:
            inp.evaluate(
                """(el, v) => {
                    try { el.focus(); } catch(e) {}
                    try { el.value = String(v); } catch(e) {}
                    try { el.dispatchEvent(new Event('input', {bubbles:true})); } catch(e) {}
                    try { el.dispatchEvent(new Event('change', {bubbles:true})); } catch(e) {}
                    try { el.blur(); } catch(e) {}
                }""",
                vtxt,
            )
            time.sleep(0.05)
            _le()
            if _bate():
                return True, valor_lido_str, valor_lido
        except Exception:
            pass

    return False, valor_lido_str, valor_lido


def _normalizar_numero_str(texto: str) -> str:
    t = (texto or "").strip()
    # remove espacos e separador de milhar comum
    t = t.replace("\u00a0", "").replace(" ", "")
    # Se tem virgula e ponto, assume ponto milhar e virgula decimal: 1.234,56
    if "," in t and "." in t:
        t = t.replace(".", "").replace(",", ".")
    elif "," in t and "." not in t:
        # Assume virgula decimal
        t = t.replace(",", ".")
    elif "." in t and "," not in t:
        # Caso comum na tela: "12.000" (milhar com ponto, sem casas decimais).
        # Se todas as partes apos o primeiro ponto tiverem 3 digitos, trata como milhar.
        parts = t.split(".")
        if len(parts) > 1 and all(len(p) == 3 for p in parts[1:] if p != ""):
            t = "".join(parts)
    return t


def _parse_float_tolerante(texto: str) -> float | None:
    t_raw = (texto or "").strip()
    if not t_raw:
        return None

    # Normaliza traços comuns para "-"
    t = (
        t_raw.replace("\u2212", "-")  # minus sign
        .replace("\u2013", "-")       # en dash
        .replace("\u2014", "-")       # em dash
    )

    # Regra solicitada:
    # - "-" colado no numero => negativo
    # - "-" no comeco com espaco apos ele => marcador (nao altera sinal)
    negativo = False

    # Ex.: "(1.234,56)" => negativo (padrao contabil)
    if re.match(r"^\(\s*.*\s*\)$", t):
        negativo = True
        t = t[1:-1].strip()

    # Remove sinais no comeco conforme regra de "colado x marcador"
    while t.startswith("-"):
        apos = t[1:]
        if apos.startswith(" "):
            # Marcador visual: "- 1.234,56" => positivo
            t = apos.lstrip()
            continue
        # Colado: "-1.234,56" / "-R$1.234,56" => negativo
        negativo = True
        t = apos
        break

    # Remove simbolo de moeda e espacos residuais
    t = t.replace("R$", "").replace("$", "").strip()

    t_norm = _normalizar_numero_str(t)
    if not t_norm:
        return None
    try:
        val = float(t_norm)
        return -abs(val) if negativo else val
    except Exception:
        return None


def _detectar_decimal_virgula(frame) -> bool:
    """
    Tenta inferir se o input na tela usa virgula como decimal.
    Se nao der para ler, default False (ponto).
    """
    try:
        # Se houver qualquer valor/total com ",dd" na tela, assume BR.
        if frame.locator("text=/\\d+,\\d{2}/").count() > 0:
            return True
        if frame.locator("text=/\\d+\\.\\d{3},\\d{2}/").count() > 0:
            return True
    except Exception:
        pass

    try:
        inp = frame.locator("input[type='text'], input[type='number']").first
        if inp.count() == 0:
            return SITE_DECIMAL_VIRGULA_PADRAO
        v = inp.input_value(timeout=1000) or ""
        return "," in v and "." not in v
    except Exception:
        return SITE_DECIMAL_VIRGULA_PADRAO


def _quase_igual(a: float | None, b: float | None, tol_abs: float = 0.02) -> bool:
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol_abs
    except Exception:
        return False


def _filtrar_divergencias_dre_dinamicas(
    divergencias: list[dict],
    comparacao: dict[str, dict[str, dict]],
    tol_abs: float = 0.02,
) -> list[dict]:
    """
    Regras dinamicas para reduzir falso-positivo em readonly do DRE:
    - N32: aceita inversao de sinal quando o modulo bater.
    - N82: aceita zero quando o site consolidar rateio como 0.
    - N83: valida por formula da tela (N15 - N32 - N82), por coluna.
    """
    if not divergencias:
        return divergencias

    def _obs(cod: str, col: str) -> float | None:
        try:
            return comparacao.get(cod, {}).get(col, {}).get("observado", None)
        except Exception:
            return None

    filtradas: list[dict] = []
    for d in divergencias:
        try:
            cod = _norm_txt(str(d.get("codigo") or ""))
            col = _norm_txt(str(d.get("coluna") or ""))
            st = _norm_txt(str(d.get("status") or ""))
            esp = d.get("esperado", None)
            obs = d.get("observado", None)

            if st != "skip_readonly":
                filtradas.append(d)
                continue

            # N32: aceita modulo (site pode inverter sinal em despesas/totais).
            if cod == "N32":
                if _quase_igual(abs(float(esp)), abs(float(obs)), tol_abs=tol_abs):
                    continue
                filtradas.append(d)
                continue

            # N82: em varias telas o rateio readonly vem zerado.
            if cod == "N82":
                if _quase_igual(float(obs), 0.0, tol_abs=tol_abs):
                    continue
                filtradas.append(d)
                continue

            # N83: resultado departamental apos rateio, usa formula da propria tela.
            if cod == "N83":
                n15 = _obs("N15", col)
                n32 = _obs("N32", col)
                n82 = _obs("N82", col)
                if n15 is not None and n32 is not None and n82 is not None:
                    calc = float(n15) - float(n32) - float(n82)
                    if _quase_igual(calc, float(obs), tol_abs=tol_abs):
                        continue
                filtradas.append(d)
                continue

            filtradas.append(d)
        except Exception:
            filtradas.append(d)

    return filtradas


def _localizar_input_por_nivel(frame, nivel: str):
    """
    Dado "L1", encontra o input correspondente na mesma linha.
    Evita o bug de sempre pegar o primeiro input da tela.
    """
    nivel = (nivel or "").strip().upper()
    if not nivel:
        return None

    def _primeiro_input_editavel(container):
        try:
            inputs = container.locator("input[type='text'], input[type='number'], input:not([type])")
            n = inputs.count()
            for i in range(n):
                inp = inputs.nth(i)
                try:
                    if not inp.is_visible():
                        continue
                except Exception:
                    pass
                try:
                    if inp.is_disabled():
                        continue
                except Exception:
                    pass
                try:
                    ro = (inp.get_attribute("readonly") or "").strip().lower()
                    if ro in ["readonly", "true", "1"]:
                        continue
                except Exception:
                    pass
                return inp
        except Exception:
            pass
        return None

    # 1) Procura o codigo exato em elementos visiveis (evita pegar texto oculto/template).
    label = None
    try:
        cand = frame.locator("td, th, span, label, strong, b, div").filter(
            has_text=re.compile(rf"^\s*{re.escape(nivel)}\s*$", re.I)
        )
        melhores = []
        for i in range(cand.count()):
            el = cand.nth(i)
            try:
                if not el.is_visible():
                    continue
                box = el.bounding_box() or {}
                x = float(box.get("x", 99999))
                y = float(box.get("y", 99999))
                # Coluna de codigo costuma ficar mais a esquerda.
                if x > 220:
                    continue
                melhores.append((y, x, el))
            except Exception:
                continue
        if melhores:
            melhores.sort(key=lambda t: (t[0], t[1]))
            label = melhores[0][2]
    except Exception:
        pass

    # 2) Fallback estrito por XPath, exigindo elemento visivel depois.
    if label is None:
        try:
            cand2 = frame.locator(f"xpath=//*[normalize-space()='{nivel}']")
            for i in range(cand2.count()):
                el = cand2.nth(i)
                try:
                    if el.is_visible():
                        label = el
                        break
                except Exception:
                    continue
        except Exception:
            pass
        if label is None:
            return None

    # 3) Tenta o TR mais proximo.
    try:
        tr = label.locator("xpath=ancestor::tr[1]").first
        if tr.count() > 0:
            inp = _primeiro_input_editavel(tr)
            if inp is not None and inp.count() > 0:
                return inp
    except Exception:
        pass

    # 4) Fallback: bloco mais proximo (evita usar div global da pagina).
    try:
        divc = label.locator("xpath=ancestor::*[self::div or self::li][1]").first
        if divc.count() > 0:
            inp = _primeiro_input_editavel(divc)
            if inp is not None and inp.count() > 0:
                return inp
    except Exception:
        pass

    # 5) Fallback: primeiros inputs apos o label (curto, evita cair em campo distante).
    try:
        inputs = label.locator("xpath=following::input")
        n = min(inputs.count(), 4)
        for i in range(n):
            inp = inputs.nth(i)
            try:
                if inp.is_disabled():
                    continue
            except Exception:
                pass
            try:
                ro = (inp.get_attribute("readonly") or "").strip().lower()
                if ro in ["readonly", "true", "1"]:
                    continue
            except Exception:
                pass
            return inp
    except Exception:
        pass

    # 6) Fallback por proximidade visual no entorno do codigo.
    try:
        if label is not None and label.count() > 0:
            b_label = label.bounding_box() or {}
            ly = float(b_label.get("y", -1))
            lx = float(b_label.get("x", -1))
            if ly >= 0:
                cands = frame.locator("input[type='text'], input[type='number'], input:not([type])")
                best = None
                best_score = float("inf")
                n = cands.count()
                for i in range(n):
                    inp = cands.nth(i)
                    try:
                        if not inp.is_visible():
                            continue
                    except Exception:
                        continue
                    try:
                        if inp.is_disabled():
                            continue
                    except Exception:
                        pass
                    try:
                        ro = (inp.get_attribute("readonly") or "").strip().lower()
                        if ro in ["readonly", "true", "1"]:
                            continue
                    except Exception:
                        pass
                    try:
                        b = inp.bounding_box() or {}
                        iy = float(b.get("y", -1))
                        ix = float(b.get("x", -1))
                        if iy < 0 or ix < 0:
                            continue
                        # mesma linha visual (tolerancia vertical curta)
                        dy = abs(iy - ly)
                        if dy > 14:
                            continue
                        # nunca pega caixas muito a esquerda do codigo (evita primeira caixa fixa)
                        if ix < (lx - 10):
                            continue
                        # penaliza distancia horizontal
                        penal = 0.0 if ix >= lx else (lx - ix) * 5.0
                        score = dy + penal + abs(ix - lx) * 0.02
                        if score < best_score:
                            best_score = score
                            best = inp
                    except Exception:
                        continue
                if best is not None:
                    return best
    except Exception:
        pass

    return None


def _scroll_to_top(frame, page) -> None:
    """
    Rola para o topo do CONTAINER certo.
    Importante: em varias telas do AEF o scroll nao e do window, e sim de um div com overflow-y.
    """
    # Preferencia: scroller vertical (overflowY=auto/scroll) com maior area visivel.
    try:
        frame.evaluate(
            """() => {
  function findBestScrollerY() {
    const els = Array.from(document.querySelectorAll('*'));
    let best = null;
    let bestArea = 0;
    for (const el of els) {
      const cs = window.getComputedStyle(el);
      if (!['auto','scroll'].includes(cs.overflowY)) continue;
      if (!el.scrollHeight || !el.clientHeight) continue;
      if (el.scrollHeight <= el.clientHeight + 10) continue;
      const r = el.getBoundingClientRect();
      const area = r.width * r.height;
      if (area > bestArea) { bestArea = area; best = el; }
    }
    return best;
  }
  const sc = findBestScrollerY();
  if (sc) { sc.scrollTop = 0; return true; }
  try { window.scrollTo(0, 0); } catch (e) {}
  return false;
}"""
        )
        return
    except Exception:
        pass

    # Fallback: tenta scroll do window e tecla Home.
    try:
        frame.evaluate("() => window.scrollTo(0, 0)")
        return
    except Exception:
        pass
    try:
        page.keyboard.press("Home")
    except Exception:
        pass


def _scroll_down(frame) -> tuple[float, float]:
    """
    Rola a pagina para baixo. Retorna (scrollY_antes, scrollY_depois).
    """
    try:
        antes, depois = frame.evaluate(
            """(step) => {
  function findBestScrollerY() {
    const els = Array.from(document.querySelectorAll('*'));
    let best = null;
    let bestArea = 0;
    for (const el of els) {
      const cs = window.getComputedStyle(el);
      if (!['auto','scroll'].includes(cs.overflowY)) continue;
      if (!el.scrollHeight || !el.clientHeight) continue;
      if (el.scrollHeight <= el.clientHeight + 10) continue;
      const r = el.getBoundingClientRect();
      const area = r.width * r.height;
      if (area > bestArea) { bestArea = area; best = el; }
    }
    return best;
  }
  const sc = findBestScrollerY();
  if (sc) {
    const y1 = sc.scrollTop || 0;
    sc.scrollBy(0, step);
    const y2 = sc.scrollTop || 0;
    return [y1, y2];
  }
  const y1 = window.scrollY || 0;
  window.scrollBy(0, step);
  const y2 = window.scrollY || 0;
  return [y1, y2];
}""",
            SCROLL_STEP_PX,
        )
        return float(antes), float(depois)
    except Exception:
        return 0.0, 0.0


def _localizar_input_por_codigo_com_scroll(frame, page, codigo: str):
    """
    Alguns formularios carregam/atualizam o DOM conforme o scroll.
    Aqui tentamos localizar e, se nao achar, rolamos para baixo algumas vezes.
    """
    _scroll_to_top(frame, page)

    for _ in range(SCROLL_MAX_PASSOS_POR_CODIGO + 1):
        inp = _localizar_input_por_nivel(frame, nivel=codigo)
        if inp is not None and inp.count() > 0:
            return inp

        antes, depois = _scroll_down(frame)
        if depois <= antes:
            break

    return None


def _competencia_mes_anterior(hoje: datetime | None = None) -> str:
    if hoje is None:
        hoje = datetime.now()
    # Mes anterior com ajuste de ano.
    if hoje.month == 1:
        ano = hoje.year - 1
        mes = 12
    else:
        ano = hoje.year
        mes = hoje.month - 1

    meses = {
        1: "jan",
        2: "fev",
        3: "mar",
        4: "abr",
        5: "mai",
        6: "jun",
        7: "jul",
        8: "ago",
        9: "set",
        10: "out",
        11: "nov",
        12: "dez",
    }
    return f"{meses[mes]} {ano}"


def _resolver_competencia(competencia_arg: str) -> tuple[str, str]:
    competencia_manual = (competencia_arg or "").strip()
    if competencia_manual:
        return competencia_manual, "manual (--competencia)"

    hoje = datetime.now()
    competencia_auto = _competencia_mes_anterior(hoje)
    # Regra operacional: sempre preencher a competencia do mes anterior.
    # Exemplo: data base 13/03/2026 -> competencia "fev 2026".
    origem = f"automatica (mes anterior; data base {hoje.strftime('%d/%m/%Y')})"
    return competencia_auto, origem


def _rodar_etapa_4_pos_script3(empresa: str) -> int:
    emp_norm = _normalizar_empresa(empresa)
    if emp_norm not in EMPRESAS_COM_ETAPA_4:
        return 0

    if not os.path.isfile(SCRIPT_4_MOVER):
        log_linha(f"ERRO: Script 4 nao encontrado: {SCRIPT_4_MOVER}")
        return 1

    cmd = [
        sys.executable,
        SCRIPT_4_MOVER,
        "--empresa",
        emp_norm,
    ]
    log_linha(f"[SCRIPT 4] Comando: {' '.join(cmd)}")

    proc = subprocess.Popen(
        cmd,
        cwd=BASE_DIR,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    assert proc.stdout is not None
    for linha in proc.stdout:
        log_linha(f"[SCRIPT 4] {linha.rstrip()}")
    return proc.wait()


def _aguardar_carregar_pagina(page) -> None:
    try:
        page.wait_for_load_state("domcontentloaded", timeout=ESPERA_DOM_MS)
    except Exception:
        pass
    try:
        # Em SPA com polling/telemetria, networkidle pode nunca estabilizar.
        page.wait_for_load_state("networkidle", timeout=ESPERA_NETWORKIDLE_MS)
    except Exception:
        pass


def _goto_com_retry(page, url: str) -> None:
    ultima_exc: Exception | None = None
    for tentativa in range(1, TENTATIVAS_GOTO + 1):
        try:
            # Evita esperar "load" (apps SPA podem manter requests pendentes por muito tempo).
            page.goto(url, wait_until="domcontentloaded", timeout=GOTO_TIMEOUT_MS)
            return
        except Exception as exc:
            ultima_exc = exc
            if tentativa < TENTATIVAS_GOTO:
                time.sleep(ESPERA_ENTRE_GOTO_S)
                continue
            raise ultima_exc


def _esperar_frame_submissions(page, timeout_ms: int) -> object:
    """
    A tela de submissions pode demorar a popular a tabela e, em alguns casos,
    o conteudo pode estar em um iframe. Aqui fazemos polling em todos os frames
    ate encontrar um elemento tipico da tabela.
    """
    inicio = time.monotonic()
    timeout_s = max(1, int(timeout_ms / 1000))

    while time.monotonic() - inicio < timeout_s:
        candidatos = []
        try:
            candidatos.append(page.main_frame)
        except Exception:
            pass
        try:
            for fr in page.frames:
                if fr not in candidatos:
                    candidatos.append(fr)
        except Exception:
            pass

        for fr in candidatos:
            try:
                if fr.locator("a:has-text('Editar')").count() > 0:
                    return fr
            except Exception:
                pass
            try:
                if fr.locator("text=Periodo").count() > 0:
                    return fr
            except Exception:
                pass
            try:
                if fr.locator("tr.normal").count() > 0:
                    return fr
            except Exception:
                pass

        time.sleep(0.2)

    return page


def _ir_para_submissions(page) -> object:
    url = "https://aefhondabr.nx-services.com/hondabr/index.html#/submissions"
    log_linha(f"Acessando: {url}")
    _goto_com_retry(page, url)
    _aguardar_carregar_pagina(page)

    frame = _esperar_frame_submissions(page, timeout_ms=PLAYWRIGHT_TIMEOUT_MS)

    # A tabela dessa tela pode variar; ja vimos pelo menos 2 layouts:
    # - layout 1: <tr class="normal"> ... <a class="editForm">Editar</a>
    # - layout 2: header "Periodo" e link "Editar" sem classe.
    try:
        page.wait_for_function(
            "() => window.location && window.location.hash && window.location.hash.toLowerCase().includes('submissions')",
            timeout=10_000,
        )
    except Exception:
        pass

    # Validacao rapida: evita somar varios waits longos em cascata.
    checks = [
        "text=Periodo",
        "a:has-text('Editar')",
        "tr.normal",
    ]
    for sel in checks:
        try:
            frame.locator(sel).first.wait_for(state="visible", timeout=4_000)
            return frame
        except Exception:
            continue

    caminho = _salvar_screenshot(page, "submissions_nao_carregou")
    raise RuntimeError(
        "Nao consegui carregar a tabela de submissions." + (f" Screenshot: {caminho}" if caminho else "")
    )


def _validar_login_ok(page) -> object:
    """
    Confirma o login automaticamente.

    Regra pratica:
    - Login OK => conseguimos acessar /submissions e renderizar a tabela.
    - Se falhar, nao segue o fluxo.
    """
    try:
        _aguardar_carregar_pagina(page)
    except Exception:
        pass

    try:
        h = (page.url or "")
        if "#/login" in h.lower() or "login" in h.lower():
            # Ainda na rota de login: tenta mesmo assim abrir submissions para confirmar.
            pass
    except Exception:
        pass

    try:
        frame = _ir_para_submissions(page)
    except Exception as exc:
        caminho = _salvar_screenshot(page, "login_nao_confirmado")
        raise RuntimeError(
            "Nao consegui confirmar o login automaticamente (submissions nao carregou)."
            + (f" Screenshot: {caminho}" if caminho else "")
            + (f" Detalhe: {exc}" if exc else "")
        )

    return frame


def _resolver_frame_conteudo_submissions(page):
    """
    Em algumas execucoes, o conteudo da tabela aparece em um iframe.
    Esta funcao tenta achar o frame onde os seletores da tabela existem.
    """
    frames = []
    try:
        frames = page.frames
    except Exception:
        return page

    candidatos = []
    try:
        candidatos.append(page.main_frame)
    except Exception:
        pass
    for fr in frames:
        if fr not in candidatos:
            candidatos.append(fr)

    # Tenta identificar o frame pela presenca de elementos tipicos da tela.
    for fr in candidatos:
        try:
            if fr.locator("text=Periodo").count() > 0:
                return fr
        except Exception:
            pass
        try:
            if fr.locator("a:has-text('Editar')").count() > 0:
                return fr
        except Exception:
            pass
        try:
            if fr.locator("tr.normal").count() > 0:
                return fr
        except Exception:
            pass

    return page


def _clicar_editar_na_competencia(frame, page, competencia: str) -> None:
    # Ex.: "jan 2026" (do HTML que voce colou).
    competencia_norm = competencia.strip().lower()

    # Tenta localizar a linha que contenha a competencia.
    # Layout A (HTML que voce colou): <tr class="normal"> ... <a>jan 2026</a> ... <a class="editForm">Editar</a>
    # Layout B (seu print): competencia aparece como texto na 1a coluna.
    td_primeira_coluna = frame.locator("td:nth-child(1)").filter(
        has_text=re.compile(rf"^{re.escape(competencia_norm)}$", re.I)
    )
    linha = frame.locator("tr").filter(has=td_primeira_coluna).first

    # Fallbacks
    if linha.count() == 0:
        link_comp = frame.locator("a").filter(has_text=re.compile(rf"^{re.escape(competencia_norm)}$", re.I)).first
        if link_comp.count() > 0:
            linha = frame.locator("tr", has=link_comp).first
        else:
            linha = frame.locator("tr").filter(
                has_text=re.compile(rf"\\b{re.escape(competencia_norm)}\\b", re.I)
            ).first

    try:
        linha.wait_for(state="visible", timeout=10_000)
    except Exception:
        caminho = _salvar_screenshot(page, "competencia_nao_encontrada")
        raise RuntimeError(
            f"Competencia nao encontrada na tabela: '{competencia}'."
            + (f" Screenshot: {caminho}" if caminho else "")
        )

    # "Editar" na mesma linha:
    # - layout 1: <a class="editForm">Editar</a>
    # - layout 2: link "Editar" sem classe
    editar = linha.locator("a.editForm").first
    if editar.count() == 0:
        editar = linha.locator("a").filter(has_text=re.compile(r"^Editar$", re.I)).first
    try:
        editar.wait_for(state="visible", timeout=10_000)
    except Exception:
        caminho = _salvar_screenshot(page, "botao_editar_nao_encontrado")
        raise RuntimeError(
            "Nao encontrei o botao 'Editar' na linha da competencia."
            + (f" Screenshot: {caminho}" if caminho else "")
        )

    log_linha(f"Clicando em Editar da competencia: {competencia_norm}")
    editar.click()
    _aguardar_carregar_pagina(page)

    caminho_ok = _salvar_screenshot(page, "apos_clicar_editar")
    if caminho_ok:
        log_linha(f"Screenshot: {caminho_ok}")


def _clicar_editar_na_competencia_com_retry(frame, page, competencia: str) -> None:
    ultima_exc: Exception | None = None
    for tentativa in range(1, TENTATIVAS_CLICAR_EDITAR + 1):
        try:
            _clicar_editar_na_competencia(frame, page, competencia=competencia)
            return
        except Exception as exc:
            ultima_exc = exc
            log_linha(f"AVISO: tentativa {tentativa}/{TENTATIVAS_CLICAR_EDITAR} falhou: {exc}")
            if tentativa < TENTATIVAS_CLICAR_EDITAR:
                time.sleep(ESPERA_ENTRE_TENTATIVAS_S)
                _aguardar_carregar_pagina(page)
                continue
            raise ultima_exc


def _esperar_frame_edicao(page, timeout_ms: int):
    """
    Apos clicar em Editar, a tela pode carregar em outro frame/contexto.
    Polling ate encontrar:
    - texto "Veiculos Novos" (com ou sem acento)
    - botao/submit "Confirmar"
    """
    inicio = time.monotonic()
    timeout_s = max(1, int(timeout_ms / 1000))

    padrao_veiculos = re.compile(r"Ve[ií]culos\\s+Novos", re.I)

    def _frame_tem_select_formulario(fr) -> bool:
        """
        Prefere o frame que tem o dropdown do formulario (Ativo/Passivo/DRE).
        Isso evita "pegar o frame errado" so porque existe um botao Confirmar em outro container.
        """
        try:
            opt_ativo = fr.locator("option").filter(has_text=re.compile(r"^Ativo$", re.I))
            opt_passivo = fr.locator("option").filter(has_text=re.compile(r"^Passivo$", re.I))
            opt_dre = fr.locator("option").filter(has_text=re.compile(r"^DRE$", re.I))
            sel = fr.locator("select").filter(has=opt_ativo).filter(has=opt_passivo).filter(has=opt_dre)
            return sel.count() > 0
        except Exception:
            return False

    while time.monotonic() - inicio < timeout_s:
        candidatos = []
        try:
            candidatos.append(page.main_frame)
        except Exception:
            pass
        try:
            for fr in page.frames:
                if fr not in candidatos:
                    candidatos.append(fr)
        except Exception:
            pass

        # Preferencia: frame que tem o dropdown Ativo/Passivo/DRE (mais preciso).
        for fr in candidatos:
            try:
                if _frame_tem_select_formulario(fr):
                    return fr
            except Exception:
                pass

        for fr in candidatos:
            try:
                if fr.locator("input[type='submit'][value='Confirmar'], input[value='Confirmar']").count() > 0:
                    # Indicadores fracos para evitar frame errado.
                    try:
                        if fr.locator("text=/\\bN\\d{1,3}\\b/i").count() > 0:
                            return fr
                    except Exception:
                        pass
                    try:
                        if fr.locator("text=/ATIVO\\s+CIRCULANTE/i").count() > 0:
                            return fr
                    except Exception:
                        pass
                    try:
                        if fr.locator("text=/PATRIMONIO\\s+LIQUIDO/i").count() > 0:
                            return fr
                    except Exception:
                        pass
                    try:
                        if fr.locator("text=/DEMONSTRATIVO\\s+DE\\s+RESULTADO/i").count() > 0:
                            return fr
                    except Exception:
                        pass
                    return fr
            except Exception:
                pass
            try:
                if fr.locator("text=/Ve[ií]culos\\s+Novos/i").count() > 0:
                    return fr
            except Exception:
                pass
            try:
                if fr.locator("text=" + TEXTO_MENU_VEICULOS_NOVOS).count() > 0:
                    return fr
            except Exception:
                pass
            try:
                if padrao_veiculos.search(fr.inner_text(timeout=200) or ""):
                    return fr
            except Exception:
                pass

        time.sleep(0.5)

    return page


def _selecionar_veiculos_novos_ativo_e_confirmar(page) -> None:
    """
    Historicamente este passo selecionava um dropdown de "Veiculos Novos".
    Na pratica, o que precisamos aqui e entrar no formulario correto (Ativo/Passivo/DRE).
    Mantido por compatibilidade com a flag --ate-confirmar.
    """
    try:
        _selecionar_formulario_e_confirmar(page, formulario="Ativo")
        return
    except Exception:
        # Fallback: tenta o modo antigo (se existir na tela).
        pass

    _aguardar_carregar_pagina(page)
    frame = _esperar_frame_edicao(page, timeout_ms=PLAYWRIGHT_TIMEOUT_MS)

    # Localiza o dropdown associado ao texto "Veiculos Novos" (com ou sem acento).
    alvo_txt = frame.locator("text=/Ve[ií]culos\\s+Novos/i").first
    if alvo_txt.count() == 0:
        alvo_txt = frame.locator(f"text={TEXTO_MENU_VEICULOS_NOVOS}").first

    if alvo_txt.count() == 0:
        caminho = _salvar_screenshot(page, "edicao_veiculos_novos_nao_encontrado")
        raise RuntimeError(
            "Nao encontrei o texto/menu 'Veiculos Novos' na tela de edicao."
            + (f" Screenshot: {caminho}" if caminho else "")
        )

    select = None
    try:
        # Tenta pegar um <select> na mesma linha/tabela.
        select = alvo_txt.locator("xpath=ancestor::tr[1]//select[1]").first
    except Exception:
        select = None

    if not select or select.count() == 0:
        try:
            select = alvo_txt.locator("xpath=ancestor::td[1]//select[1]").first
        except Exception:
            select = None

    if not select or select.count() == 0:
        try:
            select = alvo_txt.locator("xpath=following::select[1]").first
        except Exception:
            select = None

    if not select or select.count() == 0:
        caminho = _salvar_screenshot(page, "edicao_select_nao_encontrado")
        raise RuntimeError(
            "Nao encontrei o dropdown (<select>) para 'Veiculos Novos'."
            + (f" Screenshot: {caminho}" if caminho else "")
        )

    # Seleciona "Ativo"
    try:
        select.wait_for(state="visible", timeout=10_000)
    except Exception:
        pass

    log_linha(f"Selecionando dropdown '{TEXTO_MENU_VEICULOS_NOVOS}' -> '{OPCAO_VEICULOS_NOVOS}'")
    try:
        select.select_option(label=OPCAO_VEICULOS_NOVOS)
    except Exception:
        # Fallback: tenta por value.
        select.select_option(value=OPCAO_VEICULOS_NOVOS)

    # Clica Confirmar
    confirmar = frame.locator(
        "input#main\\:j_id86, input[type='submit'][value='Confirmar'], input[value='Confirmar']"
    ).first
    if confirmar.count() == 0:
        caminho = _salvar_screenshot(page, "edicao_confirmar_nao_encontrado")
        raise RuntimeError(
            "Nao encontrei o botao 'Confirmar'."
            + (f" Screenshot: {caminho}" if caminho else "")
        )

    for tentativa in range(1, TENTATIVAS_CONFIRMAR + 1):
        try:
            confirmar.wait_for(state="visible", timeout=10_000)
            log_linha("Clicando em Confirmar.")
            confirmar.click()
            _aguardar_carregar_pagina(page)
            caminho_ok = _salvar_screenshot(page, "apos_confirmar")
            if caminho_ok:
                log_linha(f"Screenshot: {caminho_ok}")
            return
        except Exception as exc:
            log_linha(f"AVISO: falha ao clicar Confirmar (tentativa {tentativa}/{TENTATIVAS_CONFIRMAR}): {exc}")
            if tentativa < TENTATIVAS_CONFIRMAR:
                time.sleep(ESPERA_ENTRE_TENTATIVAS_CONFIRMAR_S)
                _aguardar_carregar_pagina(page)
                continue
            raise


def _aguardar_dre_carregar(page) -> None:
    """
    Aguarda o DRE renderizar (algumas vezes a tela fica branca por alguns segundos).
    Condicoes de sucesso:
    - aparece algum codigo Nxx (ex.: N16)
    - ou aparece header VN/VSN/...
    - ou aparece uma grade com varios inputs
    """
    inicio = time.monotonic()
    timeout_s = max(1, int(DRE_ESPERAR_TIMEOUT_MS / 1000))

    while time.monotonic() - inicio < timeout_s:
        _aguardar_carregar_pagina(page)
        frame = _esperar_frame_edicao(page, timeout_ms=5_000)

        try:
            # Aplica zoom-out (algumas grades so renderizam apos "caber" na tela).
            if DRE_ZOOM_PERCENT:
                frame.evaluate(
                    "(z) => { try { document.documentElement.style.zoom = String(z) + '%'; } catch (e) {} }",
                    DRE_ZOOM_PERCENT,
                )
        except Exception:
            pass

        # OBS: XPath do browser nao suporta regex (ex.: re:test). Use seletor de texto do Playwright.
        try:
            if frame.locator("text=/\\bN\\d{1,3}\\b/i").count() > 0:
                return
        except Exception:
            pass

        try:
            if frame.locator("xpath=//*[normalize-space()='VN' or normalize-space()='VSN' or normalize-space()='VED']").count() > 0:
                return
        except Exception:
            pass

        time.sleep(1)

    caminho = _salvar_screenshot(page, "dre_nao_carregou")
    raise RuntimeError(
        "Tela DRE nao carregou a grade (ficou vazia/sem headers)."
        + (f" Screenshot: {caminho}" if caminho else "")
    )


def _preencher_ativo_no_site(page, caminho_xlsx: str) -> None:
    """
    Preenche os valores do Ativo no site, batendo por codigo L1/L2/... e preenchendo apenas o valor.
    """
    linhas = ler_linhas_planilha_xlsx(caminho_xlsx, aba=ABA_ATIVO)
    if not linhas:
        raise RuntimeError(
            "Nao encontrei linhas com codigo (ex.: L1, L2, ...) e valor calculado na aba Ativo do XLSX. "
            "Se o arquivo tiver formulas, confirme se foi salvo com os valores calculados."
        )

    _aguardar_carregar_pagina(page)
    frame = _esperar_frame_edicao(page, timeout_ms=PLAYWRIGHT_TIMEOUT_MS)
    usa_virgula = _detectar_decimal_virgula(frame)
    _scroll_to_top(frame, page)

    log_linha(f"Preenchendo Ativo (linhas: {len(linhas)}; decimal_virgula={usa_virgula}).")

    faltando: list[str] = []
    preenchidas = 0
    repetidas: dict[str, int] = {}

    itens = linhas
    if MAX_LINHAS_POR_TESTE and MAX_LINHAS_POR_TESTE > 0:
        itens = linhas[:MAX_LINHAS_POR_TESTE]

    for item in itens:
        nivel = item["codigo"]
        valor = _ajustar_valor_por_descricao_para_site(
            codigo=item["codigo"],
            descricao=item.get("descricao", ""),
            valor=item["valor"],
        )

        inp = _localizar_input_por_nivel(frame, nivel=nivel)
        if not inp or inp.count() == 0:
            inp = _localizar_input_por_codigo_com_scroll(frame, page, codigo=nivel)
        if not inp or inp.count() == 0:
            faltando.append(nivel)
            continue

        # Pula campos bloqueados (totais / calculados)
        try:
            if inp.is_disabled():
                continue
        except Exception:
            pass
        try:
            ro = (inp.get_attribute("readonly") or "").strip().lower()
            if ro in ["readonly", "true", "1"]:
                continue
        except Exception:
            pass

        txt_val = _formatar_valor_para_site(valor, usa_virgula=usa_virgula)
        try:
            inp.scroll_into_view_if_needed(timeout=2000)
        except Exception:
            pass
        ok_write, lido_txt, lido_val = _escrever_input_robusto(inp, txt_val, valor_esperado=float(valor))
        if not ok_write:
            faltando.append(nivel)
            continue
        preenchidas += 1

        # Verificacao rapida: le de volta e compara (tolerancia pequena).
        try:
            f_lido = lido_val if lido_val is not None else _parse_float_tolerante(lido_txt)
            if f_lido is not None:
                if abs(f_lido - float(valor)) > 0.03:
                    repetidas[nivel] = repetidas.get(nivel, 0) + 1
        except Exception:
            pass

    log_linha(f"Ativo preenchido: {preenchidas}/{len(itens)}.")
    if preenchidas == 0:
        caminho_zero = _salvar_screenshot(page, "ativo_zero_preenchido")
        raise RuntimeError(
            f"Ativo preenchido: 0 linhas. Niveis sem escrita: {len(faltando)}/{len(itens)}. "
            "Tela provavelmente sem campos editaveis carregados, formato decimal incompativel ou mapeamento incorreto."
            + (f" Screenshot: {caminho_zero}" if caminho_zero else "")
        )
    if faltando:
        log_linha(f"AVISO: niveis nao preenchidos (nao encontrados/sem input): {', '.join(faltando[:30])}")
        if len(faltando) > 30:
            log_linha(f"AVISO: (+{len(faltando) - 30} niveis omitidos do log)")
    if repetidas:
        # Indicador de possivel mapeamento errado (campo nao refletiu o valor esperado).
        tops = sorted(repetidas.items(), key=lambda kv: kv[1], reverse=True)[:10]
        log_linha("AVISO: divergencias ao ler de volta (top 10): " + ", ".join([f"{k}={v}" for k, v in tops]))

    caminho_ok = _salvar_screenshot(page, "apos_preencher_ativo")
    if caminho_ok:
        log_linha(f"Screenshot: {caminho_ok}")

    # Dump de comparacao (site vs XLSX) para debug.
    try:
        comparacao: dict[str, dict] = {}
        divergencias: list[dict] = []

        for item in itens:
            codigo = item["codigo"]
            esp = _ajustar_valor_por_descricao_para_site(
                codigo=item["codigo"],
                descricao=item.get("descricao", ""),
                valor=item["valor"],
            )
            try:
                esp_f = float(esp)
            except Exception:
                continue

            comparacao[codigo] = {"esperado": esp_f}

            inp = _localizar_input_por_nivel(frame, nivel=codigo)
            if not inp or inp.count() == 0:
                inp = _localizar_input_por_codigo_com_scroll(frame, page, codigo=codigo)

            if not inp or inp.count() == 0:
                comparacao[codigo]["status"] = "skip_no_input"
                if abs(esp_f) > 0.005:
                    divergencias.append({"codigo": codigo, "status": "skip_no_input", "esperado": esp_f})
                continue

            try:
                comparacao[codigo]["readonly_attr"] = inp.get_attribute("readonly") or ""
            except Exception:
                pass
            try:
                comparacao[codigo]["disabled"] = bool(inp.is_disabled())
            except Exception:
                pass

            try:
                vtxt = inp.input_value(timeout=1000) or ""
                comparacao[codigo]["observado_str"] = vtxt
                obs_f = _parse_float_tolerante(vtxt)
                comparacao[codigo]["observado"] = obs_f
            except Exception:
                obs_f = None

            try:
                if inp.is_disabled():
                    comparacao[codigo]["status"] = "skip_disabled"
                    if abs(esp_f) > 0.005:
                        divergencias.append({"codigo": codigo, "status": "skip_disabled", "esperado": esp_f, "observado": obs_f})
                    continue
            except Exception:
                pass

            try:
                ro = (inp.get_attribute("readonly") or "").strip().lower()
                if ro in ["readonly", "true", "1"]:
                    comparacao[codigo]["status"] = "skip_readonly"
                    if abs(esp_f) > 0.005 and not _quase_igual(esp_f, obs_f):
                        divergencias.append({"codigo": codigo, "status": "skip_readonly", "esperado": esp_f, "observado": obs_f})
                    continue
            except Exception:
                pass

            if _quase_igual(esp_f, obs_f):
                comparacao[codigo]["status"] = "ok"
            else:
                comparacao[codigo]["status"] = "mismatch"
                divergencias.append({"codigo": codigo, "status": "mismatch", "esperado": esp_f, "observado": obs_f})

        pjson = _salvar_dump_json("ativo_compare_full", comparacao)
        if pjson:
            log_linha(f"Dump Ativo compare (json): {pjson}")
        pjson2 = _salvar_dump_json("ativo_compare_divergencias", divergencias)
        if pjson2:
            log_linha(f"Dump Ativo divergencias (json): {pjson2}")

        ptsv = ""
        try:
            linhas_tsv = ["codigo\tstatus\tesperado\tobservado\tobservado_str"]
            for d in divergencias:
                linhas_tsv.append(
                    f"{d.get('codigo','')}\t{d.get('status','')}\t{d.get('esperado','')}\t{d.get('observado','')}\t{str(d.get('observado_str','')).replace(chr(9),' ')}"
                )
            ptsv = _salvar_dump_texto("ativo_compare_divergencias", "\n".join(linhas_tsv) + "\n")
            if ptsv:
                log_linha(f"Dump Ativo divergencias (tsv): {ptsv}")
        except Exception:
            pass

        # Mantem diagnostico, mas por padrao nao bloqueia o fluxo no Ativo/Passivo.
        if divergencias:
            caminho = _salvar_screenshot(page, "ativo_divergencias")
            msg = (
                "Ativo: valores do site nao batem com o XLSX (incluindo totais/readonly). "
                + (f"Veja: {ptsv or pjson2 or pjson}. " if (ptsv or pjson2 or pjson) else "")
                + (f"Screenshot: {caminho}" if caminho else "")
            )
            if FALHAR_DIVERGENCIA_ATIVO_PASSIVO:
                raise RuntimeError(msg)
            log_linha("AVISO: " + msg)
    except RuntimeError:
        raise
    except Exception:
        pass


def _selecionar_formulario_e_confirmar(page, formulario: str) -> None:
    """
    Seleciona o formulario no dropdown principal (Ativo/Passivo/etc) e clica Confirmar.
    """
    alvo_norm = _norm_header(formulario)

    def _texto_opcao_selecionada(sel):
        try:
            txt = sel.evaluate(
                """el => {
                    const o = el && el.selectedOptions && el.selectedOptions.length ? el.selectedOptions[0] : null;
                    return (o && o.textContent ? o.textContent : '').trim();
                }"""
            )
            return (txt or "").strip()
        except Exception:
            return ""

    def _frame_esta_no_formulario(fr) -> bool:
        # 1) Validacao pelo dropdown selecionado (topo/rodape).
        try:
            opt_ativo = fr.locator("option").filter(has_text=re.compile(r"^Ativo$", re.I))
            opt_passivo = fr.locator("option").filter(has_text=re.compile(r"^Passivo$", re.I))
            opt_dre = fr.locator("option").filter(has_text=re.compile(r"^DRE$", re.I))
            sels = fr.locator("select").filter(has=opt_ativo).filter(has=opt_passivo).filter(has=opt_dre)
            n = sels.count()
            for i in range(n):
                txt = _texto_opcao_selecionada(sels.nth(i))
                if _norm_header(txt) == alvo_norm:
                    return True
            else:
                return False
        except Exception:
            return False
        return False

    def _frame_tem_inputs_editaveis(fr, prefixo: str) -> bool:
        """
        Garante que a grade do formulario terminou de renderizar:
        precisa existir ao menos alguns inputs editaveis nas linhas do prefixo (L/M).
        """
        try:
            return bool(
                fr.evaluate(
                    """(pref) => {
                        const re = new RegExp("\\\\b" + String(pref || "") + "\\\\d{1,3}\\\\b", "i");
                        let editaveis = 0;
                        const linhas = Array.from(document.querySelectorAll("tr, div"));
                        for (const ln of linhas) {
                            const txt = (ln.innerText || "").toUpperCase();
                            if (!re.test(txt)) continue;
                            const ins = ln.querySelectorAll("input");
                            for (const inp of ins) {
                                if (!inp) continue;
                                if (inp.disabled) continue;
                                const ro = String(inp.getAttribute("readonly") || "").trim().toLowerCase();
                                if (ro === "readonly" || ro === "true" || ro === "1") continue;
                                editaveis += 1;
                                if (editaveis >= 3) return true;
                            }
                        }
                        return false;
                    }""",
                    prefixo,
                )
            )
        except Exception:
            return False

    def _aguardar_inputs_editaveis(fr, alvo: str, timeout_s: int = 15) -> bool:
        inicio = time.monotonic()
        while time.monotonic() - inicio < timeout_s:
            if alvo == "ATIVO":
                if _frame_tem_inputs_editaveis(fr, "L"):
                    return True
            elif alvo == "PASSIVO":
                if _frame_tem_inputs_editaveis(fr, "M"):
                    return True
            else:
                return True
            time.sleep(0.2)
        return False

    ultima_exc = None
    tentativas = 4
    for tentativa in range(1, tentativas + 1):
        try:
            _aguardar_carregar_pagina(page)
            frame = _esperar_frame_edicao(page, timeout_ms=PLAYWRIGHT_TIMEOUT_MS)

            # Acha o select correto do formulario (precisa ter Ativo/Passivo/DRE).
            opt_ativo = frame.locator("option").filter(has_text=re.compile(r"^Ativo$", re.I))
            opt_passivo = frame.locator("option").filter(has_text=re.compile(r"^Passivo$", re.I))
            opt_dre = frame.locator("option").filter(has_text=re.compile(r"^DRE$", re.I))
            sels = frame.locator("select").filter(has=opt_ativo).filter(has=opt_passivo).filter(has=opt_dre)
            if sels.count() == 0:
                # Fallback: algum select que tenha DRE (melhor do que "qualquer select").
                sels = frame.locator("select").filter(has=opt_dre)

            if sels.count() == 0:
                caminho = _salvar_screenshot(page, "select_formulario_nao_encontrado")
                raise RuntimeError(
                    "Nao encontrei o dropdown do formulario (Ativo/Passivo)."
                    + (f" Screenshot: {caminho}" if caminho else "")
                )

            # Prefere select visivel mais acima (toolbar superior).
            select_form = None
            melhores = []
            for i in range(sels.count()):
                sel = sels.nth(i)
                try:
                    if not sel.is_visible():
                        continue
                    box = sel.bounding_box() or {}
                    melhores.append((float(box.get("y", 99999)), float(box.get("x", 99999)), sel))
                except Exception:
                    continue
            if melhores:
                melhores.sort(key=lambda t: (t[0], t[1]))
                select_form = melhores[0][2]
            else:
                select_form = sels.first

            if tentativa == 1:
                log_linha(f"Trocando formulario -> {formulario}")
            else:
                log_linha(f"Trocando formulario -> {formulario} (tentativa {tentativa}/{tentativas})")

            try:
                select_form.select_option(label=formulario)
            except Exception:
                # Fallback: busca option pelo texto e usa value real.
                alvo_value = select_form.evaluate(
                    """(el, alvo) => {
                        const norm = s => (s || '').normalize('NFKD').replace(/[\\u0300-\\u036f]/g, '').toUpperCase().replace(/[^A-Z0-9]+/g, '');
                        const t = norm(alvo);
                        const opts = Array.from(el.options || []);
                        const op = opts.find(o => norm(o.textContent || '') === t);
                        return op ? op.value : '';
                    }""",
                    formulario,
                )
                if not alvo_value:
                    raise RuntimeError(f"Opcao '{formulario}' nao encontrada no dropdown.")
                select_form.select_option(value=str(alvo_value))

            # Botao Confirmar (pode ter id diferente; usa value).
            confirmar = select_form.locator("xpath=ancestor::form[1]//input[@value='Confirmar']").first
            if confirmar.count() == 0:
                confirmar = select_form.locator(
                    "xpath=ancestor::div[1]//input[@value='Confirmar'] | xpath=ancestor::div[2]//input[@value='Confirmar']"
                ).first
            if confirmar.count() == 0:
                confirmar = frame.locator("input[type='submit'][value='Confirmar'], input[value='Confirmar']").first

            if confirmar.count() == 0:
                caminho = _salvar_screenshot(page, "confirmar_formulario_nao_encontrado")
                raise RuntimeError(
                    "Nao encontrei o botao Confirmar apos trocar formulario."
                    + (f" Screenshot: {caminho}" if caminho else "")
                )

            confirmar.click()
            _aguardar_carregar_pagina(page)

            frame2 = _esperar_frame_edicao(page, timeout_ms=PLAYWRIGHT_TIMEOUT_MS)
            try:
                _scroll_to_top(frame2, page)
            except Exception:
                pass

            if _frame_esta_no_formulario(frame2):
                if alvo_norm in ("ATIVO", "PASSIVO"):
                    if not _aguardar_inputs_editaveis(frame2, alvo_norm, timeout_s=15):
                        caminho = _salvar_screenshot(page, "formulario_sem_inputs_editaveis")
                        raise RuntimeError(
                            f"Formulario '{formulario}' abriu, mas os campos editaveis nao carregaram a tempo."
                            + (f" Screenshot: {caminho}" if caminho else "")
                        )
                if alvo_norm == "DRE":
                    _aguardar_dre_carregar(page)
                return

            caminho = _salvar_screenshot(page, "troca_formulario_nao_aplicou")
            msg = (
                f"Troca para '{formulario}' nao aplicou (pagina ficou em outro formulario)."
                + (f" Screenshot: {caminho}" if caminho else "")
            )
            raise RuntimeError(msg)
        except Exception as exc:
            ultima_exc = exc
            if tentativa < tentativas:
                log_linha(f"AVISO: falha ao trocar para {formulario} ({tentativa}/{tentativas}): {exc}")
                time.sleep(1.0)
                continue
            raise ultima_exc


def _clicar_guardar_topo(page, obrigatorio: bool = False) -> None:
    """
    Clica no botao Guardar do topo da tela.
    Deve ser chamado apos preencher cada formulario, antes de trocar de pagina/formulario.
    """
    _aguardar_carregar_pagina(page)
    frame = _esperar_frame_edicao(page, timeout_ms=PLAYWRIGHT_TIMEOUT_MS)

    def _listar_guardar_habilitados():
        base = frame.locator("input[type='submit'][value='Guardar'], input[value='Guardar'], button:has-text('Guardar')")
        saida = []
        try:
            n = base.count()
        except Exception:
            n = 0
        for i in range(n):
            b = base.nth(i)
            try:
                if not b.is_visible():
                    continue
                if not b.is_enabled():
                    continue
                box = b.bounding_box() or {}
                y = float(box.get("y", 99999))
                x = float(box.get("x", 99999))
                saida.append((y, x, b))
            except Exception:
                continue
        # Preferencia visual: barra superior (menor Y).
        saida.sort(key=lambda t: (t[0], t[1]))
        return saida

    def _clicar_confirmar_habilitado() -> bool:
        cands = frame.locator("input[type='submit'][value='Confirmar'], input[value='Confirmar']")
        try:
            n = cands.count()
        except Exception:
            n = 0
        for i in range(n):
            c = cands.nth(i)
            try:
                if c.is_visible() and c.is_enabled():
                    c.click(timeout=7000)
                    _aguardar_carregar_pagina(page)
                    return True
            except Exception:
                continue
        return False

    cands = _listar_guardar_habilitados()
    if not cands:
        # Em algumas telas o Guardar so habilita apos Confirmar.
        confirmou = _clicar_confirmar_habilitado()
        if confirmou:
            cands = _listar_guardar_habilitados()

    if not cands:
        caminho = _salvar_screenshot(page, "guardar_desabilitado_ou_nao_encontrado")
        msg = "Botao 'Guardar' nao encontrado habilitado (topo)."
        if caminho:
            msg += f" Screenshot: {caminho}"
        if obrigatorio:
            raise RuntimeError(msg)
        log_linha("AVISO: " + msg)
        return

    _, _, btn = cands[0]
    try:
        btn.scroll_into_view_if_needed(timeout=2000)
    except Exception:
        pass

    try:
        btn.click(timeout=10_000)
    except Exception as exc:
        caminho = _salvar_screenshot(page, "guardar_falhou")
        msg = f"Falha ao clicar em Guardar: {exc}"
        if caminho:
            msg += f" Screenshot: {caminho}"
        if obrigatorio:
            raise RuntimeError(msg)
        log_linha("AVISO: " + msg)
        return

    log_linha("Clicando em Guardar (botao habilitado).")
    _aguardar_carregar_pagina(page)
    caminho = _salvar_screenshot(page, "apos_guardar")
    if caminho:
        log_linha(f"Screenshot: {caminho}")


def _preencher_passivo_no_site(page, caminho_xlsx: str) -> None:
    linhas = ler_linhas_planilha_xlsx(caminho_xlsx, aba=ABA_PASSIVO)
    if not linhas:
        raise RuntimeError(
            "Nao encontrei linhas com codigo (ex.: M1, M2, ...) e valor calculado na aba Passivo do XLSX. "
            "Se o arquivo tiver formulas, confirme se foi salvo com os valores calculados."
        )

    _aguardar_carregar_pagina(page)
    frame = _esperar_frame_edicao(page, timeout_ms=PLAYWRIGHT_TIMEOUT_MS)
    usa_virgula = _detectar_decimal_virgula(frame)
    _scroll_to_top(frame, page)

    log_linha(f"Preenchendo Passivo (linhas: {len(linhas)}; decimal_virgula={usa_virgula}).")

    faltando: list[str] = []
    preenchidas = 0

    itens = linhas
    if MAX_LINHAS_POR_TESTE and MAX_LINHAS_POR_TESTE > 0:
        itens = linhas[:MAX_LINHAS_POR_TESTE]

    for item in itens:
        codigo = item["codigo"]
        valor = _ajustar_valor_por_descricao_para_site(
            codigo=item["codigo"],
            descricao=item.get("descricao", ""),
            valor=item["valor"],
        )

        inp = _localizar_input_por_nivel(frame, nivel=codigo)
        if not inp or inp.count() == 0:
            inp = _localizar_input_por_codigo_com_scroll(frame, page, codigo=codigo)
        if not inp or inp.count() == 0:
            faltando.append(codigo)
            continue

        try:
            if inp.is_disabled():
                continue
        except Exception:
            pass
        try:
            ro = (inp.get_attribute("readonly") or "").strip().lower()
            if ro in ["readonly", "true", "1"]:
                continue
        except Exception:
            pass

        txt_val = _formatar_valor_para_site(valor, usa_virgula=usa_virgula)
        try:
            inp.scroll_into_view_if_needed(timeout=2000)
        except Exception:
            pass
        ok_write, _, _ = _escrever_input_robusto(inp, txt_val, valor_esperado=float(valor))
        if ok_write:
            preenchidas += 1
        else:
            faltando.append(codigo)

    log_linha(f"Passivo preenchido: {preenchidas}/{len(itens)}.")
    if faltando:
        log_linha(f"AVISO: codigos nao preenchidos (nao encontrados/sem input): {', '.join(faltando[:30])}")
        if len(faltando) > 30:
            log_linha(f"AVISO: (+{len(faltando) - 30} codigos omitidos do log)")

    caminho_ok = _salvar_screenshot(page, "apos_preencher_passivo")
    if caminho_ok:
        log_linha(f"Screenshot: {caminho_ok}")

    # Dump de comparacao (site vs XLSX) para debug.
    try:
        comparacao: dict[str, dict] = {}
        divergencias: list[dict] = []

        for item in itens:
            codigo = item["codigo"]
            esp = _ajustar_valor_por_descricao_para_site(
                codigo=item["codigo"],
                descricao=item.get("descricao", ""),
                valor=item["valor"],
            )
            try:
                esp_f = float(esp)
            except Exception:
                continue

            comparacao[codigo] = {"esperado": esp_f}

            inp = _localizar_input_por_nivel(frame, nivel=codigo)
            if not inp or inp.count() == 0:
                inp = _localizar_input_por_codigo_com_scroll(frame, page, codigo=codigo)

            if not inp or inp.count() == 0:
                comparacao[codigo]["status"] = "skip_no_input"
                if abs(esp_f) > 0.005:
                    divergencias.append({"codigo": codigo, "status": "skip_no_input", "esperado": esp_f})
                continue

            try:
                comparacao[codigo]["readonly_attr"] = inp.get_attribute("readonly") or ""
            except Exception:
                pass
            try:
                comparacao[codigo]["disabled"] = bool(inp.is_disabled())
            except Exception:
                pass

            try:
                vtxt = inp.input_value(timeout=1000) or ""
                comparacao[codigo]["observado_str"] = vtxt
                obs_f = _parse_float_tolerante(vtxt)
                comparacao[codigo]["observado"] = obs_f
            except Exception:
                obs_f = None

            try:
                if inp.is_disabled():
                    comparacao[codigo]["status"] = "skip_disabled"
                    if abs(esp_f) > 0.005:
                        divergencias.append({"codigo": codigo, "status": "skip_disabled", "esperado": esp_f, "observado": obs_f})
                    continue
            except Exception:
                pass

            try:
                ro = (inp.get_attribute("readonly") or "").strip().lower()
                if ro in ["readonly", "true", "1"]:
                    comparacao[codigo]["status"] = "skip_readonly"
                    if abs(esp_f) > 0.005 and not _quase_igual(esp_f, obs_f):
                        divergencias.append({"codigo": codigo, "status": "skip_readonly", "esperado": esp_f, "observado": obs_f})
                    continue
            except Exception:
                pass

            if _quase_igual(esp_f, obs_f):
                comparacao[codigo]["status"] = "ok"
            else:
                comparacao[codigo]["status"] = "mismatch"
                divergencias.append({"codigo": codigo, "status": "mismatch", "esperado": esp_f, "observado": obs_f})

        pjson = _salvar_dump_json("passivo_compare_full", comparacao)
        if pjson:
            log_linha(f"Dump Passivo compare (json): {pjson}")
        pjson2 = _salvar_dump_json("passivo_compare_divergencias", divergencias)
        if pjson2:
            log_linha(f"Dump Passivo divergencias (json): {pjson2}")

        ptsv = ""
        try:
            linhas_tsv = ["codigo\tstatus\tesperado\tobservado\tobservado_str"]
            for d in divergencias:
                linhas_tsv.append(
                    f"{d.get('codigo','')}\t{d.get('status','')}\t{d.get('esperado','')}\t{d.get('observado','')}\t{str(d.get('observado_str','')).replace(chr(9),' ')}"
                )
            ptsv = _salvar_dump_texto("passivo_compare_divergencias", "\n".join(linhas_tsv) + "\n")
            if ptsv:
                log_linha(f"Dump Passivo divergencias (tsv): {ptsv}")
        except Exception:
            pass

        # Mantem diagnostico, mas por padrao nao bloqueia o fluxo no Ativo/Passivo.
        if divergencias:
            caminho = _salvar_screenshot(page, "passivo_divergencias")
            msg = (
                "Passivo: valores do site nao batem com o XLSX (incluindo totais/readonly). "
                + (f"Veja: {ptsv or pjson2 or pjson}. " if (ptsv or pjson2 or pjson) else "")
                + (f"Screenshot: {caminho}" if caminho else "")
            )
            if FALHAR_DIVERGENCIA_ATIVO_PASSIVO:
                raise RuntimeError(msg)
            log_linha("AVISO: " + msg)
    except RuntimeError:
        raise
    except Exception:
        pass


def _obter_mapa_colunas_dre_site_por_js(frame) -> dict[str, int]:
    """
    Fallback ultra-robusto:
    - roda JS no frame
    - acha a melhor linha de header (TOTAL/VN/VED/...)
    - retorna header canonico -> indice de CELULA (th/td) (0-based)

    Motivo: em alguns runs o Playwright falha/intermitente para "achar" o header por locators,
    apesar do HTML estar presente (race/virtualizacao/scroll container).
    """
    try:
        mapa = frame.evaluate(
            """() => {
  function norm(s) {
    s = (s || '').toString().trim().toUpperCase();
    try { s = s.normalize('NFD').replace(/[\\u0300-\\u036f]/g, ''); } catch (e) {}
    s = s.replace(/[^A-Z0-9 ]+/g, ' ');
    s = s.replace(/\\s+/g, ' ').trim();
    return s;
  }

  function canon(h) {
    h = norm(h);
    if (!h) return '';
    if (h === 'PEÇAS' || h === 'PECAS' || h === 'PE A S' || h === 'PEAS') return 'PECAS';
    if (h === 'ACESSORIOS' || h === 'ACESSORIO' || h === 'ACESS' || h === 'ACESSORIOS ') return 'ACESS';
    if (h === 'ADM' || h === 'ADMINISTRATIVO') return 'ADMINISTRATIVO';
    if (h === 'SEMUSOATUAL' || h === 'SEM USO ATUAL') return 'SEM USO ATUAL';
    if (['TOTAL','VN','VED','VSN','CNH','ATV','PF','BOUTIQUE','OFICINA'].includes(h)) return h;
    return '';
  }

  const alvo = new Set(['TOTAL','VN','VED','SEM USO ATUAL','VSN','CNH','ATV','PF','PECAS','ACESS','BOUTIQUE','OFICINA','ADMINISTRATIVO']);

  // Busca por tabelas provaveis do DRE (idealmente a de #inputForm).
  const tables = Array.from(document.querySelectorAll('#inputForm table, table'));

  let best = null; // {tr, hits, table}
  for (const tb of tables) {
    const trs = Array.from(tb.querySelectorAll('tr'));
    for (const tr of trs) {
      // Header tende a nao ter input.
      if (tr.querySelector('input')) continue;
      const cells = Array.from(tr.querySelectorAll('th,td'));
      if (cells.length < 10) continue;

      const keys = cells.map(c => canon(c.textContent));
      let hits = 0;
      let hasTotal = false, hasVN = false, hasVED = false;
      for (const k of keys) {
        if (!k) continue;
        if (alvo.has(k)) hits += 1;
        if (k === 'TOTAL') hasTotal = true;
        if (k === 'VN') hasVN = true;
        if (k === 'VED') hasVED = true;
      }

      // Header principal normalmente tem TOTAL+VN+VED.
      if (hits >= 4 && hasTotal && hasVN && hasVED) {
        if (!best || hits > best.hits) {
          best = { tr, hits, table: tb };
        }
      }
    }
  }

  if (!best) return {};

  const cells = Array.from(best.tr.querySelectorAll('th,td'));
  const mapa = {};
  for (let i = 0; i < cells.length && i < 200; i++) {
    const k = canon(cells[i].textContent);
    if (!k) continue;
    if (!(k in mapa)) mapa[k] = i;
  }
  return mapa;
}"""
        )
        if isinstance(mapa, dict):
            # Garantia de tipos.
            out: dict[str, int] = {}
            for k, v in mapa.items():
                try:
                    out[str(k)] = int(v)
                except Exception:
                    continue
            return out
    except Exception:
        pass
    return {}


def _obter_mapa_colunas_dre_site(frame) -> dict[str, int]:
    """
    Tenta mapear nome da coluna (VN/VSN/...) -> indice na linha (0-based) na tabela do site.
    """
    # 1) Acha UMA linha de header com a maior quantidade de colunas conhecidas.
    alvo = set([_norm_header(x) for x in DRE_COLUNAS_ALVO] + ["PF"])
    variantes = {
        "PECAS": ["PEÇAS", "PECAS"],
        "ACESS": ["ACESS", "ACESSÓRIOS", "ACESSORIOS"],
    }

    melhor_tr = None
    melhor_hits = 0

    trs = frame.locator("tr")
    total = 0
    try:
        total = trs.count()
    except Exception:
        total = 0

    # Limita para nao ficar caro se o DOM for enorme.
    limite = min(total, 200)

    for i in range(limite):
        tr = trs.nth(i)
        try:
            # Header row geralmente nao tem input.
            if tr.locator("input").count() > 0:
                continue
        except Exception:
            pass

        try:
            cells = tr.locator("th,td")
            n = cells.count()
        except Exception:
            continue

        hits = 0
        for j in range(min(n, 60)):
            try:
                tx = cells.nth(j).inner_text(timeout=200).strip()
            except Exception:
                continue
            hn = _norm_header(tx)
            if hn in alvo:
                hits += 1
                continue
            if hn in ["PECAS", "PEÇAS"]:
                hits += 1
                continue
            if hn in ["ACESS", "ACESSORIOS", "ACESSÓRIOS"]:
                hits += 1
                continue

        if hits > melhor_hits:
            melhor_hits = hits
            melhor_tr = tr

    if melhor_tr is None or melhor_hits < 4:
        return {}

    # 2) Monta o mapa baseado nessa linha.
    mapa: dict[str, int] = {}
    cells = melhor_tr.locator("th,td")
    try:
        n = cells.count()
    except Exception:
        n = 0

    for idx in range(min(n, 80)):
        try:
            tx = cells.nth(idx).inner_text(timeout=200).strip()
        except Exception:
            continue
        hn = _norm_header(tx)

        if hn in alvo:
            # Mantem o nome canonico usado no XLSX (PECAS/ACESS etc)
            if hn in ["PEÇAS", "PECAS"]:
                hn = "PECAS"
            if hn in ["ACESSORIOS", "ACESSÓRIOS", "ACESS"]:
                hn = "ACESS"
            if hn in ["ADMINISTRATIVO", "ADM"]:
                hn = "ADMINISTRATIVO"
            if hn == "SEMUSOATUAL":
                hn = "SEM USO ATUAL"
            mapa[hn] = idx
            continue

        if hn in ["PEÇAS", "PECAS"]:
            mapa["PECAS"] = idx
            continue
        if hn in ["ACESS", "ACESSORIOS", "ACESSÓRIOS"]:
            mapa["ACESS"] = idx
            continue

    return mapa


def _canon_col_dre_site(hn: str) -> str:
    """
    Normaliza o header do site para as chaves usadas no XLSX/config.
    hn deve vir de _norm_header().
    Retorna "" se nao for uma coluna alvo.
    """
    if not hn:
        return ""

    # Forma canonica (chaves do script)
    if hn in ["PEÇAS", "PECAS"]:
        return "PECAS"
    if hn in ["ACESS", "ACESSORIOS", "ACESSÓRIOS"]:
        return "ACESS"
    if hn in ["ADMINISTRATIVO", "ADM"]:
        return "ADMINISTRATIVO"
    if hn == "SEMUSOATUAL":
        return "SEM USO ATUAL"
    if hn in ["TOTAL", "VN", "VED", "VSN", "CNH", "ATV", "PF", "BOUTIQUE", "OFICINA"]:
        return hn
    return ""


def _obter_mapa_colunas_dre_site_por_header_cells(frame, page, codigo_referencia: str) -> dict[str, int]:
    """
    Versao mais robusta: mapeia header -> indice de celula (th/td) na linha, e NAO por indice de input.
    Isso evita preencher coluna errada quando existem colunas sem input / colunas somente-leitura.
    """
    # Ancora: pega a tabela/linha de um codigo conhecido.
    lab = _localizar_label_codigo_com_scroll(frame, page, codigo=codigo_referencia)
    if lab is None or lab.count() == 0:
        return {}

    try:
        tr_ref = lab.locator("xpath=ancestor::tr[1]").first
    except Exception:
        tr_ref = None
    if tr_ref is None or tr_ref.count() == 0:
        return {}

    try:
        n_cells_ref = tr_ref.locator("th,td").count()
    except Exception:
        n_cells_ref = 0

    try:
        # Prefere a tabela ancestral que claramente e o grid do DRE (tem TOTAL/VN no header).
        table = lab.locator(
            "xpath=ancestor::table[.//td[normalize-space()='TOTAL'] and .//td[normalize-space()='VN']][1]"
        ).first
        if table.count() == 0:
            table = tr_ref.locator("xpath=ancestor::table[1]").first
    except Exception:
        table = None
    if table is None or table.count() == 0:
        # Fallback: usa o frame inteiro.
        table = frame

    # Scroller horizontal (para colunas a direita).
    try:
        h_sc = frame.evaluate_handle(
            """() => {
  const els = Array.from(document.querySelectorAll('*'));
  let best = null;
  let bestArea = 0;
  for (const el of els) {
    const cs = window.getComputedStyle(el);
    if (!['auto','scroll'].includes(cs.overflowX)) continue;
    if (!el.scrollWidth || !el.clientWidth) continue;
    if (el.scrollWidth <= el.clientWidth + 10) continue;
    const r = el.getBoundingClientRect();
    const area = r.width * r.height;
    if (area > bestArea) { bestArea = area; best = el; }
  }
  return best;
}"""
        )
        scroller = h_sc.as_element()
    except Exception:
        scroller = None

    max_scroll = 0
    if scroller is not None:
        try:
            max_scroll = int(
                scroller.evaluate("(el) => Math.max(0, (el.scrollWidth || 0) - (el.clientWidth || 0))")
            )
        except Exception:
            max_scroll = 0

    scroll_pos = [0]
    if max_scroll > 0:
        scroll_pos = [
            0,
            int(max_scroll * 0.25),
            int(max_scroll * 0.5),
            int(max_scroll * 0.75),
            int(max_scroll),
        ]

    def _set_scroll(x: int) -> None:
        if scroller is not None:
            try:
                scroller.evaluate("(el, x) => { el.scrollLeft = x; }", x)
                return
            except Exception:
                pass
        try:
            frame.evaluate(
                "(x) => { const se = document.scrollingElement || document.documentElement; se.scrollLeft = x; }",
                x,
            )
        except Exception:
            pass

    # Procura uma linha de header com o maior numero de hits.
    def _achar_melhor_header_tr() -> object | None:
        alvo = set([_norm_header(x) for x in DRE_COLUNAS_ALVO] + ["PF"])
        trs = table.locator("tr")
        try:
            total = trs.count()
        except Exception:
            total = 0

        limite = min(total, 250)
        melhor_tr = None
        melhor_hits = 0

        # Fast-path: linha que contem TOTAL+VN+VED normalmente e o header principal.
        try:
            cand = trs.filter(has_text=re.compile(r"\bTOTAL\b", re.I)).filter(
                has_text=re.compile(r"\bVN\b", re.I)
            ).filter(has_text=re.compile(r"\bVED\b", re.I)).first
            if cand.count() > 0:
                try:
                    if cand.locator("input[type='text'], input[type='number']").count() == 0:
                        if cand.locator("th,td").count() >= 10:
                            return cand
                except Exception:
                    pass
        except Exception:
            pass

        for i in range(limite):
            tr = trs.nth(i)
            try:
                # Header row geralmente nao tem input.
                if tr.locator("input[type='text'], input[type='number']").count() > 0:
                    continue
            except Exception:
                pass

            try:
                cells = tr.locator("th,td")
                n = cells.count()
            except Exception:
                continue

            # Evita pegar header de outra tabela/grade com largura diferente.
            if n_cells_ref and abs(n - n_cells_ref) > 6:
                continue
            if n < 10:
                continue

            hits = 0
            for j in range(min(n, 80)):
                try:
                    tx = cells.nth(j).inner_text(timeout=200).strip()
                except Exception:
                    continue
                hn = _norm_header(tx)
                if hn in alvo:
                    hits += 1
                    continue
                # Variantes comuns
                if hn in ["PECAS", "PEÇAS"]:
                    hits += 1
                    continue
                if hn in ["ACESS", "ACESSORIOS", "ACESSÓRIOS"]:
                    hits += 1
                    continue
                if hn == "SEMUSOATUAL":
                    hits += 1
                    continue

            if hits > melhor_hits:
                melhor_hits = hits
                melhor_tr = tr

        if melhor_tr is None or melhor_hits < 4:
            return None
        return melhor_tr

    mapa: dict[str, int] = {}
    for sp in scroll_pos:
        _set_scroll(sp)
        time.sleep(0.15)

        htr = _achar_melhor_header_tr()
        if htr is None:
            continue

        cells = htr.locator("th,td")
        try:
            n = cells.count()
        except Exception:
            n = 0

        for idx in range(min(n, 120)):
            try:
                tx = cells.nth(idx).inner_text(timeout=200).strip()
            except Exception:
                continue
            key = _canon_col_dre_site(_norm_header(tx))
            if not key:
                continue
            # Guarda o indice de CELULA, nao de input.
            if key not in mapa:
                mapa[key] = idx

    _set_scroll(0)
    return mapa


def _obter_mapa_colunas_dre_site_por_bbox(frame, page, codigo_referencia: str) -> dict[str, int]:
    """
    Mapeia header -> indice da celula (th/td) na linha usando posicao X (bbox).
    Faz scanning de scroll horizontal para enxergar colunas a direita (ex.: OFICINA/ADMINISTRATIVO).

    Importante:
    - O codigo_referencia deve ser de um bloco com a grade completa (ex.: N17 em Pessoal).
    """

    # Inputs da linha de referencia (precisa estar visivel para bounding_box).
    lab = _localizar_label_codigo_com_scroll(frame, page, codigo=codigo_referencia)
    if lab is None or lab.count() == 0:
        return {}

    try:
        tr = lab.locator("xpath=ancestor::tr[1]").first
    except Exception:
        tr = None
    if tr is None or tr.count() == 0:
        return {}

    cells_row = tr.locator("th,td")
    try:
        n_cells = cells_row.count()
    except Exception:
        n_cells = 0
    if n_cells == 0:
        return {}

    # Tenta achar um scroller horizontal grande (para colunas a direita).
    try:
        h_sc = frame.evaluate_handle(
            """() => {
  const els = Array.from(document.querySelectorAll('*'));
  let best = null;
  let bestArea = 0;
  for (const el of els) {
    const cs = window.getComputedStyle(el);
    if (!['auto','scroll'].includes(cs.overflowX)) continue;
    if (!el.scrollWidth || !el.clientWidth) continue;
    if (el.scrollWidth <= el.clientWidth + 10) continue;
    const r = el.getBoundingClientRect();
    const area = r.width * r.height;
    if (area > bestArea) { bestArea = area; best = el; }
  }
  return best;
}"""
        )
        scroller = h_sc.as_element()
    except Exception:
        scroller = None

    max_scroll = 0
    if scroller is not None:
        try:
            max_scroll = int(
                scroller.evaluate("(el) => Math.max(0, (el.scrollWidth || 0) - (el.clientWidth || 0))")
            )
        except Exception:
            max_scroll = 0

    scroll_pos = [0]
    if max_scroll > 0:
        scroll_pos = [
            0,
            int(max_scroll * 0.25),
            int(max_scroll * 0.5),
            int(max_scroll * 0.75),
            int(max_scroll),
        ]

    def _set_scroll(x: int) -> None:
        if scroller is not None:
            try:
                scroller.evaluate("(el, x) => { el.scrollLeft = x; }", x)
                return
            except Exception:
                pass
        try:
            frame.evaluate(
                "(x) => { const se = document.scrollingElement || document.documentElement; se.scrollLeft = x; }",
                x,
            )
        except Exception:
            pass

    def _achar_header(col: str):
        variantes = [col]
        if col == "PECAS":
            variantes = ["PEÇAS", "PECAS", "PEÃ‡AS"]
        elif col == "ACESS":
            variantes = ["ACESS", "ACESSÓRIOS", "ACESSORIOS", "ACESSÃ“RIOS"]
        elif col == "ADMINISTRATIVO":
            variantes = ["ADMINISTRATIVO", "ADM"]
        elif col == "SEM USO ATUAL":
            variantes = ["Sem uso atual", "SEM USO ATUAL"]
        for v in variantes:
            loc = frame.locator(f"xpath=//*[normalize-space()='{v}']").first
            if loc.count() > 0:
                return loc
        return None

    colunas = [
        "TOTAL",
        "VN",
        "VED",
        "SEM USO ATUAL",
        "VSN",
        "CNH",
        "ATV",
        "PF",
        "PECAS",
        "ACESS",
        "BOUTIQUE",
        "OFICINA",
        "ADMINISTRATIVO",
    ]

    mapa: dict[str, int] = {}
    for col in colunas:
        for sp in scroll_pos:
            _set_scroll(sp)
            time.sleep(0.15)

            head = _achar_header(col)
            if head is None:
                continue

            try:
                box = head.bounding_box()
            except Exception:
                box = None
            if not box:
                continue

            cx = float(box["x"]) + float(box["width"]) / 2.0

            best_i = None  # indice de CELULA
            best_d = None
            for i in range(min(n_cells, 200)):
                try:
                    cell = cells_row.nth(i)
                    inp = cell.locator("input[type='text'], input[type='number'], input:not([type])").first
                    if inp.count() == 0:
                        continue
                    b = inp.bounding_box()
                except Exception:
                    b = None
                if not b:
                    continue
                ix = float(b["x"]) + float(b["width"]) / 2.0
                d = abs(ix - cx)
                if best_d is None or d < best_d:
                    best_d = d
                    best_i = i

            if best_i is not None:
                mapa[col] = int(best_i)
                break

    _set_scroll(0)
    return mapa

def _localizar_label_codigo_com_scroll(frame, page, codigo: str):
    """
    Localiza o label do codigo (ex.: N16) rolando a pagina se necessario.
    """
    codigo = _norm_txt(codigo)
    _scroll_to_top(frame, page)

    for _ in range(SCROLL_MAX_PASSOS_POR_CODIGO + 1):
        # Preferencia: match exato do codigo (ex.: "N10").
        lab = frame.locator(f"xpath=//*[normalize-space()='{codigo}']").first
        if lab.count() > 0:
            return lab

        # Fallback: as vezes o codigo vem junto com descricao no mesmo elemento (ex.: "N10 Lucro Bruto HDA").
        # Aqui buscamos por token em texto visivel.
        lab = frame.locator(f"text=/\\b{re.escape(codigo)}\\b/i").first
        if lab.count() > 0:
            return lab
        antes, depois = _scroll_down(frame)
        if depois <= antes:
            break
    return None


def _preencher_dre_no_site(page, caminho_xlsx: str, debug: bool = False) -> None:
    dados = ler_dre_xlsx_com_descricoes(caminho_xlsx)
    if not dados:
        raise RuntimeError("Nao consegui extrair dados da aba DRE do XLSX.")

    _aguardar_carregar_pagina(page)
    frame = _esperar_frame_edicao(page, timeout_ms=PLAYWRIGHT_TIMEOUT_MS)
    _scroll_to_top(frame, page)
    # Garante que a grade do DRE esta realmente renderizada antes de tentar detectar colunas.
    # Isso reduz falhas intermitentes de "colunas nao detectadas".
    try:
        _aguardar_dre_carregar(page)
    except Exception:
        pass

    codigos = sorted(dados.keys(), key=lambda x: int(x[1:]) if x[1:].isdigit() else 999999)
    # Referencias:
    # - Header: N1 costuma estar sempre visivel (menos scroll, menos chance de falhar).
    # - BBox: N17 e uma boa ancora quando "Pessoal" existe (grade costuma estar completa).
    codigo_ref_bbox = "N17" if "N17" in dados else (codigos[0] if codigos else "N1")
    codigo_ref_header = "N1" if "N1" in dados else codigo_ref_bbox

    # Preferencia: JS (mais resiliente contra race/virtualizacao/scroll container).
    # Depois: header cells (indice de th/td). Fallback: bbox. Ultimo fallback: heuristica global.
    mapa_col = _obter_mapa_colunas_dre_site_por_js(frame)
    if not mapa_col:
        mapa_col = _obter_mapa_colunas_dre_site_por_header_cells(frame, page, codigo_referencia=codigo_ref_header)
    if not mapa_col:
        mapa_col = _obter_mapa_colunas_dre_site_por_header_cells(frame, page, codigo_referencia=codigo_ref_bbox)
    if not mapa_col:
        mapa_col = _obter_mapa_colunas_dre_site_por_bbox(frame, page, codigo_referencia=codigo_ref_bbox)
    if not mapa_col:
        mapa_col = _obter_mapa_colunas_dre_site(frame)
    if not mapa_col:
        # Gera o maximo de debug possivel antes de falhar.
        try:
            _dump_dre_tela(page, debug_blocos=True)
        except Exception:
            pass
        try:
            _dump_estado_pagina(page, "estado_dre_colunas_nao_detectadas")
        except Exception:
            pass
        caminho = _salvar_screenshot(page, "dre_colunas_nao_detectadas")
        raise RuntimeError(
            "Nao consegui detectar as colunas do DRE no site (VN/VSN/...)." + (f" Screenshot: {caminho}" if caminho else "")
        )

    if debug:
        log_linha("DEBUG DRE: mapa_col (header->idx_cell): " + ", ".join([f"{k}={v}" for k, v in sorted(mapa_col.items(), key=lambda kv: kv[1])]))

    usa_virgula = _detectar_decimal_virgula(frame)

    total_celulas = 0
    preenchidas = 0
    faltando: list[str] = []

    log_linha(f"Preenchendo DRE (codigos: {len(codigos)}; cols_detectadas: {sorted(mapa_col.keys())}).")

    # Colunas a direita (normalmente exigem scroll horizontal para renderizar inputs).
    cols_direita = set(["PECAS", "ACESS", "BOUTIQUE", "OFICINA", "ADMINISTRATIVO"])

    def _obter_scroller_horizontal():
        try:
            h_sc = frame.evaluate_handle(
                """() => {
  const els = Array.from(document.querySelectorAll('*'));
  let best = null;
  let bestArea = 0;
  for (const el of els) {
    const cs = window.getComputedStyle(el);
    if (!['auto','scroll'].includes(cs.overflowX)) continue;
    if (!el.scrollWidth || !el.clientWidth) continue;
    if (el.scrollWidth <= el.clientWidth + 10) continue;
    const r = el.getBoundingClientRect();
    const area = r.width * r.height;
    if (area > bestArea) { bestArea = area; best = el; }
  }
  return best;
}"""
            )
            sc = h_sc.as_element()
        except Exception:
            sc = None
        max_scroll = 0
        if sc is not None:
            try:
                max_scroll = int(
                    sc.evaluate("(el) => Math.max(0, (el.scrollWidth || 0) - (el.clientWidth || 0))")
                )
            except Exception:
                max_scroll = 0
        return sc, max_scroll

    scroller, max_scroll = _obter_scroller_horizontal()

    def _set_scroll_x(x: int) -> None:
        if scroller is not None:
            try:
                scroller.evaluate("(el, x) => { el.scrollLeft = x; }", x)
                return
            except Exception:
                pass
        try:
            frame.evaluate(
                "(x) => { const se = document.scrollingElement || document.documentElement; se.scrollLeft = x; }",
                x,
            )
        except Exception:
            pass

    # Para debug: coleta comparacao site vs XLSX.
    # Estrutura:
    # {
    #   "N17": {
    #      "VN": {"esperado": -2264.34, "observado": -2264.34, "status": "ok|mismatch|skip_readonly|skip_no_input|skip_disabled", ...},
    #   }
    # }
    comparacao: dict[str, dict[str, dict]] = {}

    for codigo in codigos:
        desc = (dados[codigo].get("descricao") or "").strip()
        # No XLSX, algumas descricoes podem vir sem espacos/pontuacao padrao
        # (ex.: "SalárioGerente"). Para comparar com o texto do site, use forma "compacta".
        desc_key = _norm_header(desc)

        lab = _localizar_label_codigo_com_scroll(frame, page, codigo=codigo)

        tr = None
        if lab is not None and lab.count() > 0:
            try:
                lab.scroll_into_view_if_needed(timeout=2000)
            except Exception:
                pass
            try:
                tr = lab.locator("xpath=ancestor::tr[1]").first
            except Exception:
                tr = None

        # Fallback: procura qualquer row com o texto do codigo, mesmo que nao tenha achado o label.
        if tr is None or tr.count() == 0:
            tr = frame.locator("tr").filter(has_text=re.compile(rf"\\b{re.escape(codigo)}\\b", re.I)).first
            if tr.count() == 0:
                faltando.append(codigo)
                continue

        # Valida descricao na mesma linha (coluna B do XLSX).
        # Se nao bater, tenta achar outro tr que tenha codigo + descricao.
        if desc_key:
            try:
                row_text = _norm_txt(tr.inner_text(timeout=500))
            except Exception:
                row_text = ""
            row_key = _norm_header(row_text)
            if row_key and desc_key not in row_key:
                token = desc_key[:10] if len(desc_key) >= 10 else desc_key
                alt = frame.locator("tr").filter(
                    has_text=re.compile(rf"\\b{re.escape(codigo)}\\b", re.I)
                ).filter(has_text=re.compile(re.escape(token), re.I)).first
                if alt.count() > 0:
                    tr = alt
                else:
                    # Nao aborta por mismatch de descricao.
                    # Em alguns XLSX a descricao vem "compactada" (sem espacos),
                    # e a linha no site continua sendo a correta pelo codigo.
                    if debug:
                        log_linha(f"DEBUG DRE: descricao divergente em {codigo} (mantendo linha por codigo).")

        # Cells na linha (IMPORTANTE: mapa_col e indice de celula th/td).
        try:
            cells = tr.locator("th,td")
        except Exception:
            cells = None

        valores = dados[codigo]["valores"]
        for col, idx in mapa_col.items():
            # Alias de chave
            chave = col
            if chave == "ADM":
                chave = "ADMINISTRATIVO"
            if chave not in valores:
                continue
            val_raw = valores.get(chave)
            if val_raw is None:
                continue
            val_oficial = float(val_raw)
            hint_sinal = 0
            try:
                hint_sinal = int(((dados.get(codigo, {}) or {}).get("hints", {}) or {}).get(chave, 0))
            except Exception:
                hint_sinal = 0
            val_input = _ajustar_valor_por_descricao_para_site(
                codigo=codigo,
                descricao=desc,
                valor=val_oficial,
                forcar_positivo=False,
            )
            val_input = _ajustar_valor_dre_para_input(
                codigo=codigo,
                descricao=desc,
                valor=val_input,
            )
            total_celulas += 1

            # Inicializa registro da comparacao por celula.
            if codigo not in comparacao:
                comparacao[codigo] = {}
            if col not in comparacao[codigo]:
                comparacao[codigo][col] = {
                    "esperado": float(val_oficial),      # oficial XLSX (para readonly/totais)
                    "esperado_raw": float(val_oficial),
                    "esperado_input": float(val_input),  # valor efetivo para campo editavel
                    "hint_sinal": int(hint_sinal),       # -1=negativo colado | +1=marcador
                }

            # Dica: algumas colunas a direita so aparecem/ganham input apos scroll horizontal.
            # Ajusta scroll conforme a coluna antes de localizar o input.
            try:
                if max_scroll > 0:
                    if col in cols_direita:
                        _set_scroll_x(max_scroll)
                    else:
                        _set_scroll_x(0)
            except Exception:
                pass

            # Re-obtem cells apos scroll horizontal (DOM pode mudar/virtualizar).
            try:
                cells = tr.locator("th,td")
            except Exception:
                cells = None
            if cells is None:
                continue
            try:
                if cells.count() <= idx:
                    continue
            except Exception:
                continue

            cell = cells.nth(idx)
            inp = cell.locator("input[type='text'], input[type='number'], input:not([type])").first
            if inp.count() == 0:
                # Retry simples: tenta o outro extremo do scroll (caso o grupo esteja errado no layout).
                if max_scroll > 0:
                    try:
                        _set_scroll_x(0 if col in cols_direita else max_scroll)
                        cells = tr.locator("th,td")
                        if cells.count() > idx:
                            cell = cells.nth(idx)
                            inp = cell.locator("input[type='text'], input[type='number'], input:not([type])").first
                    except Exception:
                        pass
                if inp.count() == 0:
                    # Heuristica: se a celula esperada nao tem input, mas a proxima tem,
                    # pode ser sinal de deslocamento por coluna nao-editavel/mesclada.
                    try:
                        if cells.count() > (idx + 1):
                            nxt = cells.nth(idx + 1).locator(
                                "input[type='text'], input[type='number'], input:not([type])"
                            ).first
                            if nxt.count() > 0:
                                comparacao[codigo][col]["possivel_offset"] = "idx+1_tem_input"
                    except Exception:
                        pass
                    comparacao[codigo][col]["status"] = "skip_no_input"
                    continue

            # Captura atributos (para debug).
            try:
                comparacao[codigo][col]["readonly_attr"] = (inp.get_attribute("readonly") or "")
            except Exception:
                pass
            try:
                comparacao[codigo][col]["disabled"] = bool(inp.is_disabled())
            except Exception:
                pass

            try:
                if inp.is_disabled():
                    comparacao[codigo][col]["status"] = "skip_disabled"
                    continue
            except Exception:
                pass
            try:
                ro = (inp.get_attribute("readonly") or "").strip().lower()
                if ro in ["readonly", "true", "1"]:
                    # Se for readonly, ainda tenta ler valor atual para comparar.
                    try:
                        vtxt = inp.input_value(timeout=500) or ""
                        comparacao[codigo][col]["observado_str"] = vtxt
                        comparacao[codigo][col]["observado"] = _parse_float_tolerante(vtxt)
                    except Exception:
                        pass
                    comparacao[codigo][col]["status"] = "skip_readonly"
                    continue
            except Exception:
                pass

            # Campo editavel: regra atual do DRE -> sempre enviar valor positivo.
            base = float(val_input)
            candidatos: list[float] = []
            if abs(base) <= 0.005:
                candidatos = [0.0]
            else:
                candidatos = [abs(base)]

            melhor_val = float(candidatos[0])
            melhor_obs = None
            melhor_obs_str = ""
            melhor_dist = float("inf")
            preencheu_algum = False

            for cand in candidatos:
                txt = _formatar_valor_para_site(cand, usa_virgula=usa_virgula)
                ok_fill, vtxt_rb, obs_rb = _escrever_input_robusto(inp, txt, valor_esperado=float(cand))
                if not ok_fill:
                    continue
                preencheu_algum = True

                # Le de volta para comparacao (mesmo que seja string com separador BR).
                vtxt = vtxt_rb or ""
                obs = obs_rb

                # Distancia para escolher melhor candidato.
                if obs is None:
                    dist = float("inf")
                else:
                    dist = abs(float(cand) - float(obs))

                if dist < melhor_dist:
                    melhor_dist = dist
                    melhor_val = float(cand)
                    melhor_obs = obs
                    melhor_obs_str = vtxt

                # Se esse candidato ja bateu, para de testar.
                if _quase_igual(float(cand), obs):
                    break

            if not preencheu_algum:
                comparacao[codigo][col]["status"] = "fill_error"
                continue

            preenchidas += 1
            comparacao[codigo][col]["esperado"] = float(melhor_val)
            comparacao[codigo][col]["observado"] = melhor_obs
            comparacao[codigo][col]["observado_str"] = melhor_obs_str

            # Status ok/mismatch
            if _quase_igual(float(melhor_val), melhor_obs):
                comparacao[codigo][col]["status"] = "ok"
            else:
                comparacao[codigo][col]["status"] = "mismatch"

    log_linha(f"DRE preenchido: {preenchidas}/{total_celulas} celulas (com valor no XLSX).")
    if faltando:
        log_linha(f"AVISO: codigos nao encontrados na tela: {', '.join(faltando[:30])}")
        if len(faltando) > 30:
            log_linha(f"AVISO: (+{len(faltando) - 30} codigos omitidos do log)")

    # Screenshot diagnostico: tenta mostrar a grade (ex.: bloco Pessoal com N17).
    try:
        alvo = "N17" if "N17" in dados else codigo_ref_bbox
        lab2 = _localizar_label_codigo_com_scroll(frame, page, codigo=alvo)
        if lab2 is not None and lab2.count() > 0:
            try:
                lab2.scroll_into_view_if_needed(timeout=2000)
            except Exception:
                pass
            caminho_diag = _salvar_screenshot(page, f"dre_bloco_{alvo.lower()}")
            if caminho_diag:
                log_linha(f"Screenshot: {caminho_diag}")
    except Exception:
        pass

    caminho_ok = _salvar_screenshot(page, "apos_preencher_dre")
    if caminho_ok:
        log_linha(f"Screenshot: {caminho_ok}")

    # Dump da comparacao (site vs XLSX) para debug.
    try:
        # Lista de divergencias (mais facil de ler).
        divergencias: list[dict] = []
        for cod, cols in comparacao.items():
            for col, info in cols.items():
                st = info.get("status", "")
                esp = info.get("esperado", None)
                obs = info.get("observado", None)
                # Reporta mismatch e skips relevantes (valor esperado != 0).
                if st == "mismatch":
                    divergencias.append(
                        {"codigo": cod, "coluna": col, "status": st, "esperado": esp, "observado": obs, "obs_str": info.get("observado_str", "")}
                    )
                elif st in ["skip_no_input", "skip_readonly", "skip_disabled", "fill_error"]:
                    try:
                        if esp is not None and abs(float(esp)) > 0.005:
                            # Para readonly/disabled, so reporta quando houver divergencia real.
                            if st in ["skip_readonly", "skip_disabled"] and _quase_igual(float(esp), obs):
                                continue
                            divergencias.append(
                                {"codigo": cod, "coluna": col, "status": st, "esperado": esp, "observado": obs, "obs_str": info.get("observado_str", ""), "possivel_offset": info.get("possivel_offset", "")}
                            )
                    except Exception:
                        pass

        # Reconciliacao dinamica de readonly com base no comportamento da tela.
        try:
            qtd_antes = len(divergencias)
            divergencias = _filtrar_divergencias_dre_dinamicas(divergencias, comparacao)
            qtd_depois = len(divergencias)
            if qtd_depois < qtd_antes:
                log_linha(
                    f"DRE reconciliacao dinamica: {qtd_antes - qtd_depois} divergencias readonly resolvidas em tempo real."
                )
        except Exception:
            pass

        pjson = _salvar_dump_json("dre_compare_full", comparacao)
        if pjson:
            log_linha(f"Dump DRE compare (json): {pjson}")
        pjson2 = _salvar_dump_json("dre_compare_divergencias", divergencias)
        if pjson2:
            log_linha(f"Dump DRE divergencias (json): {pjson2}")

        # TSV curto
        ptsv = ""
        try:
            linhas = ["codigo\tcoluna\tstatus\tesperado\tobservado\tobservado_str\tpossivel_offset"]
            for d in divergencias:
                linhas.append(
                    f"{d.get('codigo','')}\t{d.get('coluna','')}\t{d.get('status','')}\t{d.get('esperado','')}\t{d.get('observado','')}\t{str(d.get('obs_str','')).replace(chr(9),' ')}\t{d.get('possivel_offset','')}"
                )
            ptsv = _salvar_dump_texto("dre_compare_divergencias", "\n".join(linhas) + "\n")
            if ptsv:
                log_linha(f"Dump DRE divergencias (tsv): {ptsv}")
        except Exception:
            pass

        # Regra: nunca editar o XLSX final por este script.
        # Se sobrar divergencia apenas em readonly (campos calculados do site),
        # apenas registra no log e segue sem alterar arquivo.
        if divergencias:
            try:
                _clicar_guardar_topo(page, obrigatorio=False)
            except Exception:
                pass

            nao_readonly = [d for d in divergencias if (d.get("status") or "") != "skip_readonly"]
            if not nao_readonly:
                caminho = _salvar_screenshot(page, "dre_divergencias_readonly_sincronizadas")
                log_linha(
                    "AVISO: DRE com divergencias apenas readonly (campos calculados do site). "
                    + "XLSX nao foi alterado por regra de seguranca. "
                    + (f"Veja: {ptsv or pjson2 or pjson}. " if (ptsv or pjson2 or pjson) else "")
                    + (f"Screenshot: {caminho}" if caminho else "")
                )
                return

            caminho = _salvar_screenshot(page, "dre_divergencias")
            raise RuntimeError(
                "DRE: valores do site nao batem com o XLSX (incluindo campos editaveis). "
                + (f"Veja: {ptsv or pjson2 or pjson}. " if (ptsv or pjson2 or pjson) else "")
                + (f"Screenshot: {caminho}" if caminho else "")
            )
    except RuntimeError:
        raise
    except Exception:
        pass


def _dump_dre_tela(page, debug_blocos: bool = True) -> None:
    """
    Captura texto (innerText) e prints por blocos do DRE para facilitar debug.
    Nao tenta preencher nada, apenas registrar estado da tela.
    """
    try:
        _aguardar_carregar_pagina(page)
        frame = _esperar_frame_edicao(page, timeout_ms=PLAYWRIGHT_TIMEOUT_MS)
    except Exception:
        return

    # Dump de texto da tela (evita seletor instavel; usa evaluate no DOM do frame).
    try:
        texto = frame.evaluate("() => (document.body && document.body.innerText) ? document.body.innerText : ''")
    except Exception:
        texto = ""
    if isinstance(texto, str) and texto.strip():
        caminho_txt = _salvar_dump_texto("dre_tela", texto)
        if caminho_txt:
            log_linha(f"Dump DRE (texto): {caminho_txt}")

    # Dump do HTML (ajuda a ajustar seletores).
    try:
        html = frame.evaluate("() => (document.documentElement && document.documentElement.outerHTML) ? document.documentElement.outerHTML : ''")
    except Exception:
        html = ""
    if isinstance(html, str) and html.strip():
        caminho_html = _salvar_dump_html("dre_tela", html)
        if caminho_html:
            log_linha(f"Dump DRE (html): {caminho_html}")

    if not debug_blocos:
        return

    # Prints por pontos de ancoragem para pegar as caixas "tabeladas".
    anchors = ["N1", "N17", "N32", "N83", "N116", "N117"]

    def _set_scroll_x(x: int) -> None:
        try:
            frame.evaluate(
                "(x) => { const se = document.scrollingElement || document.documentElement; se.scrollLeft = x; }",
                x,
            )
        except Exception:
            pass

    max_scroll = 0
    try:
        max_scroll = int(
            frame.evaluate(
                """() => {
  const se = document.scrollingElement || document.documentElement;
  return Math.max(0, (se && se.scrollWidth ? se.scrollWidth : 0) - (se && se.clientWidth ? se.clientWidth : 0));
}"""
            )
        )
    except Exception:
        max_scroll = 0

    for a in anchors:
        for pos in ([0, max_scroll] if max_scroll > 0 else [0]):
            try:
                _set_scroll_x(pos)
                time.sleep(0.15)
                lab = _localizar_label_codigo_com_scroll(frame, page, codigo=a)
                if lab is None or lab.count() == 0:
                    continue
                try:
                    lab.scroll_into_view_if_needed(timeout=2000)
                except Exception:
                    pass
                sufx = "right" if (max_scroll > 0 and pos == max_scroll) else "left"
                caminho = _salvar_screenshot(page, f"dre_bloco_{a.lower()}_{sufx}")
                if caminho:
                    log_linha(f"Screenshot: {caminho}")
            except Exception:
                continue


def executar_fluxo_playwright(
    url_login: str,
    usuario: str,
    senha: str,
    somente_login: bool,
    ate_editar: bool,
    ate_confirmar: bool,
    preencher_ativo: bool,
    preencher_passivo: bool,
    preencher_dre: bool,
    caminho_xlsx_ativo: str,
    competencia: str,
    fechar_apos: bool,
    pausar: bool,
    debug_dre: bool = False,
    dump_dre: bool = False,
) -> None:
    _importar_dependencias_playwright()
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=PLAYWRIGHT_HEADLESS)
        context = browser.new_context()
        page = context.new_page()
        page.set_default_timeout(PLAYWRIGHT_TIMEOUT_MS)

        # Captura logs do navegador para debug.
        if dump_dre:
            try:
                def _on_console(msg):
                    try:
                        log_linha(f"[console] {msg.type}: {msg.text}")
                    except Exception:
                        pass
                def _on_pageerror(err):
                    try:
                        log_linha(f"[pageerror] {err}")
                    except Exception:
                        pass
                page.on("console", _on_console)
                page.on("pageerror", _on_pageerror)
            except Exception:
                pass

        log_linha(f"Abrindo: {url_login}")
        _goto_com_retry(page, url_login)
        _aguardar_carregar_pagina(page)

        # Campos do print: placeholders "Username" e "Password"
        user = page.locator(
            "input[placeholder='Username'], input[placeholder='User'], input[name*='user' i], input[id*='user' i]"
        ).first
        pwd = page.locator(
            "input[placeholder='Password'], input[type='password'], input[name*='pass' i], input[id*='pass' i]"
        ).first

        try:
            user.wait_for(state="visible", timeout=20_000)
            pwd.wait_for(state="visible", timeout=20_000)
        except PWTimeoutError:
            caminho = _salvar_screenshot(page, "login_campos_nao_encontrados")
            raise RuntimeError(
                "Nao encontrei os campos de login (Username/Password)."
                + (f" Screenshot: {caminho}" if caminho else "")
            )

        user.fill(usuario)
        pwd.fill(senha)

        btn = page.get_by_role("button", name=re.compile(r"^login$", re.I))
        if btn.count() == 0:
            btn = page.locator("button:has-text('LOGIN'), button:has-text('Login'), input[type='submit']").first
            btn.click()
        else:
            btn.first.click()

        # SPA (hash). Em geral sai de "#/login" para outra rota.
        try:
            page.wait_for_function(
                "() => window.location && window.location.hash && !window.location.hash.toLowerCase().includes('login')",
                timeout=20_000,
            )
        except PWTimeoutError:
            # Fallback: se os campos ainda estiverem visiveis, assume falha
            try:
                if user.is_visible() or pwd.is_visible():
                    caminho = _salvar_screenshot(page, "login_falhou")
                    raise RuntimeError(
                        "Login nao saiu da tela de login (possivel usuario/senha invalidos ou captcha/politica)."
                        + (f" Screenshot: {caminho}" if caminho else "")
                    )
            except Exception:
                pass

        _aguardar_carregar_pagina(page)

        # Confirma login automaticamente abrindo a tela de submissions.
        frame_submissions = None
        try:
            frame_submissions = _validar_login_ok(page)
            log_linha("Login detectado automaticamente (submissions carregou).")
        except Exception as exc:
            log_linha(f"ERRO: {exc}")
            raise RuntimeError("Falha ao detectar login automaticamente.")

        caminho_ok = _salvar_screenshot(page, "login_ok")
        if caminho_ok:
            log_linha(f"Screenshot: {caminho_ok}")
        if dump_dre:
            _dump_estado_pagina(page, "estado_login_ok")

        # Processo assistido: se solicitado, pausa logo apos o login para o usuario confirmar que entrou.
        # Nao fecha o navegador aqui.
        # Modo nao interativo: nao pausa aguardando ENTER.
        if pausar:
            log_linha("Aviso: pausa apos login ignorada (execucao nao interativa).")

        if not somente_login:
            try:
                frame = frame_submissions or _ir_para_submissions(page)
                if ate_editar:
                    _clicar_editar_na_competencia_com_retry(frame, page, competencia=competencia)
                    if dump_dre:
                        _dump_estado_pagina(page, "estado_apos_editar")
                    if ate_confirmar:
                        # Entra no formulario inicial (Ativo) antes de preencher.
                        _selecionar_formulario_e_confirmar(page, formulario="Ativo")
                        if dump_dre:
                            _dump_estado_pagina(page, "estado_apos_confirmar")

                        # Importante: permitir rodar Passivo/DRE mesmo se Ativo estiver desabilitado via flag.
                        if preencher_ativo:
                            _preencher_ativo_no_site(page, caminho_xlsx=caminho_xlsx_ativo)
                            _clicar_guardar_topo(page, obrigatorio=False)
                            if dump_dre:
                                _dump_estado_pagina(page, "estado_apos_preencher_ativo")

                        if preencher_passivo:
                            _selecionar_formulario_e_confirmar(page, formulario="Passivo")
                            _preencher_passivo_no_site(page, caminho_xlsx=caminho_xlsx_ativo)
                            _clicar_guardar_topo(page, obrigatorio=False)
                            if dump_dre:
                                _dump_estado_pagina(page, "estado_apos_preencher_passivo")

                        if preencher_dre:
                            _selecionar_formulario_e_confirmar(page, formulario="DRE")
                            _preencher_dre_no_site(page, caminho_xlsx=caminho_xlsx_ativo, debug=debug_dre)
                            _clicar_guardar_topo(page, obrigatorio=False)
                            if dump_dre:
                                _dump_dre_tela(page, debug_blocos=True)
                                _dump_estado_pagina(page, "estado_apos_preencher_dre")
            except Exception as exc:
                log_linha(f"ERRO: {exc}")
                try:
                    caminho = _salvar_screenshot(page, "erro_fluxo")
                    if caminho:
                        log_linha(f"Screenshot: {caminho}")
                except Exception:
                    pass
                raise

        # Sempre encerra automaticamente para permitir reinicio limpo por empresa/login.
        try:
            context.close()
        except Exception:
            pass
        try:
            browser.close()
        except Exception:
            pass
        return


def organizar_no_site(empresa: str, caminho_arquivo: str) -> None:
    """
    TODO: implementar o fluxo do site.

    Para completar, preciso que voce confirme:
    - qual e o site (URL / sistema)
    - qual e o passo a passo (login, menu, tela, campos)
    - se usa Chrome/Edge e se ha extensoes/popup
    - onde anexar o arquivo (upload) e como confirmar sucesso
    """
    _importar_dependencias_ui()
    raise NotImplementedError("Fluxo do site ainda nao implementado (apos login).")


# =========================
# Main
# =========================


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Organiza/Envia arquivos finais no site (assistido).")
    p.add_argument(
        "--empresa",
        help="Processa apenas uma empresa (ex.: 22). Se omitido, usa empresas.txt.",
        default="",
    )
    p.add_argument(
        "--perfil",
        default=PERFIL_PADRAO,
        help="Perfil de login (LOBO/CASTRO/TELEMACO/MOTOACAO/RIO BRANCO).",
    )
    p.add_argument(
        "--url",
        default="",
        help="URL do login (sobrescreve URL_LOGIN).",
    )
    p.add_argument(
        "--somente-login",
        action="store_true",
        help="Apenas abre o site e tenta logar (nao processa empresas/arquivos).",
    )
    p.add_argument(
        "--ate-editar",
        action="store_true",
        help="Apos logar, vai em /submissions e clica em Editar na competencia do mes anterior (ou --competencia).",
    )
    p.add_argument(
        "--ate-confirmar",
        action="store_true",
        help="Apos clicar em Editar, entra no formulario 'Ativo' e clica em Confirmar.",
    )
    p.add_argument(
        "--preencher-ativo",
        action="store_true",
        help="Apos Confirmar, preenche o Ativo a partir do XLSX da empresa (BALANCETE AEF {empresa}.xlsx / final_{empresa}.xlsx / balancete.xlsx).",
    )
    p.add_argument(
        "--preencher-passivo",
        action="store_true",
        help="Depois de preencher o Ativo, troca para Passivo e preenche a partir do mesmo XLSX.",
    )
    p.add_argument(
        "--preencher-dre",
        action="store_true",
        help="Depois de preencher o Passivo, troca para DRE e preenche a partir do mesmo XLSX.",
    )
    p.add_argument(
        "--competencia",
        default="",
        help="Override da competencia (ex.: \"jan 2026\"). Se vazio, usa mes anterior.",
    )
    p.add_argument(
        "--fechar-apos-login",
        action="store_true",
        help="Fecha o navegador automaticamente apos o login.",
    )
    p.add_argument(
        "--pausar-apos-login",
        action="store_true",
        help="Pausa apos detectar login (ENTER para continuar).",
    )
    p.add_argument(
        "--nao-pausar-apos-login",
        action="store_true",
        help="Nao pausa aguardando ENTER apos logar.",
    )
    p.add_argument(
        "--somente-listar",
        action="store_true",
        help="Apenas lista as empresas/arquivos encontrados (nao executa automacao).",
    )
    p.add_argument(
        "--debug-dre",
        action="store_true",
        help="Log extra do mapeamento do DRE (colunas detectadas) e validacoes por descricao.",
    )
    p.add_argument(
        "--dump-dre",
        action="store_true",
        help="Salva dumps/prints para debug do DRE (por padrao em logs\\runs\\<run_id>\\dumps e logs\\runs\\<run_id>\\prints).",
    )
    p.add_argument(
        "--sem-debug",
        action="store_true",
        help="Desativa geracao extra de logs/prints/dumps por execucao (nao recomendado).",
    )
    p.add_argument(
        "--nao-limpar-runs",
        action="store_true",
        help="Nao remove runs antigos em logs\\runs apos executar (nao recomendado).",
    )
    p.add_argument(
        "--manter-runs",
        type=int,
        default=MANTER_ULTIMOS_RUNS_PADRAO,
        help=f"Quantos runs manter em logs\\runs (padrao: {MANTER_ULTIMOS_RUNS_PADRAO}).",
    )
    return p.parse_args()


def main() -> int:
    args = parse_args()

    # Por padrao, gera debug completo em logs\\runs\\... (prints, dumps, html, log por execucao).
    if not args.sem_debug:
        try:
            _init_debug_run(args)
            log_linha(f"Run dir: {RUN_DIR}")
        except Exception:
            pass

    if not args.somente_listar:
        # Modo padrao solicitado: sem flags, roda fluxo completo.
        if not args.somente_login and not args.ate_editar:
            args.ate_editar = True
            args.ate_confirmar = True
            args.preencher_ativo = True
            args.preencher_passivo = True
            args.preencher_dre = True
            if args.empresa:
                log_linha(
                    "Modo padrao ativado para empresa unica: --ate-editar --ate-confirmar --preencher-ativo --preencher-passivo --preencher-dre."
                )
            else:
                log_linha(
                    "Modo padrao ativado para todas as empresas: --ate-editar --ate-confirmar --preencher-ativo --preencher-passivo --preencher-dre."
                )

        if args.ate_confirmar and not args.ate_editar:
            log_linha("ERRO: --ate-confirmar exige --ate-editar.")
            return 1

        if (args.preencher_ativo or args.preencher_passivo or args.preencher_dre) and not args.ate_confirmar:
            log_linha("ERRO: --preencher-* exige --ate-confirmar (para entrar no formulario antes de preencher).")
            return 1

        perfis_norm = {_normalizar_perfil(p): p for p in PERFIS_LOGIN}
        perfil_in = args.perfil.strip()
        perfil_norm = _normalizar_perfil(perfil_in)
        if perfil_norm not in perfis_norm:
            log_linha(f"ERRO: perfil invalido: {args.perfil}. Perfis: {', '.join(PERFIS_LOGIN)}")
            return 1
        perfil_ok = perfis_norm[perfil_norm]

        url = (args.url or URL_LOGIN).strip()
        if not url:
            log_linha("ERRO: informe --url (link do login) ou preencha URL_LOGIN no topo do script.")
            return 1

        usuario = ""
        senha = ""
        if args.somente_login:
            try:
                usuario, senha = obter_credenciais_por_perfil(perfil_ok)
                log_linha(f"Perfil de login: {perfil_ok}.")
            except Exception as exc:
                log_linha(f"ERRO: {exc}")
                return 1

        if args.somente_login:
            try:
                competencia, origem_comp = _resolver_competencia(args.competencia)
                log_linha(f"Competencia alvo: {competencia} | origem: {origem_comp}")
                # Default: manter aberto. Para fechar automatico, use --fechar-apos-login.
                pausar = (not args.fechar_apos_login) and (
                    bool(args.pausar_apos_login) or (PAUSAR_APOS_LOGIN and (not args.nao_pausar_apos_login))
                )
                debug_dre = bool(args.debug_dre or (not args.sem_debug))
                dump_dre = bool(args.dump_dre or (not args.sem_debug))
                executar_fluxo_playwright(
                    url_login=url,
                    usuario=usuario,
                    senha=senha,
                    somente_login=True,
                    ate_editar=False,
                    ate_confirmar=False,
                    preencher_ativo=False,
                    preencher_passivo=False,
                    preencher_dre=False,
                    caminho_xlsx_ativo="",
                    competencia=competencia,
                    fechar_apos=args.fechar_apos_login,
                    pausar=pausar,
                    debug_dre=debug_dre,
                    dump_dre=dump_dre,
                )
                if not args.nao_limpar_runs:
                    removidos = _limpar_runs_antigos(args.manter_runs)
                    if removidos:
                        log_linha(f"Limpeza: removidos {len(removidos)} runs antigos de logs\\runs.")
                return 0
            except Exception as exc:
                log_linha(f"ERRO: {exc}")
                if not args.nao_limpar_runs:
                    removidos = _limpar_runs_antigos(args.manter_runs)
                    if removidos:
                        log_linha(f"Limpeza: removidos {len(removidos)} runs antigos de logs\\runs.")
                return 2

        if args.ate_editar:
            try:
                competencia, origem_comp = _resolver_competencia(args.competencia)
                log_linha(f"Competencia alvo: {competencia} | origem: {origem_comp}")
                pausar = (not args.fechar_apos_login) and (
                    bool(args.pausar_apos_login) or (PAUSAR_APOS_LOGIN and (not args.nao_pausar_apos_login))
                )
                debug_dre = bool(args.debug_dre or (not args.sem_debug))
                dump_dre = bool(args.dump_dre or (not args.sem_debug))

                if args.empresa:
                    empresas_exec = [_normalizar_empresa(args.empresa)]
                else:
                    empresas_exec = carregar_empresas(CAMINHO_EMPRESAS)

                houve_erro = False
                for idx_emp, emp in enumerate(empresas_exec, start=1):
                    log_linha(f"[{idx_emp}/{len(empresas_exec)}] Empresa alvo: {emp}")

                    caminho_xlsx = ""
                    if args.preencher_ativo or args.preencher_passivo or args.preencher_dre:
                        caminho_xlsx = localizar_arquivo_final(emp) or ""
                        if not caminho_xlsx:
                            log_linha(f"ERRO: nao encontrei o arquivo final da empresa {emp}.")
                            houve_erro = True
                            continue
                        try:
                            caminho_xlsx = _validar_xlsx_existente(caminho_xlsx, contexto="XLSX final")
                        except Exception as exc:
                            log_linha(f"ERRO: {exc}")
                            houve_erro = True
                            continue
                        log_linha(f"XLSX: {caminho_xlsx}")

                    perfil_emp = obter_perfil_por_empresa(emp) or perfil_ok
                    if perfil_emp != perfil_ok:
                        log_linha(f"Perfil de login (por codigo da empresa): {perfil_emp}.")
                    try:
                        usuario_emp, senha_emp = obter_credenciais_por_perfil(perfil_emp)
                    except Exception as exc:
                        log_linha(f"ERRO: {exc}")
                        houve_erro = True
                        continue

                    try:
                        executar_fluxo_playwright(
                            url_login=url,
                            usuario=usuario_emp,
                            senha=senha_emp,
                            somente_login=False,
                            ate_editar=True,
                            ate_confirmar=bool(args.ate_confirmar),
                            preencher_ativo=bool(args.preencher_ativo),
                            preencher_passivo=bool(args.preencher_passivo),
                            preencher_dre=bool(args.preencher_dre),
                            caminho_xlsx_ativo=caminho_xlsx,
                            competencia=competencia,
                            fechar_apos=args.fechar_apos_login,
                            pausar=pausar,
                            debug_dre=debug_dre,
                            dump_dre=dump_dre,
                        )
                        if args.preencher_ativo or args.preencher_passivo or args.preencher_dre:
                            rc4 = _rodar_etapa_4_pos_script3(emp)
                            log_linha(f"[SCRIPT 4] Retorno: {rc4}")
                            if rc4 != 0:
                                houve_erro = True
                    except Exception as exc:
                        log_linha(f"ERRO: falha na empresa {emp}: {exc}")
                        houve_erro = True
                        continue

                if not args.nao_limpar_runs:
                    removidos = _limpar_runs_antigos(args.manter_runs)
                    if removidos:
                        log_linha(f"Limpeza: removidos {len(removidos)} runs antigos de logs\\runs.")
                return 2 if houve_erro else 0
            except Exception as exc:
                log_linha(f"ERRO: {exc}")
                if not args.nao_limpar_runs:
                    removidos = _limpar_runs_antigos(args.manter_runs)
                    if removidos:
                        log_linha(f"Limpeza: removidos {len(removidos)} runs antigos de logs\\runs.")
                return 2

    if args.empresa:
        empresas = [_normalizar_empresa(args.empresa)]
    else:
        empresas = carregar_empresas(CAMINHO_EMPRESAS)

    tarefas = montar_tarefas(empresas)
    if not tarefas:
        log_linha("ERRO: nenhuma tarefa encontrada (nenhum arquivo final localizado).")
        return 1

    if args.somente_listar:
        log_linha("Modo somente-listar.")
        for emp, arq in tarefas:
            log_linha(f"OK: {emp} -> {arq}")
        return 0

    log_linha(f"Iniciando automacao (tarefas: {len(tarefas)}).")
    for i, (emp, arq) in enumerate(tarefas, start=1):
        log_linha(f"[{i}/{len(tarefas)}] Empresa {emp}: {arq}")
        try:
            organizar_no_site(emp, arq)
            log_linha(f"OK: empresa {emp}.")
        except NotImplementedError as exc:
            log_linha(f"ERRO: {exc}")
            return 2
        except Exception as exc:
            log_linha(f"ERRO: falha na empresa {emp}: {exc}")
            return 3

        time.sleep(PAUSA_ENTRE_EMPRESAS)

    log_linha("Finalizado.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

