import os
import re
import sys
import json
import asyncio
from datetime import datetime, date, timedelta
from typing import Optional, Tuple
from pathlib import Path
from urllib.parse import quote

from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError

# Mantém as dependências do relatório original
from decimal import Decimal, InvalidOperation
from xml.etree import ElementTree as ET
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ===================== CONFIGURAÇÃO =====================

# Pasta base para salvar: W:\XML PREFEITURA\<EMPRESA>\<PASTA_PERIODO>\
BASE_DOWNLOAD_DIR = r"W:\XML PREFEITURA"
BASE_DOWNLOAD_DIR_FALLBACK = r"\\192.0.0.251\Arquivos\XML PREFEITURA"

# Período (DD/MM/AAAA) — sempre do 1º ao último dia do mês anterior ao corrente
def periodo_mes_anterior(hoje: Optional[date] = None) -> Tuple[str, str]:
    hoje = hoje or date.today()
    primeiro_dia_mes_atual = hoje.replace(day=1)
    ultimo_dia_mes_anterior = primeiro_dia_mes_atual - timedelta(days=1)
    primeiro_dia_mes_anterior = ultimo_dia_mes_anterior.replace(day=1)
    return (
        primeiro_dia_mes_anterior.strftime("%d/%m/%Y"),
        ultimo_dia_mes_anterior.strftime("%d/%m/%Y"),
    )

DATA_INICIO = "01/01/2026"
DATA_FIM = "31/01/2026"

# Concorrência de downloads (XML/PDF) por página
MAX_CONCORRENCIA_DOWNLOADS = 15

# Se True, rebaixa arquivo mesmo se já existir
SOBRESCREVER_ARQUIVOS = False

# Playwright
HEADLESS = False
TIMEOUT_MS = 60_000
TIMEOUT_LOGIN_MS = 10 * 60_000  # 10 min
BROWSER_CHANNEL = "chrome"  # use Chrome instalado para aplicar policy de certificado

# Login automatico (gov.br)
# metodos: "certificado", "credenciais", "manual"
METODO_LOGIN = "certificado"
AUTO_LOGIN = True
CREDENCIAIS_JSON = Path(__file__).with_name("credenciais_nfse.json")
NFSE_CERT_CN = (os.getenv("NFSE_CERT_CN") or "").strip()


# ===================== UTILITÁRIOS =====================

def sanitize_folder(name: str) -> str:
    """Remove caracteres inválidos em nomes de pasta no Windows e normaliza espaços."""
    name = re.sub(r'[<>:"/\\|?*]+', ' ', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name if name else "Empresa_Desconhecida"


def resolver_base_download_dir() -> str:
    """
    Usa BASE_DOWNLOAD_DIR por padrao.
    Se o drive W: nao estiver disponivel na sessao, usa fallback UNC.
    Permite override por variavel de ambiente NFSE_BASE_DOWNLOAD_DIR.
    """
    env_dir = (os.getenv("NFSE_BASE_DOWNLOAD_DIR") or "").strip()
    if env_dir:
        return env_dir

    drive = os.path.splitdrive(BASE_DOWNLOAD_DIR)[0]
    if drive and not os.path.exists(drive + "\\"):
        return BASE_DOWNLOAD_DIR_FALLBACK
    return BASE_DOWNLOAD_DIR


def normalize_compare_text(s: str) -> str:
    s = (s or "").upper()
    s = re.sub(r"[^A-Z0-9]+", "", s)
    return s


def empresa_esperada_from_cert_cn(cert_cn: str) -> str:
    """
    Extrai nome da empresa a partir do CN do certificado.
    Ex.: "EMPRESA XYZ LTDA:12345678000199" -> "EMPRESA XYZ LTDA"
    """
    if not cert_cn:
        return ""
    parts = cert_cn.split(":", 1)
    return parts[0].strip()


def _try_parse_date(s: str) -> datetime:
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    raise ValueError(f"Data inválida: {s!r} (use DD/MM/AAAA)")


def pasta_periodo(data_inicio: str, data_fim: str) -> str:
    """Nomeia a pasta do período. Se for o mesmo mês/ano -> 'MM-AAAA', senão -> 'PERIODO_YYYYMMDD_YYYYMMDD'."""
    di = _try_parse_date(data_inicio)
    df = _try_parse_date(data_fim)

    if di.year == df.year and di.month == df.month:
        return f"{di.month:02d}-{di.year}"
    return f"PERIODO_{di.strftime('%Y%m%d')}_{df.strftime('%Y%m%d')}"


def carregar_credenciais() -> tuple[str, str] | None:
    """
    Carrega credenciais por prioridade:
      1) Variaveis de ambiente NFSE_CPF e NFSE_SENHA
      2) Arquivo credenciais_nfse.json ao lado do script:
         {"cpf": "...", "senha": "..."}
    """
    cpf_env = (os.getenv("NFSE_CPF") or "").strip()
    senha_env = (os.getenv("NFSE_SENHA") or "").strip()
    if cpf_env and senha_env:
        return cpf_env, senha_env

    if CREDENCIAIS_JSON.exists():
        try:
            with open(CREDENCIAIS_JSON, "r", encoding="utf-8") as f:
                data = json.load(f)
            cpf = str(data.get("cpf", "")).strip()
            senha = str(data.get("senha", "")).strip()
            if cpf and senha:
                return cpf, senha
        except Exception as e:
            print(f"[AVISO] Nao consegui ler credenciais em {CREDENCIAIS_JSON}: {e}")

    return None


async def _first_locator(page, selectors: list[str]):
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if await loc.count() > 0:
                return loc
        except Exception:
            pass
    return None


async def tentar_login_com_credenciais(page) -> bool:
    """
    Tenta login automatico por CPF/senha no gov.br.
    Retorna True se tentou preencher/enviar credenciais.
    """
    if not AUTO_LOGIN:
        return False

    creds = carregar_credenciais()
    if not creds:
        print("[INFO] Credenciais nao encontradas; seguindo com login manual.")
        return False

    cpf, senha = creds
    tentou = False

    # Tenta abrir o fluxo gov.br, se houver botao na pagina da NFSe
    try:
        btn_gov = await _first_locator(page, [
            "button:has-text('Entrar com gov.br')",
            "a:has-text('Entrar com gov.br')",
            "button:has-text('gov.br')",
            "a:has-text('gov.br')",
        ])
        if btn_gov:
            await btn_gov.click()
            await page.wait_for_load_state("domcontentloaded")
    except Exception:
        pass

    # CPF
    try:
        cpf_input = await _first_locator(page, [
            "input#accountId",
            "input[name='accountId']",
            "input[name='cpf']",
            "input[type='tel']",
            "input[inputmode='numeric']",
        ])
        if cpf_input:
            await cpf_input.fill(cpf)
            tentou = True
            btn_continuar = await _first_locator(page, [
                "button:has-text('Continuar')",
                "button:has-text('Avancar')",
                "button:has-text('Próxima')",
                "button:has-text('Proxima')",
                "input[type='submit']",
            ])
            if btn_continuar:
                await btn_continuar.click()
    except Exception as e:
        print(f"[AVISO] Falha ao preencher CPF no login automatico: {e}")

    # Senha
    try:
        senha_input = await _first_locator(page, [
            "input#password",
            "input[name='password']",
            "input[type='password']",
        ])
        if senha_input:
            await senha_input.fill(senha)
            tentou = True
            btn_entrar = await _first_locator(page, [
                "button:has-text('Entrar')",
                "button:has-text('Login')",
                "input[type='submit']",
            ])
            if btn_entrar:
                await btn_entrar.click()
    except Exception as e:
        print(f"[AVISO] Falha ao preencher senha no login automatico: {e}")

    if tentou:
        print("[INFO] Login automatico disparado; aguardando possivel validacao adicional (MFA/certificado).")
    return tentou


async def tentar_login_com_certificado(page) -> bool:
    """
    Tenta iniciar login por certificado digital no gov.br.
    Retorna True se conseguiu disparar o fluxo de certificado.
    """
    # Primeiro tenta o botao de imagem da propria tela NFSe
    try:
        btn_cert_nfse = await _first_locator(page, [
            "img[alt*='Certificado']",
            "img[title*='Certificado']",
            "img[src*='certificado']",
            "img[src*='Certificado']",
            "img[src*='e-CNPJ']",
            "img[src*='ecnpj']",
        ])
        if btn_cert_nfse:
            await btn_cert_nfse.click()
            await page.wait_for_load_state("domcontentloaded")
            print("[INFO] Clique no botao de certificado da tela NFSe realizado.")
            return True
    except Exception as e:
        print(f"[AVISO] Falha ao clicar no botao de certificado da NFSe: {e}")

    # Fallback para fluxo gov.br (caso a tela mude)
    try:
        btn_cert = await _first_locator(page, [
            "button:has-text('Seu certificado digital')",
            "a:has-text('Seu certificado digital')",
            "button:has-text('Entrar com gov.br')",
            "a:has-text('Entrar com gov.br')",
            "button:has-text('certificado digital')",
            "a:has-text('certificado digital')",
            "#login-certificate",
        ])
        if not btn_cert:
            print("[INFO] Botao de certificado nao encontrado automaticamente; seguindo com login manual.")
            return False

        await btn_cert.click()
        print("[INFO] Login por certificado disparado. Se abrir o seletor nativo, selecione o certificado.")
        return True
    except Exception as e:
        print(f"[AVISO] Falha ao disparar login por certificado: {e}")
        return False


async def tentar_login_automatico(page) -> bool:
    if not AUTO_LOGIN:
        return False

    metodo = (METODO_LOGIN or "").strip().lower()
    if metodo == "manual":
        return False
    if metodo == "credenciais":
        return await tentar_login_com_credenciais(page)

    # default: certificado
    return await tentar_login_com_certificado(page)


def build_list_url(base: str, pg: int, executar, data_inicio: str, data_fim: str, busca: str = "") -> str:
    """
    Monta a URL do portal com:
      pg, executar (pode repetir), busca, datainicio, datafim

    - 'executar' pode ser str ou lista[str]. Se lista, repete o parâmetro:
        executar=["1","1"]  -> ...&executar=1&executar=1
    - Mantém vírgula em valores (ex.: '1,1')
    - Escapa '/' nas datas (vira %2F)
    """
    if isinstance(executar, (list, tuple)):
        executar_list = list(executar)
    else:
        executar_list = [str(executar)]

    busca_enc = quote(busca or "", safe="")
    di_enc = quote(data_inicio, safe="")
    df_enc = quote(data_fim, safe="")

    executar_part = "".join([f"&executar={quote(v, safe=',')}" for v in executar_list])

    return (
        f"{base}?"
        f"pg={pg}"
        f"{executar_part}"
        f"&busca={busca_enc}"
        f"&datainicio={di_enc}"
        f"&datafim={df_enc}"
    )

    return (
        f"{base}?"
        f"pg={pg}"
        f"&executar={executar_enc}"
        f"&busca={busca_enc}"
        f"&datainicio={di_enc}"
        f"&datafim={df_enc}"
    )


async def wait_table_or_empty(page):
    """Espera a tabela carregar (linhas ou estado vazio)."""
    await page.wait_for_selector("table.table", timeout=TIMEOUT_MS)
    await page.wait_for_selector("table.table tbody", timeout=TIMEOUT_MS)

    # Aguarda o corpo popular (linhas) ou algum sinal de conteúdo.
    # Não usamos `networkidle` porque o site pode manter requisições em background.
    try:
        await page.wait_for_function(
            """() => {
                const tbody = document.querySelector('table.table tbody');
                if (!tbody) return false;
                const hasRow = tbody.querySelectorAll('tr').length > 0;
                const hasDownload = tbody.querySelector("a[href*='/Notas/Download/']") !== null;
                return hasRow || hasDownload;
            }""",
            timeout=3_000,
        )
    except PlaywrightTimeoutError:
        # Se não detectar, segue com o que já carregou
        pass


async def get_last_page_number(page) -> int:
    """
    Lê o máximo número de página no componente de paginação.
    Se não encontrar, assume 1.
    """
    try:
        texts = await page.eval_on_selector_all(
            "ul.pagination a",
            "els => els.map(e => (e.textContent || '').trim())"
        )
        nums = []
        for t in texts:
            if t.isdigit():
                nums.append(int(t))
        return max(nums) if nums else 1
    except Exception:
        return 1


async def obter_empresa_no_dashboard(page) -> str:
    """
    Lê o parágrafo que contém 'Nome:' no bloco 'Meus dados' do Dashboard
    e retorna o texto após 'Nome:' (o nome da empresa).
    """
    candidatos = [
        "section:has-text('Meus dados') p:has-text('Nome')",
        "div:has-text('Meus dados') p:has-text('Nome')",
        "p:has-text('Nome')",
        "xpath=(//p[contains(normalize-space(.),'Nome')])[1]",
    ]

    for sel in candidatos:
        try:
            p = page.locator(sel).first
            if await p.count() > 0:
                raw = (await p.inner_text()).strip()
                m = re.search(r"Nome\s*:\s*(.+)$", raw, flags=re.IGNORECASE)
                if m:
                    return sanitize_folder(m.group(1).strip())
        except Exception:
            pass

    # Fallbacks
    for label in ["Razão Social", "Razao Social", "Empresa"]:
        try:
            p = page.locator(f"p:has-text('{label}')").first
            if await p.count() > 0:
                raw = (await p.inner_text()).strip()
                parts = raw.split(":", 1)
                if len(parts) == 2:
                    return sanitize_folder(parts[1].strip())
        except Exception:
            pass

    return "Empresa_Desconhecida"


# ===================== DOWNLOAD CONCORRENTE =====================

async def baixar_binario(request, url: str, out_path: str, sem: asyncio.Semaphore, max_retries: int = 3) -> bool:
    """Baixa um arquivo via request.get (mesma sessão/cookies do navegador) com retries e limite de concorrência."""
    if (not SOBRESCREVER_ARQUIVOS) and os.path.exists(out_path):
        return True

    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    async with sem:
        for tent in range(1, max_retries + 1):
            try:
                resp = await request.get(url, timeout=TIMEOUT_MS)
                if not resp.ok:
                    raise RuntimeError(f"HTTP {resp.status}")
                data = await resp.body()
                with open(out_path, "wb") as f:
                    f.write(data)
                return True
            except Exception as e:
                if tent == max_retries:
                    print(f"❌ Falhou ({tent}/{max_retries}): {url} -> {e}")
                    return False
                await asyncio.sleep(0.6 * tent)
    return False


async def processa_pagina_download(page, request, pasta_destino: str, prefixo: str, sem: asyncio.Semaphore) -> int:
    """
    Baixa todos os XMLs/PDFs da página atual (sem filtrar por data — o filtro já vem na URL).
    Retorna o número de downloads disparados.
    """
    await wait_table_or_empty(page)

    rows = page.locator("table.table tbody tr")
    n = await rows.count()
    if n == 0:
        return 0

    tarefas = []
    vistos = set()  # evita duplicar por algum bug de DOM

    for i in range(n):
        row = rows.nth(i)

        # Situação (cancelada / substituida) pelo tooltip do ícone na coluna de situação
        status_suf = ""
        try:
            img = row.locator("td.td-situacao img").first
            tooltip = None
            if await img.count() > 0:
                tooltip = (await img.get_attribute("data-original-title")) or (await img.get_attribute("title"))
            if tooltip:
                t = tooltip.lower()
                if "cancelada" in t:
                    status_suf = " - cancelada"
                elif "substitui" in t:
                    status_suf = " - substituida"
        except Exception:
            pass

        # XML
        link_xml = row.locator("a[href*='/Notas/Download/NFSe/']").first
        if await link_xml.count() > 0:
            href = await link_xml.get_attribute("href")
            if href:
                url = href if href.startswith("http") else ("https://www.nfse.gov.br" + href)
                if url not in vistos:
                    vistos.add(url)
                    nome_padrao = href.split("/")[-1] + ".xml"
                    base, ext = os.path.splitext(nome_padrao)
                    novo_nome = f"{prefixo} {base}{status_suf}{ext}"
                    out_path = os.path.join(pasta_destino, novo_nome)
                    tarefas.append(baixar_binario(request, url, out_path, sem))

        # PDF
        link_pdf = row.locator("a[href*='/Notas/Download/DANFSe/']").first
        if await link_pdf.count() > 0:
            href = await link_pdf.get_attribute("href")
            if href:
                url = href if href.startswith("http") else ("https://www.nfse.gov.br" + href)
                if url not in vistos:
                    vistos.add(url)
                    nome_padrao = href.split("/")[-1] + ".pdf"
                    base, ext = os.path.splitext(nome_padrao)
                    novo_nome = f"{prefixo} {base}{status_suf}{ext}"
                    out_path = os.path.join(pasta_destino, novo_nome)
                    tarefas.append(baixar_binario(request, url, out_path, sem))

    if tarefas:
        await asyncio.gather(*tarefas)

    return len(tarefas)


async def baixar_todas_paginas(page, request, base_url: str, executar: str, pasta_destino: str, prefixo: str):
    """
    Percorre pg=1..N alterando a URL (sem clicar em 'fa-angle-right'),
    baixa tudo de cada página e para quando a página não tiver nenhuma nota (sem links de download).
    """
    sem = asyncio.Semaphore(MAX_CONCORRENCIA_DOWNLOADS)
    pg = 1
    while True:
        url = build_list_url(base_url, pg, executar, DATA_INICIO, DATA_FIM, busca="")
        print(f"\n--- {prefixo} | PÁGINA {pg} ---")
        print(url)

        await page.goto(url, wait_until="domcontentloaded", timeout=TIMEOUT_MS)

        try:
            await wait_table_or_empty(page)
        except PlaywrightTimeoutError:
            print("❌ Timeout ao carregar a tabela.")
            break

        rows = page.locator("table.table tbody tr")
        if await rows.count() == 0:
            print("✔ Lista vazia. Encerrando paginação.")
            break

        qtd = await processa_pagina_download(page, request, pasta_destino, prefixo=prefixo, sem=sem)
        print(f"✔ Downloads disparados nesta página: {qtd}")

        if qtd == 0:
            print("✔ Nenhuma nota detectada nesta página (sem links de download). Encerrando paginação.")
            break

        pg += 1


# ========== FUNÇÕES PARA GERAR RELATÓRIO (EX- nfse_extractor) ==========

OUTPUT_XLSX = "nfse_resumo.xlsx"
NS = {"n": "http://www.sped.fazenda.gov.br/nfse"}  # namespace padrão dos XMLs

def to_decimal(s: str) -> Decimal:
    if s is None or s == "":
        return Decimal("0")
    # normaliza vírgula/ponto
    s = s.strip().replace(".", "").replace(",", ".") if "," in s and "." in s else s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")

def get_text(root: ET.Element, path: str):
    el = root.find(path, NS)
    return el.text.strip() if el is not None and el.text is not None else None

def parse_xml(fp: Path) -> dict | None:
    try:
        root = ET.parse(fp).getroot()
    except Exception as e:
        print(f"[AVISO] Não consegui ler '{fp.name}': {e}")
        return None

    # -------- CAMPOS BÁSICOS --------
    n_nfse = get_text(root, ".//n:nNFSe")

    # data de emissão (DPS)
    dh_emi_raw = get_text(root, ".//n:DPS/n:infDPS/n:dhEmi")
    data_emissao_fmt = None
    if dh_emi_raw:
        try:
            data_emissao_fmt = datetime.fromisoformat(dh_emi_raw.split("T")[0]).strftime("%d/%m/%Y")
        except Exception:
            data_emissao_fmt = dh_emi_raw[:10]

    # razão social / CNPJ/CPF do TOMADOR (dentro do DPS)
    razao_tomador = get_text(root, ".//n:DPS/n:infDPS/n:toma/n:xNome")
    cnpj_tomador = get_text(root, ".//n:DPS/n:infDPS/n:toma/n:CNPJ")
    cpf_tomador  = get_text(root, ".//n:DPS/n:infDPS/n:toma/n:CPF")


    # razão social / CNPJ do EMITENTE (infNFSe/emit)
    razao_emitente = get_text(root, ".//n:infNFSe/n:emit/n:xNome")
    cnpj_emitente = get_text(root, ".//n:infNFSe/n:emit/n:CNPJ")

    # -------- VALORES --------
    # valor do serviço (DPS)
    v_serv = to_decimal(get_text(root, ".//n:DPS/n:infDPS/n:valores/n:vServPrest/n:vServ") or "0")

    # valor líquido da NFS-e (infNFSe)
    v_liq_xml = to_decimal(get_text(root, ".//n:infNFSe/n:valores/n:vLiq") or "0")

    # tributos federais / municipais (retidos)
    irrf = to_decimal(get_text(root, ".//n:DPS/n:infDPS/n:valores/n:trib/n:tribFed/n:vRetIRRF") or "0")
    pis = to_decimal(get_text(root, ".//n:DPS/n:infDPS/n:valores/n:trib/n:tribFed/n:piscofins/n:vPis") or "0")
    cofins = to_decimal(get_text(root, ".//n:DPS/n:infDPS/n:valores/n:trib/n:tribFed/n:piscofins/n:vCofins") or "0")
    csll = to_decimal(get_text(root, ".//n:DPS/n:infDPS/n:valores/n:trib/n:tribFed/n:vRetCSLL") or "0")

    # flag: tipo de retenção de PIS/COFINS (0/1)
    tp_ret_piscofins = get_text(
        root,
        ".//n:DPS/n:infDPS/n:valores/n:trib/n:tribFed/n:piscofins/n:tpRetPisCofins"
    )

    # NOVOS: INSS e ISSQN retidos (paths típicos do layout nacional)
    inss = to_decimal(get_text(root, ".//n:DPS/n:infDPS/n:valores/n:trib/n:tribFed/n:vRetCP") or "0")
    issqn = to_decimal(get_text(root, ".//n:DPS/n:infDPS/n:valores/n:trib/n:tribMun/n:tpRetISSQN") or "0")

    # por compatibilidade, mantemos um líquido "calculado"
    v_liq_calc = v_serv - (irrf + pis + cofins + csll + inss + issqn)

    # identificar notas canceladas/substituídas pelo nome do arquivo
    stem_l = fp.stem.lower()  # nome sem extensão
    situacao = ""
    if "- cancelada" in stem_l:
        situacao = "cancelada"
    elif "- substituida" in stem_l:
        situacao = "substituida"

    if situacao:
        # zera tudo para não impactar o resumo
        v_serv = Decimal("0")
        irrf = Decimal("0")
        pis = Decimal("0")
        cofins = Decimal("0")
        csll = Decimal("0")
        inss = Decimal("0")
        issqn = Decimal("0")
        v_liq_xml = Decimal("0")
        v_liq_calc = Decimal("0")

    return {
        "numero_nfse": n_nfse,
        "data_emissao": data_emissao_fmt,

        "situacao": situacao,

        "razao_social_tomador": razao_tomador,
        "cnpj_tomador": cnpj_tomador,
        "cpf_tomador": cpf_tomador,

        "razao_social_emitente": razao_emitente,
        "cnpj_emitente": cnpj_emitente,

        "valor_total_servicos": float(v_serv),

        "irrf_retido": float(irrf),
        "pis_retido": float(pis),
        "cofins_retido": float(cofins),
        "csll_retido": float(csll),
        "inss_retido": float(inss),
        "issqn_retido": float(issqn),

        "valor_liquido_servico": float(v_liq_xml),

        "tp_ret_piscofins": tp_ret_piscofins,   # <-- NOVO

        "arquivo": fp.name,
    }

def gerar_relatorio(target_dir: Path):
    """
    Gera nfse_resumo.xlsx com:
      - Aba EMITIDAS   (arquivos prefixo 'SAIDA ')
      - Aba RECEBIDAS  (arquivos prefixo 'ENTRADA ')

    Regras especiais:
      - EMITIDAS: usa razão social / CNPJ do TOMADOR
      - RECEBIDAS: usa razão social / CNPJ do EMITENTE
      - RECEBIDAS: só preenche colunas de retidos quando
                   valor_liquido_servico != valor_total_servicos
    """
    if not target_dir.exists() or not target_dir.is_dir():
        print(f"[ERRO] Pasta não encontrada para gerar relatório: {target_dir}")
        return

    xml_files = [p for p in target_dir.iterdir() if p.is_file() and p.suffix.lower() == ".xml"]
    if not xml_files:
        print("[AVISO] Nenhum XML encontrado na pasta para gerar relatório.")
        return

    rows = []
    for fp in xml_files:
        row = parse_xml(fp)
        if row:
            # tipo pela convenção de nome
            if fp.name.startswith("SAIDA "):
                tipo = "EMITIDAS"
            elif fp.name.startswith("ENTRADA "):
                tipo = "RECEBIDAS"
            else:
                tipo = "DESCONHECIDO"
            row["tipo"] = tipo
            rows.append(row)

    if not rows:
        print("[AVISO] Nenhum XML válido para relatório.")
        return

    df = pd.DataFrame(rows)

    # separa emitidas x recebidas
    df_emitidas = df[df["tipo"] == "EMITIDAS"].copy()
    df_recebidas = df[df["tipo"] == "RECEBIDAS"].copy()

    # -------- REGRAS ESPECIAIS PARA RECEBIDAS --------
    if not df_recebidas.empty:
        # 1) nas RECEBIDAS só mostra QUALQUER retido se valor liq != valor do serviço
        mask_iguais = df_recebidas["valor_total_servicos"] == df_recebidas["valor_liquido_servico"]
        ret_cols = [
            "irrf_retido",
            "pis_retido",
            "cofins_retido",
            "csll_retido",
            "inss_retido",
            "issqn_retido",
        ]
        df_recebidas.loc[mask_iguais, ret_cols] = 0

        # 2) PIS/COFINS só aparecem se tpRetPisCofins == '1'
        if "tp_ret_piscofins" in df_recebidas.columns:
            mask_piscofins = df_recebidas["tp_ret_piscofins"] == "1"
            df_recebidas.loc[~mask_piscofins, ["pis_retido", "cofins_retido"]] = 0

    # -------- REGRAS PARA EMITIDAS (apenas flag PIS/COFINS) --------
    if not df_emitidas.empty and "tp_ret_piscofins" in df_emitidas.columns:
        mask_piscofins_em = df_emitidas["tp_ret_piscofins"] == "1"
        df_emitidas.loc[~mask_piscofins_em, ["pis_retido", "cofins_retido"]] = 0


    # -------- MONTAGEM DAS ABAS --------

    # ABA EMITIDAS
    if not df_emitidas.empty:
        # cria coluna combinada CNPJ/CPF do tomador
        # prioridade: CNPJ; se vazio/nulo, usa CPF
        if "cpf_tomador" not in df_emitidas.columns:
            df_emitidas["cpf_tomador"] = None

        df_emitidas["cnpj_cpf_tomador"] = df_emitidas["cnpj_tomador"]
        mask_vazio = df_emitidas["cnpj_tomador"].isna() | (df_emitidas["cnpj_tomador"] == "")
        df_emitidas.loc[mask_vazio, "cnpj_cpf_tomador"] = df_emitidas["cpf_tomador"]

        col_emitidas = [
            "numero_nfse",
            "data_emissao",
            "situacao",
            "razao_social_tomador",
            "cnpj_cpf_tomador",        # <-- usa coluna combinada
            "valor_total_servicos",
            "irrf_retido",
            "pis_retido",
            "cofins_retido",
            "csll_retido",
            "inss_retido",
            "issqn_retido",
            "valor_liquido_servico",
            "arquivo",
        ]
        df_emitidas = df_emitidas[col_emitidas].rename(columns={
            "cnpj_cpf_tomador": "CNPJ/CPF",   # <-- nome da coluna no Excel
        })


    # ABA RECEBIDAS
    if not df_recebidas.empty:
        col_recebidas = [
            "numero_nfse",
            "data_emissao",
            "situacao",
            "razao_social_emitente",   # aqui mostramos o EMITENTE
            "cnpj_emitente",           # CNPJ do EMITENTE
            "valor_total_servicos",
            "irrf_retido",
            "pis_retido",
            "cofins_retido",
            "csll_retido",
            "inss_retido",
            "issqn_retido",
            "valor_liquido_servico",
            "arquivo",
        ]
        df_recebidas = df_recebidas[col_recebidas].rename(columns={
            "razao_social_emitente": "razao_social_emitente",
            "cnpj_emitente": "cnpj_emitente",
        })

    out_path = target_dir / OUTPUT_XLSX
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        if not df_emitidas.empty:
            df_emitidas.to_excel(xw, index=False, sheet_name="EMITIDAS")
        if not df_recebidas.empty:
            df_recebidas.to_excel(xw, index=False, sheet_name="RECEBIDAS")

    # Destaque visual no Excel para notas canceladas/substituídas (coluna 'situacao')
    try:
        wb = load_workbook(out_path)
        fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for sheet in ("EMITIDAS", "RECEBIDAS"):
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            headers = [c.value for c in ws[1]]
            if "situacao" not in headers:
                continue
            col_idx = headers.index("situacao") + 1
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=col_idx).value
                if val:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = fill
        wb.save(out_path)
    except Exception as e:
        print(f"[AVISO] Não consegui aplicar destaque no Excel: {e}")

    print(f"[OK] Relatório gerado: {out_path}")


# ===================== MAIN =====================

async def main():
    periodo_dirname = pasta_periodo(DATA_INICIO, DATA_FIM)
    base_download_dir = resolver_base_download_dir()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS, channel=BROWSER_CHANNEL)
        context = await browser.new_context()
        page = await context.new_page()

        # Login (manual)
        await page.goto("https://www.nfse.gov.br/EmissorNacional/Login", wait_until="domcontentloaded", timeout=TIMEOUT_MS)
        auto_ok = await tentar_login_automatico(page)
        if auto_ok:
            print("Aguardando conclusao do login...")
        else:
            print("Aguardando login...")

        # Aguarda o Dashboard
        await page.wait_for_url("**/Dashboard", timeout=TIMEOUT_LOGIN_MS)

        # Identifica a empresa
        empresa = await obter_empresa_no_dashboard(page)
        print(f"Empresa identificada: {empresa}")

        # Valida empresa esperada (quando executado pelo orquestrador por certificado)
        esperado = empresa_esperada_from_cert_cn(NFSE_CERT_CN)
        if esperado:
            emp_n = normalize_compare_text(empresa)
            esp_n = normalize_compare_text(esperado)
            if esp_n and (esp_n not in emp_n) and (emp_n not in esp_n):
                raise RuntimeError(
                    f"Empresa logada '{empresa}' nao confere com certificado esperado '{esperado}'."
                )

        # Cria pasta base W:\XML PREFEITURA\<EMPRESA>\<PERIODO>
        pasta_empresa = os.path.join(base_download_dir, empresa)
        pasta_destino = os.path.join(pasta_empresa, periodo_dirname)
        os.makedirs(pasta_destino, exist_ok=True)

        # ===================== EMITIDAS =====================
        await baixar_todas_paginas(
            page=page,
            request=context.request,
            base_url="https://www.nfse.gov.br/EmissorNacional/Notas/Emitidas",
            executar=["1","1"],
            pasta_destino=pasta_destino,
            prefixo="SAIDA",
        )

        # ===================== RECEBIDAS =====================
        await baixar_todas_paginas(
            page=page,
            request=context.request,
            base_url="https://www.nfse.gov.br/EmissorNacional/Notas/Recebidas",
            executar="1,1",  # conforme exemplo informado
            pasta_destino=pasta_destino,
            prefixo="ENTRADA",
        )

        await browser.close()

    # Relatório (usa os XMLs do período recém baixado)
    print(f"\nIniciando processamento do relatório na pasta: {pasta_destino}")
    gerar_relatorio(Path(pasta_destino))


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        sys.exit(0)
