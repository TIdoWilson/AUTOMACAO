import calendar
import re
import time
import unicodedata
from datetime import date, datetime
from pathlib import Path

import pyautogui
import pygetwindow as gw
import pytesseract
import tkinter as tk
from openpyxl import Workbook, load_workbook
from PIL import ImageEnhance, ImageOps

# =========================
# CONFIGURACAO
# =========================
PASTA_TXT = Path(__file__).resolve().parent / "BUSCAR RECIBO"
ARQUIVO_EXCEL_CONFERIDOS = Path(__file__).resolve().parent / "cnpjs_conferidos.xlsx"
NOME_ABA_EXCEL = "Conferidos"

ATRASO_INICIAL_SEGUNDOS = 5
ATRASO_CURTO = 0.4
ATRASO_APOS_PESQUISAR = 2.0
TIMEOUT_AVISO_CONFIRMACAO = 20.0

COORD_TROCAR_PERFIL = (237, 132)
COORD_SELECIONAR_PROCURADOR = (1026, 684)
COORD_TROCAR_CPF = (812, 682)
COORD_DIGITAR_CNPJ = (932, 685)
COORD_CONFIRMAR_PERFIL = (1126, 724)
COORD_SELECIONAR_SISTEMA = (885, 207)
COORD_TIPO_ARQUIVO = (588, 234)
COORD_TIPO_PESQUISA = (467, 262)
COORD_DATA_INICIO = (448, 301)
COORD_DATA_FIM = (343, 321)
COORD_APOS_DATA_FIM = (243, 437)
COORD_PESQUISAR = (929, 877)
COORD_MARCAR_TODOS = (232, 555)
COORD_SOLICITAR_MARCADOS = (1045, 892)

TITULO_JANELA = "Receitanet BX"
FORMATO_DATA = "%d/%m/%Y"

# Se necessario, descomente e ajuste para o caminho do tesseract instalado.
# Exemplo: pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

TERMOS_ERRO_OCR = [
    "nao existe procuracao eletronica",
    "solicite procuracao eletronica",
    "centro virtual de atendimento ao contribuinte",
]
TERMOS_EXTRA_ERRO_OCR = [
    "certificado digital",
    "e-cac",
    "(7)",
]


# =========================
# UTILITARIOS
# =========================
def obter_todos_txt(pasta: Path) -> list[Path]:
    arquivos = sorted(pasta.glob("*.txt"))
    if not arquivos:
        raise FileNotFoundError(f"Nenhum .txt encontrado em: {pasta}")
    return arquivos


def extrair_cnpj_primeira_linha(arquivo_txt: Path) -> str:
    with arquivo_txt.open("r", encoding="utf-8", errors="ignore") as arquivo:
        primeira_linha = arquivo.readline().strip()

    cnpjs = re.findall(r"\d{14}", primeira_linha)
    if len(cnpjs) != 1:
        raise ValueError(
            f"Esperado 1 CNPJ (14 digitos) na primeira linha, encontrado {len(cnpjs)} em: {arquivo_txt.name}"
        )

    return cnpjs[0]


def obter_periodo_digitado() -> tuple[str, str]:
    entrada = input("Informe o periodo (aceita MM/AAAA, MMAAAA ou MMAA): ").strip()
    entrada_sem_espaco = re.sub(r"\s+", "", entrada)

    mes_str = ""
    ano_str = ""

    if "/" in entrada_sem_espaco:
        partes = [p for p in entrada_sem_espaco.split("/") if p]
        if len(partes) != 2 or not partes[0].isdigit() or not partes[1].isdigit():
            raise ValueError("Periodo invalido. Exemplos: 02/2026, 022026, 0226")
        mes_str, ano_str = partes
    else:
        apenas_digitos = re.sub(r"\D", "", entrada_sem_espaco)
        if len(apenas_digitos) == 6:
            mes_str = apenas_digitos[:2]
            ano_str = apenas_digitos[2:]
        elif len(apenas_digitos) == 4:
            mes_str = apenas_digitos[:2]
            ano_str = f"20{apenas_digitos[2:]}"
        else:
            raise ValueError("Periodo invalido. Exemplos: 02/2026, 022026, 0226")

    mes_int = int(mes_str)
    ano_int = int(ano_str)
    if mes_int < 1 or mes_int > 12:
        raise ValueError("Mes invalido. Informe um mes entre 01 e 12.")
    if ano_int < 2000 or ano_int > 2099:
        raise ValueError("Ano invalido. Informe entre 2000 e 2099.")

    primeiro_dia = date(ano_int, mes_int, 1)
    ultimo_dia_num = calendar.monthrange(ano_int, mes_int)[1]
    ultimo_dia = date(ano_int, mes_int, ultimo_dia_num)

    return primeiro_dia.strftime(FORMATO_DATA), ultimo_dia.strftime(FORMATO_DATA)


def copiar_para_area_transferencia(texto: str) -> None:
    raiz = tk.Tk()
    raiz.withdraw()
    raiz.clipboard_clear()
    raiz.clipboard_append(texto)
    raiz.update()
    raiz.destroy()


def clicar(xy: tuple[int, int], descricao: str) -> None:
    print(f"Clicando em {descricao}: {xy}")
    pyautogui.click(xy[0], xy[1])
    time.sleep(ATRASO_CURTO)


def seta_baixo_e_enter() -> None:
    pyautogui.press("down")
    time.sleep(ATRASO_CURTO)
    pyautogui.press("enter")
    time.sleep(ATRASO_CURTO)


def preencher_campo_texto(xy: tuple[int, int], texto: str, descricao: str) -> None:
    clicar(xy, descricao)
    pyautogui.hotkey("ctrl", "a")
    time.sleep(ATRASO_CURTO)
    pyautogui.press("backspace")
    time.sleep(ATRASO_CURTO)
    pyautogui.write(texto, interval=0.02)
    time.sleep(ATRASO_CURTO)


def normalizar_texto(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return texto.lower()


def detectar_erro_sem_procuracao(texto_ocr: str) -> bool:
    texto_norm = normalizar_texto(texto_ocr)
    principais = sum(1 for termo in TERMOS_ERRO_OCR if termo in texto_norm)
    extras = sum(1 for termo in TERMOS_EXTRA_ERRO_OCR if termo in texto_norm)

    # OCR costuma falhar em 1 trecho da frase; por isso usamos regra por pontuacao.
    return principais >= 2 or (principais >= 1 and extras >= 1)


def capturar_texto_tela() -> str:
    screenshot = pyautogui.screenshot()
    largura, altura = screenshot.size

    # Foca no centro da tela, onde o popup de erro aparece.
    x1 = int(largura * 0.18)
    y1 = int(altura * 0.35)
    x2 = int(largura * 0.82)
    y2 = int(altura * 0.70)
    recorte = screenshot.crop((x1, y1, x2, y2))

    cinza = ImageOps.grayscale(recorte)
    contraste = ImageEnhance.Contrast(cinza).enhance(2.5)
    ampliada = contraste.resize((contraste.width * 2, contraste.height * 2))
    config = "--oem 3 --psm 6"
    return pytesseract.image_to_string(ampliada, lang="por", config=config)


def focar_janela_receitanet() -> None:
    titulo_alvo = TITULO_JANELA.lower()
    janelas = [j for j in gw.getAllWindows() if j.title and titulo_alvo in j.title.lower()]
    if not janelas:
        raise RuntimeError(f"Nao encontrei janela com titulo contendo: {TITULO_JANELA}")

    janela = next((j for j in janelas if j.title.strip().lower() == titulo_alvo), None)
    if janela is None:
        janela = next((j for j in janelas if j.title.strip().lower().startswith(titulo_alvo)), None)
    if janela is None:
        janela = janelas[0]

    if janela.isMinimized:
        janela.restore()
        time.sleep(0.3)

    try:
        janela.activate()
    except Exception:
        pyautogui.click(janela.left + 30, janela.top + 15)
        time.sleep(0.2)
        janela.activate()

    time.sleep(0.3)
    ativa = gw.getActiveWindow()
    titulo_ativo = (ativa.title or "") if ativa else ""
    if titulo_alvo not in titulo_ativo.lower():
        raise RuntimeError(f"Nao foi possivel focar '{TITULO_JANELA}'. Janela ativa: '{titulo_ativo}'")


def garantir_excel_conferidos(caminho_excel: Path) -> None:
    if caminho_excel.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = NOME_ABA_EXCEL
    ws.append(["DataHora", "CNPJ", "ArquivoTXT", "Conferido", "Status", "TextoOCR"])
    wb.save(caminho_excel)


def carregar_cnpjs_conferidos(caminho_excel: Path) -> set[str]:
    garantir_excel_conferidos(caminho_excel)
    wb = load_workbook(caminho_excel)
    ws = wb[NOME_ABA_EXCEL]

    conferidos: set[str] = set()
    for linha in ws.iter_rows(min_row=2, values_only=True):
        cnpj = str(linha[1]).strip() if linha[1] is not None else ""
        conferido = str(linha[3]).strip().upper() if linha[3] is not None else ""
        if cnpj and conferido == "SIM":
            conferidos.add(cnpj)

    wb.close()
    return conferidos


def registrar_resultado_excel(
    caminho_excel: Path,
    cnpj: str,
    arquivo_txt: str,
    conferido: bool,
    status: str,
    texto_ocr: str,
) -> None:
    garantir_excel_conferidos(caminho_excel)
    wb = load_workbook(caminho_excel)
    ws = wb[NOME_ABA_EXCEL]
    ws.append(
        [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            cnpj,
            arquivo_txt,
            "SIM" if conferido else "NAO",
            status,
            texto_ocr[:3000],
        ]
    )
    wb.save(caminho_excel)
    wb.close()


def executar_fluxo_para_cnpj(cnpj: str, data_inicio: str, data_fim: str) -> tuple[bool, str]:
    clicar(COORD_TROCAR_PERFIL, "Trocar Perfil")

    clicar(COORD_SELECIONAR_PROCURADOR, "Selecionar Procurador")
    seta_baixo_e_enter()

    clicar(COORD_TROCAR_CPF, "Trocar CPF")
    seta_baixo_e_enter()

    copiar_para_area_transferencia(cnpj)
    clicar(COORD_DIGITAR_CNPJ, "Digitar CNPJ")
    pyautogui.hotkey("ctrl", "v")
    time.sleep(ATRASO_CURTO)

    clicar(COORD_CONFIRMAR_PERFIL, "Confirmar Perfil")

    clicar(COORD_SELECIONAR_SISTEMA, "Selecione um Sistema")
    seta_baixo_e_enter()

    clicar(COORD_TIPO_ARQUIVO, "Selecione um Tipo de Arquivo")
    seta_baixo_e_enter()

    clicar(COORD_TIPO_PESQUISA, "Selecione um Tipo de Pesquisa")
    seta_baixo_e_enter()

    preencher_campo_texto(COORD_DATA_INICIO, data_inicio, "Data de inicio")
    preencher_campo_texto(COORD_DATA_FIM, data_fim, "Data de fim")
    pyautogui.press("enter")
    time.sleep(ATRASO_CURTO)
    pyautogui.press("enter")
    time.sleep(ATRASO_CURTO)
    clicar(COORD_APOS_DATA_FIM, "Clique apos Data de fim")

    clicar(COORD_PESQUISAR, "Pesquisar")
    time.sleep(ATRASO_APOS_PESQUISAR)

    texto_ocr = capturar_texto_tela()
    tem_erro_sem_procuracao = detectar_erro_sem_procuracao(texto_ocr)
    if tem_erro_sem_procuracao:
        pyautogui.press("enter")
        time.sleep(ATRASO_CURTO)
    else:
        trecho = " ".join(texto_ocr.split())[:220]
        print(f"OCR capturado (amostra): {trecho}")

    return tem_erro_sem_procuracao, texto_ocr


def solicitar_arquivos_marcados() -> None:
    clicar(COORD_MARCAR_TODOS, "Marcar todos")
    clicar(COORD_SOLICITAR_MARCADOS, "Solicitar arquivos marcados acima")
    time.sleep(ATRASO_CURTO)


def aguardar_aviso_e_confirmar() -> None:
    inicio = time.time()
    titulo_alvo = TITULO_JANELA.lower()

    while (time.time() - inicio) < TIMEOUT_AVISO_CONFIRMACAO:
        janelas = [j for j in gw.getAllWindows() if j.title and titulo_alvo in j.title.lower()]
        popup = next((j for j in janelas if j.width < 900 and j.height < 500), None)

        if popup is not None:
            try:
                popup.activate()
            except Exception:
                pass
            time.sleep(0.2)
            pyautogui.press("enter")
            time.sleep(ATRASO_CURTO)
            print("Aviso detectado e confirmado com Enter.")
            return

        time.sleep(0.5)

    # Fallback: alguns avisos nao aparecem como janela separada.
    pyautogui.press("enter")
    time.sleep(ATRASO_CURTO)
    print("Aviso nao detectado como popup; Enter enviado mesmo assim.")


# =========================
# FLUXO PRINCIPAL
# =========================
def main() -> None:
    pyautogui.FAILSAFE = True
    pyautogui.PAUSE = 0.15

    try:
        pytesseract.get_tesseract_version()
    except Exception as erro:
        raise RuntimeError(
            "Tesseract nao esta acessivel. Instale o Tesseract OCR e ajuste o caminho em pytesseract.pytesseract.tesseract_cmd."
        ) from erro

    data_inicio, data_fim = obter_periodo_digitado()
    cnpjs_conferidos = carregar_cnpjs_conferidos(ARQUIVO_EXCEL_CONFERIDOS)

    arquivos_txt = obter_todos_txt(PASTA_TXT)
    fila_processamento: list[tuple[Path, str]] = []
    for arquivo_txt in arquivos_txt:
        cnpj = extrair_cnpj_primeira_linha(arquivo_txt)
        if cnpj not in cnpjs_conferidos:
            fila_processamento.append((arquivo_txt, cnpj))

    if not fila_processamento:
        print("Nenhum CNPJ pendente. Todos ja estao conferidos no Excel.")
        return

    print(f"Data de inicio: {data_inicio}")
    print(f"Data de fim: {data_fim}")
    print(f"Total pendente: {len(fila_processamento)}")
    print(f"Iniciando em {ATRASO_INICIAL_SEGUNDOS}s. Tentando focar automaticamente a janela Receitanet BX.")
    time.sleep(ATRASO_INICIAL_SEGUNDOS)

    for indice, (arquivo_txt, cnpj) in enumerate(fila_processamento, start=1):
        print(f"\n[{indice}/{len(fila_processamento)}] Processando {arquivo_txt.name}")
        print(f"CNPJ: {cnpj}")

        focar_janela_receitanet()
        time.sleep(0.3)

        tem_erro_sem_procuracao, texto_ocr = executar_fluxo_para_cnpj(cnpj, data_inicio, data_fim)

        if tem_erro_sem_procuracao:
            registrar_resultado_excel(
                ARQUIVO_EXCEL_CONFERIDOS,
                cnpj,
                arquivo_txt.name,
                conferido=True,
                status="SEM_PROCURACAO",
                texto_ocr=texto_ocr,
            )
            print("Erro detectado por OCR. CNPJ marcado como conferido e registrado no Excel.")
            continue

        solicitar_arquivos_marcados()
        aguardar_aviso_e_confirmar()
        registrar_resultado_excel(
            ARQUIVO_EXCEL_CONFERIDOS,
            cnpj,
            arquivo_txt.name,
            conferido=True,
            status="SOLICITADO_ARQUIVO",
            texto_ocr=texto_ocr,
        )
        print("Sem erro de procuracao. Solicitei arquivos marcados e registrei no Excel.")
        continue

    print("Fluxo finalizado.")


if __name__ == "__main__":
    main()
